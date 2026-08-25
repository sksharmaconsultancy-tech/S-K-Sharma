# Iter 730 — fill module status + estimates into user's Payroll.xlsx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = "/tmp/payroll.xlsx"
DST = "/app/Payroll_Status_SKS.xlsx"

DONE = "✅ बना हुआ"
PART = "🟡 आंशिक बना"
NO = "❌ नहीं बना"

def N(s):
    return " ".join(str(s or "").lower().split())

# key (normalized substring) -> (status, remark/estimate)
M = [
    # 1 Recruitment
    ("current opening post", NO, "बन सकता है — Recruitment module के साथ (कुल 3-4 दिन)"),
    ("job description", NO, "Recruitment module में शामिल"),
    ("hod approval, budget", NO, "Recruitment module में शामिल"),
    ("interview schedule", NO, "Recruitment module में शामिल"),
    ("recruitment metrics", NO, "Recruitment module के बाद 1 दिन"),
    ("offer letters", DONE, "HR Letters module में मौजूद"),
    ("appointment letter", DONE, "HR Letters module में मौजूद"),
    ("promotion letter", DONE, "HR Letters module में मौजूद"),
    ("probation to confirmation letter", DONE, "HR Letters module में मौजूद"),
    ("probation to extension letter", DONE, "HR Letters module में मौजूद"),
    # 2 Performance Management
    ("employees life cycle & trend", PART, "Joining/Exit data मौजूद; trend report — 1-2 दिन"),
    ("apprisal", NO, "Performance module के साथ (कुल 4-6 दिन)"),
    ("kra", NO, "Performance module में शामिल"),
    ("kpa", NO, "Performance module में शामिल"),
    ("goals", NO, "Performance module में शामिल"),
    ("feedback", NO, "Performance module में शामिल"),
    ("job rotation / transfer", PART, "Branch/Dept transfer मौजूद; rotation history report — 1 दिन"),
    ("user satisfaction survey", NO, "बन सकता है — 1-2 दिन"),
    ("tracking performance metrics", NO, "Performance module के बाद 1 दिन"),
    # 3-7
    ("employee relations & grievances", PART, "Ticket/complaint system मौजूद; grievance escalation workflow — 1-2 दिन"),
    ("employee exit & analysis", PART, "Exit/Resign/Rejoin process मौजूद; exit analysis report — 1 दिन"),
    ("benefits", NO, "बन सकता है — 2-3 दिन (scope पर निर्भर)"),
    ("employees assets approval", NO, "बन सकता है — 2 दिन (asset issue/return + approval)"),
    ("tranning & development", NO, "बन सकता है — 3-4 दिन"),
    # 8 HR Analytics
    ("payroll expenditure to revenue ratio", NO, "बन सकता है — 1 दिन (revenue input चाहिए)"),
    ("absenteeism rate", PART, "Attendance data मौजूद; ratio report — 1 दिन"),
    ("employees attrition rate", NO, "बन सकता है — 1 दिन"),
    ("total overtime cost vs", PART, "OT cost reports मौजूद; comparison view — 1 दिन"),
    ("salary virance from previous month", PART, "Salary audit/compare मौजूद; variance report polish — 1 दिन"),
    ("recruitment cost & budget", NO, "Recruitment module के बाद"),
    # 9 HRIS Dashboard
    ("total manpower-cetegory wise", DONE, "Dashboard + group counts मौजूद"),
    ("new joined", DONE, "DOJ reports मौजूद"),
    ("left", DONE, "Exit/DOL reports मौजूद"),
    ("total hrs", DONE, "Attendance grid totals मौजूद"),
    ("ot hrs", DONE, "OT reports मौजूद"),
    ("depttwise cost & budget", PART, "Labour cost report मौजूद; budget tracking — 1-2 दिन"),
    ("depttwise present & absent summary", DONE, "Department-wise reports मौजूद"),
    ("freqent late commers", PART, "Late marks data मौजूद; frequent-list report — 1 दिन"),
    ("freqent absenters", PART, "Absent data मौजूद; frequent-list report — 1 दिन"),
    ("compare chart current month & previous", PART, "बन सकता है — 1-2 दिन"),
    # 10 Compliance
    ("man days work report", DONE, "Factory returns / labour statistics मौजूद"),
    ("average number of workers", DONE, "Annual returns में मौजूद"),
    ("number of man hours worked", DONE, "मौजूद"),
    ("number of man hours ot worked", DONE, "OT hours reports मौजूद"),
    ("total wages paid during the year", DONE, "Wage registers मौजूद"),
    ("total find/deduction during the year", DONE, "Fine/Deduction registers मौजूद"),
    ("statuary bonus report", DONE, "Bonus module + register मौजूद"),
    ("customized salary/wages register", DONE, "Report Formats में customized registers"),
    ("customized attendance register", DONE, "Customized attendance registers मौजूद"),
    ("esi accident recrods report", NO, "बन सकता है — 1 दिन"),
    ("traking hrs worked", DONE, "मौजूद"),
    # Masters & Policy
    ("employees master data", DONE, "पूरा Employee Master मौजूद"),
    ("employee probation to confirmation process report", PART, "Letters मौजूद; auto-tracking report — 1 दिन"),
    ("default shift assignment", DONE, "Shift Master + assignment मौजूद"),
    ("shift policy configuration", DONE, "Multi-shift config मौजूद"),
    ("leave policy, master & opening", DONE, "Leave policy + ESIC leave मौजूद"),
    ("holiday master", DONE, "मौजूद"),
    ("week off assign", DONE, "Firm + per-employee week off मौजूद"),
    ("alternate weekoff assign policy", PART, "Fixed weekoff मौजूद; alternate (2nd/4th Sat) — 1 दिन"),
    ("c-off policy configuration", DONE, "Comp-Off module मौजूद"),
    ("priviledge/rights assign", DONE, "RBAC roles & permissions मौजूद"),
    ("reporting manager & working flow", PART, "Approvals engine मौजूद; reporting-manager hierarchy — 2 दिन"),
    ("deptt hierarchy process & approval", PART, "Dept/branch मौजूद; hierarchy approvals — 2 दिन"),
    ("incremnetal/promotion type effective date", DONE, "CTC increment + arrear module मौजूद"),
    ("org. structure", NO, "बन सकता है — 1-2 दिन (org chart)"),
    ("employees dashboard, matrix & static analysis", DONE, "Dashboards + analytics मौजूद"),
    # Attendance & Policy
    ("biometric  in-out attendnace configuration", DONE, "ZKTeco push + USB import मौजूद"),
    ("biometric in-out attendnace configuration", DONE, "ZKTeco push + USB import मौजूद"),
    ("live attendance fatching", DONE, "Live sync engine + dashboard मौजूद"),
    ("count of present & absent as per time/shift", DONE, "Shift-wise present/absent मौजूद"),
    ("shift rosters daily/weekly/monthly", DONE, "Shift change/roster module मौजूद"),
    ("attendnace regularization policy", DONE, "Punch approvals + repair मौजूद"),
    ("gate pass with in-out time deduction", NO, "बन सकता है — 1-2 दिन"),
    ("miss punch/multi punch regularization", DONE, "Repair Punches + dedupe engine मौजूद"),
    ("out station employees attendnace", DONE, "Tours/OD module मौजूद"),
    ("computing of actual hrs work", DONE, "मौजूद"),
    ("computing of total hrs work", DONE, "मौजूद"),
    ("late mark policy/penatly", PART, "Late marks + grace मौजूद; auto salary penalty — 1 दिन"),
    ("employees live attend. dashboard", DONE, "मौजूद"),
    ("various attendnace format /reports", DONE, "कई formats (Excel/PDF) मौजूद"),
    # Overtime & Policy
    ("over time policy (upto 48", PART, "OT policies मौजूद; 48-hr cap + overflow-to-other-earning rule — 1 दिन"),
    ("over time count, hrs shift start & end", DONE, "Shift-based OT calc मौजूद"),
    ("minimum ot count 00:30", PART, "OT rounding मौजूद; 30-min minimum config — आधा दिन"),
    ("miss punch regularization with ot hrs", DONE, "मौजूद (cross-midnight OT सहित)"),
    ("overtime hrs calculation & policy", DONE, "मौजूद"),
    ("overtime approval process", DONE, "Punch/OT approvals मौजूद"),
    ("various ot reports & hrs report", DONE, "OT reports मौजूद"),
    ("customized ot report", DONE, "मौजूद"),
    # Salary Process & Policy
    ("salary strucutre assigned", DONE, "Salary structure + components मौजूद"),
    ("personal/loan advance", DONE, "Advances module मौजूद"),
    ("salary advance", DONE, "मौजूद"),
    ("tour expenditure & approval process", DONE, "Expense claims + approvals मौजूद"),
    ("pf calculation, implementation of abry", PART, "PF engine पूरा मौजूद; ABRY scheme flag — 1 दिन"),
    ("esi calculation", DONE, "ESIC engine (portal-matched rounding) मौजूद"),
    ("ot hrs. calculation & process", DONE, "मौजूद"),
    ("bonus, ex gratia, incentive", DONE, "Bonus module + incentive columns मौजूद"),
    ("deduction ( fine, uniform, assets", DONE, "Deduction heads मौजूद"),
    ("tds calculation, form 16", PART, "TDS column + Form 16 मौजूद; full tax planning (old/new regime) — 2-3 दिन"),
    ("f& f calculator", NO, "बन सकता है — 2 दिन (F&F settlement + format)"),
    ("required various format", DONE, "Report Formats hub मौजूद"),
    # ESS & Other
    ("emmployees self service portal", DONE, "ESS PWA (attendance/leave/OD) मौजूद"),
    ("mobile app self services", DONE, "PWA mobile app मौजूद"),
    ("auto mail to all hod for present & absent", DONE, "Daily report email scheduler मौजूद"),
    ("auto mail &  notificaton for approval", DONE, "Email + in-app notifications मौजूद"),
    ("auto mail & notificaton for approval", DONE, "Email + in-app notifications मौजूद"),
    ("employees greeting mail for birthday", NO, "बन सकता है — आधा-1 दिन"),
]

FILL = {
    DONE: PatternFill("solid", fgColor="C6EFCE"),
    PART: PatternFill("solid", fgColor="FFEB9C"),
    NO:   PatternFill("solid", fgColor="FFC7CE"),
}

def lookup(text):
    t = N(text)
    if not t:
        return None
    for key, st, rem in M:
        if key in t:
            return st, rem
    return None

wb = openpyxl.load_workbook(SRC)
for ws in wb.worksheets:
    # find which column holds the module text: HRIS uses B(section)/C(function);
    # HR Modules uses C. We match on BOTH B and C text per row.
    st_col, rm_col = ws.max_column + 1, ws.max_column + 2
    hdr_row = 1
    ws.cell(row=hdr_row, column=st_col, value="Status (SKS Portal)").font = Font(bold=True)
    ws.cell(row=hdr_row, column=rm_col, value="Remark / अनुमानित समय").font = Font(bold=True)
    for r in range(hdr_row + 1, ws.max_row + 1):
        hit = None
        for c in range(1, ws.max_column - 1):
            v = ws.cell(row=r, column=c).value
            got = lookup(v)
            if got:
                hit = got
                break
        if hit:
            st, rem = hit
            cs = ws.cell(row=r, column=st_col, value=st)
            cs.fill = FILL[st]
            cs.alignment = Alignment(vertical="center")
            ws.cell(row=r, column=rm_col, value=rem).alignment = Alignment(wrap_text=True, vertical="center")
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(st_col)].width = 16
    ws.column_dimensions[get_column_letter(rm_col)].width = 52

wb.save(DST)

# coverage check — report unmatched non-empty module rows
wb2 = openpyxl.load_workbook(DST)
miss = 0
for ws in wb2.worksheets:
    stc = ws.max_column - 1
    for r in range(2, ws.max_row + 1):
        texts = [ws.cell(row=r, column=c).value for c in range(1, stc)]
        has_text = any(N(t) and not N(t).isdigit() and N(t) not in ("policy",) for t in texts)
        # skip section-header-only rows (single word section labels)
        joined = N(" ".join(str(t or "") for t in texts))
        if has_text and not ws.cell(row=r, column=stc).value:
            if joined not in ("policy", "masters & policy 1 employees master data"):
                print("UNMATCHED [%s r%d]: %s" % (ws.title, r, joined[:90]))
                miss += 1
print("done, unmatched:", miss)
