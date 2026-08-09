"""Iter 533 (user request) — "Present / Absent + Daily OT Report".

A NEW report format inside the Present / Absent Report screen. The OLD
report (routes/present_absent_report.py) is NOT touched — this module is a
separate template that reuses the same attendance pipeline
(``_compute_monthly_grid_data``, which already applies the firm's
"Count Present Day @ 8 HRS — Compliance" policy: OT = worked − 8).

Layout: TWO rows per employee —
  Row 1: P / HD / A / WO / H per calendar day
  Row 2: that day's OT hours
plus monthly totals (Present Days · Absent Days · OT Hours).

Endpoints:
  GET /api/admin/reports/present-absent-ot         (JSON)
  GET /api/admin/reports/present-absent-ot.xlsx    (Excel, 2-row layout)
  GET /api/admin/reports/present-absent-ot.pdf     (landscape PDF, 2-row)
"""
import io
import sys
from datetime import date
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, Query
from fastapi.responses import Response

sys.path.append("/app/backend")
from server import db, _compute_monthly_grid_data  # noqa: E402
from routes.present_absent_report import _authz, _day_status  # noqa: E402

router = APIRouter(prefix="/api/admin/reports", tags=["present-absent-ot"])


def _fmt_ot(v: float):
    if not v:
        return 0
    return int(v) if float(v).is_integer() else round(float(v), 2)


async def _build_ot(company_id: str, month: str, department: str = "",
                    search: str = "") -> Dict[str, Any]:
    data = await _compute_monthly_grid_data(company_id=company_id, month=month)
    day_labels: List[str] = data.get("day_labels") or []
    weekday_labels = data.get("weekday_labels") or []
    day_full_dates: List[str] = data.get("day_full_dates") or [
        f"{month}-{str(dl)[:2]}" for dl in day_labels]
    today_iso = date.today().isoformat()

    comp = await db.companies.find_one(
        {"company_id": company_id},
        {"_id": 0, "name": 1, "address": 1, "attendance_policy": 1})
    _pol = (comp or {}).get("attendance_policy") or {}
    _fd = float(_pol.get("full_day_hours") or _pol.get("standard_working_hours") or 8.0)
    policy_line = (f"As per Firm Attendance Policy — Full Day {_fd:g} hrs · "
                   f"OT = worked hours beyond {_fd:g} per day · "
                   "P=Present  HD=Half Day  A=Absent  WO=Weekly Off  H=Holiday")

    # father names come from the Employee Master
    fathers = {u.get("employee_code"): (u.get("father_name") or "")
               async for u in db.users.find(
                   {"company_id": company_id, "role": "employee"},
                   {"_id": 0, "employee_code": 1, "father_name": 1})}

    term = (search or "").strip().lower()
    dep = (department or "").strip()
    rows: List[Dict[str, Any]] = []
    for e in data.get("employees") or []:
        if dep and (e.get("department") or "") != dep:
            continue
        if term and term not in f"{e.get('name', '')} {e.get('employee_code', '')}".lower():
            continue
        days: Dict[str, Dict[str, Any]] = {}
        absent = 0
        ot_total = 0.0
        for i, dl in enumerate(day_labels):
            cell = (e.get("days") or {}).get(dl) or {}
            iso = day_full_dates[i] if i < len(day_full_dates) else f"{month}-{str(dl)[:2]}"
            st = _day_status(cell, iso, today_iso)
            ot = float(cell.get("ot_hours") or 0) if st else 0.0
            days[dl] = {"st": st, "ot": _fmt_ot(ot)}
            if st == "A":
                absent += 1
            ot_total += ot
        rows.append({
            "employee_code": e.get("employee_code"),
            "name": e.get("name"),
            "father_name": fathers.get(e.get("employee_code")) or "",
            "department": e.get("department"),
            "designation": e.get("designation") or e.get("position"),
            "days": days,
            # policy Present Days exactly as payroll counts them
            "present_days": (e.get("totals") or {}).get("present_days_policy", 0),
            "absent_days": absent,
            "ot_total": _fmt_ot(round(ot_total, 2)),
        })
    rows.sort(key=lambda r: ((r.get("department") or ""),
                             str(r.get("employee_code") or "")))
    return {
        "company": {"company_id": company_id,
                    "name": (comp or {}).get("name"),
                    "address": (comp or {}).get("address") or ""},
        "month": month,
        "policy_line": policy_line,
        "day_labels": day_labels,
        "weekday_labels": weekday_labels,
        "employees": rows,
        "departments": sorted({r.get("department") or "" for r in rows} - {""}),
        "grand_totals": {
            "present_days": round(sum(float(r["present_days"] or 0) for r in rows), 1),
            "absent_days": sum(r["absent_days"] for r in rows),
            "ot_total": _fmt_ot(round(sum(float(r["ot_total"] or 0) for r in rows), 2)),
        },
    }


@router.get("/present-absent-ot")
async def pa_ot_json(company_id: str = Query(...), month: str = Query(...),
                     department: str = "", search: str = "",
                     authorization: Optional[str] = Header(None)):
    await _authz(authorization, company_id)
    return await _build_ot(company_id, month, department, search)


_FILL = {"P": "DCFCE7", "HD": "FEF9C3", "A": "FEE2E2",
         "WO": "E0F2FE", "H": "FED7AA"}


@router.get("/present-absent-ot.xlsx")
async def pa_ot_xlsx(company_id: str = Query(...), month: str = Query(...),
                     department: str = "", search: str = "",
                     authorization: Optional[str] = Header(None),
                     token: Optional[str] = Query(None)):
    await _authz(authorization or (f"Bearer {token}" if token else None), company_id)
    d = await _build_ot(company_id, month, department, search)
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "P-A + Daily OT"
    ws.cell(1, 1, d["company"]["name"]).font = Font(bold=True, size=14)
    ws.cell(2, 1, "Present / Absent + Daily OT Report — " + month).font = Font(bold=True, size=11)
    ws.cell(3, 1, d["policy_line"]).font = Font(size=9)
    hdr = ["Employee Name", "Father Name", "Designation", ""]
    dls = d["day_labels"]
    head_row = 5
    heads = (["Employee Name", "Father Name", "Designation", ""]
             + [str(int(str(dl)[:2])) for dl in dls]
             + ["Present", "Absent", "OT Hrs"])
    for c, h in enumerate(heads, start=1):
        cell = ws.cell(head_row, c, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 8
    from openpyxl.utils import get_column_letter
    for i in range(len(dls)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 4.4
    r = head_row
    for e in d["employees"]:
        r += 1
        r2 = r + 1
        # merged employee info cells across the 2 rows
        for col, val in ((1, e["name"]), (2, e["father_name"]), (3, e["designation"] or "")):
            ws.merge_cells(start_row=r, start_column=col, end_row=r2, end_column=col)
            cell = ws.cell(r, col, val)
            cell.alignment = Alignment(vertical="center")
            if col == 1:
                cell.font = Font(bold=True)
        ws.cell(r, 4, "Present").font = Font(size=8)
        ws.cell(r2, 4, "OT Hrs").font = Font(size=8, italic=True)
        for i, dl in enumerate(dls):
            cd = e["days"].get(dl) or {}
            st = cd.get("st") or ""
            c1 = ws.cell(r, 5 + i, st)
            c1.alignment = Alignment(horizontal="center")
            if st in _FILL:
                c1.fill = PatternFill("solid", fgColor=_FILL[st])
            c2 = ws.cell(r2, 5 + i, cd.get("ot") if st else "")
            c2.alignment = Alignment(horizontal="center")
            c2.font = Font(size=9, color="7C3AED", bold=bool(cd.get("ot")))
        base = 5 + len(dls)
        for col, val in ((base, e["present_days"]), (base + 1, e["absent_days"]),
                         (base + 2, e["ot_total"])):
            ws.merge_cells(start_row=r, start_column=col, end_row=r2, end_column=col)
            cell = ws.cell(r, col, val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
        r = r2
    # monthly summary
    gt = d["grand_totals"]
    r += 2
    for line in (f"Total Present Days: {gt['present_days']}",
                 f"Total Absent Days: {gt['absent_days']}",
                 f"Total OT Hours: {gt['ot_total']}"):
        ws.cell(r, 1, line).font = Font(bold=True)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="present-absent-ot-{month}.xlsx"'})


@router.get("/present-absent-ot.pdf")
async def pa_ot_pdf(company_id: str = Query(...), month: str = Query(...),
                    department: str = "", search: str = "",
                    authorization: Optional[str] = Header(None),
                    token: Optional[str] = Query(None)):
    await _authz(authorization or (f"Bearer {token}" if token else None), company_id)
    d = await _build_ot(company_id, month, department, search)
    from reportlab.lib import colors as rc
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)
    page = landscape(A4)
    buf = io.BytesIO()

    def on_page(canvas, doc):
        canvas.saveState()
        top = page[1] - 6 * mm
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(page[0] / 2, top - 5 * mm, d["company"]["name"] or "")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(page[0] / 2, top - 9 * mm, d["company"].get("address") or "")
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(page[0] / 2, top - 13.5 * mm,
                                 f"Present / Absent + Daily OT Report — {month}")
        canvas.setFont("Helvetica", 6.5)
        canvas.drawCentredString(page[0] / 2, top - 17 * mm, d["policy_line"])
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(page[0] - 10 * mm, 7 * mm, f"Page {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=page, leftMargin=7 * mm,
                            rightMargin=7 * mm, topMargin=25 * mm,
                            bottomMargin=11 * mm)
    dls = d["day_labels"]
    n_days = len(dls)
    info_w = [8 * mm, 32 * mm, 25 * mm, 22 * mm, 10 * mm]
    tot_w = [12 * mm, 12 * mm, 12 * mm]
    day_w = (page[0] - 14 * mm - sum(info_w) - sum(tot_w)) / max(1, n_days)
    widths = info_w + [day_w] * n_days + tot_w

    head = (["S.No.", "Employee Name", "Father Name", "Designation", ""]
            + [str(int(str(dl)[:2])) for dl in dls]
            + ["Pres", "Abs", "OT"])
    body = [head]
    styles: List[Any] = [
        ("FONTSIZE", (0, 0), (-1, -1), 5.8),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, rc.HexColor("#94A3B8")),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.3),
        ("LEFTPADDING", (0, 0), (-1, -1), 1),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        ("BACKGROUND", (0, 0), (-1, 0), rc.white),
    ]
    fill = {"P": "#DCFCE7", "HD": "#FEF9C3", "A": "#FEE2E2",
            "WO": "#E0F2FE", "H": "#FED7AA"}
    p_name = ParagraphStyle("n", fontSize=5.8, leading=6.6, fontName="Helvetica-Bold")
    p_sm = ParagraphStyle("m", fontSize=5.4, leading=6.2)
    for sno, e in enumerate(d["employees"], start=1):
        ri = len(body)
        r1 = ([str(sno),
               Paragraph(str(e["name"] or ""), p_name),
               Paragraph(str(e["father_name"] or ""), p_sm),
               Paragraph(str(e["designation"] or ""), p_sm), "P/A"]
              + [(e["days"].get(dl) or {}).get("st") or "" for dl in dls]
              + [e["present_days"], e["absent_days"], e["ot_total"]])
        r2 = (["", "", "", "", "OT"]
              + [((e["days"].get(dl) or {}).get("ot") if (e["days"].get(dl) or {}).get("st") else "")
                 for dl in dls]
              + ["", "", ""])
        body.append(r1)
        body.append(r2)
        # vertical merges: S.No. + info cells + totals span the 2 rows
        for c in (0, 1, 2, 3):
            styles.append(("SPAN", (c, ri), (c, ri + 1)))
        base = 5 + n_days
        for c in (base, base + 1, base + 2):
            styles.append(("SPAN", (c, ri), (c, ri + 1)))
            styles.append(("FONT", (c, ri), (c, ri), "Helvetica-Bold"))
        for i, dl in enumerate(dls):
            st = (e["days"].get(dl) or {}).get("st") or ""
            if st in fill:
                styles.append(("BACKGROUND", (5 + i, ri), (5 + i, ri),
                               rc.HexColor(fill[st])))
        styles.append(("TEXTCOLOR", (5, ri + 1), (4 + n_days, ri + 1),
                       rc.HexColor("#7C3AED")))
        styles.append(("BACKGROUND", (4, ri + 1), (4, ri + 1),
                       rc.HexColor("#F1F5F9")))
    tbl = Table(body, colWidths=widths, repeatRows=1)
    tbl.setStyle(TableStyle(styles))
    gt = d["grand_totals"]
    sum_style = ParagraphStyle("s", fontSize=9, leading=12, fontName="Helvetica-Bold")
    story = [tbl, Spacer(1, 6),
             Paragraph(f"Total Present Days: {gt['present_days']} &nbsp;&nbsp;·&nbsp;&nbsp; "
                       f"Total Absent Days: {gt['absent_days']} &nbsp;&nbsp;·&nbsp;&nbsp; "
                       f"Total OT Hours: {gt['ot_total']}", sum_style)]
    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition":
                             f'inline; filename="present-absent-ot-{month}.pdf"'})
etvalue(), media_type="application/pdf",
                    headers={"Content-Disposition":
                             f'inline; filename="present-absent-ot-{month}.pdf"'})
