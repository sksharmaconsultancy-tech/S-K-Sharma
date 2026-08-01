"""Iter 388 — PF & ESIC VALIDATION ENGINE (Phase 3 of the configurable
statutory module).

Validates every row of a Compliance Salary run BEFORE Salary Lock
(finalize). Issue levels:
  * ``error``   — always BLOCKS the lock.
  * ``warning`` — blocks unless a Super Admin locks with the explicit
    "allow_warnings" override (user-approved policy).

Exposed as ``GET /api/admin/compliance-salary-runs/{run_id}/validate``
and reused by the finalize endpoint in server.py.
"""
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
    now_iso,
)

router = APIRouter(prefix="/api")


def _num(v: Any, d: float = 0.0) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def _issue(code: str, level: str, message: str, suggestion: str) -> Dict[str, str]:
    return {"code": code, "level": level, "message": message, "suggestion": suggestion}


async def validate_compliance_run(run: Dict[str, Any]) -> Dict[str, Any]:
    """Employee-wise PF/ESIC validation of a computed run."""
    rows: List[Dict[str, Any]] = run.get("rows") or []
    month = str(run.get("month") or "")
    stat = run.get("statutory_effective") or run.get("statutory_cfg") or {}
    pf_cap = _num(stat.get("pf_wage_cap"), 15000.0)
    esic_ceiling = _num(stat.get("esic_gross_threshold"), 21000.0)
    wage_rule_on = stat.get("wage_definition_rule_enabled") is not False
    disable_above = stat.get("esic_disable_above_ceiling") is not False
    head_map = stat.get("head_mapping") if isinstance(stat.get("head_mapping"), dict) else None

    # Iter 423 (user bug — "Finalize Lock still showing validation error") —
    # the SALARY LOCK validation must respect the FIRM MASTER policy:
    # when the firm disabled EPF / ESIC, those checks are SKIPPED entirely
    # (same Iter 369 authoritative-flag logic as the salary engine —
    # explicit ``applicable`` wins, else the Deductions catalog decides).
    firm_pf_on = firm_esi_on = True
    _cid = run.get("company_id")
    if _cid:
        _fm = await db.firm_masters.find_one(
            {"company_id": _cid},
            {"_id": 0, "epf.applicable": 1, "esi.applicable": 1, "deductions": 1})
        if _fm is not None:
            _fm_ded = _fm.get("deductions") or {}
            _epf_ap = (_fm.get("epf") or {}).get("applicable")
            _esi_ap = (_fm.get("esi") or {}).get("applicable")
            firm_pf_on = bool(_epf_ap) if _epf_ap is not None else bool(_fm_ded.get("PF"))
            firm_esi_on = bool(_esi_ap) if _esi_ap is not None else bool(_fm_ded.get("ESI"))

    # Current Employee Master docs — detects post-calculation master changes
    # and missing/duplicate statutory IDs on the LIVE master.
    user_ids = [r.get("user_id") for r in rows if r.get("user_id")]
    masters: Dict[str, Dict[str, Any]] = {}
    if user_ids:
        async for u in db.users.find(
            {"user_id": {"$in": user_ids}},
            {"_id": 0, "user_id": 1, "uan_no": 1, "esi_ip_no": 1, "pf_basic": 1,
             "compliance_basic": 1, "pf_applicable": 1, "esic_applicable": 1,
             "higher_pension": 1, "eps_disabled": 1, "exit_date": 1,
             "esic_exit_date": 1, "excluded_employee": 1,
             # Iter 408 — Higher PF / VPF validation inputs.
             "pf_contribution_type": 1, "higher_pf_wage": 1, "higher_pf_from": 1,
             "higher_pf_to": 1, "pf_approval_status": 1, "pf_approval_required": 1,
             "pf_declaration_available": 1, "vpf_enabled": 1, "vpf_percent": 1,
             "vpf_amount": 1},
        ):
            masters[u["user_id"]] = u

    # Duplicate UAN / IP detection (within this run, non-blank only).
    uan_owners: Dict[str, List[str]] = {}
    ip_owners: Dict[str, List[str]] = {}
    for r in rows:
        uan = str(r.get("uan_no") or "").strip()
        ip = str(r.get("esi_ip_no") or "").strip()
        nm = str(r.get("name") or r.get("employee_code") or "?")
        if uan:
            uan_owners.setdefault(uan, []).append(nm)
        if ip:
            ip_owners.setdefault(ip, []).append(nm)

    global_issues: List[Dict[str, str]] = []
    if firm_esi_on and not wage_rule_on and head_map and not any(
            (head_map.get(k) or {}).get("esic") for k in head_map):
        global_issues.append(_issue(
            "ESIC_MAPPING_MISSING", "error",
            "Wage Definition Rule is OFF but NO earning head is flagged as ESIC Wage.",
            "Open Standard Compliance Settings → Salary Head Mapping and enable at least one ESIC head."))

    out_rows: List[Dict[str, Any]] = []
    errors = warnings = 0
    for r in rows:
        issues: List[Dict[str, str]] = []
        uid = r.get("user_id")
        m = masters.get(uid) or {}
        name = str(r.get("name") or "?")
        gross = _num(r.get("gross_paid"))
        pd = _num(r.get("present_days"))
        pf_emp = _num(r.get("pf_employee"))
        pf_wages = _num(r.get("pf_wages"))
        esic_emp = _num(r.get("esic_employee"))
        esic_base = _num(r.get("esic_wage_base"))
        uan = str(r.get("uan_no") or m.get("uan_no") or "").strip()
        ip = str(r.get("esi_ip_no") or m.get("esi_ip_no") or "").strip()

        # ---------------- PF checks ----------------
        if r.get("pf_applicable") and gross > 0 and pf_emp <= 0:
            issues.append(_issue(
                "PF_ZERO", "error",
                "PF is applicable but the PF deduction is 0.",
                "Re-run Salary Process, or check PF Basic Salary / rates in Compliance Settings."))
        if pf_emp > 0 and not uan:
            issues.append(_issue(
                "PF_MISSING_UAN", "error",
                "PF deducted but UAN is missing on the Employee Master.",
                "Fill the 12-digit UAN in Employee Master → Statutory, or generate via EPFO."))
        if uan and len(uan_owners.get(uan, [])) > 1:
            others = [n for n in uan_owners[uan] if n != name][:2]
            issues.append(_issue(
                "PF_DUP_UAN", "error",
                f"UAN {uan} is shared with {', '.join(others) or 'another employee'}.",
                "Each employee must have a unique UAN — correct the Employee Master."))
        if pf_wages < 0:
            issues.append(_issue(
                "PF_WAGE_INVALID", "error",
                f"PF wages are negative (₹{pf_wages:,.0f}).",
                "Re-run Salary Process for this month."))
        elif pf_emp > 0 and gross > 0 and pf_wages > gross + 0.5 and not r.get("intl_worker"):
            issues.append(_issue(
                "PF_WAGE_INVALID", "warning",
                f"PF wages ₹{pf_wages:,.0f} exceed the Gross Earning ₹{gross:,.0f}.",
                "Check PF Basic Salary / proration method — wages above gross are unusual."))
        if pf_emp > 0 and pf_wages > pf_cap + 0.5 and not r.get("intl_worker") \
                and not r.get("pf_higher_active"):
            issues.append(_issue(
                "PF_ABOVE_CEILING", "error",
                f"PF wages ₹{pf_wages:,.0f} exceed the ceiling ₹{pf_cap:,.0f}.",
                "Only International Workers / approved Higher PF may cross the EPF ceiling — re-run Salary Process."))

        # ---------------- Higher PF / VPF checks (Iter 408) ----------------
        _pft = str(r.get("pf_contribution_type")
                   or m.get("pf_contribution_type") or "statutory").lower()
        if _pft == "higher":
            if not stat.get("allow_higher_pf"):
                issues.append(_issue(
                    "HIGHER_PF_NOT_ALLOWED", "error",
                    "Higher PF selected but the company policy does not allow Higher PF.",
                    "Enable 'Allow Higher PF' in Standard Compliance Settings, or set the employee back to Statutory PF."))
            if not (r.get("pf_declaration_available")
                    or m.get("pf_declaration_available")):
                issues.append(_issue(
                    "HIGHER_PF_NO_DECLARATION", "error",
                    "Higher PF selected but the Employee Declaration is missing.",
                    "Tick 'Employee Declaration Available' on the Employee Master after collecting the joint declaration."))
            if (m.get("pf_approval_required") is not False
                    and str(m.get("pf_approval_status") or "").lower() != "approved"):
                issues.append(_issue(
                    "HIGHER_PF_APPROVAL_PENDING", "error",
                    f"Higher PF approval status is '{m.get('pf_approval_status') or 'pending'}' — management approval required.",
                    "Set Approval Status = Approved on the Employee Master (or untick Management Approval Required)."))
            _hf, _ht = str(m.get("higher_pf_from") or "")[:7], str(m.get("higher_pf_to") or "")[:7]
            if _hf and _ht and _hf > _ht:
                issues.append(_issue(
                    "HIGHER_PF_INVALID_DATES", "error",
                    f"Higher PF Effective From ({_hf}) is AFTER Effective To ({_ht}).",
                    "Correct the effective dates on the Employee Master."))
            elif month and ((_hf and month[:7] < _hf) or (_ht and month[:7] > _ht)):
                issues.append(_issue(
                    "HIGHER_PF_OUT_OF_WINDOW", "warning",
                    f"Higher PF window ({_hf or '…'} → {_ht or '…'}) does not cover {month} — statutory ceiling was used.",
                    "Extend the effective window or ignore if intentional."))
            if _num(m.get("higher_pf_wage")) <= 0:
                issues.append(_issue(
                    "HIGHER_PF_WAGE_BLANK", "warning",
                    "Higher PF Wage is blank — the ACTUAL PF wage was used instead.",
                    "Fill Higher PF Wage on the Employee Master if a fixed approved wage applies."))
        _vpf_on = _pft == "vpf" or bool(m.get("vpf_enabled") or r.get("vpf_enabled"))
        if _vpf_on:
            if stat.get("allow_vpf") is False:
                issues.append(_issue(
                    "VPF_NOT_ALLOWED", "error",
                    "VPF selected but the company policy does not allow VPF.",
                    "Enable 'Allow VPF' in Standard Compliance Settings, or remove VPF from the Employee Master."))
            _vp, _va = _num(m.get("vpf_percent")), _num(m.get("vpf_amount"))
            if _vp < 0 or _va < 0:
                issues.append(_issue(
                    "VPF_NEGATIVE", "error",
                    "VPF percentage / amount is negative.",
                    "Correct the VPF value on the Employee Master."))
            _vlim = _num(stat.get("vpf_max_percent"))
            if _vlim > 0 and _vp > _vlim + 0.001:
                issues.append(_issue(
                    "VPF_ABOVE_LIMIT", "error",
                    f"VPF {_vp:g}% exceeds the company limit of {_vlim:g}%.",
                    f"Reduce the employee's VPF % to {_vlim:g}% or raise the limit in Compliance Settings."))
            if _pft == "vpf" and _vp <= 0 and _va <= 0:
                issues.append(_issue(
                    "VPF_VALUE_MISSING", "warning",
                    "Contribution Type is VPF but neither VPF % nor a fixed amount is set — no VPF was deducted.",
                    "Set VPF Percentage or VPF Fixed Amount on the Employee Master."))
        # PF Basic blank on the master (user rule: PF intentionally skipped).
        _m_pf_basic = _num(m.get("pf_basic"))
        if (m and m.get("pf_applicable") is not False
                and not m.get("excluded_employee")
                and _m_pf_basic <= 0 and gross > 0 and pd > 0):
            issues.append(_issue(
                "PF_MISSING_BASIC", "warning",
                "PF Basic Salary is blank on the Employee Master — NO PF is deducted (your standing rule).",
                "Fill PF Basic Salary if PF is intended, or tick Excluded Employee / PF Applicable = No."))
        # Higher pension mismatch.
        if r.get("higher_pension") and (r.get("eps_disabled") or (
                pf_emp > 0 and _num(r.get("pf_employer_eps")) <= 0)):
            issues.append(_issue(
                "PF_HIGHER_PENSION_MISMATCH", "warning",
                "Higher Pension is ticked but EPS contribution is 0 / EPS is disabled.",
                "Untick either Higher Pension or EPS Disable on the Employee Master, then re-run."))
        # Master salary changed AFTER this calculation.
        if m and pf_emp > 0 and abs(_m_pf_basic - _num(r.get("pf_basic"))) > 0.5:
            issues.append(_issue(
                "PF_SALARY_CHANGED", "warning",
                f"PF Basic changed on the Employee Master (₹{_num(r.get('pf_basic')):,.0f} → ₹{_m_pf_basic:,.0f}) after this calculation.",
                "Re-run Salary Process so PF follows the updated master."))

        # ---------------- ESIC checks ----------------
        if r.get("esic_applicable") and gross > 0 and esic_emp <= 0:
            issues.append(_issue(
                "ESIC_ZERO", "error",
                "ESIC is applicable but the ESIC deduction is 0.",
                "Re-run Salary Process, or check the ESIC rates / Head Mapping."))
        if esic_emp > 0 and not ip:
            issues.append(_issue(
                "ESIC_MISSING_IP", "error",
                "ESIC deducted but the IP Number is missing on the Employee Master.",
                "Fill the ESI IP No. in Employee Master → Statutory, or generate via the ESIC portal."))
        if ip and len(ip_owners.get(ip, [])) > 1:
            others = [n for n in ip_owners[ip] if n != name][:2]
            issues.append(_issue(
                "ESIC_DUP_IP", "error",
                f"ESI IP No. {ip} is shared with {', '.join(others) or 'another employee'}.",
                "Each employee must have a unique IP number — correct the Employee Master."))
        if esic_base < 0:
            issues.append(_issue(
                "ESIC_WAGE_INVALID", "error",
                f"ESIC wage base is negative (₹{esic_base:,.0f}).",
                "Re-run Salary Process for this month."))
        elif esic_emp > 0:
            _ge = gross + _num(r.get("ot_pay"))
            if _ge > 0 and esic_base > _ge + 0.5:
                issues.append(_issue(
                    "ESIC_WAGE_INVALID", "warning",
                    f"ESIC wage base ₹{esic_base:,.0f} exceeds the Gross Earning ₹{_ge:,.0f}.",
                    "Check the Head Mapping / proration — a base above gross is unusual."))
        # Above ceiling but still calculated (only wrong when the rule says
        # coverage must stop above the ceiling).
        _elig_basic = _num(r.get("compliance_basic")) or _num(r.get("basic_master"))
        if esic_emp > 0 and disable_above and _elig_basic > esic_ceiling + 0.5:
            issues.append(_issue(
                "ESIC_ABOVE_CEILING", "error",
                f"ESIC deducted although Basic ₹{_elig_basic:,.0f} is above the ceiling ₹{esic_ceiling:,.0f}.",
                "Re-run Salary Process, or disable the above-ceiling rule in Compliance Settings if coverage must continue."))
        # Incorrectly EXCLUDED: master says Not Applicable but within ceiling.
        if (m and m.get("esic_applicable") is False and gross > 0
                and 0 < _elig_basic <= esic_ceiling):
            issues.append(_issue(
                "ESIC_WRONG_EXCLUSION", "warning",
                f"ESIC Applicable = No although Basic ₹{_elig_basic:,.0f} is within the ceiling.",
                "Confirm the exclusion or set ESIC Applicable = Yes on the Employee Master."))
        # Exit before the salary month but still paid/deducted.
        for _exit_field in ("exit_date", "esic_exit_date"):
            _ex = str((m.get(_exit_field) if m else "") or "")[:7]
            if month and _ex and _ex < month[:7] and esic_emp > 0:
                issues.append(_issue(
                    "ESIC_EXIT_BEFORE_MONTH", "error",
                    f"{'Exit Date' if _exit_field == 'exit_date' else 'ESIC Exit Date'} ({m.get(_exit_field)}) is before {month} but ESIC was deducted.",
                    "Remove the employee from the run or correct the exit date, then re-run."))
                break

        # Iter 423 — Firm Master policy gate: drop PF / ESIC findings when
        # the firm disabled that statute (nothing to validate).
        if not firm_pf_on:
            issues = [i for i in issues if not i["code"].startswith(
                ("PF_", "HIGHER_PF", "VPF_"))]
        if not firm_esi_on:
            issues = [i for i in issues if not i["code"].startswith("ESIC_")]

        if issues:
            errors += sum(1 for i in issues if i["level"] == "error")
            warnings += sum(1 for i in issues if i["level"] == "warning")
            out_rows.append({
                "user_id": uid,
                "employee_code": r.get("employee_code"),
                "name": name,
                "gross_paid": gross,
                "present_days": pd,
                "issues": issues,
            })

    errors += sum(1 for i in global_issues if i["level"] == "error")
    warnings += sum(1 for i in global_issues if i["level"] == "warning")
    return {
        "ok": errors == 0,
        "errors_count": errors,
        "warnings_count": warnings,
        "employees_flagged": len(out_rows),
        "employees_total": len(rows),
        "rows": out_rows,
        "global_issues": global_issues,
        "checked_at": now_iso(),
    }


@router.get("/admin/compliance-salary-runs/{run_id}/validate")
async def validate_run_endpoint(
    run_id: str, authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    run = await db.compliance_salary_runs.find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Compliance salary run not found")
    if admin["role"] == "company_admin" and run.get("company_id") != admin.get("company_id"):
        raise HTTPException(status_code=403, detail="Not authorised for this run")
    return await validate_compliance_run(run)


# ---------------------------------------------------------------------------
# Iter 388 (Phase 4) — PF & ESIC AUDIT DASHBOARD + "View Calculation".
# Merges the run rows with the validation issues into one colour-coded
# per-employee audit view (Green OK / Yellow Warning / Red Error).
# ---------------------------------------------------------------------------
@router.get("/admin/compliance-salary-runs/{run_id}/audit-dashboard")
async def audit_dashboard(
    run_id: str, authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    run = await db.compliance_salary_runs.find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Compliance salary run not found")
    if admin["role"] == "company_admin" and run.get("company_id") != admin.get("company_id"):
        raise HTTPException(status_code=403, detail="Not authorised for this run")

    validation = await validate_compliance_run(run)
    issues_by_user: Dict[str, List[Dict[str, str]]] = {
        r["user_id"]: r["issues"] for r in validation["rows"]}
    stat = run.get("statutory_effective") or {}

    rows_out: List[Dict[str, Any]] = []
    counts = {"ok": 0, "warning": 0, "error": 0}
    for r in (run.get("rows") or []):
        issues = issues_by_user.get(r.get("user_id"), [])
        if any(i["level"] == "error" for i in issues):
            status = "error"
        elif issues:
            status = "warning"
        else:
            status = "ok"
        counts[status] += 1
        reason = (issues[0]["message"] if issues
                  else (r.get("pf_reason") or r.get("esic_reason") or "OK"))
        rows_out.append({
            "user_id": r.get("user_id"),
            "employee_code": r.get("employee_code"),
            "name": r.get("name"),
            "gross_paid": _num(r.get("gross_paid")),
            "present_days": _num(r.get("present_days")),
            "pf_wages": _num(r.get("pf_wages")),
            "pf_employee": _num(r.get("pf_employee")),
            "pf_employer_epf": _num(r.get("pf_employer_epf")),
            "pf_employer_eps": _num(r.get("pf_employer_eps")),
            "esic_wage_base": _num(r.get("esic_wage_base")),
            "esic_employee": _num(r.get("esic_employee")),
            "esic_employer": _num(r.get("esic_employer")),
            "status": status,
            "reason": reason,
            "issues": issues,
            "pf_reason": r.get("pf_reason") or "",
            "esic_reason": r.get("esic_reason") or "",
            "calc_snapshot": r.get("calc_snapshot") or None,
        })

    return {
        "run_id": run_id,
        "month": run.get("month"),
        "company_id": run.get("company_id"),
        "finalized": bool(run.get("finalized")),
        "lock_validation": run.get("lock_validation"),
        "rule_version": str(stat.get("rule_version") or ""),
        "summary": {**counts, "total": len(rows_out),
                    "errors_count": validation["errors_count"],
                    "warnings_count": validation["warnings_count"]},
        "global_issues": validation["global_issues"],
        "rows": rows_out,
    }


async def write_monthly_snapshot(run: Dict[str, Any], admin: Dict[str, Any],
                                 lock_validation: Dict[str, Any]) -> str:
    """Iter 388 (Phase 4) — APPEND-ONLY monthly statutory snapshot written
    at Salary Lock. Historical documents are never modified."""
    import uuid
    stat = run.get("statutory_effective") or {}
    snap_rows = []
    for r in (run.get("rows") or []):
        snap_rows.append({
            "user_id": r.get("user_id"),
            "employee_code": r.get("employee_code"),
            "name": r.get("name"),
            "uan_no": r.get("uan_no"),
            "esi_ip_no": r.get("esi_ip_no"),
            "present_days": _num(r.get("present_days")),
            "gross_paid": _num(r.get("gross_paid")),
            "pf_wages": _num(r.get("pf_wages")),
            "pf_employee": _num(r.get("pf_employee")),
            "pf_employer_epf": _num(r.get("pf_employer_epf")),
            "pf_employer_eps": _num(r.get("pf_employer_eps")),
            "esic_wage_base": _num(r.get("esic_wage_base")),
            "esic_employee": _num(r.get("esic_employee")),
            "esic_employer": _num(r.get("esic_employer")),
            # Iter 408 — PF contribution type snapshot (never changes after lock).
            "pf_contribution_type": r.get("pf_contribution_type") or "statutory",
            "pf_higher_active": bool(r.get("pf_higher_active")),
            "pf_ceiling_applied": bool(r.get("pf_ceiling_applied")),
            "vpf_amount": _num(r.get("vpf_amount")),
        })
    doc = {
        "snapshot_id": f"csnap_{uuid.uuid4().hex[:12]}",
        "run_id": run.get("run_id"),
        "month": run.get("month"),
        "company_id": run.get("company_id"),
        "rule_version": str(stat.get("rule_version") or ""),
        "calculated_at": run.get("generated_at"),
        "locked_at": now_iso(),
        "locked_by": admin.get("user_id"),
        "locked_by_name": admin.get("name") or admin.get("email") or "",
        "lock_validation": lock_validation,
        "rows": snap_rows,
    }
    await db.compliance_monthly_snapshots.insert_one(dict(doc))
    return doc["snapshot_id"]


# ---------------------------------------------------------------------------
# Iter 388 (Phase 5) — PF/ESIC AUDIT & EXCEPTION REPORTS (Excel + PDF).
#   kind = pf          → PF Audit Report (all rows)
#          esic        → ESIC Audit Report (all rows)
#          exceptions  → PF & ESIC Exception / Salary-Lock Error Report
#   Plus a master-level Missing UAN / Missing IP report.
# ---------------------------------------------------------------------------
import io
import re as _re_mod

from fastapi import Query, Response

_re_sheet = _re_mod.compile(r"[\[\]*?/\\:]")


def _xlsx_bytes(title: str, headers: List[str], data: List[List[Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    # openpyxl forbids [ ] * ? / \ : in sheet titles.
    ws.title = _re_sheet.sub(" ", title)[:31]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F172A")
    for row in data:
        ws.append(row)
    for i, h in enumerate(headers, start=1):
        width = max([len(str(h))] + [len(str(r[i - 1])) for r in data[:200]] or [10])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(46, width + 2)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_bytes(title: str, subtitle: str, headers: List[str],
               data: List[List[Any]], level_col: Optional[int] = None) -> bytes:
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=8 * mm, rightMargin=8 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=6.8, leading=8.2)
    head = ParagraphStyle("head", parent=styles["Normal"], fontSize=7, leading=8.5,
                          textColor=rl.white, fontName="Helvetica-Bold")
    story = [Paragraph(title, styles["Title"]),
             Paragraph(subtitle, styles["Normal"]), Spacer(1, 4 * mm)]
    tbl_data = [[Paragraph(str(h), head) for h in headers]]
    for r in data:
        tbl_data.append([Paragraph(str(v if v is not None else ""), cell) for v in r])
    t = Table(tbl_data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#0F172A")),
        ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl.white, rl.HexColor("#F8FAFC")]),
    ]
    if level_col is not None:
        for i, r in enumerate(data, start=1):
            lv = str(r[level_col]).lower()
            if "error" in lv:
                style.append(("BACKGROUND", (0, i), (-1, i), rl.HexColor("#FEE2E2")))
            elif "warn" in lv:
                style.append(("BACKGROUND", (0, i), (-1, i), rl.HexColor("#FEF3C7")))
    t.setStyle(TableStyle(style))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def _file_response(content: bytes, filename: str, fmt: str) -> Response:
    mt = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
          if fmt == "xlsx" else "application/pdf")
    return Response(content=content, media_type=mt, headers={
        "Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/admin/compliance-salary-runs/{run_id}/audit-export")
async def audit_export(
    run_id: str,
    kind: str = Query("pf", pattern="^(pf|esic|exceptions)$"),
    format: str = Query("xlsx", pattern="^(xlsx|pdf)$"),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    run = await db.compliance_salary_runs.find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Compliance salary run not found")
    if admin["role"] == "company_admin" and run.get("company_id") != admin.get("company_id"):
        raise HTTPException(status_code=403, detail="Not authorised for this run")

    validation = await validate_compliance_run(run)
    issues_by_user = {r["user_id"]: r["issues"] for r in validation["rows"]}
    month = str(run.get("month") or "")
    rows = run.get("rows") or []

    def _status(uid: str) -> str:
        iss = issues_by_user.get(uid, [])
        if any(i["level"] == "error" for i in iss):
            return "ERROR"
        return "WARNING" if iss else "OK"

    def _reason(uid: str, fallback: str) -> str:
        iss = issues_by_user.get(uid, [])
        return "; ".join(i["message"] for i in iss) if iss else (fallback or "OK")

    if kind == "pf":
        headers = ["Code", "Employee Name", "UAN", "Gross", "Days", "PF Wage",
                   "EE PF", "ER EPF", "ER EPS", "Status", "Reason"]
        data = [[r.get("employee_code") or "", r.get("name") or "",
                 r.get("uan_no") or "", _num(r.get("gross_paid")),
                 _num(r.get("present_days")), _num(r.get("pf_wages")),
                 _num(r.get("pf_employee")), _num(r.get("pf_employer_epf")),
                 _num(r.get("pf_employer_eps")), _status(r.get("user_id")),
                 _reason(r.get("user_id"), r.get("pf_reason") or "")]
                for r in rows]
        title, level_col = f"PF Audit Report — {month}", 9
    elif kind == "esic":
        headers = ["Code", "Employee Name", "ESI IP No", "Gross", "Days",
                   "ESIC Wage", "EE ESIC", "ER ESIC", "Status", "Reason"]
        data = [[r.get("employee_code") or "", r.get("name") or "",
                 r.get("esi_ip_no") or "", _num(r.get("gross_paid")),
                 _num(r.get("present_days")), _num(r.get("esic_wage_base")),
                 _num(r.get("esic_employee")), _num(r.get("esic_employer")),
                 _status(r.get("user_id")),
                 _reason(r.get("user_id"), r.get("esic_reason") or "")]
                for r in rows]
        title, level_col = f"ESIC Audit Report — {month}", 8
    else:  # exceptions == Salary Lock Error Report
        headers = ["Code", "Employee Name", "Level", "Issue", "Message", "Suggested Fix"]
        data = []
        for g in validation["global_issues"]:
            data.append(["—", "GLOBAL", g["level"].upper(), g["code"],
                         g["message"], g["suggestion"]])
        for vr in validation["rows"]:
            for i in vr["issues"]:
                data.append([vr.get("employee_code") or "", vr.get("name") or "",
                             i["level"].upper(), i["code"], i["message"], i["suggestion"]])
        title, level_col = f"PF & ESIC Exception / Lock Error Report — {month}", 2

    fname = f"{kind}_audit_{month}.{format}"
    subtitle = (f"Run {run_id} · {len(rows)} employees · "
                f"{validation['errors_count']} error(s), {validation['warnings_count']} warning(s)"
                f"{' · LOCKED' if run.get('finalized') else ''}")
    if format == "xlsx":
        return _file_response(_xlsx_bytes(title, headers, data), fname, "xlsx")
    return _file_response(_pdf_bytes(title, subtitle, headers, data, level_col), fname, "pdf")


@router.get("/admin/compliance-reports/missing-ids")
async def missing_ids_report(
    which: str = Query("uan", pattern="^(uan|ip)$"),
    company_id: Optional[str] = Query(None),
    format: str = Query("xlsx", pattern="^(xlsx|pdf)$"),
    authorization: Optional[str] = Header(None),
):
    """Iter 388 (Phase 5) — Missing UAN / Missing ESI IP Number report
    straight from the LIVE Employee Master (active employees only)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    q: Dict[str, Any] = {"role": "employee", "disabled": {"$ne": True}}
    if admin["role"] == "company_admin":
        q["company_id"] = admin.get("company_id")
    elif company_id:
        q["company_id"] = company_id
    field = "uan_no" if which == "uan" else "esi_ip_no"
    q["$or"] = [{field: {"$exists": False}}, {field: None}, {field: ""}]
    emps = await db.users.find(
        q, {"_id": 0, "employee_code": 1, "name": 1, "designation": 1,
            "department": 1, "doj": 1, "phone": 1, "company_id": 1}
    ).sort("employee_code", 1).to_list(5000)

    label = "UAN" if which == "uan" else "ESI IP Number"
    headers = ["Code", "Employee Name", "Designation", "Department", "DOJ", "Mobile"]
    data = [[e.get("employee_code") or "", e.get("name") or "",
             e.get("designation") or "", e.get("department") or "",
             str(e.get("doj") or "")[:10], e.get("phone") or ""] for e in emps]
    title = f"Missing {label} Report"
    subtitle = f"{len(data)} active employee(s) without a {label}"
    fname = f"missing_{which}.{format}"
    if format == "xlsx":
        return _file_response(_xlsx_bytes(title, headers, data), fname, "xlsx")
    return _file_response(_pdf_bytes(title, subtitle, headers, data), fname, "pdf")


# ---------------------------------------------------------------------------
# Iter 388 (Phase 6) — AI COMPLIANCE ASSISTANT.
# Per-employee plain-language explanation of WHY PF/ESIC was (or was not)
# calculated, which heads/rules/ceilings applied, which validations failed
# and the recommended corrective action. Powered by the Emergent LLM key
# (same pattern as ai_salary_compliance.py). The AI NEVER recalculates —
# it only explains the engine's stored snapshot.
# ---------------------------------------------------------------------------
import json as _json
import os as _os
import re as _re
import uuid as _uuid

_EMERGENT_LLM_KEY = _os.environ.get("EMERGENT_LLM_KEY", "")


@router.post("/admin/compliance-salary-runs/{run_id}/ai-explain/{user_id}")
async def ai_explain_employee(
    run_id: str, user_id: str,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    run = await db.compliance_salary_runs.find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Compliance salary run not found")
    if admin["role"] == "company_admin" and run.get("company_id") != admin.get("company_id"):
        raise HTTPException(status_code=403, detail="Not authorised for this run")
    row = next((r for r in (run.get("rows") or []) if r.get("user_id") == user_id), None)
    if not row:
        raise HTTPException(status_code=404, detail="Employee not found in this run")
    if not _EMERGENT_LLM_KEY:
        raise HTTPException(status_code=503, detail="AI key not configured")

    validation = await validate_compliance_run(run)
    issues = next((r["issues"] for r in validation["rows"]
                   if r["user_id"] == user_id), [])
    facts = {
        "month": run.get("month"),
        "name": row.get("name"),
        "present_days": row.get("present_days"),
        "gross_earned": row.get("gross_paid"),
        "pf": {
            "applicable": row.get("pf_applicable"),
            "wages": row.get("pf_wages"),
            "employee": row.get("pf_employee"),
            "employer_epf": row.get("pf_employer_epf"),
            "employer_eps": row.get("pf_employer_eps"),
            "engine_reason": row.get("pf_reason"),
        },
        "esic": {
            "applicable": row.get("esic_applicable"),
            "wage_base": row.get("esic_wage_base"),
            "employee": row.get("esic_employee"),
            "employer": row.get("esic_employer"),
            "engine_reason": row.get("esic_reason"),
        },
        "flags": {k: bool(row.get(k)) for k in (
            "higher_pension", "intl_worker", "excluded_employee",
            "esic_temp_exempt", "eps_disabled")},
        "calc_snapshot": row.get("calc_snapshot"),
        "validation_issues": issues,
    }
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=_EMERGENT_LLM_KEY,
            session_id=f"pf-esic-ai-{_uuid.uuid4().hex[:8]}",
            system_message=(
                "You are a Senior Indian Payroll & EPF/ESIC Compliance Expert. "
                "The amounts are ALREADY calculated by the firm's statutory "
                "engine — NEVER recalculate or change any number; only EXPLAIN "
                "the given facts. Answer in SHORT plain text (no markdown) with "
                "exactly these numbered sections: 1. PF — WHY (why PF was or "
                "was not calculated, which wage base, ceiling and rule applied), "
                "2. ESIC — WHY (same for ESIC), 3. SALARY HEADS CONSIDERED "
                "(which heads counted as PF/ESIC wages per the mapping), "
                "4. VALIDATION (which checks failed, if any), 5. RECOMMENDED "
                "ACTION (one or two concrete corrective steps, or 'No action "
                "needed'). Keep it under 220 words."),
        ).with_model("openai", "gpt-5.4-mini")
        resp = await chat.send_message(UserMessage(
            text="Employee statutory facts (JSON):\n" + _json.dumps(facts, default=str)))
        text = _re.sub(r"[*#`|]", "", str(resp)).strip()
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"AI explanation failed: {str(e)[:160]}")
    return {"ok": True, "explanation": text, "name": row.get("name")}


# ---------------------------------------------------------------------------
# Iter 389 (user request) — PRINTABLE A4 CALCULATION EXPLANATION SHEET.
# One-page portrait PDF per employee for PF/ESIC inspector queries:
# every figure of the stored snapshot + rules applied + validation result.
# ---------------------------------------------------------------------------
@router.get("/admin/compliance-salary-runs/{run_id}/calc-sheet/{user_id}")
async def calc_explanation_sheet(
    run_id: str, user_id: str,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    run = await db.compliance_salary_runs.find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Compliance salary run not found")
    if admin["role"] == "company_admin" and run.get("company_id") != admin.get("company_id"):
        raise HTTPException(status_code=403, detail="Not authorised for this run")
    row = next((r for r in (run.get("rows") or []) if r.get("user_id") == user_id), None)
    if not row:
        raise HTTPException(status_code=404, detail="Employee not found in this run")

    validation = await validate_compliance_run(run)
    issues = next((r["issues"] for r in validation["rows"]
                   if r["user_id"] == user_id), [])
    company = await db.companies.find_one(
        {"company_id": run.get("company_id")}, {"_id": 0, "name": 1}) or {}

    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)

    s = row.get("calc_snapshot") or {}
    pf = s.get("pf") or {}
    es = s.get("esic") or {}
    heads = s.get("heads_considered") or {}
    month = str(run.get("month") or "")

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=13, leading=16,
                        spaceAfter=1)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=8.5,
                         textColor=rl.HexColor("#475569"), alignment=1)
    sec = ParagraphStyle("sec", parent=styles["Normal"], fontSize=9.5,
                         fontName="Helvetica-Bold",
                         textColor=rl.HexColor("#0F172A"), spaceBefore=7,
                         spaceAfter=2)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=10)
    ital = ParagraphStyle("ital", parent=cell, fontName="Helvetica-Oblique",
                          textColor=rl.HexColor("#475569"))

    def _rs(v):
        return f"Rs. {_num(v):,.2f}"

    def _kv_table(pairs, col1=62 * mm):
        t = Table([[Paragraph(str(a), cell), Paragraph(str(b), cell)]
                   for a, b in pairs], colWidths=[col1, None])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#CBD5E1")),
            ("BACKGROUND", (0, 0), (0, -1), rl.HexColor("#F1F5F9")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        return t

    story = [
        Paragraph(str(company.get("name") or "S.K. Sharma & Co."), h1),
        Paragraph("PF & ESIC CALCULATION EXPLANATION SHEET — "
                  f"Salary Month {month}", sub),
        Spacer(1, 3 * mm),
        _kv_table([
            ("Employee", f"{row.get('name') or ''}  "
                         f"(Code {row.get('employee_code') or '—'})"),
            ("UAN / ESI IP No.", f"{row.get('uan_no') or '—'}  /  "
                                 f"{row.get('esi_ip_no') or '—'}"),
            ("Paid Days / Gross Earned",
             f"{_num(row.get('present_days')):g} days  /  {_rs(row.get('gross_paid'))}"),
            ("Rule Version / Wage Definition Rule",
             f"{s.get('rule_version') or '—'}  /  "
             + ("ON — max(Basic, floor% of Gross)"
                if s.get("wage_definition_rule") is not False else "OFF — Head Mapping")),
            ("Proration (PF / ESIC)",
             f"{s.get('pf_proration_method') or 'calendar_days'} / "
             f"{s.get('esic_proration_method') or 'calendar_days'}"),
        ]),
        Paragraph("1. SALARY HEADS CONSIDERED", sec),
    ]
    head_rows = [["Head", "Earned Amount", "PF Wage?", "ESIC Wage?"]]
    for k in ("basic", "hra", "conveyance", "medical", "special", "others", "ot"):
        h = heads.get(k) or {}
        head_rows.append([k.upper(), _rs(h.get("amount")),
                          "Yes" if h.get("pf_wage") else "No",
                          "Yes" if h.get("esic_wage") else "No"])
    ht = Table(head_rows, colWidths=[38 * mm, 42 * mm, 30 * mm, 30 * mm])
    ht.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#CBD5E1")),
        ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story += [ht, Paragraph("2. PROVIDENT FUND (EPF)", sec), _kv_table([
        ("PF Basic (Employee Master)", _rs(pf.get("pf_basic_master"))),
        ("PF Basic after proration", _rs(pf.get("pf_basic_prorated"))),
        ("Wage Base (after wage rule)", _rs(pf.get("wage_base"))),
        ("Statutory Ceiling", _rs(pf.get("ceiling"))),
        ("PF Wages (final)", _rs(row.get("pf_wages"))),
        (f"Employee PF @ {pf.get('rate_employee', 12):g}%", _rs(row.get("pf_employee"))),
        (f"Employer EPF @ {pf.get('rate_epf', 3.67):g}%", _rs(row.get("pf_employer_epf"))),
        (f"Employer EPS @ {pf.get('rate_eps', 8.33):g}%", _rs(row.get("pf_employer_eps"))),
        ("Rounding", str(pf.get("rounding") or "nearest")),
    ]), Paragraph("Engine reason: " + (row.get("pf_reason") or "—"), ital),
        Paragraph("3. ESIC", sec), _kv_table([
        ("Eligibility Basic", _rs(es.get("eligibility_basic"))),
        ("ESIC Ceiling", _rs(es.get("ceiling"))),
        ("ESIC Wage Base", _rs(row.get("esic_wage_base"))),
        (f"Employee ESIC @ {es.get('rate_employee', 0.75):g}%", _rs(row.get("esic_employee"))),
        (f"Employer ESIC @ {es.get('rate_employer', 3.25):g}%", _rs(row.get("esic_employer"))),
        ("Rounding", str(es.get("rounding") or "ceil")),
    ]), Paragraph("Engine reason: " + (row.get("esic_reason") or "—"), ital),
        Paragraph("4. VALIDATION RESULT", sec)]
    if not issues:
        story.append(Paragraph("All PF/ESIC checks passed — no issues.", cell))
    else:
        vt_rows = [["Level", "Check", "Finding", "Suggested Fix"]]
        for i in issues:
            vt_rows.append([i["level"].upper(), i["code"], i["message"], i["suggestion"]])
        vt = Table([[Paragraph(str(c), cell) for c in r] for r in vt_rows],
                   colWidths=[16 * mm, 34 * mm, 65 * mm, None])
        st = [("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#CBD5E1")),
              ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#0F172A")),
              ("VALIGN", (0, 0), (-1, -1), "TOP")]
        for idx, i in enumerate(issues, start=1):
            st.append(("BACKGROUND", (0, idx), (-1, idx),
                       rl.HexColor("#FEE2E2" if i["level"] == "error" else "#FEF3C7")))
        vt.setStyle(TableStyle(st))
        story.append(vt)
    story += [Spacer(1, 5 * mm),
              Paragraph(f"Computer-generated explanation sheet · Run {run_id}"
                        f"{' · LOCKED' if run.get('finalized') else ''} · "
                        f"Generated {now_iso()[:16].replace('T', ' ')}", ital)]

    buf = io.BytesIO()
    SimpleDocTemplate(buf, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
                      topMargin=12 * mm, bottomMargin=12 * mm).build(story)
    fname = f"calc_sheet_{row.get('employee_code') or user_id}_{month}.pdf"
    return _file_response(buf.getvalue(), fname, "pdf")
