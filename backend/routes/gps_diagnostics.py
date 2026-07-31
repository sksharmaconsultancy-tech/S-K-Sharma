"""Iter 417 — GPS DIAGNOSTICS module (Smart Punch GPS revamp).

POST /api/gps-diagnostics          — employees' devices log every GPS
                                     attempt (success / weak / failed).
GET  /api/admin/gps-diagnostics    — admin dashboard feed (filters).
GET  /api/admin/gps-diagnostics.xlsx — Excel export.
"""
import io
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, Optional

from fastapi import APIRouter, Header, HTTPException, Query
from pydantic import BaseModel

from server import (  # noqa: E402
    db,
    get_user_from_token,
    now_iso,
    require_role,
)

router = APIRouter(prefix="/api")
api = router


class GpsDiagnosticIn(BaseModel):
    outcome: str                       # success | weak | failed
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    accuracy: Optional[float] = None
    retry_count: Optional[int] = 0
    failure_reason: Optional[str] = None
    permission_status: Optional[str] = None
    gps_enabled: Optional[bool] = None
    mock_location: Optional[bool] = None
    device: Optional[str] = None
    platform: Optional[str] = None
    network_status: Optional[str] = None
    gps_time: Optional[str] = None
    context: Optional[str] = None      # punch | punch_warm | warmup ...


@api.post("/gps-diagnostics")
async def log_gps_diagnostic(
    payload: GpsDiagnosticIn,
    authorization: Optional[str] = Header(None),
):
    """Device-side GPS attempt log. Any authenticated user."""
    user = await get_user_from_token(authorization)
    doc = {
        "diag_id": f"gps_{uuid.uuid4().hex[:12]}",
        "user_id": user.get("user_id"),
        "name": user.get("name"),
        "employee_code": user.get("employee_code"),
        "company_id": user.get("company_id"),
        "outcome": (payload.outcome or "failed")[:20],
        "latitude": payload.latitude,
        "longitude": payload.longitude,
        "accuracy": payload.accuracy,
        "retry_count": int(payload.retry_count or 0),
        "failure_reason": (payload.failure_reason or None),
        "permission_status": (payload.permission_status or None),
        "gps_enabled": payload.gps_enabled,
        "mock_location": bool(payload.mock_location or False),
        "device": (payload.device or "")[:300],
        "platform": (payload.platform or "")[:20],
        "network_status": (payload.network_status or None),
        "gps_time": (payload.gps_time or None),
        "context": (payload.context or None),
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "created_at": now_iso(),
    }
    await db.gps_diagnostics.insert_one(doc)
    return {"ok": True}


def _diag_query(
    admin: dict, company_id: Optional[str], date: Optional[str],
    outcome: Optional[str],
) -> Dict[str, Any]:
    q: Dict[str, Any] = {}
    if admin.get("role") == "company_admin":
        q["company_id"] = admin.get("company_id")
    elif company_id:
        q["company_id"] = company_id
    if date:
        q["date"] = date
    if outcome:
        q["outcome"] = outcome
    return q


@api.get("/admin/gps-diagnostics")
async def list_gps_diagnostics(
    company_id: Optional[str] = Query(None),
    date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    outcome: Optional[str] = Query(None),
    limit: int = Query(300, ge=1, le=1000),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    q = _diag_query(admin, company_id, date, outcome)
    rows = await db.gps_diagnostics.find(q, {"_id": 0}) \
        .sort("created_at", -1).to_list(limit)
    # Summary counts for the header cards.
    counts: Dict[str, int] = {"success": 0, "weak": 0, "failed": 0}
    async for row in db.gps_diagnostics.aggregate([
        {"$match": q},
        {"$group": {"_id": "$outcome", "n": {"$sum": 1}}},
    ]):
        if row["_id"] in counts:
            counts[row["_id"]] = row["n"]
    total = sum(counts.values()) or 1
    return {
        "rows": rows,
        "counts": counts,
        "success_rate": round(100.0 * (counts["success"] + counts["weak"]) / total, 1),
    }


@api.get("/admin/gps-diagnostics.xlsx")
async def export_gps_diagnostics(
    company_id: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    outcome: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    q = _diag_query(admin, company_id, date, outcome)
    rows = await db.gps_diagnostics.find(q, {"_id": 0}) \
        .sort("created_at", -1).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "GPS Diagnostics"
    headers = ["Time", "Employee", "Code", "Status", "Accuracy (m)",
               "Retries", "Permission", "GPS On", "Mock GPS", "Network",
               "Reason", "Platform", "Device"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F766E")
    fills = {
        "success": PatternFill("solid", fgColor="DCFCE7"),
        "weak": PatternFill("solid", fgColor="FEF9C3"),
        "failed": PatternFill("solid", fgColor="FEE2E2"),
    }
    for r in rows:
        ws.append([
            (r.get("created_at") or "")[:19].replace("T", " "),
            r.get("name") or "", r.get("employee_code") or "",
            (r.get("outcome") or "").upper(),
            r.get("accuracy"), r.get("retry_count"),
            r.get("permission_status") or "",
            "" if r.get("gps_enabled") is None else ("Yes" if r.get("gps_enabled") else "No"),
            "Yes" if r.get("mock_location") else "",
            r.get("network_status") or "",
            r.get("failure_reason") or "",
            r.get("platform") or "",
            (r.get("device") or "")[:80],
        ])
        f = fills.get(r.get("outcome") or "")
        if f:
            ws.cell(row=ws.max_row, column=4).fill = f
    widths = [19, 24, 8, 10, 12, 8, 12, 8, 9, 9, 22, 9, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="GPS_Diagnostics.xlsx"'},
    )
