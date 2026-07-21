"""Iter 161 — PF Reports (user request, formats per uploaded EPFO samples).

Replaces the old "P.F. Contribution Sheet" nav entry with a "PF Reports"
hub offering, for a MANUALLY selected month/year period (From → To):

  * PF Challan Report  — PDF (EPFO provisional-challan layout: A/c 01, 02,
    10, 21, 22 columns × EE/ER/Admin/7Q/14B rows) and Excel.
  * PF ECR             — PDF + Excel (EPFO "Return Statement (Regular
    Return)" layout: UAN / name / Gross-EPF-EPS-EDLI wages /
    EE-EPS-ER-Refunds / NCP days).

Data source: the LATEST compliance salary run of each month in the range.
A/c 2 / 21 / 22 use the Standard Compliance Settings percentages effective
for each month (with firm overrides).
"""
import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException
from fastapi.responses import Response

from server import db, get_user_from_token, require_role  # noqa: E402

router = APIRouter(prefix="/api/admin/pf-reports", tags=["pf-reports"])


# --------------------------------------------------------------- helpers
def _months_range(m_from: str, m_to: str) -> List[str]:
    try:
        y1, mo1 = int(m_from[:4]), int(m_from[5:7])
        y2, mo2 = int(m_to[:4]), int(m_to[5:7])
    except Exception:
        raise HTTPException(status_code=400, detail="from/to must be YYYY-MM")
    if (y1, mo1) > (y2, mo2):
        raise HTTPException(status_code=400, detail="'From' month is after 'To' month")
    out = []
    y, mo = y1, mo1
    while (y, mo) <= (y2, mo2):
        out.append(f"{y:04d}-{mo:02d}")
        mo += 1
        if mo > 12:
            mo, y = 1, y + 1
        if len(out) > 24:
            raise HTTPException(status_code=400, detail="Period too long (max 24 months)")
    return out


def _mon_label(month: str) -> str:
    try:
        return datetime(int(month[:4]), int(month[5:7]), 1).strftime("%b %Y").upper()
    except Exception:
        return month


def _r0(v: Any) -> int:
    try:
        return int(round(float(v or 0)))
    except Exception:
        return 0


async def _latest_run(company_id: str, month: str) -> Optional[Dict[str, Any]]:
    return await db.compliance_salary_runs.find_one(
        {"company_id": company_id, "month": month}, {"_id": 0},
        sort=[("created_at", -1)])


async def _month_challan(company_id: str, month: str) -> Optional[Dict[str, Any]]:
    """Aggregated challan numbers for one month (None if no run)."""
    run = await _latest_run(company_id, month)
    if not run:
        return None
    from routes.compliance_settings import (
        get_standard_compliance_cfg, get_firm_statutory_overrides)
    cfg = {**(await get_standard_compliance_cfg(on_date=f"{month}-31")),
           **(await get_firm_statutory_overrides(company_id))}
    cap = float(cfg.get("pf_wage_cap") or 15000)
    ee = er_epf = eps = 0
    epf_wages = edli_wages = 0.0
    subs = eps_subs = 0
    # Iter 230 (user sample) — Combined Challan extras: total employees in
    # the month, non-PF employee count + their wages.
    total_emps = 0
    non_pf = 0
    non_pf_wages = 0.0
    for r in run.get("rows") or []:
        total_emps += 1
        if not r.get("pf_applicable"):
            non_pf += 1
            non_pf_wages += float(r.get("gross_paid") or r.get("monthly_gross") or 0)
            continue
        subs += 1
        ee += _r0(r.get("pf_employee"))
        er_epf += _r0(r.get("pf_employer_epf"))
        eps += _r0(r.get("pf_employer_eps"))
        if _r0(r.get("pf_employer_eps")) > 0:
            eps_subs += 1
        w = float(r.get("pf_wages") or 0)
        epf_wages += w
        edli_wages += min(w, cap)
    ac2 = max(_r0(epf_wages * float(cfg.get("pf_admin_percent") or 0) / 100.0),
              500 if epf_wages > 0 else 0)
    ac21 = _r0(edli_wages * float(cfg.get("pf_edli_percent") or 0) / 100.0)
    ac22 = _r0(edli_wages * float(cfg.get("pf_edli_admin_percent") or 0) / 100.0)
    return {
        "month": month, "label": _mon_label(month), "subscribers": subs,
        "eps_subscribers": eps_subs or subs,
        "ee": ee, "er_epf": er_epf, "eps": eps,
        "ac2": ac2, "ac21": ac21, "ac22": ac22,
        "epf_wages": _r0(epf_wages), "edli_wages": _r0(edli_wages),
        "total": ee + er_epf + eps + ac2 + ac21 + ac22,
        "total_emps": total_emps, "non_pf": non_pf,
        "non_pf_wages": _r0(non_pf_wages),
        "run_id": run.get("run_id"),
    }


async def _company(company_id: str) -> Dict[str, Any]:
    c = await db.companies.find_one({"company_id": company_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Firm not found")
    return c


async def _auth(authorization: Optional[str], company_id: str):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if admin.get("role") == "company_admin" and admin.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Not your firm")
    return admin


# --------------------------------------------------------------- summary
@router.get("/summary")
async def pf_reports_summary(company_id: str, month_from: str, month_to: str,
                             authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    months = _months_range(month_from, month_to)
    rows, missing = [], []
    for m in months:
        c = await _month_challan(company_id, m)
        (rows.append(c) if c else missing.append(m))
    tot = {k: sum(r[k] for r in rows) for k in
           ("ee", "er_epf", "eps", "ac2", "ac21", "ac22", "total")}
    return {"months": rows, "missing_months": missing, "totals": tot}


# --------------------------------------------------------------- challan
def _challan_pdf(firm: Dict[str, Any], months: List[Dict[str, Any]],
                 fmt: Optional[Dict[str, Any]] = None) -> bytes:
    """Iter 230 (user sample) — SBI 'COMBINED CHALLAN OF A/C NO. 01, 02,
    10, 21 & 22 (WITH ECR)' layout: header block, TOTAL SUBSCRIBERS +
    TOTAL WAGES rows, particulars grid, Grand Total in words, FOR BANK /
    FOR ESTABLISHMENT boxes and the Non-PF footer lines."""
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)
    from utils.salary_register_pdf import _num_to_words_inr
    fmt = fmt or {}
    fs = float(fmt.get("font_size") or 8)
    pagesize = landscape(A4) if (fmt.get("orientation") == "landscape") else A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=10 * mm,
                            rightMargin=10 * mm, topMargin=10 * mm,
                            bottomMargin=10 * mm)
    H1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=11, alignment=1)
    H2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=9.5, alignment=1)
    LBL = ParagraphStyle("lbl", fontName="Helvetica", fontSize=fs + 0.5)
    LBLB = ParagraphStyle("lblb", fontName="Helvetica-Bold", fontSize=fs + 0.5)
    SMALL = ParagraphStyle("small", fontName="Helvetica", fontSize=fs)
    story: List[Any] = []
    BL = ""  # blank cell

    def month_page(c: Dict[str, Any]) -> List[Any]:
        el: List[Any] = []
        el.append(Paragraph(
            "COMBINED CHALLAN OF A/C NO. 01, 02, 10, 21 &amp; 22 (WITH ECR)", H1))
        el.append(Paragraph("(STATE BANK OF INDIA)", H2))
        el.append(Paragraph("EMPLOYEES' PROVIDENT FUND ORGANISATION", H2))
        if firm.get("pf_office_city") or firm.get("city"):
            el.append(Paragraph(str(firm.get("pf_office_city") or firm.get("city")).upper(), H2))
        el.append(Spacer(1, 3 * mm))
        el.append(Paragraph(
            f"<b>ESTABLISHMENT CODE &amp; NAME :</b> "
            f"{firm.get('pf_code') or '—'}, {firm.get('name') or ''}", LBL))
        el.append(Paragraph(
            f"<b>ADDRESS :</b> {firm.get('address') or '—'}", LBL))
        el.append(Paragraph(
            f"<b>DUES FOR THE WAGES MONTH OF :</b> {c['label']}", LBL))
        el.append(Spacer(1, 3 * mm))

        # --- TOTAL SUBSCRIBERS / TOTAL WAGES ---
        sub_rows = [
            ["TOTAL SUBSCRIBERS", f"A/C. 01 :  {c['subscribers']}",
             f"A/C. 10 :  {c.get('eps_subscribers', c['subscribers'])}",
             f"A/C. 21 :  {c['subscribers']}"],
            ["TOTAL WAGES", f"A/C. 01 :  {c['epf_wages']}",
             f"A/C. 10 :  {c['edli_wages']}",
             f"A/C. 21 :  {c['edli_wages']}"],
        ]
        st = Table(sub_rows, colWidths=[45 * mm, 48 * mm, 48 * mm, 48 * mm])
        st.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), fs),
            ("GRID", (0, 0), (-1, -1), 0.5, rl.black),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        el.append(st)
        el.append(Spacer(1, 3 * mm))

        # --- Particulars grid ---
        er_total = c["er_epf"] + c["eps"] + c["ac21"]
        adm_total = c["ac2"] + c["ac22"]
        g01 = c["ee"] + c["er_epf"]
        rows = [
            ["SL.", "PARTICULARS", "A/C. 01", "A/C. 02", "A/C. 10",
             "A/C. 21", "A/C. 22", "TOTAL"],
            ["1", "EMPLOYER'S SHARE OF CONTRIBUTION",
             c["er_epf"], BL, c["eps"], c["ac21"], BL, er_total],
            ["2", "EMPLOYEE'S SHARE OF CONTRIBUTION",
             c["ee"], BL, BL, BL, 0, c["ee"]],
            ["3", "ADMIN CHARGES", BL, c["ac2"], BL, BL, c["ac22"], adm_total],
            ["4", "INSPECTION CHARGES", BL, BL, BL, BL, BL, BL],
            ["5", "PENAL DAMAGES", BL, BL, BL, BL, BL, BL],
            ["6", "MISC. PAYMENT (INTEREST U/S 7Q)", BL, BL, BL, BL, BL, BL],
            ["", "GRAND TOTAL",
             g01, c["ac2"], c["eps"], c["ac21"], c["ac22"], c["total"]],
        ]
        t = Table(rows, colWidths=[8 * mm, 63 * mm, 20 * mm, 18 * mm,
                                   18 * mm, 18 * mm, 18 * mm, 26 * mm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), fs),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, rl.black),
            ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#EEEEEE")),
            ("BACKGROUND", (0, -1), (-1, -1), rl.HexColor("#F5F5F5")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        el.append(t)
        el.append(Spacer(1, 2.5 * mm))
        el.append(Paragraph(
            f"<b><u>GRAND TOTAL (IN WORDS)</u> :</b> RUPEES : "
            f"{_num_to_words_inr(int(c['total'])).upper()}", LBLB))
        el.append(Spacer(1, 4 * mm))

        # --- FOR BANK / FOR ESTABLISHMENT boxes ---
        bank_lines = [
            "FOR BANKS USE ONLY", "",
            "Amount Received Rs. : ______________",
            "Date of presentation of Cheque/DD : ______________",
            "Date of Realisation of Cheque/DD : ______________",
            "SBI Branch Name : ______________",
            "SBI Branch Code : ______________",
        ]
        est_lines = [
            "FOR ESTABLISHMENT USE ONLY (To be manually filled by Employer)", "",
            "Cheque/DD No. : ______________    Date : __________",
            "Cheque/DD drawn bank & Branch : ______________",
            "Name of the Depositer : ______________",
            "Date of Deposit : ____________  Mobile No. : __________",
            "",
            "Signature of Depositor : ______________",
        ]
        bt = Table(
            [[Paragraph("<br/>".join(bank_lines), SMALL),
              Paragraph("<br/>".join(est_lines), SMALL)]],
            colWidths=[92 * mm, 97 * mm])
        bt.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, rl.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        el.append(bt)
        el.append(Spacer(1, 3 * mm))
        ft = Table([[
            Paragraph(
                f"No Of Total Employee's in this month : <b>{c.get('total_emps', 0)}</b><br/>"
                f"No Of Non PF Employee's in this month : <b>{c.get('non_pf', 0)}</b><br/>"
                f"Wages Of Non PF Employee's in this month : <b>{c.get('non_pf_wages', 0)}</b>",
                SMALL),
            Paragraph(
                "KINDLY SUBMIT CHEQUE/DEMAND DRAFT &amp; CHALLAN AT SBI COUNTER ONLY",
                ParagraphStyle("foot", fontName="Helvetica-Bold",
                               fontSize=fs, alignment=1)),
        ]], colWidths=[92 * mm, 97 * mm])
        ft.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, rl.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        el.append(ft)
        return el

    for i, c in enumerate(months):
        story.extend(month_page(c))
        if i < len(months) - 1:
            story.append(PageBreak())

    if len(months) > 1:
        story.append(PageBreak())
        story.append(Paragraph("PERIOD SUMMARY — PF CHALLAN", H1))
        story.append(Spacer(1, 4 * mm))
        rows = [["MONTH", "EE (A/c 1)", "ER EPF (A/c 1)", "EPS (A/c 10)",
                 "ADMIN (A/c 2)", "EDLI (A/c 21)", "EDLI ADM (A/c 22)", "TOTAL"]]
        for c in months:
            rows.append([c["label"], c["ee"], c["er_epf"], c["eps"],
                         c["ac2"], c["ac21"], c["ac22"], c["total"]])
        rows.append(["GRAND TOTAL",
                     sum(c["ee"] for c in months), sum(c["er_epf"] for c in months),
                     sum(c["eps"] for c in months), sum(c["ac2"] for c in months),
                     sum(c["ac21"] for c in months), sum(c["ac22"] for c in months),
                     sum(c["total"] for c in months)])
        t = Table(rows, colWidths=[28 * mm] + [22 * mm] * 7)
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), fs),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, rl.black),
            ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#EEEEEE")),
        ]))
        story.append(t)
    doc.build(story)
    return buf.getvalue()


@router.get("/challan.pdf")
async def challan_pdf(company_id: str, month_from: str, month_to: str,
                      authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    firm = await _company(company_id)
    months = [c for m in _months_range(month_from, month_to)
              if (c := await _month_challan(company_id, m))]
    if not months:
        raise HTTPException(status_code=404,
                            detail="No compliance salary run found in this period")
    from routes.report_formats import get_report_format
    pdf = _challan_pdf(firm, months, await get_report_format("pf_challan"))
    fn = f"PF_Challan_{firm.get('name', '')}_{month_from}_{month_to}.pdf".replace(" ", "_")
    return Response(pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@router.get("/challan.xlsx")
async def challan_xlsx(company_id: str, month_from: str, month_to: str,
                       authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    firm = await _company(company_id)
    months = [c for m in _months_range(month_from, month_to)
              if (c := await _month_challan(company_id, m))]
    if not months:
        raise HTTPException(status_code=404,
                            detail="No compliance salary run found in this period")
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from utils.salary_register_pdf import _num_to_words_inr
    wb = Workbook()
    ws = wb.active
    ws.title = "PF Challan"
    bold = Font(bold=True)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center")

    def _row(vals, *, b=False, grid=False):
        ws.append(vals)
        for cell in ws[ws.max_row]:
            if b:
                cell.font = bold
            if grid:
                cell.border = box
        return ws.max_row

    first = True
    for c in months:
        if not first:
            _row([])
            _row([])
        first = False
        r = _row(["COMBINED CHALLAN OF A/C NO. 01, 02, 10, 21 & 22 (WITH ECR)"], b=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(row=r, column=1).alignment = ctr
        r = _row(["(STATE BANK OF INDIA) — EMPLOYEES' PROVIDENT FUND ORGANISATION"], b=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(row=r, column=1).alignment = ctr
        _row([f"ESTABLISHMENT CODE & NAME : {firm.get('pf_code') or '—'}, {firm.get('name') or ''}"])
        _row([f"ADDRESS : {firm.get('address') or '—'}"])
        _row([f"DUES FOR THE WAGES MONTH OF : {c['label']}"], b=True)
        _row([])
        _row(["TOTAL SUBSCRIBERS", f"A/C. 01 : {c['subscribers']}",
              f"A/C. 10 : {c.get('eps_subscribers', c['subscribers'])}",
              f"A/C. 21 : {c['subscribers']}"], b=True, grid=True)
        _row(["TOTAL WAGES", f"A/C. 01 : {c['epf_wages']}",
              f"A/C. 10 : {c['edli_wages']}",
              f"A/C. 21 : {c['edli_wages']}"], b=True, grid=True)
        _row([])
        _row(["SL.", "PARTICULARS", "A/C. 01", "A/C. 02", "A/C. 10",
              "A/C. 21", "A/C. 22", "TOTAL"], b=True, grid=True)
        er_total = c["er_epf"] + c["eps"] + c["ac21"]
        _row(["1", "EMPLOYER'S SHARE OF CONTRIBUTION",
              c["er_epf"], None, c["eps"], c["ac21"], None, er_total], grid=True)
        _row(["2", "EMPLOYEE'S SHARE OF CONTRIBUTION",
              c["ee"], None, None, None, 0, c["ee"]], grid=True)
        _row(["3", "ADMIN CHARGES",
              None, c["ac2"], None, None, c["ac22"], c["ac2"] + c["ac22"]], grid=True)
        _row(["4", "INSPECTION CHARGES", None, None, None, None, None, None], grid=True)
        _row(["5", "PENAL DAMAGES", None, None, None, None, None, None], grid=True)
        _row(["6", "MISC. PAYMENT (INTEREST U/S 7Q)",
              None, None, None, None, None, None], grid=True)
        _row(["", "GRAND TOTAL", c["ee"] + c["er_epf"], c["ac2"], c["eps"],
              c["ac21"], c["ac22"], c["total"]], b=True, grid=True)
        _row([f"GRAND TOTAL (IN WORDS) : RUPEES : "
              f"{_num_to_words_inr(int(c['total'])).upper()}"], b=True)
        _row([])
        _row([f"No Of Total Employee's in this month : {c.get('total_emps', 0)}"])
        _row([f"No Of Non PF Employee's in this month : {c.get('non_pf', 0)}"])
        _row([f"Wages Of Non PF Employee's in this month : {c.get('non_pf_wages', 0)}"])
        _row(["KINDLY SUBMIT CHEQUE/DEMAND DRAFT & CHALLAN AT SBI COUNTER ONLY"], b=True)

    for col, w in zip("ABCDEFGH", (6, 38, 12, 12, 12, 12, 12, 14)):
        ws.column_dimensions[col].width = w
    out = io.BytesIO()
    wb.save(out)
    fn = f"PF_Challan_{firm.get('name', '')}_{month_from}_{month_to}.xlsx".replace(" ", "_")
    return Response(out.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# ------------------------------------------------------------------- ECR
async def _ecr_month(company_id: str, month: str):
    """(run, member-lines) for a month — reuses challans._ecr_lines."""
    run = await _latest_run(company_id, month)
    if not run:
        return None, []
    from routes.challans import _ecr_lines, _uan_esic_map
    extra = await _uan_esic_map(run.get("rows") or [])
    return run, _ecr_lines(run, extra)


_ECR_HDR = ["Sl.", "UAN", "Member Name", "Gross Wages", "EPF Wages",
            "EPS Wages", "EDLI Wages", "EE Share", "EPS Contri.",
            "ER Share (Diff)", "Refunds", "NCP Days"]


def _ecr_pdf(firm: Dict[str, Any], sections: List[Dict[str, Any]],
             rate: float, fmt: Optional[Dict[str, Any]] = None) -> bytes:
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)
    from routes.report_formats import resolve_columns
    # Iter 163 — saved global format (Utilities → PDF Report Formats).
    fmt = fmt or {}
    cols = resolve_columns("pf_ecr", fmt)
    fs = float(fmt.get("font_size") or 7.5)
    title = fmt.get("title") or "EMPLOYEE'S PROVIDENT FUND ORGANISATION"
    pagesize = A4 if (fmt.get("orientation") == "portrait") else landscape(A4)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=10 * mm,
                            rightMargin=10 * mm, topMargin=10 * mm,
                            bottomMargin=10 * mm)
    H1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=12, alignment=1)
    H2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=10, alignment=1)
    LBL = ParagraphStyle("lbl", fontName="Helvetica", fontSize=8.5)
    # widths are proportions — stretch to the full printable width
    avail = pagesize[0] - 20 * mm
    wsum = sum(c["width"] for c in cols) or 1
    col_widths = [avail * c["width"] / wsum for c in cols]
    story: List[Any] = []
    for si, sec in enumerate(sections):
        lines = sec["lines"]
        story.append(Paragraph(title, H1))
        story.append(Paragraph(
            f"RETURN STATEMENT (Regular Return) : {sec['label']}", H2))
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(
            f"<b>Name of Establishment :</b> {firm.get('name') or ''}"
            f"&nbsp;&nbsp;&nbsp;<b>Establishment Id :</b> {firm.get('pf_code') or '—'}"
            f"&nbsp;&nbsp;&nbsp;<b>Contribution Rate (%) :</b> {rate:g}"
            f"&nbsp;&nbsp;&nbsp;<b>Total Members :</b> {len(lines)}", LBL))
        ee_t = sum(x["epf_ee"] for x in lines)
        eps_t = sum(x["eps_er"] for x in lines)
        er_t = sum(x["diff_er"] for x in lines)
        story.append(Paragraph(
            f"<b>Total EPF Contribution (EE) :</b> {ee_t}"
            f"&nbsp;&nbsp;&nbsp;<b>Total EPS Contribution :</b> {eps_t}"
            f"&nbsp;&nbsp;&nbsp;<b>Total EPF-EPS (ER Diff.) :</b> {er_t}"
            f"&nbsp;&nbsp;&nbsp;<b>Total Refund of Advances :</b> 0", LBL))
        story.append(Spacer(1, 3 * mm))

        def _cell(x: Dict[str, Any], i: int, key: str) -> Any:
            if key == "sl":
                return i + 1
            if key == "uan":
                return x["uan"] or "—"
            return x.get(key, "")

        tot = {"name": "TOTAL",
               "gross": sum(x["gross"] for x in lines),
               "epf_wages": sum(x["epf_wages"] for x in lines),
               "eps_wages": sum(x["eps_wages"] for x in lines),
               "edli_wages": sum(x["edli_wages"] for x in lines),
               "epf_ee": ee_t, "eps_er": eps_t, "diff_er": er_t, "refund": 0}
        data = [[c["heading"] for c in cols]] + [
            [_cell(x, i, c["key"]) for c in cols] for i, x in enumerate(lines)]
        data.append([tot.get(c["key"], "") for c in cols])
        t = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), fs),
            ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#888888")),
            ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#EEEEEE")),
            ("BACKGROUND", (0, -1), (-1, -1), rl.HexColor("#F5F5F5")),
        ]
        for ci, c in enumerate(cols):
            if c["numeric"]:
                style.append(("ALIGN", (ci, 0), (ci, -1), "RIGHT"))
        t.setStyle(TableStyle(style))
        story.append(t)
        if si < len(sections) - 1:
            story.append(PageBreak())
    doc.build(story)
    return buf.getvalue()


@router.get("/ecr.pdf")
async def ecr_pdf(company_id: str, month_from: str, month_to: str,
                  authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    firm = await _company(company_id)
    from routes.compliance_settings import get_standard_compliance_cfg
    rate = float((await get_standard_compliance_cfg()).get("pf_percent_employee") or 12)
    sections = []
    for m in _months_range(month_from, month_to):
        run, lines = await _ecr_month(company_id, m)
        if run and lines:
            sections.append({"month": m, "label": _mon_label(m), "lines": lines})
    if not sections:
        raise HTTPException(status_code=404,
                            detail="No PF members / compliance run found in this period")
    from routes.report_formats import get_report_format
    pdf = _ecr_pdf(firm, sections, rate, await get_report_format("pf_ecr"))
    fn = f"PF_ECR_{firm.get('name', '')}_{month_from}_{month_to}.pdf".replace(" ", "_")
    return Response(pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@router.get("/ecr.xlsx")
async def ecr_xlsx(company_id: str, month_from: str, month_to: str,
                   authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    firm = await _company(company_id)
    from routes.compliance_settings import get_standard_compliance_cfg
    rate = float((await get_standard_compliance_cfg()).get("pf_percent_employee") or 12)
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    any_data = False
    for m in _months_range(month_from, month_to):
        run, lines = await _ecr_month(company_id, m)
        if not run or not lines:
            continue
        any_data = True
        ws = wb.create_sheet(title=_mon_label(m).replace(" ", "-"))
        ws.append(["EMPLOYEE'S PROVIDENT FUND ORGANISATION"])
        ws.append([f"RETURN STATEMENT (Regular Return) : {_mon_label(m)}"])
        ws.append([f"Name of Establishment: {firm.get('name')}",
                   f"Establishment Id: {firm.get('pf_code') or '—'}",
                   f"Contribution Rate (%): {rate:g}",
                   f"Total Members: {len(lines)}"])
        ws.append([])
        ws.append(_ECR_HDR)
        for cell in ws[ws.max_row]:
            cell.font = bold
        for i, x in enumerate(lines):
            ws.append([i + 1, x["uan"] or "", x["name"], x["gross"],
                       x["epf_wages"], x["eps_wages"], x["edli_wages"],
                       x["epf_ee"], x["eps_er"], x["diff_er"], x["refund"], x["ncp"]])
        ws.append(["", "", "TOTAL",
                   sum(x["gross"] for x in lines),
                   sum(x["epf_wages"] for x in lines),
                   sum(x["eps_wages"] for x in lines),
                   sum(x["edli_wages"] for x in lines),
                   sum(x["epf_ee"] for x in lines),
                   sum(x["eps_er"] for x in lines),
                   sum(x["diff_er"] for x in lines), 0, ""])
        for cell in ws[ws.max_row]:
            cell.font = bold
        for col, w in zip("ABCDEFGHIJKL", (6, 16, 30, 12, 12, 12, 12, 10, 11, 13, 9, 10)):
            ws.column_dimensions[col].width = w
    if not any_data:
        raise HTTPException(status_code=404,
                            detail="No PF members / compliance run found in this period")
    out = io.BytesIO()
    wb.save(out)
    fn = f"PF_ECR_{firm.get('name', '')}_{month_from}_{month_to}.xlsx".replace(" ", "_")
    return Response(out.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# ================================================================== ESIC
async def _month_esic(company_id: str, month: str) -> Optional[Dict[str, Any]]:
    run = await _latest_run(company_id, month)
    if not run:
        return None
    from routes.challans import _uan_esic_map
    extra = await _uan_esic_map(run.get("rows") or [])
    lines: List[Dict[str, Any]] = []
    for r in run.get("rows") or []:
        if not r.get("esic_applicable"):
            continue
        days = float(r.get("present_days") or 0)
        wages = round(float(r.get("gross_paid") or 0), 2)
        lines.append({
            "ip_no": str((extra.get(r.get("user_id"), {}) or {}).get("esi_ip_no") or "").strip(),
            "name": (r.get("name") or "").upper(),
            "days": int(days) if days.is_integer() else days,
            "wages": wages,
            "ee": _r0(r.get("esic_employee")),
            "er": _r0(r.get("esic_employer")),
            "reason": "On Leave" if days <= 0 else "",
        })
    return {
        "month": month, "label": _mon_label(month), "lines": lines,
        "employees": len(lines),
        "wages": round(sum(x["wages"] for x in lines), 2),
        "ee": sum(x["ee"] for x in lines),
        "er": sum(x["er"] for x in lines),
        "total": sum(x["ee"] for x in lines) + sum(x["er"] for x in lines),
    }


def _esic_sheet_pdf(firm: Dict[str, Any], sections: List[Dict[str, Any]],
                    fmt: Optional[Dict[str, Any]] = None) -> bytes:
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)
    from routes.report_formats import resolve_columns
    # Iter 163 — saved global format (Utilities → PDF Report Formats).
    fmt = fmt or {}
    cols = resolve_columns("esic_contribution", fmt)
    fs = float(fmt.get("font_size") or 7.5)
    title = fmt.get("title") or "Employees' State Insurance Corporation"
    pagesize = landscape(A4) if (fmt.get("orientation") == "landscape") else A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=10 * mm,
                            rightMargin=10 * mm, topMargin=10 * mm,
                            bottomMargin=10 * mm)
    H1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=12, alignment=1)
    H2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=9.5, alignment=1)
    LBL = ParagraphStyle("lbl", fontName="Helvetica", fontSize=8.5)
    avail = pagesize[0] - 20 * mm
    wsum = sum(c["width"] for c in cols) or 1
    col_widths = [avail * c["width"] / wsum for c in cols]
    story: List[Any] = []
    for si, sec in enumerate(sections):
        lines = sec["lines"]
        story.append(Paragraph(title, H1))
        story.append(Paragraph(
            f"Contribution History Of {firm.get('esi_code') or firm.get('name') or ''} "
            f"for {sec['label']}", H2))
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(
            f"<b>Total IP Contribution :</b> {sec['ee']:,}"
            f"&nbsp;&nbsp;&nbsp;<b>Total Employer Contribution :</b> {sec['er']:,}"
            f"&nbsp;&nbsp;&nbsp;<b>Total Contribution :</b> {sec['total']:,}"
            f"&nbsp;&nbsp;&nbsp;<b>Total Government Contribution :</b> 0.00"
            f"&nbsp;&nbsp;&nbsp;<b>Total Monthly Wages :</b> {sec['wages']:,.2f}", LBL))
        story.append(Spacer(1, 3 * mm))

        def _cell(x: Dict[str, Any], i: int, key: str) -> Any:
            if key == "sl":
                return i + 1
            if key == "disable":
                return ""
            if key == "ip_no":
                return x["ip_no"] or "—"
            if key == "wages":
                return f"{x['wages']:.2f}"
            if key == "ee":
                return f"{x['ee']:.2f}"
            return x.get(key, "")

        tot = {"name": "TOTAL", "wages": f"{sec['wages']:.2f}",
               "ee": f"{sec['ee']:.2f}"}
        data = [[c["heading"] for c in cols]] + [
            [_cell(x, i, c["key"]) for c in cols] for i, x in enumerate(lines)]
        data.append([tot.get(c["key"], "") for c in cols])
        t = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), fs),
            ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#888888")),
            ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#EEEEEE")),
        ]
        for ci, c in enumerate(cols):
            if c["numeric"]:
                style.append(("ALIGN", (ci, 0), (ci, -1), "RIGHT"))
        t.setStyle(TableStyle(style))
        story.append(t)
        if si < len(sections) - 1:
            story.append(PageBreak())
    doc.build(story)
    return buf.getvalue()


def _esic_challan_pdf(firm: Dict[str, Any], months: List[Dict[str, Any]],
                      fmt: Optional[Dict[str, Any]] = None) -> bytes:
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)
    from utils.salary_register_pdf import _num_to_words_inr
    # Iter 163 — saved global format (Utilities → PDF Report Formats).
    fmt = fmt or {}
    fs = float(fmt.get("font_size") or 9)
    org_title = fmt.get("title") or "EMPLOYEE STATE INSURANCE CORPORATION"
    pagesize = landscape(A4) if (fmt.get("orientation") == "landscape") else A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=14 * mm,
                            rightMargin=14 * mm, topMargin=12 * mm,
                            bottomMargin=12 * mm)
    H1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=13, alignment=1)
    H2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=10, alignment=1)
    LBL = ParagraphStyle("lbl", fontName="Helvetica", fontSize=fs, leading=fs + 5)
    SM = ParagraphStyle("sm", fontName="Helvetica", fontSize=8,
                        textColor=rl.HexColor("#444444"), leading=12)
    story: List[Any] = []
    for i, c in enumerate(months):
        story.append(Paragraph(
            f"Code No. : {firm.get('esi_code') or '—'}", LBL))
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(org_title, H1))
        story.append(Paragraph("Challan For Deposit In A/C No. 1", H2))
        story.append(Spacer(1, 5 * mm))
        story.append(Paragraph(
            f"<b>Employer's Code :</b> {firm.get('esi_code') or '—'}"
            f"&nbsp;&nbsp;&nbsp;&nbsp;<b>Date of Payment :</b> ____________", LBL))
        story.append(Paragraph(
            f"<b>Name of Factory :</b> {firm.get('name') or ''}", LBL))
        story.append(Paragraph(
            f"<b>Estt. &amp; Address :</b> {firm.get('address') or '—'}", LBL))
        story.append(Paragraph(
            "<b>Mode of Payment :</b> CHEQUE / CASH"
            "&nbsp;&nbsp;&nbsp;&nbsp;<b>Details of Payment :</b> Regular Contribution", LBL))
        story.append(Paragraph(
            f"<b>Period of Contribution :</b> {c['label'].title()}"
            f"&nbsp;&nbsp;&nbsp;&nbsp;<b>No. of Employees :</b> {c['employees']}"
            f"&nbsp;&nbsp;&nbsp;&nbsp;<b>Total Wages :</b> Rs. {c['wages']:,.0f}", LBL))
        story.append(Spacer(1, 5 * mm))
        t = Table([
            ["Employees Contribution :", f"Rs. {c['ee']:,.2f}"],
            ["Employer's Contribution :", f"Rs. {c['er']:,.2f}"],
            ["TOTAL", f"Rs. {c['total']:,.0f}"],
        ], colWidths=[100 * mm, 60 * mm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica"),
            ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), fs + 0.5),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, rl.black),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t)
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(
            f"<b>Total Amount (In words) :</b> RUPEES: "
            f"{_num_to_words_inr(int(round(c['total'])))} ONLY", LBL))
        story.append(Spacer(1, 10 * mm))
        story.append(Paragraph(
            f"FOR {(firm.get('name') or '').upper()}", LBL))
        story.append(Spacer(1, 10 * mm))
        story.append(Paragraph("(Authorised Signatory)", LBL))
        story.append(Spacer(1, 6 * mm))
        story.append(Paragraph("<b>(Acknowledgement)</b>", H2))
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(
            f"Received Rs. : {c['total']:,.0f} — RUPEES: "
            f"{_num_to_words_inr(int(round(c['total'])))} ONLY", SM))
        story.append(Paragraph(
            "In cash/by Cheque/ D.D. No. : .................... "
            "Dated ............ drawn on .................... (subject to Realisation)", SM))
        story.append(Paragraph(
            "Bank Scroll No. .................... Dated ....................", SM))
        story.append(Paragraph("(Bank) In favour of ESIC A/C No. 1.", SM))
        story.append(Spacer(1, 8 * mm))
        story.append(Paragraph("(Authorised signature)", SM))
        if i < len(months) - 1:
            story.append(PageBreak())
    doc.build(story)
    return buf.getvalue()


@router.get("/esic-sheet.pdf")
async def esic_sheet_pdf(company_id: str, month_from: str, month_to: str,
                         authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    firm = await _company(company_id)
    sections = [s for m in _months_range(month_from, month_to)
                if (s := await _month_esic(company_id, m)) and s["lines"]]
    if not sections:
        raise HTTPException(status_code=404,
                            detail="No ESIC members / compliance run found in this period")
    from routes.report_formats import get_report_format
    pdf = _esic_sheet_pdf(firm, sections,
                          await get_report_format("esic_contribution"))
    fn = f"ESIC_Contribution_{firm.get('name', '')}_{month_from}_{month_to}.pdf".replace(" ", "_")
    return Response(pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@router.get("/esic-sheet.xlsx")
async def esic_sheet_xlsx(company_id: str, month_from: str, month_to: str,
                          authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    firm = await _company(company_id)
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    any_data = False
    for m in _months_range(month_from, month_to):
        sec = await _month_esic(company_id, m)
        if not sec or not sec["lines"]:
            continue
        any_data = True
        ws = wb.create_sheet(title=_mon_label(m).replace(" ", "-"))
        ws.append(["Employees' State Insurance Corporation"])
        ws.append([f"Contribution History Of {firm.get('esi_code') or firm.get('name')} for {sec['label']}"])
        ws.append([f"Total IP Contribution: {sec['ee']}",
                   f"Total Employer Contribution: {sec['er']}",
                   f"Total Contribution: {sec['total']}",
                   "Total Government Contribution: 0.00",
                   f"Total Monthly Wages: {sec['wages']:.2f}"])
        ws.append([])
        ws.append(["SNo.", "Is Disable", "IP Number", "IP Name",
                   "No. Of Days", "Total Wages", "IP Contribution", "Reason"])
        for cell in ws[ws.max_row]:
            cell.font = bold
        for i, x in enumerate(sec["lines"]):
            ws.append([i + 1, "", x["ip_no"], x["name"], x["days"],
                       x["wages"], x["ee"], x["reason"]])
        ws.append(["", "", "", "TOTAL", "", sec["wages"], sec["ee"], ""])
        for cell in ws[ws.max_row]:
            cell.font = bold
        for col, w in zip("ABCDEFGH", (6, 10, 15, 30, 11, 13, 15, 14)):
            ws.column_dimensions[col].width = w
    if not any_data:
        raise HTTPException(status_code=404,
                            detail="No ESIC members / compliance run found in this period")
    out = io.BytesIO()
    wb.save(out)
    fn = f"ESIC_Contribution_{firm.get('name', '')}_{month_from}_{month_to}.xlsx".replace(" ", "_")
    return Response(out.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@router.get("/esic-challan.pdf")
async def esic_challan_pdf(company_id: str, month_from: str, month_to: str,
                           authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    firm = await _company(company_id)
    months = [s for m in _months_range(month_from, month_to)
              if (s := await _month_esic(company_id, m)) and s["lines"]]
    if not months:
        raise HTTPException(status_code=404,
                            detail="No ESIC members / compliance run found in this period")
    from routes.report_formats import get_report_format
    pdf = _esic_challan_pdf(firm, months,
                            await get_report_format("esic_challan"))
    fn = f"ESIC_Challan_{firm.get('name', '')}_{month_from}_{month_to}.pdf".replace(" ", "_")
    return Response(pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@router.get("/esic-challan.xlsx")
async def esic_challan_xlsx(company_id: str, month_from: str, month_to: str,
                            authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    firm = await _company(company_id)
    months = [s for m in _months_range(month_from, month_to)
              if (s := await _month_esic(company_id, m)) and s["lines"]]
    if not months:
        raise HTTPException(status_code=404,
                            detail="No ESIC members / compliance run found in this period")
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "ESIC Challan"
    bold = Font(bold=True)
    ws.append(["EMPLOYEE STATE INSURANCE CORPORATION — Challan For Deposit In A/C No. 1"])
    ws.append([f"Employer's Code: {firm.get('esi_code') or '—'}   {firm.get('name')}"])
    ws.append([])
    ws.append(["Month", "No. of Employees", "Total Wages",
               "Employees Contribution", "Employer's Contribution", "TOTAL"])
    for cell in ws[ws.max_row]:
        cell.font = bold
    for c in months:
        ws.append([c["label"], c["employees"], c["wages"], c["ee"], c["er"], c["total"]])
    ws.append(["GRAND TOTAL", "", round(sum(c["wages"] for c in months), 2),
               sum(c["ee"] for c in months), sum(c["er"] for c in months),
               sum(c["total"] for c in months)])
    for cell in ws[ws.max_row]:
        cell.font = bold
    for col, w in zip("ABCDEF", (14, 16, 14, 21, 21, 12)):
        ws.column_dimensions[col].width = w
    out = io.BytesIO()
    wb.save(out)
    fn = f"ESIC_Challan_{firm.get('name', '')}_{month_from}_{month_to}.xlsx".replace(" ", "_")
    return Response(out.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@router.get("/esic-summary")
async def esic_summary(company_id: str, month_from: str, month_to: str,
                       authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    rows, missing = [], []
    for m in _months_range(month_from, month_to):
        s = await _month_esic(company_id, m)
        if s and s["lines"]:
            rows.append({k: s[k] for k in
                         ("month", "label", "employees", "wages", "ee", "er", "total")})
        else:
            missing.append(m)
    tot = {k: round(sum(r[k] for r in rows), 2) for k in ("wages", "ee", "er", "total")}
    return {"months": rows, "missing_months": missing, "totals": tot}
