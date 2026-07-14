"""Iter 92 — Route module: Compliance Reports (single firm).

Reports → Compliance Reports. Three ACT-wise report types for ONE firm,
broken down EMPLOYEE-GROUP-wise, over Monthly / Half-Yearly / Annual
periods:

  * ``contributions`` — PF / ESIC / PT / TDS totals sourced from the
    LATEST compliance salary run of each month in the period.
  * ``leave``         — leave-register style report (approved leaves by
    type per employee + Earned-Leave entitlement @ 1 day / 20 present
    days from the biometric attendance, Factories Act s.79).
  * ``gratuity``      — Payment of Gratuity Act calculation per employee:
    (15/26) × monthly Basic × completed years of service, eligibility at
    5 years (or on exit), ceiling ₹20,00,000.

  GET /api/admin/reports/compliance?company_id=&report=&year=&period=monthly|half1|half2|annual&month=&employee_type=
  GET /api/admin/reports/compliance.xlsx?...same...
"""
from datetime import date
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException
from fastapi.responses import StreamingResponse

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
)

router = APIRouter(prefix="/api/admin/reports", tags=["compliance-reports"])

GRATUITY_CEILING = 2_000_000.0


def _months_for(year: int, period: str, month: Optional[int]) -> List[str]:
    if period == "monthly":
        if not month or not 1 <= month <= 12:
            raise HTTPException(status_code=400, detail="month (1-12) required for monthly period")
        return [f"{year:04d}-{month:02d}"]
    if period == "half1":
        return [f"{year:04d}-{m:02d}" for m in range(1, 7)]
    if period == "half2":
        return [f"{year:04d}-{m:02d}" for m in range(7, 13)]
    if period == "annual":
        return [f"{year:04d}-{m:02d}" for m in range(1, 13)]
    raise HTTPException(status_code=400, detail="period must be monthly | half1 | half2 | annual")


def _resolve_company(admin: Dict[str, Any], company_id: Optional[str]) -> str:
    if admin["role"] == "company_admin":
        return admin.get("company_id") or ""
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required (pick one firm)")
    return company_id


async def _latest_runs(company_id: str, months: List[str]) -> Dict[str, Dict[str, Any]]:
    """month → latest compliance run doc (rows included)."""
    out: Dict[str, Dict[str, Any]] = {}
    for m in months:
        run = await db.compliance_salary_runs.find_one(
            {"company_id": company_id, "month": m},
            {"_id": 0},
            sort=[("generated_at", -1)],
        )
        if run:
            out[m] = run
    return out


# ---------------------------------------------------------------------------
# Report builders
# ---------------------------------------------------------------------------

_CONTRIB_KEYS = [
    ("gross_paid", "Gross Paid"),
    ("pf_wages", "PF Wages"),
    ("pf_employee", "PF (Employee 12%)"),
    ("pf_employer_epf", "PF Employer EPF"),
    ("pf_employer_eps", "PF Employer EPS"),
    ("esic_employee", "ESIC (Employee)"),
    ("esic_employer", "ESIC (Employer)"),
    ("pt", "Prof. Tax"),
    ("tds", "TDS"),
    ("net", "Net Pay"),
]


async def _report_contributions(
    company_id: str, months: List[str], employee_type: Optional[str],
) -> Dict[str, Any]:
    runs = await _latest_runs(company_id, months)
    groups: Dict[str, Dict[str, Any]] = {}
    month_totals: Dict[str, Dict[str, float]] = {}
    for m, run in runs.items():
        mt = month_totals.setdefault(m, {k: 0.0 for k, _ in _CONTRIB_KEYS})
        for row in run.get("rows") or []:
            grp = (row.get("employee_type") or row.get("employee_group") or "Ungrouped")
            if employee_type and grp != employee_type:
                continue
            g = groups.setdefault(grp, {
                "group": grp, "employees": set(),
                **{k: 0.0 for k, _ in _CONTRIB_KEYS},
            })
            g["employees"].add(row.get("user_id"))
            for k, _ in _CONTRIB_KEYS:
                v = float(row.get(k) or 0.0)
                g[k] += v
                mt[k] += v
    rows = []
    for g in sorted(groups.values(), key=lambda x: x["group"]):
        rows.append({
            "group": g["group"],
            "employees": len(g["employees"]),
            **{k: round(g[k], 2) for k, _ in _CONTRIB_KEYS},
        })
    return {
        "columns": [{"key": "group", "label": "Employee Group"},
                    {"key": "employees", "label": "Employees"}] +
                   [{"key": k, "label": lbl} for k, lbl in _CONTRIB_KEYS],
        "rows": rows,
        "months_covered": sorted(runs.keys()),
        "months_missing": [m for m in months if m not in runs],
        "month_totals": {m: {k: round(v, 2) for k, v in t.items()}
                         for m, t in sorted(month_totals.items())},
    }


_LEAVE_COLS = [
    ("employee_code", "Emp Code"), ("name", "Name"), ("group", "Group"),
    ("present_days", "Present Days"),
    ("el_entitled", "EL Entitled (1/20)"),
    ("leaves_taken", "Leaves Taken (approved)"),
    ("leave_breakup", "Break-up by Type"),
    ("el_balance", "EL Balance"),
]


def _days_between(a: str, b: str) -> int:
    try:
        d1 = date.fromisoformat(a[:10])
        d2 = date.fromisoformat(b[:10])
        return max((d2 - d1).days + 1, 1)
    except ValueError:
        return 1


async def _report_leave(
    company_id: str, months: List[str], employee_type: Optional[str],
) -> Dict[str, Any]:
    q: Dict[str, Any] = {"role": "employee", "company_id": company_id}
    if employee_type:
        q["employee_type"] = employee_type
    emps = await db.users.find(
        q, {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1, "employee_type": 1},
    ).sort("name", 1).to_list(3000)
    uid_set = {e["user_id"] for e in emps}

    start, end = f"{months[0]}-01", f"{months[-1]}-31"

    # Present days from attendance punches (distinct days with an IN punch).
    present: Dict[str, set] = {}
    async for a in db.attendance.find(
        {"company_id": company_id, "date": {"$gte": start, "$lte": end}},
        {"_id": 0, "user_id": 1, "date": 1},
    ):
        if a.get("user_id") in uid_set:
            present.setdefault(a["user_id"], set()).add(a.get("date"))

    # Approved leaves overlapping the window.
    taken: Dict[str, Dict[str, int]] = {}
    async for lv in db.leaves.find(
        {"company_id": company_id, "status": "approved",
         "from_date": {"$lte": end}, "to_date": {"$gte": start}},
        {"_id": 0, "user_id": 1, "leave_type": 1, "from_date": 1, "to_date": 1},
    ):
        if lv.get("user_id") not in uid_set:
            continue
        days = _days_between(max(lv["from_date"], start), min(lv["to_date"], end))
        bt = taken.setdefault(lv["user_id"], {})
        bt[lv.get("leave_type") or "other"] = bt.get(lv.get("leave_type") or "other", 0) + days

    rows = []
    for e in emps:
        pd = len(present.get(e["user_id"], set()))
        el = pd // 20  # Factories Act s.79 — 1 EL per 20 days worked
        bt = taken.get(e["user_id"], {})
        tk = sum(bt.values())
        rows.append({
            "employee_code": e.get("employee_code"),
            "name": e.get("name"),
            "group": e.get("employee_type") or "Ungrouped",
            "present_days": pd,
            "el_entitled": el,
            "leaves_taken": tk,
            "leave_breakup": ", ".join(f"{k}: {v}" for k, v in sorted(bt.items())) or "—",
            "el_balance": max(el - tk, 0),
        })
    return {
        "columns": [{"key": k, "label": lbl} for k, lbl in _LEAVE_COLS],
        "rows": rows,
    }


_GRATUITY_COLS = [
    ("employee_code", "Emp Code"), ("name", "Name"), ("group", "Group"),
    ("doj", "DOJ"), ("as_of", "As-of / Exit"),
    ("service_years", "Service (yrs)"),
    ("monthly_basic", "Monthly Basic"),
    ("eligible", "Eligible (≥5 yrs)"),
    ("gratuity", "Gratuity (15/26 × Basic × yrs)"),
]


def _parse_any_date(v: Optional[str]) -> Optional[date]:
    s = (v or "").strip()
    if not s:
        return None
    try:
        if "-" in s and len(s.split("-")[0]) == 4:
            return date.fromisoformat(s[:10])
        dd, mm, yy = s.split("-")[:3]
        return date(int(yy), int(mm), int(dd))
    except (ValueError, IndexError):
        return None


def _emp_basic(emp: Dict[str, Any]) -> float:
    for key in ("salary_structure_compliance", "salary_structure_actual"):
        for r in emp.get(key) or []:
            if isinstance(r, dict) and str(r.get("head", "")).strip().lower().startswith("basic"):
                amt = float(r.get("amount") or 0.0)
                if amt > 0:
                    return amt
    return float(emp.get("salary_monthly") or 0.0)


async def _report_gratuity(
    company_id: str, employee_type: Optional[str],
) -> Dict[str, Any]:
    q: Dict[str, Any] = {"role": "employee", "company_id": company_id}
    if employee_type:
        q["employee_type"] = employee_type
    emps = await db.users.find(q, {"_id": 0}).sort("name", 1).to_list(3000)
    today = date.today()
    rows = []
    for e in emps:
        doj = _parse_any_date(e.get("doj"))
        exit_d = _parse_any_date(e.get("exit_date"))
        as_of = exit_d or today
        if not doj:
            years = 0.0
        else:
            days = (as_of - doj).days
            years = max(days / 365.25, 0.0)
        completed = int(years + (1 if (years - int(years)) >= 0.75 else 0))  # ≥6mo in final yr rounds up per Act practice (240 days)
        basic = _emp_basic(e)
        eligible = years >= 4.75  # judicial 4y 190/240d relaxation
        gratuity = round(min((15.0 / 26.0) * basic * completed, GRATUITY_CEILING), 2) if eligible else 0.0
        rows.append({
            "employee_code": e.get("employee_code"),
            "name": e.get("name"),
            "group": e.get("employee_type") or "Ungrouped",
            "doj": e.get("doj"),
            "as_of": as_of.isoformat(),
            "service_years": round(years, 2),
            "monthly_basic": round(basic, 2),
            "eligible": "Yes" if eligible else "No",
            "gratuity": gratuity,
        })
    return {
        "columns": [{"key": k, "label": lbl} for k, lbl in _GRATUITY_COLS],
        "rows": rows,
    }


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

async def _build(
    admin: Dict[str, Any],
    report: str,
    company_id: Optional[str],
    year: int,
    period: str,
    month: Optional[int],
    employee_type: Optional[str],
) -> Dict[str, Any]:
    cid = _resolve_company(admin, company_id)
    if report == "contributions":
        months = _months_for(year, period, month)
        data = await _report_contributions(cid, months, employee_type)
    elif report == "leave":
        months = _months_for(year, period, month)
        data = await _report_leave(cid, months, employee_type)
    elif report == "gratuity":
        months = []
        data = await _report_gratuity(cid, employee_type)
    else:
        raise HTTPException(status_code=400, detail="report must be contributions | leave | gratuity")
    company = await db.companies.find_one(
        {"company_id": cid},
        {"_id": 0, "name": 1, "address": 1, "city": 1, "state": 1},
    )
    company = company or {}
    addr_bits = [company.get("address"), company.get("city"), company.get("state")]
    return {
        "report": report,
        "company_id": cid,
        "company_name": company.get("name"),
        "company_address": ", ".join(b for b in addr_bits if b) or None,
        "year": year,
        "period": period,
        "months": months,
        "employee_type": employee_type,
        **data,
    }


@router.get("/compliance")
async def compliance_report(
    report: str = "contributions",
    company_id: Optional[str] = None,
    year: int = 0,
    period: str = "annual",
    month: Optional[int] = None,
    employee_type: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    y = year or date.today().year
    return await _build(admin, report, company_id, y, period, month, employee_type)


@router.get("/compliance.xlsx")
async def compliance_report_xlsx(
    report: str = "contributions",
    company_id: Optional[str] = None,
    year: int = 0,
    period: str = "annual",
    month: Optional[int] = None,
    employee_type: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    y = year or date.today().year
    data = await _build(admin, report, company_id, y, period, month, employee_type)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = {"contributions": "PF-ESIC-PT-TDS", "leave": "Leave Register", "gratuity": "Gratuity"}[report]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1E3A8A")
    cols = data["columns"]
    ncols = max(len(cols), 3)
    firm = data.get("company_name") or data["company_id"]

    # -- Letterhead ---------------------------------------------------------
    def _merged(row: int, value: str, font: Font):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=value)
        c.font = font
        c.alignment = Alignment(horizontal="center")

    _merged(1, firm, Font(bold=True, size=16, color="1E3A8A"))
    if data.get("company_address"):
        _merged(2, data["company_address"], Font(size=10, color="475569"))
    _merged(3, (
        f"{ws.title} — {period.upper()} {y}"
        + (f" (month {month})" if period == "monthly" and month else "")
        + (f" — Group: {employee_type}" if employee_type else "")
    ), Font(bold=True, size=12))
    _merged(4, f"Generated on {date.today().strftime('%d-%m-%Y')} · S.K. Sharma & Co.",
            Font(size=9, italic=True, color="94A3B8"))

    # -- Table --------------------------------------------------------------
    head_row = 6
    for ci, c in enumerate(cols, start=1):
        cell = ws.cell(row=head_row, column=ci, value=c["label"])
        cell.font = hdr_font
        cell.fill = hdr_fill
        ws.column_dimensions[cell.column_letter].width = 18
    for ri, r in enumerate(data["rows"], start=head_row + 1):
        for ci, c in enumerate(cols, start=1):
            ws.cell(row=ri, column=ci, value=r.get(c["key"]))

    # -- Signature block ----------------------------------------------------
    sig_row = head_row + len(data["rows"]) + 3
    sig_col = max(ncols - 1, 1)
    ws.cell(row=sig_row, column=sig_col, value=f"For {firm}").font = Font(bold=True, size=11)
    ws.cell(row=sig_row + 3, column=sig_col, value="Authorised Signatory").font = Font(size=10, color="475569")

    # Contributions: month-wise sheet too
    if report == "contributions" and data.get("month_totals"):
        ws2 = wb.create_sheet("Month-wise Totals")
        ws2.cell(row=1, column=1, value="Month").font = hdr_font
        ws2.cell(row=1, column=1).fill = hdr_fill
        for ci, (k, lbl) in enumerate(_CONTRIB_KEYS, start=2):
            c = ws2.cell(row=1, column=ci, value=lbl)
            c.font = hdr_font
            c.fill = hdr_fill
            ws2.column_dimensions[c.column_letter].width = 18
        for ri, (m, t) in enumerate(sorted(data["month_totals"].items()), start=2):
            ws2.cell(row=ri, column=1, value=m)
            for ci, (k, _) in enumerate(_CONTRIB_KEYS, start=2):
                ws2.cell(row=ri, column=ci, value=t.get(k, 0.0))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="Compliance_{report}_{period}_{y}.xlsx"'},
    )
