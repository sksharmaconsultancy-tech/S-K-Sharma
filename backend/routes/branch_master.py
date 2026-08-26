"""Iter 737 — BRANCH MASTER (complete enhancement, user 25-section spec).

Single source of truth stays ``db.branches`` (same collection used by the
standalone Branches screen, geofence punching, branch-management hub and
branch-extras compliance reports — NO duplicate branch data).

This module ONLY adds master-data fields + mapping/config sections and
read-only dashboards built from EXISTING system data. It does NOT touch:
attendance calculation, duty hours, compliance salary, actual salary,
F&F math, or any report calculation engine.

New nested config sections stored on the branch document:
  * compliance_config  — PT / LWF / Min-Wage / PF / ESIC / Establishment
  * payroll_config     — branch-level payroll DEFAULTS (mapping only)
  * attendance_config  — branch-level attendance DEFAULTS
                         (cross_midnight defaults to True for NEW branches)

Employee ⇄ branch mapping reuses users.home_branch_id and the existing
``branch_transfers`` collection (effective-dated, history never rewritten
— same lazy applier as routes/branch_management.py).
Documents live in ``branch_documents`` (base64, replaced docs kept as
history — mirrors the firm compliance-doc architecture).
Audit trail goes to the existing ``branch_audit`` collection with
field-level old → new values.
"""
import base64
import re
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
    now_iso,
)
from routes.branch_management import _apply_due_transfers  # noqa: E402

router = APIRouter(prefix="/api/admin/branch-master", tags=["branch-master"])

_ADMIN_ROLES = ["super_admin", "sub_admin", "company_admin"]

BRANCH_TYPES = ["Head Office", "Branch", "Factory", "Site",
                "Warehouse", "Remote Office", "Other"]

DOC_TYPES = ["Shops & Establishment Certificate", "Labour License",
             "PF Registration", "ESIC Registration", "PT Registration",
             "LWF Registration", "Factory License", "Rent Agreement", "Other"]

# Flat editable fields (top-level on the branch doc).
_BASIC_FIELDS = ("name", "code", "branch_type", "active", "manager_name",
                 "contact_person", "mobile", "email")
_ADDR_FIELDS = ("address1", "address2", "area", "city", "district",
                "state", "pin_code", "country")
_GPS_FIELDS = ("office_lat", "office_lng", "geofence_enabled",
               "geofence_radius_m", "allow_punch_inside", "gps_accuracy_m")
_CONFIG_SECTIONS = ("compliance_config", "payroll_config", "attendance_config")

_ALL_FLAT = _BASIC_FIELDS + _ADDR_FIELDS + _GPS_FIELDS


def _ist_today() -> str:
    return (datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)).strftime("%Y-%m-%d")


async def _gate(authorization: Optional[str], company_id: Optional[str]) -> tuple[dict, str]:
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    cid = admin.get("company_id") if admin["role"] == "company_admin" else company_id
    if not cid:
        raise HTTPException(status_code=400, detail="company_id required")
    return admin, cid


async def _get_branch(branch_id: str, admin: dict) -> dict:
    b = await db.branches.find_one({"branch_id": branch_id}, {"_id": 0})
    if not b:
        raise HTTPException(status_code=404, detail="Branch not found")
    if admin["role"] == "company_admin" and b.get("company_id") != admin.get("company_id"):
        raise HTTPException(status_code=403, detail="Not your branch")
    # Branch RBAC (existing framework — sub-admins restricted to branches).
    scope = admin.get("branch_admin_branch_ids") or []
    if scope and branch_id not in [str(s) for s in scope]:
        raise HTTPException(status_code=403, detail="Branch not in your scope")
    return b


def _validate(patch: Dict[str, Any]) -> None:
    """Master-data validation with clear messages."""
    if "name" in patch and not str(patch["name"] or "").strip():
        raise HTTPException(status_code=400, detail="Branch Name is required")
    if patch.get("branch_type") and patch["branch_type"] not in BRANCH_TYPES:
        raise HTTPException(status_code=400,
                            detail=f"Branch Type must be one of: {', '.join(BRANCH_TYPES)}")
    pin = str(patch.get("pin_code") or "").strip()
    if pin and not re.fullmatch(r"\d{6}", pin):
        raise HTTPException(status_code=400, detail="PIN Code must be exactly 6 digits")
    email = str(patch.get("email") or "").strip()
    if email and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
        raise HTTPException(status_code=400, detail="Branch Email is not a valid email address")
    if "office_lat" in patch and patch["office_lat"] is not None:
        try:
            lat = float(patch["office_lat"])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Latitude must be a number")
        if not -90 <= lat <= 90:
            raise HTTPException(status_code=400, detail="Latitude must be between -90 and 90")
    if "office_lng" in patch and patch["office_lng"] is not None:
        try:
            lng = float(patch["office_lng"])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Longitude must be a number")
        if not -180 <= lng <= 180:
            raise HTTPException(status_code=400, detail="Longitude must be between -180 and 180")
    if "geofence_radius_m" in patch and patch["geofence_radius_m"] is not None:
        try:
            rad = float(patch["geofence_radius_m"])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Geofence radius must be a number")
        if rad <= 0:
            raise HTTPException(status_code=400, detail="Geofence radius must be positive")


async def _check_code_unique(code: str, company_id: str, exclude_branch_id: Optional[str]) -> None:
    q: dict = {"code": code,
               "$or": [{"company_id": company_id}, {"linked_company_ids": company_id}]}
    if exclude_branch_id:
        q["branch_id"] = {"$ne": exclude_branch_id}
    dup = await db.branches.find_one(q, {"_id": 0, "branch_id": 1, "name": 1})
    if dup:
        raise HTTPException(
            status_code=409,
            detail=f"Branch Code '{code}' is already used by '{dup.get('name')}' in this firm")


async def _audit(admin: dict, branch_id: str, action: str,
                 changes: List[dict]) -> None:
    if not changes:
        return
    await db.branch_audit.insert_one({
        "audit_id": f"bra_{uuid.uuid4().hex[:10]}",
        "action": action, "branch_id": branch_id,
        "by": admin["user_id"], "by_name": admin.get("name") or admin.get("email"),
        "at": now_iso(), "changes": changes,
    })


def _diff(old: dict, patch: dict) -> List[dict]:
    out = []
    for k, v in patch.items():
        if k in ("updated_at",):
            continue
        ov = old.get(k)
        if isinstance(v, dict) and isinstance(ov, dict):
            for sk, sv in v.items():
                if ov.get(sk) != sv:
                    out.append({"field": f"{k}.{sk}", "old": ov.get(sk), "new": sv})
        elif ov != v:
            out.append({"field": k, "old": ov, "new": v})
    return out


# ---------------------------------------------------------------- list
@router.get("/list")
async def bm_list(company_id: Optional[str] = Query(None),
                  search: Optional[str] = Query(None),
                  status: Optional[str] = Query(None),      # active|inactive
                  state: Optional[str] = Query(None),
                  branch_type: Optional[str] = Query(None),
                  authorization: Optional[str] = Header(None)):
    """Branch list + cached-cheap summary counts (employees / present /
    absent today). One users query + one attendance query — no per-branch
    API calls."""
    admin, cid = await _gate(authorization, company_id)
    q: dict = {"$or": [{"company_id": cid}, {"linked_company_ids": cid}]}
    branches = await db.branches.find(q, {"_id": 0}).sort("created_at", 1).to_list(500)
    scope = admin.get("branch_admin_branch_ids") or []
    if scope:
        scope = [str(s) for s in scope]
        branches = [b for b in branches if b["branch_id"] in scope]
    # server-side filters
    if status in ("active", "inactive"):
        want = status == "active"
        branches = [b for b in branches if bool(b.get("active", True)) == want]
    if state:
        branches = [b for b in branches if (b.get("state") or "") == state]
    if branch_type:
        branches = [b for b in branches if (b.get("branch_type") or "") == branch_type]
    if search:
        s = search.strip().lower()
        branches = [b for b in branches if s in (b.get("name") or "").lower()
                    or s in (b.get("code") or "").lower()
                    or s in (b.get("city") or "").lower()]
    # summary counts (single pass)
    today = _ist_today()
    punched = set()
    async for a in db.attendance.find({"company_id": cid, "date": today, "kind": "in"},
                                      {"_id": 0, "user_id": 1}):
        punched.add(a["user_id"])
    emp_total: Dict[str, int] = {}
    emp_active: Dict[str, int] = {}
    present: Dict[str, int] = {}
    async for u in db.users.find({"company_id": cid, "role": "employee"},
                                 {"_id": 0, "user_id": 1, "home_branch_id": 1, "status": 1}):
        hb = u.get("home_branch_id") or ""
        emp_total[hb] = emp_total.get(hb, 0) + 1
        if u.get("status") != "inactive":
            emp_active[hb] = emp_active.get(hb, 0) + 1
            if u["user_id"] in punched:
                present[hb] = present.get(hb, 0) + 1
    for b in branches:
        bid = b["branch_id"]
        b["emp_count"] = emp_total.get(bid, 0)
        b["active_employees"] = emp_active.get(bid, 0)
        b["present_today"] = present.get(bid, 0)
        b["absent_today"] = max(0, emp_active.get(bid, 0) - present.get(bid, 0))
    return {"branches": branches, "branch_types": BRANCH_TYPES,
            "unassigned_employees": emp_total.get("", 0)}


# ---------------------------------------------------------------- create
@router.post("/create")
async def bm_create(body: Dict[str, Any] = Body(...),
                    authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    cid = admin.get("company_id") if admin["role"] == "company_admin" \
        else str(body.get("company_id") or "")
    if not cid:
        raise HTTPException(status_code=400, detail="company_id required")
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0, "company_id": 1})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    patch = {k: body.get(k) for k in _ALL_FLAT if k in body}
    patch["name"] = str(body.get("name") or "").strip()
    _validate(patch)
    if not patch["name"]:
        raise HTTPException(status_code=400, detail="Branch Name is required")
    code = str(patch.get("code") or "").strip().upper()
    if code:
        patch["code"] = code
        await _check_code_unique(code, cid, None)

    branch_id = f"br_{uuid.uuid4().hex[:10]}"
    doc = {
        "branch_id": branch_id, "company_id": cid,
        "branch_type": body.get("branch_type") or "Branch",
        "active": True,
        "address": (str(body.get("address1") or "").strip() or None),  # legacy field kept in sync
        "geofence_radius_m": body.get("geofence_radius_m") or 200,
        "geofence_enabled": bool(body.get("geofence_enabled", True)),
        "allow_punch_inside": bool(body.get("allow_punch_inside", True)),
        # Cross-midnight ON by default for NEW branches (user spec §8) —
        # mapping/default only, attendance engine untouched.
        "attendance_config": {"cross_midnight": True,
                              **(body.get("attendance_config") or {})},
        "compliance_config": body.get("compliance_config") or {},
        "payroll_config": body.get("payroll_config") or {},
        "created_at": now_iso(), "created_by_user_id": admin["user_id"],
        "created_by_name": admin.get("name") or admin.get("email"),
    }
    doc.update({k: v for k, v in patch.items() if v is not None})
    await db.branches.insert_one(doc)
    doc.pop("_id", None)
    await _audit(admin, branch_id, "branch_create",
                 [{"field": "branch", "old": None, "new": doc.get("name")}])
    return {"ok": True, "branch": doc}


# ---------------------------------------------------------------- detail
@router.get("/{branch_id}")
async def bm_detail(branch_id: str, authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    b = await _get_branch(branch_id, admin)
    audit = await db.branch_audit.find(
        {"branch_id": branch_id}, {"_id": 0}).sort("at", -1).to_list(30)
    return {"branch": b, "audit": audit, "branch_types": BRANCH_TYPES,
            "doc_types": DOC_TYPES}


@router.patch("/{branch_id}")
async def bm_patch(branch_id: str, body: Dict[str, Any] = Body(...),
                   authorization: Optional[str] = Header(None)):
    """Update any master section. Nested configs are MERGED (partial save).
    Field-level audit old → new."""
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    b = await _get_branch(branch_id, admin)

    patch: Dict[str, Any] = {k: body[k] for k in _ALL_FLAT if k in body}
    _validate(patch)
    if "name" in patch:
        patch["name"] = str(patch["name"]).strip()
    if "code" in patch:
        code = str(patch.get("code") or "").strip().upper()
        patch["code"] = code or None
        if code:
            await _check_code_unique(code, b["company_id"], branch_id)
    if "address1" in patch:  # keep legacy single-line address in sync
        parts = [str(body.get(k) or "").strip()
                 for k in ("address1", "address2", "area", "city")]
        patch["address"] = ", ".join([p for p in parts if p]) or None
    # merge nested config sections
    for sec in _CONFIG_SECTIONS:
        if sec in body and isinstance(body[sec], dict):
            merged = dict(b.get(sec) or {})
            merged.update(body[sec])
            patch[sec] = merged
    # document-date sanity used by establishment section
    cc = patch.get("compliance_config") or {}
    rd, ed = str(cc.get("sne_regn_date") or ""), str(cc.get("sne_expiry_date") or "")
    if rd and ed and ed < rd:
        raise HTTPException(status_code=400,
                            detail="Registration Expiry Date cannot be before Registration Date")
    if not patch:
        return {"ok": True, "branch": b}
    changes = _diff(b, patch)
    patch["updated_at"] = now_iso()
    patch["updated_by_name"] = admin.get("name") or admin.get("email")
    await db.branches.update_one({"branch_id": branch_id}, {"$set": patch})
    # Branch renamed? keep the report dimension users.branch_name in sync
    # for CURRENT home employees (old run rows keep their snapshots).
    if "name" in patch and patch["name"] and patch["name"] != b.get("name"):
        await db.users.update_many(
            {"home_branch_id": branch_id},
            {"$set": {"branch_name": patch["name"]}})
    await _audit(admin, branch_id, "branch_update", changes)
    fresh = await db.branches.find_one({"branch_id": branch_id}, {"_id": 0})
    return {"ok": True, "branch": fresh}


# ---------------------------------------------------------------- dashboard
@router.get("/{branch_id}/dashboard")
async def bm_branch_dashboard(branch_id: str,
                              authorization: Optional[str] = Header(None)):
    """Compact per-branch dashboard from EXISTING system data — no new
    calculation engine. Count queries only (fast)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    b = await _get_branch(branch_id, admin)
    cid = b["company_id"]
    today = _ist_today()
    month = today[:7]

    uids_all: List[str] = []
    uids_active: List[str] = []
    joiners = exits = 0
    async for u in db.users.find({"company_id": cid, "role": "employee",
                                  "home_branch_id": branch_id},
                                 {"_id": 0, "user_id": 1, "status": 1, "doj": 1,
                                  "date_of_joining": 1, "dol": 1, "date_of_leaving": 1}):
        uids_all.append(u["user_id"])
        if u.get("status") != "inactive":
            uids_active.append(u["user_id"])
        doj = str(u.get("doj") or u.get("date_of_joining") or "")
        dol = str(u.get("dol") or u.get("date_of_leaving") or "")
        if doj.startswith(month):
            joiners += 1
        if dol.startswith(month):
            exits += 1

    present = len(await db.attendance.distinct(
        "user_id", {"company_id": cid, "date": today, "kind": "in",
                    "user_id": {"$in": uids_active}})) if uids_active else 0
    on_leave = await db.leaves.count_documents(
        {"user_id": {"$in": uids_active}, "status": "approved",
         "from_date": {"$lte": today}, "to_date": {"$gte": today}}) if uids_active else 0
    on_duty = await db.branch_temp_assignments.count_documents(
        {"user_id": {"$in": uids_all}, "status": "approved",
         "from_date": {"$lte": today}, "to_date": {"$gte": today}}) if uids_all else 0
    pending_leaves = await db.leaves.count_documents(
        {"user_id": {"$in": uids_all}, "status": "pending"}) if uids_all else 0
    pending_advances = await db.advances.count_documents(
        {"user_id": {"$in": uids_all}, "status": "pending"}) if uids_all else 0
    open_fnf = await db.fnf_settlements.count_documents(
        {"user_id": {"$in": uids_all},
         "status": {"$nin": ["paid", "cancelled", "rejected"]}}) if uids_all else 0
    run = await db.compliance_salary_runs.find_one(
        {"company_id": cid, "month": month}, {"_id": 0, "status": 1},
        sort=[("generated_at", -1)])
    # Iter 739 — statutory document counts for the dashboard card
    warn = await _warn_days(cid)
    soon = (datetime.now(timezone.utc) + timedelta(days=warn, hours=5, minutes=30)).strftime("%Y-%m-%d")
    docs = await db.branch_documents.find(
        {"branch_id": branch_id, "status": {"$nin": ["deleted", "replaced"]}},
        {"_id": 0, "expiry_date": 1, "no_expiry": 1, "applicable": 1, "status": 1}).to_list(500)
    st_counts = {"active": 0, "expiring_soon": 0, "expired": 0}
    for d in docs:
        s = _doc_status(d, today, soon)
        if s in st_counts:
            st_counts[s] += 1

    active = len(uids_active)
    absent = max(0, active - present - on_leave)
    cc = b.get("compliance_config") or {}
    compliance_ok = bool(b.get("state")) and (
        cc.get("pt_applicable") is not None or cc.get("pf_applicable") is not None)
    return {"branch_id": branch_id, "date": today, "month": month,
            "total_employees": len(uids_all), "active_employees": active,
            "inactive_employees": len(uids_all) - active,
            "present_today": present, "absent_today": absent,
            "on_leave": on_leave, "on_duty": on_duty,
            "attendance_pct": round(present / active * 100, 1) if active else 0,
            "payroll_status": (run or {}).get("status") or "not_generated",
            "compliance_status": "configured" if compliance_ok
            else ("state_set" if b.get("state") else "pending"),
            "pending_approvals": pending_leaves + pending_advances,
            "pending_leaves": pending_leaves, "pending_advances": pending_advances,
            "new_joiners": joiners, "exits": exits, "open_fnf": open_fnf,
            "statutory_docs": st_counts}


# ---------------------------------------------------------------- employees
@router.get("/{branch_id}/employees")
async def bm_branch_employees(branch_id: str,
                              q: Optional[str] = Query(None),
                              page: int = Query(1, ge=1),
                              page_size: int = Query(50, ge=1, le=200),
                              authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    b = await _get_branch(branch_id, admin)
    query: dict = {"company_id": b["company_id"], "role": "employee",
                   "home_branch_id": branch_id}
    if q:
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        query["$or"] = [{"name": rx}, {"employee_code": rx}]
    total = await db.users.count_documents(query)
    emps = await db.users.find(query, {
        "_id": 0, "user_id": 1, "name": 1, "employee_code": 1, "status": 1,
        "designation": 1, "department": 1, "doj": 1, "date_of_joining": 1,
        "mobile": 1, "phone": 1,
    }).sort("name", 1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    return {"employees": emps, "total": total, "page": page, "page_size": page_size}


@router.get("/{branch_id}/employees-export")
async def bm_branch_employees_export(branch_id: str,
                                     authorization: Optional[str] = Header(None)):
    import io
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    b = await _get_branch(branch_id, admin)
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append([f"Branch Employees — {b.get('name')} ({b.get('code') or b['branch_id']})"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["Code", "Name", "Designation", "Department", "DOJ", "Mobile", "Status"])
    for c in ws[3]:
        c.font = Font(bold=True)
    async for u in db.users.find({"company_id": b["company_id"], "role": "employee",
                                  "home_branch_id": branch_id},
                                 {"_id": 0}).sort("name", 1):
        ws.append([u.get("employee_code"), u.get("name"), u.get("designation"),
                   u.get("department"), u.get("doj") or u.get("date_of_joining"),
                   u.get("mobile") or u.get("phone"),
                   "Inactive" if u.get("status") == "inactive" else "Active"])
    out = io.BytesIO()
    wb.save(out)
    fname = f"Branch_Employees_{(b.get('code') or b.get('name') or 'branch').replace(' ', '_')}.xlsx"
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------------------------------------------------------- transfer
@router.post("/transfer")
async def bm_transfer(body: Dict[str, Any] = Body(...),
                      authorization: Optional[str] = Header(None)):
    """Single or BULK controlled transfer. Writes to the existing
    ``branch_transfers`` collection (effective-dated; same lazy applier as
    the branch-management hub) → history is created automatically and old
    payroll/attendance records are NEVER rewritten."""
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    uids = [str(u) for u in (body.get("user_ids") or []) if u]
    if body.get("user_id"):
        uids.append(str(body["user_id"]))
    uids = list(dict.fromkeys(uids))
    new_b = str(body.get("new_branch_id") or "")
    eff = str(body.get("effective_date") or "")
    if not (uids and new_b and eff):
        raise HTTPException(status_code=400,
                            detail="user_ids, new_branch_id and effective_date are required")
    nb = await db.branches.find_one({"branch_id": new_b}, {"_id": 0, "company_id": 1,
                                                           "name": 1, "active": 1})
    if not nb:
        raise HTTPException(status_code=404, detail="Target branch not found")
    if nb.get("active") is False:
        raise HTTPException(status_code=400, detail="Target branch is Inactive — activate it first")
    created = []
    for uid in uids:
        emp = await db.users.find_one({"user_id": uid},
                                      {"_id": 0, "company_id": 1, "name": 1,
                                       "home_branch_id": 1})
        if not emp:
            continue
        if admin["role"] == "company_admin" and emp.get("company_id") != admin.get("company_id"):
            continue
        if emp.get("home_branch_id") == new_b:
            continue  # already there
        doc = {
            "transfer_id": f"btr_{uuid.uuid4().hex[:10]}",
            "company_id": emp["company_id"], "user_id": uid,
            "employee_name": emp.get("name"),
            "prev_branch_id": emp.get("home_branch_id"),
            "new_branch_id": new_b, "effective_date": eff,
            "reason": str(body.get("reason") or "").strip() or None,
            "remarks": str(body.get("remarks") or "").strip() or None,
            "approved_by": admin["user_id"], "status": "pending",
            "created_at": now_iso(),
        }
        await db.branch_transfers.insert_one(doc)
        doc.pop("_id", None)
        created.append(doc)
        await _audit(admin, new_b, "employee_transfer",
                     [{"field": f"employee:{emp.get('name') or uid}",
                       "old": emp.get("home_branch_id"), "new": new_b}])
    if created:
        await _apply_due_transfers(created[0]["company_id"])
    return {"ok": True, "created": len(created), "transfers": created}


# ---------------------------------------------------------------- history
@router.get("/{branch_id}/history")
async def bm_branch_history(branch_id: str,
                            authorization: Optional[str] = Header(None)):
    """Transfers in / out of this branch (never deleted)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    await _get_branch(branch_id, admin)
    rows = await db.branch_transfers.find(
        {"$or": [{"prev_branch_id": branch_id}, {"new_branch_id": branch_id}]},
        {"_id": 0}).sort("created_at", -1).to_list(300)
    names = {b["branch_id"]: b.get("name") for b in await db.branches.find(
        {}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(500)}
    for r in rows:
        r["prev_branch_name"] = names.get(r.get("prev_branch_id")) or "Main / Unassigned"
        r["new_branch_name"] = names.get(r.get("new_branch_id")) or "-"
        r["direction"] = "IN" if r.get("new_branch_id") == branch_id else "OUT"
    return {"history": rows}


@router.get("/employee-history/{user_id}")
async def bm_employee_history(user_id: str,
                              authorization: Optional[str] = Header(None)):
    """Employee → Branch History timeline (effective from/to) built from
    the immutable branch_transfers register."""
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    emp = await db.users.find_one({"user_id": user_id},
                                  {"_id": 0, "company_id": 1, "name": 1,
                                   "home_branch_id": 1, "doj": 1, "date_of_joining": 1})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    if admin["role"] == "company_admin" and emp.get("company_id") != admin.get("company_id"):
        raise HTTPException(status_code=403, detail="Not your employee")
    names = {b["branch_id"]: b.get("name") for b in await db.branches.find(
        {}, {"_id": 0, "branch_id": 1, "name": 1}).to_list(500)}
    transfers = await db.branch_transfers.find(
        {"user_id": user_id, "status": {"$ne": "cancelled"}},
        {"_id": 0}).sort("effective_date", 1).to_list(200)
    timeline = []
    start = str(emp.get("doj") or emp.get("date_of_joining") or "") or None
    prev_branch = transfers[0].get("prev_branch_id") if transfers \
        else emp.get("home_branch_id")
    cur_from = start
    for t in transfers:
        eff = t.get("effective_date")
        timeline.append({
            "from": cur_from, "to": eff,
            "branch_id": prev_branch,
            "branch": names.get(prev_branch) or "Main / Unassigned",
            "transfer_reason": None,
        })
        prev_branch = t.get("new_branch_id")
        cur_from = eff
        timeline[-1]["next_reason"] = t.get("reason")
    timeline.append({"from": cur_from, "to": None,
                     "branch_id": prev_branch,
                     "branch": names.get(prev_branch) or "Main / Unassigned"})
    return {"employee": {"user_id": user_id, "name": emp.get("name")},
            "timeline": timeline, "transfers": [
                {**t, "prev_branch_name": names.get(t.get("prev_branch_id")) or "Main / Unassigned",
                 "new_branch_name": names.get(t.get("new_branch_id")) or "-"}
                for t in transfers]}


# ---------------------------------------------------------------- documents
@router.get("/{branch_id}/documents")
async def bm_docs_list(branch_id: str, authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    await _get_branch(branch_id, admin)
    today = _ist_today()
    soon = (datetime.now(timezone.utc) + timedelta(days=60)).strftime("%Y-%m-%d")
    docs = await db.branch_documents.find(
        {"branch_id": branch_id, "status": {"$ne": "deleted"}},
        {"_id": 0, "file_base64": 0}).sort("created_at", -1).to_list(200)
    for d in docs:
        exp = str(d.get("expiry_date") or "")
        if d.get("status") == "replaced":
            d["expiry_status"] = "replaced"
        elif exp and exp < today:
            d["expiry_status"] = "expired"
        elif exp and exp <= soon:
            d["expiry_status"] = "expiring_soon"
        else:
            d["expiry_status"] = "active"
    return {"documents": docs, "doc_types": DOC_TYPES}


@router.post("/{branch_id}/documents")
async def bm_docs_add(branch_id: str, body: Dict[str, Any] = Body(...),
                      authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    await _get_branch(branch_id, admin)
    doc_type = str(body.get("doc_type") or "").strip()
    if not doc_type:
        raise HTTPException(status_code=400, detail="Document Type is required")
    issue, exp = str(body.get("issue_date") or ""), str(body.get("expiry_date") or "")
    if issue and exp and exp < issue:
        raise HTTPException(status_code=400,
                            detail="Expiry date cannot be before issue date")
    f64 = body.get("file_base64")
    if f64:
        try:
            raw = base64.b64decode(f64.split(",")[-1], validate=False)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid file data")
        if len(raw) > 8 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large (max 8 MB)")
    # replacing an older doc of the same type? keep history — mark replaced.
    if body.get("replace_same_type"):
        await db.branch_documents.update_many(
            {"branch_id": branch_id, "doc_type": doc_type,
             "status": {"$nin": ["deleted", "replaced"]}},
            {"$set": {"status": "replaced", "replaced_at": now_iso(),
                      "replaced_by": admin["user_id"]}})
    doc = {
        "doc_id": f"brdoc_{uuid.uuid4().hex[:10]}", "branch_id": branch_id,
        "doc_type": doc_type,
        "doc_number": str(body.get("doc_number") or "").strip() or None,
        "issue_date": issue or None, "expiry_date": exp or None,
        "remarks": str(body.get("remarks") or "").strip() or None,
        "file_name": str(body.get("file_name") or "").strip() or None,
        "file_base64": f64 or None,
        "status": "active",
        "created_at": now_iso(),
        "created_by": admin["user_id"],
        "created_by_name": admin.get("name") or admin.get("email"),
    }
    await db.branch_documents.insert_one(doc)
    doc.pop("_id", None)
    doc.pop("file_base64", None)
    await _audit(admin, branch_id, "document_add",
                 [{"field": f"document:{doc_type}", "old": None,
                   "new": doc.get("doc_number") or doc.get("file_name") or "added"}])
    return {"ok": True, "document": doc}


@router.get("/documents/{doc_id}/file")
async def bm_docs_file(doc_id: str, authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    d = await db.branch_documents.find_one({"doc_id": doc_id}, {"_id": 0})
    if not d or not d.get("file_base64"):
        raise HTTPException(status_code=404, detail="File not found")
    await _get_branch(d["branch_id"], admin)
    return {"file_name": d.get("file_name") or f"{d.get('doc_type')}.pdf",
            "file_base64": d["file_base64"]}


@router.delete("/documents/{doc_id}")
async def bm_docs_delete(doc_id: str, authorization: Optional[str] = Header(None)):
    """Soft delete — document history is preserved."""
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    d = await db.branch_documents.find_one({"doc_id": doc_id}, {"_id": 0, "branch_id": 1,
                                                                "doc_type": 1})
    if not d:
        raise HTTPException(status_code=404, detail="Document not found")
    await _get_branch(d["branch_id"], admin)
    await db.branch_documents.update_one(
        {"doc_id": doc_id},
        {"$set": {"status": "deleted", "deleted_at": now_iso(),
                  "deleted_by": admin["user_id"]}})
    await _audit(admin, d["branch_id"], "document_delete",
                 [{"field": f"document:{d.get('doc_type')}", "old": "active",
                   "new": "deleted"}])
    return {"ok": True}


# ═══════════════════ Iter 739 — LICENSES & STATUTORY COMPLIANCE ═══════════════════
# Branch-wise statutory registrations/licenses with categories, effective/
# expiry dates, applicability, attachments (history kept), configurable
# expiry warning, compliance-wise summary + alerts, document-type MASTER
# (configurable, state-aware) and a firm-wide register report.
# NO change to any payroll/compliance calculation — master data only.

STATUTORY_CATALOG: Dict[str, List[str]] = {
    "EPFO / PF": [
        "PF Registration Certificate", "PF Code / Establishment Letter",
        "PF Registration Details", "PF Exemption Certificate",
        "PF Trust Registration", "PF Authorization / Approval Letter",
        "PF Amendment / Change Letter", "PF Other Document"],
    "ESIC": [
        "ESIC Registration Certificate", "ESIC Employer Code Letter",
        "ESIC Registration Details", "ESIC Exemption Certificate",
        "ESIC Amendment / Change Letter", "ESIC Other Document"],
    "Professional Tax (PT)": [
        "PT Registration Certificate", "PT Enrollment Certificate",
        "PT Employer Registration", "PT Registration Number Letter",
        "PT Amendment / Change Letter", "PT Exemption Certificate",
        "PT Other Document"],
    "Labour Welfare Fund (LWF)": [
        "LWF Registration Certificate", "LWF Registration / Code Letter",
        "LWF Enrollment Certificate", "LWF Exemption Certificate",
        "LWF Amendment / Change Letter", "LWF Other Document"],
    "Shops & Establishment": [
        "Shops & Establishment Registration Certificate", "Renewal Certificate",
        "Amendment Certificate", "Registration Approval Letter",
        "Establishment License", "Exemption Certificate", "Other Document"],
    "Minimum Wages": [
        "Minimum Wage Notification", "State Minimum Wage Notification",
        "Minimum Wage Order", "Wage Revision Notification",
        "Minimum Wage Category / Schedule", "Zone / Area Notification",
        "Other Wage Notification"],
    "Labour Department": [
        "Labour License", "Labour Registration Certificate",
        "Labour Department Registration", "Labour Contractor License",
        "Principal Employer Registration", "Renewal Certificate",
        "Amendment Certificate", "Exemption Certificate",
        "Labour Department Order", "Other Labour Document"],
    "Factory Compliance": [
        "Factory License", "Factory Registration Certificate",
        "Factory License Renewal", "Factory Amendment Certificate",
        "Factory Plan Approval", "Occupier Approval",
        "Factory Inspector Approval", "Factory Exemption Certificate",
        "Other Factory Document"],
    "Contract Labour": [
        "Principal Employer Registration", "Contractor License",
        "Contract Labour Registration", "Contractor Registration Certificate",
        "Renewal Certificate", "Amendment Certificate",
        "Exemption Certificate", "Other Contract Labour Document"],
    "Gratuity": [
        "Gratuity Registration", "Gratuity Insurance Policy",
        "Gratuity Trust Deed", "Gratuity Trust Registration",
        "Gratuity Exemption Approval", "Gratuity Scheme",
        "Other Gratuity Document"],
    "Bonus": [
        "Bonus Registration", "Bonus Exemption Certificate",
        "Bonus Scheme / Approval", "Other Bonus Document"],
    "Maternity Benefit": [
        "Maternity Benefit Registration", "Maternity Benefit Approval",
        "Maternity Benefit Exemption", "Other Maternity Benefit Document"],
    "Employment / Labour Registrations": [
        "Labour Department Registration", "Employment Exchange Registration",
        "Labour Welfare Registration", "Worker Registration",
        "Establishment Registration", "Renewal Certificate",
        "Amendment Certificate", "Other Registration"],
    "Local / Municipal Compliance": [
        "Trade License", "Municipal Registration",
        "Commercial Establishment License", "Fire NOC",
        "Building / Occupancy Certificate", "Local Authority Registration",
        "Pollution / Environmental Approval", "Other Local License"],
    "Other Statutory Compliance": ["Other Statutory Compliance Document"],
}

# Compliance-summary mapping: category → compliance_config applicability key
_COMPLIANCE_MAP = [
    ("PF", "EPFO / PF", "pf_applicable"),
    ("ESIC", "ESIC", "esic_applicable"),
    ("PT", "Professional Tax (PT)", "pt_applicable"),
    ("LWF", "Labour Welfare Fund (LWF)", "lwf_applicable"),
    ("Shops & Establishment", "Shops & Establishment", None),
    ("Labour License", "Labour Department", None),
    ("Factory License", "Factory Compliance", None),
    ("Contract Labour", "Contract Labour", None),
    ("Minimum Wages", "Minimum Wages", None),
    ("Other", "Other Statutory Compliance", None),
]


async def _warn_days(company_id: str) -> int:
    c = await db.companies.find_one({"company_id": company_id},
                                    {"_id": 0, "statutory_alert_days": 1})
    d = int((c or {}).get("statutory_alert_days") or 60)
    return d if d in (30, 60, 90) else 60


def _doc_status(d: dict, today: str, soon: str) -> str:
    if d.get("status") == "replaced":
        return "replaced"
    if d.get("applicable") is False:
        return "not_applicable"
    exp = str(d.get("expiry_date") or "")
    if d.get("no_expiry") or not exp:
        return "active"
    if exp < today:
        return "expired"
    if exp <= soon:
        return "expiring_soon"
    return "active"


@router.get("/statutory/catalog")
async def statutory_catalog(state: Optional[str] = Query(None),
                            company_id: Optional[str] = Query(None),
                            authorization: Optional[str] = Header(None)):
    """Built-in catalog + custom types from statutory_doc_master, with
    active flags and optional state-specific entries."""
    admin, cid = await _gate(authorization, company_id)
    entries = await db.statutory_doc_master.find({}, {"_id": 0}).to_list(2000)
    overrides = {(e.get("category"), e.get("doc_type")): e for e in entries}
    out = []
    for cat, types in STATUTORY_CATALOG.items():
        rows = []
        for t in types:
            ov = overrides.pop((cat, t), None)
            rows.append({"doc_type": t, "active": (ov or {}).get("active", True),
                         "custom": False, "state": (ov or {}).get("state")})
        out.append({"category": cat, "types": rows})
    # custom types (state-filtered: no state OR matching state)
    for (cat, t), e in overrides.items():
        st_val = e.get("state")
        if state and st_val and st_val != state:
            continue
        grp = next((g for g in out if g["category"] == cat), None)
        if not grp:
            grp = {"category": cat, "types": []}
            out.append(grp)
        grp["types"].append({"doc_type": t, "active": e.get("active", True),
                             "custom": True, "state": st_val})
    return {"categories": out, "warn_days": await _warn_days(cid)}


@router.post("/statutory/catalog")
async def statutory_catalog_add(body: Dict[str, Any] = Body(...),
                                authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    cat = str(body.get("category") or "").strip()
    dt = str(body.get("doc_type") or "").strip()
    if not cat or not dt:
        raise HTTPException(status_code=400, detail="category and doc_type are required")
    await db.statutory_doc_master.update_one(
        {"category": cat, "doc_type": dt},
        {"$set": {"category": cat, "doc_type": dt,
                  "state": str(body.get("state") or "").strip() or None,
                  "active": True, "updated_by": admin["user_id"],
                  "updated_at": now_iso()},
         "$setOnInsert": {"created_at": now_iso(), "created_by": admin["user_id"]}},
        upsert=True)
    return {"ok": True}


@router.patch("/statutory/catalog")
async def statutory_catalog_toggle(body: Dict[str, Any] = Body(...),
                                   authorization: Optional[str] = Header(None)):
    """Activate/deactivate a document type. NEVER deletes — historical
    records keep their type."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    cat = str(body.get("category") or "").strip()
    dt = str(body.get("doc_type") or "").strip()
    await db.statutory_doc_master.update_one(
        {"category": cat, "doc_type": dt},
        {"$set": {"category": cat, "doc_type": dt,
                  "active": bool(body.get("active", True)),
                  "updated_by": admin["user_id"], "updated_at": now_iso()}},
        upsert=True)
    return {"ok": True}


@router.post("/statutory/warn-days")
async def statutory_warn_days(body: Dict[str, Any] = Body(...),
                              authorization: Optional[str] = Header(None)):
    admin, cid = await _gate(authorization, body.get("company_id"))
    days = int(body.get("days") or 60)
    if days not in (30, 60, 90):
        raise HTTPException(status_code=400, detail="days must be 30, 60 or 90")
    await db.companies.update_one({"company_id": cid},
                                  {"$set": {"statutory_alert_days": days}})
    return {"ok": True, "days": days}


@router.get("/{branch_id}/licenses")
async def bm_licenses(branch_id: str,
                      search: Optional[str] = Query(None),
                      category: Optional[str] = Query(None),
                      status: Optional[str] = Query(None),
                      doc_type: Optional[str] = Query(None),
                      state: Optional[str] = Query(None),
                      authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    b = await _get_branch(branch_id, admin)
    warn = await _warn_days(b["company_id"])
    today = _ist_today()
    soon = (datetime.now(timezone.utc) + timedelta(days=warn, hours=5, minutes=30)).strftime("%Y-%m-%d")
    docs = await db.branch_documents.find(
        {"branch_id": branch_id, "status": {"$ne": "deleted"}},
        {"_id": 0, "file_base64": 0}).sort("created_at", -1).to_list(500)
    for d in docs:
        d["expiry_status"] = _doc_status(d, today, soon)
        d["has_file"] = bool(d.get("file_name"))
    if category:
        docs = [d for d in docs if (d.get("category") or "") == category]
    if doc_type:
        docs = [d for d in docs if (d.get("doc_type") or "") == doc_type]
    if state:
        docs = [d for d in docs if (d.get("state") or "") == state]
    if status:
        docs = [d for d in docs if d["expiry_status"] == status]
    if search:
        s = search.strip().lower()
        docs = [d for d in docs if s in (d.get("doc_name") or "").lower()
                or s in (d.get("doc_type") or "").lower()
                or s in (d.get("doc_number") or "").lower()
                or s in (d.get("issuing_authority") or "").lower()]
    live = [d for d in docs if d["expiry_status"] not in ("replaced",)]
    summary = {"active": sum(1 for d in live if d["expiry_status"] == "active"),
               "expiring_soon": sum(1 for d in live if d["expiry_status"] == "expiring_soon"),
               "expired": sum(1 for d in live if d["expiry_status"] == "expired"),
               "not_applicable": sum(1 for d in live if d["expiry_status"] == "not_applicable"),
               "total": len(live)}
    return {"documents": docs, "summary": summary, "warn_days": warn}


def _license_validate(body: dict) -> None:
    eff = str(body.get("effective_from") or "")
    exp = str(body.get("expiry_date") or "")
    if eff and exp and exp < eff:
        raise HTTPException(status_code=400,
                            detail="Expiry Date cannot be before Effective From")
    f64 = body.get("file_base64")
    if f64:
        try:
            raw = base64.b64decode(f64.split(",")[-1], validate=False)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid file data")
        if len(raw) > 8 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large (max 8 MB)")


_LICENSE_FIELDS = ("category", "doc_type", "doc_name", "doc_number",
                   "establishment_code", "issuing_authority", "state",
                   "effective_from", "expiry_date", "no_expiry", "applicable",
                   "remarks")


@router.post("/{branch_id}/licenses")
async def bm_license_add(branch_id: str, body: Dict[str, Any] = Body(...),
                         authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    await _get_branch(branch_id, admin)
    if not str(body.get("doc_type") or "").strip():
        raise HTTPException(status_code=400, detail="Document Type is required")
    _license_validate(body)
    # Renewal: keep the previous record as history (spec §19)
    if body.get("replace_same_type"):
        await db.branch_documents.update_many(
            {"branch_id": branch_id, "doc_type": body["doc_type"],
             "category": body.get("category"),
             "status": {"$nin": ["deleted", "replaced"]}},
            {"$set": {"status": "replaced", "replaced_at": now_iso(),
                      "replaced_by": admin["user_id"]}})
    doc = {"doc_id": f"brdoc_{uuid.uuid4().hex[:10]}", "branch_id": branch_id,
           "status": "active", "created_at": now_iso(),
           "created_by": admin["user_id"],
           "created_by_name": admin.get("name") or admin.get("email")}
    for k in _LICENSE_FIELDS:
        v = body.get(k)
        doc[k] = (str(v).strip() or None) if isinstance(v, str) else v
    doc["applicable"] = body.get("applicable", True) is not False
    doc["no_expiry"] = bool(body.get("no_expiry"))
    if body.get("file_base64"):
        doc["file_base64"] = body["file_base64"]
        doc["file_name"] = str(body.get("file_name") or "").strip() or "document"
        doc["uploaded_at"] = now_iso()
        doc["uploaded_by_name"] = admin.get("name") or admin.get("email")
    await db.branch_documents.insert_one(doc)
    doc.pop("_id", None)
    doc.pop("file_base64", None)
    await _audit(admin, branch_id, "license_add",
                 [{"field": f"{doc.get('category') or 'Document'}:{doc.get('doc_type')}",
                   "old": None, "new": doc.get("doc_number") or "added"}])
    return {"ok": True, "document": doc}


@router.patch("/licenses/{doc_id}")
async def bm_license_patch(doc_id: str, body: Dict[str, Any] = Body(...),
                           authorization: Optional[str] = Header(None)):
    """Edit fields and/or REPLACE the attachment (old file kept in
    attachment_history — spec §8)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    d = await db.branch_documents.find_one({"doc_id": doc_id}, {"_id": 0})
    if not d:
        raise HTTPException(status_code=404, detail="Document not found")
    await _get_branch(d["branch_id"], admin)
    _license_validate({**d, **body})
    patch: Dict[str, Any] = {}
    changes = []
    for k in _LICENSE_FIELDS:
        if k in body:
            v = body[k]
            v = (str(v).strip() or None) if isinstance(v, str) else v
            if d.get(k) != v:
                changes.append({"field": k, "old": d.get(k), "new": v})
            patch[k] = v
    if body.get("file_base64"):
        if d.get("file_name"):
            hist = d.get("attachment_history") or []
            hist.append({"file_name": d.get("file_name"),
                         "uploaded_at": d.get("uploaded_at"),
                         "uploaded_by_name": d.get("uploaded_by_name"),
                         "replaced_at": now_iso(),
                         "replaced_by_name": admin.get("name") or admin.get("email")})
            patch["attachment_history"] = hist[-10:]
            changes.append({"field": "attachment", "old": d.get("file_name"),
                            "new": body.get("file_name")})
        patch["file_base64"] = body["file_base64"]
        patch["file_name"] = str(body.get("file_name") or "").strip() or "document"
        patch["uploaded_at"] = now_iso()
        patch["uploaded_by_name"] = admin.get("name") or admin.get("email")
    if not patch:
        return {"ok": True}
    patch["updated_at"] = now_iso()
    patch["updated_by_name"] = admin.get("name") or admin.get("email")
    await db.branch_documents.update_one({"doc_id": doc_id}, {"$set": patch})
    await _audit(admin, d["branch_id"], "license_update", changes)
    fresh = await db.branch_documents.find_one({"doc_id": doc_id},
                                               {"_id": 0, "file_base64": 0})
    return {"ok": True, "document": fresh}


@router.get("/{branch_id}/compliance-summary")
async def bm_compliance_summary(branch_id: str,
                                authorization: Optional[str] = Header(None)):
    """Compliance-wise applicability vs registration vs attachment (spec
    §11/§14/§22) + branch-specific alerts (§10/§15)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, _ADMIN_ROLES)
    b = await _get_branch(branch_id, admin)
    cc = b.get("compliance_config") or {}
    warn = await _warn_days(b["company_id"])
    today = _ist_today()
    soon = (datetime.now(timezone.utc) + timedelta(days=warn, hours=5, minutes=30)).strftime("%Y-%m-%d")
    docs = await db.branch_documents.find(
        {"branch_id": branch_id, "status": {"$nin": ["deleted", "replaced"]}},
        {"_id": 0, "file_base64": 0}).to_list(500)
    for d in docs:
        d["expiry_status"] = _doc_status(d, today, soon)
    rows, alerts = [], []
    for label, cat, app_key in _COMPLIANCE_MAP:
        cat_docs = [d for d in docs if (d.get("category") or "") == cat]
        applicable = bool(cc.get(app_key)) if app_key else (
            True if cat_docs else None)
        best = next((d for d in cat_docs if d["expiry_status"] == "active"), None) \
            or next((d for d in cat_docs if d["expiry_status"] == "expiring_soon"), None) \
            or (cat_docs[0] if cat_docs else None)
        if applicable is False:
            reg = "not_applicable"
        elif not best:
            reg = "missing" if applicable else "none"
        else:
            reg = best["expiry_status"]
        rows.append({
            "compliance": label, "category": cat,
            "applicable": applicable,
            "registration": reg,
            "doc_number": (best or {}).get("doc_number"),
            "effective_from": (best or {}).get("effective_from"),
            "expiry_date": (best or {}).get("expiry_date"),
            "attached": bool((best or {}).get("file_name")),
            "docs_count": len(cat_docs),
        })
        if applicable and not best:
            alerts.append(f"{label} registration missing")
        if applicable and best and not best.get("file_name"):
            alerts.append(f"{label} document not attached")
    for d in docs:
        if d["expiry_status"] == "expired":
            alerts.append(f"{d.get('doc_type')} EXPIRED on {d.get('expiry_date')}")
        elif d["expiry_status"] == "expiring_soon":
            try:
                days_left = (datetime.strptime(d["expiry_date"], "%Y-%m-%d")
                             - datetime.strptime(today, "%Y-%m-%d")).days
                alerts.append(f"{d.get('doc_type')} expires in {days_left} days")
            except Exception:
                alerts.append(f"{d.get('doc_type')} expiring soon")
    counts = {"active": sum(1 for d in docs if d["expiry_status"] == "active"),
              "expiring_soon": sum(1 for d in docs if d["expiry_status"] == "expiring_soon"),
              "expired": sum(1 for d in docs if d["expiry_status"] == "expired")}
    return {"compliances": rows, "alerts": alerts, "counts": counts,
            "warn_days": warn}


@router.get("/statutory/register")
async def statutory_register(company_id: Optional[str] = Query(None),
                             fmt: str = Query("json"),
                             authorization: Optional[str] = Header(None)):
    """Firm-wide branch-wise statutory document register (spec §18)."""
    admin, cid = await _gate(authorization, company_id)
    branches = await db.branches.find(
        {"$or": [{"company_id": cid}, {"linked_company_ids": cid}]},
        {"_id": 0, "branch_id": 1, "name": 1, "code": 1, "state": 1}).to_list(500)
    warn = await _warn_days(cid)
    today = _ist_today()
    soon = (datetime.now(timezone.utc) + timedelta(days=warn, hours=5, minutes=30)).strftime("%Y-%m-%d")
    rows = []
    for b in branches:
        docs = await db.branch_documents.find(
            {"branch_id": b["branch_id"], "status": {"$nin": ["deleted"]}},
            {"_id": 0, "file_base64": 0}).to_list(500)
        for d in docs:
            rows.append({
                "branch": b.get("name"), "branch_code": b.get("code"),
                "branch_state": b.get("state"),
                "category": d.get("category"), "doc_type": d.get("doc_type"),
                "doc_number": d.get("doc_number"),
                "issuing_authority": d.get("issuing_authority"),
                "state": d.get("state"),
                "effective_from": d.get("effective_from"),
                "expiry_date": "No Expiry" if d.get("no_expiry") else d.get("expiry_date"),
                "status": _doc_status(d, today, soon),
                "attached": "Yes" if d.get("file_name") else "No",
            })
    if fmt == "json":
        return {"rows": rows}
    import io
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Statutory Register"
    heads = ["Branch", "Code", "Branch State", "Compliance", "Document Type",
             "Registration No.", "Issuing Authority", "Doc State",
             "Effective From", "Expiry", "Status", "Attached"]
    ws.append(heads)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["branch"], r["branch_code"], r["branch_state"],
                   r["category"], r["doc_type"], r["doc_number"],
                   r["issuing_authority"], r["state"], r["effective_from"],
                   r["expiry_date"], r["status"], r["attached"]])
    out = io.BytesIO()
    wb.save(out)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 'attachment; filename="Statutory_Compliance_Register.xlsx"'})
