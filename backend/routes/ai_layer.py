"""AI Layer (Iter 346) — analysis engine on top of the existing payroll system.

Modular, read-only analysis of existing data. AI recommends & explains; any
change requires explicit user approval (Apply Fix) and is audit-logged.
Existing payroll calculations remain the source of truth.

Endpoints (all /api prefixed):
  GET  /admin/ai/analysis?company_id&month[&refresh=1]  — full analysis
       (compliance findings, audit severities, attendance anomalies, scores,
        trends, forecast, reconciliation, compliance calendar, smart insights,
        recommendations). Cached in `ai_analyses`; refresh=1 re-analyses.
  POST /admin/ai/apply-fix        {company_id, month, finding_id} — apply an
       approved auto-fix (only safe, normalising fixes). Logged.
  POST /admin/ai/feedback         {company_id, finding_key, verdict} — learning
       engine: "false_positive" suppresses that finding in future analyses.
  GET  /admin/ai/salary-diff?company_id&month — month-vs-month explanation.
  GET  /admin/ai/audit-report.xlsx|.pdf?company_id&month — audit export.
  POST /admin/ai/map-columns      {headers: []} — AI Excel column mapping.
  GET/POST /admin/ai/import-templates — saved mapping templates.

New collections only: ai_analyses, ai_action_log, ai_feedback,
ai_import_templates. No existing tables are modified.
"""
import io
import json
import logging
import os
import re
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from dotenv import load_dotenv
from fastapi import APIRouter, Body, Header, HTTPException, Query
from fastapi.responses import StreamingResponse

from server import db, get_user_from_token, require_role  # noqa: E402

load_dotenv()

router = APIRouter(prefix="/api", tags=["ai-layer"])
logger = logging.getLogger("ai-layer")

EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")

_PAN_RX = re.compile(r"^[A-Z]{5}[0-9]{4}[A-Z]$")
_IFSC_RX = re.compile(r"^[A-Z]{4}0[A-Z0-9]{6}$")


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _money(n: Any) -> str:
    try:
        return f"₹{round(float(n or 0)):,}"
    except Exception:
        return "₹0"


async def _admin(authorization):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    return admin


def _f(code: str, severity: str, emp: Optional[dict], issue: str, reason: str,
       impact: str, fix: str, confidence: int, fixable: bool = False,
       fix_route: Optional[str] = None) -> dict:
    name = (emp or {}).get("name")
    ecode = (emp or {}).get("employee_code")
    return {
        "finding_id": f"fnd_{uuid.uuid4().hex[:10]}",
        # stable key for learning-engine suppression (per employee + check)
        "key": f"{code}:{(emp or {}).get('user_id') or 'firm'}",
        "code": code, "severity": severity,
        "user_id": (emp or {}).get("user_id"),
        "employee": f"{name} (Code {ecode})" if name else None,
        "issue": issue, "reason": reason, "impact": impact,
        "fix": fix, "confidence": confidence,
        "fixable": fixable, "fix_route": fix_route,
    }


# ---------------------------------------------------------------------------
# Compliance Checker + Payroll Auditor — rule engine
# ---------------------------------------------------------------------------

async def _employee_checks(cid: str) -> List[dict]:
    out: List[dict] = []
    emps = await db.users.find(
        {"role": "employee", "company_id": cid, "active": {"$ne": False},
         "employment_status": {"$nin": ["resigned", "exited", "terminated",
                                        "inactive", "left"]}},
        {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1, "uan_no": 1,
         "esi_ip_no": 1, "aadhaar_no": 1, "aadhar_number": 1, "pan_no": 1,
         "pan_number": 1, "bank_account": 1, "bank_ifsc": 1, "pay_mode": 1,
         "doj": 1, "compliance_gross": 1, "father_name": 1,
         "exit_date": 1}).to_list(5000)
    emps = [e for e in emps if not (e.get("exit_date") or "").strip()]
    # Iter 421 (user rule) — validations follow the FIRM MASTER policy:
    # when EPF (or ESI) is DISABLED for the firm, don't demand UAN / ESIC
    # IP numbers from its employees.
    _fm = await db.firm_masters.find_one(
        {"company_id": cid}, {"_id": 0, "epf": 1, "esi": 1, "deductions": 1}) or {}
    _fm_ded = _fm.get("deductions") or {}
    _epf_ap = (_fm.get("epf") or {}).get("applicable")
    _esi_ap = (_fm.get("esi") or {}).get("applicable")
    _pf_on = bool(_epf_ap) if _epf_ap is not None else bool(_fm_ded.get("PF", True))
    _esi_on = bool(_esi_ap) if _esi_ap is not None else bool(_fm_ded.get("ESI", True))
    seen_bank: Dict[str, dict] = {}
    seen_uan: Dict[str, dict] = {}
    seen_namefather: Dict[str, dict] = {}
    today = datetime.now()
    for e in emps:
        gross = float(e.get("compliance_gross") or 0)
        prof = f"/employee-detail-slip?user_id={e['user_id']}"
        if _pf_on and not (e.get("uan_no") or "").strip() and gross > 0:
            out.append(_f("missing_uan", "high", e, "Missing UAN",
                          "Employee has a compliance gross but no UAN number.",
                          "PF ECR upload will reject this member.",
                          "Collect the UAN (or generate via EPFO) and update the Employee Master.",
                          100, fix_route=prof))
        if _esi_on and not (e.get("esi_ip_no") or "").strip() and 0 < gross <= 21000:
            out.append(_f("missing_esic_no", "high", e, "Missing ESIC IP Number",
                          f"Gross {_money(gross)} is within the ESIC wage limit (₹21,000) but no IP number is on record.",
                          "ESIC monthly contribution filing will fail for this employee.",
                          "Register the employee on the ESIC portal and update the IP number.",
                          95, fix_route=prof))
        if not (e.get("aadhaar_no") or e.get("aadhar_number") or "").strip():
            out.append(_f("missing_aadhaar", "medium", e, "Missing Aadhaar",
                          "No Aadhaar number in the Employee Master.",
                          "UAN KYC seeding and ESIC registration require Aadhaar.",
                          "Collect and enter the employee's Aadhaar number.",
                          100, fix_route=prof))
        pay_mode = (e.get("pay_mode") or "").lower()
        if pay_mode in ("", "bank"):
            if not (e.get("bank_account") or "").strip():
                out.append(_f("missing_bank", "high", e, "Missing Bank Account",
                              "Pay mode is Bank but no account number is stored.",
                              "Employee will be skipped in the Bank Transfer sheet.",
                              "Enter the bank account number (verify with a cancelled cheque).",
                              95, fix_route=prof))
            ifsc = (e.get("bank_ifsc") or "").strip()
            if e.get("bank_account") and not ifsc:
                out.append(_f("missing_ifsc", "high", e, "Missing IFSC Code",
                              "Bank account exists but the IFSC code is empty.",
                              "Bank transfer file will be rejected by the bank.",
                              "Add the branch IFSC code.", 95, fix_route=prof))
            elif ifsc and not _IFSC_RX.match(ifsc.upper().replace(" ", "")):
                fixable = bool(_IFSC_RX.match(ifsc.upper().replace(" ", "").replace("-", "")))
                out.append(_f("invalid_ifsc", "medium", e, "Invalid IFSC Format",
                              f"IFSC '{ifsc}' does not match the RBI format (AAAA0XXXXXX).",
                              "Bank transfer rows with this IFSC will bounce.",
                              "Normalise the IFSC (uppercase, remove spaces) or correct it.",
                              90, fixable=fixable, fix_route=prof))
        pan = (e.get("pan_no") or e.get("pan_number") or "").strip()
        if pan and not _PAN_RX.match(pan):
            fixable = bool(_PAN_RX.match(pan.upper().replace(" ", "")))
            out.append(_f("invalid_pan", "medium", e, "Invalid PAN Format",
                          f"PAN '{pan}' does not match AAAAA9999A.",
                          "TDS returns (24Q) will report an invalid PAN — higher TDS may apply.",
                          "Normalise (uppercase, no spaces) or correct the PAN.",
                          92, fixable=fixable, fix_route=prof))
        # Bonus / Gratuity eligibility (informational)
        if 0 < gross <= 21000:
            out.append(_f("bonus_eligible", "low", e, "Bonus Act Eligible",
                          f"Gross {_money(gross)} ≤ ₹21,000 (Payment of Bonus Act limit).",
                          "Statutory bonus (min 8.33%) provision applies for this employee.",
                          "Include in the annual Bonus Process run.", 90))
        doj = (e.get("doj") or "").strip()
        try:
            if doj and (today - datetime.strptime(doj[:10], "%Y-%m-%d")).days >= int(4.8 * 365):
                out.append(_f("gratuity_eligible", "low", e, "Gratuity Eligible",
                              f"Service since {doj[:10]} exceeds 4.8 years.",
                              "Gratuity liability accrues (15/26 × last drawn basic × years).",
                              "Maintain a gratuity provision for this employee.", 90))
        except ValueError:
            pass
        # Duplicate detection
        acct = (e.get("bank_account") or "").strip()
        if acct:
            if acct in seen_bank:
                out.append(_f("duplicate_employee", "critical", e,
                              "Duplicate Bank Account",
                              f"Same account no. as {seen_bank[acct].get('name')} "
                              f"(Code {seen_bank[acct].get('employee_code')}).",
                              "Salary may be paid twice to one account.",
                              "Verify both records — deactivate/merge the duplicate.",
                              85, fix_route="/admin"))
            seen_bank[acct] = e
        uan = (e.get("uan_no") or "").strip()
        if uan:
            if uan in seen_uan:
                out.append(_f("duplicate_uan", "critical", e, "Duplicate UAN",
                              f"Same UAN as {seen_uan[uan].get('name')} "
                              f"(Code {seen_uan[uan].get('employee_code')}).",
                              "PF ECR will double-report contributions on one UAN.",
                              "Correct the UAN of one of the two employees.",
                              90, fix_route="/admin"))
            seen_uan[uan] = e
        nf = f"{(e.get('name') or '').strip().lower()}|{(e.get('father_name') or '').strip().lower()}"
        if nf.strip("|"):
            if nf in seen_namefather:
                out.append(_f("duplicate_employee", "high", e,
                              "Possible Duplicate Employee",
                              f"Same name + father name as Code {seen_namefather[nf].get('employee_code')}.",
                              "Double payroll cost if both records get salary.",
                              "Verify and mark one record resigned/disabled.",
                              70, fix_route="/admin"))
            seen_namefather[nf] = e
    return out


def _run_checks(run: dict, month: str) -> List[dict]:
    """Payroll checks on the latest compliance run rows."""
    out: List[dict] = []
    rows = run.get("rows") or []
    stat = run.get("statutory_cfg") or {}
    pf_cap = float(stat.get("pf_wage_cap") or 15000)
    esi_lim = float(stat.get("esic_gross_threshold") or 21000)
    seen_uid: Dict[str, int] = {}
    for r in rows:
        emp = {"user_id": r.get("user_id"), "name": r.get("name"),
               "employee_code": r.get("employee_code")}
        gross = float(r.get("gross_paid") or r.get("monthly_gross") or 0)
        net = float(r.get("net") or 0)
        seen_uid[r.get("user_id")] = seen_uid.get(r.get("user_id"), 0) + 1
        # Negative salary
        if net < 0:
            out.append(_f("negative_salary", "critical", emp, "Negative Net Salary",
                          f"Net is {_money(net)} — deductions exceed earnings.",
                          "Employee cannot be paid a negative amount; register is invalid.",
                          "Review the deductions/advance for this row and reprocess.",
                          100, fix_route="/compliance-salary-run"))
        # Excess deductions
        ded = float(r.get("total_deduction") or 0)
        if gross > 0 and ded > 0.5 * gross:
            out.append(_f("excess_deduction", "high", emp, "Excess Deductions",
                          f"Deductions {_money(ded)} exceed 50% of gross {_money(gross)}.",
                          "Payment of Wages Act caps deductions at 50% of wages.",
                          "Split the recovery over multiple months.", 90,
                          fix_route="/compliance-salary-run"))
        # PF checks
        pf_wages = float(r.get("pf_wages") or 0)
        pf_emp = float(r.get("pf_employee") or 0)
        if r.get("pf_applicable"):
            if pf_wages > pf_cap + 1:
                out.append(_f("pf_wage_mismatch", "high", emp, "PF Wages Above Ceiling",
                              f"PF wages {_money(pf_wages)} exceed the ₹{int(pf_cap):,} statutory ceiling.",
                              "ECR will be inconsistent with the ceiling rule.",
                              "Cap PF wages at the ceiling (or confirm voluntary higher contribution).",
                              95, fix_route="/compliance-salary-run"))
            expected = round(min(pf_wages, pf_cap) * 0.12)
            if pf_wages > 0 and abs(pf_emp - expected) > 2:
                out.append(_f("pf_calc_mismatch", "high", emp, "PF Deduction Mismatch",
                              f"EE PF {_money(pf_emp)} ≠ 12% of PF wages ({_money(expected)}).",
                              "Short/excess PF deposit — interest & damages risk (7Q/14B).",
                              "Reprocess the row so PF = 12% of PF wages.", 92,
                              fix_route="/compliance-salary-run"))
            if pf_wages > 0 and pf_emp == 0:
                out.append(_f("missing_pf_deduction", "critical", emp,
                              "Missing PF Deduction",
                              f"PF applicable with wages {_money(pf_wages)} but EE PF is 0.",
                              "Non-deposit of statutory PF — heavy penalty exposure.",
                              "Reprocess this employee's row with PF enabled.", 95,
                              fix_route="/compliance-salary-run"))
        # ESIC checks
        esic_emp = float(r.get("esic_employee") or 0)
        esic_base = float(r.get("esic_wage_base") or 0)
        if r.get("esic_applicable"):
            expected = round(esic_base * 0.0075)
            if esic_base > 0 and abs(esic_emp - expected) > 2:
                out.append(_f("esic_calc_mismatch", "high", emp, "ESIC Deduction Mismatch",
                              f"EE ESIC {_money(esic_emp)} ≠ 0.75% of wage base ({_money(expected)}).",
                              "ESIC contribution filing will not reconcile.",
                              "Reprocess the row so ESIC = 0.75% of ESIC wages.", 92,
                              fix_route="/compliance-salary-run"))
        elif 0 < gross <= esi_lim and esic_emp == 0:
            out.append(_f("esic_wage_mismatch", "medium", emp, "ESIC Possibly Missed",
                          f"Gross {_money(gross)} is within the ESIC limit (₹{int(esi_lim):,}) "
                          "but the row is marked not applicable.",
                          "Coverage gap — inspection risk for ESIC-eligible wages.",
                          "Verify ESIC applicability for this employee.", 75,
                          fix_route="/compliance-salary-run"))
        # OT checks
        ot_pay = float(r.get("ot_pay") or 0)
        ot_hours = float(r.get("ot_hours") or 0)
        ot_rate = float(r.get("ot_hourly_rate") or 0)
        if ot_pay > 0 and ot_hours == 0:
            out.append(_f("ot_calc_error", "medium", emp, "OT Amount Without OT Hours",
                          f"OT pay {_money(ot_pay)} exists but OT hours are 0.",
                          "OT register and Form-B wages register will not match.",
                          "Enter the OT hours or move the amount to Other Allowances.",
                          85, fix_route="/compliance-salary-run"))
        elif ot_pay > 0 and ot_rate > 0 and abs(ot_pay - ot_hours * ot_rate) > max(10, 0.05 * ot_pay):
            out.append(_f("ot_calc_error", "medium", emp, "OT Calculation Mismatch",
                          f"OT pay {_money(ot_pay)} ≠ hours {ot_hours:g} × rate {_money(ot_rate)}.",
                          "OT wages register mismatch.",
                          "Recompute OT for this row.", 80,
                          fix_route="/compliance-salary-run"))
        if ot_hours > 60:
            out.append(_f("ot_anomaly", "medium", emp, "Excessive OT Hours",
                          f"{ot_hours:g} OT hours in {month} exceed the Factories Act "
                          "quarterly guidance pro-rata (~50h/quarter historically, state limits vary).",
                          "Inspection observation risk for excess overtime.",
                          "Verify punches; distribute workload or take state-specific exemption.",
                          70))
        # Zero days but paid
        pd_ = float(r.get("present_days") or 0)
        if pd_ == 0 and net > 0:
            out.append(_f("paid_without_days", "high", emp, "Salary Without Present Days",
                          f"Net {_money(net)} paid with 0 present days.",
                          "Wages register shows payment without attendance.",
                          "Verify attendance import or remove the row.", 90,
                          fix_route="/compliance-salary-run"))
    for uid, n in seen_uid.items():
        if n > 1 and uid:
            row = next(r for r in rows if r.get("user_id") == uid)
            out.append(_f("duplicate_salary", "critical",
                          {"user_id": uid, "name": row.get("name"),
                           "employee_code": row.get("employee_code")},
                          "Duplicate Salary Row",
                          f"Employee appears {n} times in the {month} run.",
                          "Salary would be paid multiple times.",
                          "Reprocess the run — duplicates are removed automatically.",
                          100, fix_route="/compliance-salary-run"))
    return out


async def _attendance_checks(cid: str, month: str) -> List[dict]:
    out: List[dict] = []
    recs = await db.attendance.find(
        {"company_id": cid, "date": {"$regex": f"^{month}"},
         "status": {"$nin": ["rejected"]}},
        {"_id": 0, "user_id": 1, "date": 1, "kind": 1}).to_list(100000)
    by_emp: Dict[str, Dict[str, Dict[str, int]]] = {}
    for r in recs:
        d = by_emp.setdefault(r["user_id"], {}).setdefault(r["date"], {"in": 0, "out": 0})
        d[r.get("kind") or "in"] = d.get(r.get("kind") or "in", 0) + 1
    if not by_emp:
        return out
    uids = list(by_emp.keys())
    names = {u["user_id"]: u for u in await db.users.find(
        {"user_id": {"$in": uids}},
        {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1}).to_list(5000)}
    for uid, days in by_emp.items():
        emp = names.get(uid) or {"user_id": uid}
        miss_out = [d for d, k in days.items() if k.get("in") and not k.get("out")]
        if len(miss_out) >= 3:
            out.append(_f("missing_punch", "medium", emp, "Missing OUT Punches",
                          f"{len(miss_out)} day(s) have an IN punch but no OUT "
                          f"(e.g. {', '.join(sorted(miss_out)[:3])}).",
                          "Работа hours & OT cannot be computed for those days.".replace("Работа", "Work"),
                          "Approve corrected OUT times via Punch Approvals.",
                          85, fix_route="/punch-approvals"))
        dups = [d for d, k in days.items() if k.get("in", 0) > 2]
        if dups:
            out.append(_f("duplicate_punch", "low", emp, "Duplicate Punches",
                          f"{len(dups)} day(s) have 3+ IN punches (machine bounce).",
                          "Inflated punch logs can distort first-IN/last-OUT.",
                          "No action needed if first-IN/last-OUT logic is used; else clean.",
                          75))
        # Continuous working streak
        ds = sorted(days.keys())
        streak = best = 1
        for i in range(1, len(ds)):
            prev = datetime.strptime(ds[i - 1], "%Y-%m-%d")
            cur = datetime.strptime(ds[i], "%Y-%m-%d")
            streak = streak + 1 if (cur - prev).days == 1 else 1
            best = max(best, streak)
        if best >= 10:
            out.append(_f("continuous_working", "medium", emp,
                          "Long Continuous Working Streak",
                          f"Worked {best} consecutive days in {month} without a weekly off.",
                          "Factories Act requires a weekly holiday — inspection risk.",
                          "Schedule a weekly off / compensatory holiday.", 88))
    return out


# ---------------------------------------------------------------------------
# Trends, forecast, reconciliation, calendar, insights
# ---------------------------------------------------------------------------

def _tget(t: Optional[dict], *keys: str) -> float:
    t = t or {}
    for k in keys:
        v = t.get(k)
        if isinstance(v, (int, float)) and v:
            return float(v)
    return 0.0


def _prev_months(month: str, n: int) -> List[str]:
    y, m = map(int, month.split("-"))
    out = []
    for _ in range(n):
        m -= 1
        if m == 0:
            y, m = y - 1, 12
        out.append(f"{y}-{m:02d}")
    return out


async def _trends_and_forecast(cid: str, month: str) -> dict:
    months = list(reversed(_prev_months(month, 5))) + [month]
    series = []
    for mon in months:
        comp = await db.compliance_salary_runs.find_one(
            {"company_id": cid, "month": mon}, {"_id": 0, "totals": 1, "employees_count": 1},
            sort=[("generated_at", -1)])
        act = await db.salary_runs.find_one(
            {"company_id": cid, "month": mon}, {"_id": 0, "totals": 1, "employees_count": 1},
            sort=[("generated_at", -1)])
        t = (comp or {}).get("totals") or {}
        ta = (act or {}).get("totals") or {}
        series.append({
            "month": mon,
            "net": round(_tget(t, "net") + _tget(ta, "net_pay", "net")),
            "gross": round(_tget(t, "gross_paid", "monthly_gross")
                           + _tget(ta, "total_gross", "gross")),
            "pf": round(_tget(t, "pf_employee")),
            "esic": round(_tget(t, "esic_employee")),
            "employees": (comp or {}).get("employees_count") or (act or {}).get("employees_count") or 0,
        })
    nz = [s for s in series if s["net"]]
    forecast = {}
    if len(nz) >= 2:
        deltas = [nz[i]["net"] - nz[i - 1]["net"] for i in range(1, len(nz))]
        trend = sum(deltas) / len(deltas)
        nxt = max(0, round(nz[-1]["net"] + trend))
        forecast = {
            "next_month_net": nxt,
            "yearly_net": round(nxt * 12),
            "next_month_pf": round(sum(s["pf"] for s in nz[-3:]) / min(3, len(nz))),
            "next_month_esic": round(sum(s["esic"] for s in nz[-3:]) / min(3, len(nz))),
            "bonus_provision": round(0.0833 * sum(s["gross"] for s in nz) / len(nz) * 12) if nz else 0,
            "basis": f"linear trend of last {len(nz)} processed months",
        }
    elif nz:
        forecast = {"next_month_net": nz[-1]["net"], "yearly_net": nz[-1]["net"] * 12,
                    "next_month_pf": nz[-1]["pf"], "next_month_esic": nz[-1]["esic"],
                    "bonus_provision": round(nz[-1]["gross"] * 0.0833 * 12),
                    "basis": "single processed month (flat projection)"}
    return {"series": series, "forecast": forecast}


async def _reconciliation(cid: str, month: str, comp: Optional[dict],
                          act: Optional[dict]) -> dict:
    items = []
    crows = (comp or {}).get("rows") or []
    arows = (act or {}).get("rows") or []
    if comp and act:
        cu = {r.get("user_id") for r in crows}
        au = {r.get("user_id") for r in arows}
        only_c = cu - au
        only_a = au - cu
        if only_c:
            items.append({"kind": "missing_in_actual", "count": len(only_c),
                          "detail": f"{len(only_c)} employee(s) exist in the Compliance run "
                                    "but not in the Actual run."})
        if only_a:
            items.append({"kind": "missing_in_compliance", "count": len(only_a),
                          "detail": f"{len(only_a)} employee(s) exist in the Actual run "
                                    "but not in the Compliance run."})
    # Freeze (imported sheet) vs computed gross differences
    frz = [r for r in crows if r.get("imported_gross") is not None
           and abs((r.get("imported_gross") or 0) - (r.get("gross_paid") or 0)) >= 1]
    if frz:
        tot = round(sum((r.get("imported_gross") or 0) - (r.get("gross_paid") or 0) for r in frz))
        items.append({"kind": "freeze_gross_diff", "count": len(frz),
                      "detail": f"{len(frz)} row(s) differ from the imported Freeze gross "
                                f"(total difference {_money(tot)})."})
    # Bank sheet coverage
    if crows or arows:
        rows = crows or arows
        uids = [r.get("user_id") for r in rows if (r.get("net") or r.get("net_pay") or 0) > 0]
        nobank = await db.users.count_documents(
            {"user_id": {"$in": uids}, "$or": [
                {"bank_account": {"$in": [None, ""]}}, {"bank_ifsc": {"$in": [None, ""]}}],
             "pay_mode": {"$nin": ["cash", "Cash", "CASH", "cheque"]}})
        if nobank:
            items.append({"kind": "bank_transfer_gap", "count": nobank,
                          "detail": f"{nobank} paid employee(s) will be missing from the "
                                    "bank transfer file (no account/IFSC)."})
    return {"items": items, "ok": not items}


def _compliance_calendar(today: Optional[datetime] = None) -> List[dict]:
    today = today or datetime.now()
    y, m = today.year, today.month
    nxt_y, nxt_m = (y + 1, 1) if m == 12 else (y, m + 1)
    def d(yy, mm, dd):  # noqa: E306
        return f"{yy}-{mm:02d}-{dd:02d}"
    items = [
        {"due": d(y, m, 7), "what": "TDS deposit (previous month)", "kind": "tds"},
        {"due": d(y, m, 15), "what": "PF payment + ECR filing (previous month)", "kind": "pf"},
        {"due": d(y, m, 15), "what": "ESIC contribution payment (previous month)", "kind": "esic"},
        {"due": d(y, m, 21), "what": "Professional Tax deposit (state-wise, typical)", "kind": "pt"},
        {"due": d(nxt_y, nxt_m, 7), "what": "TDS deposit (this month)", "kind": "tds"},
        {"due": d(nxt_y, nxt_m, 15), "what": "PF payment + ECR filing (this month)", "kind": "pf"},
        {"due": d(nxt_y, nxt_m, 15), "what": "ESIC contribution (this month)", "kind": "esic"},
    ]
    if m == 12 or m == 1:
        items.append({"due": d(y if m == 12 else y - 1, 12, 30),
                      "what": "LWF contribution (Dec — most states, half-yearly/annual)", "kind": "lwf"})
    if m in (10, 11):
        items.append({"due": d(y, 11, 30),
                      "what": "Bonus payment (within 8 months of FY close)", "kind": "bonus"})
    if m in (1, 2):
        items.append({"due": d(y, 2, 1), "what": "Annual Factory Return (Form 21/state form)",
                      "kind": "factory_return"})
    ts = today.strftime("%Y-%m-%d")
    for it in items:
        days = (datetime.strptime(it["due"], "%Y-%m-%d") - today).days
        it["days_left"] = days
        it["status"] = "overdue" if it["due"] < ts else ("soon" if days <= 5 else "upcoming")
    return sorted([i for i in items if i["days_left"] >= -10], key=lambda i: i["due"])


def _smart_insights(comp: Optional[dict], prev: Optional[dict]) -> List[dict]:
    out = []
    rows = (comp or {}).get("rows") or []
    if rows:
        ot = sorted(rows, key=lambda r: r.get("ot_hours") or 0, reverse=True)[:5]
        ot = [r for r in ot if (r.get("ot_hours") or 0) > 0]
        if ot:
            out.append({"title": "Top OT Employees",
                        "lines": [f"{r.get('name')} (Code {r.get('employee_code')}) — "
                                  f"{r.get('ot_hours'):g} h · {_money(r.get('ot_pay'))}" for r in ot]})
        by_dept: Dict[str, float] = {}
        for r in rows:
            by_dept[(r.get("department") or r.get("employee_type") or "—")] = \
                by_dept.get(r.get("department") or r.get("employee_type") or "—", 0) \
                + (r.get("ot_hours") or 0)
        top_dept = sorted(by_dept.items(), key=lambda x: -x[1])[:3]
        if any(v for _, v in top_dept):
            out.append({"title": "Departments with Maximum OT",
                        "lines": [f"{k}: {v:g} OT hours" for k, v in top_dept if v]})
    if rows and prev and prev.get("rows"):
        pmap = {r.get("user_id"): r for r in prev["rows"]}
        ups = []
        for r in rows:
            p = pmap.get(r.get("user_id"))
            if p:
                diff = (r.get("net") or 0) - (p.get("net") or 0)
                if diff > 0:
                    ups.append((diff, r))
        ups.sort(key=lambda x: -x[0])
        if ups[:5]:
            out.append({"title": "Highest Salary Increase (vs last month)",
                        "lines": [f"{r.get('name')} (Code {r.get('employee_code')}) — +{_money(d)}"
                                  for d, r in ups[:5]]})
    return out


# ---------------------------------------------------------------------------
# Master analysis
# ---------------------------------------------------------------------------

async def _analyze(admin: dict, cid: str, month: str) -> dict:
    comp = await db.compliance_salary_runs.find_one(
        {"company_id": cid, "month": month}, {"_id": 0}, sort=[("generated_at", -1)])
    act = await db.salary_runs.find_one(
        {"company_id": cid, "month": month}, {"_id": 0}, sort=[("generated_at", -1)])
    prev_mon = _prev_months(month, 1)[0]
    prev = await db.compliance_salary_runs.find_one(
        {"company_id": cid, "month": prev_mon}, {"_id": 0, "rows": 1},
        sort=[("generated_at", -1)])

    findings: List[dict] = []
    findings += await _employee_checks(cid)
    if comp:
        findings += _run_checks(comp, month)
    findings += await _attendance_checks(cid, month)

    # Learning engine — suppress user-confirmed false positives.
    fps = {f["finding_key"] async for f in db.ai_feedback.find(
        {"company_id": cid, "verdict": "false_positive"}, {"_id": 0, "finding_key": 1})}
    suppressed = [f for f in findings if f["key"] in fps]
    findings = [f for f in findings if f["key"] not in fps]

    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    findings.sort(key=lambda f: (sev_order.get(f["severity"], 9), -f["confidence"]))
    counts = {s: sum(1 for f in findings if f["severity"] == s) for s in sev_order}

    # Scores
    penalty = counts["critical"] * 15 + counts["high"] * 6 + counts["medium"] * 2 + counts["low"] * 0.5
    compliance_score = max(0, round(100 - penalty))
    att_issues = sum(1 for f in findings if f["code"] in
                     ("missing_punch", "duplicate_punch", "continuous_working"))
    pay_errors = sum(1 for f in findings if f["code"] in
                     ("negative_salary", "duplicate_salary", "pf_calc_mismatch",
                      "esic_calc_mismatch", "ot_calc_error", "excess_deduction",
                      "paid_without_days", "missing_pf_deduction"))
    health = max(0, round(100 - counts["critical"] * 20 - pay_errors * 5
                          - (0 if comp else 15)))
    risk = ("red" if counts["critical"] or compliance_score < 60
            else "yellow" if counts["high"] or compliance_score < 85 else "green")

    tf = await _trends_and_forecast(cid, month)
    recon = await _reconciliation(cid, month, comp, act)
    calendar = _compliance_calendar()
    insights = _smart_insights(comp, prev)

    pending = [c for c in calendar if c["status"] in ("overdue", "soon")]
    recommendations = []
    for f in findings[:6]:
        recommendations.append({"text": f"{f['issue']}"
                                        f"{' — ' + f['employee'] if f.get('employee') else ''}: {f['fix']}",
                                "confidence": f["confidence"], "severity": f["severity"]})
    if not comp:
        recommendations.insert(0, {"text": f"No Compliance Salary run exists for {month} — "
                                           "process it to enable payroll checks.",
                                   "confidence": 100, "severity": "high"})

    analysis = {
        "analysis_id": f"ana_{uuid.uuid4().hex[:10]}",
        "company_id": cid, "month": month, "generated_at": _now_iso(),
        "generated_by": admin["user_id"],
        "scores": {"payroll_health": health, "compliance_score": compliance_score,
                   "risk_level": risk, "ai_alerts": counts["critical"] + counts["high"],
                   "attendance_issues": att_issues, "payroll_errors": pay_errors,
                   "pending_compliance": len(pending)},
        "severity_counts": counts,
        "findings": findings,
        "suppressed_count": len(suppressed),
        "recommendations": recommendations,
        "trends": tf["series"], "forecast": tf["forecast"],
        "reconciliation": recon,
        "calendar": calendar,
        "insights": insights,
        "run_status": {
            "compliance": {"exists": bool(comp), "finalized": bool((comp or {}).get("finalized")),
                           "employees": (comp or {}).get("employees_count")},
            "actual": {"exists": bool(act), "finalized": bool((act or {}).get("finalized")),
                       "employees": (act or {}).get("employees_count")},
        },
    }
    await db.ai_analyses.replace_one(
        {"company_id": cid, "month": month}, analysis, upsert=True)
    # Notification engine — surface red alerts in the existing notification bell.
    try:
        if risk == "red":
            exists = await db.notifications.find_one(
                {"company_id": cid, "title": f"AI Alert — {month}"})
            if not exists:
                await db.notifications.insert_one({
                    "notification_id": f"ntf_{uuid.uuid4().hex[:10]}",
                    "company_id": cid, "audience": "admins",
                    "title": f"AI Alert — {month}",
                    "body": (f"AI analysis found {counts['critical']} critical and "
                             f"{counts['high']} high issues. Open AI Payroll Assistant."),
                    "created_at": _now_iso(), "created_by": "ai-layer", "read_by": []})
    except Exception:
        pass
    return analysis


@router.get("/admin/ai/analysis")
async def ai_analysis(company_id: str = Query(...), month: str = Query(...),
                      refresh: int = Query(0),
                      authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    if admin["role"] == "company_admin" and admin.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Not authorised for this firm")
    if not refresh:
        cached = await db.ai_analyses.find_one(
            {"company_id": company_id, "month": month}, {"_id": 0})
        if cached:
            cached["cached"] = True
            return cached
    return await _analyze(admin, company_id, month)


@router.post("/admin/ai/apply-fix")
async def ai_apply_fix(payload: Dict[str, Any] = Body(...),
                       authorization: Optional[str] = Header(None)):
    """Apply an APPROVED safe fix. Only normalising fixes are automated
    (PAN/IFSC uppercase + strip); everything else returns the screen to fix
    manually. Every action is audit-logged."""
    admin = await _admin(authorization)
    cid, month = payload.get("company_id"), payload.get("month")
    fid = payload.get("finding_id")
    ana = await db.ai_analyses.find_one({"company_id": cid, "month": month}, {"_id": 0})
    if not ana:
        raise HTTPException(status_code=404, detail="Run the AI analysis first")
    f = next((x for x in ana.get("findings", []) if x["finding_id"] == fid), None)
    if not f:
        raise HTTPException(status_code=404, detail="Finding not found — refresh the analysis")
    applied = False
    detail = ""
    if f["code"] == "invalid_pan" and f.get("user_id"):
        u = await db.users.find_one({"user_id": f["user_id"]},
                                    {"_id": 0, "pan_no": 1, "pan_number": 1})
        pan = ((u or {}).get("pan_no") or (u or {}).get("pan_number") or "")
        norm = pan.upper().replace(" ", "")
        if _PAN_RX.match(norm):
            await db.users.update_one({"user_id": f["user_id"]}, {"$set": {"pan_no": norm}})
            applied, detail = True, f"PAN normalised to {norm}"
    elif f["code"] == "invalid_ifsc" and f.get("user_id"):
        u = await db.users.find_one({"user_id": f["user_id"]}, {"_id": 0, "bank_ifsc": 1})
        norm = ((u or {}).get("bank_ifsc") or "").upper().replace(" ", "").replace("-", "")
        if _IFSC_RX.match(norm):
            await db.users.update_one({"user_id": f["user_id"]}, {"$set": {"bank_ifsc": norm}})
            applied, detail = True, f"IFSC normalised to {norm}"
    await db.ai_action_log.insert_one({
        "at": _now_iso(), "by": admin["user_id"], "company_id": cid, "month": month,
        "action": "apply_fix", "finding": {k: f.get(k) for k in
                                           ("code", "issue", "employee", "confidence")},
        "applied": applied, "detail": detail})
    if applied:
        await _analyze(admin, cid, month)  # re-analyse so the finding clears
        return {"ok": True, "applied": True, "detail": detail}
    return {"ok": True, "applied": False,
            "detail": "This fix needs a manual change — opening the right screen.",
            "fix_route": f.get("fix_route") or "/admin"}


@router.post("/admin/ai/feedback")
async def ai_feedback(payload: Dict[str, Any] = Body(...),
                      authorization: Optional[str] = Header(None)):
    """Learning engine: mark a finding correct / false_positive. False
    positives are suppressed from future analyses for that employee+check."""
    admin = await _admin(authorization)
    verdict = payload.get("verdict")
    if verdict not in ("correct", "false_positive"):
        raise HTTPException(status_code=400, detail="verdict must be correct|false_positive")
    await db.ai_feedback.replace_one(
        {"company_id": payload.get("company_id"), "finding_key": payload.get("finding_key")},
        {"company_id": payload.get("company_id"), "finding_key": payload.get("finding_key"),
         "verdict": verdict, "by": admin["user_id"], "at": _now_iso()}, upsert=True)
    await db.ai_action_log.insert_one({
        "at": _now_iso(), "by": admin["user_id"], "company_id": payload.get("company_id"),
        "action": "feedback", "finding_key": payload.get("finding_key"), "verdict": verdict})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Salary Difference Analysis
# ---------------------------------------------------------------------------

@router.get("/admin/ai/salary-diff")
async def ai_salary_diff(company_id: str = Query(...), month: str = Query(...),
                         authorization: Optional[str] = Header(None)):
    await _admin(authorization)
    prev_mon = _prev_months(month, 1)[0]
    cur = await db.compliance_salary_runs.find_one(
        {"company_id": company_id, "month": month}, {"_id": 0, "rows": 1, "totals": 1},
        sort=[("generated_at", -1)])
    prv = await db.compliance_salary_runs.find_one(
        {"company_id": company_id, "month": prev_mon}, {"_id": 0, "rows": 1, "totals": 1},
        sort=[("generated_at", -1)])
    if not cur or not prv:
        return {"month": month, "prev_month": prev_mon, "rows": [],
                "summary": f"Need processed compliance runs for BOTH {prev_mon} and {month} to compare."}
    pmap = {r.get("user_id"): r for r in prv.get("rows") or []}
    out = []
    for r in cur.get("rows") or []:
        p = pmap.get(r.get("user_id"))
        if not p:
            out.append({"name": r.get("name"), "employee_code": r.get("employee_code"),
                        "net_diff": round(r.get("net") or 0), "reasons": ["New joiner this month"]})
            continue
        nd = round((r.get("net") or 0) - (p.get("net") or 0))
        if abs(nd) < 1:
            continue
        reasons = []
        dd = (r.get("present_days") or 0) - (p.get("present_days") or 0)
        if abs(dd) >= 0.5:
            reasons.append(f"Present days {'+' if dd > 0 else ''}{dd:g} "
                           f"({p.get('present_days'):g} → {r.get('present_days'):g})")
        od = round((r.get("ot_pay") or 0) - (p.get("ot_pay") or 0))
        if abs(od) >= 1:
            reasons.append(f"OT {'+' if od > 0 else ''}{_money(od)}")
        ad = round((r.get("others") or 0) - (p.get("others") or 0))
        if abs(ad) >= 1:
            reasons.append(f"Other allowances {'+' if ad > 0 else ''}{_money(ad)}")
        gd = round((r.get("monthly_gross") or 0) - (p.get("monthly_gross") or 0))
        if abs(gd) >= 1 and abs(dd) < 0.5:
            reasons.append(f"Gross rate change {'+' if gd > 0 else ''}{_money(gd)} (increment/arrear?)")
        xd = round((r.get("total_deduction") or 0) - (p.get("total_deduction") or 0))
        if abs(xd) >= 1:
            reasons.append(f"Deductions {'+' if xd > 0 else ''}{_money(xd)}")
        ld = (r.get("esic_leave_days") or 0) - (p.get("esic_leave_days") or 0)
        if abs(ld) >= 0.5:
            reasons.append(f"ESIC leave days {'+' if ld > 0 else ''}{ld:g}")
        out.append({"name": r.get("name"), "employee_code": r.get("employee_code"),
                    "net_prev": round(p.get("net") or 0), "net_cur": round(r.get("net") or 0),
                    "net_diff": nd, "reasons": reasons or ["Mixed small changes"]})
    left = [{"name": p.get("name"), "employee_code": p.get("employee_code"),
             "net_diff": -round(p.get("net") or 0), "reasons": ["Not in this month's run (exit?)"]}
            for uid, p in pmap.items()
            if uid not in {r.get("user_id") for r in cur.get("rows") or []}
            and round(p.get("net") or 0) != 0]
    out = [r for r in out if r["net_diff"] != 0]
    out += left
    out.sort(key=lambda x: -abs(x["net_diff"]))
    tot_d = round(_tget(cur.get("totals"), "net") - _tget(prv.get("totals"), "net"))
    ups = sum(1 for r in out if r["net_diff"] > 0)
    downs = sum(1 for r in out if r["net_diff"] < 0)
    summary = (f"{month} vs {prev_mon}: total Net changed by {_money(tot_d)} "
               f"({ups} employees up, {downs} down, {len(out)} changed rows). "
               "Top reasons are attendance-day changes, OT and allowance edits.")
    return {"month": month, "prev_month": prev_mon, "total_net_diff": tot_d,
            "rows": out[:100], "summary": summary}


# ---------------------------------------------------------------------------
# Audit report exports
# ---------------------------------------------------------------------------

async def _get_analysis_or_404(cid: str, month: str) -> dict:
    ana = await db.ai_analyses.find_one({"company_id": cid, "month": month}, {"_id": 0})
    if not ana:
        raise HTTPException(status_code=404, detail="Run the AI analysis first")
    return ana


@router.get("/admin/ai/audit-report.xlsx")
async def ai_audit_xlsx(company_id: str = Query(...), month: str = Query(...),
                        authorization: Optional[str] = Header(None)):
    await _admin(authorization)
    ana = await _get_analysis_or_404(company_id, month)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "AI Audit"
    ws.append(["Severity", "Issue", "Employee", "Reason", "Impact",
               "Recommended Fix", "Confidence %"])
    fill = PatternFill(start_color="0F2E3D", end_color="0F2E3D", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
    colors = {"critical": "FEE2E2", "high": "FFEDD5", "medium": "FEF9C3", "low": "DCFCE7"}
    for f in ana.get("findings", []):
        ws.append([f["severity"].upper(), f["issue"], f.get("employee") or "—",
                   f["reason"], f["impact"], f["fix"], f["confidence"]])
        ws.cell(row=ws.max_row, column=1).fill = PatternFill(
            start_color=colors.get(f["severity"], "FFFFFF"),
            end_color=colors.get(f["severity"], "FFFFFF"), fill_type="solid")
    for i, w in enumerate([12, 34, 30, 48, 44, 48, 12], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="ai-audit-{company_id}-{month}.xlsx"'})


@router.get("/admin/ai/audit-report.pdf")
async def ai_audit_pdf(company_id: str = Query(...), month: str = Query(...),
                       authorization: Optional[str] = Header(None)):
    await _admin(authorization)
    ana = await _get_analysis_or_404(company_id, month)
    co = await db.companies.find_one({"company_id": company_id}, {"_id": 0, "name": 1})
    from reportlab.lib import colors as rc
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)
    st = getSampleStyleSheet()
    cell = st["BodyText"]; cell.fontSize = 7.2; cell.leading = 9
    els = [Paragraph(f"<b>AI Payroll Audit — {(co or {}).get('name')} — {month}</b>", st["Title"]),
           Paragraph(f"Health {ana['scores']['payroll_health']} · Compliance "
                     f"{ana['scores']['compliance_score']} · Risk {ana['scores']['risk_level'].upper()}"
                     f" · Generated {ana['generated_at'][:16]}", st["Normal"]),
           Spacer(1, 4 * mm)]
    data = [["Severity", "Issue", "Employee", "Reason", "Recommended Fix", "Conf %"]]
    for f in ana.get("findings", [])[:400]:
        data.append([f["severity"].upper(), Paragraph(f["issue"], cell),
                     Paragraph(f.get("employee") or "—", cell),
                     Paragraph(f["reason"], cell), Paragraph(f["fix"], cell),
                     str(f["confidence"])])
    tbl = Table(data, colWidths=[20 * mm, 45 * mm, 45 * mm, 80 * mm, 75 * mm, 12 * mm],
                repeatRows=1)
    sev_col = {"CRITICAL": rc.HexColor("#FEE2E2"), "HIGH": rc.HexColor("#FFEDD5"),
               "MEDIUM": rc.HexColor("#FEF9C3"), "LOW": rc.HexColor("#DCFCE7")}
    style = [("BACKGROUND", (0, 0), (-1, 0), rc.HexColor("#0F2E3D")),
             ("TEXTCOLOR", (0, 0), (-1, 0), rc.white),
             ("FONTSIZE", (0, 0), (-1, -1), 7.2),
             ("GRID", (0, 0), (-1, -1), 0.4, rc.HexColor("#CBD5E1")),
             ("VALIGN", (0, 0), (-1, -1), "TOP")]
    for i, row in enumerate(data[1:], start=1):
        style.append(("BACKGROUND", (0, i), (0, i), sev_col.get(row[0], rc.white)))
    tbl.setStyle(TableStyle(style))
    els.append(tbl)
    doc.build(els)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition":
                                      f'attachment; filename="ai-audit-{company_id}-{month}.pdf"'})


# ---------------------------------------------------------------------------
# AI Excel Import Assistant — column mapping + templates
# ---------------------------------------------------------------------------

_CANON_FIELDS = [
    "employee_code", "name", "father_name", "designation", "department",
    "employee_type", "doj", "exit_date", "phone", "email", "uan_no",
    "esi_ip_no", "pan_no", "aadhaar_no", "bank_name", "bank_account",
    "bank_ifsc", "gross_salary", "basic_salary", "present_days", "ot_hours",
    "ot_amount", "in_time", "out_time", "date", "net_salary", "advance",
    "other_allowance", "other_deduction",
]

_MAP_HINTS = {
    "employee_code": ["emp code", "employee id", "emp id", "code", "empno", "emp no", "token"],
    "name": ["worker name", "employee name", "emp name", "name of employee"],
    "father_name": ["father", "guardian"],
    "gross_salary": ["pay", "gross", "salary", "wages", "total salary"],
    "in_time": ["punch in", "in time", "intime", "time in"],
    "out_time": ["punch out", "out time", "outtime", "time out"],
    "doj": ["date joined", "joining", "doj", "date of joining"],
    "present_days": ["days", "present", "p days", "attendance"],
    "ot_hours": ["ot hrs", "overtime hours", "ot hours"],
    "ot_amount": ["ot amt", "overtime amount", "ot amount"],
    "uan_no": ["uan"],
    "esi_ip_no": ["esic", "esi no", "ip no"],
    "aadhaar_no": ["aadhaar", "aadhar"],
    "pan_no": ["pan"],
    "bank_account": ["account", "a/c", "acc no"],
    "bank_ifsc": ["ifsc"],
    "phone": ["mobile", "contact"],
    "net_salary": ["net", "take home", "net pay"],
    "date": ["date", "att date"],
}


@router.post("/admin/ai/map-columns")
async def ai_map_columns(payload: Dict[str, Any] = Body(...),
                         authorization: Optional[str] = Header(None)):
    """Map raw Excel headers to canonical payroll fields. Deterministic
    keyword pass first (fast, 100% repeatable), LLM only for leftovers."""
    await _admin(authorization)
    headers: List[str] = [str(h) for h in (payload.get("headers") or []) if str(h).strip()]
    if not headers:
        raise HTTPException(status_code=400, detail="headers list is required")
    mapping: Dict[str, dict] = {}
    unresolved = []
    for h in headers:
        hl = h.lower().strip()
        hit = None
        for field, hints in _MAP_HINTS.items():
            if hl == field or hl.replace(" ", "_") == field or any(k in hl for k in hints):
                hit = field
                break
        if hit:
            mapping[h] = {"field": hit, "confidence": 92, "source": "rules"}
        else:
            unresolved.append(h)
    if unresolved and EMERGENT_LLM_KEY:
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            chat = LlmChat(api_key=EMERGENT_LLM_KEY,
                           session_id=f"ai-map-{uuid.uuid4().hex[:8]}",
                           system_message=(
                               "Map Excel column headers of an Indian payroll sheet to these "
                               "canonical fields: " + ", ".join(_CANON_FIELDS) +
                               '. Reply STRICT JSON {"<header>": "<field or null>"} only.'
                           )).with_model("openai", "gpt-5.4")
            resp = await chat.send_message(UserMessage(text=json.dumps(unresolved)))
            m = re.search(r"\{.*\}", str(resp), re.DOTALL)
            if m:
                for h, fld in json.loads(m.group(0)).items():
                    if fld in _CANON_FIELDS:
                        mapping[h] = {"field": fld, "confidence": 78, "source": "llm"}
        except Exception as e:
            logger.warning("[ai-map] LLM mapping failed: %s", e)
    unmapped = [h for h in headers if h not in mapping]
    return {"mapping": mapping, "unmapped": unmapped, "fields": _CANON_FIELDS}


@router.get("/admin/ai/import-templates")
async def ai_templates_list(authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    rows = await db.ai_import_templates.find(
        {}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"templates": rows}


@router.post("/admin/ai/import-templates")
async def ai_templates_save(payload: Dict[str, Any] = Body(...),
                            authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    name = (payload.get("name") or "").strip()
    mapping = payload.get("mapping") or {}
    if not name or not mapping:
        raise HTTPException(status_code=400, detail="name and mapping are required")
    doc = {"template_id": f"tpl_{uuid.uuid4().hex[:10]}", "name": name,
           "mapping": mapping, "created_by": admin["user_id"], "created_at": _now_iso()}
    await db.ai_import_templates.replace_one({"name": name}, doc, upsert=True)
    return {"ok": True, "template_id": doc["template_id"]}


@router.get("/admin/ai/action-log")
async def ai_action_log(company_id: Optional[str] = Query(None),
                        authorization: Optional[str] = Header(None)):
    await _admin(authorization)
    q = {"company_id": company_id} if company_id else {}
    rows = await db.ai_action_log.find(q, {"_id": 0}).sort("at", -1).to_list(100)
    return {"entries": rows}
