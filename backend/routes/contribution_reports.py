"""PF / ESI Contribution Sheets + Bonus Yearly Summary (Reports).

  * P.F. Contribution Sheet  — month-wise per-employee + employee-wise yearly
  * E.S.I. Contribution Sheet — month-wise per-employee + employee-wise yearly
  * Bonus Yearly Summary      — employee-wise, month-wise earned wages across
    the FY with the earning-allowance heads enabled in Firm Master.

Data source: the LATEST compliance salary run saved for each month
(``db.compliance_salary_runs``).

  GET /api/admin/reports/contribution?kind=pf|esi&company_id=&month=YYYY-MM
  GET /api/admin/reports/contribution.xlsx?...same...
  GET /api/admin/reports/contribution-yearly?kind=pf|esi&company_id=&fy_start_year=
  GET /api/admin/reports/contribution-yearly.xlsx?...same...
  GET /api/admin/reports/bonus-yearly-summary?company_id=&fy_start_year=
  GET /api/admin/reports/bonus-yearly-summary.xlsx?...same...
"""
from datetime import date
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException
from fastapi.responses import StreamingResponse

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
)

router = APIRouter(prefix="/api/admin/reports", tags=["contribution-reports"])

_ALL_HEADS = ["basic", "hra", "conveyance", "medical", "special", "others"]
_HEAD_LABELS = {
    "basic": "Basic", "hra": "HRA", "conveyance": "Conveyance",
    "medical": "Medical", "special": "Special", "others": "Others",
}


def _fy_months(fy_start_year: int) -> List[Dict[str, str]]:
    """April..March of the FY as [{key: 'YYYY-MM', label: 'Apr-25'}]."""
    out = []
    for i in range(12):
        m = 4 + i
        y = fy_start_year
        if m > 12:
            m -= 12
            y += 1
        out.append({
            "key": f"{y:04d}-{m:02d}",
            "label": f"{date(y, m, 1).strftime('%b')}-{str(y)[-2:]}",
        })
    return out


def _resolve_company(admin: Dict[str, Any], company_id: Optional[str]) -> str:
    if admin["role"] == "company_admin":
        return admin.get("company_id") or ""
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required (pick one firm)")
    return company_id


async def _latest_run(company_id: str, month: str) -> Optional[Dict[str, Any]]:
    return await db.compliance_salary_runs.find_one(
        {"company_id": company_id, "month": month},
        {"_id": 0},
        sort=[("generated_at", -1)],
    )


async def _emp_lookup(user_ids: List[str]) -> Dict[str, Dict[str, Any]]:
    from utils.relation import father_or_spouse_display
    out: Dict[str, Dict[str, Any]] = {}
    if not user_ids:
        return out
    async for u in db.users.find(
        {"user_id": {"$in": user_ids}},
        {"_id": 0, "user_id": 1, "employee_code": 1, "name": 1,
         "father_name": 1, "doj": 1, "uan_no": 1, "esi_ip_no": 1,
         "designation": 1, "gender": 1, "marital_status": 1, "spouse_name": 1},
    ):
        # Relation-aware display (D/O for unmarried females, spouse for married).
        u["father_name"] = father_or_spouse_display(u)
        out[u["user_id"]] = u
    return out


def _kind_ok(kind: str) -> str:
    k = (kind or "pf").lower()
    if k not in ("pf", "esi"):
        raise HTTPException(status_code=400, detail="kind must be pf | esi")
    return k


# ---------------------------------------------------------------------------
# Month-wise contribution sheet
# ---------------------------------------------------------------------------

async def _monthly_contribution(company_id: str, month: str, kind: str) -> Dict[str, Any]:
    run = await _latest_run(company_id, month)
    rows_out: List[Dict[str, Any]] = []
    if kind == "pf":
        columns = [
            {"key": "sr", "label": "Sr."},
            {"key": "name", "label": "Employee Name"},
            {"key": "uan_no", "label": "UAN No."},
            {"key": "pf_wages", "label": "PF Wages"},
            {"key": "pf_employee", "label": "Employee PF (12%)"},
            {"key": "pf_employer_eps", "label": "Employer EPS (8.33%)"},
            {"key": "pf_employer_epf", "label": "Employer EPF (3.67%)"},
            {"key": "pf_employer_total", "label": "Employer Total"},
            {"key": "total", "label": "Total Contribution"},
        ]
    else:
        columns = [
            {"key": "sr", "label": "Sr."},
            {"key": "name", "label": "Employee Name"},
            {"key": "esi_ip_no", "label": "ESIC IP No."},
            {"key": "esic_wage_base", "label": "ESI Wages"},
            {"key": "esic_employee", "label": "Employee ESI (0.75%)"},
            {"key": "esic_employer", "label": "Employer ESI (3.25%)"},
            {"key": "total", "label": "Total Contribution"},
        ]

    if run:
        raw = [
            r for r in (run.get("rows") or [])
            if (kind == "pf" and (float(r.get("pf_wages") or 0) > 0 or float(r.get("pf_employee") or 0) > 0))
            or (kind == "esi" and (float(r.get("esic_wage_base") or 0) > 0 or float(r.get("esic_employee") or 0) > 0))
        ]
        emp = await _emp_lookup([r.get("user_id") for r in raw if r.get("user_id")])
        raw.sort(key=lambda r: str(r.get("employee_code") or ""))
        for i, r in enumerate(raw, start=1):
            u = emp.get(r.get("user_id") or "", {})
            if kind == "pf":
                ee = float(r.get("pf_employee") or 0)
                er = float(r.get("pf_employer_total") or 0)
                rows_out.append({
                    "sr": i,
                    "user_id": r.get("user_id"),
                    "employee_code": r.get("employee_code") or u.get("employee_code") or "",
                    "name": r.get("name") or u.get("name") or "",
                    "uan_no": u.get("uan_no") or "",
                    "pf_wages": float(r.get("pf_wages") or 0),
                    "pf_employee": ee,
                    "pf_employer_eps": float(r.get("pf_employer_eps") or 0),
                    "pf_employer_epf": float(r.get("pf_employer_epf") or 0),
                    "pf_employer_total": er,
                    "total": round(ee + er, 2),
                })
            else:
                ee = float(r.get("esic_employee") or 0)
                er = float(r.get("esic_employer") or 0)
                rows_out.append({
                    "sr": i,
                    "user_id": r.get("user_id"),
                    "employee_code": r.get("employee_code") or u.get("employee_code") or "",
                    "name": r.get("name") or u.get("name") or "",
                    "esi_ip_no": u.get("esi_ip_no") or "",
                    "esic_wage_base": float(r.get("esic_wage_base") or 0),
                    "esic_employee": ee,
                    "esic_employer": er,
                    "total": round(ee + er, 2),
                })

    num_keys = [c["key"] for c in columns if c["key"] not in ("sr", "name", "uan_no", "esi_ip_no")]
    totals = {k: round(sum(float(r.get(k) or 0) for r in rows_out), 2) for k in num_keys}
    return {
        "kind": kind,
        "month": month,
        "company_id": company_id,
        "run_found": bool(run),
        "columns": columns,
        "rows": rows_out,
        "totals": totals,
        "employees_count": len(rows_out),
    }


# ---------------------------------------------------------------------------
# Employee-wise yearly contribution report
# ---------------------------------------------------------------------------

async def _yearly_contribution(company_id: str, fy_start_year: int, kind: str) -> Dict[str, Any]:
    months = _fy_months(fy_start_year)
    per_user: Dict[str, Dict[str, Any]] = {}
    months_covered: List[str] = []
    for m in months:
        run = await _latest_run(company_id, m["key"])
        if not run:
            continue
        months_covered.append(m["key"])
        for r in run.get("rows") or []:
            if kind == "pf":
                ee = float(r.get("pf_employee") or 0)
                er = float(r.get("pf_employer_total") or 0)
                wages = float(r.get("pf_wages") or 0)
            else:
                ee = float(r.get("esic_employee") or 0)
                er = float(r.get("esic_employer") or 0)
                wages = float(r.get("esic_wage_base") or 0)
            if ee <= 0 and er <= 0:
                continue
            uid = r.get("user_id") or ""
            agg = per_user.setdefault(uid, {
                "user_id": uid,
                "employee_code": r.get("employee_code") or "",
                "name": r.get("name") or "",
                "monthly": {},
                "wages_total": 0.0, "ee_total": 0.0, "er_total": 0.0,
            })
            agg["monthly"][m["key"]] = round(ee + er, 2)
            agg["wages_total"] = round(agg["wages_total"] + wages, 2)
            agg["ee_total"] = round(agg["ee_total"] + ee, 2)
            agg["er_total"] = round(agg["er_total"] + er, 2)

    emp = await _emp_lookup(list(per_user.keys()))
    rows = []
    for uid, agg in per_user.items():
        u = emp.get(uid, {})
        agg["uan_no"] = u.get("uan_no") or ""
        agg["esi_ip_no"] = u.get("esi_ip_no") or ""
        agg["grand_total"] = round(agg["ee_total"] + agg["er_total"], 2)
        rows.append(agg)
    rows.sort(key=lambda r: str(r.get("employee_code") or ""))
    for i, r in enumerate(rows, start=1):
        r["sr"] = i

    totals = {
        "wages_total": round(sum(r["wages_total"] for r in rows), 2),
        "ee_total": round(sum(r["ee_total"] for r in rows), 2),
        "er_total": round(sum(r["er_total"] for r in rows), 2),
        "grand_total": round(sum(r["grand_total"] for r in rows), 2),
        "monthly": {
            m["key"]: round(sum(float(r["monthly"].get(m["key"]) or 0) for r in rows), 2)
            for m in months
        },
    }
    return {
        "kind": kind,
        "fy_start_year": fy_start_year,
        "fy_label": f"FY {fy_start_year}-{str(fy_start_year + 1)[-2:]}",
        "company_id": company_id,
        "months": months,
        "months_covered": months_covered,
        "rows": rows,
        "totals": totals,
        "employees_count": len(rows),
    }


# ---------------------------------------------------------------------------
# Bonus Yearly Summary
# ---------------------------------------------------------------------------

async def _bonus_yearly_summary(company_id: str, fy_start_year: int) -> Dict[str, Any]:
    months = _fy_months(fy_start_year)

    # Earning allowance heads enabled in Firm Master (compliance policy).
    comp = await db.companies.find_one(
        {"company_id": company_id}, {"_id": 0, "compliance_policy": 1, "name": 1},
    )
    enabled = ((comp or {}).get("compliance_policy") or {}).get("enabled_allowances")
    if enabled and isinstance(enabled, list):
        heads = [h for h in _ALL_HEADS if h in {str(x).lower() for x in enabled} or h == "basic"]
    else:
        heads = list(_ALL_HEADS)

    per_user: Dict[str, Dict[str, Any]] = {}
    months_covered: List[str] = []
    for m in months:
        run = await _latest_run(company_id, m["key"])
        if not run:
            continue
        months_covered.append(m["key"])
        for r in run.get("rows") or []:
            earned = float(r.get("gross_paid") or 0)
            days = float(r.get("present_days") or 0)
            if earned <= 0 and days <= 0:
                continue
            uid = r.get("user_id") or ""
            agg = per_user.setdefault(uid, {
                "user_id": uid,
                "employee_code": r.get("employee_code") or "",
                "name": r.get("name") or "",
                "monthly": {},
                "total_days": 0.0,
                "total_earned": 0.0,
                **{f"head_{h}": 0.0 for h in heads},
            })
            agg["monthly"][m["key"]] = {"days": days, "earned": round(earned, 2)}
            agg["total_days"] = round(agg["total_days"] + days, 2)
            agg["total_earned"] = round(agg["total_earned"] + earned, 2)
            for h in heads:
                agg[f"head_{h}"] = round(agg[f"head_{h}"] + float(r.get(h) or 0), 2)

    emp = await _emp_lookup(list(per_user.keys()))
    rows = []
    for uid, agg in per_user.items():
        u = emp.get(uid, {})
        agg["father_name"] = u.get("father_name") or ""
        agg["doj"] = u.get("doj") or ""
        rows.append(agg)
    rows.sort(key=lambda r: str(r.get("employee_code") or ""))
    for i, r in enumerate(rows, start=1):
        r["sr"] = i

    totals = {
        "total_days": round(sum(r["total_days"] for r in rows), 2),
        "total_earned": round(sum(r["total_earned"] for r in rows), 2),
        **{f"head_{h}": round(sum(r.get(f"head_{h}") or 0 for r in rows), 2) for h in heads},
        "monthly": {
            m["key"]: {
                "days": round(sum(float((r["monthly"].get(m["key"]) or {}).get("days") or 0) for r in rows), 2),
                "earned": round(sum(float((r["monthly"].get(m["key"]) or {}).get("earned") or 0) for r in rows), 2),
            }
            for m in months
        },
    }
    return {
        "fy_start_year": fy_start_year,
        "fy_label": f"FY {fy_start_year}-{str(fy_start_year + 1)[-2:]}",
        "company_id": company_id,
        "company_name": (comp or {}).get("name"),
        "months": months,
        "months_covered": months_covered,
        "heads": [{"key": f"head_{h}", "label": _HEAD_LABELS[h]} for h in heads],
        "rows": rows,
        "totals": totals,
        "employees_count": len(rows),
    }


# ---------------------------------------------------------------------------
# XLSX helper
# ---------------------------------------------------------------------------

def _xlsx_response(title: str, header: List[str], data_rows: List[List[Any]],
                   totals_row: Optional[List[Any]], filename: str) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1E3A8A")

    ncols = max(len(header), 3)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=13, color="1E3A8A")
    c.alignment = Alignment(horizontal="center")

    head_row = 3
    for ci, lbl in enumerate(header, start=1):
        cell = ws.cell(row=head_row, column=ci, value=lbl)
        cell.font = hdr_font
        cell.fill = hdr_fill
        ws.column_dimensions[cell.column_letter].width = max(12, min(24, len(str(lbl)) + 4))
    for ri, row in enumerate(data_rows, start=head_row + 1):
        for ci, v in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=v)
    if totals_row:
        ri = head_row + len(data_rows) + 1
        for ci, v in enumerate(totals_row, start=1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/contribution")
async def contribution_sheet(
    kind: str = "pf",
    company_id: Optional[str] = None,
    month: str = "",
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    k = _kind_ok(kind)
    cid = _resolve_company(admin, company_id)
    if not month or len(month) != 7:
        raise HTTPException(status_code=400, detail="month (YYYY-MM) is required")
    return await _monthly_contribution(cid, month, k)


@router.get("/contribution.xlsx")
async def contribution_sheet_xlsx(
    kind: str = "pf",
    company_id: Optional[str] = None,
    month: str = "",
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    k = _kind_ok(kind)
    cid = _resolve_company(admin, company_id)
    data = await _monthly_contribution(cid, month, k)
    cols = data["columns"]
    header = [c["label"] for c in cols]
    rows = [[r.get(c["key"]) for c in cols] for r in data["rows"]]
    totals_row = ["", "TOTAL"] + [
        data["totals"].get(c["key"], "") for c in cols[2:]
    ]
    label = "P.F." if k == "pf" else "E.S.I."
    return _xlsx_response(
        f"{label} Contribution Sheet — {month}", header, rows, totals_row,
        f"{'PF' if k == 'pf' else 'ESI'}_Contribution_{month}.xlsx",
    )


@router.get("/contribution-yearly")
async def contribution_yearly(
    kind: str = "pf",
    company_id: Optional[str] = None,
    fy_start_year: int = 0,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    k = _kind_ok(kind)
    cid = _resolve_company(admin, company_id)
    fy = fy_start_year or (date.today().year if date.today().month >= 4 else date.today().year - 1)
    return await _yearly_contribution(cid, fy, k)


@router.get("/contribution-yearly.xlsx")
async def contribution_yearly_xlsx(
    kind: str = "pf",
    company_id: Optional[str] = None,
    fy_start_year: int = 0,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    k = _kind_ok(kind)
    cid = _resolve_company(admin, company_id)
    fy = fy_start_year or (date.today().year if date.today().month >= 4 else date.today().year - 1)
    data = await _yearly_contribution(cid, fy, k)
    id_label = "UAN No." if k == "pf" else "ESIC IP No."
    id_key = "uan_no" if k == "pf" else "esi_ip_no"
    header = (["Sr.", "Emp Code", "Employee Name", id_label]
              + [m["label"] for m in data["months"]]
              + ["Wages Total", "Employee Total", "Employer Total", "Grand Total"])
    rows = []
    for r in data["rows"]:
        rows.append([r["sr"], r["employee_code"], r["name"], r.get(id_key) or ""]
                    + [r["monthly"].get(m["key"]) or 0 for m in data["months"]]
                    + [r["wages_total"], r["ee_total"], r["er_total"], r["grand_total"]])
    t = data["totals"]
    totals_row = (["", "", "TOTAL", ""]
                  + [t["monthly"].get(m["key"]) or 0 for m in data["months"]]
                  + [t["wages_total"], t["ee_total"], t["er_total"], t["grand_total"]])
    label = "P.F." if k == "pf" else "E.S.I."
    return _xlsx_response(
        f"{label} Contribution — Employee-wise Yearly ({data['fy_label']})",
        header, rows, totals_row,
        f"{'PF' if k == 'pf' else 'ESI'}_Contribution_Yearly_{fy}.xlsx",
    )


@router.get("/bonus-yearly-summary")
async def bonus_yearly_summary(
    company_id: Optional[str] = None,
    fy_start_year: int = 0,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _resolve_company(admin, company_id)
    fy = fy_start_year or (date.today().year if date.today().month >= 4 else date.today().year - 1)
    return await _bonus_yearly_summary(cid, fy)


@router.get("/bonus-yearly-summary.xlsx")
async def bonus_yearly_summary_xlsx(
    company_id: Optional[str] = None,
    fy_start_year: int = 0,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _resolve_company(admin, company_id)
    fy = fy_start_year or (date.today().year if date.today().month >= 4 else date.today().year - 1)
    data = await _bonus_yearly_summary(cid, fy)
    header = (["Sr.", "Emp Code", "Employee Name", "Father Name", "Date of Join"]
              + [f"{m['label']} Days" for m in data["months"]]
              + [f"{m['label']} Earned" for m in data["months"]]
              + [h["label"] + " (Yr)" for h in data["heads"]]
              + ["Total Working Days", "Total Earned"])
    rows = []
    for r in data["rows"]:
        rows.append(
            [r["sr"], r["employee_code"], r["name"], r["father_name"], r["doj"]]
            + [(r["monthly"].get(m["key"]) or {}).get("days") or 0 for m in data["months"]]
            + [(r["monthly"].get(m["key"]) or {}).get("earned") or 0 for m in data["months"]]
            + [r.get(h["key"]) or 0 for h in data["heads"]]
            + [r["total_days"], r["total_earned"]]
        )
    t = data["totals"]
    totals_row = (["", "", "TOTAL", "", ""]
                  + [(t["monthly"].get(m["key"]) or {}).get("days") or 0 for m in data["months"]]
                  + [(t["monthly"].get(m["key"]) or {}).get("earned") or 0 for m in data["months"]]
                  + [t.get(h["key"]) or 0 for h in data["heads"]]
                  + [t["total_days"], t["total_earned"]])
    return _xlsx_response(
        f"Bonus Yearly Summary — {data['fy_label']} — {data.get('company_name') or ''}",
        header, rows, totals_row,
        f"Bonus_Yearly_Summary_{fy}.xlsx",
    )
