"""Iter 501 — CLIENT ATTENDANCE IMPORT (Attendance Summary Excel).

Additive module — imports client-supplied attendance SUMMARY sheets
(Employee Code, Date, In/Out, Shift, Late/Early/OT/Work hours, Present /
Absent / Leave / Paid Days / WO / Holiday / CL / SL / PL / OD / CO)
alongside the existing Biometric Sync and Punch Import modules. Nothing in
the attendance calculation engine, salary process or device sync changes.

Data written:
  * db.attendance          — IN/OUT punches (source="client_import",
                             auto-approved) so times flow into every
                             existing report exactly like the punch import.
  * db.client_attendance_days — the client's day-summary AS SUPPLIED
                             (hours are NOT recalculated).
  * db.client_attendance_imports — import log (re-downloadable error report,
                             delete = full rollback of that import).
  * db.client_import_templates — saved column-mapping templates.
  * db.compliance_import_entries — OPTIONAL (opt-in): monthly Present Days
                             for the existing "Imported Sheet" compliance
                             run mode. Off by default.
"""
import base64
import io
import re
import sys
import uuid
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Body, Header, HTTPException, Query

sys.path.append("/app/backend")
from server import db, get_user_from_token, require_role, now_iso  # noqa: E402
from routes.punch_import import (  # noqa: E402
    _read_grid, _parse_date_cell, _parse_time_cell,
)

router = APIRouter(prefix="/api/admin/client-attendance", tags=["client-attendance"])

MAX_BYTES = 20 * 1024 * 1024  # 20 MB

# ---------------------------------------------------------------------------
# Field registry + header synonyms (auto-detect)
# ---------------------------------------------------------------------------
FIELDS: List[Tuple[str, str]] = [
    ("employee_code", "Employee Code"),
    ("name", "Employee Name"),
    ("department", "Department"),
    ("date", "Attendance Date"),
    ("in_time", "In Time"),
    ("out_time", "Out Time"),
    ("shift", "Shift"),
    ("late_hours", "Late Time"),
    ("early_hours", "Early Exit"),
    ("ot_hours", "OT Hours"),
    ("work_hours", "Working Hours"),
    ("present", "Present"),
    ("absent", "Absent"),
    ("leave", "Leave"),
    ("paid_days", "Paid Days"),
    ("weekly_off", "Weekly Off"),
    ("holiday", "Holiday"),
    ("cl", "CL"), ("sl", "SL"), ("pl", "PL"), ("od", "OD"), ("co", "CO"),
]
SYNONYMS: Dict[str, List[str]] = {
    "employee_code": ["employeecode", "empcode", "code", "empno", "employeeno",
                      "empid", "employeeid", "biocode", "bio", "cardno",
                      "punchcode", "machinecode", "paycode"],
    "name": ["employeename", "empname", "name", "staffname", "workername"],
    "department": ["department", "dept", "deptt"],
    "date": ["attendancedate", "attdate", "date", "day", "punchdate"],
    "in_time": ["intime", "firstpunch", "punchin", "timein", "in", "first",
                "arrival"],
    "out_time": ["outtime", "lastpunch", "punchout", "timeout", "out", "last",
                 "departure"],
    "shift": ["shift", "shiftname", "shiftcode"],
    "late_hours": ["latehrs", "latetime", "lateby", "late", "latehours",
                   "latecoming"],
    "early_hours": ["earlyhrs", "earlyexit", "earlyby", "early", "earlygoing",
                    "earlyhours", "earlydeparture"],
    "ot_hours": ["othrs", "othours", "overtime", "otamounthrs", "ot",
                 "overtimehours", "othr"],
    "work_hours": ["workinghours", "workhrs", "workhours", "workhr",
                   "totalhours", "totalhrs", "duration", "workdur", "hrs",
                   "hoursworked", "workedhours", "totalduration"],
    "present": ["presentdays", "present", "presnt", "prsnt"],
    "absent": ["absentdays", "absent", "abst"],
    "leave": ["leavetype", "leave", "lv", "attstatus", "status", "attendance"],
    "paid_days": ["paiddays", "payabledays", "paydays", "payable"],
    "weekly_off": ["weeklyoff", "weekoff", "wo", "woff"],
    "holiday": ["holidays", "holiday", "hl", "hd"],
    "cl": ["cl", "casualleave"],
    "sl": ["sl", "sickleave"],
    "pl": ["pl", "privilegeleave", "el", "earnedleave"],
    "od": ["od", "onduty", "outdoorduty"],
    "co": ["co", "compoff", "compensatoryoff"],
}
# leave-code → status (mapped to the existing leave vocabulary)
LEAVE_CODES = {
    "CL": "casual_leave", "SL": "sick_leave", "PL": "privilege_leave",
    "EL": "earned_leave", "OD": "on_duty", "CO": "comp_off",
    "WO": "weekly_off", "WH": "holiday", "H": "holiday",
    "P": "present", "A": "absent", "L": "leave", "LWP": "leave_without_pay",
    "ML": "medical_leave", "HD": "half_day",
}
STATUS_LABEL = {
    "present": "Present", "absent": "Absent", "casual_leave": "Casual Leave",
    "sick_leave": "Sick Leave", "privilege_leave": "Privilege Leave",
    "earned_leave": "Earned Leave", "on_duty": "On Duty",
    "comp_off": "Comp Off", "weekly_off": "Weekly Off", "holiday": "Holiday",
    "leave": "Leave", "leave_without_pay": "Leave Without Pay",
    "medical_leave": "Medical Leave", "half_day": "Half Day",
}


def _norm(h: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


def _auto_map(headers: List[Any]) -> Dict[str, int]:
    """field → column index; longest synonym wins, one column per field."""
    normed = [_norm(h) for h in headers]
    mapping: Dict[str, int] = {}
    used: set = set()
    # exact match first, then contains
    for exact in (True, False):
        for field, syns in SYNONYMS.items():
            if field in mapping:
                continue
            for syn in syns:
                hit = None
                for i, h in enumerate(normed):
                    if i in used or not h:
                        continue
                    if (h == syn) if exact else (syn in h):
                        hit = i
                        break
                if hit is not None:
                    mapping[field] = hit
                    used.add(hit)
                    break
    return mapping


def _num(v: Any) -> Optional[float]:
    """Parse hour/day cells: 1.5, '1.30', '01:30', '2', '' → float | None."""
    if v is None or v == "":
        return None
    s = str(v).strip()
    if not s:
        return None
    if ":" in s:
        try:
            parts = s.split(":")
            return round(float(parts[0]) + float(parts[1]) / 60.0
                         + (float(parts[2]) / 3600.0 if len(parts) > 2 else 0), 2)
        except (ValueError, IndexError):
            return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def _code_keys(v: Any) -> List[str]:
    """Normalised employee-code keys ('0101', 101, '101.0' all match)."""
    s = str(v or "").strip()
    if not s:
        return []
    keys = [s.upper()]
    try:
        keys.append(str(int(float(s))))
    except (TypeError, ValueError):
        pass
    return keys


def _detect_header_row(grid: List[List[Any]]) -> int:
    """First row (within the top 10) that auto-maps ≥3 known fields."""
    best_i, best_n = 0, -1
    for i, row in enumerate(grid[:10]):
        n = len(_auto_map(row))
        if n > best_n:
            best_i, best_n = i, n
    return best_i


def _row_status(vals: Dict[str, Any]) -> Tuple[str, str]:
    """(status, leave_code) per the requirement's precedence."""
    if (_num(vals.get("present")) or 0) > 0:
        return "present", ""
    if (_num(vals.get("absent")) or 0) > 0:
        return "absent", ""
    lv = str(vals.get("leave") or "").strip().upper()
    if lv:
        st = LEAVE_CODES.get(lv)
        if st:
            return st, lv if st not in ("present", "absent") else ""
        return "leave", lv  # unknown code → generic leave, code preserved
    for f, st in (("cl", "casual_leave"), ("sl", "sick_leave"),
                  ("pl", "privilege_leave"), ("od", "on_duty"),
                  ("co", "comp_off")):
        if (_num(vals.get(f)) or 0) > 0:
            return st, f.upper()
    if (_num(vals.get("weekly_off")) or 0) > 0:
        return "weekly_off", ""
    if (_num(vals.get("holiday")) or 0) > 0:
        return "holiday", ""
    if vals.get("in_time") or vals.get("out_time"):
        return "present", ""
    return "", ""


async def _auth(authorization: Optional[str], company_id: str = ""):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    if company_id:
        if admin.get("role") == "sub_admin":
            from server import sub_admin_can_touch_company
            if not sub_admin_can_touch_company(admin, company_id):
                raise HTTPException(status_code=403, detail="Firm not in your scope")
        elif (admin.get("role") == "company_admin"
                and admin.get("company_id") != company_id):
            raise HTTPException(status_code=403, detail="Not authorised for this firm")
    return admin


# ---------------------------------------------------------------------------
# Sample template
# ---------------------------------------------------------------------------
@router.get("/template")
async def download_template(authorization: Optional[str] = Header(None)):
    await _auth(authorization)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = ["Employee Code", "Employee Name", "Department", "Date",
               "Intime", "Outtime", "Shift", "Late", "Early", "OT",
               "Working Hours", "Present", "Absent", "Leave", "Paid Days",
               "Weekly Off", "Holiday"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1D4ED8")
    ws.append(["101", "RAKESH KUMAR", "PRODUCTION", "01-06-2026", "09:00",
               "18:05", "GEN", "0:00", "0:00", "1:05", "9:05", 1, "", "", 1,
               "", ""])
    ws.append(["102", "SUNITA DEVI", "PACKING", "01-06-2026", "", "", "GEN",
               "", "", "", "", "", "", "CL", 1, "", ""])
    ws.append(["103", "MOHAN LAL", "PRODUCTION", "01-06-2026", "", "", "GEN",
               "", "", "", "", "", 1, "", 0, "", ""])
    for i, w in enumerate([14, 22, 14, 12, 8, 8, 8, 8, 8, 8, 12, 8, 8, 8,
                           10, 10, 8], start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(38 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return {"filename": "client_attendance_template.xlsx",
            "file_base64": base64.b64encode(buf.getvalue()).decode()}


# ---------------------------------------------------------------------------
# Preview (parse + auto-detect + validate + stage)
# ---------------------------------------------------------------------------
@router.post("/preview")
async def preview(payload: Dict[str, Any] = Body(...),
                  authorization: Optional[str] = Header(None)):
    company_id = str(payload.get("company_id") or "").strip()
    admin = await _auth(authorization, company_id)
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id required")
    b64 = payload.get("file_base64") or ""
    staging_id = payload.get("staging_id")
    mapping_in = payload.get("mapping")  # {field: col_index} — optional

    if staging_id and not b64:
        # re-validate a previously uploaded file with a NEW mapping —
        # no re-upload needed (large files).
        stg = await db.client_import_staging.find_one(
            {"staging_id": staging_id}, {"_id": 0, "file_b64": 1,
                                         "filename": 1})
        if not stg:
            raise HTTPException(status_code=404, detail="Upload expired — upload the file again")
        b64 = stg["file_b64"]
        filename = stg.get("filename") or "attendance.xlsx"
    else:
        filename = str(payload.get("filename") or "attendance.xlsx")

    try:
        data = base64.b64decode(b64)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid file data")
    if len(data) > MAX_BYTES:
        raise HTTPException(status_code=400, detail="File exceeds the 20 MB limit")
    grid = _read_grid(data)
    if not grid:
        raise HTTPException(status_code=400, detail="Empty sheet")

    hrow = _detect_header_row(grid)
    headers = [str(h or "").strip() for h in grid[hrow]]
    mapping: Dict[str, int] = {}
    if isinstance(mapping_in, dict) and mapping_in:
        for f, idx in mapping_in.items():
            try:
                idx = int(idx)
            except (TypeError, ValueError):
                continue
            if 0 <= idx < len(headers):
                mapping[f] = idx
    else:
        mapping = _auto_map(headers)
    if "employee_code" not in mapping and "name" not in mapping:
        raise HTTPException(
            status_code=400,
            detail="Could not find an Employee Code or Name column — map it manually")
    if "date" not in mapping:
        raise HTTPException(status_code=400,
                            detail="Could not find a Date column — map it manually")

    # employees of the firm for matching
    by_code: Dict[str, dict] = {}
    by_name: Dict[str, dict] = {}
    async for u in db.users.find(
            {"company_id": company_id, "role": "employee",
             "deleted": {"$ne": True}},
            {"_id": 0, "user_id": 1, "employee_code": 1, "name": 1,
             "bio_code": 1}):
        for k in _code_keys(u.get("employee_code")) + _code_keys(u.get("bio_code")):
            by_code.setdefault(k, u)
        nm = re.sub(r"\s+", " ", str(u.get("name") or "").strip().upper())
        if nm:
            by_name.setdefault(nm, u)

    valid: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    missing: Dict[str, dict] = {}
    counts: Dict[str, int] = defaultdict(int)
    seen_in_file: set = set()

    def cell(row: List[Any], f: str) -> Any:
        i = mapping.get(f)
        return row[i] if i is not None and i < len(row) else None

    for rn, row in enumerate(grid[hrow + 1:], start=hrow + 2):
        if not any(str(c or "").strip() for c in row):
            continue
        vals = {f: cell(row, f) for f, _ in FIELDS}
        code_raw = str(vals.get("employee_code") or "").strip()
        name_raw = re.sub(r"\s+", " ", str(vals.get("name") or "").strip().upper())
        if not code_raw and not name_raw:
            invalid.append({"row": rn, "code": "", "name": "",
                            "date": str(vals.get("date") or ""),
                            "reason": "Blank employee code & name"})
            continue
        emp = None
        for k in _code_keys(code_raw):
            emp = by_code.get(k)
            if emp:
                break
        if emp is None and name_raw:
            emp = by_name.get(name_raw)
        if emp is None:
            key = code_raw or name_raw
            missing[key] = {"code": code_raw, "name": str(vals.get("name") or "").strip(),
                            "rows": missing.get(key, {}).get("rows", 0) + 1}
            invalid.append({"row": rn, "code": code_raw,
                            "name": str(vals.get("name") or ""),
                            "date": str(vals.get("date") or ""),
                            "reason": "Employee not found"})
            continue
        date = _parse_date_cell(vals.get("date"))
        if not date:
            invalid.append({"row": rn, "code": code_raw, "name": emp.get("name"),
                            "date": str(vals.get("date") or ""),
                            "reason": "Invalid date"})
            continue
        in_t = _parse_time_cell(vals.get("in_time"))
        out_t = _parse_time_cell(vals.get("out_time"))
        if vals.get("in_time") not in (None, "") and in_t is None:
            invalid.append({"row": rn, "code": code_raw, "name": emp.get("name"),
                            "date": date, "reason": f"Invalid In Time '{vals.get('in_time')}'"})
            continue
        if vals.get("out_time") not in (None, "") and out_t is None:
            invalid.append({"row": rn, "code": code_raw, "name": emp.get("name"),
                            "date": date, "reason": f"Invalid Out Time '{vals.get('out_time')}'"})
            continue
        hours_bad = None
        hours: Dict[str, Optional[float]] = {}
        for hf in ("work_hours", "late_hours", "early_hours", "ot_hours",
                   "paid_days"):
            raw = vals.get(hf)
            hv = _num(raw)
            if raw not in (None, "") and hv is None:
                hours_bad = f"Invalid {hf.replace('_', ' ')} '{raw}'"
                break
            if hv is not None and (hv < 0 or hv > (31 if hf == "paid_days" else 24)):
                hours_bad = f"Out-of-range {hf.replace('_', ' ')} '{raw}'"
                break
            hours[hf] = hv
        if hours_bad:
            invalid.append({"row": rn, "code": code_raw, "name": emp.get("name"),
                            "date": date, "reason": hours_bad})
            continue
        status, leave_code = _row_status({**vals, "in_time": in_t, "out_time": out_t})
        if not status:
            invalid.append({"row": rn, "code": code_raw, "name": emp.get("name"),
                            "date": date, "reason": "No attendance data (no P/A/Leave/times)"})
            continue
        fkey = (emp["user_id"], date)
        if fkey in seen_in_file:
            invalid.append({"row": rn, "code": code_raw, "name": emp.get("name"),
                            "date": date, "reason": "Duplicate row in file (same employee + date)"})
            continue
        seen_in_file.add(fkey)
        missing_punch = ""
        if in_t and not out_t:
            missing_punch = "out"
        elif out_t and not in_t:
            missing_punch = "in"
        counts[status] += 1
        valid.append({
            "row": rn, "user_id": emp["user_id"],
            "employee_code": emp.get("employee_code") or code_raw,
            "name": emp.get("name") or "", "date": date,
            "in_time": in_t or "", "out_time": out_t or "",
            "missing_punch": missing_punch,
            "shift": str(vals.get("shift") or "").strip(),
            "department": str(vals.get("department") or "").strip(),
            "status": status, "leave_code": leave_code,
            "work_hours": hours.get("work_hours"),
            "late_hours": hours.get("late_hours"),
            "early_hours": hours.get("early_hours"),
            "ot_hours": hours.get("ot_hours"),
            "paid_days": hours.get("paid_days"),
        })

    # duplicates against the DATABASE (existing imported day OR any punch)
    dup_keys: set = set()
    if valid:
        uids = list({r["user_id"] for r in valid})
        dmin = min(r["date"] for r in valid)
        dmax = max(r["date"] for r in valid)
        async for d in db.client_attendance_days.find(
                {"company_id": company_id, "user_id": {"$in": uids},
                 "date": {"$gte": dmin, "$lte": dmax}},
                {"_id": 0, "user_id": 1, "date": 1}):
            dup_keys.add((d["user_id"], d["date"]))
        async for p in db.attendance.find(
                {"company_id": company_id, "user_id": {"$in": uids},
                 "date": {"$gte": dmin, "$lte": dmax}},
                {"_id": 0, "user_id": 1, "date": 1}):
            dup_keys.add((p["user_id"], p["date"]))
    dups = sum(1 for r in valid if (r["user_id"], r["date"]) in dup_keys)
    for r in valid:
        r["duplicate"] = (r["user_id"], r["date"]) in dup_keys

    stats = {
        "total": len(valid) + len(invalid),
        "valid": len(valid),
        "invalid": len(invalid),
        "duplicates": dups,
        "missing_employees": len(missing),
        "present": counts.get("present", 0),
        "absent": counts.get("absent", 0),
        "leave": sum(v for k, v in counts.items()
                     if k not in ("present", "absent", "weekly_off", "holiday")),
        "weekly_off": counts.get("weekly_off", 0),
        "holiday": counts.get("holiday", 0),
    }
    new_staging = f"cstg_{uuid.uuid4().hex[:12]}"
    await db.client_import_staging.delete_many(
        {"company_id": company_id, "created_by": admin["user_id"],
         "consumed": {"$ne": True}})
    await db.client_import_staging.insert_one({
        "staging_id": new_staging, "company_id": company_id,
        "filename": filename, "file_b64": b64, "mapping": mapping,
        "rows": valid, "invalid": invalid,
        "stats": stats, "consumed": False,
        "created_by": admin["user_id"], "created_at": now_iso()})
    return {
        "staging_id": new_staging,
        "headers": headers,
        "header_row": hrow,
        "mapping": mapping,
        "fields": [{"key": k, "label": lb} for k, lb in FIELDS],
        "stats": stats,
        "sample": valid[:15],
        "invalid_rows": invalid[:200],
        "missing_employees": [
            {"code": v["code"], "name": v["name"], "rows": v["rows"]}
            for v in list(missing.values())[:200]],
    }


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------
@router.post("/commit")
async def commit(payload: Dict[str, Any] = Body(...),
                 authorization: Optional[str] = Header(None)):
    staging_id = str(payload.get("staging_id") or "")
    dup_mode = str(payload.get("duplicate_mode") or "skip").lower()
    if dup_mode not in ("replace", "skip", "merge"):
        raise HTTPException(status_code=400,
                            detail="duplicate_mode must be replace|skip|merge")
    sync_compliance = bool(payload.get("sync_compliance"))
    stg = await db.client_import_staging.find_one(
        {"staging_id": staging_id}, {"_id": 0})
    if not stg:
        raise HTTPException(status_code=404, detail="Upload expired — run Preview again")
    company_id = stg["company_id"]
    admin = await _auth(authorization, company_id)
    if stg.get("consumed"):
        # refresh-safe: the same staging can never import twice.
        log = await db.client_attendance_imports.find_one(
            {"staging_id": staging_id}, {"_id": 0, "errors": 0})
        return {"ok": True, "already_imported": True, "log": log}
    rows: List[Dict[str, Any]] = stg.get("rows") or []
    if not rows:
        raise HTTPException(status_code=400, detail="No valid rows to import")

    import_id = f"cimp_{uuid.uuid4().hex[:12]}"
    uids = list({r["user_id"] for r in rows})
    dmin = min(r["date"] for r in rows)
    dmax = max(r["date"] for r in rows)

    # existing punches (any source) for merge/dedupe decisions
    existing_kinds: Dict[Tuple[str, str], set] = defaultdict(set)
    async for p in db.attendance.find(
            {"company_id": company_id, "user_id": {"$in": uids},
             "date": {"$gte": dmin, "$lte": dmax}},
            {"_id": 0, "user_id": 1, "date": 1, "kind": 1}):
        existing_kinds[(p["user_id"], p["date"])].add(p["kind"])
    existing_days: set = set()
    async for d in db.client_attendance_days.find(
            {"company_id": company_id, "user_id": {"$in": uids},
             "date": {"$gte": dmin, "$lte": dmax}},
            {"_id": 0, "user_id": 1, "date": 1}):
        existing_days.add((d["user_id"], d["date"]))

    imported = skipped = 0
    punches_created = 0
    replaced_pairs: List[Dict[str, str]] = []
    day_docs: List[Dict[str, Any]] = []
    punch_docs: List[Dict[str, Any]] = []
    now = now_iso()

    for r in rows:
        key = (r["user_id"], r["date"])
        is_dup = key in existing_days or key in existing_kinds
        if is_dup and dup_mode == "skip":
            skipped += 1
            continue
        if is_dup and dup_mode == "replace":
            replaced_pairs.append({"user_id": r["user_id"], "date": r["date"]})
        add_in = bool(r.get("in_time"))
        add_out = bool(r.get("out_time"))
        if is_dup and dup_mode == "merge":
            kinds = existing_kinds.get(key) or set()
            add_in = add_in and "in" not in kinds
            add_out = add_out and "out" not in kinds
        day_docs.append({
            "rec_id": f"cday_{uuid.uuid4().hex[:12]}",
            "import_id": import_id, "company_id": company_id,
            "user_id": r["user_id"], "employee_code": r["employee_code"],
            "name": r["name"], "date": r["date"],
            "status": r["status"],
            "status_label": STATUS_LABEL.get(r["status"], r["status"]),
            "leave_code": r.get("leave_code") or "",
            "in_time": r.get("in_time") or "",
            "out_time": r.get("out_time") or "",
            "missing_punch": r.get("missing_punch") or "",
            "shift": r.get("shift") or "",
            "department": r.get("department") or "",
            "work_hours": r.get("work_hours"),
            "late_hours": r.get("late_hours"),
            "early_hours": r.get("early_hours"),
            "ot_hours": r.get("ot_hours"),
            "paid_days": r.get("paid_days"),
            "source": "client_import", "created_at": now,
            "created_by": admin["user_id"],
        })
        for kind, tval, add in (("in", r.get("in_time"), add_in),
                                ("out", r.get("out_time"), add_out)):
            if not (tval and add):
                continue
            punch_docs.append({
                "record_id": f"att_{uuid.uuid4().hex[:12]}",
                "user_id": r["user_id"], "company_id": company_id,
                "date": r["date"], "kind": kind,
                "at": f"{r['date']}T{tval}:00Z",
                "source": "client_import", "status": "approved",
                "approved_by": admin["user_id"],
                "manual_reason": "Client attendance import",
                "import_batch_id": import_id,
                "created_by": admin["user_id"], "created_at": now,
            })
        imported += 1

    # REPLACE mode: remove the previous client-imported data for those
    # employee+date pairs (biometric punches are NEVER touched).
    for i in range(0, len(replaced_pairs), 400):
        chunk = replaced_pairs[i:i + 400]
        ors = [{"user_id": p["user_id"], "date": p["date"]} for p in chunk]
        await db.client_attendance_days.delete_many(
            {"company_id": company_id, "$or": ors})
        await db.attendance.delete_many(
            {"company_id": company_id, "source": "client_import", "$or": ors})

    for i in range(0, len(day_docs), 1000):
        await db.client_attendance_days.insert_many(day_docs[i:i + 1000])
    for i in range(0, len(punch_docs), 1000):
        await db.attendance.insert_many(punch_docs[i:i + 1000])
        punches_created += len(punch_docs[i:i + 1000])

    # OPT-IN: monthly Present Days → compliance "Imported Sheet" entries.
    synced_months: List[str] = []
    if sync_compliance:
        month_days: Dict[Tuple[str, str], float] = defaultdict(float)
        for d in day_docs:
            month = d["date"][:7]
            add = 0.0
            if d["status"] == "present":
                add = 1.0
            elif d["status"] == "half_day":
                add = 0.5
            elif d.get("paid_days") is not None:
                add = min(float(d["paid_days"]), 1.0)
            month_days[(d["user_id"], month)] += add
        for (uid, month), pdays in month_days.items():
            await db.compliance_import_entries.update_one(
                {"company_id": company_id, "user_id": uid, "month": month},
                {"$set": {"present_days": round(pdays, 1),
                          "source": "client_attendance_import",
                          "import_id": import_id, "updated_at": now}},
                upsert=True)
            if month not in synced_months:
                synced_months.append(month)

    log = {
        "import_id": import_id, "staging_id": staging_id,
        "company_id": company_id, "filename": stg.get("filename") or "",
        "imported_by": admin["user_id"],
        "imported_by_name": admin.get("name") or admin.get("email") or "",
        "at": now, "date_from": dmin, "date_to": dmax,
        "total_rows": (stg.get("stats") or {}).get("total") or len(rows),
        "imported": imported, "skipped": skipped,
        "failed": (stg.get("stats") or {}).get("invalid") or 0,
        "punches_created": punches_created,
        "duplicate_mode": dup_mode,
        "sync_compliance": sync_compliance,
        "synced_months": sorted(synced_months),
        "errors": stg.get("invalid") or [],
        "remarks": str(payload.get("remarks") or ""),
        "status": "imported",
    }
    await db.client_attendance_imports.insert_one({**log})
    await db.client_import_staging.update_one(
        {"staging_id": staging_id},
        {"$set": {"consumed": True, "rows": [], "invalid": [],
                  "file_b64": ""}})
    log.pop("errors", None)
    return {"ok": True, "log": log}


# ---------------------------------------------------------------------------
# Import log / error report / rollback
# ---------------------------------------------------------------------------
@router.get("/logs")
async def import_logs(company_id: str = Query(...),
                      authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    items = await db.client_attendance_imports.find(
        {"company_id": company_id}, {"_id": 0, "errors": 0}).sort(
        "at", -1).to_list(200)
    return {"logs": items}


@router.get("/logs/{import_id}/errors")
async def error_report(import_id: str,
                       authorization: Optional[str] = Header(None)):
    log = await db.client_attendance_imports.find_one(
        {"import_id": import_id}, {"_id": 0})
    if not log:
        raise HTTPException(status_code=404, detail="Import not found")
    await _auth(authorization, log["company_id"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    ws.append(["Row", "Employee Code", "Name", "Date", "Error"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="B91C1C")
    for e in log.get("errors") or []:
        ws.append([e.get("row"), e.get("code"), e.get("name"),
                   e.get("date"), e.get("reason")])
    for i, w in enumerate([8, 16, 26, 12, 48], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return {"filename": f"import_errors_{import_id}.xlsx",
            "file_base64": base64.b64encode(buf.getvalue()).decode()}


@router.delete("/logs/{import_id}")
async def delete_import(import_id: str,
                        authorization: Optional[str] = Header(None)):
    """Full rollback of one import: its day records + its punches. The
    biometric / other-source punches are never touched."""
    log = await db.client_attendance_imports.find_one(
        {"import_id": import_id}, {"_id": 0})
    if not log:
        raise HTTPException(status_code=404, detail="Import not found")
    admin = await _auth(authorization, log["company_id"])
    d1 = await db.client_attendance_days.delete_many({"import_id": import_id})
    d2 = await db.attendance.delete_many({"import_batch_id": import_id,
                                          "source": "client_import"})
    await db.compliance_import_entries.delete_many(
        {"import_id": import_id, "source": "client_attendance_import"})
    await db.client_attendance_imports.update_one(
        {"import_id": import_id},
        {"$set": {"status": "deleted", "deleted_at": now_iso(),
                  "deleted_by": admin["user_id"],
                  "deleted_days": d1.deleted_count,
                  "deleted_punches": d2.deleted_count}})
    return {"ok": True, "deleted_days": d1.deleted_count,
            "deleted_punches": d2.deleted_count}


# ---------------------------------------------------------------------------
# Imported day records (viewer)
# ---------------------------------------------------------------------------
@router.get("/days")
async def imported_days(company_id: str = Query(...),
                        month: Optional[str] = Query(None),
                        import_id: Optional[str] = Query(None),
                        authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    q: Dict[str, Any] = {"company_id": company_id}
    if import_id:
        q["import_id"] = import_id
    if month:
        q["date"] = {"$gte": f"{month}-01", "$lte": f"{month}-31"}
    items = await db.client_attendance_days.find(q, {"_id": 0}).sort(
        [("date", 1), ("employee_code", 1)]).to_list(5000)
    return {"days": items}


# ---------------------------------------------------------------------------
# Mapping templates
# ---------------------------------------------------------------------------
@router.get("/templates")
async def list_templates(company_id: str = Query(...),
                         authorization: Optional[str] = Header(None)):
    await _auth(authorization, company_id)
    items = await db.client_import_templates.find(
        {"company_id": company_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"templates": items}


@router.post("/templates")
async def save_template(payload: Dict[str, Any] = Body(...),
                        authorization: Optional[str] = Header(None)):
    company_id = str(payload.get("company_id") or "")
    admin = await _auth(authorization, company_id)
    name = str(payload.get("name") or "").strip()
    mapping = payload.get("mapping") or {}
    headers = payload.get("headers") or []
    if not name or not mapping:
        raise HTTPException(status_code=400, detail="name and mapping required")
    doc = {"template_id": f"ctpl_{uuid.uuid4().hex[:10]}",
           "company_id": company_id, "name": name,
           # store header NAMES so the template survives column re-ordering
           "mapping_headers": {f: str(headers[i]) if i < len(headers) else ""
                               for f, i in mapping.items()},
           "mapping": mapping,
           "created_by": admin["user_id"], "created_at": now_iso()}
    await db.client_import_templates.insert_one({**doc})
    return {"ok": True, "template": doc}


@router.delete("/templates/{template_id}")
async def delete_template(template_id: str,
                          authorization: Optional[str] = Header(None)):
    t = await db.client_import_templates.find_one(
        {"template_id": template_id}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    await _auth(authorization, t["company_id"])
    await db.client_import_templates.delete_one({"template_id": template_id})
    return {"ok": True}
