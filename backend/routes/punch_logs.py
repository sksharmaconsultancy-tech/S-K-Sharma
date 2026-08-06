"""Iter 145 — Punch Log Report (Utility).

Full audit trail of every biometric / app / manual punch, filterable by
date range, machine (device serial or source category) and firm.

Endpoints
---------
GET /api/admin/punch-logs         → JSON rows + machine list for filters
GET /api/admin/punch-logs.xlsx    → Excel download (same filters)
"""
from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException, Query
from fastapi.responses import StreamingResponse

from server import (  # noqa: E402
    db,
    get_user_from_token,
    sub_admin_can_touch_company,
)

router = APIRouter(prefix="/api/admin", tags=["punch-logs"])

MAX_JSON_ROWS = 2000
MAX_XLSX_ROWS = 100000


def _source_label(rec: dict) -> str:
    """Human machine/source label for a punch record."""
    serial = str(rec.get("device_serial") or "")
    src = str(rec.get("source") or "")
    if serial.startswith("import:") or src.startswith("import:") or src.startswith("zk_dat"):
        return "Import (.dat/.TXT)"
    if serial:
        return f"Device {serial}"
    if src.startswith("zkteco:"):
        return f"Device {src.split(':', 1)[1]}"
    if src == "manual_admin":
        return "Manual (Admin)"
    if src == "roster":
        return "Roster"
    return "Mobile App"


def _machine_key(rec: dict) -> str:
    serial = str(rec.get("device_serial") or "")
    src = str(rec.get("source") or "")
    if serial.startswith("import:") or src.startswith("import:") or src.startswith("zk_dat"):
        return "import"
    if serial:
        return f"device:{serial}"
    if src.startswith("zkteco:"):
        return f"device:{src.split(':', 1)[1]}"
    if src == "manual_admin":
        return "manual_admin"
    return "app"


async def _query_rows(
    admin: dict,
    company_id: Optional[str],
    machine: Optional[str],
    from_date: Optional[str],
    to_date: Optional[str],
    limit: int,
) -> Dict[str, Any]:
    q: Dict[str, Any] = {"kind": {"$in": ["in", "out"]}}
    if company_id:
        if not sub_admin_can_touch_company(admin, company_id):
            raise HTTPException(403, "No access to this firm")
        q["company_id"] = company_id
    elif (admin.get("role") == "sub_admin"
          and (admin.get("sub_admin_company_scope") or "all") != "all"):
        q["company_id"] = {"$in": admin.get("sub_admin_company_ids") or []}
    date_q: Dict[str, str] = {}
    if from_date:
        date_q["$gte"] = from_date
    if to_date:
        date_q["$lte"] = to_date
    if date_q:
        q["date"] = date_q

    recs = await db.attendance.find(
        q,
        {"_id": 0, "record_id": 1, "user_id": 1, "company_id": 1, "date": 1,
         "at": 1, "kind": 1, "source": 1, "device_serial": 1, "status": 1,
         "branch_name": 1},
    ).sort([("at", -1)]).to_list(MAX_XLSX_ROWS)

    # Machine filter applied post-query (source label is derived).
    if machine:
        recs = [r for r in recs if _machine_key(r) == machine]

    # Resolve employee + firm names.
    uids = list({r.get("user_id") for r in recs if r.get("user_id")})
    users: Dict[str, dict] = {}
    if uids:
        async for u in db.users.find(
            {"user_id": {"$in": uids}},
            {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1, "bio_code": 1,
             # Iter 341 — "NEW REGISTRATION" flag needs the creation date.
             "created_at": 1},
        ):
            users[u["user_id"]] = u
    cids = list({r.get("company_id") for r in recs if r.get("company_id")})
    firms: Dict[str, str] = {}
    if cids:
        async for c in db.companies.find(
            {"company_id": {"$in": cids}}, {"_id": 0, "company_id": 1, "name": 1},
        ):
            firms[c["company_id"]] = c.get("name") or c["company_id"]

    rows: List[Dict[str, Any]] = []
    machines: Dict[str, str] = {}
    # Iter 419 (user request) — two extra columns:
    #   • Machine Name  — the friendly device name from Device Setup.
    #   • Name In Machine — the employee name AS STORED ON THE MACHINE
    #     (harvested via machine sync into biometric_machine_users).
    _dev_map: Dict[str, dict] = {}
    async for d in db.biometric_devices.find(
        {}, {"_id": 0, "serial_number": 1, "company_id": 1, "name": 1, "location": 1},
    ):
        _dev_map[str(d.get("serial_number") or "")] = d
    _mu_map: Dict[tuple, str] = {}
    async for mu in db.biometric_machine_users.find(
        {}, {"_id": 0, "company_id": 1, "pin": 1, "name": 1},
    ):
        _p = str(mu.get("pin") or "").strip()
        _mu_map[(mu.get("company_id"), _p)] = mu.get("name") or ""
        # Iter 494 — some machines zero-pad the PIN ("050" vs bio "50"):
        # index the normalized form too so the name always resolves.
        if _p.lstrip("0") and _p.lstrip("0") != _p:
            _mu_map.setdefault((mu.get("company_id"), _p.lstrip("0")), mu.get("name") or "")

    def _serial_of(rec: dict) -> str:
        s = str(rec.get("device_serial") or "")
        if s and not s.startswith("import:"):
            return s
        src = str(rec.get("source") or "")
        if src.startswith("zkteco:"):
            return src.split(":", 1)[1]
        return ""

    # Iter 250 — which records carry a punch photo (machine ATTPHOTO or
    # mobile selfie)? Cheap id-only lookup, photos themselves stay out of
    # the JSON payload.
    photo_ids = set(await db.attendance.distinct(
        "record_id", {**q, "selfie_base64": {"$exists": True, "$nin": [None, ""]}}))
    # Iter 419 (user request) — flag OT punches: the system convention is
    # that a 2nd IN→OUT pair on the same day is the OT session (same rule
    # as the Day-wise Present/OT count). Everything from the 2nd IN of the
    # day onwards is marked "OT PUNCH".
    _by_day: Dict[tuple, List[dict]] = {}
    for r in recs:
        _by_day.setdefault((r.get("user_id"), r.get("date")), []).append(r)
    _ot_threshold: Dict[tuple, str] = {}
    for key, plist in _by_day.items():
        ins = sorted(str(p.get("at") or "") for p in plist if p.get("kind") == "in")
        if len(ins) >= 2:
            _ot_threshold[key] = ins[1]  # 2nd IN of the day

    def _is_ot(rec: dict) -> bool:
        th = _ot_threshold.get((rec.get("user_id"), rec.get("date")))
        return bool(th) and str(rec.get("at") or "") >= th

    # Iter 341 (user request) — flags: "not_found" when the punch's user is
    # missing from the Employee Master, "new_registration" when the
    # employee was registered TODAY.
    _today_ist = datetime.now(
        timezone(timedelta(hours=5, minutes=30))).strftime("%Y-%m-%d")
    for r in recs:
        u = users.get(r.get("user_id") or "", {})
        at = str(r.get("at") or "")
        mkey = _machine_key(r)
        mlabel = _source_label(r)
        machines.setdefault(mkey, mlabel)
        _flag = ""
        if not u:
            _flag = "not_found"
        elif str(u.get("created_at") or "")[:10] == _today_ist:
            _flag = "new_registration"
        _serial = _serial_of(r)
        _pin = str(u.get("bio_code") or "").strip() or (r.get("user_id") or "" if not u else "")
        rows.append({
            "record_id": r.get("record_id"),
            # Iter 494 — employee photo support (None when NOT FOUND).
            "user_id": r.get("user_id") if u else None,
            "date": r.get("date") or at[:10],
            "time": at[11:19] if len(at) >= 19 else at[11:16],
            "kind": r.get("kind"),
            "ot": _is_ot(r),
            "employee_code": u.get("employee_code") or "",
            "name": u.get("name") or ("NOT FOUND" if not u else r.get("user_id") or ""),
            "name_in_machine": (_mu_map.get((r.get("company_id"), _pin))
                                or _mu_map.get((r.get("company_id"), _pin.lstrip("0")))
                                or ""),
            "bio_code": u.get("bio_code") or ("" if u else (r.get("user_id") or "")),
            "machine_name": (_dev_map.get(_serial) or {}).get("name") or "",
            "machine": mlabel,
            "machine_key": mkey,
            "company_name": firms.get(r.get("company_id") or "", ""),
            "status": "not found" if _flag == "not_found" else (r.get("status") or ""),
            "source": r.get("source") or "",
            "has_photo": r.get("record_id") in photo_ids,
            "photo_ref": (r.get("record_id")
                          if r.get("record_id") in photo_ids else None),
            "flag": _flag,
        })
    # Iter 341 (user request) — device punches whose BIO CODE matched NO
    # employee in the Master (stored in biometric_unmapped) appear in the
    # report marked "NOT FOUND".
    _allowed: Optional[set] = None
    if isinstance(q.get("company_id"), str):
        _allowed = {q["company_id"]}
    elif isinstance(q.get("company_id"), dict):
        _allowed = set(q["company_id"].get("$in") or [])
    _unm_q: Dict[str, Any] = {}
    if from_date:
        _unm_q.setdefault("at", {})["$gte"] = f"{from_date}T00:00:00"
    if to_date:
        _unm_q.setdefault("at", {})["$lte"] = f"{to_date}T23:59:59.999999"
    _dev_cids = {d.get("company_id") for d in _dev_map.values()
                 if d.get("company_id") and d.get("company_id") not in firms}
    if _dev_cids:
        async for c in db.companies.find(
            {"company_id": {"$in": list(_dev_cids)}},
            {"_id": 0, "company_id": 1, "name": 1},
        ):
            firms[c["company_id"]] = c.get("name") or c["company_id"]
    # Iter 503 (user bug + request) — parked machine photos (ATTPHOTO that
    # matched no employee) so a NOT-FOUND punch can still show its photo.
    _park: Dict[tuple, List[str]] = {}
    async for p_ in db.biometric_photos.find(
            _unm_q, {"_id": 0, "device_serial": 1, "device_user_id": 1,
                     "at": 1}).limit(4000):
        _park.setdefault((str(p_.get("device_serial") or ""),
                          str(p_.get("device_user_id") or "")),
                         []).append(str(p_.get("at") or ""))

    def _parked_photo_at(serial: str, pin: str, at: str) -> Optional[str]:
        for pat in _park.get((serial, pin), []):
            try:
                d1 = datetime.fromisoformat(pat.replace("Z", "+00:00"))
                d2 = datetime.fromisoformat(at.replace("Z", "+00:00"))
                if abs((d1 - d2).total_seconds()) <= 120:
                    return pat
            except ValueError:
                continue
        return None

    async for m_ in db.biometric_unmapped.find(
        _unm_q, {"_id": 0, "device_serial": 1, "device_user_id": 1, "at": 1},
    ).sort([("at", -1)]).limit(2000):
        d = _dev_map.get(str(m_.get("device_serial") or ""), {})
        _cid = d.get("company_id")
        # Iter 503 (user bug: "NOT FOUND not working on some machines") —
        # punches from a machine that is NOT registered in the Devices
        # master have no firm, so the firm filter silently dropped them.
        # They now ALWAYS show, marked "Unregistered Device".
        if _allowed is not None and _cid is not None and _cid not in _allowed:
            continue
        _serial = str(m_.get("device_serial") or "")
        mkey = f"device:{_serial}" if _serial else "app"
        if machine and mkey != machine:
            continue
        mlabel = f"Device {_serial}" if _serial else "Device"
        machines.setdefault(mkey, mlabel)
        at = str(m_.get("at") or "")
        _pin = str(m_.get("device_user_id") or "")
        _ph_at = _parked_photo_at(_serial, _pin, at)
        rows.append({
            "record_id": None,
            "date": at[:10],
            "time": at[11:19] if len(at) >= 19 else at[11:16],
            "kind": "",
            "ot": False,
            "employee_code": "",
            "name": "NOT FOUND IN MASTER",
            "name_in_machine": _mu_map.get(
                (_cid, _pin.strip()), ""),
            "bio_code": _pin,
            "machine_name": d.get("name") or "",
            "machine": mlabel,
            "machine_key": mkey,
            "company_name": firms.get(_cid or "", "")
            or (d.get("name") or "")
            or "⚠ Unregistered Device",
            "status": "not found",
            "source": "device",
            "has_photo": bool(_ph_at),
            "photo_ref": (f"unmapped|{_serial}|{_pin}|{_ph_at}"
                          if _ph_at else None),
            "flag": "not_found",
        })
    rows.sort(key=lambda x: (x.get("date") or "", x.get("time") or ""), reverse=True)
    return {"rows": rows, "machines": machines}


@router.get("/punch-logs")
async def punch_logs(
    company_id: Optional[str] = Query(None),
    machine: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    if admin.get("role") not in ("super_admin", "sub_admin", "company_admin"):
        raise HTTPException(403, "Admin only")
    if admin.get("role") == "company_admin":
        company_id = admin.get("company_id")
    data = await _query_rows(admin, company_id, machine, from_date, to_date, MAX_JSON_ROWS)
    rows = data["rows"]
    return {
        "total": len(rows),
        "truncated": len(rows) > MAX_JSON_ROWS,
        "rows": rows[:MAX_JSON_ROWS],
        "machines": [
            {"key": k, "label": v} for k, v in sorted(data["machines"].items())
        ],
    }


@router.get("/punch-logs/photo")
async def punch_log_photo(
    ref: str = Query(..., description="record_id or unmapped|serial|pin|at"),
    authorization: Optional[str] = Header(None),
):
    """Iter 503 (user request) — view the machine photo of a punch row,
    INCLUDING 'NOT FOUND IN MASTER' rows (parked ATTPHOTO of an unmapped
    device user)."""
    admin = await get_user_from_token(authorization)
    if admin.get("role") not in ("super_admin", "sub_admin", "company_admin"):
        raise HTTPException(403, "Admin only")
    if ref.startswith("unmapped|"):
        try:
            _, serial, pin, at = ref.split("|", 3)
        except ValueError:
            raise HTTPException(400, "Bad photo reference")
        p = await db.biometric_photos.find_one(
            {"device_serial": serial, "device_user_id": pin, "at": at},
            {"_id": 0, "photo_base64": 1, "at": 1})
        if not p or not p.get("photo_base64"):
            raise HTTPException(404, "Photo not found")
        return {"photo_base64": p["photo_base64"], "at": p.get("at"),
                "caption": f"Machine user {pin} · {serial}"}
    rec = await db.attendance.find_one(
        {"record_id": ref},
        {"_id": 0, "selfie_base64": 1, "at": 1, "user_id": 1, "company_id": 1})
    if not rec or not rec.get("selfie_base64"):
        raise HTTPException(404, "Photo not found")
    if admin.get("role") == "company_admin" \
            and rec.get("company_id") != admin.get("company_id"):
        raise HTTPException(403, "Not your firm's punch")
    if admin.get("role") == "sub_admin" \
            and not sub_admin_can_touch_company(admin, rec.get("company_id")):
        raise HTTPException(403, "Firm not in your scope")
    u = await db.users.find_one({"user_id": rec.get("user_id")},
                                {"_id": 0, "name": 1, "employee_code": 1})
    return {"photo_base64": rec["selfie_base64"], "at": rec.get("at"),
            "caption": f"{(u or {}).get('name') or ''} ({(u or {}).get('employee_code') or ''})"}


@router.get("/daily-attendance")
async def daily_attendance(
    date: str = Query(..., description="YYYY-MM-DD"),
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Iter 148 — Date-wise attendance summary, firm-wise (employer PWA
    dashboard). One row per employee with all their punches for the day,
    first-IN / last-OUT, worked hours and Present/Absent status."""
    admin = await get_user_from_token(authorization)
    if admin.get("role") not in ("super_admin", "sub_admin", "company_admin"):
        raise HTTPException(403, "Admin only")
    if admin.get("role") == "company_admin":
        company_id = admin.get("company_id")

    uq: Dict[str, Any] = {"role": "employee", "approval_status": "approved"}
    if company_id:
        if not sub_admin_can_touch_company(admin, company_id):
            raise HTTPException(403, "No access to this firm")
        uq["company_id"] = company_id
    elif (admin.get("role") == "sub_admin"
          and (admin.get("sub_admin_company_scope") or "all") != "all"):
        uq["company_id"] = {"$in": admin.get("sub_admin_company_ids") or []}

    employees = await db.users.find(
        uq, {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1,
             "bio_code": 1, "company_id": 1},
    ).to_list(3000)

    aq: Dict[str, Any] = {"date": date, "kind": {"$in": ["in", "out"]}}
    if company_id:
        aq["company_id"] = company_id
    recs = await db.attendance.find(
        aq, {"_id": 0, "user_id": 1, "at": 1, "kind": 1,
             "source": 1, "device_serial": 1},
    ).sort([("at", 1)]).to_list(50000)

    def _is_machine(src: str) -> bool:
        s = (src or "").lower()
        return s.startswith("zkteco") or s.startswith("import:") or "machine" in s

    by_user: Dict[str, List[dict]] = {}
    for r in recs:
        by_user.setdefault(r.get("user_id") or "", []).append(r)

    cids = list({e.get("company_id") for e in employees if e.get("company_id")})
    firms: Dict[str, str] = {}
    if cids:
        async for c in db.companies.find(
            {"company_id": {"$in": cids}}, {"_id": 0, "company_id": 1, "name": 1},
        ):
            firms[c["company_id"]] = c.get("name") or c["company_id"]

    def _mins(at: str) -> Optional[int]:
        # Wall-clock convention: read HH:MM verbatim from the ISO string.
        try:
            return int(at[11:13]) * 60 + int(at[14:16])
        except Exception:
            return None

    rows: List[Dict[str, Any]] = []
    present = 0
    for e in employees:
        precs = by_user.get(e["user_id"], [])
        punches = [{
            "time": str(p.get("at") or "")[11:16],
            "kind": p.get("kind"),
            "machine": _source_label(p),
        } for p in precs]
        first_in = next((p for p in precs if p.get("kind") == "in"), None)
        last_out = next((p for p in reversed(precs) if p.get("kind") == "out"), None)
        # Worked minutes: sum of IN→next-OUT pairs.
        worked = 0
        open_in: Optional[int] = None
        for p in precs:
            m = _mins(str(p.get("at") or ""))
            if m is None:
                continue
            if p.get("kind") == "in":
                open_in = m
            elif p.get("kind") == "out" and open_in is not None:
                if m >= open_in:
                    worked += m - open_in
                open_in = None
        status = "present" if precs else "absent"
        if precs:
            present += 1
        # Iter 155 — expected punching mode vs how they actually punched.
        expected = "biometric" if (e.get("bio_code") or "").strip() else "manual"
        machine_cnt = sum(1 for p in precs if _is_machine(p.get("source") or ""))
        if not precs:
            actual = None
        elif machine_cnt == len(precs):
            actual = "machine"
        elif machine_cnt == 0:
            actual = "app"
        else:
            actual = "mixed"
        mode_mismatch = expected == "biometric" and actual in ("app", "mixed")
        rows.append({
            "user_id": e["user_id"],
            "name": e.get("name") or "",
            "employee_code": e.get("employee_code") or "",
            "company_id": e.get("company_id"),
            "company_name": firms.get(e.get("company_id") or "", ""),
            "status": status,
            "punches": punches,
            "first_in": str(first_in.get("at"))[11:16] if first_in else None,
            "last_out": str(last_out.get("at"))[11:16] if last_out else None,
            "worked_hrs": round(worked / 60, 2) if worked else 0,
            "still_in": bool(precs) and precs[-1].get("kind") == "in",
            "mode_expected": expected,
            "mode_actual": actual,
            "mode_mismatch": mode_mismatch,
        })

    rows.sort(key=lambda r: (r["status"] != "present",
                             r.get("first_in") or "99:99", r["name"].lower()))
    return {
        "date": date,
        "total": len(rows),
        "present": present,
        "absent": len(rows) - present,
        "mismatch": sum(1 for r in rows if r.get("mode_mismatch")),
        "rows": rows,
    }


@router.get("/attendance-report/day-counts")
async def attendance_day_counts(
    month: str = Query(..., description="YYYY-MM"),
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Iter 154 — day-wise present count for a month (1–31), including how
    many employees worked OT (2nd IN→OUT pair) each day. Tapping a count in
    the UI deep-links to /daily-attendance for the full employee list."""
    admin = await get_user_from_token(authorization)
    if admin.get("role") not in ("super_admin", "sub_admin", "company_admin"):
        raise HTTPException(403, "Admin only")
    if admin.get("role") == "company_admin":
        company_id = admin.get("company_id")
    import re as _re
    if not _re.fullmatch(r"\d{4}-\d{2}", month or ""):
        raise HTTPException(400, "month must be YYYY-MM")

    aq: Dict[str, Any] = {
        "date": {"$gte": f"{month}-01", "$lte": f"{month}-31"},
        "kind": {"$in": ["in", "out"]},
        "status": {"$ne": "rejected"},
    }
    if company_id:
        if not sub_admin_can_touch_company(admin, company_id):
            raise HTTPException(403, "No access to this firm")
        aq["company_id"] = company_id
    elif (admin.get("role") == "sub_admin"
          and (admin.get("sub_admin_company_scope") or "all") != "all"):
        aq["company_id"] = {"$in": admin.get("sub_admin_company_ids") or []}

    recs = await db.attendance.find(
        aq, {"_id": 0, "date": 1, "user_id": 1, "kind": 1},
    ).to_list(200000)

    # per day: distinct present users + users with ≥2 IN punches (OT pair).
    per_day: Dict[str, Dict[str, Any]] = {}
    for r in recs:
        d = per_day.setdefault(r["date"], {"users": set(), "ins": {}})
        d["users"].add(r["user_id"])
        if r["kind"] == "in":
            d["ins"][r["user_id"]] = d["ins"].get(r["user_id"], 0) + 1

    from calendar import monthrange
    y, m = int(month[:4]), int(month[5:7])
    ndays = monthrange(y, m)[1]
    days = []
    tot_present = tot_ot = 0
    for i in range(1, ndays + 1):
        date = f"{month}-{i:02d}"
        d = per_day.get(date)
        present = len(d["users"]) if d else 0
        ot = sum(1 for c in (d["ins"].values() if d else []) if c >= 2)
        tot_present += present
        tot_ot += ot
        days.append({"date": date, "present": present, "ot_count": ot})
    return {"month": month, "days": days,
            "total_present_mandays": tot_present, "total_ot_mandays": tot_ot}


@router.get("/punch-logs.xlsx")
async def punch_logs_xlsx(
    company_id: Optional[str] = Query(None),
    machine: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    if admin.get("role") not in ("super_admin", "sub_admin", "company_admin"):
        raise HTTPException(403, "Admin only")
    if admin.get("role") == "company_admin":
        company_id = admin.get("company_id")
    data = await _query_rows(admin, company_id, machine, from_date, to_date, MAX_XLSX_ROWS)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Punch Log"
    headers = ["Sr", "Date", "Time", "IN/OUT",
               # Iter 419 — OT punches (2nd IN→OUT pair of the day) marked.
               "OT",
               "Emp Code", "Employee Name",
               # Iter 419 — employee name AS STORED ON THE MACHINE + device name.
               "Name In Machine", "Bio Code", "Machine Name",
               "Machine / Source", "Firm", "Status", "Photo",
               # Iter 341 — NOT FOUND / NEW REGISTRATION marker.
               "Remark"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E79")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill

    # Iter 250 (user request) — punches WITH PHOTO: embed the actual punch
    # photo (machine ATTPHOTO / mobile selfie) as a thumbnail in the row.
    MAX_EMBEDDED_PHOTOS = 2000
    _photo_col = headers.index("Photo") + 1
    photo_rows: List[tuple] = []  # (excel_row, record_id)
    for i, r in enumerate(data["rows"], start=1):
        _rm = ("NOT FOUND" if r.get("flag") == "not_found"
               else "NEW REGISTRATION" if r.get("flag") == "new_registration" else "")
        ws.append([i, r["date"], r["time"], (r["kind"] or "").upper(),
                   "OT PUNCH" if r.get("ot") else "",
                   r["employee_code"], r["name"], r.get("name_in_machine") or "",
                   r["bio_code"], r.get("machine_name") or "",
                   r["machine"], r["company_name"], r["status"],
                   "" if r.get("has_photo") else "—", _rm])
        if r.get("ot"):
            _oc = ws.cell(row=i + 1, column=5)
            _oc.font = Font(bold=True, color="B45309")
        if _rm:
            _c = ws.cell(row=i + 1, column=len(headers))
            _c.font = Font(bold=True,
                           color="B91C1C" if _rm == "NOT FOUND" else "15803D")
        if r.get("has_photo") and len(photo_rows) < MAX_EMBEDDED_PHOTOS:
            photo_rows.append((i + 1, r.get("record_id")))
    if photo_rows:
        import base64 as _b64

        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        for xl_row, rid in photo_rows:
            try:
                rec = await db.attendance.find_one(
                    {"record_id": rid}, {"_id": 0, "selfie_base64": 1})
                b64 = (rec or {}).get("selfie_base64") or ""
                if b64.startswith("data:"):
                    b64 = b64.split(",", 1)[-1]
                img_bytes = _b64.b64decode(b64)
                pim = PILImage.open(BytesIO(img_bytes))
                pim.thumbnail((72, 54))
                out = BytesIO()
                pim.convert("RGB").save(out, format="PNG")
                out.seek(0)
                xli = XLImage(out)
                ws.add_image(xli, f"{get_column_letter(_photo_col)}{xl_row}")
                ws.row_dimensions[xl_row].height = 45
            except Exception:
                ws.cell(row=xl_row, column=_photo_col, value="YES")
    widths = [6, 12, 10, 8, 11, 10, 26, 22, 9, 18, 22, 24, 10, 12, 16]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Punch_Log_{from_date or 'all'}_{to_date or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# Iter 277 (user request) — "Rectified Punches" audit: which duplicate /
# rapid scans were auto-ignored per employee per day, and why.
# ---------------------------------------------------------------------------
@router.get("/attendance/rectified-punches/{company_id}/{month}")
async def rectified_punches(
    company_id: str,
    month: str,
    authorization: Optional[str] = Header(None),
):
    from server import (
        dedupe_rapid_punches,
        dedupe_same_machine_punches,
        merge_out_in_bounces,
    )

    admin = await get_user_from_token(authorization)
    if admin.get("role") not in ("super_admin", "sub_admin", "company_admin"):
        raise HTTPException(403, "Admin only")
    if admin.get("role") == "company_admin":
        company_id = admin.get("company_id")
    if admin.get("role") == "sub_admin" and not sub_admin_can_touch_company(admin, company_id):
        raise HTTPException(403, "Not authorised for this firm")

    emp_map: Dict[str, Dict[str, Any]] = {}
    async for u in db.users.find(
        {"company_id": company_id, "role": "employee"},
        {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1},
    ):
        emp_map[u["user_id"]] = u

    by_user_day: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}
    async for r in db.attendance.find(
        {"user_id": {"$in": list(emp_map.keys())},
         "date": {"$gte": f"{month}-01", "$lte": f"{month}-31"},
         "status": "approved"},
        {"_id": 0, "user_id": 1, "date": 1, "kind": 1, "at": 1, "source": 1},
    ).sort([("user_id", 1), ("at", 1)]):
        by_user_day.setdefault(r["user_id"], {}).setdefault(r["date"], []).append(r)

    def _ist(at: Any) -> str:
        try:
            dt = datetime.fromisoformat(str(at).replace("Z", "+00:00"))
            return (dt + timedelta(hours=5, minutes=30)).strftime("%H:%M:%S")
        except Exception:
            return str(at or "")

    def _stage(punches, fn, args, reason, dropped):
        kept = fn(punches, *args)
        kept_ids = {id(p) for p in kept}
        for p in punches:
            if id(p) not in kept_ids:
                dropped.append({
                    "time": _ist(p.get("at")),
                    "kind": (p.get("kind") or "").upper(),
                    "source": p.get("source") or "",
                    "reason": reason,
                })
        return kept

    rows: List[Dict[str, Any]] = []
    total_dropped = 0
    for uid, days in by_user_day.items():
        emp = emp_map.get(uid) or {}
        for date_s in sorted(days.keys()):
            raw = days[date_s]
            dropped: List[Dict[str, Any]] = []
            kept = _stage(raw, dedupe_rapid_punches, (30,),
                          "Double-scan within 30 seconds", dropped)
            kept = _stage(kept, dedupe_same_machine_punches, (15,),
                          "Duplicate same-direction punch within 15 min", dropped)
            kept = _stage(kept, merge_out_in_bounces, (60,),
                          "OUT-IN bounce within 60 seconds (device stutter)", dropped)
            if not dropped:
                continue
            total_dropped += len(dropped)
            rows.append({
                "user_id": uid,
                "name": emp.get("name") or "",
                "employee_code": emp.get("employee_code") or "",
                "date": date_s,
                "raw_count": len(raw),
                "kept_count": len(kept),
                "kept": [{"time": _ist(p.get("at")),
                          "kind": (p.get("kind") or "").upper()} for p in kept],
                "dropped": dropped,
            })
    rows.sort(key=lambda r: (r["date"], r["name"]))
    return {"month": month, "days_affected": len(rows),
            "punches_ignored": total_dropped, "rows": rows}
