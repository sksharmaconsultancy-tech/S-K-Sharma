"""Iter 392 (user spec) — ATTENDANCE SYNCHRONIZATION DASHBOARD.

Single-page reconciliation between the Employee Master, the Biometric
Machines and the Attendance data. READ-ONLY analytics — no existing
module or collection is modified. Auto-detection rules:

  Rule 1  in Machine, not in Master        → machine_only  (biometric_unmapped)
  Rule 2  in Master, not in Machine        → master_only
  Rule 3  in both, punches stopped         → attendance_missing
  Rule 4  DOJ within the selected range    → new_joining
  Rule 5  machine-registered, never punched→ never_punched
  Rule 6  punch stopped N+ days            → inactive_attendance

"AI Smart Analysis" remarks are deterministic rule-based texts (no LLM
cost) as per the user's examples.
"""
import difflib
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException, Query

from server import db, get_user_from_token, require_role

router = APIRouter(prefix="/api")

_ZK_PREFIX = "zkteco"


def _today() -> datetime:
    # Attendance dates are stored as device-local wall clock labelled UTC —
    # use IST calendar date to match the punches.
    return datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)


def _iso_date(d: datetime) -> str:
    return d.strftime("%Y-%m-%d")


def _range_from_preset(preset: str, date_from: Optional[str],
                       date_to: Optional[str]):
    t = _today()
    today = _iso_date(t)
    if preset == "today":
        return today, today
    if preset == "yesterday":
        y = _iso_date(t - timedelta(days=1))
        return y, y
    if preset == "week":
        start = t - timedelta(days=t.weekday())
        return _iso_date(start), today
    if preset == "custom" and date_from and date_to:
        return date_from[:10], date_to[:10]
    # default: this month
    return today[:8] + "01", today


def _doj_iso(u: Dict[str, Any]) -> str:
    """Normalise DOJ (may be DD-MM-YYYY legacy or ISO) to YYYY-MM-DD."""
    raw = str(u.get("doj") or "")[:10]
    if len(raw) == 10 and raw[2] == "-" and raw[5] == "-":
        return f"{raw[6:10]}-{raw[3:5]}-{raw[0:2]}"
    return raw


def _is_active(u: Dict[str, Any], today: str) -> bool:
    if u.get("disabled"):
        return False
    ex = str(u.get("exit_date") or "")[:10]
    return not (ex and ex <= today)


@router.get("/admin/attendance-sync-dashboard")
async def attendance_sync_dashboard(
    company_id: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    designation: Optional[str] = Query(None),
    machine: Optional[str] = Query(None, description="Device serial filter"),
    status: str = Query("active", pattern="^(active|left|all)$"),
    preset: str = Query("month", pattern="^(today|yesterday|week|month|custom)$"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    missing_days: int = Query(3, ge=1, le=90),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if admin["role"] == "company_admin":
        company_id = admin.get("company_id")  # company-wise access enforced

    d_from, d_to = _range_from_preset(preset, date_from, date_to)
    today = _iso_date(_today())

    # ---------------- Employee Master ----------------
    uq: Dict[str, Any] = {"role": "employee"}
    if company_id:
        uq["company_id"] = company_id
    if department:
        uq["department"] = department
    if designation:
        uq["designation"] = designation
    users = await db.users.find(uq, {
        "_id": 0, "user_id": 1, "employee_code": 1, "bio_code": 1, "name": 1,
        "company_id": 1, "department": 1, "designation": 1, "doj": 1,
        "disabled": 1, "exit_date": 1, "is_onroll": 1, "phone": 1,
    }).to_list(120000)
    if status == "active":
        users = [u for u in users if _is_active(u, today)]
    elif status == "left":
        users = [u for u in users if not _is_active(u, today)]
    uid_list = [u["user_id"] for u in users]
    by_uid = {u["user_id"]: u for u in users}

    # Company names for labels
    comp_names = {c["company_id"]: c.get("name") for c in await db.companies.find(
        {}, {"_id": 0, "company_id": 1, "name": 1}).to_list(500)}

    # ---------------- Punch aggregates (optimised) ----------------
    # last/first punch + zkteco flag per employee — single aggregation.
    aq_match: Dict[str, Any] = {"user_id": {"$in": uid_list}, "status": {"$ne": "rejected"}}
    if machine:
        aq_match["device_serial"] = machine
    punch_stats: Dict[str, Dict[str, Any]] = {}
    async for a in db.attendance.aggregate([
        {"$match": aq_match},
        {"$group": {
            "_id": "$user_id",
            "first_punch": {"$min": "$date"},
            "last_punch": {"$max": "$date"},
            "zk": {"$max": {"$cond": [
                {"$regexMatch": {"input": {"$ifNull": ["$source", ""]},
                                 "regex": f"^{_ZK_PREFIX}"}}, 1, 0]}},
        }},
    ], allowDiskUse=True):
        punch_stats[a["_id"]] = a

    # Punches within the selected window (attendance % + missing check)
    in_range: Dict[str, int] = {}
    async for a in db.attendance.aggregate([
        {"$match": {**aq_match, "date": {"$gte": d_from, "$lte": d_to}}},
        {"$group": {"_id": "$user_id", "days": {"$addToSet": "$date"}}},
        {"$project": {"n": {"$size": "$days"}}},
    ], allowDiskUse=True):
        in_range[a["_id"]] = a["n"]

    # Approved leave overlapping "now" (Section 4/5 exclusions)
    on_leave: Dict[str, str] = {}
    try:
        lq: Dict[str, Any] = {"status": "approved", "user_id": {"$in": uid_list},
                              "to_date": {"$gte": _iso_date(_today() - timedelta(days=45))}}
        async for lv in db.leaves.find(lq, {"_id": 0, "user_id": 1, "from_date": 1,
                                            "to_date": 1, "leave_type": 1}).limit(20000):
            on_leave[lv["user_id"]] = str(lv.get("leave_type") or "Leave")
    except Exception:
        pass

    def _machine_registered(u: Dict[str, Any]) -> bool:
        return bool(str(u.get("bio_code") or "").strip()) or bool(
            (punch_stats.get(u["user_id"]) or {}).get("zk"))

    # ---------------- Section 1 — New Joining ----------------
    new_joining: List[Dict[str, Any]] = []
    for u in users:
        doj = _doj_iso(u)
        if not (doj and d_from <= doj <= d_to):
            continue
        ps = punch_stats.get(u["user_id"]) or {}
        mreg = _machine_registered(u)
        if mreg and ps.get("first_punch"):
            state, color = "Complete", "green"
            remark = "Registered and punching — payroll ready."
        elif mreg:
            state, color = "Never Punched", "orange"
            remark = "Employee joined but never attended — check enrolment/shift."
        else:
            state, color = "Machine Pending", "orange"
            remark = "Machine registration pending."
        new_joining.append({
            "user_id": u["user_id"], "employee_code": u.get("employee_code"),
            "name": u.get("name"), "doj": doj,
            "company": comp_names.get(u.get("company_id")),
            "department": u.get("department"), "designation": u.get("designation"),
            "machine_registered": mreg, "master_available": True,
            "first_punch": ps.get("first_punch"),
            "status": state, "color": color, "remark": remark,
        })
    new_joining.sort(key=lambda r: r["doj"] or "", reverse=True)

    # ---------------- Section 2 — Machine only (Rule 1) ----------------
    codes = [str(u.get("employee_code") or "") for u in users if u.get("employee_code")]
    umq: Dict[str, Any] = {}
    if machine:
        umq["device_serial"] = machine
    machine_only: List[Dict[str, Any]] = []
    async for m in db.biometric_unmapped.aggregate([
        {"$match": umq},
        {"$group": {"_id": {"pin": "$device_user_id", "sn": "$device_serial"},
                    "first": {"$min": "$at"}, "last": {"$max": "$at"},
                    "count": {"$sum": 1}}},
        {"$sort": {"last": -1}}, {"$limit": 300},
    ]):
        pin = str(m["_id"]["pin"] or "")
        close = difflib.get_close_matches(pin, codes, n=1, cutoff=0.7)
        dev = await db.biometric_devices.find_one(
            {"serial_number": m["_id"]["sn"]}, {"_id": 0, "name": 1, "company_id": 1})
        machine_only.append({
            "machine_id": pin, "machine": m["_id"]["sn"],
            "machine_name": (dev or {}).get("name"),
            "company": comp_names.get((dev or {}).get("company_id")),
            "first_punch": str(m["first"])[:16], "last_punch": str(m["last"])[:16],
            "punch_count": m["count"],
            "suggested_match": close[0] if close else None,
            "remark": ("Possible duplicate machine registration — code close to "
                       f"employee {close[0]}." if close
                       else "Employee probably not enrolled in payroll — create the Employee Master."),
        })

    # ---------------- Section 3 — Master only (Rule 2) ----------------
    master_only: List[Dict[str, Any]] = []
    for u in users:
        if _machine_registered(u) or not _is_active(u, today):
            continue
        doj = _doj_iso(u)
        days_since = 0
        try:
            days_since = (datetime.strptime(today, "%Y-%m-%d")
                          - datetime.strptime(doj, "%Y-%m-%d")).days if doj else 0
        except ValueError:
            pass
        master_only.append({
            "user_id": u["user_id"], "employee_code": u.get("employee_code"),
            "name": u.get("name"), "department": u.get("department"),
            "doj": doj, "days_since_joining": days_since,
            "machine_status": "Never Registered",
            "remark": "Machine registration pending — enrol on the biometric device.",
        })
    master_only.sort(key=lambda r: -r["days_since_joining"])

    # -------- Section 4 + 5 — Attendance missing / continuous absence ----
    attendance_missing: List[Dict[str, Any]] = []
    continuous_buckets = {"3": 0, "5": 0, "7": 0, "15": 0, "30": 0}
    never_punched = 0
    t_dt = datetime.strptime(today, "%Y-%m-%d")
    for u in users:
        if not _is_active(u, today):
            continue
        if not _machine_registered(u):
            continue
        ps = punch_stats.get(u["user_id"]) or {}
        last = ps.get("last_punch")
        if not last:
            never_punched += 1
            days_missing = 9999
        else:
            try:
                days_missing = (t_dt - datetime.strptime(last, "%Y-%m-%d")).days
            except ValueError:
                continue
        if days_missing < missing_days:
            continue
        leave = on_leave.get(u["user_id"])
        if days_missing >= 30:
            color = "red"
        elif days_missing >= 7:
            color = "red"
        elif days_missing >= 5:
            color = "orange"
        else:
            color = "orange" if not leave else "green"
        if leave:
            remark = f"On approved {leave} — no action needed."
        elif not last:
            remark = "Employee joined but never attended — verify enrolment and shift."
        elif days_missing >= 15:
            remark = "Employee not punching regularly — possible unrecorded exit; notify HR."
        elif days_missing >= 7:
            remark = "Attendance missing for a week — check machine sync or employee absence."
        else:
            remark = "Attendance missing — verify machine synchronisation or leave."
        for b in ("3", "5", "7", "15", "30"):
            if last and days_missing >= int(b):
                continuous_buckets[b] += 1
        attendance_missing.append({
            "user_id": u["user_id"], "employee_code": u.get("employee_code"),
            "name": u.get("name"), "department": u.get("department"),
            "last_punch": last, "days_missing": days_missing if last else None,
            "never_punched": not last,
            "leave_status": leave, "color": color, "remark": remark,
            "status": "Never Punched" if not last else "Attendance Missing",
        })
    attendance_missing.sort(key=lambda r: -(r["days_missing"] or 99999))

    # ---------------- Machines (Rule 6 — sync issues) ----------------
    dq: Dict[str, Any] = {}
    if company_id:
        dq["company_id"] = company_id
    machines: List[Dict[str, Any]] = []
    now_utc = datetime.now(timezone.utc)
    sync_ok = 0
    for d in await db.biometric_devices.find(dq, {
            "_id": 0, "serial_number": 1, "name": 1, "company_id": 1,
            "last_seen_at": 1, "enabled": 1,
            # Iter 512 — Direct SDK pull channel fields
            "connection_mode": 1, "sdk_vendor": 1, "sdk_last_pull_at": 1,
            "sdk_last_error": 1, "auto_pull_minutes": 1}).to_list(200):
        seen = d.get("last_seen_at")
        online = False
        secs = None
        if seen:
            try:
                secs = (now_utc - datetime.fromisoformat(
                    str(seen).replace("Z", "+00:00"))).total_seconds()
                online = secs < 180
            except ValueError:
                pass
        is_sdk = (d.get("connection_mode") == "sdk")
        if is_sdk:
            # SDK-pull machines are contacted BY the server — "healthy" means
            # the last pull/test worked (they don't hold a live connection).
            online = bool(d.get("sdk_last_pull_at")) and not d.get("sdk_last_error")
        if online:
            sync_ok += 1
        if is_sdk:
            _remark = (
                d.get("sdk_last_error")
                or (None if d.get("sdk_last_pull_at")
                    else "SDK pull device — never pulled yet. Use 'Pull punches' on the device card."))
        else:
            _remark = None if online else (
                "Biometric device may not be synchronised — check power/network."
                if seen else "Attendance synchronisation failed — device never connected.")
        machines.append({
            "serial_number": d.get("serial_number"), "name": d.get("name"),
            "company": comp_names.get(d.get("company_id")),
            "last_seen_at": seen, "online": online,
            # Iter 512 — SDK pull surfacing
            "connection_mode": d.get("connection_mode") or "push",
            "sdk_vendor": d.get("sdk_vendor"),
            "sdk_last_pull_at": d.get("sdk_last_pull_at"),
            "auto_pull_minutes": d.get("auto_pull_minutes") or 0,
            "remark": _remark,
        })

    # ---------------- KPIs / health ----------------
    total = len(users)
    active = sum(1 for u in users if _is_active(u, today))
    machine_reg = sum(1 for u in users if _machine_registered(u))
    att_days_expected = max(1, (datetime.strptime(d_to, "%Y-%m-%d")
                                - datetime.strptime(d_from, "%Y-%m-%d")).days + 1)
    workers = [u for u in users if _is_active(u, today)]
    punched_in_range = sum(1 for u in workers if in_range.get(u["user_id"], 0) > 0)
    att_pct = round(100.0 * sum(min(in_range.get(u["user_id"], 0), att_days_expected)
                                for u in workers)
                    / max(1, len(workers) * att_days_expected), 1)
    machine_sync_pct = round(100.0 * sync_ok / max(1, len(machines)), 1) if machines else 100.0
    master_sync_pct = round(100.0 * machine_reg / max(1, active), 1)
    compliance_pct = round(100.0 * punched_in_range / max(1, len(workers)), 1)
    overall = round((att_pct + machine_sync_pct + master_sync_pct + compliance_pct) / 4.0, 1)
    last_sync = max([str(m.get("last_seen_at") or "") for m in machines] or [""])

    # ---------------- Trend (weekly joins + daily punch %) ----------------
    weekly_joins: List[Dict[str, Any]] = []
    for w in range(7, -1, -1):
        ws = _today() - timedelta(days=_today().weekday() + 7 * w)
        we = ws + timedelta(days=6)
        n = sum(1 for u in users
                if _iso_date(ws) <= _doj_iso(u) <= _iso_date(we))
        weekly_joins.append({"week": _iso_date(ws), "joins": n})
    daily_punch: List[Dict[str, Any]] = []
    day_counts: Dict[str, int] = {}
    async for a in db.attendance.aggregate([
        {"$match": {"user_id": {"$in": uid_list},
                    "date": {"$gte": _iso_date(_today() - timedelta(days=13)),
                             "$lte": today},
                    "status": {"$ne": "rejected"}}},
        {"$group": {"_id": "$date", "u": {"$addToSet": "$user_id"}}},
        {"$project": {"n": {"$size": "$u"}}},
    ], allowDiskUse=True):
        day_counts[a["_id"]] = a["n"]
    for i in range(13, -1, -1):
        dd = _iso_date(_today() - timedelta(days=i))
        daily_punch.append({
            "date": dd,
            "pct": round(100.0 * day_counts.get(dd, 0) / max(1, len(workers)), 1),
        })

    return {
        "range": {"from": d_from, "to": d_to, "preset": preset},
        "kpis": {
            "total_employees": total,
            "master_employees": total,
            "machine_registered": machine_reg,
            "active_employees": active,
            "new_joining": len(new_joining),
            "machine_pending": len(master_only),
            "master_pending": len(machine_only),
            "attendance_missing": len(attendance_missing),
            "never_punched": never_punched,
            "attendance_pct": att_pct,
            "machine_sync_pct": machine_sync_pct,
            "master_sync_pct": master_sync_pct,
            "compliance_pct": compliance_pct,
            "overall_health": overall,
            "last_sync_at": last_sync or None,
            "machines_total": len(machines),
            "machines_online": sync_ok,
        },
        "new_joining": new_joining[:500],
        "machine_only": machine_only,
        "master_only": master_only[:500],
        "attendance_missing": attendance_missing[:500],
        "continuous_absence": continuous_buckets,
        "machines": machines,
        "trend": {"weekly_joins": weekly_joins, "daily_punch_pct": daily_punch},
        "generated_at": now_utc.replace(microsecond=0).isoformat(),
    }


@router.get("/admin/attendance-sync-dashboard/export")
async def attendance_sync_export(
    section: str = Query("attendance_missing",
                         pattern="^(new_joining|machine_only|master_only|attendance_missing|full)$"),
    format: str = Query("xlsx", pattern="^(xlsx|pdf|csv)$"),
    company_id: Optional[str] = Query(None),
    preset: str = Query("month", pattern="^(today|yesterday|week|month|custom)$"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    missing_days: int = Query(3, ge=1, le=90),
    authorization: Optional[str] = Header(None),
):
    """Excel / PDF / CSV export of any dashboard section."""
    data = await attendance_sync_dashboard(
        company_id=company_id, department=None, designation=None, machine=None,
        status="active", preset=preset, date_from=date_from, date_to=date_to,
        missing_days=missing_days, authorization=authorization)
    cfg = {
        "new_joining": ("New Joining Report",
                        ["Code", "Name", "DOJ", "Company", "Department",
                         "Machine Registered", "First Punch", "Status"],
                        [[r.get("employee_code"), r.get("name"), r.get("doj"),
                          r.get("company"), r.get("department"),
                          "Yes" if r.get("machine_registered") else "No",
                          r.get("first_punch") or "—", r.get("status")]
                         for r in data["new_joining"]]),
        "machine_only": ("Registered in Machine but NOT in Master",
                         ["Machine ID", "Machine", "First Punch", "Last Punch",
                          "Punches", "Suggested Match", "Remark"],
                         [[r.get("machine_id"), r.get("machine_name") or r.get("machine"),
                           r.get("first_punch"), r.get("last_punch"),
                           r.get("punch_count"), r.get("suggested_match") or "—",
                           r.get("remark")] for r in data["machine_only"]]),
        "master_only": ("Registered in Master but NOT in Machine",
                        ["Code", "Name", "Department", "DOJ",
                         "Days Since Joining", "Machine Status", "Remark"],
                        [[r.get("employee_code"), r.get("name"), r.get("department"),
                          r.get("doj"), r.get("days_since_joining"),
                          r.get("machine_status"), r.get("remark")]
                         for r in data["master_only"]]),
        "attendance_missing": ("Attendance Missing Report",
                               ["Code", "Name", "Department", "Last Punch",
                                "Days Missing", "Leave", "Status", "Remark"],
                               [[r.get("employee_code"), r.get("name"),
                                 r.get("department"), r.get("last_punch") or "Never",
                                 r.get("days_missing") if r.get("days_missing") is not None else "—",
                                 r.get("leave_status") or "—", r.get("status"),
                                 r.get("remark")] for r in data["attendance_missing"]]),
    }
    title, headers, rows = cfg[section] if section != "full" else (None, None, None)
    from routes.compliance_validation import _file_response, _pdf_bytes, _xlsx_bytes

    # ---- Iter 393 (user request) — FULL report: every section in ONE
    # single sheet / one PDF document. ----
    if section == "full":
        k = data["kpis"]
        kpi_pairs = [
            ("Total Employees", k["total_employees"]),
            ("Machine Registered", k["machine_registered"]),
            ("Active Employees", k["active_employees"]),
            ("New Joining", k["new_joining"]),
            ("Master Pending (machine-only users)", k["master_pending"]),
            ("Machine Pending (not enrolled)", k["machine_pending"]),
            ("Attendance Missing", k["attendance_missing"]),
            ("Never Punched", k["never_punched"]),
            ("Attendance %", f"{k['attendance_pct']}%"),
            ("Machine Sync %", f"{k['machine_sync_pct']}%"),
            ("Master Sync %", f"{k['master_sync_pct']}%"),
            ("Overall Health Score", f"{k['overall_health']}%"),
            ("Machines Online", f"{k['machines_online']}/{k['machines_total']}"),
            ("Last Synchronization", str(k.get("last_sync_at") or "—")[:16]),
        ]
        cont = data["continuous_absence"]
        sections = [
            ("SUMMARY (KPI)", ["Indicator", "Value"],
             [[a, b] for a, b in kpi_pairs]),
            *[(cfg[s][0].upper(), cfg[s][1], cfg[s][2])
              for s in ("new_joining", "machine_only", "master_only",
                        "attendance_missing")],
            ("CONTINUOUS ABSENCE", ["Consecutive Days", "Employees"],
             [[f"{d}+ days", n] for d, n in cont.items()]),
            ("MACHINE SYNCHRONIZATION", ["Machine", "Serial", "Status", "Last Seen", "Remark"],
             [[m.get("name") or "—", m.get("serial_number"),
               "ONLINE" if m.get("online") else "OFFLINE",
               str(m.get("last_seen_at") or "never")[:16],
               m.get("remark") or "OK"] for m in data["machines"]]),
        ]
        fname = f"attendance_sync_full_{data['range']['from']}_{data['range']['to']}.{format}"
        sub = (f"{data['range']['from']} → {data['range']['to']} · "
               f"Overall Health {k['overall_health']}%")
        if format == "xlsx":
            # SINGLE SHEET: all sections stacked with styled section bands.
            import io as _io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Sync Report"
            ws.append(["ATTENDANCE SYNCHRONIZATION DASHBOARD"])
            ws.cell(row=1, column=1).font = Font(bold=True, size=13)
            ws.append([sub])
            maxw: Dict[int, int] = {}
            for title2, hdrs, rows2 in sections:
                ws.append([])
                ws.append([title2])
                c = ws.cell(row=ws.max_row, column=1)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1D4ED8")
                ws.append(hdrs)
                for cc in ws[ws.max_row]:
                    cc.font = Font(bold=True, color="FFFFFF")
                    cc.fill = PatternFill("solid", fgColor="0F172A")
                for r in rows2:
                    ws.append(r)
                for i2, h in enumerate(hdrs, start=1):
                    w = max([len(str(h))] + [len(str(r[i2 - 1])) for r in rows2[:150]] or [8])
                    maxw[i2] = max(maxw.get(i2, 8), min(44, w + 2))
            for i2, w in maxw.items():
                ws.column_dimensions[ws.cell(row=1, column=i2).column_letter].width = w
            buf = _io.BytesIO()
            wb.save(buf)
            return _file_response(buf.getvalue(), fname, "xlsx")
        if format == "pdf":
            # One PDF with all section tables stacked.
            import io as _io
            from reportlab.lib import colors as rl
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import (Paragraph, SimpleDocTemplate,
                                            Spacer, Table, TableStyle)
            styles = getSampleStyleSheet()
            cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=6.8, leading=8.2)
            headp = ParagraphStyle("head", parent=cell, textColor=rl.white,
                                   fontName="Helvetica-Bold")
            secp = ParagraphStyle("sec", parent=styles["Normal"], fontSize=10,
                                  fontName="Helvetica-Bold", spaceBefore=8,
                                  textColor=rl.HexColor("#1D4ED8"))
            story = [Paragraph("Attendance Synchronization Dashboard", styles["Title"]),
                     Paragraph(sub, styles["Normal"]), Spacer(1, 3 * mm)]
            for title2, hdrs, rows2 in sections:
                story.append(Paragraph(title2, secp))
                tdata = [[Paragraph(str(h), headp) for h in hdrs]]
                for r in rows2[:400]:
                    tdata.append([Paragraph(str(v if v is not None else ""), cell) for v in r])
                t = Table(tdata, repeatRows=1)
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#0F172A")),
                    ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [rl.white, rl.HexColor("#F8FAFC")]),
                ]))
                story.append(t)
            buf = _io.BytesIO()
            SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=8 * mm,
                              rightMargin=8 * mm, topMargin=10 * mm,
                              bottomMargin=10 * mm).build(story)
            return _file_response(buf.getvalue(), fname, "pdf")
        # CSV: sections stacked with blank separators.
        import csv as _csv
        import io as _io
        buf2 = _io.StringIO()
        w = _csv.writer(buf2)
        w.writerow(["ATTENDANCE SYNCHRONIZATION DASHBOARD"])
        w.writerow([sub])
        for title2, hdrs, rows2 in sections:
            w.writerow([])
            w.writerow([title2])
            w.writerow(hdrs)
            w.writerows(rows2)
        from fastapi import Response
        return Response(content=buf2.getvalue().encode("utf-8-sig"),
                        media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    fname = f"{section}_{data['range']['from']}_{data['range']['to']}.{format}"
    if format == "csv":
        import csv as _csv
        import io as _io
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow(headers)
        w.writerows(rows)
        from fastapi import Response
        return Response(content=buf.getvalue().encode("utf-8-sig"),
                        media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    if format == "xlsx":
        return _file_response(_xlsx_bytes(title, headers, rows), fname, "xlsx")
    sub = f"{data['range']['from']} → {data['range']['to']} · {len(rows)} record(s)"
    return _file_response(_pdf_bytes(title, sub, headers, rows), fname, "pdf")
