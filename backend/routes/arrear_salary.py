"""Iter 102 — Salary Process (Arrear).

Revised-wage arrears: when an employee's compliance salary (Govt salary)
is revised with back-effect, this module computes — for each past month
in the selected range — the difference between what WAS paid (the stored
compliance salary run rows) and what SHOULD have been paid at the current
(revised) master salary, using the SAME present days.

PF / ESIC on arrears follow the EPFO Arrear ECR help file:
  * ARREAR_EPF_WAGES  = new PF wages − old PF wages (capped)
  * EPF due (EE)      = 12%  of arrear EPF wages
  * EPS due           = 8.33% of arrear EPS wages
  * ER due (EPF-EPS)  = 12% − EPS  (3.67% diff)
  * ESIC EE/ER        = 0.75% / 3.25% of arrear gross (when applicable)

ECR text export format (8 fields, `#~#` separated):
  UAN#~#MEMBER_NAME#~#ARREAR_EPF_WAGES#~#ARREAR_EPS_WAGES#~#ARREAR_EDLI_WAGES
  #~#ARREAR_EPF_CONTRIBUTION_DUE#~#ARREAR_EPS_CONTRIBUTION_DUE#~#ARREAR_ER_CONTRIBUTION_DUE
"""
import io
import uuid
from typing import Optional

from fastapi import APIRouter, Body, Header, HTTPException
from fastapi.responses import Response

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
    require_employer_permission,
    now_iso,
    sub_admin_can_touch_company,
)

router = APIRouter(prefix="/api", tags=["arrear-salary"])


def _months_between(from_month: str, to_month: str) -> list:
    """Inclusive list of YYYY-MM strings from..to (max 24)."""
    try:
        fy, fm = int(from_month[:4]), int(from_month[5:7])
        ty, tm = int(to_month[:4]), int(to_month[5:7])
        if not (1 <= fm <= 12 and 1 <= tm <= 12):
            raise ValueError
    except (ValueError, TypeError, IndexError):
        raise HTTPException(status_code=400, detail="Months must be YYYY-MM")
    out = []
    y, m = fy, fm
    while (y, m) <= (ty, tm):
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
        if len(out) > 24:
            raise HTTPException(status_code=400, detail="Range too large (max 24 months)")
    if not out:
        raise HTTPException(status_code=400, detail="From month must be before To month")
    return out


async def _check_admin(authorization, company_id):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if admin["role"] == "company_admin":
        if admin.get("company_id") != company_id:
            raise HTTPException(status_code=403, detail="You can only process your own firm")
    elif admin["role"] == "sub_admin":
        if not sub_admin_can_touch_company(admin, company_id):
            raise HTTPException(status_code=403, detail="Firm not in your scope")
    return admin


@router.post("/admin/arrear-salary-runs")
async def create_arrear_run(
    payload: dict = Body(...),
    authorization: Optional[str] = Header(None),
):
    from utils.compliance_salary import compute_compliance_row
    company_id = payload.get("company_id")
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    admin = await _check_admin(authorization, company_id)
    # Employer Access Rights — company_admin needs the firm-level
    # "salary_process:write" grant to process arrear salary runs.
    await require_employer_permission(admin, "salary_process:write", db)
    months = _months_between(payload.get("from_month") or "", payload.get("to_month") or "")

    company = await db.companies.find_one(
        {"company_id": company_id},
        {"_id": 0, "name": 1, "attendance_policy": 1, "compliance_policy": 1},
    )
    if not company:
        raise HTTPException(status_code=404, detail="Firm not found")
    att_pol = company.get("attendance_policy") or {}

    # Current Firm Master EPF/ESI applicability
    fm = await db.firm_masters.find_one(
        {"company_id": company_id}, {"_id": 0, "epf": 1, "esi": 1})
    firm_pf = bool(((fm or {}).get("epf") or {}).get("applicable"))
    firm_esic = bool(((fm or {}).get("esi") or {}).get("applicable"))

    # Current (revised) employee master
    employees: dict = {}
    async for u in db.users.find(
        {"role": "employee", "company_id": company_id}, {"_id": 0}):
        u.pop("pin_hash", None)
        u.pop("password_hash", None)
        employees[u["user_id"]] = u

    per_emp: dict = {}
    months_used = []
    months_skipped = []
    for month in months:
        # OLD = latest compliance run for that month (prefer finalized)
        run = await db.compliance_salary_runs.find_one(
            {"company_id": company_id, "month": month, "finalized": True},
            {"_id": 0}, sort=[("generated_at", -1)],
        ) or await db.compliance_salary_runs.find_one(
            {"company_id": company_id, "month": month},
            {"_id": 0}, sort=[("generated_at", -1)],
        )
        if not run or not run.get("rows"):
            months_skipped.append(month)
            continue
        months_used.append(month)
        month_days = int(run.get("month_days") or 26)
        structure_pct = run.get("structure_pct") or None
        statutory_cfg = run.get("statutory_cfg") or None

        for old in run["rows"]:
            uid = old.get("user_id")
            emp = employees.get(uid)
            if not emp:
                continue
            pol = {**att_pol, **(emp.get("employee_policy") or {})}
            half = int(old.get("half_days") or 0)
            # Reconstruct effective_present EXACTLY from what was paid:
            #   gross_paid = per_day_rate × effective + ot_pay
            old_rate = float(old.get("rate") or 0)
            mode = (old.get("salary_mode") or "monthly").lower()
            base_paid = float(old.get("gross_paid") or 0) - float(old.get("ot_pay") or 0)
            per_day_old = (old_rate / month_days) if mode == "monthly" else old_rate
            if per_day_old > 0:
                eff = max(0.0, round(base_paid / per_day_old, 4))
            else:
                eff = float(old.get("present_days") or 0)
            stats = {
                "present_days": int(old.get("present_days") or 0),
                "half_days": half,
                "effective_present": eff,
                "duty_hours": float(old.get("duty_hours") or 0),
                "ot_hours": float(old.get("ot_hours") or 0),
            }
            new = compute_compliance_row(
                emp, pol, month_days, stats,
                company_structure_pct=structure_pct,
                statutory_cfg=statutory_cfg,
                firm_pf_enabled=firm_pf,
                firm_esic_enabled=firm_esic,
            )
            old_gross = float(old.get("gross_paid") or 0)
            new_gross = float(new.get("gross_paid") or 0)
            arrear_gross = max(0.0, round(new_gross - old_gross, 2))
            old_pfw = float(old.get("pf_wages") or 0)
            new_pfw = float(new.get("pf_wages") or 0)
            arrear_pfw = max(0.0, round(new_pfw - old_pfw, 2)) if firm_pf and new.get("pf_applicable") else 0.0
            if arrear_gross <= 0 and arrear_pfw <= 0:
                continue
            epf_due = round(arrear_pfw * 0.12)
            eps_due = round(arrear_pfw * 0.0833)
            er_due = epf_due - eps_due
            # ESIC on arrears only when the revised salary actually
            # attracts ESIC (≤ ceiling → row has a contribution).
            esic_on = firm_esic and float(new.get("esic_employee") or 0) > 0
            esic_ee = round(arrear_gross * 0.0075) if esic_on else 0
            esic_er = round(arrear_gross * 0.0325) if esic_on else 0

            agg = per_emp.setdefault(uid, {
                "user_id": uid,
                "employee_code": emp.get("employee_code"),
                "name": emp.get("name"),
                "uan_no": emp.get("uan_no"),
                "esic_no": emp.get("esi_ip_no"),
                "months": [],
                "old_gross": 0.0, "new_gross": 0.0, "arrear_gross": 0.0,
                "arrear_epf_wages": 0.0, "arrear_eps_wages": 0.0,
                "arrear_edli_wages": 0.0,
                "epf_due": 0, "eps_due": 0, "er_due": 0,
                "esic_employee": 0, "esic_employer": 0,
            })
            agg["months"].append({
                "month": month,
                "present_days": stats["present_days"],
                "old_gross": round(old_gross, 2),
                "new_gross": round(new_gross, 2),
                "arrear_gross": arrear_gross,
                "arrear_epf_wages": arrear_pfw,
                "epf_due": epf_due, "eps_due": eps_due, "er_due": er_due,
                "esic_employee": esic_ee, "esic_employer": esic_er,
            })
            agg["old_gross"] = round(agg["old_gross"] + old_gross, 2)
            agg["new_gross"] = round(agg["new_gross"] + new_gross, 2)
            agg["arrear_gross"] = round(agg["arrear_gross"] + arrear_gross, 2)
            agg["arrear_epf_wages"] = round(agg["arrear_epf_wages"] + arrear_pfw, 2)
            agg["arrear_eps_wages"] = agg["arrear_epf_wages"]
            agg["arrear_edli_wages"] = agg["arrear_epf_wages"]
            agg["epf_due"] += epf_due
            agg["eps_due"] += eps_due
            agg["er_due"] += er_due
            agg["esic_employee"] += esic_ee
            agg["esic_employer"] += esic_er

    def _code_key(r):
        c = str(r.get("employee_code") or "").strip()
        try:
            return (0, float(c), "")
        except ValueError:
            return (1, 0.0, c.lower())

    rows = sorted(per_emp.values(), key=_code_key)
    totals = {
        "arrear_gross": round(sum(r["arrear_gross"] for r in rows), 2),
        "arrear_epf_wages": round(sum(r["arrear_epf_wages"] for r in rows), 2),
        "epf_due": sum(r["epf_due"] for r in rows),
        "eps_due": sum(r["eps_due"] for r in rows),
        "er_due": sum(r["er_due"] for r in rows),
        "esic_employee": sum(r["esic_employee"] for r in rows),
        "esic_employer": sum(r["esic_employer"] for r in rows),
    }
    doc = {
        "run_id": f"arr_{uuid.uuid4().hex[:12]}",
        "company_id": company_id,
        "company_name": company.get("name"),
        "from_month": months[0],
        "to_month": months[-1],
        "months_used": months_used,
        "months_skipped": months_skipped,
        "employees_count": len(rows),
        "rows": rows,
        "totals": totals,
        "generated_by": admin["user_id"],
        "generated_at": now_iso(),
    }
    await db.arrear_salary_runs.insert_one(dict(doc))
    return {"run": doc}


@router.get("/admin/arrear-salary-runs")
async def list_arrear_runs(
    company_id: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    await require_employer_permission(admin, "salary_process:read", db)
    q: dict = {}
    if admin["role"] == "company_admin":
        q["company_id"] = admin.get("company_id")
    elif company_id:
        q["company_id"] = company_id
    runs = await db.arrear_salary_runs.find(
        q, {"_id": 0, "rows": 0},
    ).sort("generated_at", -1).to_list(100)
    return {"runs": runs}


@router.get("/admin/arrear-salary-runs/{run_id}")
async def get_arrear_run(run_id: str, authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    await require_employer_permission(admin, "salary_process:read", db)
    run = await db.arrear_salary_runs.find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Arrear run not found")
    if admin["role"] == "company_admin" and admin.get("company_id") != run["company_id"]:
        raise HTTPException(status_code=403, detail="Not your firm")
    return {"run": run}


@router.delete("/admin/arrear-salary-runs/{run_id}")
async def delete_arrear_run(run_id: str, authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin"])
    r = await db.arrear_salary_runs.delete_one({"run_id": run_id})
    if not r.deleted_count:
        raise HTTPException(status_code=404, detail="Arrear run not found")
    return {"ok": True}


@router.get("/admin/arrear-salary-runs/{run_id}/ecr.txt")
async def arrear_ecr_txt(run_id: str, authorization: Optional[str] = Header(None)):
    """EPFO Arrear ECR text file — 8 fields, `#~#` separated."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    await require_employer_permission(admin, "salary_process:read", db)
    run = await db.arrear_salary_runs.find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Arrear run not found")
    if admin["role"] == "company_admin" and admin.get("company_id") != run["company_id"]:
        raise HTTPException(status_code=403, detail="Not your firm")
    lines = []
    for r in run.get("rows", []):
        if not r.get("uan_no") or float(r.get("arrear_epf_wages") or 0) <= 0:
            continue
        name = str(r.get("name") or "").upper().replace("#~#", " ")
        lines.append("#~#".join([
            str(r["uan_no"]).strip(),
            name,
            str(int(round(float(r.get("arrear_epf_wages") or 0)))),
            str(int(round(float(r.get("arrear_eps_wages") or 0)))),
            str(int(round(float(r.get("arrear_edli_wages") or 0)))),
            str(int(r.get("epf_due") or 0)),
            str(int(r.get("eps_due") or 0)),
            str(int(r.get("er_due") or 0)),
        ]))
    body = "\n".join(lines)
    fname = f"arrear_ecr_{run['from_month']}_{run['to_month']}.txt"
    return Response(
        content=body.encode("utf-8"),
        media_type="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/admin/arrear-salary-runs/{run_id}/export.xlsx")
async def arrear_export_xlsx(run_id: str, authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    await require_employer_permission(admin, "salary_process:read", db)
    run = await db.arrear_salary_runs.find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Arrear run not found")
    if admin["role"] == "company_admin" and admin.get("company_id") != run["company_id"]:
        raise HTTPException(status_code=403, detail="Not your firm")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Arrear Register"
    ws.append([f"ARREAR REGISTER — {run.get('company_name') or run['company_id']}"])
    ws.append([f"Period: {run['from_month']} to {run['to_month']}  ·  Generated: {run.get('generated_at', '')[:10]}"])
    ws.append([])
    hdr = ["SN", "Emp Code", "Name", "UAN", "ESIC No",
           "Old Gross (Paid)", "Revised Gross", "Arrear Gross",
           "Arrear EPF Wages", "Arrear EPS Wages", "Arrear EDLI Wages",
           "EPF Due (EE)", "EPS Due", "ER Due (3.67)",
           "ESIC EE", "ESIC ER", "Months"]
    ws.append(hdr)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F2E3D")
    for i, r in enumerate(run.get("rows", []), start=1):
        ws.append([
            i, r.get("employee_code"), r.get("name"), r.get("uan_no"), r.get("esic_no"),
            r.get("old_gross"), r.get("new_gross"), r.get("arrear_gross"),
            r.get("arrear_epf_wages"), r.get("arrear_eps_wages"), r.get("arrear_edli_wages"),
            r.get("epf_due"), r.get("eps_due"), r.get("er_due"),
            r.get("esic_employee"), r.get("esic_employer"),
            ", ".join(m["month"] for m in (r.get("months") or [])),
        ])
    t = run.get("totals") or {}
    ws.append([])
    ws.append(["", "", "TOTAL", "", "",
               "", "", t.get("arrear_gross"),
               t.get("arrear_epf_wages"), t.get("arrear_epf_wages"), t.get("arrear_epf_wages"),
               t.get("epf_due"), t.get("eps_due"), t.get("er_due"),
               t.get("esic_employee"), t.get("esic_employer"), ""])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    widths = [5, 10, 24, 15, 13, 14, 14, 13, 15, 15, 15, 12, 12, 12, 10, 10, 24]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + j) if j <= 26 else "A" + chr(64 + j - 26)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    fname = f"arrear_register_{run['from_month']}_{run['to_month']}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
