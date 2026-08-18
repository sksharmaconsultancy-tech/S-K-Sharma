"""ZKTeco biometric device integration (ADMS / iClock push protocol).

Two-device topology chosen by the client:
  • Device A (Serial X) — installed at the entry gate; every punch is IN
  • Device B (Serial Y) — installed at the exit gate; every punch is OUT
We identify the device by the `SN` query-string parameter that ZKTeco firmware
always sends. Punches are auto-approved (skip the mobile approval queue) and
linked to the app user via the pre-existing `bio_code` field on the User
master. Legacy field: `employee_code` is used as a fallback.

iClock endpoints — the ZKTeco firmware calls these paths verbatim. They are
mounted under /api/iclock/ so they follow the standard ingress rule that
maps /api/* to the backend. When deploying, configure the device with:
  Comm → Cloud Server → Server URL: https://<your-host>/api
  (Some firmwares split into Server Address + URL Path — use both fields.)
"""
import asyncio
import base64
import difflib
import logging
import random
import re
import secrets
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Literal, Optional, Tuple

from fastapi import APIRouter, Body, Header, HTTPException, Query, Request
from fastapi.responses import PlainTextResponse
from pydantic import BaseModel

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
)

router = APIRouter(prefix="/api", tags=["biometric-devices"])
logger = logging.getLogger("biometric-devices")


class BiometricDeviceCreate(BaseModel):
    serial_number: str
    name: str
    kind: Literal["in", "out", "both"]
    company_id: Optional[str] = None
    location: Optional[str] = None
    enabled: bool = True
    gmt_offset: Optional[str] = "+05:30"   # Iter 263 — machine time zone
    brand: Optional[str] = None            # Iter 294 — multi-brand (zkteco/essl/matrix/mantra/other)
    branch_name: Optional[str] = None      # Iter 298 — two-branch firms: punches from this machine count as this branch's duty
    # Iter 512 — Direct SDK pull channel (ADDITIVE — push/ADMS unchanged)
    connection_mode: Optional[str] = "push"  # push (ADMS) | sdk (server pulls)
    sdk_vendor: Optional[str] = None         # adapter id from /biometric/sdks
    device_ip: Optional[str] = None          # device IP / DDNS host
    device_port: Optional[int] = None        # vendor default when empty
    comm_key: Optional[str] = None           # device comm key / password
    auto_pull_minutes: Optional[int] = 0     # 0 = manual pulls only


class BiometricDeviceUpdate(BaseModel):
    name: Optional[str] = None
    kind: Optional[Literal["in", "out", "both"]] = None
    company_id: Optional[str] = None
    location: Optional[str] = None
    enabled: Optional[bool] = None
    gmt_offset: Optional[str] = None       # Iter 263 — machine time zone
    brand: Optional[str] = None            # Iter 294 — multi-brand
    branch_name: Optional[str] = None      # Iter 298 — branch tag
    # Iter 512 — Direct SDK pull channel fields
    connection_mode: Optional[str] = None
    sdk_vendor: Optional[str] = None
    device_ip: Optional[str] = None
    device_port: Optional[int] = None
    comm_key: Optional[str] = None
    auto_pull_minutes: Optional[int] = None


# Iter 263 — GMT / time-zone handling for machines.
_GMT_RE = re.compile(r"^([+-]?)(\d{1,2})(?::(\d{2})|\.(\d+))?$")


def _parse_gmt_offset_minutes(raw: Optional[str]) -> int:
    """Parse a GMT offset like '+05:30', '5:30', '-04:00', '+8' or '5.5'
    into signed MINUTES. Defaults to India (+05:30 → 330) when blank or
    invalid."""
    s = str(raw or "").strip().upper().replace("GMT", "").replace("UTC", "").strip()
    m = _GMT_RE.match(s)
    if not m:
        return 330
    sign = -1 if m.group(1) == "-" else 1
    hours = int(m.group(2))
    if m.group(3) is not None:
        mins = int(m.group(3))
    elif m.group(4) is not None:
        mins = int(round(float(f"0.{m.group(4)}") * 60))
    else:
        mins = 0
    total = sign * (hours * 60 + mins)
    if total < -12 * 60 or total > 14 * 60:
        return 330
    return total


def _zk_timezone_value(device: dict) -> str:
    """ZKTeco handshake TimeZone value: whole hours as plain hours
    (e.g. '8'), half/quarter zones as signed MINUTES (e.g. '330' for
    GMT+5:30, '-270' for GMT-4:30) per the Push SDK convention."""
    mins = _parse_gmt_offset_minutes((device or {}).get("gmt_offset"))
    if mins % 60 == 0:
        return str(mins // 60)
    return str(mins)


def _now_iso_z() -> str:
    """UTC ISO timestamp with a trailing Z (some ZKTeco firmwares fussy)."""
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


async def _match_employee_for_bio(
    device_user_id: str, company_id: Optional[str]
) -> Optional[dict]:
    """Look up the app User for a device-reported user id. Matches on
    `bio_code` first (case-insensitive), then falls back to `employee_code`
    so unmapped early rollouts still work. Scoped to the device's company
    when we know it, else global."""
    device_user_id = (device_user_id or "").strip()
    if not device_user_id:
        return None
    def _scope(q: dict) -> dict:
        if company_id:
            q["company_id"] = company_id
        return q
    # Fast path: exact bio_code
    user = await db.users.find_one(_scope({"bio_code": device_user_id}), {"_id": 0})
    if user:
        return user
    # Case-insensitive bio_code (some sites use alphanumeric IDs)
    user = await db.users.find_one(
        _scope({"bio_code": {"$regex": f"^{re.escape(device_user_id)}$", "$options": "i"}}),
        {"_id": 0},
    )
    if user:
        return user
    # Fallback: employee_code (many firms punch the same number into device + app)
    user = await db.users.find_one(_scope({"employee_code": device_user_id}), {"_id": 0})
    return user


def _parse_zk_timestamp(raw: str) -> Optional[datetime]:
    """ZKTeco ATTLOG timestamps arrive as 'YYYY-MM-DD HH:MM:SS' in the
    device's LOCAL clock (IST on-site). The whole attendance pipeline —
    .dat/.TXT imports, the monthly grid and reports — stores and displays
    device-local wall-clock time (labelled UTC, no conversion). Iter 143:
    stop shifting live ADMS pushes by -5:30 so the punch time shown in
    Attendance matches the machine display exactly."""
    raw = (raw or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S"):
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


async def _ingest_attlog_line(
    line: str, device: dict
) -> Tuple[bool, Optional[str]]:
    """Ingest a single ATTLOG line from a ZKTeco push. Returns (ok, reason).
    Format (tab-separated): user_id\ttimestamp\tstatus\tverify_type\tworkcode\treserved
    """
    if not line or not line.strip():
        return False, "empty"
    parts = line.split("\t")
    if len(parts) < 2:
        # Some devices use spaces or ';' — try a permissive split
        parts = re.split(r"\s+", line.strip(), maxsplit=5)
    if len(parts) < 2:
        return False, "malformed"
    device_user_id = parts[0].strip()
    ts_raw = parts[1].strip() if len(parts) > 1 else ""
    verify_type = parts[3].strip() if len(parts) > 3 else ""
    dt = _parse_zk_timestamp(ts_raw)
    if not dt:
        return False, f"bad_timestamp:{ts_raw}"
    user = await _match_employee_for_bio(device_user_id, device.get("company_id"))
    if not user:
        # Log an unmapped punch so admins can create the mapping later —
        # this keeps the audit trail without breaking the ingest loop.
        await db.biometric_unmapped.insert_one({
            "device_serial": device["serial_number"],
            "device_id": device["device_id"],
            "device_user_id": device_user_id,
            "at": dt.isoformat(),
            "raw": line,
            "seen_at": _now_iso_z(),
        })
        return False, f"unmapped_user:{device_user_id}"
    # Iter 481 (user request) — duplicate punches within a configurable
    # window (Iter 583: was hardcoded 5 min, now Attendance Policy →
    # dedup_window_minutes, 0 = off).
    # Iter 488 (user: "Multi Punch Within the Same time") — the guard was
    # scoped to the SAME machine, so the same punch arriving from a second
    # registered device / webhook / re-sync was stored AGAIN at the same
    # time. Now ANY existing punch for the employee within the window
    # (any device, app or manual) marks the new machine punch as a
    # DUPLICATE. Per user rule the raw punch is still STORED in the punch
    # log (never deleted) — it is simply excluded from every calculation.
    _co_dedup = await db.companies.find_one(
        {"company_id": user.get("company_id") or device.get("company_id")},
        {"_id": 0, "attendance_policy.dedup_window_minutes": 1}) or {}
    try:
        _dedup_min = int((_co_dedup.get("attendance_policy") or {}).get("dedup_window_minutes", 5))
    except (TypeError, ValueError):
        _dedup_min = 5
    _dup5 = None
    if _dedup_min > 0:
        _win_lo = (dt - timedelta(minutes=_dedup_min)).isoformat()
        _win_hi = (dt + timedelta(minutes=_dedup_min)).isoformat()
        _dup5 = await db.attendance.find_one({
            "user_id": user["user_id"],
            "at": {"$gte": _win_lo, "$lte": _win_hi},
            "status": {"$in": ["approved", "pending"]},
        }, {"_id": 0, "record_id": 1})
    _is_duplicate = bool(_dup5)
    record_id = f"zk_{uuid.uuid4().hex[:12]}"
    # Iter 486 (user bug — "Still Facing Issue", missing OUT everywhere) —
    # honour the MACHINE'S OWN punch-state (ATTLOG col 3). ZKTeco/BIOFACE
    # devices with IN/OUT function keys send: 0=Check-In, 1=Check-Out,
    # 2=Break-Out, 3=Break-In, 4=OT-In, 5=OT-Out. Previously this column
    # was IGNORED and every punch took the device-level kind (default
    # "in"), so a single-machine site produced IN-only days.
    _state_raw = parts[2].strip() if len(parts) > 2 else ""
    _STATE_KIND = {"0": "in", "1": "out", "2": "out",
                   "3": "in", "4": "in", "5": "out"}
    machine_kind = _STATE_KIND.get(_state_raw)
    # Iter 143 (user spec) — single-machine "Both IN/OUT" mode: the punch
    # direction alternates per employee per day (first punch = IN, next =
    # OUT, then IN again …), based on the latest earlier punch that day.
    punch_kind = device.get("kind", "in")
    if punch_kind == "both":
        if machine_kind:
            # Machine reported a direction → use it directly.
            punch_kind = machine_kind
        else:
            last = await db.attendance.find_one(
                {
                    "user_id": user["user_id"],
                    "date": dt.strftime("%Y-%m-%d"),
                    "kind": {"$in": ["in", "out"]},
                    # Iter 488 — duplicate-marked punches must not flip
                    # the IN/OUT alternation.
                    "status": "approved",
                    "at": {"$lt": dt.isoformat()},
                },
                {"_id": 0, "kind": 1},
                sort=[("at", -1)],
            )
            punch_kind = "out" if (last and last.get("kind") == "in") else "in"
    elif machine_kind and machine_kind != punch_kind and _state_raw != "0":
        # Fixed-direction device BUT the worker pressed an EXPLICIT
        # non-default state key (Check-Out / Break / OT) — trust the
        # machine. State "0" is the ambiguous default on state-less
        # devices, so it never overrides the admin's device config.
        punch_kind = machine_kind
    _direction_corrected = False
    if punch_kind == "in" and device.get("kind") == "in" and _state_raw in ("", "0"):
        # Iter 518 (user choice C — RANJAN case) — SMART DIRECTION
        # CORRECTION, opt-in per firm: when the OUT machine is down,
        # workers punch their evening OUT on the IN machine. A state-less
        # punch landing >= N hours (default 4) after the employee's
        # day-first IN is recorded as the OUT instead.
        try:
            _sdc_comp = await db.companies.find_one(
                {"company_id": device.get("company_id")},
                {"_id": 0, "attendance_config": 1}) or {}
            _sdc_cfg = _sdc_comp.get("attendance_config") or {}
            if _sdc_cfg.get("smart_direction"):
                _gap_h = float(_sdc_cfg.get("smart_direction_gap_hrs") or 4)
                _first_in = await db.attendance.find_one({
                    "user_id": user["user_id"],
                    "date": dt.strftime("%Y-%m-%d"),
                    "kind": "in",
                    "status": "approved",
                    "at": {"$lt": dt.isoformat()},
                }, {"_id": 0, "at": 1}, sort=[("at", 1)])
                if _first_in:
                    _fdt = datetime.fromisoformat(
                        str(_first_in["at"]).replace("Z", "+00:00"))
                    _cur = datetime.fromisoformat(dt.isoformat())
                    if _fdt.tzinfo is None and _cur.tzinfo is not None:
                        _fdt = _fdt.replace(tzinfo=_cur.tzinfo)
                    if _cur.tzinfo is None and _fdt.tzinfo is not None:
                        _cur = _cur.replace(tzinfo=_fdt.tzinfo)
                    if (_cur - _fdt).total_seconds() >= _gap_h * 3600:
                        punch_kind = "out"
                        _direction_corrected = True
        except Exception:
            logger.exception("[smart-direction] check failed for %s", user.get("user_id"))
    # Iter 545 (user spec) — Maximum Punch / Multiple Punch policy for
    # MACHINE punches. Machine data is NEVER dropped: a punch beyond the
    # firm's limit is stored with status="exception" (visible in the punch
    # log + Punch Exception Report) but excluded from every attendance
    # calculation.
    _policy_exception: Optional[str] = None
    _policy_reason: Optional[str] = None
    if not _is_duplicate:
        try:
            from utils.punch_policy import (
                log_punch_exception, resolve_punch_policy)
            _pp_comp = await db.companies.find_one(
                {"company_id": user.get("company_id")},
                {"_id": 0, "attendance_policy.policy_master": 1}) or {}
            _ppol = resolve_punch_policy(user, _pp_comp)
            _pmax = int(_ppol.get("effective_max") or 0)
            if _pmax:
                _pcount = await db.attendance.count_documents({
                    "user_id": user["user_id"],
                    "date": dt.strftime("%Y-%m-%d"),
                    "kind": {"$in": ["in", "out"]},
                    "status": {"$in": ["approved", "pending"]},
                })
                if _pcount >= _pmax:
                    _policy_exception = (
                        "multiple_punch_not_allowed"
                        if not _ppol["multiple_punch_allowed"]
                        else "max_punch_limit")
                    _policy_reason = (
                        f"Maximum {_pmax} punches allowed per day by the "
                        f"Attendance Policy — this machine punch is stored "
                        f"as an EXCEPTION and ignored in attendance "
                        f"calculations.")
                    await log_punch_exception(
                        db, user=user, company_id=user.get("company_id"),
                        date=dt.strftime("%Y-%m-%d"), at=dt.isoformat(),
                        kind=punch_kind, exception_type=_policy_exception,
                        reason=_policy_reason, policy=_ppol,
                        existing_count=_pcount,
                        source=f"zkteco:{device['serial_number']}",
                        device_serial=device["serial_number"])
        except Exception:
            logger.exception("[punch-policy] machine punch check failed")

    record = {
        "record_id": record_id,
        "user_id": user["user_id"],
        "company_id": user.get("company_id"),
        "branch_id": None,
        "branch_name": device.get("location") or device.get("name"),
        "date": dt.strftime("%Y-%m-%d"),
        "kind": punch_kind,
        "direction_corrected": _direction_corrected,  # Iter 518 audit flag
        "at": dt.isoformat(),
        "original_at": dt.isoformat(),
        "latitude": None,
        "longitude": None,
        "distance_m": None,
        "source": f"zkteco:{device['serial_number']}",
        "outside_geofence": False,
        # Machine punches are considered trusted → auto-approved (user chose 4B)
        "status": ("duplicate" if _is_duplicate
                   else ("exception" if _policy_exception else "approved")),
        "exception_type": _policy_exception,
        "decision_by": "system:zkteco",
        "decision_at": _now_iso_z(),
        "decision_reason": (
            "Duplicate punch within 5 min of an existing punch — stored in "
            "the punch log but ignored in attendance calculations."
            if _is_duplicate else
            (_policy_reason if _policy_exception else
             f"Auto-approved from ZKTeco device '{device.get('name')}'")),
        "device_serial": device["serial_number"],
        "device_id": device["device_id"],
        "device_verify_type": verify_type or None,
        "selfie_base64": None,
    }
    # Idempotency guard: avoid duplicating the same push if the device retries
    exists = await db.attendance.find_one({
        "user_id": user["user_id"],
        "at": record["at"],
        "device_serial": device["serial_number"],
        "kind": record["kind"],
    }, {"_id": 0, "record_id": 1})
    if exists:
        return True, "duplicate_ignored"
    # Iter 175 — contractual employees: machine punches must be approved
    # by the company first (Contractor Punch approvals). Duplicates keep
    # their "duplicate" status (never enter any approval queue).
    if not _is_duplicate and not _policy_exception:
        from routes.attendance_core import apply_contractual_gate
        await apply_contractual_gate(record)
    # Iter 250 — attach a parked machine photo (ATTPHOTO that arrived
    # before this ATTLOG line) to the new punch record.
    try:
        ph = await db.biometric_photos.find_one_and_delete({
            "device_serial": device["serial_number"],
            "device_user_id": device_user_id,
            "at": {"$gte": (dt - timedelta(seconds=90)).isoformat(),
                   "$lte": (dt + timedelta(seconds=90)).isoformat()},
        })
        if ph and ph.get("photo_base64"):
            record["selfie_base64"] = ph["photo_base64"]
            record["photo_source"] = "zkteco_attphoto"
    except Exception:
        pass
    # Iter 581 — central onboarding eligibility engine: machine punches
    # from employees with missing mandatory onboarding data are stored
    # but HELD/BLOCKED (never deleted) until HR releases them.
    if record.get("status") in ("approved", "pending"):
        try:
            from shared.attendance_eligibility import (
                apply_to_record as _elig_apply,
                evaluate_for_punch as _elig_eval,
            )
            _elig = await _elig_eval(
                db, user["user_id"], company_id=record.get("company_id"),
                punch_date=record["date"])
            _elig_apply(record, _elig)
        except Exception:
            logger.exception("[iter581] biometric eligibility check failed")
    await db.attendance.insert_one(record)
    return True, (f"duplicate_within_{_dedup_min}min_stored" if _is_duplicate else None)


def _resync_active(device: dict) -> bool:
    """True while an admin-triggered 'fetch old data' window is open."""
    ru = (device or {}).get("resync_until")
    if not ru:
        return False
    try:
        return datetime.fromisoformat(str(ru).replace("Z", "+00:00")) > datetime.now(timezone.utc)
    except Exception:
        return False


async def _ingest_attphoto(raw: bytes, device: dict) -> int:
    """Iter 250 (user request) — ingest a punch PHOTO pushed by the machine
    (table=ATTPHOTO). Body: text headers (PIN=YYYYMMDDHHMMSS-<pin>.jpg,
    size=, CMD=uploadphoto) followed by raw JPEG bytes. The photo is
    attached to the matching attendance record (same employee, punch time
    within ±90s) as selfie_base64 — the same field mobile selfie punches
    use, so every existing report/detail view shows it natively."""
    jpg_at = raw.find(b"\xff\xd8\xff")
    if jpg_at < 0:
        return 0
    header = raw[:jpg_at].decode("utf-8", errors="ignore")
    m = re.search(r"PIN=(\d{14})-([^.\s&]+)\.jpg", header)
    if not m:
        # Iter 543 — firmware variants omit the "PIN=" prefix and just
        # send the photo filename: 20260811093012-42.jpg
        m = re.search(r"(\d{14})-([0-9A-Za-z_]+)\.jpg", header)
    if not m:
        return 0
    ts_raw, pin = m.group(1), m.group(2)
    try:
        dt = datetime.strptime(ts_raw, "%Y%m%d%H%M%S").replace(tzinfo=timezone.utc)
    except ValueError:
        return 0
    photo_b64 = base64.b64encode(raw[jpg_at:]).decode()
    user = await _match_employee_for_bio(pin, device.get("company_id"))
    attached = False
    if user:
        rec = await db.attendance.find_one(
            {
                "user_id": user["user_id"],
                "device_serial": device["serial_number"],
                "at": {"$gte": (dt - timedelta(seconds=90)).isoformat(),
                       "$lte": (dt + timedelta(seconds=90)).isoformat()},
            },
            {"_id": 0, "record_id": 1, "selfie_base64": 1},
            sort=[("at", 1)],
        )
        if rec and not rec.get("selfie_base64"):
            await db.attendance.update_one(
                {"record_id": rec["record_id"]},
                {"$set": {"selfie_base64": photo_b64,
                          "photo_source": "zkteco_attphoto"}},
            )
            attached = True
    if not attached:
        # Photo arrived before its ATTLOG line (or unmapped user) — park it;
        # the ATTLOG ingest picks it up right after inserting the punch.
        await db.biometric_photos.insert_one({
            "device_serial": device["serial_number"],
            "device_user_id": pin,
            "at": dt.isoformat(),
            "photo_base64": photo_b64,
            "received_at": _now_iso_z(),
        })
    return 1


# ---------------------------------------------------------------------------
# Iter 261 — Phase 2: fingerprint / face template capture & cross-device sync.
# ---------------------------------------------------------------------------
_TMPL_LINE_RE = re.compile(r"^(FP|FACE|BIODATA)\s+", re.IGNORECASE)


def _kv_parse(chunk: str) -> dict:
    """Parse 'K=V<TAB>K=V…' (or '&'-separated) into a lower-cased dict."""
    out: dict = {}
    parts = re.split(r"[\t&]", chunk)
    for p in parts:
        if "=" not in p:
            continue
        k, v = p.split("=", 1)
        out[k.strip().lower()] = v.strip()
    return out


async def _ingest_templates(raw: str, device: dict) -> int:
    """Capture fingerprint (FP / FINGERTMP), face (FACE) and unified
    bio-data (BIODATA — newer firmware, Type 1=FP, 2/8/9=Face) templates
    pushed by the machine, so admins can re-push ("sync") them to any
    other machine of the firm. Upserts into ``db.biometric_templates``
    keyed by (company_id, pin, kind, slot)."""
    saved = 0
    for line in (raw or "").splitlines():
        line = line.strip()
        if not line:
            continue
        # Iter 419 — MACHINE-ONLY SYNC: capture the machine's own user rows
        # (PIN + Name + Card/Passwd) so machines can be synchronized with
        # each other even BEFORE the Employee Master is fed in the portal.
        if re.match(r"^USER\s+PIN=", line, re.IGNORECASE):
            uf = _kv_parse(line[4:].strip())
            upin = (uf.get("pin") or "").strip()
            if upin:
                await db.biometric_machine_users.update_one(
                    {"company_id": device.get("company_id"), "pin": upin},
                    {"$set": {
                        "company_id": device.get("company_id"),
                        "pin": upin,
                        "name": (uf.get("name") or "").strip(),
                        "pri": uf.get("pri") or "0",
                        "passwd": uf.get("passwd") or "",
                        "card": uf.get("card") or "",
                        "grp": uf.get("grp") or "",
                        "device_serial": device.get("serial_number"),
                        "captured_at": _now_iso_z(),
                    }},
                    upsert=True,
                )
                saved += 1
            continue
        # Iter 495 (user request — "use SDK to get photo of registered
        # employee") — capture USERPIC / BIOPHOTO pushes: the face photo
        # registered ON THE MACHINE. Stored on the machine-user row and
        # synced into the Employee Master photo when the employee has no
        # portal photo yet, so the Punch Log / grids show real faces.
        if re.match(r"^(USERPIC|BIOPHOTO)\s+PIN=", line, re.IGNORECASE):
            _kw, _, _rest = line.partition(" ")
            pf = _kv_parse(_rest.strip())
            ppin = (pf.get("pin") or "").strip()
            pb64 = (pf.get("content") or "").strip()
            if ppin and pb64:
                await db.biometric_machine_users.update_one(
                    {"company_id": device.get("company_id"), "pin": ppin},
                    {"$set": {
                        "company_id": device.get("company_id"),
                        "pin": ppin,
                        "photo_b64": pb64,
                        "photo_wire": _kw.lower(),
                        "photo_at": _now_iso_z(),
                        "device_serial": device.get("serial_number"),
                    }},
                    upsert=True,
                )
                _emp = await _match_employee_for_bio(ppin, device.get("company_id"))
                if not _emp and ppin.lstrip("0") and ppin.lstrip("0") != ppin:
                    _emp = await _match_employee_for_bio(
                        ppin.lstrip("0"), device.get("company_id"))
                if _emp and not _emp.get("profile_photo_base64"):
                    await db.users.update_one(
                        {"user_id": _emp["user_id"]},
                        {"$set": {
                            "profile_photo_base64": pb64,
                            "profile_photo_source": "machine",
                            "profile_photo_updated_at": _now_iso_z(),
                        },
                         "$unset": {"profile_photo_thumb": ""}})
                saved += 1
            continue
        m = _TMPL_LINE_RE.match(line)
        if m:
            prefix = m.group(1).upper()
            fields = _kv_parse(line[m.end():])
        elif line.lower().startswith("pin=") and "tmp=" in line.lower():
            # table=BIODATA bodies sometimes omit the leading keyword.
            prefix = "BIODATA"
            fields = _kv_parse(line)
        else:
            continue
        pin = fields.get("pin")
        tmp = fields.get("tmp")
        if not pin or not tmp:
            continue
        if prefix == "FP":
            kind, slot = "fp", fields.get("fid") or "0"
        elif prefix == "FACE":
            kind, slot = "face", fields.get("fid") or "0"
        else:  # BIODATA
            btype = fields.get("type") or ""
            kind = "fp" if btype == "1" else "face"
            slot = f"{fields.get('no') or '0'}:{fields.get('index') or '0'}"
        doc = {
            "company_id": device.get("company_id"),
            "pin": pin,
            "kind": kind,
            "slot": slot,
            "wire": prefix.lower(),          # fp / face / biodata push format
            "size": fields.get("size"),
            "valid": fields.get("valid") or "1",
            "tmp": tmp,
            # BIODATA extras (needed to re-push in the same wire format)
            "no": fields.get("no"),
            "index": fields.get("index"),
            "duress": fields.get("duress") or "0",
            "bio_type": fields.get("type"),
            "major_ver": fields.get("majorver"),
            "minor_ver": fields.get("minorver"),
            "format": fields.get("format") or "0",
            "device_serial": device.get("serial_number"),
            "captured_at": _now_iso_z(),
        }
        await db.biometric_templates.update_one(
            {"company_id": doc["company_id"], "pin": pin,
             "kind": kind, "slot": slot},
            {"$set": doc},
            upsert=True,
        )
        saved += 1
        # Iter 267 — Sync Engine conflict: a machine holding biometric data
        # the portal has nowhere else is logged for admin review.
        try:
            from routes.sync_engine import log_template_conflict
            await log_template_conflict(
                doc["company_id"], pin, device.get("serial_number"), kind)
        except Exception:
            pass
    if saved:
        await db.biometric_devices.update_one(
            {"serial_number": device["serial_number"]},
            {"$set": {"last_template_at": _now_iso_z()},
             "$inc": {"templates_captured": saved}},
        )
        logger.info("[zkteco] SN=%s captured %d template(s)",
                    device.get("serial_number"), saved)
    return saved


def _template_to_cmd(t: dict) -> str:
    """Rebuild the exact DATA UPDATE wire command for a stored template."""
    if t.get("wire") == "biodata":
        return (
            "DATA UPDATE BIODATA "
            f"Pin={t['pin']}\tNo={t.get('no') or '0'}\tIndex={t.get('index') or '0'}"
            f"\tValid={t.get('valid') or '1'}\tDuress={t.get('duress') or '0'}"
            f"\tType={t.get('bio_type') or ('1' if t.get('kind') == 'fp' else '9')}"
            f"\tMajorVer={t.get('major_ver') or '0'}\tMinorVer={t.get('minor_ver') or '0'}"
            f"\tFormat={t.get('format') or '0'}\tTmp={t['tmp']}"
        )
    if t.get("kind") == "face":
        return (
            "DATA UPDATE FACE "
            f"PIN={t['pin']}\tFID={t.get('slot') or '0'}\tSIZE={t.get('size') or len(t['tmp'])}"
            f"\tVALID={t.get('valid') or '1'}\tTMP={t['tmp']}"
        )
    return (
        "DATA UPDATE FINGERTMP "
        f"PIN={t['pin']}\tFID={t.get('slot') or '0'}\tSize={t.get('size') or len(t['tmp'])}"
        f"\tValid={t.get('valid') or '1'}\tTMP={t['tmp']}"
    )


async def _get_device_or_404(sn: str) -> dict:
    device = await db.biometric_devices.find_one(
        {"serial_number": sn}, {"_id": 0}
    )
    if not device or not device.get("enabled", True):
        # We return a plain-text response with status 200 anyway so the device
        # keeps retrying — but log the unknown serial for admin visibility.
        await db.biometric_unknown.update_one(
            {"serial_number": sn},
            {"$setOnInsert": {"first_seen_at": _now_iso_z()},
             # Iter 608 — a fresh ping from a real device un-dismisses it.
             "$set": {"last_seen_at": _now_iso_z(), "dismissed": False},
             "$inc": {"hits": 1}},
            upsert=True,
        )
        raise HTTPException(status_code=404, detail=f"Unknown device {sn}")
    return device


def _client_ip(request: Request) -> Optional[str]:
    """Best-effort real client IP behind the k8s / nginx proxy."""
    xff = request.headers.get("x-forwarded-for") or request.headers.get("X-Forwarded-For")
    if xff:
        return xff.split(",")[0].strip()
    xr = request.headers.get("x-real-ip")
    if xr:
        return xr.strip()
    return request.client.host if request.client else None


async def _record_source_ip(sn: str, ip: Optional[str]) -> None:
    """SEC-002 — remember where a device's traffic comes from, for the
    admin's visibility and the optional per-device IP lock."""
    if not ip:
        return
    await db.biometric_devices.update_one(
        {"serial_number": sn},
        {"$set": {"last_source_ip": ip},
         "$addToSet": {"seen_ips": ip}},
    )


def _ip_blocked(device: dict, ip: Optional[str]) -> bool:
    """SEC-002 — when the admin has locked a device to specific IP(s),
    reject traffic from anywhere else. Default (no lock) allows all, so
    existing live machines keep working untouched."""
    if not device.get("ip_lock"):
        return False
    allow = device.get("ip_allowlist") or []
    if not allow:
        return False
    return ip not in allow


# ---------------------------------------------------------------------------
# iClock endpoints (called by the ZKTeco firmware — no auth header)
# ---------------------------------------------------------------------------
@router.get("/iclock/cdata")
async def iclock_handshake(
    request: Request,
    SN: str = Query(..., description="Device serial number"),
    options: Optional[str] = Query(None),
    pushver: Optional[str] = Query(None),
    language: Optional[str] = Query(None),
    PushOptionsFlag: Optional[str] = Query(None),
):
    """Initial handshake — device calls this when it comes online. We reply
    the config block telling it how often to push logs, what
    tables we accept and what the server clock currently is. This is what
    turns ADMS into a *real-time* channel: the device holds an HTTP long-poll
    open and pushes each new punch within a couple of seconds."""
    device = await _get_device_or_404(SN)
    await _record_source_ip(SN, _client_ip(request))
    await db.biometric_devices.update_one(
        {"serial_number": SN},
        {"$set": {
            "last_seen_at": _now_iso_z(),
            "last_handshake_at": _now_iso_z(),
            "firmware_pushver": pushver,
        }},
    )
    # Iter 250 (user bug: old machine data never downloaded) — while a
    # re-sync window is active we answer ATTLOGStamp=0, which resets the
    # device's upload cursor so it re-transmits EVERY attendance log stored
    # in its memory (idempotency guard skips duplicates server-side).
    # Iter 474 (user report — "registered 4 machines, showing ONLINE but no
    # data"): a FRESHLY REGISTERED machine also gets ATTLOGStamp=0 until its
    # first data push lands — otherwise answering "None" tells the machine
    # the server is up-to-date and it NEVER uploads the punches already
    # stored in its memory (it would only send punches made after
    # registration). Once any push arrives, normal stamping resumes.
    _first_sync = not (device.get("last_push_at"))
    att_stamp = "0" if (_resync_active(device) or _first_sync) else "None"
    # Standard ADMS response — see ZKTeco Push SDK docs
    body_lines = [
        f"GET OPTION FROM: {SN}",
        f"ATTLOGStamp={att_stamp}",
        "OPERLOGStamp=9999",
        f"ATTPHOTOStamp={att_stamp}",
        "ErrorDelay=30",
        "Delay=10",
        "TransTimes=00:00;14:05",
        "TransInterval=1",
        "TransFlag=TransData AttLog OpLog AttPhoto EnrollUser ChgUser EnrollFP ChgFP UserPic FvFingerVein",
        # Iter 263 — per-device GMT setting (default India GMT+5:30).
        # Previously hardcoded TimeZone=8 (China) which could drift clocks.
        f"TimeZone={_zk_timezone_value(device)}",
        "Realtime=1",
        "Encrypt=None",
        "ServerVer=SKSharma-1.0",
    ]
    return PlainTextResponse("\n".join(body_lines) + "\n")


@router.post("/iclock/cdata")
async def iclock_push(
    request: Request,
    SN: str = Query(...),
    table: Optional[str] = Query(None),
    Stamp: Optional[str] = Query(None),
):
    """Punch push endpoint. The device POSTs blocks of ATTLOG / OPERLOG /
    ATTPHOTO lines here. We parse ATTLOG lines and insert them into the same
    `attendance` collection the mobile app writes to — so reports blend both
    sources natively."""
    device = await _get_device_or_404(SN)
    src_ip = _client_ip(request)
    await _record_source_ip(SN, src_ip)
    raw_bytes = await request.body()
    raw = raw_bytes.decode("utf-8", errors="ignore")
    inserted = 0
    skipped = 0
    reasons: List[str] = []
    # SEC-002 — reject payloads from unauthorized IPs when the device is
    # IP-locked (park for audit, never ingest into attendance/payroll).
    if _ip_blocked(device, src_ip):
        logger.warning(
            "[zkteco][SEC] push from unauthorized IP %s for SN=%s — parked, not ingested",
            src_ip, SN)
        await db.biometric_locked_punches.insert_one({
            "device_serial": SN,
            "raw": raw[:8000],
            "source_ip": src_ip,
            "reason": "ip_blocked",
            "received_at": _now_iso_z(),
        })
        return PlainTextResponse("OK: 0\n")
    if (table or "").upper() == "ATTLOG":
        if device.get("locked"):
            # Iter 261 — device is LOCKED from the portal: punches are
            # parked for audit and NOT ingested into attendance.
            await db.biometric_locked_punches.insert_one({
                "device_serial": SN,
                "raw": raw[:8000],
                "received_at": _now_iso_z(),
            })
            skipped = len([ln for ln in raw.splitlines() if ln.strip()])
            reasons.append("device_locked")
        else:
            for line in raw.splitlines():
                ok, reason = await _ingest_attlog_line(line, device)
                if ok:
                    if reason == "duplicate_ignored":
                        skipped += 1
                    else:
                        inserted += 1
                else:
                    skipped += 1
                    if reason:
                        reasons.append(reason)
    elif (table or "").upper() == "ATTPHOTO":
        # Iter 250 — punch photos from the machine (attached to the punch).
        # Iter 543 — photo DIAGNOSTICS: count every ATTPHOTO push and keep
        # the last header + error on the device doc so admins can see
        # exactly what the machine sends (Biometric Devices screen).
        _ph_err = None
        try:
            inserted = await _ingest_attphoto(raw_bytes, device)
            if not inserted:
                _ph_err = ("no JPEG marker in body"
                           if raw_bytes.find(b"\xff\xd8\xff") < 0
                           else "photo header did not match any known "
                                "PIN/filename format")
        except Exception as e:  # noqa: BLE001
            _ph_err = str(e)[:200]
            logger.warning("[zkteco] ATTPHOTO ingest failed", exc_info=True)
        _jpg_i = raw_bytes.find(b"\xff\xd8\xff")
        _head = raw_bytes[:_jpg_i if _jpg_i > 0 else 200][:200]
        await db.biometric_devices.update_one(
            {"serial_number": SN},
            {"$inc": {"photo_recv_count": 1,
                      "photo_saved_count": 1 if inserted else 0},
             "$set": {"photo_last_at": _now_iso_z(),
                      "photo_last_error": _ph_err,
                      "photo_last_head": _head.decode("utf-8",
                                                      errors="ignore")}},
        )
    else:
        # Iter 261 — Phase 2: OPERLOG / BIODATA pushes may carry enrolled
        # fingerprint / face templates — capture them for cross-device sync.
        # Iter 495 — run the parser for EVERY non-ATTLOG/ATTPHOTO table
        # (some firmwares answer DATA QUERY USERINFO with table=USERINFO or
        # no table at all). The parser is prefix-guarded so unknown lines
        # are simply skipped.
        try:
            await _ingest_templates(raw, device)
        except Exception:
            logger.warning("[zkteco] template ingest failed", exc_info=True)
        # OPERLOG / EnrollUser etc. — log the receipt.
        await db.biometric_operlog.insert_one({
            "device_serial": SN,
            "table": table,
            "stamp": Stamp,
            "raw": raw[:8000],  # cap to keep the doc size reasonable
            "received_at": _now_iso_z(),
        })
    await db.biometric_devices.update_one(
        {"serial_number": SN},
        {"$set": {
            "last_seen_at": _now_iso_z(),
            "last_push_at": _now_iso_z(),
            "last_push_table": table,
        },
         "$inc": {"total_pushes": 1, "total_punches_ingested": inserted}},
    )
    # Iter 77n — Broadcast to the firm channel whenever a push adds
    # punches so admin dashboards refresh in real time.
    if inserted > 0:
        try:
            from utils.ws_broker import broker as _ws
            firm_id = device.get("company_id") if isinstance(device, dict) else None
            if firm_id:
                await _ws.broadcast_firm(firm_id, {
                    "type": "attendance.zk-pushed",
                    "device_serial": SN,
                    "inserted": inserted,
                    "table": table,
                })
        except Exception:
            pass
    # ZKTeco expects a plain "OK" line and the stamp advancement
    # so it can move its cursor forward.
    logger.info(
        "[zkteco] SN=%s table=%s inserted=%d skipped=%d reasons=%s",
        SN, table, inserted, skipped, reasons[:5],
    )
    return PlainTextResponse(f"OK: {inserted}\n")


@router.post("/iclock/fdata")
async def iclock_fdata(
    request: Request,
    SN: str = Query(...),
    table: Optional[str] = Query(None),
):
    """Iter 543 — some eSSL / ZKTeco firmware uploads punch PHOTOS to
    /iclock/fdata instead of /iclock/cdata?table=ATTPHOTO. Accept both:
    if the body carries a JPEG we run the same ATTPHOTO ingest (with the
    same diagnostics counters); anything else is acknowledged."""
    device = await _get_device_or_404(SN)
    raw_bytes = await request.body()
    if raw_bytes.find(b"\xff\xd8\xff") >= 0:
        _ph_err = None
        saved = 0
        try:
            saved = await _ingest_attphoto(raw_bytes, device)
            if not saved:
                _ph_err = ("photo header did not match any known "
                           "PIN/filename format (fdata)")
        except Exception as e:  # noqa: BLE001
            _ph_err = str(e)[:200]
            logger.warning("[zkteco] fdata photo ingest failed",
                           exc_info=True)
        _jpg_i = raw_bytes.find(b"\xff\xd8\xff")
        await db.biometric_devices.update_one(
            {"serial_number": SN},
            {"$inc": {"photo_recv_count": 1,
                      "photo_saved_count": 1 if saved else 0},
             "$set": {"photo_last_at": _now_iso_z(),
                      "photo_last_error": _ph_err,
                      "photo_last_head": raw_bytes[:_jpg_i][:200].decode(
                          "utf-8", errors="ignore")}},
        )
    return PlainTextResponse("OK")


@router.get("/iclock/getrequest")
async def iclock_getrequest(
    request: Request,
    SN: str = Query(...),
    INFO: Optional[str] = Query(None),
):
    """Command-request long-poll — the device asks the server if there are
    any pending commands (enroll user, delete user, sync time, reboot).
    Iter 250: while a re-sync window is active we answer ONE `CHECK`
    command, which makes the device immediately re-handshake and re-upload
    its stored attendance logs (combined with ATTLOGStamp=0 above). Every
    successful call still refreshes the device heartbeat so the admin UI can
    show it as online."""
    try:
        device = await _get_device_or_404(SN)
    except HTTPException:
        return PlainTextResponse("OK\n")
    src_ip = _client_ip(request)
    await _record_source_ip(SN, src_ip)
    # SEC-002 — if the device is IP-locked, never hand queued commands
    # (which can include template/PII sync) to a caller from a foreign IP.
    if _ip_blocked(device, src_ip):
        logger.warning(
            "[zkteco][SEC] getrequest from unauthorized IP %s for SN=%s — withholding commands",
            src_ip, SN)
        return PlainTextResponse("OK\n")
    await db.biometric_devices.update_one(
        {"serial_number": SN},
        {"$set": {
            "last_seen_at": _now_iso_z(),
            "last_getrequest_info": INFO,
            **_parse_info(INFO),
        }},
    )
    # Iter 494 (user: "Name in Machine — some companies' data not showing
    # on NEW registered machines") — automatically ask every machine ONCE
    # for its user database (PIN + Name) so the Punch Log's "Name in
    # Machine" column fills without the manual 'Fetch from machine' step.
    if not device.get("userinfo_query_sent"):
        await _queue_cmd(SN, "DATA QUERY USERINFO", "system:auto",
                         "Auto user-name sync (new machine)")
        await db.biometric_devices.update_one(
            {"serial_number": SN},
            {"$set": {"userinfo_query_sent": True,
                      "userinfo_query_sent_at": _now_iso_z()}},
        )
    # Iter 495 (user request) — also ask ONCE for the registered employee
    # PHOTOS on the machine (USERPIC / BIOPHOTO) so the Punch Log & grids
    # can show real faces without portal uploads.
    if not device.get("userpic_query_sent"):
        await _queue_cmd(SN, "DATA QUERY USERPIC", "system:auto",
                         "Auto user-photo sync (registered face photos)")
        await _queue_cmd(SN, "DATA QUERY BIOPHOTO", "system:auto",
                         "Auto bio-photo sync (registered face photos)")
        await db.biometric_devices.update_one(
            {"serial_number": SN},
            {"$set": {"userpic_query_sent": True,
                      "userpic_query_sent_at": _now_iso_z()}},
        )
    if _resync_active(device) and not device.get("resync_check_sent"):
        await db.biometric_devices.update_one(
            {"serial_number": SN},
            {"$set": {"resync_check_sent": True,
                      "resync_check_sent_at": _now_iso_z()}},
        )
        cmd_id = int(datetime.now(timezone.utc).timestamp())
        logger.info("[zkteco] SN=%s issuing CHECK for old-data re-sync", SN)
        return PlainTextResponse(f"C:{cmd_id}:CHECK\n")
    # Iter 258 — centralized device management: deliver queued remote
    # commands (restart / sync / clear-log / push-employee ...) to the
    # device, oldest first. Iter 419 (user: "make it direct — no waiting"):
    # batch raised 5 → 30 per poll so full-firm syncs finish in a few
    # polls instead of trickling for many minutes. There is NO approval
    # step — commands transfer the moment the machine contacts the server.
    pending = await db.biometric_device_cmds.find(
        {"device_serial": SN, "status": "pending"}, {"_id": 0}
    ).sort("created_at", 1).to_list(30)
    if pending:
        lines = []
        for c in pending:
            lines.append(f"C:{c['cmd_id']}:{c['command']}")
            await db.biometric_device_cmds.update_one(
                {"cmd_id": c["cmd_id"]},
                {"$set": {"status": "sent", "sent_at": _now_iso_z()}},
            )
        logger.info("[zkteco] SN=%s delivering %d command(s)", SN, len(lines))
        return PlainTextResponse("\n".join(lines) + "\n")
    return PlainTextResponse("OK\n")


def _parse_info(info: Optional[str]) -> dict:
    """Best-effort parse of the getrequest INFO param:
    'FWVer,UserCount,FpCount,AttLogCount,DeviceIP[,...]'."""
    if not info:
        return {}
    parts = [p.strip() for p in str(info).split(",")]
    out: dict = {}
    if parts and parts[0]:
        out["firmware"] = parts[0][:60]
    for idx, key in ((1, "user_count"), (2, "fp_count"), (3, "att_log_count")):
        if len(parts) > idx and parts[idx].isdigit():
            out[key] = int(parts[idx])
    if len(parts) > 4 and re.match(r"^\d{1,3}(\.\d{1,3}){3}$", parts[4]):
        out["device_ip"] = parts[4]
    return out


@router.get("/iclock/ping")
async def iclock_ping(SN: Optional[str] = Query(None)):
    """Heartbeat used by some firmwares between long-polls."""
    if SN:
        await db.biometric_devices.update_one(
            {"serial_number": SN},
            {"$set": {"last_seen_at": _now_iso_z()}},
        )
    return PlainTextResponse("OK\n")


@router.post("/iclock/devicecmd")
async def iclock_devicecmd(request: Request, SN: str = Query(...)):
    """Command-result reporting — device tells us the outcome of any command
    we previously issued via getrequest. We just accept and log."""
    raw = (await request.body()).decode("utf-8", errors="ignore")
    await db.biometric_devices.update_one(
        {"serial_number": SN},
        {"$set": {"last_seen_at": _now_iso_z()}},
    )
    await db.biometric_cmd_results.insert_one({
        "device_serial": SN,
        "raw": raw[:4000],
        "received_at": _now_iso_z(),
    })
    # Iter 258 — mark queued commands done/failed: body like "ID=123&Return=0&CMD=DATA"
    try:
        m_id = re.search(r"ID=(\d+)", raw)
        m_ret = re.search(r"Return=(-?\d+)", raw)
        if m_id:
            ok = (m_ret and m_ret.group(1) == "0")
            await db.biometric_device_cmds.update_one(
                {"cmd_id": m_id.group(1)},
                {"$set": {"status": "done" if ok else "failed",
                          "result_return": m_ret.group(1) if m_ret else None,
                          "result_at": _now_iso_z()}},
            )
    except Exception:
        pass
    return PlainTextResponse("OK\n")


# ---------------------------------------------------------------------------
# Iter 258 — Centralized device management: remote commands + employee push.
# ---------------------------------------------------------------------------
async def _queue_cmd(serial: str, command: str, queued_by: str, label: str) -> str:
    # Iter 261 fix — ms-timestamp alone collides when many commands are
    # queued in the same millisecond (template sync); add a random suffix
    # so every command gets a UNIQUE numeric ID for devicecmd correlation.
    cmd_id = (str(int(datetime.now(timezone.utc).timestamp() * 1000))[-9:]
              + str(random.randint(100, 999)))
    await db.biometric_device_cmds.insert_one({
        "cmd_id": cmd_id,
        "device_serial": serial,
        "command": command,
        "label": label,
        "status": "pending",
        "queued_by": queued_by,
        "created_at": _now_iso_z(),
    })
    return cmd_id


_REMOTE_ACTIONS = {
    "restart": ("REBOOT", "Restart device"),
    "sync_data": ("CHECK", "Synchronize data"),
    "refresh_info": ("INFO", "Refresh device information"),
    "clear_attlog": ("CLEAR LOG", "Clear attendance logs on device"),
    # Iter 261 — Phase 2: remote door unlock (access-control relay).
    "unlock_door": ("AC_UNLOCK", "Unlock door (relay)"),
}


def _zk_encode_datetime(dt: datetime) -> int:
    """ZKTeco option-encoding of a wall-clock datetime:
    ((y-2000)*12*31 + (m-1)*31 + (d-1)) * 86400 + h*3600 + min*60 + s."""
    return (((dt.year - 2000) * 12 * 31 + (dt.month - 1) * 31 + (dt.day - 1)) * 86400
            + dt.hour * 3600 + dt.minute * 60 + dt.second)


def _zk_datetime_now(device: Optional[dict] = None) -> int:
    """Encoding of the CURRENT wall-clock time in the device's configured
    GMT zone (default India +05:30)."""
    mins = _parse_gmt_offset_minutes((device or {}).get("gmt_offset") if device else None)
    local = datetime.now(timezone.utc) + timedelta(minutes=mins)
    return _zk_encode_datetime(local)


# ---------------------------------------------------------------------------
# Iter 294 — Generic JSON punch webhook (Matrix COSEC / Mantra / middleware).
#
# Devices (or their vendor middleware, e.g. Matrix COSEC server event export,
# Mantra ADMS bridge) POST simple JSON punches here using the per-device
# secret key generated at registration:
#
#   POST /api/device-webhook/{webhook_key}
#   { "punches": [ {"user_id": "72", "timestamp": "2026-06-25 09:01:00",
#                   "direction": "in" } ] }
#
# Punches flow through the exact same ingest pipeline as ZKTeco ATTLOG lines
# (bio_code matching, IN/OUT alternation, dedupe, contractual gate).
# ---------------------------------------------------------------------------

class WebhookPunch(BaseModel):
    user_id: str
    timestamp: str                      # ISO 8601 or "YYYY-MM-DD HH:MM:SS"
    direction: Optional[Literal["in", "out"]] = None


class WebhookPushBody(BaseModel):
    punches: List[WebhookPunch]


def _normalize_webhook_ts(raw: str) -> Optional[str]:
    """Accept ISO ('2026-06-25T09:01:00', with/without Z/offset) or plain
    'YYYY-MM-DD HH:MM:SS' and return the ZK wall-clock format."""
    s = (raw or "").strip().replace("T", " ")
    s = re.sub(r"(Z|[+-]\d{2}:?\d{2})$", "", s).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M:%S",
                "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None


@router.post("/device-webhook/{webhook_key}")
async def device_webhook_push(webhook_key: str, body: WebhookPushBody):
    device = await db.biometric_devices.find_one(
        {"webhook_key": webhook_key}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Unknown webhook key")
    if not device.get("enabled", True) or device.get("locked"):
        raise HTTPException(status_code=403, detail="Device is disabled or locked")
    inserted = 0
    skipped = 0
    reasons: List[str] = []
    for p in body.punches[:2000]:
        ts = _normalize_webhook_ts(p.timestamp)
        if not ts:
            skipped += 1
            reasons.append(f"bad_timestamp:{p.timestamp}")
            continue
        dev = dict(device)
        if p.direction in ("in", "out"):
            dev["kind"] = p.direction
        ok, reason = await _ingest_attlog_line(f"{p.user_id}\t{ts}", dev)
        if ok:
            if reason == "duplicate_ignored":
                skipped += 1
            else:
                inserted += 1
        else:
            skipped += 1
            if reason:
                reasons.append(reason)
    await db.biometric_devices.update_one(
        {"device_id": device["device_id"]},
        {"$set": {"last_seen_at": _now_iso_z(), "last_push_at": _now_iso_z(),
                  "last_push_table": "WEBHOOK"},
         "$inc": {"total_pushes": 1, "total_punches_ingested": inserted}},
    )
    logger.info("[webhook] device=%s inserted=%d skipped=%d reasons=%s",
                device["serial_number"], inserted, skipped, reasons[:5])
    return {"ok": True, "inserted": inserted, "skipped": skipped,
            "reasons": reasons[:10]}


@router.post("/biometric/devices/{device_id}/regenerate-webhook-key")
async def regenerate_webhook_key(
    device_id: str,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    device = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    key = secrets.token_hex(12)
    await db.biometric_devices.update_one(
        {"device_id": device_id}, {"$set": {"webhook_key": key}})
    return {"ok": True, "webhook_key": key}


@router.post("/biometric/devices/{device_id}/command")
async def send_device_command(
    device_id: str,
    payload: dict,
    authorization: Optional[str] = Header(None),
):
    """Queue a remote control command; the device picks it up on its next
    getrequest poll (seconds when online). Supported: restart, sync_data,
    refresh_info, clear_attlog."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    action = str(payload.get("action") or "")
    if action not in _REMOTE_ACTIONS and action not in ("sync_time", "lock", "unlock"):
        raise HTTPException(status_code=400, detail=f"Unknown action '{action}'")
    device = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    if action in ("lock", "unlock"):
        # Iter 261 — Phase 2: portal-side device LOCK. While locked, every
        # punch the machine pushes is parked (audit trail) instead of
        # entering attendance — effectively taking the machine out of
        # service without touching the hardware.
        locked = action == "lock"
        await db.biometric_devices.update_one(
            {"device_id": device_id},
            {"$set": {"locked": locked,
                      "locked_by": admin["user_id"] if locked else None,
                      "locked_at": _now_iso_z() if locked else None}},
        )
        return {
            "ok": True, "locked": locked,
            "message": (
                "Device LOCKED — punches from this machine are parked and "
                "will NOT enter attendance until you unlock it."
                if locked else
                "Device UNLOCKED — punches flow into attendance again."
            ),
        }
    if action == "sync_time":
        # Iter 259 — set the machine's DATE & TIME. Iter 262 (user
        # request): the admin can EDIT the exact date & time in the portal
        # dialog before applying; when no date/time is supplied we fall
        # back to the current Indian (IST) wall-clock time.
        d_raw = str(payload.get("date") or "").strip()   # DD-MM-YYYY
        t_raw = str(payload.get("time") or "").strip()   # HH:MM[:SS]
        target = None
        if d_raw or t_raw:
            for dfmt in ("%d-%m-%Y", "%Y-%m-%d"):
                for tfmt in ("%H:%M:%S", "%H:%M"):
                    try:
                        target = datetime.strptime(f"{d_raw} {t_raw}", f"{dfmt} {tfmt}")
                        break
                    except ValueError:
                        continue
                if target:
                    break
            if target is None:
                raise HTTPException(
                    status_code=400,
                    detail="Invalid date/time — use DD-MM-YYYY and HH:MM (24h)")
        if target is not None:
            cmd = f"SET OPTION DateTime={_zk_encode_datetime(target)}"
            label = f"Set device date & time to {target.strftime('%d-%b-%Y %H:%M:%S')}"
        else:
            cmd = f"SET OPTION DateTime={_zk_datetime_now(device)}"
            label = (
                "Set device date & time (current, GMT"
                f"{device.get('gmt_offset') or '+05:30'})"
            )
    else:
        cmd, label = _REMOTE_ACTIONS[action]
    cmd_id = await _queue_cmd(device["serial_number"], cmd, admin["user_id"], label)
    return {
        "ok": True, "cmd_id": cmd_id,
        "message": f"{label} queued — the machine executes it within seconds while online.",
    }


@router.post("/biometric/devices/push-employees")
async def push_employees_to_devices(
    payload: dict,
    authorization: Optional[str] = Header(None),
):
    """Iter 258 (user request) — push employee names (by Bio Code) INTO the
    machines so the device display shows the correct name. Body:
    { company_id, user_id? , device_id? }. Without user_id pushes ALL
    employees that have a bio_code. Errors clearly when the firm has no
    registered machine."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    company_id = admin["company_id"] if admin["role"] == "company_admin" else payload.get("company_id")
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    dev_q: dict = {"company_id": company_id, "enabled": {"$ne": False}}
    if payload.get("device_id"):
        dev_q = {"device_id": payload["device_id"]}
    devices = await db.biometric_devices.find(dev_q, {"_id": 0}).to_list(50)
    if not devices:
        raise HTTPException(
            status_code=404,
            detail="No biometric machine is registered for this company — "
                   "register the device first in ZKTeco Device Setup "
                   "(Biometric Devices screen).",
        )
    emp_q: dict = {"role": "employee", "company_id": company_id,
                   "bio_code": {"$exists": True, "$nin": [None, ""]}}
    if payload.get("user_id"):
        emp_q = {"user_id": payload["user_id"]}
    emps = await db.users.find(
        emp_q, {"_id": 0, "user_id": 1, "name": 1, "bio_code": 1}).to_list(3000)
    if payload.get("user_id") and emps and not (emps[0].get("bio_code") or "").strip():
        raise HTTPException(
            status_code=400,
            detail="This employee has no Bio Code — set the machine punch "
                   "number (Bio Code) in the Employee Master first.",
        )
    if not emps:
        raise HTTPException(status_code=404, detail="No employees with a Bio Code found")
    queued = 0
    for d in devices:
        for e in emps:
            pin = str(e.get("bio_code") or "").strip()
            if not pin:
                continue
            nm = (e.get("name") or "")[:24].replace("\t", " ")
            await _queue_cmd(
                d["serial_number"],
                f"DATA UPDATE USERINFO PIN={pin}\tName={nm}\tPri=0",
                admin["user_id"],
                f"Push employee {nm} ({pin})",
            )
            queued += 1
    return {
        "ok": True, "queued": queued, "devices": len(devices),
        "message": (
            f"{queued} update(s) queued for {len(devices)} machine(s). "
            "Names appear on the device within a minute while it is online."
        ),
    }


# ---------------------------------------------------------------------------
# Iter 261 — Phase 2: fingerprint / face template sync + live dashboard.
# ---------------------------------------------------------------------------
@router.post("/biometric/devices/{device_id}/fetch-templates")
async def fetch_templates_from_device(
    device_id: str,
    authorization: Optional[str] = Header(None),
):
    """Ask the machine to upload its user database + fingerprint / face
    templates to the portal (DATA QUERY commands). Captured templates are
    stored per employee PIN and can then be synced to any other machine."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    device = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    sn = device["serial_number"]
    for cmd, label in (
        ("DATA QUERY USERINFO", "Query user database"),
        ("DATA QUERY FINGERTMP", "Query fingerprint templates"),
        ("DATA QUERY BIODATA", "Query face / bio-data templates"),
        # Iter 600 (user bug) — also pull the registered face PHOTOS so
        # they can be synced to other machines.
        ("DATA QUERY USERPIC", "Query registered user photos"),
        ("DATA QUERY BIOPHOTO", "Query registered bio photos"),
    ):
        await _queue_cmd(sn, cmd, admin["user_id"], label)
    return {
        "ok": True,
        "message": (
            "Template fetch queued — the machine uploads its users, "
            "fingerprints, face data AND registered photos within a "
            "minute while online. "
            "Then use 'Sync FP/Face' on another machine to copy them over."
        ),
    }


@router.post("/biometric/devices/{device_id}/sync-templates")
async def sync_templates_to_device(
    device_id: str,
    payload: Optional[dict] = Body(None),
    authorization: Optional[str] = Header(None),
):
    """Push ALL stored fingerprint / face templates of the firm to this
    machine (skipping ones originally captured from it). Optional body:
    { user_id } to sync a single employee only."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    device = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    company_id = device.get("company_id")
    tq: dict = {"company_id": company_id}
    payload = payload or {}
    if payload.get("user_id"):
        u = await db.users.find_one(
            {"user_id": payload["user_id"]}, {"_id": 0, "bio_code": 1})
        pin = str((u or {}).get("bio_code") or "").strip()
        if not pin:
            raise HTTPException(
                status_code=400,
                detail="This employee has no Bio Code — set it in the "
                       "Employee Master first.")
        tq["pin"] = pin
    templates = await db.biometric_templates.find(tq, {"_id": 0}).to_list(20000)
    templates = [t for t in templates
                 if t.get("device_serial") != device["serial_number"]]
    # Iter 600 (user bug) — the registered FACE PHOTOS captured from other
    # machines (USERPIC / BIOPHOTO) were NEVER included in the sync, so
    # "photos didn't sync from machine 1 to machine 2". Push them too,
    # skipping photos originally captured FROM this device.
    pq: dict = {"company_id": company_id,
                "photo_b64": {"$exists": True, "$nin": [None, ""]}}
    if tq.get("pin"):
        pq["pin"] = tq["pin"]
    photos = await db.biometric_machine_users.find(
        pq, {"_id": 0, "pin": 1, "photo_b64": 1, "photo_wire": 1,
             "device_serial": 1},
    ).to_list(20000)
    photos = [p for p in photos
              if p.get("device_serial") != device["serial_number"]]
    if not templates and not photos:
        raise HTTPException(
            status_code=404,
            detail="No stored templates or photos to sync — first press "
                   "'Fetch FP/Face' on the machine where employees are "
                   "enrolled.",
        )
    sn = device["serial_number"]
    # Push USERINFO (name) once per PIN so the template lands on a user.
    pins = sorted({t["pin"] for t in templates} | {p["pin"] for p in photos})
    emp_by_pin = {}
    async for u in db.users.find(
            {"company_id": company_id, "bio_code": {"$in": pins}},
            {"_id": 0, "bio_code": 1, "name": 1}):
        emp_by_pin[str(u["bio_code"]).strip()] = (u.get("name") or "")[:24].replace("\t", " ")
    queued = 0
    for pin in pins:
        nm = emp_by_pin.get(pin, "")
        await _queue_cmd(
            sn, f"DATA UPDATE USERINFO PIN={pin}\tName={nm}\tPri=0",
            admin["user_id"], f"Sync user {nm or pin}")
        queued += 1
    for t in templates:
        await _queue_cmd(
            sn, _template_to_cmd(t), admin["user_id"],
            f"Sync {t.get('kind')} template PIN {t.get('pin')}")
        queued += 1
    # Iter 600 — queue the face photos (same wire the source machine used:
    # USERPIC for legacy models, BIOPHOTO Type=9 for newer face devices).
    for p in photos:
        pb = p["photo_b64"]
        if str(p.get("photo_wire") or "userpic").lower() == "biophoto":
            cmd = ("DATA UPDATE BIOPHOTO "
                   f"PIN={p['pin']}\tFileName={p['pin']}.jpg\tType=9"
                   f"\tSize={len(pb)}\tContent={pb}")
        else:
            cmd = (f"DATA UPDATE USERPIC PIN={p['pin']}"
                   f"\tSize={len(pb)}\tContent={pb}")
        await _queue_cmd(sn, cmd, admin["user_id"],
                         f"Sync face photo PIN {p['pin']}")
        queued += 1
    return {
        "ok": True, "queued": queued,
        "employees": len(pins), "templates": len(templates),
        "photos": len(photos),
        "message": (
            f"{len(templates)} template(s) + {len(photos)} face photo(s) "
            f"for {len(pins)} employee(s) queued to "
            f"'{device.get('name')}' — they install within a minute while "
            "the machine is online."
        ),
    }


@router.get("/biometric/templates-summary")
async def biometric_templates_summary(
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Per-employee count of stored fingerprint / face templates."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    rows: dict = {}
    async for t in db.biometric_templates.find(q, {"_id": 0, "tmp": 0}):
        key = (t.get("company_id"), t.get("pin"))
        r = rows.setdefault(key, {
            "company_id": t.get("company_id"), "pin": t.get("pin"),
            "fp": 0, "face": 0, "devices": set(), "last_captured": "",
        })
        r["fp" if t.get("kind") == "fp" else "face"] += 1
        if t.get("device_serial"):
            r["devices"].add(t["device_serial"])
        if (t.get("captured_at") or "") > r["last_captured"]:
            r["last_captured"] = t.get("captured_at") or ""
    out = []
    pins = [r["pin"] for r in rows.values()]
    names = {}
    if pins:
        async for u in db.users.find(
                {"bio_code": {"$in": pins}}, {"_id": 0, "bio_code": 1, "name": 1, "company_id": 1}):
            names[(u.get("company_id"), str(u.get("bio_code")).strip())] = u.get("name")
    for key, r in sorted(rows.items(), key=lambda kv: kv[0][1] or ""):
        r["devices"] = sorted(r["devices"])
        r["name"] = names.get(key)
        out.append(r)
    return {"templates": out, "total_templates": sum(r["fp"] + r["face"] for r in out)}


@router.post("/biometric/devices/{device_id}/ip-lock")
async def biometric_ip_lock(
    device_id: str,
    payload: Optional[dict] = Body(None),
    authorization: Optional[str] = Header(None),
):
    """SEC-002 — lock a device so ONLY its current source IP (or a
    supplied allowlist) may push punches / receive commands; or unlock it.
    Body: { mode: 'lock'|'unlock', ips?: ["1.2.3.4", ...] }."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    device = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    payload = payload or {}
    mode = str(payload.get("mode") or "lock")
    if mode == "unlock":
        await db.biometric_devices.update_one(
            {"device_id": device_id},
            {"$set": {"ip_lock": False}},
        )
        return {"ok": True, "ip_lock": False,
                "message": "IP lock removed — punches accepted from any IP again."}
    ips = payload.get("ips") or []
    if not ips:
        cur = device.get("last_source_ip")
        if not cur:
            raise HTTPException(
                status_code=400,
                detail="No source IP seen yet — let the machine push once, "
                       "then lock it to that IP.")
        ips = [cur]
    await db.biometric_devices.update_one(
        {"device_id": device_id},
        {"$set": {"ip_lock": True, "ip_allowlist": ips}},
    )
    return {
        "ok": True, "ip_lock": True, "ip_allowlist": ips,
        "message": (
            f"Device locked to {', '.join(ips)} — punches/commands from any "
            "other IP are now rejected."
        ),
    }


@router.get("/biometric/live-feed")
async def biometric_live_feed(
    company_id: Optional[str] = Query(None),
    limit: int = Query(30, ge=1, le=100),
    authorization: Optional[str] = Header(None),
):
    """Iter 261 — Live Dashboard feed: the most recent machine punches
    (joined with employee + device names). The frontend refreshes this on
    every `attendance.zk-pushed` WebSocket event + a polling fallback."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {"source": {"$regex": "^zkteco:"}}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    recs = await db.attendance.find(
        q,
        {"_id": 0, "user_id": 1, "company_id": 1, "date": 1, "kind": 1,
         "at": 1, "device_serial": 1, "status": 1},
    ).sort("at", -1).to_list(limit)
    uids = list({r["user_id"] for r in recs})
    names = {}
    if uids:
        async for u in db.users.find(
                {"user_id": {"$in": uids}},
                {"_id": 0, "user_id": 1, "name": 1, "bio_code": 1}):
            names[u["user_id"]] = u
    dev_names = {}
    async for d in db.biometric_devices.find({}, {"_id": 0, "serial_number": 1, "name": 1}):
        dev_names[d["serial_number"]] = d.get("name")
    feed = []
    for r in recs:
        u = names.get(r["user_id"]) or {}
        feed.append({
            "at": r.get("at"), "date": r.get("date"), "kind": r.get("kind"),
            "status": r.get("status"),
            # Iter 494 — user_id lets the UI show the employee photo.
            "user_id": r.get("user_id") if u else None,
            "name": u.get("name") or r["user_id"],
            "bio_code": u.get("bio_code"),
            "device": dev_names.get(r.get("device_serial")) or r.get("device_serial"),
        })
    return {"feed": feed}


# ---------------------------------------------------------------------------
# Admin management APIs (auth-protected) — used by the
# /biometric-devices frontend screen.
# ---------------------------------------------------------------------------
@router.post("/biometric/devices")
async def register_biometric_device(
    payload: BiometricDeviceCreate,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    sn = payload.serial_number.strip()
    if not sn:
        raise HTTPException(status_code=400, detail="Serial number is required")
    company_id = payload.company_id
    if admin["role"] == "company_admin":
        company_id = admin["company_id"]  # ignore whatever client sent
    if not company_id:
        raise HTTPException(status_code=400, detail="Please pick a company for this device")
    existing = await db.biometric_devices.find_one({"serial_number": sn})
    if existing:
        raise HTTPException(status_code=409, detail=f"Device {sn} is already registered")
    device = {
        "device_id": f"dev_{uuid.uuid4().hex[:10]}",
        "serial_number": sn,
        "name": payload.name.strip() or f"Device {sn}",
        "kind": payload.kind,
        "company_id": company_id,
        "location": (payload.location or "").strip() or None,
        # Iter 298 — branch tag for two-branch payroll splits.
        "branch_name": (payload.branch_name or "").strip() or None,
        "enabled": payload.enabled,
        # Iter 267 — participates in the multi-device sync engine by default.
        "sync_enabled": True,
        # Iter 263 — machine time zone (validated; default India +05:30).
        "gmt_offset": (payload.gmt_offset or "+05:30").strip() or "+05:30",
        # Iter 294 — multi-brand + JSON webhook key (Matrix / Mantra push).
        "brand": payload.brand or "zkteco",
        # Iter 512 — Direct SDK pull channel (additive; ADMS push unchanged)
        "connection_mode": payload.connection_mode or "push",
        "sdk_vendor": (payload.sdk_vendor or "").strip() or None,
        "device_ip": (payload.device_ip or "").strip() or None,
        "device_port": payload.device_port,
        "comm_key": (payload.comm_key or "").strip() or None,
        "auto_pull_minutes": int(payload.auto_pull_minutes or 0),
        "webhook_key": secrets.token_hex(12),
        "created_at": _now_iso_z(),
        "created_by": admin["user_id"],
        "model": {
            "zkteco": "ZKTeco AC Mini Plus",
            "essl": "eSSL (iClock/ADMS)",
            "bioface": "BIOFACE-MSD1K (iClock/ADMS)",
            "matrix": "Matrix COSEC",
            "mantra": "Mantra",
        }.get(payload.brand or "zkteco", "Generic"),
        "last_seen_at": None,
        "total_pushes": 0,
        "total_punches_ingested": 0,
    }
    await db.biometric_devices.insert_one(device)
    device.pop("_id", None)
    # Iter 474 — the machine may have handshaked BEFORE it was registered
    # (or only handshakes on reboot). Queue a CHECK so it re-initialises
    # within a minute and picks up the first-sync ATTLOGStamp=0, uploading
    # the punches already stored in its memory without a manual reboot.
    try:
        await _queue_cmd(sn, "CHECK", admin["user_id"],
                         "Initial sync after registration")
    except Exception:
        logger.warning("[zkteco] initial CHECK queue failed", exc_info=True)
    return {"ok": True, "device": device}


@router.post("/biometric/devices/{device_id}/resync")
async def resync_biometric_device(
    device_id: str,
    authorization: Optional[str] = Header(None),
):
    """Iter 250 (user request) — 'Fetch old data'. Opens a 6-hour re-sync
    window for the device: the next handshake answers ATTLOGStamp=0 and a
    one-time CHECK command is issued, making the machine re-upload EVERY
    attendance log stored in its memory. Duplicates are skipped by the
    idempotency guard, so this is safe to run any number of times."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    device = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    until = (datetime.now(timezone.utc) + timedelta(hours=6)).isoformat().replace("+00:00", "Z")
    await db.biometric_devices.update_one(
        {"device_id": device_id},
        {"$set": {"resync_until": until,
                  "resync_check_sent": False,
                  "resync_requested_by": admin["user_id"],
                  "resync_requested_at": _now_iso_z()}},
    )
    return {
        "ok": True,
        "resync_until": until,
        "message": (
            "Old-data fetch started. Keep the machine powered ON and "
            "connected to the internet — it will re-upload all stored punches "
            "within the next few minutes (large logs can take longer). "
            "Already-imported punches are skipped automatically."
        ),
    }


@router.post("/biometric/devices/{device_id}/fetch-range")
async def fetch_range_biometric_device(
    device_id: str,
    range: str = Query("current_month",
                       description="current_month | last_month"),
    authorization: Optional[str] = Header(None),
):
    """Iter 473 (user request — 'fetch at least the current month or one
    month back'): queues a targeted ``DATA QUERY ATTLOG StartTime=…
    EndTime=…`` push-protocol command so the machine re-uploads ONLY the
    chosen month's punches (much lighter than the full 'Fetch old data'
    re-sync). Duplicates are skipped by the idempotency guard."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    device = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    # "today" in the MACHINE's configured timezone so month edges are right
    mins = _parse_gmt_offset_minutes(device.get("gmt_offset"))
    today = (datetime.now(timezone.utc) + timedelta(minutes=mins)).date()
    if str(range).strip().lower() == "last_month":
        first_this = today.replace(day=1)
        end = first_this - timedelta(days=1)
        start = end.replace(day=1)
        label_month = start.strftime("%B %Y")
    else:
        start = today.replace(day=1)
        end = today
        label_month = start.strftime("%B %Y")
    cmd = (f"DATA QUERY ATTLOG StartTime={start.isoformat()} 00:00:00\t"
           f"EndTime={end.isoformat()} 23:59:59")
    cmd_id = await _queue_cmd(
        device["serial_number"], cmd, admin["user_id"],
        f"Fetch punches — {label_month}")
    # a CHECK right after makes some firmwares flush pending uploads faster
    await _queue_cmd(device["serial_number"], "CHECK",
                     admin["user_id"], "Flush after month fetch")
    return {
        "ok": True,
        "cmd_id": cmd_id,
        "from": start.isoformat(),
        "to": end.isoformat(),
        "message": (
            f"Fetch queued for {label_month} ({start.strftime('%d-%m-%Y')} to "
            f"{end.strftime('%d-%m-%Y')}). Keep the machine ON and online — "
            "it picks the command up within ~1 minute and re-uploads that "
            "month's punches. Already-imported punches are skipped. "
            "If nothing arrives in 10 minutes your firmware may not support "
            "month queries — use 'Fetch old data' instead (full re-upload)."
        ),
    }


@router.post("/biometric/devices/resync-all")
async def resync_all_biometric_devices(
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Iter 252 (user request) — one-tap 'sync machine punches up to date'
    from the Punch Log Report: opens a re-sync window on EVERY enabled
    device in scope so each machine re-uploads all stored punches."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {"enabled": {"$ne": False}}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    devices = await db.biometric_devices.find(q, {"_id": 0}).to_list(200)
    if not devices:
        raise HTTPException(status_code=404, detail="No biometric machines registered for this scope")
    until = (datetime.now(timezone.utc) + timedelta(hours=6)).isoformat().replace("+00:00", "Z")
    now = datetime.now(timezone.utc)
    online = 0
    for d in devices:
        last = d.get("last_seen_at")
        try:
            if last and (now - datetime.fromisoformat(str(last).replace("Z", "+00:00"))).total_seconds() < 180:
                online += 1
        except Exception:
            pass
        await db.biometric_devices.update_one(
            {"device_id": d["device_id"]},
            {"$set": {"resync_until": until,
                      "resync_check_sent": False,
                      "resync_requested_by": admin["user_id"],
                      "resync_requested_at": _now_iso_z()}},
        )
    return {
        "ok": True,
        "devices": len(devices),
        "online": online,
        "message": (
            f"Sync requested on {len(devices)} machine(s) ({online} online now). "
            "Each ONLINE machine will re-upload every stored punch within a few "
            "minutes — press Apply to refresh. Offline machines will sync as soon "
            "as they come online (window open for 6 hours)."
        ),
    }


@router.get("/biometric/connection-doctor")
async def biometric_connection_doctor(
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Iter 297 (user issue: "machine not connecting") — one-click
    diagnosis. For every registered device: ONLINE / OFFLINE /
    NEVER-CONNECTED verdict with plain-language advice. Also lists
    UNKNOWN serials that DID reach the server but are not registered
    (catches serial typos instantly, with a fuzzy-match hint)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    devices = await db.biometric_devices.find(q, {"_id": 0}).to_list(200)
    now = datetime.now(timezone.utc)

    def _ago(iso: Optional[str]) -> Optional[float]:
        if not iso:
            return None
        try:
            return (now - datetime.fromisoformat(iso.replace("Z", "+00:00"))).total_seconds()
        except Exception:
            return None

    out = []
    for d in devices:
        secs = _ago(d.get("last_seen_at"))
        if secs is None:
            verdict, advice = "never_connected", (
                "This machine has NEVER reached the server. The problem is on "
                "the machine / network side: on the device open COMM → ADMS "
                "(Cloud Server) and re-check Server Mode=ADMS, Server Address, "
                "Port 80, Proxy OFF (see Setup Guide), confirm the serial shown "
                "on the machine (Menu → System Info) matches what you registered, "
                "make sure the machine has working internet (gateway + DNS), "
                "then RESTART the machine."
            )
        elif secs < 180:
            verdict, advice = "online", "Connected ✓ — the machine is talking to the server."
        else:
            verdict, advice = "offline", (
                "The machine HAS connected before but is offline now — check "
                "its power and internet connection, then restart it."
            )
        out.append({
            "device_id": d.get("device_id"),
            "serial_number": d.get("serial_number"),
            "name": d.get("name"),
            "brand": d.get("brand") or "zkteco",
            "company_id": d.get("company_id"),
            "enabled": d.get("enabled", True),
            "verdict": verdict,
            "advice": advice,
            "last_seen_at": d.get("last_seen_at"),
            "last_seen_secs": int(secs) if secs is not None else None,
            "last_source_ip": d.get("last_source_ip"),
        })

    # Unknown serials — machines that reached the server but aren't
    # registered (or were registered with a slightly different serial).
    unknown = await _unknown_devices_payload()
    return {"devices": out, "unknown_devices": unknown, "checked_at": _now_iso_z()}


def _is_probe_serial(sn: str) -> bool:
    """Iter 608 — serials generated by our own connectivity checks and
    deploy verifications (never real ZKTeco machines)."""
    s = (sn or "").strip().upper()
    return (s.startswith("PROBE-") or s.startswith("TEST-")
            or s in ("TEST123", "TEST", "DEPLOY-CHECK", "HEALTHCHECK"))


async def _unknown_devices_payload() -> list:
    """Iter 473 (user report — 'new machines not showing in the list'):
    serials that reached the server but are NOT registered, with a typo
    hint when a registered serial is close. Shown on the Machine List
    itself (previously only inside the Connection Doctor).
    Iter 608 (user report — 'why testing machines show on live server'):
    our own internal health-check probes (PROBE-*, TEST-DEPLOY-CHECK,
    TEST123…) are auto-hidden, and any entry can be dismissed by admins."""
    registered = [
        str(x.get("serial_number") or "")
        for x in await db.biometric_devices.find(
            {}, {"_id": 0, "serial_number": 1}).to_list(500)
    ]
    out = []
    _docs = await db.biometric_unknown.find(
        {}, {"_id": 0}).sort("last_seen_at", -1).to_list(50)
    for u in _docs:
        sn = str(u.get("serial_number") or "")
        if sn in registered:
            continue  # registered since — no longer a problem
        if u.get("dismissed"):
            continue  # Iter 608 — admin dismissed this entry
        if _is_probe_serial(sn):
            continue  # Iter 608 — internal test/deploy probe, never a machine
        hint = None
        close = difflib.get_close_matches(sn, registered, n=1, cutoff=0.75)
        if close:
            hint = (
                f"Looks like a TYPO: the machine reports '{sn}' but you "
                f"registered '{close[0]}'. Edit the device and fix the serial."
            )
        out.append({
            "serial_number": sn,
            "first_seen_at": u.get("first_seen_at"),
            "last_seen_at": u.get("last_seen_at"),
            "hits": u.get("hits") or 0,
            "hint": hint or (
                "This machine IS reaching the server but is NOT registered. "
                "Register it with this exact serial to start receiving punches."
            ),
        })
    return out


@router.post("/biometric/unknown/dismiss")
async def dismiss_unknown_device(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    """Iter 608 — hide a stray/unwanted serial from 'New machines detected'.
    If the machine ever pings again AFTER dismissal, it reappears (the
    ingest path clears the flag on a fresh hit from a real device)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    sn = str(payload.get("serial_number") or "").strip()
    if not sn:
        raise HTTPException(status_code=400, detail="serial_number required")
    r = await db.biometric_unknown.update_one(
        {"serial_number": sn},
        {"$set": {"dismissed": True, "dismissed_at": _now_iso_z(),
                  "dismissed_by": admin.get("email") or admin.get("user_id")}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Serial not found")
    return {"ok": True}


@router.get("/biometric/devices")
async def list_biometric_devices(
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    devices = await db.biometric_devices.find(q, {"_id": 0}).sort("created_at", -1).to_list(200)
    # Attach freshness — "online" if seen in the last 3 minutes
    now = datetime.now(timezone.utc)
    for d in devices:
        last = d.get("last_seen_at")
        online = False
        if last:
            try:
                lt = datetime.fromisoformat(last.replace("Z", "+00:00"))
                online = (now - lt).total_seconds() < 180
            except Exception:
                online = False
        d["online"] = online
    unmapped = await db.biometric_unmapped.count_documents({}) if devices else 0
    # Iter 473 — surface NEW (unregistered) machines on the list itself.
    unknown = await _unknown_devices_payload()
    return {"devices": devices, "unmapped_count": unmapped,
            "unknown_devices": unknown}


@router.patch("/biometric/devices/{device_id}")
async def update_biometric_device(
    device_id: str,
    payload: BiometricDeviceUpdate,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    device = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    updates = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")
    if "company_id" in updates and admin["role"] == "company_admin":
        updates.pop("company_id")  # company_admin can't move devices between firms
    await db.biometric_devices.update_one({"device_id": device_id}, {"$set": updates})
    updated = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    return {"ok": True, "device": updated}


@router.delete("/biometric/devices/{device_id}")
async def delete_biometric_device(
    device_id: str,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    device = await db.biometric_devices.find_one({"device_id": device_id})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    await db.biometric_devices.delete_one({"device_id": device_id})
    return {"ok": True}


@router.get("/biometric/devices/{device_id}/logs")
async def biometric_device_logs(
    device_id: str,
    limit: int = Query(50, ge=1, le=500),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    device = await db.biometric_devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    logs = await db.attendance.find(
        {"device_serial": device["serial_number"]},
        {"_id": 0, "selfie_base64": 0},
    ).sort("at", -1).to_list(limit)
    return {"device": device, "logs": logs}


@router.get("/biometric/unmapped")
async def biometric_unmapped_punches(
    limit: int = Query(100, ge=1, le=500),
    authorization: Optional[str] = Header(None),
):
    """Punches that arrived from a device but couldn't be mapped to any user
    (bio_code / employee_code not yet set). Admin uses this list to enrol
    workers on the mobile app or add the missing bio_code."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {}
    if admin["role"] == "company_admin":
        # scope by devices belonging to this company
        my_sns = [d["serial_number"] async for d in db.biometric_devices.find(
            {"company_id": admin["company_id"]}, {"_id": 0, "serial_number": 1}
        )]
        q["device_serial"] = {"$in": my_sns}
    logs = await db.biometric_unmapped.find(q, {"_id": 0}).sort("seen_at", -1).to_list(limit)
    return {"unmapped": logs}


@router.post("/biometric/smart-direction-repair")
async def smart_direction_repair(
    payload: dict = Body(...),
    authorization: Optional[str] = Header(None),
):
    """Iter 518 (user choice C) — repair PAST days broken by the down OUT
    machine: for each employee-day in the range that has 2+ machine IN
    punches from an IN-kind device and NO machine OUT, the LAST IN punch
    (>= gap hours after the first) is converted to the OUT, and that day's
    synthetic server_auto_close OUT records are removed."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = str(payload.get("company_id") or "").strip()
    d_from = str(payload.get("from_date") or "").strip()
    d_to = str(payload.get("to_date") or "").strip()
    dry_run = bool(payload.get("dry_run"))
    if not cid or not d_from or not d_to:
        raise HTTPException(status_code=400, detail="company_id, from_date and to_date are required")
    if admin["role"] == "company_admin" and cid != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this firm")
    comp = await db.companies.find_one({"company_id": cid}, {"_id": 0, "attendance_config": 1}) or {}
    gap_h = float((comp.get("attendance_config") or {}).get("smart_direction_gap_hrs") or 4)
    in_serials = [d["serial_number"] async for d in db.biometric_devices.find(
        {"company_id": cid, "kind": "in"}, {"_id": 0, "serial_number": 1})]
    if not in_serials:
        raise HTTPException(status_code=400, detail="This firm has no IN-kind machine")
    srcs = [f"zkteco:{sn}" for sn in in_serials]
    # group machine punches per user-day
    pipeline = [
        {"$match": {"company_id": cid, "date": {"$gte": d_from, "$lte": d_to},
                    "status": "approved", "kind": {"$in": ["in", "out"]},
                    "source": {"$regex": "^zkteco:"}}},
        {"$group": {"_id": {"u": "$user_id", "d": "$date"},
                    "punches": {"$push": {"record_id": "$record_id", "at": "$at",
                                          "kind": "$kind", "source": "$source"}}}},
    ]
    fixed = skipped = 0
    async for grp in db.attendance.aggregate(pipeline):
        punches = sorted(grp["punches"], key=lambda p: p["at"])
        outs = [p for p in punches if p["kind"] == "out"]
        ins = [p for p in punches if p["kind"] == "in" and p["source"] in srcs]
        if outs or len(ins) < 2:
            skipped += 1
            continue
        first, last = ins[0], ins[-1]
        try:
            _f = datetime.fromisoformat(str(first["at"]).replace("Z", "+00:00"))
            _l = datetime.fromisoformat(str(last["at"]).replace("Z", "+00:00"))
            if (_l - _f).total_seconds() < gap_h * 3600:
                skipped += 1
                continue
        except ValueError:
            skipped += 1
            continue
        fixed += 1
        if dry_run:
            continue
        await db.attendance.update_one(
            {"record_id": last["record_id"]},
            {"$set": {"kind": "out", "direction_corrected": True,
                      "direction_corrected_at": _now_iso_z(),
                      "direction_corrected_by": admin["user_id"]}})
        await db.attendance.delete_many({
            "user_id": grp["_id"]["u"], "date": grp["_id"]["d"],
            "source": "server_auto_close"})
    logger.info("[smart-direction-repair] firm=%s %s..%s fixed=%d skipped=%d dry=%s",
                cid, d_from, d_to, fixed, skipped, dry_run)
    return {"ok": True, "fixed_days": fixed, "skipped_days": skipped,
            "gap_hours": gap_h, "dry_run": dry_run}


@router.post("/biometric/create-master-from-pin")
async def create_master_from_pin(
    payload: dict = Body(...),
    authorization: Optional[str] = Header(None),
):
    """Iter 514 (user request) — ONE-TAP Employee Master creation from an
    unmapped machine PIN. Creates a minimal, already-approved employee
    (bio_code = machine PIN, name pre-filled from the machine's own user
    table when captured) under the device's firm, then RE-MAPS all of that
    PIN's parked unmapped punches into attendance. The admin completes the
    remaining master details (phone, salary, DOJ…) in Employee Master."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    sn = str(payload.get("device_serial") or "").strip()
    pin = str(payload.get("device_user_id") or "").strip()
    if not sn or not pin:
        raise HTTPException(status_code=400, detail="device_serial and device_user_id are required")
    device = await db.biometric_devices.find_one({"serial_number": sn}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail=f"Device {sn} is not registered")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    cid = device.get("company_id")
    if not cid:
        raise HTTPException(status_code=400, detail="Device has no company — assign it to a firm first")

    created = False
    user = await _match_employee_for_bio(pin, cid)
    if not user:
        # Name: prefer the machine's own user table (USERINFO capture).
        mu = await db.biometric_machine_users.find_one(
            {"company_id": cid, "pin": pin}, {"_id": 0, "name": 1})
        name = (str(payload.get("name") or "").strip()
                or (mu or {}).get("name") or f"EMP {pin}").strip().upper()
        emp_code = (str(payload.get("employee_code") or "").strip() or pin)
        if await db.users.find_one({"company_id": cid, "employee_code": emp_code},
                                   {"_id": 0, "user_id": 1}):
            emp_code = None  # code taken — bio_code still maps the punches
        from server import now_iso  # same helper employees_admin uses
        user = {
            "user_id": f"user_{uuid.uuid4().hex[:12]}",
            "email": None, "phone": None, "picture": None,
            "name": name, "role": "employee", "company_id": cid,
            "employee_code": emp_code, "bio_code": pin,
            "is_onroll": True, "pay_mode": "Bank",
            "onboarded": True, "onboarded_at": now_iso(),
            "approval_status": "approved", "approved_at": now_iso(),
            "approved_by": admin.get("user_id"),
            "has_pin": False,
            "created_at": now_iso(),
            "created_by_admin": admin.get("user_id"),
            "created_from_machine_pin": sn,
        }
        await db.users.insert_one(dict(user))
        user.pop("_id", None)
        created = True

    # Re-map ALL parked punches for this PIN across the firm's machines.
    firm_devices = {
        d["serial_number"]: d
        async for d in db.biometric_devices.find({"company_id": cid}, {"_id": 0})
    }
    remapped = 0
    failed = 0
    unmapped_rows = await db.biometric_unmapped.find({
        "device_user_id": pin,
        "device_serial": {"$in": list(firm_devices.keys())},
    }).sort("seen_at", 1).to_list(5000)
    for doc in unmapped_rows:
        dev = firm_devices.get(doc.get("device_serial"))
        if not dev:
            failed += 1
            continue
        ok, _reason = await _ingest_attlog_line(doc.get("raw") or "", dev)
        if ok:
            await db.biometric_unmapped.delete_one({"_id": doc["_id"]})
            remapped += 1
        else:
            failed += 1
    logger.info("[create-master-from-pin] pin=%s firm=%s created=%s remapped=%d failed=%d",
                pin, cid, created, remapped, failed)
    return {"ok": True, "created": created, "user_id": user["user_id"],
            "name": user.get("name"), "employee_code": user.get("employee_code"),
            "bio_code": pin, "remapped": remapped, "still_unmapped": failed}


@router.post("/biometric/remap-unmapped")
async def biometric_remap_unmapped(
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Iter 93 — Re-map previously-unmapped device punches after an admin
    fixes/updates an employee's bio code in the Employee Master. Each
    stored raw ATTLOG line is re-matched against the CURRENT bio_code /
    employee_code mapping; matches are ingested as normal attendance
    records and removed from the unmapped queue."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    dev_q: dict = {}
    if admin["role"] == "company_admin":
        dev_q["company_id"] = admin["company_id"]
    elif company_id:
        dev_q["company_id"] = company_id
    devices = {
        d["serial_number"]: d
        async for d in db.biometric_devices.find(dev_q, {"_id": 0})
    }
    q: dict = {}
    if dev_q:  # scope unmapped punches to this firm's devices
        q["device_serial"] = {"$in": list(devices.keys())}
    unmapped = await db.biometric_unmapped.find(q).sort("seen_at", 1).to_list(5000)

    checked = len(unmapped)
    remapped = 0
    still_unmapped = 0
    for doc in unmapped:
        device = devices.get(doc.get("device_serial"))
        if not device:
            device = await db.biometric_devices.find_one(
                {"serial_number": doc.get("device_serial")}, {"_id": 0},
            )
        if not device:
            still_unmapped += 1
            continue
        user = await _match_employee_for_bio(
            doc.get("device_user_id"), device.get("company_id"),
        )
        if not user:
            still_unmapped += 1
            continue
        # Matched now → remove from queue FIRST (ingest re-queues on miss),
        # then run the standard ingest path for dedupe + record shape.
        await db.biometric_unmapped.delete_one({"_id": doc["_id"]})
        ok, _reason = await _ingest_attlog_line(doc.get("raw") or "", device)
        if ok:
            remapped += 1
        else:
            still_unmapped += 1

    # Iter 93 — ALSO re-read every stored .dat import for this scope so
    # punches that were skipped as "unmapped" during the original upload
    # are recovered once the bio code exists. Re-running is idempotent:
    # import_zk_dat_bytes dedupes on (user, at, kind, source_tag).
    from utils.zk_dat_import import import_zk_dat_bytes
    dat_q: dict = {}
    if admin["role"] == "company_admin":
        dat_q["company_id"] = admin["company_id"]
    elif company_id:
        dat_q["company_id"] = company_id
    dat_imports = await db.zk_dat_imports.find(dat_q).sort("uploaded_at", -1).to_list(20)
    dat_files_reread = 0
    dat_recovered = 0
    for imp in dat_imports:
        try:
            stats = await import_zk_dat_bytes(
                db,
                company_id=imp["company_id"],
                in_bytes=(imp.get("in_text") or "").encode() or None,
                out_bytes=(imp.get("out_text") or "").encode() or None,
                combined_bytes=(imp.get("combined_text") or "").encode() or None,
                from_date=imp.get("from_date"),
                to_date=imp.get("to_date"),
                source_tag=imp.get("source_tag"),  # SAME tag → dedupe works
            )
            dat_files_reread += 1
            dat_recovered += int(stats.get("inserted") or 0)
            await db.zk_dat_imports.update_one(
                {"_id": imp["_id"]},
                {"$set": {
                    "last_reread_at": _now_iso_z(),
                    "last_stats": {k: v for k, v in stats.items() if k != "unmapped_bio_codes"},
                }},
            )
        except Exception as exc:
            logger.warning("[remap] .dat re-read failed for %s: %s", imp.get("import_id"), exc)

    return {
        "ok": True,
        "checked": checked,
        "remapped": remapped,
        "still_unmapped": still_unmapped,
        "dat_files_reread": dat_files_reread,
        "dat_recovered": dat_recovered,
    }


@router.post("/biometric/devices/simulate-punch")
async def biometric_simulate_punch(
    payload: dict = Body(...),
    authorization: Optional[str] = Header(None),
):
    """Dev / QA helper — creates a synthetic ATTLOG line for a registered
    device so admins can rehearse the end-to-end flow without a physical
    machine present."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    sn = (payload.get("serial_number") or "").strip()
    device_user_id = (payload.get("device_user_id") or "").strip()
    if not sn or not device_user_id:
        raise HTTPException(status_code=400, detail="serial_number and device_user_id are required")
    device = await db.biometric_devices.find_one({"serial_number": sn}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not registered")
    if admin["role"] == "company_admin" and device.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised for this device")
    # Craft an ATTLOG line in the exact format the device pushes.
    ist = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    line = f"{device_user_id}\t{ist.strftime('%Y-%m-%d %H:%M:%S')}\t0\t1\t0\t0"
    ok, reason = await _ingest_attlog_line(line, device)
    return {"ok": ok, "reason": reason, "line": line}


# ---------------------------------------------------------------------------
# Iter 96b — System Health: biometric last-sync summary for the dashboard.
# ---------------------------------------------------------------------------

@router.get("/admin/system-health/biometric")
async def biometric_system_health(
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Last biometric sync info for the dashboard badge.

    Combines the newest .dat import upload, the newest live-device
    heartbeat (ADMS ``last_seen_at``) and the newest biometric-sourced
    punch record. ``status``: ok (<24h), warn (<48h), stale (older/never).
    """
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    if admin.get("role") == "company_admin":
        company_id = admin.get("company_id")
    q = {"company_id": company_id} if company_id else {}

    last_import = await db.zk_dat_imports.find_one(
        q, {"_id": 0, "uploaded_at": 1}, sort=[("uploaded_at", -1)],
    )
    dev_q = dict(q)
    last_device = await db.biometric_devices.find_one(
        dev_q, {"_id": 0, "last_seen_at": 1, "name": 1, "serial": 1},
        sort=[("last_seen_at", -1)],
    )
    punch_q = {**q, "source": {"$regex": "^(import|zkteco|bio)"}}
    last_punch = await db.attendance.find_one(
        punch_q, {"_id": 0, "created_at": 1, "at": 1}, sort=[("created_at", -1)],
    )

    candidates = [
        ("dat_import", (last_import or {}).get("uploaded_at")),
        ("device", (last_device or {}).get("last_seen_at")),
        ("punch", (last_punch or {}).get("created_at")),
    ]
    best_kind, best_at = None, None
    for kind, iso in candidates:
        if not iso:
            continue
        if best_at is None or str(iso) > str(best_at):
            best_kind, best_at = kind, str(iso)

    status = "never"
    hours_ago = None
    if best_at:
        try:
            dt = datetime.fromisoformat(best_at.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            hours_ago = round(
                (datetime.now(timezone.utc) - dt).total_seconds() / 3600.0, 1,
            )
            status = "ok" if hours_ago < 24 else ("warn" if hours_ago < 48 else "stale")
        except Exception:
            status = "unknown"

    return {
        "status": status,
        "last_sync_at": best_at,
        "last_sync_kind": best_kind,   # dat_import | device | punch
        "hours_ago": hours_ago,
        "last_import_at": (last_import or {}).get("uploaded_at"),
        "last_device_seen_at": (last_device or {}).get("last_seen_at"),
        "last_punch_created_at": (last_punch or {}).get("created_at"),
        "devices_registered": await db.biometric_devices.count_documents(dev_q),
    }


# ---------------------------------------------------------------------------
# Iter 259 — Device OFFLINE alerts + Device Health Report (Excel).
# ---------------------------------------------------------------------------
OFFLINE_ALERT_AFTER_MIN = 15


async def device_offline_alert_loop():
    """Background loop (every 5 min): a device silent > 15 min raises ONE
    admin notification (company admins + super admins); coming back online
    resets the flag so a future outage alerts again."""
    while True:
        try:
            now = datetime.now(timezone.utc)
            async for d in db.biometric_devices.find(
                {"enabled": {"$ne": False}}, {"_id": 0}
            ):
                last = d.get("last_seen_at")
                offline = True
                if last:
                    try:
                        dt = datetime.fromisoformat(str(last).replace("Z", "+00:00"))
                        offline = (now - dt).total_seconds() > OFFLINE_ALERT_AFTER_MIN * 60
                    except ValueError:
                        pass
                if offline and last and not d.get("offline_alerted"):
                    await db.biometric_devices.update_one(
                        {"device_id": d["device_id"]},
                        {"$set": {"offline_alerted": True,
                                  "offline_alerted_at": _now_iso_z()}},
                    )
                    for audience, cid in (("admins", d.get("company_id")),
                                          ("super_admins", None)):
                        await db.notifications.insert_one({
                            "notification_id": f"n_{uuid.uuid4().hex[:10]}",
                            "company_id": cid,
                            "audience": audience,
                            "type": "device.offline",
                            "title": f"⚠️ Machine OFFLINE — {d.get('name') or d.get('serial_number')}",
                            "body": (
                                f"Biometric machine '{d.get('name')}' (SN {d.get('serial_number')}) "
                                f"has been offline for over {OFFLINE_ALERT_AFTER_MIN} minutes. "
                                "Punches are NOT syncing — check its power and internet."
                            ),
                            "created_at": _now_iso_z(),
                            "created_by": "system",
                        })
                    logger.warning("[zkteco] OFFLINE alert raised for %s", d.get("serial_number"))
                    # Iter 410 (user accepted) — also EMAIL the alert to the
                    # super admins + the firm's company admins so nobody
                    # silently loses punches. Best-effort: any email failure
                    # never breaks the loop (in-app notification above is
                    # already raised).
                    try:
                        from routes.email_notifications import _get_settings, _send_and_log
                        settings = await _get_settings()
                        if settings and settings.get("enabled") and settings.get("username"):
                            recipients: set = set()
                            async for a in db.users.find(
                                {"role": {"$in": ["super_admin", "company_admin"]},
                                 "email": {"$nin": [None, ""]}},
                                {"_id": 0, "email": 1, "role": 1, "company_id": 1},
                            ):
                                if a["role"] == "super_admin" or \
                                        a.get("company_id") == d.get("company_id"):
                                    recipients.add(a["email"])
                            subject = (f"⚠️ Biometric machine OFFLINE — "
                                       f"{d.get('name') or d.get('serial_number')}")
                            body = (
                                f"Biometric machine '{d.get('name')}' "
                                f"(SN {d.get('serial_number')}) has been OFFLINE for over "
                                f"{OFFLINE_ALERT_AFTER_MIN} minutes.\n\n"
                                f"Last seen: {last}\n"
                                "Punches are NOT syncing — please check the machine's "
                                "power and internet connection.\n\n"
                                "You will not receive another email for this outage; "
                                "alerts re-arm automatically once the machine reconnects."
                            )
                            for to in sorted(recipients):
                                await _send_and_log(settings, to, subject, body,
                                                    "device_offline")
                    except Exception:
                        logger.warning("[zkteco] offline alert EMAIL failed",
                                       exc_info=True)
                elif not offline and d.get("offline_alerted"):
                    await db.biometric_devices.update_one(
                        {"device_id": d["device_id"]},
                        {"$set": {"offline_alerted": False}},
                    )
        except Exception:
            logger.warning("[zkteco] offline alert loop error", exc_info=True)
        await asyncio.sleep(300)


@router.get("/biometric/devices/health-report.xlsx")
async def device_health_report_xlsx(
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Device Health Report — one row per machine with status, heartbeat,
    firmware, counters and last command result."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    devices = await db.biometric_devices.find(q, {"_id": 0}).to_list(200)
    firms = {c["company_id"]: c.get("name", "")
             async for c in db.companies.find({}, {"_id": 0, "company_id": 1, "name": 1})}
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import Response
    wb = Workbook()
    ws = wb.active
    ws.title = "Device Health"
    ws.append(["Firm", "Device", "Serial No", "Direction", "Location",
               "Status", "Last Heartbeat", "Firmware", "Users", "Fingerprints",
               "Logs on Device", "Device IP", "Punches Synced", "Enabled"])
    fill = PatternFill("solid", fgColor="1F4E79")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
    now = datetime.now(timezone.utc)
    for d in devices:
        last = d.get("last_seen_at")
        online = False
        if last:
            try:
                online = (now - datetime.fromisoformat(str(last).replace("Z", "+00:00"))).total_seconds() < 180
            except ValueError:
                pass
        ws.append([
            firms.get(d.get("company_id") or "", ""),
            d.get("name") or "", d.get("serial_number") or "",
            (d.get("kind") or "").upper(), d.get("location") or "",
            "ONLINE" if online else "OFFLINE",
            (str(last).replace("T", " ")[:19] if last else "Never"),
            d.get("firmware") or "", d.get("user_count", ""),
            d.get("fp_count", ""), d.get("att_log_count", ""),
            d.get("device_ip") or "",
            d.get("total_punches_ingested", 0),
            "YES" if d.get("enabled", True) else "NO",
        ])
        ws.cell(row=ws.max_row, column=6).font = Font(
            bold=True, color="16A34A" if online else "DC2626")
    for col, w in zip("ABCDEFGHIJKLMN", (22, 18, 14, 10, 14, 10, 20, 18, 8, 12, 13, 14, 14, 8)):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="device-health-report.xlsx"'},
    )

