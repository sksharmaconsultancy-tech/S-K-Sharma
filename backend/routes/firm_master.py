"""Iter 89 — Route module: Firm Master (Web Portal only).

Comprehensive firm profile with 17 sections migrated from the user's
legacy Windows Firm Master screen:

  * Header: start_date, category, business_nature, emails
  * Registered / Office / Factory addresses
  * Allowance & Deduction checklists (fixed labels — reusable across firms)
  * Bank details
  * Firm settings (salary structure, toggles)
  * Contact persons (repeatable)
  * Salary process settings + CL/PL + EPF + ESI + Bonus + Report Order
  * Compliance documents grid (13 fixed rows)
  * Portal login credentials grid (5 fixed rows — passwords encrypted at
    rest via ``fernet`` when the key is available; otherwise stored
    plaintext for MVP so an ops admin can re-enter them later.)

Endpoints:
  * GET   /api/admin/firm-master/{company_id}      - Read profile
  * PATCH /api/admin/firm-master/{company_id}      - Upsert profile

The frontend renders the whole thing as a single scrollable form; the
backend does not gate individual fields with validation errors — it
accepts partial payloads and merges into the persisted subdocument so
Save works section-by-section without forcing all-or-nothing input.
"""
from datetime import datetime, timezone
import uuid
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Request

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
    require_super_admin_strict,
    now_iso,
    logger,
)


router = APIRouter(prefix="/api/admin", tags=["firm-master"])


# ---------------------------------------------------------------------------
# Fixed catalogs — mirror the exact labels from the legacy screen. Storing
# them as constants (not per-firm records) means every firm inherits the
# same rows automatically, and adding a new label later is a single-file
# code change instead of a data migration.
# ---------------------------------------------------------------------------
ALLOWANCE_LABELS: List[str] = [
    "HRA", "CONV.", "OTH. ALLOW.", "OVER TIME", "INCENTIVE",
    "OTHER MISC.ALLOWANCE", "BONUS", "MEDICAL ALLOWANCES",
    "FOOD ALLOWANCES", "GRATUITY", "LEAVE", "DA",
]

DEDUCTION_LABELS: List[str] = [
    "PF", "ESI", "I. TAX", "TDS", "OTH. DEDUC.",
    "ADVANCE", "UNIFORM", "CLUB", "CANTEEN", "PT",
]

COMPLIANCE_DOC_LABELS: List[str] = [
    "TIN REG. NO.", "FIRM REG. NO.", "POWER CONN. NO.", "E.S.I NO.",
    "SHOP ACT. REG. NO.", "COMPANY PAN NO.", "TAN NO.", "EPF CODE NO.",
    "INCORPORATION CER. NO.", "FACTORY & BOILER LICENCE NO.",
    "LIN NO.", "DIGITAL SIGNATURE", "LABOURE LICENCE",
]

PORTAL_LOGIN_LABELS: List[str] = [
    "PF LOGIN", "ESI Login", "SSO Login", "Rajfeb Login", "PT Login",
]

SALARY_STRUCTURES: List[str] = [
    "Standard Monthly",
    "Per-Day Wages",
    "Piece-Rate",
    "Contractual",
    "Mixed",
]

REPORT_ORDER_OPTIONS: List[str] = [
    "Employee Code",
    "Name (A → Z)",
    "Designation",
    "Department",
    "Date of Joining",
]


def _empty_master(company_id: str, company_name: str = "") -> Dict[str, Any]:
    """Return a fully-populated default firm-master doc so the frontend
    can render every section even for a brand-new firm."""
    return {
        "company_id": company_id,
        "company_name": company_name,
        # Iter 89 — Firm logo (base64 PNG/JPEG data URL). Synced to
        # ``companies.logo_base64`` on save so the admin shell + mobile
        # app can render it without a second fetch.
        "logo": {
            "image_base64": None,
            "mime_type": None,
        },
        "header": {
            "start_date": None,
            "category": None,
            "business_nature": None,
            "email_1": None,
            "email_2": None,
        },
        "registered_address": {
            "address1": None, "address2": None,
            "city": None, "state": None, "pin_code": None,
        },
        "office_address": {
            "same_as_firm": True,
            "address1": None, "address2": None,
            "city": None, "state": None, "pin_code": None,
        },
        "factory_address": {
            "same_as_firm": True,
            "address1": None, "address2": None,
            "city": None, "state": None, "pin_code": None,
        },
        # Fixed allowance/deduction catalogs — value is bool enabled flag.
        "allowances": {label: False for label in ALLOWANCE_LABELS},
        "deductions": {label: False for label in DEDUCTION_LABELS},
        "bank": {
            "account_no": None, "account_name": None,
            "bank_name": None, "branch_name": None, "ifsc": None,
        },
        "settings": {
            "salary_structure": None,
            "reference_by": None,
            "firm_active": True,
            "whatsapp_enable": False,
            "auto_email_process": False,
            "email_enable": False,
            "allow_category_rate": False,
            # Iter 158 — when ON, Employee Code is ALWAYS auto-assigned and
            # the manual code field is locked in Add/Edit Employee.
            "auto_employee_code": False,
        },
        "contact_persons": [],  # {name, mobile, position}
        "salary_process": {
            "online_salary": True,  # Iter 114 — Compliance salary is DEFAULT for every firm
            "offline_salary": False,
            "bio_matrix_attendance": False,
            "gratuity_applicable": False,
            # Iter 142 — firm-wide OT gate. False = NO overtime is
            # calculated for ANY employee of this firm.
            "ot_allowed": True,
            "online_process_days": 0,
            "offline_process_days": 0,
            # Iter 337 (user request) — Days Calculation Method for the
            # Compliance Salary import: attendance | gross_based |
            # freeze_based | fixed | attendance_gross_validation.
            "days_calc_method": "attendance",
            "days_calc_fixed": 26,       # used when method = fixed
            # Iter 337 (user) — Compliance Days round to HALF day (0.5) or
            # FULL day (1) only.
            "days_calc_rounding": 0.5,
            # Iter 338 (user) — frozen Compliance gross becomes the ACTUAL
            # Salary gross for On-Roll employees.
            "freeze_to_actual": False,
            # Iter 763 — Dummy Shift Report toggle (see firm-master.tsx).
            "dummy_shift_report": False,
            # Iter 764 (user request) — SALARY DAYS SOURCE POLICY (mutually
            # exclusive — only ONE active at a time):
            #   "biometric"       → Policy 1: biometric punches (8h + OT)
            #   "imported_sheet"  → Policy 2: Excel imported sheet
            #   "manual_editable" → Policy 3: Monthly Attendance Editable
            # None/"" = legacy (no enforcement; Policy 1 & 2 both allowed).
            "attendance_source": None,
        },
        "leave_policy": {
            "cl_pl_applicable": False,
            "cl_day_limit": 0,
            "pl_day_limit": 0,
        },
        "epf": {
            "applicable": False,
            "applicable_date": None,
            "edli_applicable": False,
            "epf_no": None,
            "group_policy_no": None,
            "epf_user_id": None,
            "epf_password": None,
        },
        "esi": {
            "applicable": False,
            "applicable_date": None,
            "esi_rate": 1,
            "esi_no": None,
            "esi_user_id": None,
            "esi_password": None,
        },
        "bonus": {
            "monthly_bonus": False,
            "gross_mode": None,   # "including" | "excluding"
            "overtime_in_report": False,
            "days_mode": None,    # "fix" | "custom"
            "custom_days": None,
        },
        "report_order": {
            "staff": None,
            "labour": None,
            "other": None,
        },
        "compliance_docs": [
            {"description": label, "number": None,
             "issue_date": None, "expiry_date": None}
            for label in COMPLIANCE_DOC_LABELS
        ],
        "portal_logins": [
            {"login_type": label, "user_name": None,
             "password": None, "unit_location": None, "login_url": None}
            for label in PORTAL_LOGIN_LABELS
        ],
        "updated_at": None,
        "updated_by": None,
    }


def _merge_master(existing: Dict[str, Any], patch: Dict[str, Any]) -> Dict[str, Any]:
    """Deep-merge helper — top-level sections are replaced wholesale if
    provided, but nested keys are merged so a Save on just one section
    doesn't wipe the rest of the profile."""
    merged: Dict[str, Any] = dict(existing)
    for k, v in (patch or {}).items():
        if k in ("company_id", "company_name"):
            continue  # never overwrite the identity keys
        if isinstance(v, dict) and isinstance(merged.get(k), dict):
            merged[k] = {**merged[k], **v}
        else:
            merged[k] = v
    return merged


def _ensure_catalogs(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Guarantee the fixed catalog labels are always present even if a
    legacy master was saved before a new label was added."""
    doc.setdefault("allowances", {})
    for lab in ALLOWANCE_LABELS:
        doc["allowances"].setdefault(lab, False)
    doc.setdefault("deductions", {})
    for lab in DEDUCTION_LABELS:
        doc["deductions"].setdefault(lab, False)

    doc.setdefault("compliance_docs", [])
    existing_desc = {r.get("description") for r in doc["compliance_docs"]}
    for lab in COMPLIANCE_DOC_LABELS:
        if lab not in existing_desc:
            doc["compliance_docs"].append({
                "description": lab, "number": None,
                "issue_date": None, "expiry_date": None,
            })

    doc.setdefault("portal_logins", [])
    existing_types = {r.get("login_type") for r in doc["portal_logins"]}
    for lab in PORTAL_LOGIN_LABELS:
        if lab not in existing_types:
            doc["portal_logins"].append({
                "login_type": lab, "user_name": None,
                "password": None, "unit_location": None, "login_url": None,
            })
    return doc


# ---------------------------------------------------------------------------
# SEC-003 — portal credential protection helpers.
# ---------------------------------------------------------------------------
_SECRET_SALARY_FIELDS = (("epf", "epf_password"), ("esi", "esi_password"))


def _mask_secrets(doc: Dict[str, Any]) -> None:
    """Replace stored passwords with a mask before returning to the client."""
    from utils.secrets_vault import MASK
    for row in doc.get("portal_logins") or []:
        if row.get("password"):
            row["password"] = MASK
            row["password_set"] = True
        else:
            row["password_set"] = False
    for sec_key, field in _SECRET_SALARY_FIELDS:
        sec = doc.get(sec_key) or {}
        if sec.get(field):
            sec[field] = MASK
            sec[f"{field}_set"] = True
        elif field in sec:
            sec[f"{field}_set"] = False


def _protect_secrets(merged: Dict[str, Any], existing: Dict[str, Any]) -> None:
    """Encrypt any NEW plaintext passwords; when the client sends back the
    mask (unchanged field) keep the previously stored ciphertext."""
    from utils.secrets_vault import MASK, encrypt_secret

    prev_rows = {str(r.get("login_type") or ""): r
                 for r in existing.get("portal_logins") or []}
    for row in merged.get("portal_logins") or []:
        row.pop("password_set", None)
        val = row.get("password")
        if val == MASK:
            prev = prev_rows.get(str(row.get("login_type") or ""), {})
            row["password"] = prev.get("password") or None
        else:
            row["password"] = encrypt_secret(val) if val else None
    for sec_key, field in _SECRET_SALARY_FIELDS:
        sec = merged.get(sec_key)
        if not isinstance(sec, dict):
            continue
        sec.pop(f"{field}_set", None)
        val = sec.get(field)
        if val == MASK:
            sec[field] = (existing.get(sec_key) or {}).get(field) or None
        else:
            sec[field] = encrypt_secret(val) if val else None


async def migrate_portal_secrets() -> int:
    """One-time idempotent migration: encrypt any legacy PLAINTEXT portal
    passwords already sitting in ``firm_masters``. Returns docs updated."""
    from utils.secrets_vault import encrypt_secret, is_encrypted
    changed = 0
    async for doc in db.firm_masters.find(
            {}, {"_id": 0, "company_id": 1, "portal_logins": 1, "epf": 1, "esi": 1}):
        upd: Dict[str, Any] = {}
        rows = doc.get("portal_logins") or []
        if any(r.get("password") and not is_encrypted(r.get("password")) for r in rows):
            for r in rows:
                if r.get("password") and not is_encrypted(r["password"]):
                    r["password"] = encrypt_secret(r["password"])
            upd["portal_logins"] = rows
        for sec_key, field in _SECRET_SALARY_FIELDS:
            sec = doc.get(sec_key) or {}
            if sec.get(field) and not is_encrypted(sec[field]):
                upd[f"{sec_key}.{field}"] = encrypt_secret(sec[field])
        if upd:
            await db.firm_masters.update_one(
                {"company_id": doc["company_id"]}, {"$set": upd})
            changed += 1
    if changed:
        logger.info("[SEC-003] encrypted legacy portal passwords in %s firm master(s)", changed)
    return changed


async def _assert_firm_access(user: Dict[str, Any], company_id: str) -> Dict[str, Any]:
    """Super Admin sees any firm; sub admins any firm in their scope;
    company_admin only their own."""
    if user["role"] not in ("super_admin", "company_admin", "sub_admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    if user["role"] == "company_admin":
        if user.get("company_id") != company_id:
            raise HTTPException(status_code=403, detail="Not your firm")
    if user["role"] == "sub_admin":
        from server import sub_admin_can_touch_company
        if not sub_admin_can_touch_company(user, company_id):
            raise HTTPException(status_code=403, detail="Firm is outside your assigned scope")
    company = await db.companies.find_one({"company_id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Firm not found")
    return company


@router.get("/firm-emails/{company_id}")
async def get_firm_emails(
    company_id: str,
    authorization: Optional[str] = Header(None),
):
    """Iter 440 (user request) — the Firm Master's registered email id(s)
    (header.email_1 / email_2) for the Download / Mail Reports popup."""
    user = await get_user_from_token(authorization)
    await _assert_firm_access(user, company_id)
    doc = await db.firm_masters.find_one(
        {"company_id": company_id}, {"_id": 0, "header": 1})
    hdr = (doc or {}).get("header") or {}
    emails = []
    for k in ("email_1", "email_2"):
        e = str(hdr.get(k) or "").strip()
        if e and "@" in e and e not in emails:
            emails.append(e)
    return {"emails": emails}


@router.get("/firm-master/{company_id}")
async def get_firm_master(
    company_id: str,
    authorization: Optional[str] = Header(None),
):
    user = await get_user_from_token(authorization)
    company = await _assert_firm_access(user, company_id)
    doc = await db.firm_masters.find_one({"company_id": company_id}, {"_id": 0})
    if not doc:
        doc = _empty_master(company_id, company.get("name", ""))
    else:
        doc = _ensure_catalogs(doc)
        doc.setdefault("company_name", company.get("name", ""))
    # SEC-003 — NEVER return portal passwords (masked; has-password flags
    # let the UI show the field as filled).
    _mask_secrets(doc)

    # Iter 107 — Firm Category auto-selects from the business category
    # picked when the firm was created (if the master has none yet).
    hdr = doc.setdefault("header", {})
    if not (hdr.get("category") or "").strip():
        cat = (company.get("business_category") or "").strip()
        sub = (company.get("business_subcategory") or "").strip()
        if cat:
            hdr["category"] = f"{cat.title()}{' — ' + sub.title() if sub else ''}"

    # Iter 484 — "General Information" section (ERP redesign). Seed sensible
    # defaults from the company doc + legacy header so nothing is re-typed.
    gen = doc.setdefault("general", {})
    gen.setdefault("company_name", company.get("name") or "")
    gen.setdefault("company_code", company.get("company_code") or "")
    gen.setdefault("short_name", company.get("short_name") or "")
    gen.setdefault("branch_code", "")
    gen.setdefault("firm_start_date", hdr.get("start_date") or "")
    gen.setdefault("company_category", "")
    gen.setdefault("business_nature", hdr.get("business_nature") or "")
    gen.setdefault("industry_type", hdr.get("category") or "")
    gen.setdefault("establishment_type", "")
    gen.setdefault("organization_type", "")
    gen.setdefault("date_of_incorporation", "")
    gen.setdefault("financial_year", "")
    gen.setdefault("assessment_year", "")
    gen.setdefault("currency", "INR")
    gen.setdefault("timezone", "Asia/Kolkata")
    gen.setdefault("language", "English")
    gen.setdefault(
        "company_status",
        "Active" if (doc.get("settings") or {}).get("firm_active", True) else "Inactive",
    )
    gen.setdefault("color_theme", "")
    # Last-modified metadata for the sticky action bar.
    if doc.get("updated_by"):
        _u = await db.users.find_one(
            {"user_id": doc["updated_by"]}, {"_id": 0, "name": 1, "email": 1})
        doc["updated_by_name"] = (_u or {}).get("name") or (_u or {}).get("email") or doc["updated_by"]

    # Iter 89 — Sections 5 (Allowances) & 6 (Deductions) are now linked
    # to the Masters registry so any custom heads the admin adds via
    # `/admin/masters?type=allowance` etc. appear here too. We merge the
    # legacy fixed labels with any global + firm-scoped custom heads.
    allow_docs = await db.masters.find(
        {"type": "allowance",
         "company_id": {"$in": [company_id, "__global__", None]}},
        {"_id": 0, "name": 1},
    ).sort("name", 1).to_list(500)
    ded_docs = await db.masters.find(
        {"type": "deduction",
         "company_id": {"$in": [company_id, "__global__", None]}},
        {"_id": 0, "name": 1},
    ).sort("name", 1).to_list(500)

    # Preserve order: fixed legacy labels first (they mirror the
    # customer's Windows screen exactly), then any custom heads.
    def _merge(base, extra):
        seen = {x.lower(): True for x in base}
        out = list(base)
        for row in extra:
            n = (row.get("name") or "").strip()
            if n and n.lower() not in seen:
                out.append(n)
                seen[n.lower()] = True
        return out

    allowance_labels = _merge(ALLOWANCE_LABELS, allow_docs)
    deduction_labels = _merge(DEDUCTION_LABELS, ded_docs)

    # Guarantee every label from the merged list has a value in the
    # doc (defaults to False) so the frontend can render both fixed
    # and custom rows uniformly.
    doc.setdefault("allowances", {})
    for lab in allowance_labels:
        doc["allowances"].setdefault(lab, False)
    doc.setdefault("deductions", {})
    for lab in deduction_labels:
        doc["deductions"].setdefault(lab, False)

    return {
        "master": doc,
        "catalogs": {
            "allowance_labels": allowance_labels,
            "deduction_labels": deduction_labels,
            "compliance_doc_labels": COMPLIANCE_DOC_LABELS,
            "portal_login_labels": PORTAL_LOGIN_LABELS,
            "salary_structures": SALARY_STRUCTURES,
            "report_order_options": REPORT_ORDER_OPTIONS,
        },
        "role": user["role"],
    }


@router.patch("/firm-master/{company_id}")
async def upsert_firm_master(
    company_id: str,
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    user = await get_user_from_token(authorization)
    company = await _assert_firm_access(user, company_id)

    existing = await db.firm_masters.find_one(
        {"company_id": company_id}, {"_id": 0},
    )
    if not existing:
        existing = _empty_master(company_id, company.get("name", ""))
    else:
        existing = _ensure_catalogs(existing)

    merged = _merge_master(existing, payload)
    merged["company_id"] = company_id

    # Iter 484 — General Information validation + companies mirror.
    _gen = merged.get("general") or {}
    if _gen:
        if not (_gen.get("company_name") or "").strip():
            raise HTTPException(status_code=400, detail="Company Name is required.")
        _sd = (_gen.get("firm_start_date") or "").strip()
        if _sd:
            try:
                if datetime.fromisoformat(_sd).date() > datetime.now(timezone.utc).date():
                    raise HTTPException(
                        status_code=400,
                        detail="Firm Start Date cannot be in the future.")
            except ValueError:
                pass
        _code = (_gen.get("company_code") or "").strip().upper()
        if _code:
            clash = await db.companies.find_one(
                {"company_code": _code, "company_id": {"$ne": company_id}},
                {"_id": 0, "name": 1})
            if clash:
                raise HTTPException(
                    status_code=400,
                    detail=f"Company Code '{_code}' is already used by {clash.get('name')}.")
            _gen["company_code"] = _code
        # Mirror identity fields onto the companies doc so the rest of the
        # portal (pickers, reports, mobile app) stays in sync.
        _mirror: Dict[str, Any] = {}
        if (_gen.get("company_name") or "").strip() and _gen["company_name"].strip() != company.get("name"):
            _mirror["name"] = _gen["company_name"].strip()
        if _code:
            _mirror["company_code"] = _code
        for _mk in ("short_name", "currency", "timezone", "language", "color_theme"):
            if _gen.get(_mk) is not None:
                _mirror[_mk] = _gen.get(_mk)
        if _gen.get("company_status"):
            _mirror["firm_status"] = _gen["company_status"]
            merged.setdefault("settings", {})["firm_active"] = _gen["company_status"] == "Active"
        if _mirror:
            await db.companies.update_one(
                {"company_id": company_id}, {"$set": _mirror})
            if _mirror.get("name"):
                company["name"] = _mirror["name"]

    merged["company_name"] = company.get("name", "")
    merged["updated_at"] = now_iso()
    # Iter 768 (user request) — 3-POLICY EXCLUSIVITY: selecting Policy 3
    # (Monthly Attendance Editable) auto-switches Policy 1 (Attendance by
    # Duty HRS) and Policy 2 (Present @ 8 HRS) to No in the Attendance
    # Policy — only ONE of the three can ever be Yes.
    _sp_in = payload.get("salary_process") if isinstance(payload, dict) else None
    if isinstance(_sp_in, dict) and \
            _sp_in.get("attendance_source") == "manual_editable":
        await db.companies.update_one(
            {"company_id": company_id},
            {"$set": {
                "attendance_policy.policy_master.attendance_by_duty_hours": False,
                "attendance_policy.policy_master.compliance_present_8hr": False,
            }})
    merged["updated_by"] = user["user_id"]
    # SEC-003 — encrypt new passwords / keep existing ones when the client
    # echoes back the mask.
    _protect_secrets(merged, existing)

    # Iter 694 (user bug ×3 — one firm's EPFO login kept spreading to other
    # firms via browser autofill + Save): DUPLICATE GUARD. A NEW/CHANGED
    # EPFO User ID that is already saved on ANOTHER firm is rejected. Values
    # that were already stored before this save are left alone (the cleanup
    # button in Automation Studio handles those).
    async def _dup_owner(uid: str) -> Optional[str]:
        other = await db.firm_masters.find_one(
            {"company_id": {"$ne": company_id},
             "$or": [
                 {"epf.epf_user_id": uid},
                 {"portal_logins": {"$elemMatch": {
                     "login_type": "PF LOGIN", "user_name": uid}}},
             ]},
            {"_id": 0, "company_id": 1})
        if not other:
            return None
        c = await db.companies.find_one(
            {"company_id": other["company_id"]}, {"_id": 0, "name": 1})
        if not c:
            # Iter 754 (user bug) — the "owner" firm was DELETED but its
            # firm_masters doc survived as an orphan. That's not a real
            # conflict: clean the orphan up and allow this save.
            try:
                await db.firm_masters.delete_one(
                    {"company_id": other["company_id"]})
                logger.info("[firm-master] orphan firm_masters of deleted "
                            "firm %s removed (EPFO dup guard)",
                            other["company_id"])
            except Exception:
                pass
            return None
        return c.get("name") or other["company_id"]

    _new_epf_uid = ((merged.get("epf") or {}).get("epf_user_id") or "").strip()
    _old_epf_uid = ((existing.get("epf") or {}).get("epf_user_id") or "").strip()
    _new_pf_uid = _old_pf_uid = ""
    for _r in merged.get("portal_logins") or []:
        if _r.get("login_type") == "PF LOGIN":
            _new_pf_uid = (_r.get("user_name") or "").strip()
            break
    for _r in existing.get("portal_logins") or []:
        if _r.get("login_type") == "PF LOGIN":
            _old_pf_uid = (_r.get("user_name") or "").strip()
            break
    _check = set()
    if _new_epf_uid and "@" not in _new_epf_uid and _new_epf_uid != _old_epf_uid:
        _check.add(_new_epf_uid)
    if _new_pf_uid and "@" not in _new_pf_uid and _new_pf_uid != _old_pf_uid:
        _check.add(_new_pf_uid)
    for _uid in _check:
        _owner = await _dup_owner(_uid)
        if _owner:
            raise HTTPException(
                status_code=400,
                detail=(f"EPFO User ID '{_uid}' is already saved on firm "
                        f"'{_owner}' — every firm has its own separate EPFO "
                        "login. Your browser may have auto-filled the old "
                        "login here. If THIS firm is the real owner, remove "
                        f"that login from '{_owner}' in Firm Master first."))

    # Iter 98 — CL/PL gate: when "CL PL Applicable" is enabled the allowed
    # number of leaves is MANDATORY.
    _lp = merged.get("leave_policy") or {}
    if _lp.get("cl_pl_applicable"):
        _cl = float(_lp.get("cl_day_limit") or 0)
        _pl = float(_lp.get("pl_day_limit") or 0)
        if _cl <= 0 and _pl <= 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    "CL/PL is enabled — please mention the allowed no. of "
                    "leaves (CL Day Limit and/or PL Day Limit)."
                ),
            )

    await db.firm_masters.update_one(
        {"company_id": company_id},
        {"$set": merged},
        upsert=True,
    )
    # Iter 420 (user rule) — when the firm's Actual Salary Process (Offline
    # Salary) is DISABLED, or Bio Matrix Attendance is ENABLED, default the
    # attendance-policy sub-point "Count Present Day @ 8 HRS — Compliance
    # Salary only" to YES for this firm.
    _sp = merged.get("salary_process") or {}
    if (not _sp.get("offline_salary")) or _sp.get("bio_matrix_attendance"):
        await db.companies.update_one(
            {"company_id": company_id},
            {"$set": {"attendance_policy.policy_master.compliance_present_8hr": True}},
        )
        logger.info("[firm-master] %s — compliance_present_8hr auto-set YES "
                    "(offline_salary=%s, bio_matrix=%s)", company_id,
                    bool(_sp.get("offline_salary")),
                    bool(_sp.get("bio_matrix_attendance")))
    # Iter 483 (user request) — "Auto-approve Mobile App Punches" toggle.
    # Mirror onto the companies doc (the punch endpoint reads it there) and,
    # when switched ON, instantly approve this firm's OLD pending app
    # punches so they show on the Attendance Grid (contractual-employee
    # pending punches keep their own approval contract).
    _auto_app = bool((merged.get("settings") or {}).get("auto_approve_mobile_punches"))
    await db.companies.update_one(
        {"company_id": company_id},
        {"$set": {"auto_approve_mobile_punches": _auto_app}},
    )
    # Iter 503 — SINGLE MACHINE ATTENDANCE MODE (user spec, Message 148):
    # mirror the Firm Master attendance capture config onto the companies
    # doc (the attendance engine reads it there). Values are validated to
    # the allowed sets; anything else falls back to safe defaults.
    _ac_raw = (merged.get("settings") or {}).get("attendance_config")
    if isinstance(_ac_raw, dict):
        _ac_mode = str(_ac_raw.get("device_mode") or "separate")
        if _ac_mode not in ("separate", "single_machine", "mobile", "gps", "qr"):
            _ac_mode = "separate"
        _ac_interp = str(_ac_raw.get("interpretation") or "alternate")
        if _ac_interp not in ("alternate", "first_last"):
            _ac_interp = "alternate"
        try:
            _ac_dup = int(_ac_raw.get("dup_window_min", 5))
        except (TypeError, ValueError):
            _ac_dup = 5
        if _ac_dup not in (0, 1, 2, 5, 10):
            _ac_dup = 5
        _ac_lunch = str(_ac_raw.get("lunch_mode") or "ignore_middle")
        if _ac_lunch not in ("ignore_middle", "actual_break", "fixed"):
            _ac_lunch = "ignore_middle"
        try:
            _ac_lunch_min = int(_ac_raw.get("lunch_fixed_min", 30))
        except (TypeError, ValueError):
            _ac_lunch_min = 30
        if _ac_lunch_min not in (30, 45, 60):
            _ac_lunch_min = 30
        _ac_clean = {
            "device_mode": _ac_mode,
            "interpretation": _ac_interp,
            "dup_window_min": _ac_dup,
            "lunch_mode": _ac_lunch,
            "lunch_fixed_min": _ac_lunch_min,
            # Iter 715 (user spec) — CROSS MIDNIGHT punching, DEFAULT YES.
            # Next-day OUT punches map to the previous day's open
            # night-shift session; only an explicit NO switches it off.
            "cross_midnight": _ac_raw.get("cross_midnight") is not False,
        }
        merged.setdefault("settings", {})["attendance_config"] = _ac_clean
        await db.companies.update_one(
            {"company_id": company_id},
            {"$set": {"attendance_config": _ac_clean}},
        )
        await db.firm_masters.update_one(
            {"company_id": company_id},
            {"$set": {"settings.attendance_config": _ac_clean}},
        )
        logger.info("[firm-master] %s — attendance_config mirrored: %s",
                    company_id, _ac_clean)
    if _auto_app:
        _bulk = await db.attendance.update_many(
            {
                "company_id": company_id,
                "status": "pending",
                "pending_reason": {"$ne": "contractual_employee"},
                "mock_location": {"$ne": True},
            },
            {"$set": {
                "status": "approved",
                "attendance_status": "approved",
                "decision_by": "system:firm-auto-approve",
                "decision_at": now_iso(),
                "decision_reason": "auto-approved (Firm Master: auto-approve mobile app punches)",
            }},
        )
        if _bulk.modified_count:
            logger.info("[firm-master] %s — auto-approved %d old pending punches",
                        company_id, _bulk.modified_count)

    # Iter 89 — Mirror the firm logo onto ``companies.logo_base64`` so
    # the admin shell + mobile app can render it via the standard
    # /api/companies feed without a second lookup.
    logo_b64 = None
    logo_mime = None
    try:
        logo_b64 = (merged.get("logo") or {}).get("image_base64") or None
        logo_mime = (merged.get("logo") or {}).get("mime_type") or None
    except Exception:
        pass
    await db.companies.update_one(
        {"company_id": company_id},
        {"$set": {
            "logo_base64": logo_b64,
            "logo_mime": logo_mime,
            "logo_updated_at": now_iso() if logo_b64 else None,
        }},
    )
    logger.info(
        "[firm-master] %s updated by %s (%s)",
        company_id, user["user_id"], user["role"],
    )
    # Iter 484 — audit trail entry with the section keys that changed.
    from routes.firm_master_v2 import write_fm_audit
    await write_fm_audit(
        company_id, user, "master_saved",
        sorted(k for k in payload.keys() if k not in ("company_id", "company_name")),
    )
    saved = await db.firm_masters.find_one({"company_id": company_id}, {"_id": 0})
    _mask_secrets(saved)
    return {"ok": True, "master": saved}


# ---------------------------------------------------------------------------
# Iter 306 (user #14) — Firms ID & Password (PF / ESIC) vault screen.
# SUPER ADMIN ONLY and gated behind the caller's login PIN: passwords are
# decrypted from the secrets vault only after the PIN checks out.
# ---------------------------------------------------------------------------

def _fm_city(fm: Dict[str, Any]) -> str:
    """Best-effort city from registered → office → factory address."""
    for k in ("registered_address", "office_address", "factory_address"):
        c = str(((fm.get(k) or {}).get("city")) or "").strip()
        if c:
            return c
    return ""


@router.post("/firm-credentials")
async def firm_credentials(
    request: Request,
    payload: Dict[str, Any] = Body(default={}),
    authorization: Optional[str] = Header(None),
):
    user = await get_user_from_token(authorization)
    # Iter 331 (user request) — Sub Super Admins can view firm credentials
    # too (verified against their OWN login PIN).
    require_role(user, ["super_admin", "sub_admin"])

    # Iter 599 (user request) — VAULT ACCESS LOG: every unlock attempt
    # (success AND failure) is recorded for the Super Admin's audit trail.
    async def _log_vault(ok: bool, reason: str) -> None:
        try:
            await db.vault_access_log.insert_one({
                "log_id": f"val_{uuid.uuid4().hex[:12]}",
                "user_id": user["user_id"],
                "name": user.get("name"),
                "email": user.get("email"),
                "role": user.get("role"),
                "ok": ok,
                "reason": reason,
                "ip": (request.client.host if request and request.client else None),
                "at": now_iso(),
            })
        except Exception:
            logger.exception("[vault-log] failed to persist access log row")

    pin = str(payload.get("pin") or "").strip()
    from server import _verify_pin
    me = await db.users.find_one({"user_id": user["user_id"]},
                                 {"_id": 0, "pin_hash": 1})
    # Iter 598 (user report) — Sub Super Admins created WITHOUT a PIN kept
    # seeing a generic "wrong PIN" error and assumed only the Super Admin's
    # PIN works. Tell them exactly what to do instead.
    if not (me or {}).get("pin_hash"):
        await _log_vault(False, "no_pin_set")
        raise HTTPException(
            status_code=403,
            detail=("No PIN is set on your account yet. Use 'Forgot PIN?' below "
                    "to email yourself a temporary PIN, or ask the Super Admin "
                    "to set your 6-digit PIN in Sub Admins → Edit."))
    if not pin or not _verify_pin(pin, me["pin_hash"]):
        await _log_vault(False, "wrong_pin")
        raise HTTPException(
            status_code=403,
            detail=("Incorrect PIN — enter YOUR OWN admin PIN. Sub Super "
                    "Admins use their own PIN, not the Super Admin's. "
                    "Forgotten it? Tap 'Forgot PIN?' below."))
    await _log_vault(True, "success")

    from utils.secrets_vault import decrypt_secret
    names: Dict[str, str] = {}
    async for c in db.companies.find({}, {"_id": 0, "company_id": 1, "name": 1}):
        names[c["company_id"]] = c.get("name") or c["company_id"]

    out: List[Dict[str, Any]] = []
    async for fm in db.firm_masters.find(
        {}, {"_id": 0, "company_id": 1, "epf": 1, "esi": 1, "portal_logins": 1,
             "registered_address": 1, "office_address": 1, "factory_address": 1},
    ):
        cid = fm.get("company_id")
        if cid not in names:
            continue  # orphaned master — firm was deleted
        epf = fm.get("epf") or {}
        esi = fm.get("esi") or {}
        extras = []
        for row in fm.get("portal_logins") or []:
            if row.get("user_name") or row.get("password"):
                extras.append({
                    "login_type": row.get("login_type"),
                    "user_name": row.get("user_name"),
                    "password": decrypt_secret(row.get("password")),
                })
        out.append({
            "company_id": cid,
            "firm_name": names.get(cid, cid),
            "city": _fm_city(fm),
            "epf_code": epf.get("epf_no"),
            "epf_user_id": epf.get("epf_user_id"),
            "epf_password": decrypt_secret(epf.get("epf_password")),
            "esi_no": esi.get("esi_no"),
            "esi_user_id": esi.get("esi_user_id"),
            "esi_password": decrypt_secret(esi.get("esi_password")),
            "other_logins": extras,
        })
    out.sort(key=lambda r: str(r.get("firm_name") or "").lower())
    logger.info("[firm-credentials] revealed to %s (%d firms)",
                user["user_id"], len(out))
    return {"firms": out}


@router.get("/firm-credentials/access-log")
async def firm_credentials_access_log(
    authorization: Optional[str] = Header(None),
):
    """Iter 599 (user request) — VAULT ACCESS LOG viewer (Super Admin only):
    who unlocked (or failed to unlock) Firms ID & Password, and when."""
    user = await get_user_from_token(authorization)
    require_super_admin_strict(user)
    rows = await db.vault_access_log.find(
        {}, {"_id": 0},
    ).sort("at", -1).to_list(200)
    return {"logs": rows}


# ---------------------------------------------------------------------------
# Iter 325 (user request) — LIST OF FIRMS: full Firm Master data as an
# Excel-style grid + downloadable XLSX.
# ---------------------------------------------------------------------------
_FIRM_LIST_COLUMNS: List[tuple] = [
    ("firm_name", "Firm Name"), ("category", "Category"),
    ("business_nature", "Business Nature"), ("start_date", "Start Date"),
    ("city", "City"), ("state", "State"), ("pin_code", "Pin Code"),
    ("address", "Registered Address"), ("email_1", "Email 1"),
    ("email_2", "Email 2"),
    ("epf_no", "PF Code"), ("epf_applicable", "PF Applicable"),
    ("esi_no", "ESIC Code"), ("esi_applicable", "ESI Applicable"),
    ("bank_name", "Bank Name"), ("account_no", "Account No"),
    ("ifsc", "IFSC"),
    ("contact_person", "Contact Person"), ("contact_phone", "Contact Phone"),
    ("firm_active", "Active"),
]


async def _firms_master_rows() -> List[Dict[str, Any]]:
    names: Dict[str, str] = {}
    async for c in db.companies.find({}, {"_id": 0, "company_id": 1, "name": 1}):
        names[c["company_id"]] = c.get("name") or c["company_id"]
    rows: List[Dict[str, Any]] = []
    seen: set = set()
    async for fm in db.firm_masters.find({}, {"_id": 0}):
        cid = fm.get("company_id")
        if cid not in names:
            continue
        seen.add(cid)
        ra = fm.get("registered_address") or {}
        hdr = fm.get("header") or {}
        epf = fm.get("epf") or {}
        esi = fm.get("esi") or {}
        bank = fm.get("bank") or {}
        cps = fm.get("contact_persons") or []
        cp = cps[0] if cps else {}
        addr = ", ".join(str(x).strip() for x in [ra.get("address1"), ra.get("address2")]
                         if x and str(x).strip())
        rows.append({
            "company_id": cid,
            "firm_name": names[cid],
            "category": hdr.get("category") or "",
            "business_nature": hdr.get("business_nature") or "",
            "start_date": hdr.get("start_date") or "",
            "city": _fm_city(fm),
            "state": ra.get("state") or "",
            "pin_code": ra.get("pin_code") or "",
            "address": addr,
            "email_1": hdr.get("email_1") or "",
            "email_2": hdr.get("email_2") or "",
            "epf_no": epf.get("epf_no") or "",
            "epf_applicable": "Yes" if epf.get("applicable") else "No",
            "esi_no": esi.get("esi_no") or "",
            "esi_applicable": "Yes" if esi.get("applicable") else "No",
            "bank_name": bank.get("bank_name") or "",
            "account_no": bank.get("account_no") or "",
            "ifsc": bank.get("ifsc") or "",
            "contact_person": (cp.get("name") or "") if isinstance(cp, dict) else "",
            "contact_phone": (cp.get("phone") or cp.get("mobile") or "") if isinstance(cp, dict) else "",
            "firm_active": "Yes" if (fm.get("settings") or {}).get("firm_active", True) else "No",
        })
    # Firms without a saved Firm Master still appear (name only).
    for cid, nm in names.items():
        if cid not in seen:
            rows.append({"company_id": cid, "firm_name": nm,
                         **{k: "" for k, _l in _FIRM_LIST_COLUMNS
                            if k not in ("firm_name",)}})
    rows.sort(key=lambda r: str(r.get("firm_name") or "").lower())
    return rows


@router.get("/firms-master-list")
async def firms_master_list(authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    require_role(user, ["super_admin", "sub_admin"])
    rows = await _firms_master_rows()
    return {"columns": [{"key": k, "label": l} for k, l in _FIRM_LIST_COLUMNS],
            "firms": rows}


@router.get("/firms-master-list/export.xlsx")
async def firms_master_list_xlsx(authorization: Optional[str] = Header(None)):
    import io as _io
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    user = await get_user_from_token(authorization)
    require_role(user, ["super_admin", "sub_admin"])
    rows = await _firms_master_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Firms Master"
    heads = ["S.No"] + [l for _k, l in _FIRM_LIST_COLUMNS]
    for i, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F3B5C")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=r - 1)
        for i, (k, _l) in enumerate(_FIRM_LIST_COLUMNS, 2):
            ws.cell(row=r, column=i, value=row.get(k) or "")
    for i, w in enumerate([6, 32, 14, 18, 12, 14, 14, 10, 34, 24, 24,
                           18, 12, 18, 12, 18, 18, 14, 20, 14, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    buf = _io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Firms_Master_List.xlsx"',
                 "Cache-Control": "no-store"})
