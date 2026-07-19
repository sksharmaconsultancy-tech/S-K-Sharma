"""Iter 202 — Remaining statutory / management reports.

  * Professional Tax (PT)  — state-wise slabs for all major states.
  * Labour Welfare Fund    — state-wise EE/ER contributions + periodicity.
  * Gratuity               — accrual & eligibility (15/26 × basic × years).
  * Full & Final (F&F)     — settlement sheet for exited employees.
  * Advance / Loan         — register with outstanding balances.
  * Management MIS         — per-firm KPI summary for a month.

All endpoints accept ``fmt=json|xlsx|pdf`` and share the firm scoping used
by the other statutory registers.
"""
import base64
import calendar
import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Header, HTTPException, Query
from fastapi.responses import Response

from server import db, get_user_from_token, require_role, sub_admin_can_touch_company  # noqa: E402
from routes.statutory_registers import (  # noqa: E402
    _check, _company, _new_pdf, _pdf_header, _table, _sig_block,
    _pdf_response, _s, _rs,
)
from utils.compliance_salary import resolve_structure, _num  # noqa: E402

router = APIRouter(prefix="/api/admin/reports", tags=["statutory-extra-reports"])


# ---------------------------------------------------------------------------
# State rules — Professional Tax (monthly slabs unless noted).
# NOTE: rates follow commonly-published 2025-26 notifications; verify with
# the latest state notification before filing.
# ---------------------------------------------------------------------------
def _pt_amount(state: str, gross: float, month_num: int) -> Tuple[float, str]:
    """Return (pt_for_month, rule_note)."""
    s = (state or "").strip().lower()
    g = gross

    if s in ("maharashtra",):
        if g <= 7500:
            return 0.0, "MH: <=7500 nil; 7501-10000 Rs.175; >10000 Rs.200 (Feb Rs.300)"
        if g <= 10000:
            return 175.0, "MH slab 7501-10000"
        return (300.0 if month_num == 2 else 200.0), "MH slab >10000 (Feb Rs.300)"
    if s in ("karnataka",):
        return (200.0 if g >= 25000 else 0.0), "KA: >=25000 Rs.200/month"
    if s in ("west bengal",):
        if g <= 10000:
            return 0.0, "WB: <=10000 nil"
        if g <= 15000:
            return 110.0, "WB slab 10001-15000"
        if g <= 25000:
            return 130.0, "WB slab 15001-25000"
        if g <= 40000:
            return 150.0, "WB slab 25001-40000"
        return 200.0, "WB slab >40000"
    if s in ("madhya pradesh",):
        if g <= 18750:
            return 0.0, "MP: <=2.25L/yr nil"
        if g <= 25000:
            return 125.0, "MP slab 2.25-3L/yr"
        if g <= 33333:
            return 167.0, "MP slab 3-4L/yr"
        return 208.0, "MP slab >4L/yr (Rs.2500/yr)"
    if s in ("gujarat",):
        return (200.0 if g >= 12000 else 0.0), "GJ: >=12000 Rs.200/month"
    if s in ("andhra pradesh", "telangana"):
        if g <= 15000:
            return 0.0, "AP/TS: <=15000 nil"
        if g <= 20000:
            return 150.0, "AP/TS slab 15001-20000"
        return 200.0, "AP/TS slab >20000"
    if s in ("tamil nadu", "tamilnadu"):
        # Half-yearly levy — shown as monthly equivalent (half-year ÷ 6).
        if g <= 21000:
            hy = 0.0
        elif g <= 30000:
            hy = 135.0
        elif g <= 45000:
            hy = 315.0
        elif g <= 60000:
            hy = 690.0
        elif g <= 75000:
            hy = 1025.0
        else:
            hy = 1250.0
        return round(hy / 6.0, 2), "TN: half-yearly levy (shown as monthly equivalent)"
    if s in ("kerala",):
        if g < 12000:
            hy = 0.0
        elif g < 18000:
            hy = 120.0
        elif g < 30000:
            hy = 180.0
        elif g < 45000:
            hy = 300.0
        elif g < 60000:
            hy = 450.0
        elif g < 75000:
            hy = 600.0
        else:
            hy = 1250.0
        return round(hy / 6.0, 2), "KL: half-yearly levy (shown as monthly equivalent)"
    if s in ("bihar",):
        annual = g * 12
        if annual <= 300000:
            a = 0.0
        elif annual <= 500000:
            a = 1000.0
        elif annual <= 1000000:
            a = 2000.0
        else:
            a = 2500.0
        return round(a / 12.0, 2), "BR: annual slab (shown as monthly equivalent)"
    if s in ("jharkhand",):
        annual = g * 12
        if annual <= 300000:
            a = 0.0
        elif annual <= 500000:
            a = 1200.0
        elif annual <= 800000:
            a = 1800.0
        elif annual <= 1000000:
            a = 2100.0
        else:
            a = 2500.0
        return round(a / 12.0, 2), "JH: annual slab (shown as monthly equivalent)"
    if s in ("odisha", "orissa"):
        annual = g * 12
        if annual <= 160000:
            return 0.0, "OD: <=1.6L/yr nil"
        if annual <= 300000:
            return 125.0, "OD slab 1.6-3L/yr"
        return (300.0 if month_num == 12 else 200.0), "OD slab >3L/yr (Dec Rs.300)"
    if s in ("assam",):
        if g <= 10000:
            return 0.0, "AS: <=10000 nil"
        if g <= 15000:
            return 150.0, "AS slab 10001-15000"
        if g <= 25000:
            return 180.0, "AS slab 15001-25000"
        return 208.0, "AS slab >25000"
    if s in ("meghalaya",):
        if g <= 4166:
            return 0.0, "ML: <=50k/yr nil"
        if g <= 6250:
            return 16.5, "ML slab"
        if g <= 8333:
            return 25.0, "ML slab"
        if g <= 12500:
            return 41.5, "ML slab"
        if g <= 16666:
            return 62.5, "ML slab"
        if g <= 20833:
            return 83.33, "ML slab"
        if g <= 25000:
            return 104.16, "ML slab"
        return 208.0, "ML top slab"
    if s in ("tripura",):
        if g <= 7500:
            return 0.0, "TR: <=7500 nil"
        if g <= 15000:
            return 150.0, "TR slab 7501-15000"
        return 208.0, "TR slab >15000"
    if s in ("sikkim",):
        if g <= 20000:
            return 0.0, "SK: <=20000 nil"
        if g <= 30000:
            return 125.0, "SK slab 20001-30000"
        if g <= 40000:
            return 150.0, "SK slab 30001-40000"
        return 200.0, "SK slab >40000"
    if s in ("punjab",):
        return (200.0 if g > 20833 else 0.0), "PB: PSDT Rs.200/month above taxable limit"
    if s in ("manipur",):
        annual = g * 12
        if annual <= 50000:
            return 0.0, "MN: <=50k/yr nil"
        if annual <= 75000:
            return 100.0, "MN slab"
        if annual <= 100000:
            return 166.66, "MN slab"
        if annual <= 125000:
            return 200.0, "MN slab"
        return 208.33, "MN top slab"
    if s in ("nagaland",):
        if g <= 4000:
            return 0.0, "NL: <=4000 nil"
        if g <= 5000:
            return 35.0, "NL slab"
        if g <= 7000:
            return 75.0, "NL slab"
        if g <= 9000:
            return 110.0, "NL slab"
        if g <= 12000:
            return 180.0, "NL slab"
        return 208.0, "NL top slab"
    if s in ("mizoram",):
        if g <= 5000:
            return 0.0, "MZ: <=5000 nil"
        if g <= 8000:
            return 75.0, "MZ slab"
        if g <= 10000:
            return 120.0, "MZ slab"
        if g <= 12000:
            return 150.0, "MZ slab"
        if g <= 15000:
            return 180.0, "MZ slab"
        return 208.0, "MZ top slab"
    if s in ("puducherry", "pondicherry"):
        hy_slabs = [(99999, 0.0), (200000, 250.0), (300000, 500.0),
                    (400000, 750.0), (500000, 1000.0), (10**9, 1250.0)]
        annual = g * 12
        hy = next(v for cap, v in hy_slabs if annual <= cap)
        return round(hy / 6.0, 2), "PY: half-yearly levy (shown as monthly equivalent)"
    return 0.0, f"No Professional Tax levied in {state or 'this state'}"


# ---------------------------------------------------------------------------
# State rules — Labour Welfare Fund. (ee, er, frequency, due_months)
# frequency: monthly | half_yearly | annual
# ---------------------------------------------------------------------------
LWF_RULES: Dict[str, Dict[str, Any]] = {
    "maharashtra":     {"ee": 12.0, "er": 36.0, "frequency": "half_yearly", "due_months": [6, 12]},
    "karnataka":       {"ee": 50.0, "er": 100.0, "frequency": "annual", "due_months": [12]},
    "gujarat":         {"ee": 6.0, "er": 12.0, "frequency": "half_yearly", "due_months": [6, 12]},
    "tamil nadu":      {"ee": 20.0, "er": 40.0, "frequency": "annual", "due_months": [12]},
    "kerala":          {"ee": 20.0, "er": 20.0, "frequency": "half_yearly", "due_months": [6, 12]},
    "madhya pradesh":  {"ee": 10.0, "er": 30.0, "frequency": "half_yearly", "due_months": [6, 12]},
    "andhra pradesh":  {"ee": 30.0, "er": 70.0, "frequency": "annual", "due_months": [12]},
    "telangana":       {"ee": 2.0, "er": 5.0, "frequency": "annual", "due_months": [12]},
    "west bengal":     {"ee": 3.0, "er": 15.0, "frequency": "half_yearly", "due_months": [6, 12]},
    "haryana":         {"ee": 31.0, "er": 62.0, "frequency": "monthly", "due_months": list(range(1, 13))},
    "punjab":          {"ee": 5.0, "er": 20.0, "frequency": "monthly", "due_months": list(range(1, 13))},
    "chandigarh":      {"ee": 5.0, "er": 20.0, "frequency": "monthly", "due_months": list(range(1, 13))},
    "chhattisgarh":    {"ee": 15.0, "er": 45.0, "frequency": "half_yearly", "due_months": [6, 12]},
    "goa":             {"ee": 60.0, "er": 180.0, "frequency": "half_yearly", "due_months": [6, 12]},
    "odisha":          {"ee": 10.0, "er": 20.0, "frequency": "half_yearly", "due_months": [6, 12]},
    "delhi":           {"ee": 0.75, "er": 2.25, "frequency": "half_yearly", "due_months": [6, 12]},
}


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _code_key(r):
    c = str(r.get("employee_code") or "").strip()
    try:
        return (0, float(c), "")
    except ValueError:
        return (1, 0.0, c.lower())


async def _emps(company_id: str, include_exited: bool = False) -> List[dict]:
    q: Dict[str, Any] = {"role": "employee", "company_id": company_id}
    rows = await db.users.find(q, {
        "_id": 0, "user_id": 1, "employee_code": 1, "name": 1, "father_name": 1,
        "designation": 1, "department": 1, "contractor_name": 1, "gender": 1,
        "doj": 1, "exit_date": 1, "exit_reason": 1, "employment_status": 1,
        "salary_monthly": 1, "compliance_gross": 1, "structure_pct": 1,
        "basic_amount": 1, "hra_amount": 1, "conv_amount": 1,
        "medical_amount": 1, "special_amount": 1, "others_amount": 1,
        "salary_structure_actual": 1, "is_onroll": 1, "pf_no": 1, "uan_no": 1,
        "compliance_salary_mode": 1,
    }).to_list(10000)
    if not include_exited:
        today = date.today().strftime("%Y-%m-%d")
        rows = [r for r in rows if not (r.get("exit_date") and str(r["exit_date"]) <= today)]
    rows.sort(key=_code_key)
    return rows


def _parse_doj(v: Any) -> Optional[date]:
    s = str(v or "").strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _service_years(doj: Any, until: Optional[date] = None) -> float:
    d = _parse_doj(doj)
    if not d:
        return 0.0
    until = until or date.today()
    return max(0.0, round((until - d).days / 365.25, 2))


async def _gross_map(company_id: str, month: str) -> Tuple[Dict[str, float], str]:
    """user_id → month gross. Prefers the compliance run, then actual run,
    then the master salary."""
    run = await db.compliance_salary_runs.find_one(
        {"company_id": company_id, "month": month, "finalized": True}, {"_id": 0, "rows": 1})
    if not run:
        run = await db.compliance_salary_runs.find_one(
            {"company_id": company_id, "month": month}, {"_id": 0, "rows": 1})
    if run and run.get("rows"):
        return ({r["user_id"]: _num(r.get("gross_paid")) or _num(r.get("monthly_gross"))
                 for r in run["rows"] if r.get("user_id")}, "compliance run")
    run = await db.salary_runs.find_one(
        {"company_id": company_id, "month": month, "finalized": True}, {"_id": 0, "rows": 1})
    if not run:
        run = await db.salary_runs.find_one(
            {"company_id": company_id, "month": month}, {"_id": 0, "rows": 1})
    if run and run.get("rows"):
        return ({r["user_id"]: _num(r.get("total_gross"))
                 for r in run["rows"] if r.get("user_id")}, "actual salary run")
    return {}, "employee master"


def _master_gross(u: dict) -> float:
    """Monthly-equivalent gross from the employee master. Daily-rated
    employees (rate_type=daily / compliance_salary_mode=daily) are
    converted at 26 working days so PT slabs & gratuity use a true
    monthly wage."""
    g = _num(u.get("compliance_gross"))
    if g > 0:
        return g
    mode_daily = str(u.get("compliance_salary_mode") or "").lower() == "daily"
    g = _num(u.get("salary_monthly"))
    if g > 0:
        return g * 26 if mode_daily else g
    for r in (u.get("salary_structure_actual") or []):
        if isinstance(r, dict) and str(r.get("head", "")).lower().startswith("basic"):
            amt = _num(r.get("amount"))
            if amt <= 0:
                continue
            if mode_daily or str(r.get("rate_type") or "").lower() == "daily":
                return amt * 26
            return amt
    return 0.0


def _xlsx_bytes(title: str, subtitle: str, cols: List[Tuple[str, str, bool]],
                rows: List[dict], totals: bool = True) -> bytes:
    """cols = [(label, key, numeric)]"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    # Sanitize sheet title — openpyxl rejects / \ * ? : [ ]
    _safe_title = title[:31]
    for _ch in "/\\*?:[]":
        _safe_title = _safe_title.replace(_ch, "-")
    ws.title = _safe_title
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13, color="1F3D7A")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(cols)))
    ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, size=10, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(cols)))
    hdr = 3
    for i, (label, _k, _n) in enumerate(cols, start=1):
        c = ws.cell(row=hdr, column=i, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3D7A")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)
    for r_idx, row in enumerate(rows, start=hdr + 1):
        for c_idx, (_l, key, numeric) in enumerate(cols, start=1):
            v = row.get(key)
            if numeric:
                cell = ws.cell(row=r_idx, column=c_idx, value=_num(v) if v not in (None, "") else None)
                cell.number_format = "#,##0.00"
            else:
                ws.cell(row=r_idx, column=c_idx, value="" if v is None else v)
    if totals and rows:
        tr = hdr + 1 + len(rows)
        ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
        for c_idx, (_l, key, numeric) in enumerate(cols, start=1):
            if not numeric:
                continue
            cell = ws.cell(row=tr, column=c_idx,
                           value=round(sum(_num(r.get(key)) for r in rows), 2))
            cell.font = Font(bold=True)
            cell.number_format = "#,##0.00"
    for i, (label, key, _n) in enumerate(cols, start=1):
        w = max(len(label), *(len(str(r.get(key) or "")) for r in rows)) if rows else len(label)
        ws.column_dimensions[get_column_letter(i)].width = min(w + 3, 40)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_response(data: bytes, fname: str) -> Response:
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _emit(fmt: str, *, title: str, subtitle: str, company: dict,
          cols: List[Tuple[str, str, bool]], rows: List[dict],
          fname_base: str, pdf_note: str = "", landscape: bool = True,
          json_extra: Optional[dict] = None):
    if fmt == "xlsx":
        return _xlsx_response(
            _xlsx_bytes(title, f"{company.get('name') or ''} — {subtitle}", cols, rows),
            f"{fname_base}.xlsx")
    if fmt == "pdf":
        pdf = _new_pdf(landscape=landscape)
        _pdf_header(pdf, title, subtitle, company)
        headers = [c[0] for c in cols]
        page_w = pdf.w - pdf.l_margin - pdf.r_margin
        weights = [max(len(c[0]), 8) for c in cols]
        # widen name-ish columns
        for i, (_l, key, _n) in enumerate(cols):
            if key in ("name", "detail", "note", "rule"):
                weights[i] = int(weights[i] * 2.2)
        total_w = sum(weights)
        widths = [round(page_w * w / total_w, 1) for w in weights]
        aligns = ["R" if c[2] else "L" for c in cols]
        body = [[(_rs(r.get(k)) if numeric and r.get(k) not in (None, "") else str(r.get(k) if r.get(k) is not None else ""))
                 for (_l, k, numeric) in cols] for r in rows]
        if rows:
            tot = ["TOTAL"] + [""] * (len(cols) - 1)
            for i, (_l, k, numeric) in enumerate(cols):
                if numeric and i > 0:
                    tot[i] = _rs(round(sum(_num(r.get(k)) for r in rows), 2))
            body.append(tot)
        _table(pdf, headers, widths, body, aligns=aligns, font_size=7.5)
        if pdf_note:
            pdf.ln(3)
            pdf.set_font("Helvetica", "I", 8.5)
            pdf.multi_cell(0, 4.5, _s(pdf_note))
        _sig_block(pdf, company.get("name") or "")
        return _pdf_response(pdf, f"{fname_base}.pdf")
    out = {"rows": rows, "count": len(rows), "title": title, "subtitle": subtitle}
    if json_extra:
        out.update(json_extra)
    return out


# ---------------------------------------------------------------------------
# 1) PROFESSIONAL TAX
# ---------------------------------------------------------------------------
@router.get("/pt")
async def pt_report(
    company_id: str = Query(...),
    month: str = Query(...),        # YYYY-MM
    fmt: str = Query("json"),
    state: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    await _check(admin, company_id)
    company = await _company(company_id)
    st = (state or company.get("state") or "").strip()
    month_num = int(month[5:7])
    gmap, source = await _gross_map(company_id, month)
    emps = await _emps(company_id)
    rows, note = [], ""
    for u in emps:
        gross = gmap.get(u["user_id"]) or _master_gross(u)
        pt, note = _pt_amount(st, gross, month_num)
        rows.append({
            "employee_code": u.get("employee_code"), "name": u.get("name"),
            "designation": u.get("designation"), "gross": round(gross, 2),
            "pt": round(pt, 2),
        })
    cols = [("Emp Code", "employee_code", False), ("Name", "name", False),
            ("Designation", "designation", False), ("Monthly Gross", "gross", True),
            ("PT Amount", "pt", True)]
    return _emit(fmt, title=f"Professional Tax Report — {month}",
                 subtitle=f"State: {st or 'Not set'} · Gross source: {source}",
                 company=company, cols=cols, rows=rows,
                 fname_base=f"PT_Report_{month}",
                 pdf_note=f"Rule applied: {note}. Verify slabs with the latest state notification.",
                 landscape=False,
                 json_extra={"state": st, "rule_note": note, "gross_source": source,
                             "total_pt": round(sum(r["pt"] for r in rows), 2)})


# ---------------------------------------------------------------------------
# 2) LABOUR WELFARE FUND
# ---------------------------------------------------------------------------
@router.get("/lwf")
async def lwf_report(
    company_id: str = Query(...),
    month: str = Query(...),
    fmt: str = Query("json"),
    state: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    await _check(admin, company_id)
    company = await _company(company_id)
    st = (state or company.get("state") or "").strip()
    rule = LWF_RULES.get(st.lower())
    month_num = int(month[5:7])
    emps = await _emps(company_id)
    is_due = bool(rule and month_num in rule["due_months"])
    rows = []
    for u in emps:
        ee = rule["ee"] if is_due else 0.0
        er = rule["er"] if is_due else 0.0
        rows.append({
            "employee_code": u.get("employee_code"), "name": u.get("name"),
            "designation": u.get("designation"),
            "ee": round(ee, 2) if rule else 0.0,
            "er": round(er, 2) if rule else 0.0,
            "total": round(ee + er, 2) if rule else 0.0,
        })
    if not rule:
        note = f"No Labour Welfare Fund levied in {st or 'this state'} (or state not configured)."
    elif is_due:
        note = (f"{st}: EE Rs.{rule['ee']} / ER Rs.{rule['er']} per employee — "
                f"{rule['frequency'].replace('_', '-')} contribution, due this month.")
    else:
        due = ", ".join(calendar.month_abbr[m] for m in rule["due_months"][:12])
        note = (f"{st}: {rule['frequency'].replace('_', '-')} contribution due in {due} — "
                f"no deposit due for {month}. Amounts shown as 0.")
    cols = [("Emp Code", "employee_code", False), ("Name", "name", False),
            ("Designation", "designation", False), ("Employee", "ee", True),
            ("Employer", "er", True), ("Total", "total", True)]
    return _emit(fmt, title=f"Labour Welfare Fund Report — {month}",
                 subtitle=f"State: {st or 'Not set'}",
                 company=company, cols=cols, rows=rows,
                 fname_base=f"LWF_Report_{month}", pdf_note=note, landscape=False,
                 json_extra={"state": st, "rule_note": note, "due_this_month": is_due,
                             "total_ee": round(sum(r["ee"] for r in rows), 2),
                             "total_er": round(sum(r["er"] for r in rows), 2)})


# ---------------------------------------------------------------------------
# 3) GRATUITY
# ---------------------------------------------------------------------------
@router.get("/gratuity")
async def gratuity_report(
    company_id: str = Query(...),
    fmt: str = Query("json"),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    await _check(admin, company_id)
    company = await _company(company_id)
    structure_pct = None
    comp_full = await db.companies.find_one(
        {"company_id": company_id}, {"_id": 0, "compliance_policy": 1, "structure_pct": 1})
    if comp_full:
        structure_pct = ((comp_full.get("compliance_policy") or {}).get("structure_pct")
                         or comp_full.get("structure_pct"))
    emps = await _emps(company_id, include_exited=True)
    rows = []
    for u in emps:
        yrs = _service_years(u.get("doj"), _parse_doj(u.get("exit_date")) or None)
        gross = _master_gross(u)
        basic = resolve_structure(u, gross, structure_pct)["basic"] if gross > 0 else 0.0
        eligible = yrs >= 4.81  # 4 years + 240 days continuous service
        # Payable years: round half-up (6+ months counts as a full year)
        pay_years = int(yrs) + (1 if (yrs - int(yrs)) >= 0.5 else 0)
        amount = round(basic * 15.0 / 26.0 * pay_years, 2) if eligible else 0.0
        accrued = round(basic * 15.0 / 26.0 * pay_years, 2)
        rows.append({
            "employee_code": u.get("employee_code"), "name": u.get("name"),
            "doj": u.get("doj"), "exit_date": u.get("exit_date") or "",
            "service_years": yrs, "monthly_basic": round(basic, 2),
            "eligible": "Yes" if eligible else "No",
            "gratuity": amount, "accrued": accrued,
        })
    cols = [("Emp Code", "employee_code", False), ("Name", "name", False),
            ("DOJ", "doj", False), ("Exit Date", "exit_date", False),
            ("Service (Yrs)", "service_years", False), ("Monthly Basic", "monthly_basic", True),
            ("Eligible", "eligible", False), ("Gratuity Payable", "gratuity", True),
            ("Accrued (Projected)", "accrued", True)]
    return _emit(fmt, title="Gratuity Report",
                 subtitle="Payment of Gratuity Act, 1972 — 15/26 x last Basic x completed years",
                 company=company, cols=cols, rows=rows, fname_base="Gratuity_Report",
                 pdf_note="Eligibility: 4 years + 240 days of continuous service. "
                          "6+ months in the final year counts as a full year.",
                 json_extra={"total_liability": round(sum(r["gratuity"] for r in rows), 2)})


# ---------------------------------------------------------------------------
# 4) FULL & FINAL SETTLEMENT
# ---------------------------------------------------------------------------
@router.get("/fnf")
async def fnf_report(
    company_id: str = Query(...),
    month: Optional[str] = Query(None),   # exit month filter (YYYY-MM)
    fmt: str = Query("json"),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    await _check(admin, company_id)
    company = await _company(company_id)
    comp_full = await db.companies.find_one(
        {"company_id": company_id}, {"_id": 0, "compliance_policy": 1, "structure_pct": 1})
    structure_pct = ((comp_full.get("compliance_policy") or {}).get("structure_pct")
                     or comp_full.get("structure_pct")) if comp_full else None

    emps = await _emps(company_id, include_exited=True)
    exited = [u for u in emps if u.get("exit_date")]
    if month:
        exited = [u for u in exited if str(u["exit_date"])[:7] == month]

    rows = []
    for u in exited:
        exit_month = str(u["exit_date"])[:7]
        gmap, _src = await _gross_map(company_id, exit_month)
        gross = gmap.get(u["user_id"]) or _master_gross(u)
        # Iter 202 (user request) — Present Days per the firm's Attendance
        # Policy (same pipeline as the grid: 8-HR sub-point, duty-hours
        # basis, week-off/holiday sub-points), not a raw punch-date count.
        from server import _compute_monthly_grid_data
        try:
            _grid = await _compute_monthly_grid_data(
                company_id=company_id, month=exit_month,
                only_user_id=u["user_id"])
            _tot = ((_grid.get("rows") or [{}])[0].get("totals") or {})
            days_worked = _num(_tot.get("present_days_policy"),
                               _num(_tot.get("present_days")))
        except Exception:
            days_worked = len(await db.attendance.distinct("date", {
                "user_id": u["user_id"], "kind": "in",
                "date": {"$gte": f"{exit_month}-01", "$lte": str(u["exit_date"])},
                "status": {"$in": ["approved", None]},
            }))
        y, m = int(exit_month[:4]), int(exit_month[5:7])
        month_days = calendar.monthrange(y, m)[1]
        earned = round(gross * days_worked / max(1, month_days), 2)
        yrs = _service_years(u.get("doj"), _parse_doj(u.get("exit_date")))
        basic = resolve_structure(u, _master_gross(u), structure_pct)["basic"] \
            if _master_gross(u) > 0 else 0.0
        eligible = yrs >= 4.81
        pay_years = int(yrs) + (1 if (yrs - int(yrs)) >= 0.5 else 0)
        gratuity = round(basic * 15.0 / 26.0 * pay_years, 2) if eligible else 0.0
        # Outstanding advances
        adv_docs = await db.advances.find(
            {"user_id": u["user_id"], "company_id": company_id},
            {"_id": 0, "remaining_balance": 1, "status": 1}).to_list(100)
        adv_out = round(sum(_num(a.get("remaining_balance")) for a in adv_docs
                            if str(a.get("status") or "").lower() not in
                            ("closed", "cancelled", "rejected")), 2)
        net = round(earned + gratuity - adv_out, 2)
        rows.append({
            "employee_code": u.get("employee_code"), "name": u.get("name"),
            "doj": u.get("doj"), "exit_date": u.get("exit_date"),
            "service_years": yrs, "days_worked": round(_num(days_worked), 2),
            "monthly_gross": round(gross, 2), "earned_salary": earned,
            "gratuity": gratuity, "advance_recovery": adv_out, "net_payable": net,
        })
    cols = [("Emp Code", "employee_code", False), ("Name", "name", False),
            ("DOJ", "doj", False), ("Exit Date", "exit_date", False),
            ("Service (Yrs)", "service_years", False), ("Present Days", "days_worked", False),
            ("Monthly Gross", "monthly_gross", True), ("Earned Salary", "earned_salary", True),
            ("Gratuity", "gratuity", True), ("Advance Recovery", "advance_recovery", True),
            ("Net Payable", "net_payable", True)]
    return _emit(fmt, title=f"Full & Final Settlement{' — ' + month if month else ''}",
                 subtitle="Earned salary (exit month) + gratuity - outstanding advances",
                 company=company, cols=cols, rows=rows,
                 fname_base=f"FnF_Report{('_' + month) if month else ''}",
                 pdf_note="Leave encashment / bonus / notice-pay adjustments (if any) are to be "
                          "added manually per company policy.",
                 json_extra={"total_net": round(sum(r["net_payable"] for r in rows), 2)})


# ---------------------------------------------------------------------------
# 5) ADVANCE / LOAN REGISTER
# ---------------------------------------------------------------------------
@router.get("/advance-loan")
async def advance_loan_report(
    company_id: str = Query(...),
    fmt: str = Query("json"),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    await _check(admin, company_id)
    company = await _company(company_id)
    advs = await db.advances.find(
        {"company_id": company_id}, {"_id": 0}).sort("advance_date", -1).to_list(5000)
    rows = []
    for a in advs:
        rows.append({
            "voucher_no": a.get("voucher_no"), "advance_date": a.get("advance_date"),
            "employee_code": a.get("employee_code"), "name": a.get("employee_name"),
            "advance_type": a.get("advance_type"), "amount": _num(a.get("amount")),
            "emi_amount": _num(a.get("emi_amount")),
            "recovered": _num(a.get("recovered_total")),
            "balance": _num(a.get("remaining_balance")),
            "status": (a.get("status") or "active").upper(),
        })
    cols = [("Voucher", "voucher_no", False), ("Date", "advance_date", False),
            ("Emp Code", "employee_code", False), ("Name", "name", False),
            ("Type", "advance_type", False), ("Amount", "amount", True),
            ("EMI", "emi_amount", True), ("Recovered", "recovered", True),
            ("Balance", "balance", True), ("Status", "status", False)]
    return _emit(fmt, title="Advance / Loan Register",
                 subtitle="All advances with recovery status",
                 company=company, cols=cols, rows=rows, fname_base="Advance_Loan_Register",
                 json_extra={"total_outstanding": round(sum(r["balance"] for r in rows), 2)})


# ---------------------------------------------------------------------------
# 6) MANAGEMENT MIS (per-firm summary)
# ---------------------------------------------------------------------------
@router.get("/mis")
async def mis_report(
    month: str = Query(...),
    company_id: Optional[str] = Query(None),
    fmt: str = Query("json"),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    q: Dict[str, Any] = {}
    if company_id:
        q["company_id"] = company_id
    companies = await db.companies.find(q, {"_id": 0, "company_id": 1, "name": 1,
                                            "state": 1}).to_list(500)
    if admin.get("role") == "sub_admin":
        companies = [c for c in companies if sub_admin_can_touch_company(admin, c["company_id"])]
    if admin.get("role") == "company_admin":
        companies = [c for c in companies if c["company_id"] == admin.get("company_id")]

    y, m = int(month[:4]), int(month[5:7])
    month_start = f"{month}-01"
    month_end = f"{month}-{calendar.monthrange(y, m)[1]:02d}"
    today = date.today().strftime("%Y-%m-%d")
    rows = []
    for c in companies:
        cid = c["company_id"]
        emps = await db.users.find(
            {"role": "employee", "company_id": cid},
            {"_id": 0, "user_id": 1, "doj": 1, "exit_date": 1, "is_onroll": 1}).to_list(10000)
        active = [e for e in emps if not (e.get("exit_date") and str(e["exit_date"]) <= today)]
        joins = 0
        for e in emps:
            d = _parse_doj(e.get("doj"))
            if d and d.strftime("%Y-%m") == month:
                joins += 1
        exits = len([e for e in emps
                     if e.get("exit_date") and str(e["exit_date"])[:7] == month])
        # Man-days: distinct (user, date) IN punches in the month
        pipe = [
            {"$match": {"company_id": cid, "kind": "in",
                        "date": {"$gte": month_start, "$lte": month_end}}},
            {"$group": {"_id": {"u": "$user_id", "d": "$date"}}},
            {"$count": "n"},
        ]
        agg = await db.attendance.aggregate(pipe).to_list(1)
        man_days = agg[0]["n"] if agg else 0
        # Salary totals
        arun = await db.salary_runs.find_one(
            {"company_id": cid, "month": month},
            {"_id": 0, "totals": 1}, sort=[("finalized", -1)])
        crun = await db.compliance_salary_runs.find_one(
            {"company_id": cid, "month": month},
            {"_id": 0, "totals": 1}, sort=[("finalized", -1)])
        at = (arun or {}).get("totals") or {}
        ct = (crun or {}).get("totals") or {}
        adv_docs = await db.advances.find(
            {"company_id": cid}, {"_id": 0, "remaining_balance": 1, "status": 1}).to_list(5000)
        adv_out = round(sum(_num(a.get("remaining_balance")) for a in adv_docs
                            if str(a.get("status") or "").lower() not in
                            ("closed", "cancelled", "rejected")), 2)
        rows.append({
            "firm": c.get("name"), "state": c.get("state") or "",
            "headcount": len(active), "joins": joins, "exits": exits,
            "man_days": man_days,
            "actual_gross": round(_num(at.get("total_gross")), 2),
            "actual_net": round(_num(at.get("net_pay")), 2),
            "compliance_gross": round(_num(ct.get("gross_paid") or ct.get("monthly_gross")), 2),
            "pf": round(_num(ct.get("pf_employee")) + _num(ct.get("pf_employer_total")), 2),
            "esic": round(_num(ct.get("esic_employee")) + _num(ct.get("esic_employer")), 2),
            "pt": round(_num(ct.get("pt")), 2),
            "advance_outstanding": adv_out,
        })
    cols = [("Firm", "firm", False), ("State", "state", False),
            ("Headcount", "headcount", False), ("Joins", "joins", False),
            ("Exits", "exits", False), ("Man-Days", "man_days", False),
            ("Actual Gross", "actual_gross", True), ("Actual Net", "actual_net", True),
            ("Compliance Gross", "compliance_gross", True), ("PF (EE+ER)", "pf", True),
            ("ESIC (EE+ER)", "esic", True), ("PT", "pt", True),
            ("Advances O/S", "advance_outstanding", True)]
    company = {"name": "All Firms" if not company_id else
               (companies[0].get("name") if companies else "")}
    return _emit(fmt, title=f"Management MIS — {month}",
                 subtitle="Headcount, attendance man-days, payroll & statutory summary per firm",
                 company=company, cols=cols, rows=rows, fname_base=f"MIS_Report_{month}",
                 pdf_note="Salary figures come from the month's processed runs "
                          "(finalized preferred). 0 = month not processed for that firm.")
