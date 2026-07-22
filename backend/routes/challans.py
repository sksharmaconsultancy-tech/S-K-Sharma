"""Iter 89 — PF / ESIC Challan uploads (Automation module).

Endpoints:
  * POST   /api/admin/challans                - Upload a challan
  * GET    /api/admin/challans                - List (with filters)
  * GET    /api/admin/challans/{challan_id}   - Fetch full doc with file
  * GET    /api/admin/challans/export.xlsx    - Excel export of the list
  * DELETE /api/admin/challans/{challan_id}   - Remove a challan

Challan doc shape:
  {
    challan_id, company_id, portal ("pf" | "esic"),
    month (YYYY-MM), amount, trrn (transaction ref), paid_on,
    notes,
    file_base64, file_mime, file_name,
    created_by, created_at
  }
"""
import base64
import io
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException
from fastapi.responses import StreamingResponse

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
    now_iso,
    logger,
)


router = APIRouter(prefix="/api/admin", tags=["challans"])


ALLOWED_MIMES = {
    "application/pdf": ".pdf",
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/webp": ".webp",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/vnd.ms-excel": ".xls",
    "text/csv": ".csv",
}


async def _assert_admin(user_id_or_token: Optional[str]) -> Dict[str, Any]:
    user = await get_user_from_token(user_id_or_token)
    require_role(user, ["super_admin", "company_admin", "sub_admin"])
    return user


def _scope_company(user: Dict[str, Any], company_id: Optional[str]) -> str:
    if user["role"] == "company_admin":
        if company_id and company_id != user.get("company_id"):
            raise HTTPException(status_code=403, detail="Not your firm")
        return user["company_id"] or ""
    if user["role"] == "sub_admin":
        # Iter 124 — sub admins work across all firms in their scope.
        from server import sub_admin_can_touch_company
        if not company_id:
            raise HTTPException(status_code=400, detail="company_id is required")
        if not sub_admin_can_touch_company(user, company_id):
            raise HTTPException(status_code=403, detail="Firm is outside your assigned scope")
        return company_id
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required for Super Admin")
    return company_id


# ---------------------------------------------------------------------------
# Iter 96u — Bank Sheet (salary transfer statement) for the Actual Salary
# Process. Net Salary is caught from the COMPLIANCE Salary run ('net' field).
# Columns: S.No, Name, Father Name, Bank Name, Name-as-per-bank, IFSC,
# Account No, Net Salary. Filters: month, employee_type, pay_mode, bank_name.
# ---------------------------------------------------------------------------
async def _bank_sheet_rows(
    user: Dict[str, Any], company_id: Optional[str], month: str,
    employee_type: Optional[str], pay_mode: Optional[str], bank_name: Optional[str],
) -> Dict[str, Any]:
    cid = _scope_company(user, company_id)
    # Latest compliance run(s) for this firm + month (optionally a type).
    q: Dict[str, Any] = {"company_id": cid, "month": month}
    runs = await db.compliance_salary_runs.find(q, {"_id": 0}).sort("created_at", -1).to_list(50)
    # Keep the most recent run per employee_type so we don't double-count.
    net_by_user: Dict[str, float] = {}
    seen_types: set = set()
    for run in runs:
        rtype = run.get("employee_type") or "__all__"
        if rtype in seen_types:
            continue
        seen_types.add(rtype)
        for r in run.get("rows", []):
            uid = r.get("user_id")
            if uid and uid not in net_by_user:
                net_by_user[uid] = float(r.get("net") or 0)

    uids = list(net_by_user.keys())
    emp_q: Dict[str, Any] = {"user_id": {"$in": uids}}
    if employee_type and employee_type != "all":
        emp_q["employee_type"] = employee_type
    emps = await db.users.find(emp_q, {
        "_id": 0, "user_id": 1, "name": 1, "father_name": 1, "employee_code": 1,
        "bank_name": 1, "bank_account_name": 1, "bank_ifsc": 1, "bank_account": 1,
        "pay_mode": 1, "employee_type": 1,
    }).to_list(5000)

    rows: List[Dict[str, Any]] = []
    for e in emps:
        pm = (e.get("pay_mode") or "Bank")
        if pay_mode and pay_mode != "all" and pm.lower() != pay_mode.lower():
            continue
        bn = (e.get("bank_name") or "")
        if bank_name and bank_name != "all" and bn != bank_name:
            continue
        rows.append({
            "user_id": e.get("user_id"),
            "employee_code": e.get("employee_code") or "",
            "name": e.get("name") or "",
            "father_name": e.get("father_name") or "",
            "bank_name": bn,
            "name_as_per_bank": e.get("bank_account_name") or "",
            "ifsc": e.get("bank_ifsc") or "",
            "account_no": e.get("bank_account") or "",
            "pay_mode": pm,
            "net_salary": round(net_by_user.get(e.get("user_id"), 0), 2),
        })

    def _key(r):
        c = r["employee_code"]
        return (0, int(c)) if c.isdigit() else (1, c or r["name"])
    rows.sort(key=_key)
    for i, r in enumerate(rows, 1):
        r["sn"] = i

    # Distinct bank names (for the filter dropdown) across the firm.
    banks = sorted({(e.get("bank_name") or "").strip() for e in emps if e.get("bank_name")})
    total = round(sum(r["net_salary"] for r in rows), 2)
    return {
        "month": month, "company_id": cid, "rows": rows,
        "count": len(rows), "total_net": total, "banks": banks,
        "has_compliance": len(net_by_user) > 0,
    }


@router.get("/bank-sheet")
async def bank_sheet(
    month: str,
    company_id: Optional[str] = None,
    employee_type: Optional[str] = None,
    pay_mode: Optional[str] = None,
    bank_name: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    return await _bank_sheet_rows(user, company_id, month, employee_type, pay_mode, bank_name)


@router.get("/bank-sheet.xlsx")
async def bank_sheet_xlsx(
    month: str,
    company_id: Optional[str] = None,
    employee_type: Optional[str] = None,
    pay_mode: Optional[str] = None,
    bank_name: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    data = await _bank_sheet_rows(user, company_id, month, employee_type, pay_mode, bank_name)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Bank Sheet"
    headers = ["S.No.", "Name", "Father Name", "Bank Name", "Name as per Bank", "IFSC Code", "Account No.", "Net Salary"]
    ws.append(headers)
    fill = PatternFill(start_color="0F2E3D", end_color="0F2E3D", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill; c.alignment = Alignment(horizontal="center")
    for r in data["rows"]:
        ws.append([r["sn"], r["name"], r["father_name"], r["bank_name"],
                   r["name_as_per_bank"], r["ifsc"], r["account_no"], r["net_salary"]])
    ws.append(["", "", "", "", "", "", "TOTAL", data["total_net"]])
    ws.cell(row=ws.max_row, column=7).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=8).font = Font(bold=True)
    for i, w in enumerate([7, 26, 24, 22, 26, 16, 20, 14], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"bank-sheet-{data['company_id']}-{month}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



@router.post("/challans")
async def create_challan(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    portal = (payload.get("portal") or "").lower()
    if portal not in ("pf", "esic"):
        raise HTTPException(status_code=400, detail="portal must be 'pf' or 'esic'")
    month = (payload.get("month") or "").strip()
    if len(month) != 7 or month[4] != "-":
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")

    company_id = _scope_company(user, payload.get("company_id"))

    # Iter 126h — challans can only be uploaded once the Compliance Salary
    # for that firm & month is FINALIZED (super admin exempt).
    if user.get("role") != "super_admin":
        fin = await db.compliance_salary_runs.find_one(
            {"company_id": company_id, "month": month, "finalized": True},
            {"run_id": 1},
        )
        if not fin:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Compliance salary for this firm & month is not finalized "
                    "yet. Finalize the salary run before uploading challans."
                ),
            )

    file_b64 = payload.get("file_base64") or ""
    file_mime = (payload.get("file_mime") or "").lower()
    file_name = (payload.get("file_name") or "").strip()
    if not file_b64:
        raise HTTPException(status_code=400, detail="file_base64 is required")
    if file_mime not in ALLOWED_MIMES:
        raise HTTPException(
            status_code=400,
            detail=f"file_mime must be one of {sorted(ALLOWED_MIMES.keys())}",
        )
    if "," in file_b64 and file_b64.startswith("data:"):
        file_b64 = file_b64.split(",", 1)[1]
    # Cap at 8 MB
    if len(file_b64) > 12 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max ~8 MB)")

    try:
        amount = float(payload.get("amount") or 0)
    except (TypeError, ValueError):
        amount = 0.0

    doc = {
        "challan_id": f"chl_{uuid.uuid4().hex[:12]}",
        "company_id": company_id,
        "portal": portal,
        "month": month,
        "amount": amount,
        "trrn": (payload.get("trrn") or "").strip() or None,
        "paid_on": (payload.get("paid_on") or "").strip() or None,
        "notes": (payload.get("notes") or "").strip() or None,
        "file_base64": file_b64,
        "file_mime": file_mime,
        "file_name": file_name or f"{portal}-{month}{ALLOWED_MIMES[file_mime]}",
        "created_by": user["user_id"],
        "created_at": now_iso(),
    }
    await db.challans.insert_one(doc)
    logger.info(
        "[challans] %s uploaded portal=%s month=%s amount=%s by %s",
        doc["challan_id"], portal, month, amount, user["user_id"],
    )
    doc.pop("_id", None)
    return {"ok": True, "challan": {k: v for k, v in doc.items() if k != "file_base64"}}


@router.get("/challans")
async def list_challans(
    company_id: Optional[str] = None,
    portal: Optional[str] = None,
    month: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    q: Dict[str, Any] = {}
    if user["role"] == "company_admin":
        q["company_id"] = user.get("company_id")
    elif company_id:
        q["company_id"] = company_id
    if portal and portal.lower() in ("pf", "esic"):
        q["portal"] = portal.lower()
    if month:
        q["month"] = month

    rows: List[Dict[str, Any]] = []
    async for c in db.challans.find(q, {"_id": 0, "file_base64": 0}).sort("month", -1):
        rows.append(c)
    return {"challans": rows}


@router.get("/challans/export.xlsx")
async def export_challans_xlsx(
    company_id: Optional[str] = None,
    portal: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    q: Dict[str, Any] = {}
    if user["role"] == "company_admin":
        q["company_id"] = user.get("company_id")
    elif company_id:
        q["company_id"] = company_id
    if portal and portal.lower() in ("pf", "esic"):
        q["portal"] = portal.lower()

    rows = []
    async for c in db.challans.find(q, {"_id": 0, "file_base64": 0}).sort("month", -1):
        rows.append(c)

    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=503, detail="openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "Challans"
    ws.append([
        "Challan ID", "Firm", "Portal", "Month", "Amount",
        "TRRN", "Paid On", "Uploaded At", "Notes", "File Name",
    ])
    # Fetch firm names in a single query
    firm_ids = {r.get("company_id") for r in rows if r.get("company_id")}
    firms = {c["company_id"]: c.get("name") async for c in
             db.companies.find({"company_id": {"$in": list(firm_ids)}},
                               {"_id": 0, "company_id": 1, "name": 1})}
    for r in rows:
        ws.append([
            r.get("challan_id"),
            firms.get(r.get("company_id"), r.get("company_id") or ""),
            (r.get("portal") or "").upper(),
            r.get("month"),
            r.get("amount") or 0,
            r.get("trrn") or "",
            r.get("paid_on") or "",
            r.get("created_at") or "",
            r.get("notes") or "",
            r.get("file_name") or "",
        ])
    for col in "ABCDEFGHIJ":
        ws.column_dimensions[col].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"challans-{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# Iter 95i — EPFO / ESIC portal upload files generated from a compliance
# salary run.
#   * ecr.txt   — EPFO ECR 2.0 text file (#~# separated, one line per member)
#   * ecr.xlsx  — the same member data in Excel for review
#   * esic.xlsx — ESIC monthly contribution upload format
# ---------------------------------------------------------------------------

async def _load_run_for_portal(run_id: str, user: Dict[str, Any]) -> Dict[str, Any]:
    run = await db.compliance_salary_runs.find_one({"run_id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Compliance run not found")
    if user["role"] in ("company_admin", "sub_admin") and run.get("company_id"):
        _scope_company(user, run.get("company_id"))
    return run


async def _uan_esic_map(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    uids = [r.get("user_id") for r in rows if r.get("user_id")]
    users = await db.users.find(
        {"user_id": {"$in": uids}},
        {"_id": 0, "user_id": 1, "uan_no": 1, "esi_ip_no": 1},
    ).to_list(5000)
    return {u["user_id"]: u for u in users}


def _r0(v: Any) -> int:
    try:
        return int(round(float(v or 0)))
    except Exception:
        return 0


def _ecr_lines(run: Dict[str, Any], extra: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """One dict per PF member with all ECR 2.0 columns computed."""
    month_days = int(run.get("month_days") or 30)
    eps_cap = 15000
    out: List[Dict[str, Any]] = []
    for r in run.get("rows") or []:
        if not r.get("pf_applicable"):
            continue
        uan = str((extra.get(r.get("user_id"), {}) or {}).get("uan_no") or "").strip()
        gross = _r0(r.get("gross_paid"))
        epf_wages = _r0(r.get("pf_wages"))
        eps_wages = min(epf_wages, eps_cap)
        edli_wages = min(epf_wages, eps_cap)
        epf_ee = _r0(r.get("pf_employee"))
        eps_er = _r0(r.get("pf_employer_eps"))
        diff_er = _r0(r.get("pf_employer_epf"))
        present = float(r.get("present_days") or 0)
        ncp = max(0, round(month_days - present, 1))
        ncp = int(ncp) if float(ncp).is_integer() else ncp
        out.append({
            "uan": uan,
            "name": (r.get("name") or "").upper(),
            "employee_code": r.get("employee_code") or "",
            "gross": gross,
            "epf_wages": epf_wages,
            "eps_wages": eps_wages,
            "edli_wages": edli_wages,
            "epf_ee": epf_ee,
            "eps_er": eps_er,
            "diff_er": diff_er,
            "ncp": ncp,
            "refund": 0,
        })
    return out


def _ecr_txt_bytes(run: Dict[str, Any], extra: Dict[str, Dict[str, Any]]) -> bytes:
    """EPFO Contribution file body (6-field #~# format).

    Members WITH a UAN are written normally; PF members WITHOUT a UAN yet
    (new joiners) are still included with a BLANK UAN field so EPFO can
    assign one on upload (per user request). Only raises 400 if there are
    no PF-applicable members at all."""
    members = _ecr_lines(run, extra)
    if not members:
        raise HTTPException(
            status_code=400,
            detail="No PF-applicable members in this run.",
        )
    lines = [
        "#~#".join(str(x) for x in (
            m["uan"], m["name"], m["epf_ee"], m["eps_er"], m["diff_er"], m["refund"],
        ))
        for m in members
    ]
    return ("\r\n".join(lines) + "\r\n").encode("utf-8")


def _esic_xls_bytes(run: Dict[str, Any], extra: Dict[str, Dict[str, Any]]) -> bytes:
    """ESIC monthly-contribution bulk sheet (.xls) body. Raises 400 if no
    ESIC-applicable member has an IP number."""
    import xlwt
    wb = xlwt.Workbook()
    ws = wb.add_sheet("ESIC")
    for col, h in enumerate(["ESI_CODE", "NAME", "DAYS", "SAL", "RE", "DATE"]):
        ws.write(0, col, h)
    rownum = 1
    for r in run.get("rows") or []:
        if not r.get("esic_applicable"):
            continue
        ip_no = str((extra.get(r.get("user_id"), {}) or {}).get("esi_ip_no") or "").strip()
        if not ip_no:
            continue
        present = float(r.get("present_days") or 0)
        days = int(present) if present.is_integer() else present
        wages = _r0(r.get("esic_wage_base") or r.get("gross_paid"))
        if days <= 0:
            wages = 0
        ws.write(rownum, 0, ip_no)
        ws.write(rownum, 1, (r.get("name") or "").upper())
        ws.write(rownum, 2, days)
        ws.write(rownum, 3, wages)
        ws.write(rownum, 4, 1 if days <= 0 else 0)
        ws.write(rownum, 5, "")
        rownum += 1
    if rownum == 1:
        raise HTTPException(
            status_code=400,
            detail="No ESIC-applicable employees with an ESIC/IP number in this run — fill ESIC numbers in the Employee Master first",
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/challans/ecr.txt")
async def download_ecr_txt(
    run_id: str,
    authorization: Optional[str] = Header(None),
):
    """EPFO Contribution file (.txt) — EXACT 6-field format from the EPFO
    portal's CONTRIBUTION_HELP_FILE:
    ``UAN#~#MEMBER NAME#~#EPF EE CONTRI#~#EPS ER CONTRI#~#EPF ER CONTRI#~#REFUND``
    e.g. ``123467198618#~#VIRAT SHARMA#~#300#~#100#~#50#~#0``.
    Members without a UAN in the Employee Master are SKIPPED."""
    user = await _assert_admin(authorization)
    run = await _load_run_for_portal(run_id, user)
    extra = await _uan_esic_map(run.get("rows") or [])
    content = _ecr_txt_bytes(run, extra)
    fname = f"ECR_{run.get('month') or 'month'}.txt"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/challans/ecr.xlsx")
async def download_ecr_xlsx(
    run_id: str,
    authorization: Optional[str] = Header(None),
):
    """The ECR member data as Excel — for checking before uploading the
    .txt on the EPFO portal (includes members missing UAN, highlighted)."""
    user = await _assert_admin(authorization)
    run = await _load_run_for_portal(run_id, user)
    extra = await _uan_esic_map(run.get("rows") or [])
    members = _ecr_lines(run, extra)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = f"ECR {run.get('month') or ''}"
    headers = ["UAN", "Member Name", "EPF Contribution (EE)",
               "EPS Contribution (ER)", "EPF Contribution (ER)",
               "Refund of Advances",
               "Emp Code", "Gross Wages", "EPF Wages", "EPS Wages",
               "EDLI Wages", "NCP Days"]
    ws.append(headers)
    hdr_fill = PatternFill(start_color="0F2E3D", end_color="0F2E3D", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
    warn_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    for m in members:
        ws.append([m["uan"] or "MISSING UAN", m["name"],
                   m["epf_ee"], m["eps_er"], m["diff_er"], m["refund"],
                   m["employee_code"], m["gross"], m["epf_wages"],
                   m["eps_wages"], m["edli_wages"], m["ncp"]])
        if not m["uan"]:
            for c in ws[ws.max_row]:
                c.fill = warn_fill
    totals_row = ["", "TOTAL",
                  sum(m["epf_ee"] for m in members),
                  sum(m["eps_er"] for m in members),
                  sum(m["diff_er"] for m in members),
                  sum(m["refund"] for m in members),
                  "",
                  sum(m["gross"] for m in members),
                  sum(m["epf_wages"] for m in members),
                  sum(m["eps_wages"] for m in members),
                  sum(m["edli_wages"] for m in members), ""]
    ws.append(totals_row)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for i, w in enumerate([16, 30, 18, 18, 18, 16, 10, 12, 12, 12, 12, 10], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ECR_{run.get('month') or 'month'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/challans/esic.xls")
async def download_esic_xls(
    run_id: str,
    authorization: Optional[str] = Header(None),
):
    """ESIC monthly-contribution upload sheet (.xls) — EXACT format from the
    firm's sample file: columns ``ESI_CODE, NAME, DAYS, SAL, RE, DATE``.
    RE = 1 for zero working days (else 0); DATE = last working day, blank
    unless the employee exited. Members without an ESIC number are SKIPPED
    (use /challans/esic-check.xlsx to review them)."""
    user = await _assert_admin(authorization)
    run = await _load_run_for_portal(run_id, user)
    extra = await _uan_esic_map(run.get("rows") or [])
    content = _esic_xls_bytes(run, extra)
    fname = f"ESIC_MC_{run.get('month') or 'month'}.xls"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.ms-excel",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/challans/esic.xlsx")
async def download_esic_xlsx(
    run_id: str,
    authorization: Optional[str] = Header(None),
):
    """ESIC review Excel (same ESI_CODE/NAME/DAYS/SAL/RE/DATE columns) —
    INCLUDES employees missing an ESIC number (highlighted) so the admin can
    fix the Employee Master before generating the portal .xls."""
    user = await _assert_admin(authorization)
    run = await _load_run_for_portal(run_id, user)
    extra = await _uan_esic_map(run.get("rows") or [])
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = f"ESIC {run.get('month') or ''}"
    ws.append(["ESI_CODE", "NAME", "DAYS", "SAL", "RE", "DATE"])
    hdr_fill = PatternFill(start_color="0F2E3D", end_color="0F2E3D", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
    warn_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    n = 0
    for r in run.get("rows") or []:
        if not r.get("esic_applicable"):
            continue
        ip_no = str((extra.get(r.get("user_id"), {}) or {}).get("esi_ip_no") or "").strip()
        present = float(r.get("present_days") or 0)
        days = int(present) if present.is_integer() else present
        wages = _r0(r.get("esic_wage_base") or r.get("gross_paid"))
        if days <= 0:
            wages = 0
        ws.append([ip_no or "MISSING IP NO", (r.get("name") or "").upper(),
                   days, wages, 1 if days <= 0 else 0, ""])
        if not ip_no:
            for c in ws[ws.max_row]:
                c.fill = warn_fill
        n += 1
    if n == 0:
        raise HTTPException(status_code=400, detail="No ESIC-applicable employees in this run")
    for i, w in enumerate([16, 32, 10, 14, 8, 16], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ESIC_MC_{run.get('month') or 'month'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# Iter 96e — Missing statutory numbers report (employees without UAN / ESI IP
# number) so the team can fill gaps before generating portal files.
# ---------------------------------------------------------------------------

async def _missing_statutory_rows(user: Dict[str, Any], company_id: Optional[str]) -> List[Dict[str, Any]]:
    q: Dict[str, Any] = {"role": "employee", "disabled": {"$ne": True}}
    if user["role"] == "company_admin":
        q["company_id"] = user.get("company_id")
    elif company_id:
        q["company_id"] = company_id
    q["$or"] = [
        {"uan_no": {"$in": [None, ""]}},
        {"esi_ip_no": {"$in": [None, ""]}},
    ]
    rows: List[Dict[str, Any]] = []
    async for u in db.users.find(q, {
        "_id": 0, "user_id": 1, "company_id": 1, "employee_code": 1,
        "name": 1, "employee_type": 1, "uan_no": 1, "esi_ip_no": 1,
    }):
        rows.append({
            "user_id": u.get("user_id"),
            "company_id": u.get("company_id"),
            "employee_code": u.get("employee_code") or "",
            "name": u.get("name") or "",
            "employee_type": u.get("employee_type") or "",
            "uan_no": (u.get("uan_no") or "").strip(),
            "esi_ip_no": (u.get("esi_ip_no") or "").strip(),
        })

    def _code_key(r: Dict[str, Any]):
        c = r["employee_code"]
        return (0, int(c)) if c.isdigit() else (1, c)
    rows.sort(key=_code_key)
    return rows


@router.get("/challans/missing-statutory")
async def missing_statutory(
    company_id: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    rows = await _missing_statutory_rows(user, company_id)
    return {
        "total": len(rows),
        "missing_uan": sum(1 for r in rows if not r["uan_no"]),
        "missing_esi": sum(1 for r in rows if not r["esi_ip_no"]),
        "employees": rows,
    }


@router.get("/challans/missing-statutory.xlsx")
async def missing_statutory_xlsx(
    company_id: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    rows = await _missing_statutory_rows(user, company_id)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl not installed")
    firm_ids = {r["company_id"] for r in rows if r.get("company_id")}
    firms = {c["company_id"]: c.get("name") async for c in
             db.companies.find({"company_id": {"$in": list(firm_ids)}},
                               {"_id": 0, "company_id": 1, "name": 1})}
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Statutory Nos"
    ws.append(["Emp Code", "Name", "Group", "Firm", "UAN", "ESI IP No"])
    hdr_fill = PatternFill(start_color="0F2E3D", end_color="0F2E3D", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
    warn_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    for r in rows:
        ws.append([
            r["employee_code"], r["name"], r["employee_type"],
            firms.get(r.get("company_id"), r.get("company_id") or ""),
            r["uan_no"] or "MISSING", r["esi_ip_no"] or "MISSING",
        ])
        for col in (5, 6):
            cell = ws.cell(row=ws.max_row, column=col)
            if cell.value == "MISSING":
                cell.fill = warn_fill
                cell.font = Font(bold=True, color="92400E")
    for i, w in enumerate([10, 30, 12, 24, 18, 18], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"missing-statutory-{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/challans/{challan_id}")
async def get_challan(
    challan_id: str,
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    doc = await db.challans.find_one({"challan_id": challan_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Challan not found")
    if user["role"] == "company_admin" and doc.get("company_id") != user.get("company_id"):
        raise HTTPException(status_code=403, detail="Not your firm's challan")
    if user["role"] == "sub_admin":
        from server import sub_admin_can_touch_company
        if not sub_admin_can_touch_company(user, doc.get("company_id")):
            raise HTTPException(status_code=403, detail="Firm is outside your assigned scope")
    return {"challan": doc}


@router.delete("/challans/{challan_id}")
async def delete_challan(
    challan_id: str,
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    doc = await db.challans.find_one({"challan_id": challan_id}, {"_id": 0, "company_id": 1})
    if not doc:
        raise HTTPException(status_code=404, detail="Challan not found")
    if user["role"] == "company_admin" and doc.get("company_id") != user.get("company_id"):
        raise HTTPException(status_code=403, detail="Not your firm's challan")
    if user["role"] == "sub_admin":
        from server import sub_admin_can_touch_company
        if not sub_admin_can_touch_company(user, doc.get("company_id")):
            raise HTTPException(status_code=403, detail="Firm is outside your assigned scope")
    await db.challans.delete_one({"challan_id": challan_id})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Portal upload AUTOMATION (user directive) — queue an RPA job that logs in
# to EPFO / ESIC with the Firm Master credentials (AI captcha reading),
# uploads the generated ECR .txt / ESIC bulk .xls, and STOPS at challan
# finalisation. Bank payment is NEVER automated.
# ---------------------------------------------------------------------------

_UPLOAD_ACTIONS = {"epfo": "upload_ecr", "esic": "upload_esic_mc"}


@router.post("/portal-upload-jobs")
async def create_portal_upload_job(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    run_id = (payload.get("run_id") or "").strip()
    portal = (payload.get("portal") or "").strip().lower()
    if portal not in _UPLOAD_ACTIONS:
        raise HTTPException(status_code=400, detail="portal must be 'epfo' or 'esic'")
    if not run_id:
        raise HTTPException(status_code=400, detail="run_id is required")
    run = await _load_run_for_portal(run_id, user)
    company_id = run.get("company_id")
    if not company_id:
        raise HTTPException(status_code=400, detail="This run has no firm attached")

    from utils.rpa_worker import _fetch_creds
    creds = await _fetch_creds(db, company_id, portal)
    if not creds:
        raise HTTPException(
            status_code=400,
            detail=(f"No {portal.upper()} User ID/Password saved in Firm Master "
                    "— fill them in the EPF/ESIC Detail section first"),
        )

    extra = await _uan_esic_map(run.get("rows") or [])
    month = run.get("month") or "month"
    if portal == "epfo":
        content = _ecr_txt_bytes(run, extra)
        file_name = f"ECR_{month}.txt"
    else:
        content = _esic_xls_bytes(run, extra)
        file_name = f"ESIC_MC_{month}.xls"

    job_id = f"puj_{uuid.uuid4().hex[:12]}"
    await db.portal_automation_jobs.insert_one({
        "job_id": job_id,
        "company_id": company_id,
        "portal": portal,
        "action_type": _UPLOAD_ACTIONS[portal],
        "run_id": run_id,
        "month": month,
        "file_name": file_name,
        "file_b64": base64.b64encode(content).decode("ascii"),
        "status": "pending",
        "steps": [{
            "at": now_iso(),
            "msg": (f"Queued {portal.upper()} upload of {file_name} "
                    "(stops at challan finalisation — no bank payment)."),
        }],
        "created_by": user["user_id"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    })
    logger.info("[challans] queued portal upload job=%s portal=%s run=%s",
                job_id, portal, run_id)
    return {"ok": True, "job_id": job_id, "file_name": file_name, "status": "pending"}


@router.get("/portal-upload-jobs")
async def list_portal_upload_jobs(
    company_id: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    user = await _assert_admin(authorization)
    q: Dict[str, Any] = {"action_type": {"$in": list(_UPLOAD_ACTIONS.values())}}
    if user["role"] == "company_admin":
        q["company_id"] = user.get("company_id")
    elif company_id:
        q["company_id"] = company_id
    jobs = await db.portal_automation_jobs.find(
        q, {"_id": 0, "file_b64": 0},
    ).sort("created_at", -1).to_list(20)
    # Strip heavy screenshots out of the step log for the list view.
    for j in jobs:
        j["steps"] = [
            {"at": s.get("at"), "msg": s.get("msg"),
             "has_screenshot": bool(s.get("screenshot_base64"))}
            for s in (j.get("steps") or [])
        ]
    return {"jobs": jobs}


@router.get("/portal-upload-jobs/{job_id}/file")
async def download_portal_upload_file(
    job_id: str,
    authorization: Optional[str] = Header(None),
):
    """Manual fallback — download the exact file the job generated."""
    user = await _assert_admin(authorization)
    job = await db.portal_automation_jobs.find_one(
        {"job_id": job_id}, {"_id": 0, "company_id": 1, "file_b64": 1, "file_name": 1})
    if not job or not job.get("file_b64"):
        raise HTTPException(status_code=404, detail="Job or file not found")
    if user["role"] == "company_admin" and job.get("company_id") != user.get("company_id"):
        raise HTTPException(status_code=403, detail="Not your firm's job")
    if user["role"] == "sub_admin":
        from server import sub_admin_can_touch_company
        if not sub_admin_can_touch_company(user, job.get("company_id")):
            raise HTTPException(status_code=403, detail="Firm is outside your assigned scope")
    content = base64.b64decode(job["file_b64"])
    fname = job.get("file_name") or "upload.bin"
    mime = "text/plain" if fname.endswith(".txt") else "application/vnd.ms-excel"
    return StreamingResponse(
        io.BytesIO(content),
        media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/challans-portal-preview")
async def portal_preview(
    run_id: str,
    kind: str = "epfo",
    authorization: Optional[str] = Header(None),
):
    """Iter 161 — on-screen data preview BEFORE the auto portal upload.
    Returns the exact member lines that will go to the portal."""
    user = await _assert_admin(authorization)
    run = await _load_run_for_portal(run_id, user)
    extra = await _uan_esic_map(run.get("rows") or [])
    if kind == "esic":
        lines = []
        for r in run.get("rows") or []:
            if not r.get("esic_applicable"):
                continue
            ex = extra.get(r.get("user_id"), {}) or {}
            days = float(r.get("present_days") or 0)
            lines.append({
                "ip_no": str(ex.get("esi_ip_no") or "").strip(),
                "name": (r.get("name") or "").upper(),
                "days": int(days) if days.is_integer() else days,
                "wages": round(float(r.get("gross_paid") or 0), 2),
                "ee": int(round(float(r.get("esic_employee") or 0))),
                "skipped": not str(ex.get("esi_ip_no") or "").strip(),
            })
        totals = {
            "members": len(lines),
            "uploadable": sum(1 for x in lines if not x["skipped"]),
            "wages": round(sum(x["wages"] for x in lines), 2),
            "ee": sum(x["ee"] for x in lines),
        }
        return {"kind": "esic", "month": run.get("month"), "lines": lines, "totals": totals}
    members = _ecr_lines(run, extra)
    lines = [{
        "uan": m["uan"] or "", "name": m["name"], "gross": m["gross"],
        "epf_wages": m["epf_wages"], "eps_wages": m["eps_wages"],
        "edli_wages": m["edli_wages"], "epf_ee": m["epf_ee"],
        "eps_er": m["eps_er"], "diff_er": m["diff_er"],
        "ncp": m["ncp"], "skipped": not (m["uan"] or "").strip(),
    } for m in members]
    totals = {
        "members": len(lines),
        "uploadable": sum(1 for x in lines if not x["skipped"]),
        "epf_ee": sum(x["epf_ee"] for x in lines),
        "eps_er": sum(x["eps_er"] for x in lines),
        "diff_er": sum(x["diff_er"] for x in lines),
    }
    return {"kind": "epfo", "month": run.get("month"), "lines": lines, "totals": totals}
