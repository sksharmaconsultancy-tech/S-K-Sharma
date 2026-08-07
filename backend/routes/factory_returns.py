"""Iter 499 — FACTORY & BOILER ANNUAL RETURN (Compliance → Statutory Returns).

Government-style annual returns computed from EXISTING payroll + attendance
data with a UNIFIED DATA LAYER:
  * CURRENT data  = compliance_salary_runs (frozen/processed months)
  * LEGACY data   = compliance_import_entries (imported old-database months)
  * COMBINED (default) merges both WITHOUT duplication — a legacy entry is
    used only when the same employee+month has no current run row.
Legacy records are READ-ONLY — this module never writes to them.

Endpoints:
  GET  /api/admin/factory-return/details/{company_id}     — statutory particulars
  PUT  /api/admin/factory-return/details/{company_id}     — save particulars
  GET  /api/admin/factory-return/{company_id}/{year}      — full computed return
  GET  /api/admin/factory-return/{company_id}/{year}.pdf  — Factory Return PDF
  GET  /api/admin/factory-return/{company_id}/{year}/boiler.pdf — Boiler Return
  GET  /api/admin/factory-return/{company_id}/{year}.xlsx — Excel workbook
"""
import io
import sys
from collections import defaultdict
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException, Query
from fastapi.responses import Response

sys.path.append("/app/backend")
from server import db, get_user_from_token, require_role, now_iso  # noqa: E402

router = APIRouter(prefix="/api/admin/factory-return", tags=["factory-return"])

ADMIN_ROLES = ["super_admin", "sub_admin", "company_admin"]

DETAIL_FIELDS = [
    "factory_name", "factory_address", "factory_license_no",
    "factory_registration_no", "boiler_registration_no", "occupier_name",
    "factory_manager", "nature_of_manufacturing", "district", "state",
]
WELFARE_FIELDS = ["canteen", "rest_room", "creche", "first_aid",
                  "ambulance_room", "drinking_water", "washing_facility"]

# Iter 520 (user upload — official FORM NO. 23, Rule 105(i)) — statutory
# particulars that cannot be computed from payroll data. Saved on the firm
# under factory_details.form23 and printed on the Form 23 PDF.
FORM23_FIELDS = [
    "application_no", "area", "dangerous_process",
    "safety_officers_required", "safety_officers_appointed",
    "ambulance_room", "canteen", "canteen_departmental", "canteen_contractor",
    "rest_rooms", "lunch_rooms", "creche",
    "welfare_officers_required", "welfare_officers_appointed",
    "safety_trainings", "safety_trained_male", "safety_trained_female",
    "acc_ret_same_count", "acc_ret_same_mandays",
    "acc_ret_prev_count", "acc_ret_prev_mandays",
    "acc_not_ret_count", "acc_not_ret_mandays",
    "fines", "deduction_damage", "deduction_breach",
    "bonus_paid", "money_concessions",
    "left_service", "wages_in_lieu_paid",
]


async def _admin(authorization: Optional[str]):
    user = await get_user_from_token(authorization)
    require_role(user, ADMIN_ROLES)
    return user


async def _company(cid: str) -> Dict[str, Any]:
    c = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Company not found")
    return c


def _details_of(c: Dict[str, Any]) -> Dict[str, Any]:
    fd = c.get("factory_details") or {}
    out = {k: fd.get(k) or "" for k in DETAIL_FIELDS}
    out["factory_name"] = out["factory_name"] or c.get("name") or ""
    out["factory_address"] = out["factory_address"] or c.get("address") or ""
    out["welfare"] = {k: bool((fd.get("welfare") or {}).get(k)) for k in WELFARE_FIELDS}
    out["accidents"] = fd.get("accidents") or {}
    out["form23"] = {k: str((fd.get("form23") or {}).get(k) or "")
                     for k in FORM23_FIELDS}
    return out


@router.get("/details/{company_id}")
async def get_details(company_id: str, authorization: Optional[str] = Header(None)):
    await _admin(authorization)
    return {"details": _details_of(await _company(company_id))}


@router.put("/details/{company_id}")
async def put_details(company_id: str, body: Dict[str, Any],
                      authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    await _company(company_id)
    fd: Dict[str, Any] = {k: str(body.get(k) or "").strip() for k in DETAIL_FIELDS}
    fd["welfare"] = {k: bool((body.get("welfare") or {}).get(k)) for k in WELFARE_FIELDS}
    fd["form23"] = {k: str((body.get("form23") or {}).get(k) or "").strip()
                    for k in FORM23_FIELDS}
    acc = body.get("accidents") or {}
    fd["accidents"] = {
        str(y): {
            "fatal": int((v or {}).get("fatal") or 0),
            "nonfatal": int((v or {}).get("nonfatal") or 0),
            "mandays_lost": int((v or {}).get("mandays_lost") or 0),
        } for y, v in acc.items() if str(y).isdigit()
    }
    await db.companies.update_one(
        {"company_id": company_id},
        {"$set": {"factory_details": fd,
                  "factory_details_updated_at": now_iso(),
                  "factory_details_updated_by": admin["user_id"]}})
    return {"ok": True, "details": fd}


# ---------------------------------------------------------------------------
# Unified data layer — merge current runs + legacy import entries
# ---------------------------------------------------------------------------
async def _collect_rows(cid: str, year: int, source: str):
    """Returns per-month merged salary rows:
    {month: {user_id: {present_days, gross, ot_hours, ot_pay, leave_days,
                       month_days, origin}}}"""
    months = [f"{year}-{m:02d}" for m in range(1, 13)]
    data: Dict[str, Dict[str, Dict[str, Any]]] = {m: {} for m in months}

    if source in ("combined", "current"):
        async for run in db.compliance_salary_runs.find(
                {"company_id": cid, "month": {"$in": months}},
                {"_id": 0, "month": 1, "month_days": 1, "rows": 1}):
            m = run["month"]
            for r in run.get("rows") or []:
                uid = r.get("user_id")
                if not uid:
                    continue
                data[m][uid] = {
                    "present_days": float(r.get("present_days") or 0),
                    "gross": float(r.get("gross_paid") or r.get("monthly_gross") or 0),
                    "basic": float(r.get("basic") or 0),
                    "ot_hours": float(r.get("ot_hours") or 0),
                    "ot_pay": float(r.get("ot_pay") or 0),
                    "leave_days": float(r.get("esic_leave_days") or 0),
                    "month_days": float(run.get("month_days") or 30),
                    "origin": "current",
                }

    if source in ("combined", "legacy"):
        async for e in db.compliance_import_entries.find(
                {"company_id": cid, "month": {"$in": months}},
                {"_id": 0, "month": 1, "user_id": 1, "present_days": 1,
                 "gross_earning": 1, "ot_hours": 1}):
            m, uid = e["month"], e.get("user_id")
            if not uid or uid in data[m]:
                continue  # duplicate prevention — current data wins
            data[m][uid] = {
                "present_days": float(e.get("present_days") or 0),
                "gross": float(e.get("gross_earning") or 0),
                "basic": 0.0,
                "ot_hours": float(e.get("ot_hours") or 0),
                "ot_pay": 0.0,
                "leave_days": 0.0,
                "month_days": 30.0,
                "origin": "legacy",
            }
    return data


async def _compute_return(cid: str, year: int, source: str) -> Dict[str, Any]:
    c = await _company(cid)
    details = _details_of(c)
    data = await _collect_rows(cid, year, source)

    # employee master lookup (indexed by user_id) — cached single fetch
    emp: Dict[str, Dict[str, Any]] = {}
    async for u in db.users.find(
            {"company_id": cid, "role": "employee"},
            {"_id": 0, "user_id": 1, "employee_code": 1, "name": 1, "gender": 1,
             "doj": 1, "employee_type": 1, "department": 1, "designation": 1,
             "contractor_name": 1, "is_onroll": 1, "approval_status": 1,
             "deleted": 1}):
        emp[u["user_id"]] = u

    monthly: List[Dict[str, Any]] = []
    per_user: Dict[str, Dict[str, float]] = defaultdict(
        lambda: {"man_days": 0.0, "wages": 0.0, "ot_hours": 0.0,
                 "ot_pay": 0.0, "leave": 0.0, "months": 0.0})
    tot = {"man_days": 0.0, "wages": 0.0, "basic": 0.0, "ot_hours": 0.0,
           "ot_pay": 0.0, "leave": 0.0, "working_days": 0.0}
    max_emp = 0
    for m in sorted(data.keys()):
        rows = data[m]
        if not rows:
            continue
        md = sum(r["present_days"] for r in rows.values())
        wg = sum(r["gross"] for r in rows.values())
        bs = sum(r["basic"] for r in rows.values())
        oh = sum(r["ot_hours"] for r in rows.values())
        op = sum(r["ot_pay"] for r in rows.values())
        lv = sum(r["leave_days"] for r in rows.values())
        wd = max((r["month_days"] for r in rows.values()), default=30)
        legacy_n = sum(1 for r in rows.values() if r["origin"] == "legacy")
        monthly.append({
            "month": m, "employees": len(rows),
            "man_days": round(md, 1), "working_days": wd,
            "avg_daily_employment": round(md / wd, 1) if wd else 0,
            "wages": round(wg, 2), "ot_hours": round(oh, 1),
            "ot_amount": round(op, 2), "leave_days": round(lv, 1),
            "current_rows": len(rows) - legacy_n, "legacy_rows": legacy_n,
        })
        max_emp = max(max_emp, len(rows))
        tot["man_days"] += md
        tot["wages"] += wg
        tot["basic"] += bs
        tot["ot_hours"] += oh
        tot["ot_pay"] += op
        tot["leave"] += lv
        tot["working_days"] += wd
        for uid, r in rows.items():
            pu = per_user[uid]
            pu["man_days"] += r["present_days"]
            pu["wages"] += r["gross"]
            pu["ot_hours"] += r["ot_hours"]
            pu["ot_pay"] += r["ot_pay"]
            pu["leave"] += r["leave_days"]
            pu["months"] += 1

    male = female = contract = 0
    # Iter 520 — FORM 23 gender-wise accumulators.
    md_men = md_women = 0.0
    oh_men = oh_women = 0.0
    ent_men = ent_women = 0    # entitled to annual leave (240+ days)
    gr_men = gr_women = 0      # granted leave during the year
    employees: List[Dict[str, Any]] = []
    dept = defaultdict(lambda: {"employees": 0, "man_days": 0.0, "wages": 0.0})
    cat = defaultdict(lambda: {"employees": 0, "man_days": 0.0, "wages": 0.0})
    for uid, pu in per_user.items():
        u = emp.get(uid) or {}
        g = (u.get("gender") or "").strip().lower()
        if g.startswith("m"):
            male += 1
            md_men += pu["man_days"]
            oh_men += pu["ot_hours"]
            ent_men += 1 if pu["man_days"] >= 240 else 0
            gr_men += 1 if pu["leave"] > 0 else 0
        elif g.startswith("f"):
            female += 1
            md_women += pu["man_days"]
            oh_women += pu["ot_hours"]
            ent_women += 1 if pu["man_days"] >= 240 else 0
            gr_women += 1 if pu["leave"] > 0 else 0
        else:
            md_men += pu["man_days"]  # unspecified counted with men
            oh_men += pu["ot_hours"]
        et = (u.get("employee_type") or "").strip() or "UNSPECIFIED"
        if "contract" in et.lower() or u.get("contractor_name"):
            contract += 1
        dp = (u.get("department") or "").strip() or "UNSPECIFIED"
        dept[dp]["employees"] += 1
        dept[dp]["man_days"] += pu["man_days"]
        dept[dp]["wages"] += pu["wages"]
        cat[et]["employees"] += 1
        cat[et]["man_days"] += pu["man_days"]
        cat[et]["wages"] += pu["wages"]
        employees.append({
            "employee_code": u.get("employee_code") or "",
            "name": u.get("name") or uid,
            "gender": (u.get("gender") or "").upper()[:1] or "—",
            "doj": u.get("doj") or "",
            "category": et,
            "department": dp,
            "designation": u.get("designation") or "",
            "status": "Deleted" if u.get("deleted") else
                      ("Active" if u.get("is_onroll", True) else "Left"),
            "man_days": round(pu["man_days"], 1),
            "wages": round(pu["wages"], 2),
            "ot_hours": round(pu["ot_hours"], 1),
            "months": int(pu["months"]),
        })
    employees.sort(key=lambda e: (e["department"], e["name"]))

    acc = (details.get("accidents") or {}).get(str(year)) or {
        "fatal": 0, "nonfatal": 0, "mandays_lost": 0}

    # ---- Iter 520 — FORM NO. 23 (Rule 105(i)) computed data points ----
    _fdh = float((c.get("attendance_policy") or {}).get("full_day_hours") or 8.0)
    _days_worked = int(round(tot["working_days"]))
    _avg_men = round(md_men / _days_worked, 0) if _days_worked else 0
    _avg_women = round(md_women / _days_worked, 0) if _days_worked else 0
    form23 = {
        "manual": details.get("form23") or {},
        "days_worked": _days_worked,
        "man_days": {"men": round(md_men, 1), "women": round(md_women, 1),
                     "children": 0,
                     "total": round(md_men + md_women, 1)},
        "avg_daily": {"men": int(_avg_men), "women": int(_avg_women),
                      "children": 0, "total": int(_avg_men + _avg_women)},
        "man_hours": {"men": round(md_men * _fdh + oh_men),
                      "women": round(md_women * _fdh + oh_women),
                      "children": 0,
                      "total": round((md_men + md_women) * _fdh
                                     + oh_men + oh_women)},
        "avg_week_hours": {"men": round(_fdh * 6) if md_men else 0,
                           "women": round(_fdh * 6) if md_women else 0},
        "leave": {
            "employed": {"men": male, "women": female, "children": 0,
                         "total": male + female},
            "entitled": {"men": ent_men, "women": ent_women, "children": 0,
                         "total": ent_men + ent_women},
            "granted": {"men": gr_men, "women": gr_women, "children": 0,
                        "total": gr_men + gr_women},
        },
        "wages": {
            "gross": round(tot["wages"], 2),
            "basic": round(tot["basic"], 2),
            "da_allowances": round(max(tot["wages"] - tot["basic"], 0.0), 2),
            "arrears": 0,
            "ot_amount": round(tot["ot_pay"], 2),
        },
    }
    return {
        "year": year,
        "source": source,
        "firm": {"company_id": cid, "name": c.get("name"), **details},
        "monthly": monthly,
        "summary": {
            "avg_daily_employment": round(tot["man_days"] / tot["working_days"], 1)
            if tot["working_days"] else 0,
            "max_employment": max_emp,
            "male": male, "female": female,
            "contract_labour": contract,
            "employees_total": len(per_user),
            "total_man_days": round(tot["man_days"], 1),
            "total_wages": round(tot["wages"], 2),
            "total_ot_hours": round(tot["ot_hours"], 1),
            "total_ot_amount": round(tot["ot_pay"], 2),
            "leave_with_wages": round(tot["leave"], 1),
        },
        "departments": [{"name": k, "employees": v["employees"],
                         "man_days": round(v["man_days"], 1),
                         "wages": round(v["wages"], 2)}
                        for k, v in sorted(dept.items())],
        "categories": [{"name": k, "employees": v["employees"],
                        "man_days": round(v["man_days"], 1),
                        "wages": round(v["wages"], 2)}
                       for k, v in sorted(cat.items())],
        "employees": employees,
        "accidents": acc,
        "welfare": details.get("welfare") or {},
        "form23": form23,
    }


def _src(source: Optional[str]) -> str:
    s = (source or "combined").lower()
    return s if s in ("combined", "current", "legacy") else "combined"


@router.get("/{company_id}/{year}.pdf")
async def factory_return_pdf(company_id: str, year: int,
                             source: Optional[str] = Query(None),
                             authorization: Optional[str] = Header(None)):
    await _admin(authorization)
    d = await _compute_return(company_id, year, _src(source))
    from utils.factory_return_pdf import build_factory_return_pdf
    pdf = build_factory_return_pdf(d, boiler=False)
    return Response(content=pdf, media_type="application/pdf", headers={
        "Content-Disposition":
            f'attachment; filename="factory-annual-return-{year}.pdf"'})


@router.get("/{company_id}/{year}/boiler.pdf")
async def boiler_return_pdf(company_id: str, year: int,
                            source: Optional[str] = Query(None),
                            authorization: Optional[str] = Header(None)):
    await _admin(authorization)
    d = await _compute_return(company_id, year, _src(source))
    from utils.factory_return_pdf import build_factory_return_pdf
    pdf = build_factory_return_pdf(d, boiler=True)
    return Response(content=pdf, media_type="application/pdf", headers={
        "Content-Disposition":
            f'attachment; filename="boiler-annual-return-{year}.pdf"'})


@router.get("/{company_id}/{year}/form23.pdf")
async def form23_pdf(company_id: str, year: int,
                     source: Optional[str] = Query(None),
                     authorization: Optional[str] = Header(None)):
    """Iter 520 (user upload) — official FORM NO. 23 Annual Return
    (Prescribed under Rule 105(i), Factories Act / Payment of Wages Act)."""
    await _admin(authorization)
    d = await _compute_return(company_id, year, _src(source))
    from utils.factory_return_pdf import build_form23_pdf
    pdf = build_form23_pdf(d)
    return Response(content=pdf, media_type="application/pdf", headers={
        "Content-Disposition":
            f'attachment; filename="form23-annual-return-{year}.pdf"'})


@router.get("/{company_id}/{year}.xlsx")
async def factory_return_xlsx(company_id: str, year: int,
                              source: Optional[str] = Query(None),
                              authorization: Optional[str] = Header(None)):
    await _admin(authorization)
    d = await _compute_return(company_id, year, _src(source))
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Firm & Summary"
    f = d["firm"]
    ws.append([f"FACTORY & BOILER ANNUAL RETURN — {year} ({d['source'].upper()} data)"])
    for lbl, key in [
        ("Factory Name", "factory_name"), ("Address", "factory_address"),
        ("Factory License No", "factory_license_no"),
        ("Registration No", "factory_registration_no"),
        ("Boiler Registration No", "boiler_registration_no"),
        ("Occupier", "occupier_name"), ("Factory Manager", "factory_manager"),
        ("Nature of Manufacturing", "nature_of_manufacturing"),
        ("District", "district"), ("State", "state"),
    ]:
        ws.append([lbl, f.get(key) or ""])
    ws.append([])
    s = d["summary"]
    for lbl, key in [
        ("Average Daily Employment", "avg_daily_employment"),
        ("Maximum Employment", "max_employment"), ("Male", "male"),
        ("Female", "female"), ("Contract Labour", "contract_labour"),
        ("Total Employees (year)", "employees_total"),
        ("Total Man Days", "total_man_days"), ("Total Wages Paid", "total_wages"),
        ("Total OT Hours", "total_ot_hours"), ("Total OT Amount", "total_ot_amount"),
        ("Leave with Wages (days)", "leave_with_wages"),
    ]:
        ws.append([lbl, s.get(key)])
    ws.append([])
    ws.append(["Accidents — Fatal", d["accidents"].get("fatal", 0)])
    ws.append(["Accidents — Non-fatal", d["accidents"].get("nonfatal", 0)])
    ws.append(["Man-days lost (accidents)", d["accidents"].get("mandays_lost", 0)])

    w2 = wb.create_sheet("Monthly Statistics")
    w2.append(["Month", "Employees", "Man Days", "Working Days",
               "Avg Daily Employment", "Wages", "OT Hours", "OT Amount",
               "Leave Days", "Current Rows", "Legacy Rows"])
    for m in d["monthly"]:
        w2.append([m["month"], m["employees"], m["man_days"], m["working_days"],
                   m["avg_daily_employment"], m["wages"], m["ot_hours"],
                   m["ot_amount"], m["leave_days"], m["current_rows"],
                   m["legacy_rows"]])

    w3 = wb.create_sheet("Employee Statistics")
    w3.append(["Code", "Name", "Gender", "DOJ", "Category", "Department",
               "Designation", "Status", "Months", "Man Days", "Wages", "OT Hours"])
    for e in d["employees"]:
        w3.append([e["employee_code"], e["name"], e["gender"], e["doj"],
                   e["category"], e["department"], e["designation"], e["status"],
                   e["months"], e["man_days"], e["wages"], e["ot_hours"]])

    for title, rows in [("Department-wise", d["departments"]),
                        ("Category-wise", d["categories"])]:
        w = wb.create_sheet(title)
        w.append(["Name", "Employees", "Man Days", "Wages"])
        for r in rows:
            w.append([r["name"], r["employees"], r["man_days"], r["wages"]])

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="factory-annual-return-{year}.xlsx"'})


@router.get("/{company_id}/{year}")
async def factory_return_json(company_id: str, year: int,
                              source: Optional[str] = Query(None),
                              authorization: Optional[str] = Header(None)):
    await _admin(authorization)
    return await _compute_return(company_id, year, _src(source))
