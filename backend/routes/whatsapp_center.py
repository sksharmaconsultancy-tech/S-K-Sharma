"""Iter 395 — WhatsApp Center APIs (config, templates, send, history,
dashboard, schedules, reports, webhook + chatbot).

All admin endpoints follow the existing auth pattern (bearer token +
require_role). The webhook endpoints are PUBLIC (Meta calls them).
"""
import hashlib
import hmac
import io
import json
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query, Request, Response

from utils import whatsapp_engine as wa
from utils.secrets_vault import decrypt_secret

router = APIRouter(prefix="/api")
api = router


# --------------------------------------------------------------- helpers
async def _admin(authorization: Optional[str]):
    from server import get_user_from_token, require_role
    user = await get_user_from_token(authorization)
    require_role(user, ["super_admin", "company_admin", "sub_admin"])
    return user


def _scope_company(admin: dict, company_id: Optional[str]) -> str:
    if admin["role"] != "super_admin":
        cid = admin.get("company_id")
        if not cid:
            raise HTTPException(status_code=400, detail="No company scope")
        return cid
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id required")
    return company_id


async def _audit(admin: dict, action: str, detail: str, company_id: str):
    from server import db
    await db.wa_audit.insert_one({
        "audit_id": f"waa_{uuid.uuid4().hex[:10]}",
        "company_id": company_id, "action": action, "detail": detail[:500],
        "by": admin.get("user_id"), "by_name": admin.get("name"),
        "at": wa.now_iso()})


# --------------------------------------------------------------- settings
@api.get("/admin/whatsapp/settings")
async def get_wa_settings(company_id: Optional[str] = Query(None),
                          authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    cid = _scope_company(admin, company_id)
    s = await wa.get_settings(cid)
    return {"settings": wa.mask_settings(s),
            "configured": wa._configured(s),
            "automation_events": wa.AUTOMATION_EVENTS,
            "webhook_url": "/api/whatsapp/webhook"}


@api.put("/admin/whatsapp/settings")
async def put_wa_settings(payload: dict = Body(...),
                          company_id: Optional[str] = Query(None),
                          authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    cid = _scope_company(admin, company_id or payload.get("company_id"))
    s = await wa.save_settings(cid, payload, admin["user_id"])
    await _audit(admin, "settings_update",
                 f"keys: {', '.join(k for k in payload.keys())}", cid)
    return {"ok": True, "settings": wa.mask_settings(s),
            "configured": wa._configured(s)}


@api.post("/admin/whatsapp/test-connection")
async def test_wa_connection(company_id: Optional[str] = Query(None),
                             authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    cid = _scope_company(admin, company_id)
    s = await wa.get_settings(cid)
    if not wa._configured(s):
        return {"ok": False, "error": "Not configured — enter Phone Number ID "
                                      "and Access Token, then enable WhatsApp."}
    import httpx
    client = wa.WhatsAppClient(s)
    try:
        async with httpx.AsyncClient(timeout=20) as c:
            r = await c.get(
                f"{wa.GRAPH_BASE}/{s.get('api_version') or 'v22.0'}/{s['phone_number_id']}",
                params={"fields": "display_phone_number,verified_name,quality_rating"},
                headers={"Authorization": f"Bearer {client.token}"})
            j = r.json()
            if r.status_code >= 400:
                return {"ok": False,
                        "error": (j.get("error") or {}).get("message") or r.text[:200]}
            return {"ok": True, "phone": j.get("display_phone_number"),
                    "verified_name": j.get("verified_name"),
                    "quality_rating": j.get("quality_rating")}
    except Exception as e:
        return {"ok": False, "error": str(e)[:200]}


# -------------------------------------------------------------- templates
@api.get("/admin/whatsapp/templates")
async def list_wa_templates(company_id: Optional[str] = Query(None),
                            authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    cid = _scope_company(admin, company_id)
    rows = await db.wa_templates.find(
        {"company_id": {"$in": [cid, "__global__"]}},
        {"_id": 0}).sort("category", 1).to_list(500)
    return {"templates": rows, "variables": wa.TEMPLATE_VARIABLES,
            "categories": sorted({t["category"] for t in wa.DEFAULT_TEMPLATES})}


@api.post("/admin/whatsapp/templates/seed-defaults")
async def seed_wa_templates(company_id: Optional[str] = Query(None),
                            authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    cid = _scope_company(admin, company_id)
    n = await wa.seed_default_templates("__global__")
    await _audit(admin, "templates_seed", f"{n} defaults created", cid)
    return {"ok": True, "created": n}


@api.post("/admin/whatsapp/templates")
async def create_wa_template(payload: dict = Body(...),
                             company_id: Optional[str] = Query(None),
                             authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    cid = _scope_company(admin, company_id or payload.get("company_id"))
    doc = {
        "template_id": f"wat_{uuid.uuid4().hex[:10]}",
        "company_id": payload.get("scope") == "global" and admin["role"] == "super_admin"
                      and "__global__" or cid,
        "category": (payload.get("category") or "custom").strip(),
        "name": (payload.get("name") or "Untitled").strip(),
        "body": payload.get("body") or "",
        "language": payload.get("language") or "en",
        "meta_template_name": (payload.get("meta_template_name") or "").strip(),
        "active": bool(payload.get("active", True)),
        "is_default": False,
        "created_at": wa.now_iso(), "created_by": admin["user_id"],
    }
    await db.wa_templates.insert_one(doc)
    doc.pop("_id", None)
    await _audit(admin, "template_create", doc["name"], cid)
    return {"ok": True, "template": doc}


@api.put("/admin/whatsapp/templates/{template_id}")
async def update_wa_template(template_id: str, payload: dict = Body(...),
                             authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    upd = {k: v for k, v in payload.items()
           if k in ("name", "body", "category", "language",
                    "meta_template_name", "active")}
    upd["updated_at"] = wa.now_iso()
    r = await db.wa_templates.update_one(
        {"template_id": template_id}, {"$set": upd})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"ok": True}


@api.delete("/admin/whatsapp/templates/{template_id}")
async def delete_wa_template(template_id: str,
                             authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    r = await db.wa_templates.delete_one({"template_id": template_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"ok": True}


@api.post("/admin/whatsapp/preview")
async def preview_wa_message(payload: dict = Body(...),
                             company_id: Optional[str] = Query(None),
                             authorization: Optional[str] = Header(None)):
    """Render a template/body against a real employee for preview."""
    admin = await _admin(authorization)
    from server import db
    cid = _scope_company(admin, company_id or payload.get("company_id"))
    body = payload.get("body")
    if not body and payload.get("template_id"):
        t = await db.wa_templates.find_one(
            {"template_id": payload["template_id"]}, {"_id": 0})
        body = (t or {}).get("body") or ""
    u = {}
    if payload.get("user_id"):
        u = await db.users.find_one({"user_id": payload["user_id"]}, {"_id": 0}) or {}
    comp = await db.companies.find_one({"company_id": cid},
                                       {"_id": 0, "name": 1}) or {}
    ctx = wa.build_context(u, comp, payload.get("extra") or {})
    return {"rendered": wa.render_template(body or "", ctx),
            "to": wa.wa_number_for(u) if u else None}


# ------------------------------------------------------------------ send
@api.post("/admin/whatsapp/send")
async def send_wa_manual(payload: dict = Body(...),
                         company_id: Optional[str] = Query(None),
                         authorization: Optional[str] = Header(None)):
    """Manual / bulk send. Body:
      { user_ids: [...] | target: {mode, department, user_ids},
        template_id | category | body, extra, attachment{filename,mime,b64},
        scheduled_at (ISO, optional) }"""
    admin = await _admin(authorization)
    from server import db
    cid = _scope_company(admin, company_id or payload.get("company_id"))
    s = await wa.get_settings(cid)

    target = payload.get("target") or {}
    if payload.get("user_ids"):
        target = {"mode": "employees", "user_ids": payload["user_ids"]}
    users = await wa.resolve_targets(cid, target)
    if not users:
        raise HTTPException(status_code=400, detail="No recipients matched")
    if len(users) > 5000:
        raise HTTPException(status_code=400,
                            detail=f"Too many recipients ({len(users)}); max 5000 per batch")

    body = payload.get("body")
    category = payload.get("category") or "custom"
    if payload.get("template_id"):
        t = await db.wa_templates.find_one(
            {"template_id": payload["template_id"]}, {"_id": 0})
        if t:
            body = t.get("body")
            category = t.get("category") or category
    att = payload.get("attachment") or None
    if att and att.get("b64"):
        limit = float(s.get("attachment_limit_mb") or 16) * 1024 * 1024
        if len(att["b64"]) * 0.75 > limit:
            raise HTTPException(status_code=400,
                                detail=f"Attachment exceeds {s.get('attachment_limit_mb')} MB")
    queued = 0
    source = "bulk" if len(users) > 1 else "manual"
    for u in users:
        m = await wa.enqueue_message(
            company_id=cid, user=u, category=category, body=body,
            extra=payload.get("extra") or {}, source=source,
            attachment=att, created_by=admin["user_id"],
            scheduled_at=payload.get("scheduled_at"), dedupe=False)
        if m:
            queued += 1
    await _audit(admin, "send", f"{source}: {queued} message(s), category={category}", cid)
    return {"ok": True, "queued": queued, "recipients": len(users),
            "configured": wa._configured(s)}


@api.post("/admin/whatsapp/send-salary-slips")
async def send_wa_salary_slips(payload: dict = Body(...),
                               company_id: Optional[str] = Query(None),
                               authorization: Optional[str] = Header(None)):
    """Queue payslip PDFs for a month. Body: {month, user_ids(optional)}."""
    admin = await _admin(authorization)
    import server as srv
    cid = _scope_company(admin, company_id or payload.get("company_id"))
    month = payload.get("month")
    if not month:
        raise HTTPException(status_code=400, detail="month (YYYY-MM) required")
    rows, _md, _src = await srv._payslip_rows_for_month(cid, month)
    if not rows:
        raise HTTPException(status_code=404,
                            detail="No processed salary found for this month")
    want = set(payload.get("user_ids") or [])
    queued, skipped = 0, 0
    for row in rows:
        uid = row.get("user_id")
        if want and uid not in want:
            continue
        u = await srv.db.users.find_one({"user_id": uid}, {"_id": 0})
        if not u:
            skipped += 1
            continue
        net = row.get("net_pay") or row.get("net") or row.get("net_salary") or ""
        m = await wa.enqueue_message(
            company_id=cid, user=u, category="salary_slip",
            extra={"Month": month, "Salary": str(net)},
            source="bulk", created_by=admin["user_id"],
            attachment={"payslip": {"company_id": cid, "user_id": uid,
                                    "month": month}},
            dedupe=False)
        if m:
            queued += 1
        else:
            skipped += 1
    await _audit(admin, "send_salary_slips", f"{month}: {queued} queued", cid)
    return {"ok": True, "queued": queued, "skipped": skipped, "total_rows": len(rows)}


# --------------------------------------------------------- history & queue
@api.get("/admin/whatsapp/messages")
async def list_wa_messages(company_id: Optional[str] = Query(None),
                           status: Optional[str] = Query(None),
                           category: Optional[str] = Query(None),
                           source: Optional[str] = Query(None),
                           q: Optional[str] = Query(None),
                           date_from: Optional[str] = Query(None),
                           date_to: Optional[str] = Query(None),
                           limit: int = Query(100, le=500),
                           skip: int = Query(0),
                           authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    cid = _scope_company(admin, company_id)
    fl: Dict[str, Any] = {"company_id": cid}
    if status:
        fl["status"] = status
    if category:
        fl["category"] = category
    if source:
        fl["source"] = source
    if q:
        fl["$or"] = [{"employee_name": {"$regex": q, "$options": "i"}},
                     {"employee_code": {"$regex": q, "$options": "i"}},
                     {"to": {"$regex": q}}]
    if date_from:
        fl["created_at"] = {"$gte": date_from}
    if date_to:
        fl.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    total = await db.wa_messages.count_documents(fl)
    rows = await db.wa_messages.find(fl, {"_id": 0, "attachment.b64": 0}) \
        .sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"messages": rows, "total": total}


@api.post("/admin/whatsapp/messages/{msg_id}/retry")
async def retry_wa_message(msg_id: str,
                           authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    r = await db.wa_messages.update_one(
        {"msg_id": msg_id, "status": {"$in": ["failed", "cancelled"]}},
        {"$set": {"status": "queued", "error": None,
                  "scheduled_at": wa.now_iso(), "retry_count": 0}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404,
                            detail="Message not found or not in failed state")
    return {"ok": True}


@api.post("/admin/whatsapp/messages/{msg_id}/cancel")
async def cancel_wa_message(msg_id: str,
                            authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    r = await db.wa_messages.update_one(
        {"msg_id": msg_id, "status": "queued"},
        {"$set": {"status": "cancelled"}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404,
                            detail="Message not found or already sent")
    return {"ok": True}


@api.delete("/admin/whatsapp/messages/{msg_id}")
async def delete_wa_message(msg_id: str,
                            authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db, require_role
    require_role(admin, ["super_admin", "company_admin"])
    r = await db.wa_messages.delete_one({"msg_id": msg_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Message not found")
    return {"ok": True}


# --------------------------------------------------------------- dashboard
@api.get("/admin/whatsapp/dashboard")
async def wa_dashboard(company_id: Optional[str] = Query(None),
                       authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    cid = _scope_company(admin, company_id)
    today0 = datetime.now(wa.IST).replace(hour=0, minute=0, second=0) \
        .astimezone(timezone.utc).isoformat()

    async def _count(fl):
        return await db.wa_messages.count_documents({"company_id": cid, **fl})

    sent_today = await _count({"sent_at": {"$gte": today0}})
    delivered = await _count({"status": {"$in": ["delivered", "read"]}})
    read = await _count({"status": "read"})
    failed = await _count({"status": "failed"})
    pending = await _count({"status": {"$in": ["queued", "sending"]}})
    total = await _count({})
    sent_total = await _count({"status": {"$in": ["sent", "delivered", "read"]}})
    top = await db.wa_messages.aggregate([
        {"$match": {"company_id": cid}},
        {"$group": {"_id": "$category", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}, {"$limit": 8}]).to_list(8)
    # last-14-day trend
    trend = await db.wa_messages.aggregate([
        {"$match": {"company_id": cid,
                    "created_at": {"$gte": (datetime.now(timezone.utc)
                                            - timedelta(days=14)).isoformat()}}},
        {"$group": {"_id": {"$substr": ["$created_at", 0, 10]},
                    "count": {"$sum": 1},
                    "failed": {"$sum": {"$cond": [
                        {"$eq": ["$status", "failed"]}, 1, 0]}}}},
        {"$sort": {"_id": 1}}]).to_list(20)
    s = await wa.get_settings(cid)
    return {"kpis": {
        "sent_today": sent_today, "delivered": delivered, "read": read,
        "failed": failed, "pending": pending, "total": total,
        "success_pct": round(sent_total / total * 100, 1) if total else 0,
        "failure_pct": round(failed / total * 100, 1) if total else 0,
        "daily_limit": s.get("daily_limit"),
        "configured": wa._configured(s), "enabled": bool(s.get("enabled")),
    }, "top_templates": [{"category": t["_id"], "count": t["count"]} for t in top],
        "trend": [{"date": t["_id"], "count": t["count"], "failed": t["failed"]}
                  for t in trend]}


# --------------------------------------------------------------- schedules
@api.get("/admin/whatsapp/schedules")
async def list_wa_schedules(company_id: Optional[str] = Query(None),
                            authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    cid = _scope_company(admin, company_id)
    rows = await db.wa_schedules.find(
        {"company_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"schedules": rows}


@api.post("/admin/whatsapp/schedules")
async def create_wa_schedule(payload: dict = Body(...),
                             company_id: Optional[str] = Query(None),
                             authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    cid = _scope_company(admin, company_id or payload.get("company_id"))
    typ = payload.get("type") or "once"
    time_s = payload.get("time") or "09:00"
    if typ == "once":
        date_s = payload.get("date") or wa.ist_today()
        run_at = datetime.strptime(f"{date_s} {time_s}", "%Y-%m-%d %H:%M") \
            .replace(tzinfo=wa.IST).astimezone(timezone.utc).isoformat()
    else:
        hh, mm = time_s.split(":")
        base = datetime.now(wa.IST).replace(hour=int(hh), minute=int(mm),
                                            second=0, microsecond=0)
        if base <= datetime.now(wa.IST):
            base += timedelta(days=1)
        run_at = base.astimezone(timezone.utc).isoformat()
    doc = {
        "schedule_id": f"was_{uuid.uuid4().hex[:10]}",
        "company_id": cid,
        "title": payload.get("title") or "Scheduled message",
        "type": typ, "time": time_s,
        "date": payload.get("date"),
        "day_of_month": payload.get("day_of_month"),
        "category": payload.get("category") or "custom",
        "custom_body": payload.get("custom_body") or None,
        "target": payload.get("target") or {"mode": "company"},
        "active": True,
        "next_run_at": run_at, "last_run_at": None,
        "created_at": wa.now_iso(), "created_by": admin["user_id"],
    }
    await db.wa_schedules.insert_one(doc)
    doc.pop("_id", None)
    await _audit(admin, "schedule_create", doc["title"], cid)
    return {"ok": True, "schedule": doc}


@api.delete("/admin/whatsapp/schedules/{schedule_id}")
async def delete_wa_schedule(schedule_id: str,
                             authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    r = await db.wa_schedules.delete_one({"schedule_id": schedule_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Schedule not found")
    return {"ok": True}


@api.put("/admin/whatsapp/schedules/{schedule_id}")
async def toggle_wa_schedule(schedule_id: str, payload: dict = Body(...),
                             authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    upd = {k: v for k, v in payload.items() if k in ("active", "title", "time")}
    r = await db.wa_schedules.update_one(
        {"schedule_id": schedule_id}, {"$set": upd})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Schedule not found")
    return {"ok": True}


# ------------------------------------------------------ employee wa fields
@api.put("/admin/whatsapp/employee-fields/{user_id}")
async def put_wa_employee_fields(user_id: str, payload: dict = Body(...),
                                 authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    upd = {k: v for k, v in payload.items()
           if k in ("whatsapp_number", "wa_country_code", "wa_verified",
                    "wa_language", "wa_consent")}
    if not upd:
        raise HTTPException(status_code=400, detail="No WhatsApp fields given")
    r = await db.users.update_one({"user_id": user_id}, {"$set": upd})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"ok": True}


# ----------------------------------------------------------------- reports
@api.get("/admin/whatsapp/report")
async def wa_report(company_id: Optional[str] = Query(None),
                    date_from: Optional[str] = Query(None),
                    date_to: Optional[str] = Query(None),
                    group_by: str = Query("date"),  # date|category|department
                    fmt: str = Query("json"),       # json|xlsx|pdf
                    authorization: Optional[str] = Header(None)):
    admin = await _admin(authorization)
    from server import db
    cid = _scope_company(admin, company_id)
    fl: Dict[str, Any] = {"company_id": cid}
    if date_from:
        fl["created_at"] = {"$gte": date_from}
    if date_to:
        fl.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    key = {"date": {"$substr": ["$created_at", 0, 10]},
           "category": "$category",
           "department": {"$ifNull": ["$department", "$category"]}}.get(
        group_by, {"$substr": ["$created_at", 0, 10]})
    rows = await db.wa_messages.aggregate([
        {"$match": fl},
        {"$group": {
            "_id": key, "total": {"$sum": 1},
            "sent": {"$sum": {"$cond": [{"$in": ["$status", ["sent", "delivered", "read"]]}, 1, 0]}},
            "delivered": {"$sum": {"$cond": [{"$in": ["$status", ["delivered", "read"]]}, 1, 0]}},
            "read": {"$sum": {"$cond": [{"$eq": ["$status", "read"]}, 1, 0]}},
            "failed": {"$sum": {"$cond": [{"$eq": ["$status", "failed"]}, 1, 0]}},
            "pending": {"$sum": {"$cond": [{"$in": ["$status", ["queued", "sending"]]}, 1, 0]}},
        }},
        {"$sort": {"_id": 1}}]).to_list(1000)
    data = [{"group": r["_id"], "total": r["total"], "sent": r["sent"],
             "delivered": r["delivered"], "read": r["read"],
             "failed": r["failed"], "pending": r["pending"]} for r in rows]
    if fmt == "json":
        return {"rows": data, "group_by": group_by}

    comp = await db.companies.find_one({"company_id": cid}, {"_id": 0, "name": 1}) or {}
    title = f"WhatsApp Delivery Report — {comp.get('name') or cid}"
    period = f"{date_from or '…'} to {date_to or wa.ist_today()}"
    headers = [group_by.title(), "Total", "Sent", "Delivered", "Read",
               "Failed", "Pending"]
    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "WA Report"
        ws.append([title]); ws.append([f"Period: {period}"]); ws.append([])
        ws.append(headers)
        for c in ws[4]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="128C7E")
        for r in data:
            ws.append([r["group"], r["total"], r["sent"], r["delivered"],
                       r["read"], r["failed"], r["pending"]])
        buf = io.BytesIO(); wb.save(buf)
        return Response(buf.getvalue(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition":
                                 'attachment; filename="whatsapp_report.xlsx"'})
    if fmt == "pdf":
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm,
                                rightMargin=15 * mm, topMargin=15 * mm)
        st = getSampleStyleSheet()
        els = [Paragraph(title, st["Title"]),
               Paragraph(f"Period: {period}", st["Normal"]), Spacer(1, 6)]
        tdata = [headers] + [[str(r["group"]), r["total"], r["sent"],
                              r["delivered"], r["read"], r["failed"],
                              r["pending"]] for r in data]
        t = Table(tdata, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#128C7E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [rl_colors.white, rl_colors.HexColor("#F2FBF9")])]))
        els.append(t)
        doc.build(els)
        return Response(buf.getvalue(), media_type="application/pdf",
                        headers={"Content-Disposition":
                                 'attachment; filename="whatsapp_report.pdf"'})
    raise HTTPException(status_code=400, detail="fmt must be json|xlsx|pdf")


# ------------------------------------------------------------------ webhook
@api.get("/whatsapp/webhook")
async def wa_webhook_verify(request: Request):
    """Meta webhook verification — PUBLIC endpoint."""
    from server import db
    qp = request.query_params
    if qp.get("hub.mode") == "subscribe":
        token = qp.get("hub.verify_token") or ""
        match = await db.wa_settings.find_one(
            {"webhook_verify_token": token, "enabled": True}, {"_id": 1}) \
            if token else None
        if match:
            return Response(content=qp.get("hub.challenge", ""),
                            media_type="text/plain")
    raise HTTPException(status_code=403, detail="Verification failed")


@api.post("/whatsapp/webhook")
async def wa_webhook(request: Request,
                     x_hub_signature_256: str = Header(default="")):
    """Meta status callbacks + inbound chatbot messages — PUBLIC endpoint."""
    from server import db, logger
    raw = await request.body()
    try:
        payload = json.loads(raw)
    except Exception:
        return {"ok": True}
    for entry in payload.get("entry", []):
        for change in entry.get("changes", []):
            value = change.get("value") or {}
            pnid = ((value.get("metadata") or {}).get("phone_number_id")) or ""
            s_doc = await db.wa_settings.find_one(
                {"phone_number_id": pnid}, {"_id": 0}) if pnid else None
            if not s_doc:
                continue
            cid = s_doc.get("company_id")
            # signature validation when app secret configured
            secret = decrypt_secret(s_doc.get("webhook_secret") or "")
            if secret:
                expected = hmac.new(secret.encode(), raw,
                                    hashlib.sha256).hexdigest()
                received = (x_hub_signature_256 or "").removeprefix("sha256=")
                if not hmac.compare_digest(expected, received):
                    logger.warning("[wa] webhook bad signature for %s", cid)
                    continue
            # delivery statuses
            for st in value.get("statuses") or []:
                wamid = st.get("id")
                status = st.get("status")
                if not wamid or status not in ("sent", "delivered", "read", "failed"):
                    continue
                upd: Dict[str, Any] = {}
                if status == "delivered":
                    upd = {"status": "delivered", "delivered_at": wa.now_iso()}
                elif status == "read":
                    upd = {"status": "read", "read_at": wa.now_iso()}
                elif status == "failed":
                    err = ((st.get("errors") or [{}])[0])
                    upd = {"status": "failed",
                           "error": f"{err.get('code')}: {err.get('title') or ''}"[:300]}
                if upd:
                    q = {"wa_message_id": wamid}
                    if status == "delivered":   # never downgrade read → delivered
                        q["status"] = {"$ne": "read"}
                    await db.wa_messages.update_one(q, {"$set": upd})
            # inbound messages → chatbot
            for msg in value.get("messages") or []:
                if msg.get("type") != "text":
                    continue
                frm = msg.get("from") or ""
                text = ((msg.get("text") or {}).get("body")) or ""
                await db.wa_inbound.insert_one({
                    "inbound_id": f"wai_{uuid.uuid4().hex[:10]}",
                    "company_id": cid, "from": frm, "text": text[:1000],
                    "wa_message_id": msg.get("id"), "at": wa.now_iso()})
                reply = await wa.chatbot_reply(cid, frm, text)
                if reply:
                    await wa.enqueue_message(
                        company_id=cid, user=None, to=frm,
                        category="chatbot", body=reply,
                        source="chatbot", created_by="chatbot", dedupe=False)
    return {"ok": True}
