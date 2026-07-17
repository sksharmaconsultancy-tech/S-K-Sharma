"""Iter 172 — Punch data import from Excel (Punch Approvals screen).

Admins upload a company-wise Excel sheet of punches; rows are matched to
employees by BIO CODE first, then by NAME. Matched In/Out punches are
inserted into db.attendance (source="excel_import", auto-approved) so they
appear in every punching report immediately.

Endpoints:
  * POST /api/admin/punch-import/preview — parse + match, nothing saved.
  * POST /api/admin/punch-import/commit  — insert approved punches.
"""
import base64
import io
import uuid
from datetime import datetime, time as dtime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException

from server import db, get_user_from_token, require_role, now_iso  # noqa: E402

router = APIRouter(prefix="/api/admin/punch-import", tags=["punch-import"])


def _norm_bio(v: Any) -> str:
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0") or ("0" if s else "")


def _parse_date_cell(v: Any) -> Optional[str]:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()[:10].replace("/", "-").replace(".", "-")
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).strftime("%Y-%m-%d")
    except Exception:
        pass
    try:
        d, m, y = s.split("-")
        if len(y) == 4:
            return datetime(int(y), int(m), int(d)).strftime("%Y-%m-%d")
    except Exception:
        pass
    return None


def _parse_time_cell(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    if isinstance(v, dtime):
        return v.strftime("%H:%M")
    if isinstance(v, (int, float)):  # Excel time fraction
        total = int(round(float(v) % 1 * 24 * 60))
        return f"{total // 60:02d}:{total % 60:02d}"
    s = str(v).strip().upper().replace(".", ":")
    if not s or s in ("-", "—"):
        return None
    ampm = None
    if s.endswith("AM") or s.endswith("PM"):
        ampm = s[-2:]
        s = s[:-2].strip()
    parts = s.split(":")
    try:
        h = int(parts[0]); mi = int(parts[1]) if len(parts) > 1 else 0
        if ampm == "PM" and h < 12:
            h += 12
        if ampm == "AM" and h == 12:
            h = 0
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return f"{h:02d}:{mi:02d}"
    except Exception:
        pass
    return None


def _parse_sheet(data: bytes) -> List[Dict[str, Any]]:
    """Detect header row + columns (bio / name / date / in / out) and
    return raw parsed rows."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header_idx, cols = None, {}
    for i, r in enumerate(rows[:15]):
        labels = [str(c or "").strip().lower() for c in r]
        cmap: Dict[str, int] = {}
        for j, lbl in enumerate(labels):
            if not lbl:
                continue
            if "bio" in lbl or lbl in ("code", "emp code", "employee code", "emp. code"):
                cmap.setdefault("bio", j)
            elif "name" in lbl:
                cmap.setdefault("name", j)
            elif "date" in lbl:
                cmap.setdefault("date", j)
            elif "out" in lbl:
                cmap.setdefault("out", j)
            elif lbl == "in" or "in time" in lbl or "punch in" in lbl or lbl.startswith("in "):
                cmap.setdefault("in", j)
        if "date" in cmap and ("bio" in cmap or "name" in cmap) and ("in" in cmap or "out" in cmap):
            header_idx, cols = i, cmap
            break
    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail=("Could not find a header row. The sheet needs columns: "
                    "Bio Code (or Name), Date, In Time, Out Time."))
    out: List[Dict[str, Any]] = []
    for rn, r in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if r is None or all(c in (None, "") for c in r):
            continue
        get = lambda k: (r[cols[k]] if k in cols and cols[k] < len(r) else None)  # noqa: E731
        out.append({
            "row_no": rn,
            "bio_code": _norm_bio(get("bio")),
            "name": str(get("name") or "").strip(),
            "date": _parse_date_cell(get("date")),
            "in_time": _parse_time_cell(get("in")),
            "out_time": _parse_time_cell(get("out")),
        })
    return out


async def _match_rows(company_id: str, raw: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    emps = await db.users.find(
        {"company_id": company_id, "role": "employee"},
        {"_id": 0, "user_id": 1, "name": 1, "bio_code": 1, "employee_code": 1},
    ).to_list(3000)
    by_bio = {_norm_bio(e.get("bio_code")): e for e in emps if _norm_bio(e.get("bio_code"))}
    by_code = {_norm_bio(e.get("employee_code")): e for e in emps if _norm_bio(e.get("employee_code"))}
    by_name = {str(e.get("name") or "").strip().upper(): e for e in emps}
    for row in raw:
        emp, how = None, None
        if row["bio_code"]:
            emp = by_bio.get(row["bio_code"])
            how = "bio_code" if emp else None
            if not emp:
                emp = by_code.get(row["bio_code"])
                how = "employee_code" if emp else None
        if not emp and row["name"]:
            emp = by_name.get(row["name"].upper())
            how = "name" if emp else None
        if not row["date"]:
            row.update({"status": "error", "error": "Invalid/missing date"})
        elif not row["in_time"] and not row["out_time"]:
            row.update({"status": "error", "error": "No In or Out time"})
        elif not emp:
            row.update({"status": "unmatched",
                        "error": "No employee with this Bio Code / Name"})
        else:
            row.update({"status": "matched", "matched_by": how,
                        "user_id": emp["user_id"], "emp_name": emp.get("name")})
    return raw


async def _auth(authorization: Optional[str], company_id: str = ""):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    if (company_id and admin.get("role") != "super_admin"
            and admin.get("company_id") != company_id):
        raise HTTPException(status_code=403, detail="Not authorised for this firm")
    return admin


@router.get("/template")
async def download_template(authorization: Optional[str] = Header(None)):
    """Sample Excel layout for the bulk punch import (base64 xlsx)."""
    await _auth(authorization)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Punches"
    headers = ["Bio Code", "Name", "Date", "In Time", "Out Time"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1D4ED8")
    ws.append(["101", "RAKESH KUMAR", "01-06-2026", "09:00", "18:00"])
    ws.append(["102", "", "01-06-2026", "09:15", "18:30"])
    ws.append(["", "SUNITA DEVI", "02-06-2026", "08:55", ""])
    for i, w in enumerate([12, 24, 14, 10, 10], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return {
        "filename": "punch_import_template.xlsx",
        "file_base64": base64.b64encode(buf.getvalue()).decode(),
    }


@router.post("/preview")
async def preview_import(payload: Dict[str, Any] = Body(...),
                         authorization: Optional[str] = Header(None)):
    company_id = str(payload.get("company_id") or "").strip()
    if not company_id:
        raise HTTPException(status_code=400, detail="Select a firm first")
    await _auth(authorization, company_id)
    try:
        data = base64.b64decode(str(payload.get("file_base64") or ""))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid file upload")
    if not data:
        raise HTTPException(status_code=400, detail="Empty file")
    rows = await _match_rows(company_id, _parse_sheet(data))
    matched = [r for r in rows if r["status"] == "matched"]
    return {
        "rows": rows,
        "summary": {
            "total": len(rows),
            "matched": len(matched),
            "unmatched": len([r for r in rows if r["status"] == "unmatched"]),
            "errors": len([r for r in rows if r["status"] == "error"]),
            "punches_to_create": sum(
                (1 if r["in_time"] else 0) + (1 if r["out_time"] else 0)
                for r in matched),
        },
    }


@router.post("/commit")
async def commit_import(payload: Dict[str, Any] = Body(...),
                        authorization: Optional[str] = Header(None)):
    company_id = str(payload.get("company_id") or "").strip()
    admin = await _auth(authorization, company_id)
    rows = payload.get("rows") or []
    if not company_id or not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=400, detail="Nothing to import")
    batch_id = f"pimp_{uuid.uuid4().hex[:10]}"
    created, skipped = 0, 0
    for r in rows:
        uid = str(r.get("user_id") or "")
        date = str(r.get("date") or "")
        if not uid or not date:
            continue
        for kind, tval in (("in", r.get("in_time")), ("out", r.get("out_time"))):
            if not tval:
                continue
            at = f"{date}T{tval}:00Z"
            dup = await db.attendance.find_one(
                {"user_id": uid, "date": date, "kind": kind, "at": at}, {"_id": 1})
            if dup:
                skipped += 1
                continue
            await db.attendance.insert_one({
                "record_id": f"att_{uuid.uuid4().hex[:12]}",
                "user_id": uid,
                "company_id": company_id,
                "date": date,
                "kind": kind,
                "at": at,
                "source": "excel_import",
                "status": "approved",
                "approved_by": admin["user_id"],
                "manual_reason": "Excel punch import",
                "import_batch_id": batch_id,
                "created_by": admin["user_id"],
                "created_at": now_iso(),
            })
            created += 1
    return {"ok": True, "batch_id": batch_id, "created": created, "skipped_duplicates": skipped}
