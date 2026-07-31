"""Iter 397 — EMPLOYEES ADMIN module (extracted from server.py).

Refactor only: employee create, bulk import (+ template/parse), delete
(and their helpers _employee_is_resigned, _dup_employee_with_orphan_heal,
delete_employee_record) MOVED verbatim from server.py."""
import re
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException
from pydantic import BaseModel

import io

from server import (  # noqa: E402
    _apply_group_policy_on_create,
    _firm_offline_salary_enabled,
    _hash_pin,
    _next_employee_code,
    _normalise_phone,
    _parse_any_date,
    build_compliance_structure,
    compliance_gross_total,
    db,
    get_user_from_token,
    logger,
    now_iso,
    require_role,
)

router = APIRouter(prefix="/api")
api = router


def _employee_is_resigned(u: Dict[str, Any]) -> bool:
    """Iter 295 (user directive) — True when the employee has resigned /
    left (exit_date, resign_date, employment_status or active=False). A
    resigned employee must NEVER block a re-joining with the same mobile
    number or email (rehire case)."""
    status = str(u.get("employment_status") or "").strip().lower()
    return bool(
        u.get("exit_date")
        or u.get("resign_date")
        or status in ("resigned", "exited", "left", "terminated")
        or u.get("active") is False
    )


async def _dup_employee_with_orphan_heal(
    phone: Optional[str], email: Optional[str]
) -> Optional[Dict[str, Any]]:
    """Return an existing EMPLOYEE user that blocks this phone/email — after
    auto-healing orphans. An orphan is an employee whose ``company_id`` points
    to a firm that no longer exists (force-deleted): such stale records are
    removed (with their sessions) so the phone/email can be reused when the
    same people are re-imported. Employees of a LIVE firm still block.

    User report (Iter 134): after Force Delete of a firm from Company Master,
    re-importing the same employees said "phone / email already registered"
    because stale user docs survived on the production DB.

    Iter 295 (user directive) — RESIGNED / LEFT employees never block:
    the old record RELEASES the matched phone/email (archived to
    released_phone / released_email) so the rehired person's logins resolve
    to the NEW record. Applies to Add Employee AND Bulk Import (shared)."""
    or_q: List[Dict[str, Any]] = []
    if phone:
        or_q.append({"phone": phone})
    if email:
        or_q.append({"email": email})
    if not or_q:
        return None
    stale_ids: List[str] = []
    release_phone_ids: List[str] = []
    release_email_ids: List[str] = []
    blocker: Optional[Dict[str, Any]] = None
    async for u in db.users.find(
        {"$or": or_q, "role": "employee"},
        {"_id": 0, "user_id": 1, "name": 1, "company_id": 1, "employee_code": 1,
         "phone": 1, "email": 1, "exit_date": 1, "resign_date": 1,
         "employment_status": 1, "active": 1},
    ):
        live = None
        if u.get("company_id"):
            live = await db.companies.find_one(
                {"company_id": u["company_id"]}, {"_id": 0, "company_id": 1})
        if not live:
            stale_ids.append(u["user_id"])
            continue
        if _employee_is_resigned(u):
            if phone and u.get("phone") == phone:
                release_phone_ids.append(u["user_id"])
            if email and (u.get("email") or "").lower() == (email or "").lower():
                release_email_ids.append(u["user_id"])
            continue
        blocker = blocker or u
    if stale_ids:
        await db.users.delete_many({"user_id": {"$in": stale_ids}})
        await db.user_sessions.delete_many({"user_id": {"$in": stale_ids}})
        logger.info(
            "[dup-heal] removed %d orphan employee record(s) for phone=%s email=%s "
            "(firms already deleted)", len(stale_ids), phone, email)
    if not blocker:
        # Free the identifiers held by resigned records so the new joining
        # owns them exclusively (old value archived for audit).
        if release_phone_ids:
            await db.users.update_many(
                {"user_id": {"$in": release_phone_ids}},
                {"$set": {"released_phone": phone, "identifier_released_at": now_iso()},
                 "$unset": {"phone": "", "phone_e164": ""}})
            logger.info("[dup-heal] released phone %s from %d resigned employee(s) for rehire",
                        phone, len(release_phone_ids))
        if release_email_ids:
            await db.users.update_many(
                {"user_id": {"$in": release_email_ids}},
                {"$set": {"released_email": email, "identifier_released_at": now_iso()},
                 "$unset": {"email": ""}})
            logger.info("[dup-heal] released email %s from %d resigned employee(s) for rehire",
                        email, len(release_email_ids))
    return blocker


# ---------------------------------------------------------------------------
# Delete employees (super_admin or company_admin scoped)
# ---------------------------------------------------------------------------
@api.post("/admin/employees")
async def admin_create_employee(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    """Admin-facing employee creation.

    Called from the web portal "Add Employee" form.  Creates an
    already-approved employee under the caller's firm (or an explicit
    ``company_id`` for super/sub-admins).  A temp 6-digit PIN is
    generated and returned so the admin can share it with the new
    hire; ``pin_must_change=True`` forces the employee to reset it on
    first login.

    Required fields: ``name`` and either ``phone`` or ``email``.
    Optional fields cover the entire master sheet — designation,
    department, employee_code, DOJ, salary, statutory IDs, etc.
    """
    admin = await get_user_from_token(authorization)
    require_role(admin, ["company_admin", "super_admin", "sub_admin"])

    name = str(payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Employee name is required")

    phone = _normalise_phone(str(payload.get("phone") or ""))
    email = (str(payload.get("email") or "").strip().lower()) or None
    # Iter 246 (user rollback) — Mobile is mandatory again for new
    # employees so they can log in.
    if not phone and not email:
        raise HTTPException(
            status_code=400,
            detail="Provide at least a phone number or email so the employee can log in.",
        )

    # Resolve company_id
    if admin["role"] == "company_admin":
        cid = admin.get("company_id")
    else:
        cid = payload.get("company_id") or admin.get("company_id")
    if not cid:
        raise HTTPException(
            status_code=400,
            detail="Company is required — pick a firm before adding an employee.",
        )
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    # Duplicate checks — ONLY employee records block (user directive: an
    # Employer/Firm-Master mobile can also be registered as an Employee).
    # Orphans (employees of a force-deleted firm) are auto-healed so the
    # same phone/email can be re-registered.
    if phone:
        if await _dup_employee_with_orphan_heal(phone, None):
            raise HTTPException(
                status_code=409,
                detail="This phone number is already registered.",
            )
    if email:
        if await _dup_employee_with_orphan_heal(None, email):
            raise HTTPException(
                status_code=409,
                detail="This email is already registered.",
            )

    # Auto-generated temp PIN — admin can share with the new hire.
    import secrets as _secrets
    temp_pin = f"{_secrets.randbelow(1_000_000):06d}"

    # Iter 164 — On/Off-roll gated by Firm Master 'Offline Salary': when
    # the firm has Offline Salary DISABLED, the employee joins On-roll
    # directly regardless of what the form sent.
    _onroll_in = payload.get("is_onroll")
    _onroll_val = True if _onroll_in is None else bool(_onroll_in)
    if _onroll_val is False and not await _firm_offline_salary_enabled(cid):
        _onroll_val = True

    # Copy over allowed employee master fields.
    _emp_code_in = (str(payload.get("employee_code") or "").strip() or None)
    # Iter 375 (user rule) — letters & digits ONLY in Employee Code.
    if _emp_code_in and not re.fullmatch(r"[A-Za-z0-9]+", _emp_code_in):
        raise HTTPException(
            status_code=400,
            detail="Employee Code can contain only letters and numbers — "
                   "no special characters or spaces",
        )
    doc: Dict[str, Any] = {
        "user_id": f"user_{uuid.uuid4().hex[:12]}",
        "email": email,
        "phone": phone,
        "name": name,
        "picture": None,
        "role": "employee",
        "company_id": cid,
        "employee_code": _emp_code_in,
        # Iter 94 — Bio Code (device enrolment no.) settable at creation
        # so Add form matches the master-sheet columns.
        "bio_code": (str(payload.get("bio_code") or "").strip() or None),
        # Iter 142 — per-employee OT flag (None = default allowed).
        "ot_applicable": (bool(payload.get("ot_applicable"))
                          if payload.get("ot_applicable") is not None else None),
        "father_name": payload.get("father_name") or None,
        "mother_name": payload.get("mother_name") or None,
        "gender": payload.get("gender") or None,
        "dob": payload.get("dob") or None,
        "doj": payload.get("doj") or None,
        "designation": payload.get("designation") or None,
        "department": payload.get("department") or None,
        "employee_type": payload.get("employee_type") or None,
        "employee_group": payload.get("employee_group") or None,
        "is_onroll": _onroll_val,
        "shift_start": payload.get("shift_start") or None,
        "shift_end": payload.get("shift_end") or None,
        "salary_mode": payload.get("salary_mode") or None,
        # Iter 94 — separate rate basis for the Compliance salary part.
        "compliance_salary_mode": payload.get("compliance_salary_mode") or None,
        "salary_monthly": payload.get("salary_monthly"),
        "compliance_gross": payload.get("compliance_gross"),
        # Iter 126g — Compliance Basic + PF Basic (EPF ceiling rule).
        "compliance_basic": payload.get("compliance_basic"),
        "pf_basic": payload.get("pf_basic"),
        # Iter 126i — VPF (Voluntary PF)
        "vpf_enabled": bool(payload.get("vpf_enabled") or False),
        "vpf_amount": payload.get("vpf_amount"),
        # Iter 341 — EPS Disable (not eligible for Pension; ECR EPS = 0).
        "eps_disabled": bool(payload.get("eps_disabled") or False),
        # Iter 387 — configurable statutory module: per-employee PF flags.
        "higher_pension": bool(payload.get("higher_pension") or False),
        "intl_worker": bool(payload.get("intl_worker") or False),
        "excluded_employee": bool(payload.get("excluded_employee") or False),
        # Iter 387 — per-employee ESIC master fields.
        "esic_temp_exempt": bool(payload.get("esic_temp_exempt") or False),
        "esic_reg_status": payload.get("esic_reg_status") or None,
        "dispensary": payload.get("dispensary") or None,
        "esic_join_date": payload.get("esic_join_date") or None,
        "esic_exit_date": payload.get("esic_exit_date") or None,
        "actual_salary_allowances": payload.get("actual_salary_allowances") or [],
        "actual_salary_deductions": payload.get("actual_salary_deductions") or [],
        # Iter 91 — fixed Actual structure saved from the Add form
        # (same shape as the Employee Master Salary Update modal).
        "salary_structure_actual": payload.get("salary_structure_actual") or [],
        "compliance_salary_allowances": payload.get("compliance_salary_allowances") or [],
        "compliance_salary_deductions": payload.get("compliance_salary_deductions") or [],
        # Iter 137 — interlinked compliance structure (single source of truth)
        "salary_structure_compliance": build_compliance_structure(
            payload.get("compliance_basic"),
            payload.get("compliance_salary_allowances") or [],
            payload.get("compliance_salary_mode"),
        ),
        "half_day_hrs": payload.get("half_day_hrs"),
        "full_day_hrs": payload.get("full_day_hrs"),
        # Statutory identifiers
        "uan_no": payload.get("uan_no") or None,
        "pf_no": payload.get("pf_no") or None,
        "esi_ip_no": payload.get("esi_ip_no") or None,
        "pan_no": payload.get("pan_no") or None,
        "pan_name": payload.get("pan_name") or None,
        # Iter 275 — Name as printed on the Aadhaar card.
        "aadhaar_name": (str(payload.get("aadhaar_name") or "").strip().upper() or None),
        "aadhaar_no": payload.get("aadhaar_no") or None,
        "bank_name": payload.get("bank_name") or None,
        # Iter 272 — Branch Name (auto-filled from IFSC lookup, editable).
        "bank_branch": (str(payload.get("bank_branch") or "").strip() or None),
        "bank_account": payload.get("bank_account") or None,
        "bank_ifsc": payload.get("bank_ifsc") or None,
        "upi_id": payload.get("upi_id") or None,
        "blood_group": payload.get("blood_group") or None,
        "marital_status": payload.get("marital_status") or None,
        "spouse_name": payload.get("spouse_name") or None,
        "pay_mode": payload.get("pay_mode") or "Bank",
        "address": (str(payload.get("address") or "").strip() or None),
        # Iter 159 — structured location (PIN auto-lookup on the form).
        "pincode": (str(payload.get("pincode") or "").strip() or None),
        "district": (str(payload.get("district") or "").strip() or None),
        "state": (str(payload.get("state") or "").strip() or None),
        # Iter 109 — extra master fields: addresses, emergency & family
        "permanent_address": (str(payload.get("permanent_address") or "").strip() or None),
        # Iter 271 (user request) — separate PIN Code for the Permanent Address.
        "permanent_pincode": (str(payload.get("permanent_pincode") or "").strip() or None),
        "emergency_contact_name": (str(payload.get("emergency_contact_name") or "").strip() or None),
        "emergency_contact_phone": (str(payload.get("emergency_contact_phone") or "").strip() or None),
        "family_members": [
            {"name": str(f.get("name") or "").strip(),
             "relation": str(f.get("relation") or "").strip(),
             "dob": str(f.get("dob") or "").strip(),
             # Iter 271 — per-member Aadhaar No. (12 digits) + Nominee tick.
             "aadhaar_no": re.sub(r"\D", "", str(f.get("aadhaar_no") or ""))[:12] or None,
             "is_nominee": bool(f.get("is_nominee"))}
            for f in (payload.get("family_members") or [])
            if isinstance(f, dict) and str(f.get("name") or "").strip()
        ],
        # Admin creation → already approved, already onboarded.
        "onboarded": True,
        "onboarded_at": now_iso(),
        "approval_status": "approved",
        "approval_requested_at": now_iso(),
        "approved_at": now_iso(),
        "approved_by": admin.get("user_id"),
        "has_pin": True,
        "pin_hash": _hash_pin(temp_pin),
        "pin_must_change": True,
        "pin_set_at": now_iso(),
        "created_at": now_iso(),
        "created_by_admin": admin.get("user_id"),
    }
    # Auto-assign employee_code if the admin left it blank.
    if not doc["employee_code"]:
        try:
            code = await _next_employee_code(cid)
            if code:
                doc["employee_code"] = code
        except Exception:
            pass

    # Iter 95e — Shift comes from the Shift Master ONLY (no free-typed
    # in/out times on the Add form). ``shift_id`` lands on the employee's
    # attendance_policy_override so the grids resolve it; the master's
    # start/end are mirrored onto shift_start/shift_end for display.
    _shift_id = str(payload.get("shift_id") or "").strip()
    if _shift_id:
        _shift = await db.shift_masters.find_one(
            {"shift_id": _shift_id}, {"_id": 0, "shift_id": 1, "start": 1, "end": 1},
        )
        if not _shift:
            raise HTTPException(status_code=400, detail="Unknown shift — pick one from the Shift Master")
        doc["attendance_policy_override"] = {"shift_id": _shift_id}
        doc["shift_start"] = _shift.get("start")
        doc["shift_end"] = _shift.get("end")

    # Iter 75 — Auto-inherit group policy if the employee was tagged with
    # a known group. Preserves any explicit policy fields already on the
    # doc (per-employee overrides win).
    if doc.get("employee_group"):
        merged = await _apply_group_policy_on_create(
            cid,
            doc["employee_group"],
            existing_policy=doc.get("employee_policy"),
        )
        if merged:
            doc["employee_policy"] = merged
            # Mirror to legacy fields the payroll loop still reads.
            if merged.get("fullday_hours") is not None and doc.get("full_day_hrs") is None:
                doc["full_day_hrs"] = float(merged["fullday_hours"])
            if merged.get("halfday_hours") is not None and doc.get("half_day_hrs") is None:
                doc["half_day_hrs"] = float(merged["halfday_hours"])

    # Iter 285 — Onboarding Approval Workflow: when the firm policy is
    # enabled, new employees start as PENDING APPROVAL instead of active.
    _ob_cfg = (company.get("onboarding_approval") or {})
    if _ob_cfg.get("enabled"):
        doc["onboarding_status"] = "pending_approval"
        doc["onboarding_pending_since"] = now_iso()

    await db.users.insert_one(doc)
    logger.info(
        f"[ADMIN CREATE EMPLOYEE] {name} ({phone or email}) → company={cid} by {admin.get('email')}"
    )
    # Iter 103 — automated email trigger
    try:
        from routes.email_notifications import fire_email_event
        await fire_email_event("employee_joined", company_id=cid,
                               employee_user_id=doc["user_id"],
                               details=f"Employee code {doc.get('employee_code') or ''}")
    except Exception:
        pass
    # Iter 267 — Sync Engine: push new employee to all sync-enabled machines.
    try:
        from routes.sync_engine import enqueue_employee_sync
        await enqueue_employee_sync(cid, doc["user_id"], "create",
                                    actor=admin.get("user_id", "system"))
    except Exception:
        pass
    # Iter 395 — WhatsApp welcome message (if automation enabled for firm).
    try:
        from utils.whatsapp_engine import notify_event as _wa_notify
        await _wa_notify("welcome", cid, doc["user_id"])
    except Exception:
        pass
    return {
        "ok": True,
        "user_id": doc["user_id"],
        "employee_code": doc.get("employee_code"),
        "temp_pin": temp_pin,
        "message": (
            f"Employee added. Share the temp PIN {temp_pin} with them — "
            "they will be forced to change it on first login."
        ),
    }


@api.post("/admin/employees/bulk-import")
async def admin_employees_bulk_import(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    """CSV bulk-import — creates many employees in one call.

    Payload:
      * ``company_id`` (required for super/sub-admin; auto-forced for
        company-admin).
      * ``rows``       — list of dicts, each row matches the same schema
        as ``POST /admin/employees`` (``name`` + ``phone`` OR ``email``
        required).

    Behaviour:
      * Idempotent per-phone/email: rows that duplicate an existing user
        are reported as ``skipped_duplicates`` (not created).
      * Each new user gets an auto-assigned employee code and a random
        6-digit temp PIN with ``pin_must_change=True`` (same rule as
        the single Add-Employee endpoint).
      * Rows missing required fields are reported in ``errors`` with a
        row number and reason — the rest of the import still succeeds.
    """
    admin = await get_user_from_token(authorization)
    require_role(admin, ["company_admin", "super_admin", "sub_admin"])

    if admin["role"] == "company_admin":
        cid = admin.get("company_id")
    else:
        cid = payload.get("company_id")
    if not cid:
        raise HTTPException(status_code=400, detail="company_id is required")
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    rows = payload.get("rows") or []
    if not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=400, detail="`rows` must be a non-empty list")

    # User directive — when a CSV row has no Allowance/Deduction columns,
    # set them up from the heads ENABLED in the Firm Master company policy
    # (Sections 5 & 6). Amounts start at 0 and are edited later per employee.
    _fm_policy = await db.firm_masters.find_one(
        {"company_id": cid}, {"_id": 0, "allowances": 1, "deductions": 1}) or {}
    policy_allow_lines = [{"head": h, "amount": 0.0}
                          for h, on in (_fm_policy.get("allowances") or {}).items() if on]
    policy_ded_lines = [{"head": h, "amount": 0.0}
                        for h, on in (_fm_policy.get("deductions") or {}).items() if on]

    import secrets as _secrets

    # Iter 125 — friendly header aliases so the CSV can use the firm's own
    # column names (EMPLOYEE PFNO, Name As Per Pan Card, Mobile1 …).
    _ALIASES: Dict[str, str] = {
        "employee pfno": "pf_no", "pf no": "pf_no", "pfno": "pf_no",
        "uan no": "uan_no", "uan": "uan_no",
        "employee esino": "esi_ip_no", "esi no": "esi_ip_no", "esino": "esi_ip_no",
        "employee name": "name",
        "employee father name": "father_name", "father name": "father_name",
        "emp type": "employee_type", "emp_type": "employee_type",
        "emp group": "employee_type", "emp_group": "employee_type",
        "employee group": "employee_type", "group": "employee_type",
        "marital status": "marital_status",
        "employee basic": "basic_salary", "basic": "basic_salary",
        "pf basic": "pf_basic",
        "conv": "conveyance",
        "over time": "over_time", "overtime": "over_time",
        "gross pay": "compliance_gross",
        "present add": "present_address", "present address": "present_address",
        "permanent add": "permanent_address", "permanent address": "permanent_address",
        # Iter 266 — City / State / Pin Code column aliases.
        "city": "city", "town": "city",
        "state": "state", "province": "state",
        "pin code": "pincode", "pincode": "pincode", "pin_code": "pincode",
        "postal code": "pincode", "zip": "pincode", "zip code": "pincode",
        "panno": "pan_no", "pan no": "pan_no", "pan": "pan_no",
        "name as per pan card": "name_as_per_pan", "name as per pan": "name_as_per_pan",
        "aadhar card no": "aadhaar_no", "aadhaar no": "aadhaar_no", "aadhar no": "aadhaar_no",
        "name on aadhar card": "name_as_per_aadhar", "name as per aadhar": "name_as_per_aadhar",
        "bank name": "bank_name",
        "bank address": "bank_address",
        "account no": "bank_account", "account number": "bank_account",
        "name on bank ac": "account_holder", "name on bank a/c": "account_holder",
        "name on bank account": "account_holder",
        "ifsc code": "bank_ifsc", "ifsc": "bank_ifsc",
        "mobile1": "phone", "mobile 1": "phone", "mobile": "phone",
        "mobile2": "phone2", "mobile 2": "phone2",
        "pay mode": "pay_mode",
        "pay basis": "pay_basis",
        "resign date": "resign_date",
        "basic salary actual": "salary_monthly",
    }

    def _norm_key(k: str) -> str:
        return re.sub(r"\s+", " ", str(k or "").strip().lower())

    def _normalise_row(raw_row: Dict[str, Any]) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        for k, v in raw_row.items():
            nk = _norm_key(k)
            out[_ALIASES.get(nk, nk.replace(" ", "_"))] = v
        return out

    def _num(v: Any) -> Optional[float]:
        try:
            s = str(v).replace(",", "").strip()
            return float(s) if s else None
        except Exception:
            return None

    def _parse_salary_lines(raw: Any) -> List[Dict[str, Any]]:
        """Iter 74 — Parse ``HRA:2000|Convey:500|SpecialAllow:1000`` into
        the same ``[{head, amount}]`` list shape used by the single-add
        endpoint. Silently drops malformed / empty entries; a caller can
        pass ``None`` / empty string safely."""
        if raw is None:
            return []
        # Accept either a pre-parsed list (JSON) or a pipe-separated
        # string produced by the CSV.
        if isinstance(raw, list):
            out: List[Dict[str, Any]] = []
            for item in raw:
                if not isinstance(item, dict):
                    continue
                head = str(item.get("head") or "").strip()
                try:
                    amount = float(item.get("amount") or 0)
                except Exception:
                    continue
                if head and amount:
                    out.append({"head": head, "amount": amount})
            return out
        text = str(raw or "").strip()
        if not text:
            return []
        # Support pipe OR semicolon as separator to be user-friendly.
        parts = [p.strip() for p in text.replace(";", "|").split("|") if p.strip()]
        out2: List[Dict[str, Any]] = []
        for part in parts:
            if ":" not in part:
                continue
            head, amt = part.split(":", 1)
            head = head.strip()
            try:
                amount = float(str(amt).replace(",", "").strip())
            except Exception:
                continue
            if head and amount:
                out2.append({"head": head, "amount": amount})
        return out2

    created: List[Dict[str, Any]] = []
    skipped_duplicates: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for idx, r in enumerate(rows, start=1):
        try:
            r = _normalise_row(r)
            name = str(r.get("name") or "").strip()
            if not name:
                errors.append({"row": idx, "reason": "name is required"})
                continue
            phone = _normalise_phone(str(r.get("phone") or ""))
            email = (str(r.get("email") or "").strip().lower()) or None
            if not phone and not email:
                errors.append({"row": idx, "reason": "phone or email required"})
                continue
            # Duplicate check — only EMPLOYEE records block; an employer/
            # admin using the same mobile is allowed (user directive).
            # Orphans (employees of a force-deleted firm) are auto-healed
            # so re-importing the same people succeeds.
            dup = await _dup_employee_with_orphan_heal(phone, email)
            if dup:
                # Tell the admin WHICH firm already holds this phone/email so
                # duplicate skips are self-explanatory in the import log.
                _dup_firm = await db.companies.find_one(
                    {"company_id": dup.get("company_id")}, {"_id": 0, "name": 1})
                _firm_label = (_dup_firm or {}).get("name") or dup.get("company_id") or "?"
                skipped_duplicates.append({
                    "row": idx,
                    "name": name,
                    "reason": (
                        f"phone / email already registered — {dup.get('name') or 'employee'}"
                        + (f" (code {dup.get('employee_code')})" if dup.get("employee_code") else "")
                        + f" in firm '{_firm_label}'"
                    ),
                    "existing_user_id": dup["user_id"],
                })
                continue
            temp_pin = f"{_secrets.randbelow(1_000_000):06d}"
            # Iter 295 (user directive) — "Resign Date" in the sheet is the
            # employee's Exit / Left date. INTERLINK: `exit_date` is the
            # canonical field the whole system reads (Employee Master
            # Resigned status, punch blocking, dashboards, salary-run
            # exclusion, reports) — so normalise the date and set BOTH
            # fields + employment_status, ensuring imported ex-employees
            # land as Resigned instead of Active.
            _resign_raw = r.get("resign_date")
            _rd_dt = _parse_any_date(_resign_raw) if str(_resign_raw or "").strip() else None
            _resign_iso = _rd_dt.strftime("%Y-%m-%d") if _rd_dt else None
            doc: Dict[str, Any] = {
                "user_id": f"user_{uuid.uuid4().hex[:12]}",
                "email": email,
                "phone": phone,
                "name": name,
                "picture": None,
                "role": "employee",
                "company_id": cid,
                "employee_code": (str(r.get("employee_code") or "").strip() or None),
                "father_name": r.get("father_name") or None,
                "gender": r.get("gender") or None,
                "dob": r.get("dob") or None,
                "doj": r.get("doj") or None,
                "designation": r.get("designation") or None,
                "department": r.get("department") or None,
                "employee_type": r.get("employee_type") or None,
                # Employee Group merged into Employee Type (user directive) —
                # the CSV no longer has an employee_group column; mirror type.
                "employee_group": r.get("employee_type") or None,
                "is_onroll": r.get("is_onroll") if r.get("is_onroll") is not None else True,
                "salary_mode": (
                    str(r.get("pay_basis") or "").strip().lower()
                    if str(r.get("pay_basis") or "").strip().lower() in ("daily", "monthly")
                    else (r.get("salary_mode") or "monthly")),
                "salary_monthly": _num(r.get("salary_monthly")),
                "compliance_gross": _num(r.get("compliance_gross")),
                # Iter 125 — extended master fields from the firm's own
                # bulk-import format.
                "marital_status": r.get("marital_status") or None,
                "basic_salary": _num(r.get("basic_salary")),
                # Iter 137 (user directive) — the import columns ARE the
                # Compliance salary: EMPLOYEE BASIC → Compliance Basic,
                # HRA/CONV → Compliance allowance heads, Gross Pay →
                # Compliance Gross (Basic + Allowances).
                "compliance_basic": _num(r.get("basic_salary")),
                "pf_basic": _num(r.get("pf_basic")),
                "hra": _num(r.get("hra")),
                "conveyance": _num(r.get("conveyance")),
                "over_time": _num(r.get("over_time")),
                "present_address": r.get("present_address") or None,
                "permanent_address": r.get("permanent_address") or None,
                # Iter 266 (user directive) — City / State / Pin Code.
                # ``district`` mirrors City to match the single Add-Employee
                # form + profile which read the district field.
                "city": r.get("city") or None,
                "district": r.get("city") or r.get("district") or None,
                "state": r.get("state") or None,
                "pincode": r.get("pincode") or None,
                "name_as_per_pan": r.get("name_as_per_pan") or None,
                "name_as_per_aadhar": r.get("name_as_per_aadhar") or None,
                "bank_address": r.get("bank_address") or None,
                "account_holder": r.get("account_holder") or None,
                "phone2": (_normalise_phone(str(r.get("phone2") or "")) or None),
                "pay_mode": r.get("pay_mode") or None,
                "pay_basis": r.get("pay_basis") or None,
                "resign_date": _resign_iso or (str(_resign_raw).strip() if str(_resign_raw or "").strip() else None),
                "exit_date": _resign_iso,
                "employment_status": ("resigned" if str(_resign_raw or "").strip() else None),
                "uan_no": r.get("uan_no") or None,
                "pf_no": r.get("pf_no") or None,
                "esi_ip_no": r.get("esi_ip_no") or None,
                "pan_no": r.get("pan_no") or None,
                "aadhaar_no": r.get("aadhaar_no") or None,
                "bank_name": r.get("bank_name") or None,
                "bank_account": r.get("bank_account") or None,
                "bank_ifsc": r.get("bank_ifsc") or None,
                "address": r.get("address") or r.get("present_address") or None,
                # Iter 74 — allowance / deduction line-items via
                # `HRA:2000|Convey:500` style CSV columns. When the CSV has
                # none, DEFAULT to the heads enabled in the Firm Master
                # company policy (user directive).
                "actual_salary_allowances": (
                    _parse_salary_lines(r.get("actual_allowances") or r.get("allowances"))
                    or [dict(x) for x in policy_allow_lines]),
                "actual_salary_deductions": (
                    _parse_salary_lines(r.get("actual_deductions") or r.get("deductions"))
                    or [dict(x) for x in policy_ded_lines]),
                "compliance_salary_allowances": (
                    _parse_salary_lines(r.get("compliance_allowances"))
                    # Iter 137 (user directive) — HRA / CONV import columns
                    # are Compliance-salary allowance heads.
                    or [ln for ln in [
                        {"head": "HRA", "amount": _num(r.get("hra")) or 0},
                        {"head": "CONV.", "amount": _num(r.get("conveyance")) or 0},
                    ] if ln["amount"] > 0]),
                "compliance_salary_deductions": _parse_salary_lines(r.get("compliance_deductions")),
                "onboarded": True,
                "onboarded_at": now_iso(),
                "approval_status": "approved",
                "approval_requested_at": now_iso(),
                "approved_at": now_iso(),
                "approved_by": admin.get("user_id"),
                "has_pin": True,
                "pin_hash": _hash_pin(temp_pin),
                "pin_must_change": True,
                "pin_set_at": now_iso(),
                "created_at": now_iso(),
                "created_by_admin": admin.get("user_id"),
                "bulk_imported": True,
            }
            if not doc["employee_code"]:
                try:
                    code = await _next_employee_code(cid)
                    if code:
                        doc["employee_code"] = code
                except Exception:
                    pass
            # Iter 137 — interlinked compliance structure + linked Gross
            # (= Compliance Basic + Σ allowance lines).
            doc["salary_structure_compliance"] = build_compliance_structure(
                doc.get("compliance_basic"),
                doc.get("compliance_salary_allowances"),
                doc.get("salary_mode"),
            )
            _cg = compliance_gross_total(
                doc.get("compliance_basic"), doc.get("compliance_salary_allowances"))
            if _cg > 0:
                doc["compliance_gross"] = _cg
            # Iter 75 — Inherit group policy when the CSV row provides
            # an `employee_group` matching an existing template.
            if doc.get("employee_group"):
                merged = await _apply_group_policy_on_create(
                    cid, doc["employee_group"], existing_policy=doc.get("employee_policy"),
                )
                if merged:
                    doc["employee_policy"] = merged
                    if merged.get("fullday_hours") is not None:
                        doc["full_day_hrs"] = float(merged["fullday_hours"])
                    if merged.get("halfday_hours") is not None:
                        doc["half_day_hrs"] = float(merged["halfday_hours"])
            await db.users.insert_one(doc)
            created.append({
                "row": idx,
                "name": name,
                "user_id": doc["user_id"],
                "employee_code": doc.get("employee_code"),
                "temp_pin": temp_pin,
            })
        except Exception as ex:
            errors.append({"row": idx, "reason": str(ex)})

    logger.info(
        f"[ADMIN BULK IMPORT] cid={cid} created={len(created)} skipped={len(skipped_duplicates)} errors={len(errors)}"
    )
    return {
        "ok": True,
        "created_count": len(created),
        "skipped_count": len(skipped_duplicates),
        "error_count": len(errors),
        "created": created,
        "skipped_duplicates": skipped_duplicates,
        "errors": errors,
    }


@api.get("/admin/employees/bulk-import-template.xlsx")
async def admin_bulk_import_template(
    authorization: Optional[str] = Header(None),
):
    """Excel (.xlsx) bulk-import template. Iter 132 (user directive):
    switched from CSV to Excel because CSV editors mangle numeric fields
    (leading zeros in UAN/ESI/account numbers, dates). Every cell is
    TEXT-formatted so Excel never converts values."""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    admin = await get_user_from_token(authorization)
    require_role(admin, ["company_admin", "super_admin", "sub_admin"])
    # Iter 125 — template matches the firm's own bulk-import format.
    # Iter 132 — "Emp Type" renamed to "Emp Group" (General Master Groups).
    columns = [
        "EMPLOYEE PFNO", "UAN_NO", "EMPLOYEE ESINO",
        "EMPLOYEE NAME", "EMPLOYEE FATHER NAME",
        "Designation", "Department", "Emp Group", "Gender", "Marital Status",
        "DOB", "DOJ",
        "EMPLOYEE BASIC", "PF_BASIC", "HRA", "CONV", "OVER_TIME", "Gross Pay",
        "Present Add", "Permanent Add",
        "City", "State", "Pin Code",
        "PANNo", "Name As Per Pan Card",
        "Aadhar Card No", "Name On Aadhar Card",
        "Bank Name", "Bank Address", "Account No", "Name On Bank Ac", "IFSC Code",
        "Mobile1", "Mobile2",
        "Pay Mode", "Pay Basis", "Resign Date",
        "Basic Salary Actual",
    ]
    sample = [
        "DL/PF/12345", "123456789012", "1122334455",
        "Ramesh Kumar", "Suresh Kumar",
        "Machine Operator", "Weaving", "Worker", "Male", "Married",
        "1990-05-14", "2024-04-01",
        "12000", "12000", "3000", "1500", "0", "18000",
        "House 12 Karol Bagh New Delhi", "Village Rampur UP",
        "New Delhi", "Delhi", "110005",
        "ABCDE1234F", "RAMESH KUMAR",
        "123412341234", "RAMESH KUMAR",
        "HDFC Bank", "Karol Bagh Branch New Delhi", "1234567890", "RAMESH KUMAR", "HDFC0001234",
        "+919812345678", "+919812300000",
        "Bank", "Monthly", "",
        "22000",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1D4ED8")
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = hdr_font
        c.fill = hdr_fill
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(28, len(col) + 4))
    for i, val in enumerate(sample, start=1):
        c = ws.cell(row=2, column=i, value=val)
        c.number_format = "@"  # TEXT — preserves leading zeros / long numbers
    # Pre-format 500 blank rows as TEXT so pasted data is never mangled.
    for r in range(3, 503):
        for i in range(1, len(columns) + 1):
            ws.cell(row=r, column=i).number_format = "@"
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="employee_bulk_import_template.xlsx"',
        },
    )


class BulkImportParseBody(BaseModel):
    file_base64: str
    filename: Optional[str] = None


@api.post("/admin/employees/bulk-import-parse")
async def admin_bulk_import_parse(
    payload: BulkImportParseBody,
    authorization: Optional[str] = Header(None),
):
    """Parse an uploaded Excel (.xlsx/.xls) bulk-import file server-side and
    return {headers, rows} of STRINGS — numeric cells are stringified
    losslessly (no scientific notation / trailing .0), dates become
    YYYY-MM-DD. Iter 132 (user directive: Excel instead of CSV)."""
    import base64 as _b64
    from openpyxl import load_workbook
    admin = await get_user_from_token(authorization)
    require_role(admin, ["company_admin", "super_admin", "sub_admin"])
    try:
        blob = _b64.b64decode(payload.file_base64 or "", validate=False)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid file encoding")
    if not blob:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Could not read the Excel file — please upload a .xlsx made from the template.",
        )
    ws = wb.worksheets[0]

    def cell_str(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        if isinstance(v, float):
            if v.is_integer():
                return str(int(v))
            return repr(v)
        return str(v).strip()

    headers: List[str] = []
    rows: List[Dict[str, str]] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            headers = [cell_str(v) for v in row]
            while headers and not headers[-1]:
                headers.pop()
            continue
        vals = [cell_str(v) for v in row[: len(headers)]]
        if not any(vals):
            continue  # skip fully blank rows
        rows.append({headers[i]: (vals[i] if i < len(vals) else "") for i in range(len(headers)) if headers[i]})
    wb.close()
    if not headers:
        raise HTTPException(status_code=400, detail="No header row found in the Excel file")
    return {"headers": [h for h in headers if h], "rows": rows, "rows_count": len(rows)}


async def delete_employee_record(user_id: str, actor: str = "system") -> Dict[str, Any]:
    """Cascade-delete one employee (attendance, leaves, tickets, payslips…).
    Shared by the direct DELETE endpoint and the Super-Admin approval flow
    (Iter 306 user #1 — sub-admin deletes need approval)."""
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        return {"note": "Employee no longer exists"}
    cascade = {}
    for col in ("attendance", "leaves", "tickets", "payslips", "notifications", "user_sessions"):
        r = await db[col].delete_many({"user_id": user_id})
        cascade[col] = r.deleted_count
    await db.users.delete_one({"user_id": user_id})
    logger.info(f"[DELETE employee] {user_id} by {actor} cascade={cascade}")
    # Iter 267 — Sync Engine: remove the employee from all machines. The
    # target's PIN/company are captured before deletion so the removal
    # command can be built even though the user record is gone.
    try:
        pin = str(target.get("bio_code") or "").strip()
        if pin and target.get("company_id"):
            from routes.sync_engine import enqueue_employee_removal
            await enqueue_employee_removal(
                target["company_id"], user_id, pin, target.get("name"),
                actor=actor)
    except Exception:
        pass
    return {"cascade": cascade}


@api.delete("/admin/employees/{user_id}")
async def delete_employee(user_id: str,
                          authorization: Optional[str] = Header(None)):
    """Remove an employee (and their attendance, leaves, tickets, payslips).

    - Super admin can delete any employee.
    - Company admin can only delete employees in their own company.
    - Sub admin deletions are queued for SUPER ADMIN approval (Iter 306).
    - Super admins cannot be deleted via this endpoint (safety guard).
    """
    admin = await get_user_from_token(authorization)
    require_role(admin, ["company_admin", "super_admin", "sub_admin"])
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Employee not found")
    if target.get("role") == "super_admin":
        raise HTTPException(status_code=403, detail="Super admin accounts cannot be deleted")
    # Iter 332 (user request) — employees locked by Legacy Salary Records
    # can NEVER be deleted until the firm's legacy import is UNDONE.
    if target.get("legacy_locked"):
        raise HTTPException(
            status_code=403,
            detail=("This employee is LOCKED — the firm's Legacy Salary "
                    "Records are locked. Click UNDO on the Legacy Import "
                    "first to unlock, then delete."))
    if admin["role"] == "company_admin":
        if not admin.get("company_id") or target.get("company_id") != admin["company_id"]:
            raise HTTPException(status_code=403, detail="Not allowed to delete employees outside your company")
    if admin.get("user_id") == user_id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")

    # Iter 344 (user request) — Sub Super Admin deletes DIRECTLY, no
    # Super Admin approval queue anymore (was Iter 306).
    result = await delete_employee_record(user_id, actor=admin.get("email") or admin.get("user_id"))
    return {"ok": True, **result}

