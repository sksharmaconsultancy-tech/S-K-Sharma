"""Iter 356 — Employee-Wise Yearly Payroll Register (Bonus-Register style).

Register layout: employee info block on the left, months (Apr..Mar) across
the top, one row per salary head, a Total line after every employee and a
Grand Total block at the end. Supports multi-FY periodic export.

Data source: latest compliance salary run per month (compliance_salary_runs)
+ Employee Master (users).

  GET /api/admin/reports/payroll-register            (JSON, paginated)
  GET /api/admin/reports/payroll-register.pdf        (A3 landscape)
  GET /api/admin/reports/payroll-register.xlsx
"""
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException
from fastapi.responses import StreamingResponse

from server import db, get_user_from_token, require_role  # noqa: E402

router = APIRouter(prefix="/api/admin/reports", tags=["payroll-register"])

# head key, label, kind (info/earn/ded/total/net/er), optional?
_HEADS = [
    ("days", "Days Worked", "info", False),
    ("basic", "Basic", "earn", False),
    ("da", "DA", "earn", True),
    ("hra", "HRA", "earn", False),
    ("conveyance", "Conveyance", "earn", False),
    ("special", "Special Allowance", "earn", True),
    ("other_allow", "Other Allowance", "earn", False),
    ("ot", "Overtime", "earn", True),
    ("incentive", "Incentive", "earn", True),
    ("arrear", "Arrear", "earn", True),
    ("gross", "Gross Salary", "total", False),
    ("pf_ee", "Employee PF", "ded", False),
    ("esic_ee", "Employee ESIC", "ded", False),
    ("pt", "Professional Tax", "ded", True),
    ("lwf", "LWF", "ded", True),
    ("tds", "TDS", "ded", True),
    ("advance", "Advance Recovery", "ded", True),
    ("loan", "Loan EMI", "ded", True),
    ("other_ded", "Other Deduction", "ded", True),
    ("total_ded", "Total Deduction", "total", False),
    ("net", "Net Salary", "net", False),
    ("pf_er", "Employer PF", "er", False),
    ("esic_er", "Employer ESIC", "er", False),
    ("ctc", "CTC Cost", "total", False),
]


def _fy_months(fy_start_year: int, fy_years: int = 1) -> List[Dict[str, str]]:
    out = []
    for j in range(max(1, min(fy_years, 5))):
        for i in range(12):
            m, y = 4 + i, fy_start_year + j
            if m > 12:
                m, y = m - 12, y + 1
            out.append({"key": f"{y:04d}-{m:02d}",
                        "label": date(y, m, 1).strftime("%b-%y").upper()})
    return out


def _f(v) -> float:
    try:
        return round(float(v or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _month_vals(r: dict) -> Dict[str, float]:
    basic = _f(r.get("basic"))
    hra = _f(r.get("hra"))
    conv = _f(r.get("conveyance"))
    special = _f(r.get("special"))
    other_allow = _f(r.get("medical")) + _f(r.get("others"))
    # dynamic per-head allowances list (legacy-import structures)
    da = incentive = arrear = 0.0
    for a in r.get("allowances") or []:
        h = str(a.get("head") or "").lower()
        amt = _f(a.get("amount"))
        if h.startswith("da") or "dearness" in h:
            da += amt
        elif "incent" in h:
            incentive += amt
        elif "arrear" in h:
            arrear += amt
    ot = _f(r.get("ot_pay"))
    gross = _f(r.get("gross_paid")) or _f(r.get("monthly_gross"))
    # Freeze-salary difference allocation: gross_paid may exceed the head-wise
    # sum (difference between Imported and Master gross allocated per firm
    # settings). Absorb that residual into Other Allowance so the register
    # balances; unexplained residuals (no import markers) stay flagged.
    comp_sum = basic + da + hra + conv + special + other_allow + ot + \
        incentive + arrear
    residual = round(gross - comp_sum, 2)
    if abs(residual) > 2 and ("imported_gross" in r or "difference" in r):
        other_allow = round(other_allow + residual, 2)
    pf_ee = _f(r.get("pf_employee"))
    esic_ee = _f(r.get("esic_employee"))
    pt = _f(r.get("pt"))
    lwf = _f(r.get("lwf"))
    tds = _f(r.get("tds"))
    advance = _f(r.get("advance")) + _f(r.get("advance_recovery"))
    loan = _f(r.get("loan_emi")) + _f(r.get("loan"))
    other_ded = _f(r.get("other_deduction"))
    total_ded = _f(r.get("total_deduction")) or round(
        pf_ee + esic_ee + pt + lwf + tds + advance + loan + other_ded, 2)
    net = _f(r.get("net")) or round(gross - total_ded, 2)
    pf_er = _f(r.get("pf_employer_total"))
    esic_er = _f(r.get("esic_employer"))
    return {
        "days": _f(r.get("present_days")), "basic": basic, "da": da,
        "hra": hra, "conveyance": conv, "special": special,
        "other_allow": other_allow, "ot": ot, "incentive": incentive,
        "arrear": arrear, "gross": gross, "pf_ee": pf_ee,
        "esic_ee": esic_ee, "pt": pt, "lwf": lwf, "tds": tds,
        "advance": advance, "loan": loan, "other_ded": other_ded,
        "total_ded": total_ded, "net": net, "pf_er": pf_er,
        "esic_er": esic_er,
        "ctc": round(gross + pf_er + esic_er, 2),
    }


def _validate(r: dict, v: Dict[str, float]) -> List[str]:
    """AI validation flags for one employee-month."""
    flags: List[str] = []
    if r.get("pf_applicable") and abs(v["pf_ee"] - round(_f(r.get("pf_wages")) * 0.12)) > 2:
        flags.append("pf_mismatch")
    if r.get("esic_applicable") and v["gross"] > 0:
        import math
        want = math.ceil(_f(r.get("esic_wage_base")) * 0.0075)
        if abs(v["esic_ee"] - want) > 2:
            flags.append("esic_mismatch")
    if v["net"] < 0:
        flags.append("negative_net")
    earn_sum = round(sum(v[k] for k in (
        "basic", "da", "hra", "conveyance", "special", "other_allow",
        "ot", "incentive", "arrear")), 2)
    if abs(earn_sum - v["gross"]) > 2:
        flags.append("gross_mismatch")
    if v["gross"] > 0 and v["days"] <= 0:
        flags.append("missing_attendance")
    if (v["advance"] + v["loan"]) > v["gross"] > 0:
        flags.append("loan_recovery_error")
    return flags


async def _register_data(company_id: str, fy_start_year: int, fy_years: int,
                         department: str = "", designation: str = "",
                         category: str = "", status: str = "",
                         bank: str = "", skip: int = 0,
                         limit: int = 0) -> Dict[str, Any]:
    company = await db.companies.find_one(
        {"company_id": company_id}, {"_id": 0, "name": 1, "logo_base64": 1})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    months = _fy_months(fy_start_year, fy_years)
    runs: Dict[str, dict] = {}
    for mo in months:
        run = await db.compliance_salary_runs.find_one(
            {"company_id": company_id, "month": mo["key"]}, {"_id": 0},
            sort=[("generated_at", -1)])
        if run:
            runs[mo["key"]] = run
    # employee master info
    q: Dict[str, Any] = {"company_id": company_id, "role": "employee"}
    if department:
        q["department"] = {"$regex": department, "$options": "i"}
    if designation:
        q["designation"] = {"$regex": designation, "$options": "i"}
    if category:
        q["$or"] = [{"employee_group": {"$regex": category, "$options": "i"}},
                    {"employee_type": {"$regex": category, "$options": "i"}}]
    if bank:
        q["bank_name"] = {"$regex": bank, "$options": "i"}
    users = {u["user_id"]: u async for u in db.users.find(q, {
        "_id": 0, "user_id": 1, "employee_code": 1, "name": 1,
        "father_name": 1, "designation": 1, "department": 1, "doj": 1,
        "exit_date": 1, "resign_date": 1, "pf_no": 1, "esi_ip_no": 1,
        "uan_no": 1, "bank_name": 1, "bank_account": 1, "bank_ifsc": 1,
        "employment_status": 1, "disabled": 1, "active": 1})}
    if status:
        left = ("exited", "resigned", "terminated", "inactive", "left")
        if status == "active":
            users = {k: u for k, u in users.items()
                     if not (u.get("disabled") or u.get("active") is False
                             or str(u.get("employment_status") or "").lower() in left)}
        elif status == "left":
            users = {k: u for k, u in users.items()
                     if u.get("disabled") or u.get("active") is False
                     or str(u.get("employment_status") or "").lower() in left}
    # employee-month values
    per_emp: Dict[str, Dict[str, Any]] = {}
    dup_codes: Dict[str, int] = {}
    for mkey, run in runs.items():
        for r in run.get("rows") or []:
            uid = r.get("user_id")
            if uid not in users:
                continue
            v = _month_vals(r)
            flags = _validate(r, v)
            e = per_emp.setdefault(uid, {"months": {}, "flags": {}})
            e["months"][mkey] = v
            if flags:
                e["flags"][mkey] = flags
            dup_codes[f"{mkey}|{r.get('employee_code')}"] = \
                dup_codes.get(f"{mkey}|{r.get('employee_code')}", 0) + 1
    # order + paginate employees (only ones having any salary row)
    def _code_key(uid):
        try:
            return (0, float(str(users[uid].get("employee_code") or "").strip() or 1e12))
        except ValueError:
            return (1, 0.0)
    uids = sorted(per_emp.keys(), key=_code_key)
    total_employees = len(uids)
    if limit:
        uids = uids[skip: skip + limit]
    rows: List[Dict[str, Any]] = []
    grand = {k: 0.0 for k, *_ in _HEADS}
    used_heads = set()
    for i, uid in enumerate(sorted(per_emp.keys(), key=_code_key), 1):
        e, u = per_emp[uid], users[uid]
        tot = {k: round(sum(m.get(k, 0) for m in e["months"].values()), 2)
               for k, *_ in _HEADS}
        for k, val in tot.items():
            grand[k] = round(grand[k] + val, 2)
            if val:
                used_heads.add(k)
        # duplicate-employee flag
        if any(dup_codes.get(f"{mk}|{u.get('employee_code')}", 0) > 1
               for mk in e["months"]):
            e["flags"]["_employee"] = ["duplicate_employee"]
        if uid in uids:
            rows.append({
                "sr": i, "user_id": uid,
                "employee_code": u.get("employee_code"),
                "name": u.get("name"), "father_name": u.get("father_name"),
                "designation": u.get("designation"),
                "department": u.get("department"), "doj": u.get("doj"),
                "dol": u.get("exit_date") or u.get("resign_date"),
                "pf_no": u.get("pf_no"), "esic_no": u.get("esi_ip_no"),
                "uan": u.get("uan_no"), "bank_name": u.get("bank_name"),
                "account_no": u.get("bank_account"), "ifsc": u.get("bank_ifsc"),
                "months": e["months"], "totals": tot, "flags": e["flags"],
            })
    heads = [{"key": k, "label": lb, "kind": kd} for k, lb, kd, opt in _HEADS
             if (not opt) or k in used_heads]
    fy_label = (f"FY {fy_start_year}-{str(fy_start_year + 1)[-2:]}"
                if fy_years <= 1 else
                f"FY {fy_start_year}-{str(fy_start_year + fy_years)[-2:]}")
    return {
        "company_name": company.get("name"),
        "logo_base64": company.get("logo_base64"),
        "fy_label": fy_label, "months": months,
        "months_covered": sorted(runs.keys()), "heads": heads,
        "rows": rows, "grand": grand,
        "total_employees": total_employees,
        "print_date": datetime.now(timezone.utc).strftime("%d-%m-%Y"),
    }


def _params(company_id, fy_start_year, fy_years):
    fy = fy_start_year or (date.today().year if date.today().month >= 4
                           else date.today().year - 1)
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    return company_id, fy, max(1, min(fy_years or 1, 5))


@router.get("/payroll-register")
async def payroll_register(
    company_id: Optional[str] = None, fy_start_year: int = 0,
    fy_years: int = 1, department: str = "", designation: str = "",
    category: str = "", status: str = "", bank: str = "",
    skip: int = 0, limit: int = 25,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if admin["role"] == "company_admin":
        company_id = admin.get("company_id")
    cid, fy, fys = _params(company_id, fy_start_year, fy_years)
    data = await _register_data(cid, fy, fys, department, designation,
                                category, status, bank, skip, limit)
    data.pop("logo_base64", None)
    return data


@router.get("/payroll-register.xlsx")
async def payroll_register_xlsx(
    company_id: Optional[str] = None, fy_start_year: int = 0,
    fy_years: int = 1, department: str = "", designation: str = "",
    category: str = "", status: str = "", bank: str = "",
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if admin["role"] == "company_admin":
        company_id = admin.get("company_id")
    cid, fy, fys = _params(company_id, fy_start_year, fy_years)
    data = await _register_data(cid, fy, fys, department, designation,
                                category, status, bank)
    buf = _build_xlsx(data)
    fn = f"Payroll_Register_{fy}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@router.get("/payroll-register.pdf")
async def payroll_register_pdf(
    company_id: Optional[str] = None, fy_start_year: int = 0,
    fy_years: int = 1, department: str = "", designation: str = "",
    category: str = "", status: str = "", bank: str = "",
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if admin["role"] == "company_admin":
        company_id = admin.get("company_id")
    cid, fy, fys = _params(company_id, fy_start_year, fy_years)
    data = await _register_data(cid, fy, fys, department, designation,
                                category, status, bank)
    buf = _build_pdf(data)
    fn = f"Payroll_Register_{fy}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# ---------------------------------------------------------------------------
# Excel builder — register style (merged month headers, employee blocks)
# ---------------------------------------------------------------------------

def _build_xlsx(data: Dict[str, Any]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Register"
    months = data["months"]
    heads = data["heads"]
    ncols = 2 + len(months) + 1  # Particulars + months + Total
    thin = Side(style="thin", color="9AA0A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DDEBF7")
    tot_fill = PatternFill("solid", fgColor="FFF2CC")
    err_fill = PatternFill("solid", fgColor="F8CBAD")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rt = Alignment(horizontal="right")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, f"{data['company_name']} — Employee-Wise Yearly Payroll "
                  f"Register ({data['fy_label']})").font = Font(bold=True, size=13)
    ws.cell(1, 1).alignment = ctr
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2, 1, f"Print date: {data['print_date']}  ·  Employees: "
                  f"{data['total_employees']}").alignment = ctr
    row = 4
    for emp in data["rows"]:
        # employee info block (2 merged lines)
        info1 = (f"Sr {emp['sr']}  ·  Code {emp['employee_code'] or '—'}  ·  "
                 f"{emp['name']}  ·  F/H: {emp['father_name'] or '—'}  ·  "
                 f"{emp['designation'] or '—'}  ·  Dept: {emp['department'] or '—'}"
                 f"  ·  DOJ {emp['doj'] or '—'}  ·  DOL {emp['dol'] or '—'}")
        info2 = (f"PF {emp['pf_no'] or '—'}  ·  ESIC {emp['esic_no'] or '—'}  ·  "
                 f"UAN {emp['uan'] or '—'}  ·  Bank {emp['bank_name'] or '—'}  ·  "
                 f"A/c {emp['account_no'] or '—'}  ·  IFSC {emp['ifsc'] or '—'}")
        for txt in (info1, info2):
            ws.merge_cells(start_row=row, start_column=1, end_row=row,
                           end_column=ncols)
            c = ws.cell(row, 1, txt)
            c.font = bold
            c.fill = hdr_fill
            c.border = border
            row += 1
        # month header
        ws.cell(row, 1, "Particulars").font = bold
        ws.cell(row, 1).border = border
        ws.cell(row, 1).fill = hdr_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        for j, mo in enumerate(months):
            c = ws.cell(row, 3 + j, mo["label"])
            c.font = bold
            c.alignment = ctr
            c.border = border
            c.fill = hdr_fill
        c = ws.cell(row, ncols, "TOTAL")
        c.font = bold
        c.alignment = ctr
        c.border = border
        c.fill = hdr_fill
        row += 1
        for h in heads:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            c = ws.cell(row, 1, h["label"])
            c.border = border
            if h["kind"] in ("total", "net"):
                c.font = bold
            for j, mo in enumerate(months):
                v = (emp["months"].get(mo["key"]) or {}).get(h["key"])
                cell = ws.cell(row, 3 + j, v if v else None)
                cell.border = border
                cell.alignment = rt
                cell.number_format = "#,##0.00" if h["key"] != "days" else "0.#"
                if h["key"] != "days" and any(
                        f in ("pf_mismatch", "esic_mismatch", "negative_net",
                              "gross_mismatch", "missing_attendance",
                              "loan_recovery_error")
                        for f in (emp["flags"].get(mo["key"]) or [])):
                    cell.fill = err_fill
            tv = emp["totals"].get(h["key"])
            cell = ws.cell(row, ncols, tv if tv else None)
            cell.border = border
            cell.alignment = rt
            cell.font = bold
            cell.number_format = "#,##0.00" if h["key"] != "days" else "0.#"
            row += 1
        # employee total line
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        t = emp["totals"]
        c = ws.cell(row, 1, f"EMPLOYEE TOTAL — Gross ₹{t['gross']:,.2f}   ·   "
                            f"Deduction ₹{t['total_ded']:,.2f}   ·   "
                            f"Net ₹{t['net']:,.2f}")
        c.font = bold
        c.fill = tot_fill
        c.border = border
        row += 2
    # grand totals
    g = data["grand"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, "GRAND TOTAL")
    c.font = Font(bold=True, size=12)
    c.fill = tot_fill
    row += 1
    pairs = [("Total Employees", data["total_employees"]),
             ("Total Days", g["days"]), ("Total Basic", g["basic"]),
             ("Total Gross", g["gross"]), ("Total Overtime", g["ot"]),
             ("Total PF", g["pf_ee"]), ("Total ESIC", g["esic_ee"]),
             ("Total PT", g["pt"]), ("Total LWF", g["lwf"]),
             ("Total TDS", g["tds"]), ("Total Advance", g["advance"]),
             ("Total Loan", g["loan"]), ("Total Other Deduction", g["other_ded"]),
             ("Grand Total Deduction", g["total_ded"]),
             ("Grand Net Salary", g["net"]), ("Employer PF", g["pf_er"]),
             ("Employer ESIC", g["esic_er"]), ("Total CTC", g["ctc"])]
    for lbl, val in pairs:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1, lbl).font = bold
        ws.cell(row, 1).border = border
        c = ws.cell(row, 3, val)
        c.border = border
        c.alignment = rt
        c.number_format = "#,##0.00"
        row += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    for j in range(len(months) + 1):
        ws.column_dimensions[get_column_letter(3 + j)].width = 11
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8  # A3
    ws.print_area = f"A1:{get_column_letter(ncols)}{row}"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF builder — A3 landscape register (ReportLab)
# ---------------------------------------------------------------------------

def _build_pdf(data: Dict[str, Any]) -> BytesIO:
    import base64
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Image, KeepTogether, Paragraph,
                                    SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A3), leftMargin=8 * mm, rightMargin=8 * mm,
        topMargin=8 * mm, bottomMargin=10 * mm,
        title="Payroll Register")
    W = landscape(A3)[0] - 16 * mm
    months = data["months"]
    heads = data["heads"]
    st_h = ParagraphStyle("h", fontSize=13, fontName="Helvetica-Bold",
                          alignment=1)
    st_sub = ParagraphStyle("s", fontSize=8.5, alignment=1)
    st_info = ParagraphStyle("i", fontSize=7.6, fontName="Helvetica-Bold")
    story: List[Any] = []
    logo_b64 = data.get("logo_base64")
    if logo_b64:
        try:
            raw = base64.b64decode(logo_b64.split(",")[-1])
            story.append(Image(BytesIO(raw), width=22 * mm, height=22 * mm))
        except Exception:  # noqa: BLE001
            pass
    story.append(Paragraph(
        f"{data['company_name']} — Employee-Wise Yearly Payroll Register "
        f"({data['fy_label']})", st_h))
    story.append(Paragraph(
        f"Print date: {data['print_date']} · Employees: "
        f"{data['total_employees']}", st_sub))
    story.append(Spacer(1, 4 * mm))
    col0 = 42 * mm
    colm = (W - col0 - 24 * mm) / len(months)
    widths = [col0] + [colm] * len(months) + [24 * mm]
    err = {"pf_mismatch", "esic_mismatch", "negative_net", "gross_mismatch",
           "missing_attendance", "loan_recovery_error"}
    for emp in data["rows"]:
        info1 = (f"Sr {emp['sr']} · Code {emp['employee_code'] or '—'} · "
                 f"<b>{emp['name']}</b> · F/H: {emp['father_name'] or '—'} · "
                 f"{emp['designation'] or '—'} · Dept {emp['department'] or '—'} · "
                 f"DOJ {emp['doj'] or '—'} · DOL {emp['dol'] or '—'}")
        info2 = (f"PF {emp['pf_no'] or '—'} · ESIC {emp['esic_no'] or '—'} · "
                 f"UAN {emp['uan'] or '—'} · Bank {emp['bank_name'] or '—'} · "
                 f"A/c {emp['account_no'] or '—'} · IFSC {emp['ifsc'] or '—'}")
        rows = [["Particulars"] + [m["label"] for m in months] + ["TOTAL"]]
        styles = [
            ("FONTSIZE", (0, 0), (-1, -1), 6.6),
            ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#9AA0A6")),
            ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#DDEBF7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (1, 0), (-1, 0), "CENTER"),
        ]
        for ri, h in enumerate(heads, start=1):
            line = [h["label"]]
            for mo in months:
                v = (emp["months"].get(mo["key"]) or {}).get(h["key"])
                line.append(f"{v:,.0f}" if v else "")
                if v and h["key"] != "days" and err & set(
                        emp["flags"].get(mo["key"]) or []):
                    styles.append(("BACKGROUND", (len(line) - 1, ri),
                                   (len(line) - 1, ri), rl.HexColor("#F8CBAD")))
            tv = emp["totals"].get(h["key"])
            line.append(f"{tv:,.0f}" if tv else "")
            if h["kind"] in ("total", "net"):
                styles.append(("FONTNAME", (0, ri), (-1, ri), "Helvetica-Bold"))
            rows.append(line)
        t = emp["totals"]
        rows.append([f"EMPLOYEE TOTAL — Gross {t['gross']:,.0f} · "
                     f"Deduction {t['total_ded']:,.0f} · Net {t['net']:,.0f}"]
                    + [""] * (len(months) + 1))
        styles += [
            ("SPAN", (0, len(rows) - 1), (-1, len(rows) - 1)),
            ("BACKGROUND", (0, len(rows) - 1), (-1, len(rows) - 1),
             rl.HexColor("#FFF2CC")),
            ("FONTNAME", (0, len(rows) - 1), (-1, len(rows) - 1),
             "Helvetica-Bold"),
        ]
        tbl = Table(rows, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle(styles))
        story.append(KeepTogether([
            Paragraph(info1, st_info), Paragraph(info2, st_info),
            Spacer(1, 1), tbl, Spacer(1, 4 * mm)]))
    # grand totals
    g = data["grand"]
    pairs = [("Total Employees", data["total_employees"]),
             ("Total Days", g["days"]), ("Total Basic", g["basic"]),
             ("Total Gross", g["gross"]), ("Total Overtime", g["ot"]),
             ("Total PF", g["pf_ee"]), ("Total ESIC", g["esic_ee"]),
             ("Total PT", g["pt"]), ("Total LWF", g["lwf"]),
             ("Total TDS", g["tds"]), ("Total Advance", g["advance"]),
             ("Total Loan", g["loan"]),
             ("Total Other Deduction", g["other_ded"]),
             ("Grand Total Deduction", g["total_ded"]),
             ("Grand Net Salary", g["net"]), ("Employer PF", g["pf_er"]),
             ("Employer ESIC", g["esic_er"]), ("Total CTC", g["ctc"])]
    grows = [["GRAND TOTAL", ""]] + [
        [lbl, f"{v:,.2f}" if isinstance(v, float) else str(v)]
        for lbl, v in pairs]
    gt = Table(grows, colWidths=[70 * mm, 50 * mm])
    gt.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#9AA0A6")),
        ("SPAN", (0, 0), (-1, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#FFF2CC")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
    ]))
    story.append(KeepTogether([gt]))

    def _page(cv, d):
        cv.setFont("Helvetica", 7)
        cv.drawRightString(landscape(A3)[0] - 8 * mm, 5 * mm,
                           f"Page {cv.getPageNumber()}")

    doc.build(story, onFirstPage=_page, onLaterPages=_page)
    buf.seek(0)
    return buf
