"""Iter 686 — CENTRAL STATISTICAL · Annual Labour Statistics (user spec).

Aggregation-ONLY layer: reuses Employee Master, attendance records and the
FINALIZED/latest compliance salary runs. NEVER recalculates payroll and
never touches any existing process.

  GET  /api/admin/central-stats/annual          (full FY aggregation JSON)
  GET  /api/admin/central-stats/annual.xlsx     (multi-sheet export)
  GET  /api/admin/central-stats/annual.pdf      (consolidated PDF)
  POST /api/admin/central-stats/finalize        (immutable snapshot)
  GET  /api/admin/central-stats/snapshots       (finalized report list)
"""
import io
import uuid
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query
from fastapi.responses import StreamingResponse

from server import db, get_user_from_token, require_role, now_iso  # noqa: E402
from routes.labour_statistics import _dt, _f, _users, _company, _run_rows  # noqa: E402

router = APIRouter(prefix="/api/admin/central-stats", tags=["central-stats"])

_MONTHS = ["April", "May", "June", "July", "August", "September", "October",
           "November", "December", "January", "February", "March"]
_LEFT = ("exited", "resigned", "terminated", "inactive", "left")


def _fy_months(fy_start: int) -> List[str]:
    return [f"{fy_start + (1 if m < 4 else 0):04d}-{m:02d}"
            for m in [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]]


def _g(u: dict) -> str:
    s = str(u.get("gender") or "").strip().lower()
    return "male" if s.startswith("m") else "female" if s.startswith("f") \
        else "other"


def _skill(u: dict) -> str:
    s = (str(u.get("skill_level") or u.get("category") or "")).strip().lower()
    if "high" in s:
        return "Highly Skilled"
    if "semi" in s:
        return "Semi-Skilled"
    if "unskil" in s or "un-skil" in s:
        return "Unskilled"
    if "skil" in s:
        return "Skilled"
    return "Unclassified"


def _etype(u: dict) -> str:
    s = (str(u.get("employee_type") or u.get("employee_group") or "")).lower()
    if "contract" in s:
        return "Contract"
    if "apprentice" in s:
        return "Apprentice"
    if "trainee" in s:
        return "Trainee"
    return "Permanent"


def _cat(u: dict) -> str:
    s = (str(u.get("employee_type") or u.get("employee_group") or "")).lower()
    return "Staff" if "staff" in s else "Worker"


def _desig_level(u: dict) -> str:
    s = str(u.get("designation") or "").lower()
    for k, lb in (("manager", "Manager"), ("engineer", "Engineer"),
                  ("executive", "Executive"), ("supervis", "Supervisor")):
        if k in s:
            return lb
    return "Other"


def _active_in_month(u: dict, mk: str) -> bool:
    doj = _dt(u.get("doj"))
    ex = _dt(u.get("exit_date") or u.get("resign_date"))
    ms = date(int(mk[:4]), int(mk[5:7]), 1)
    me = date(ms.year + (ms.month == 12), (ms.month % 12) + 1, 1)
    if doj and doj >= me:
        return False
    if ex and ex < ms:
        return False
    return True


async def _aggregate(cid: str, fy: int, flt: dict) -> dict:
    comp = await _company(cid)
    users = await db.users.find(
        {"company_id": cid, "role": "employee"}, {"_id": 0}).to_list(200000)
    # filters
    def _keep(u):
        if flt.get("department") and (u.get("department") or "") != flt["department"]:
            return False
        if flt.get("gender") and _g(u) != flt["gender"]:
            return False
        if flt.get("employment_type") and _etype(u) != flt["employment_type"]:
            return False
        if flt.get("category") and _cat(u) != flt["category"]:
            return False
        if flt.get("skill") and _skill(u) != flt["skill"]:
            return False
        return True
    users = [u for u in users if _keep(u)]
    uids = {u["user_id"] for u in users}
    months = _fy_months(fy)

    # ── salary rows (finalized/latest existing runs — NEVER recalculated)
    runs: Dict[str, Dict[str, dict]] = {}
    for mk in months:
        runs[mk] = await _run_rows(cid, mk)

    # ── attendance (existing records, monthly bulk)
    att: Dict[str, Dict[str, dict]] = {mk: {} for mk in months}
    fy_s, fy_e = f"{fy:04d}-04-01", f"{fy + 1:04d}-03-31"
    async for a in db.attendance.find(
            {"company_id": cid, "date": {"$gte": fy_s, "$lte": fy_e}},
            {"_id": 0, "user_id": 1, "date": 1}):
        mk = a["date"][:7]
        if mk in att and a.get("user_id") in uids:
            att[mk].setdefault(a["user_id"], {"days": set()})["days"].add(a["date"])

    # ── per-employee annual accumulation from salary rows
    emp_rows: List[dict] = []
    monthly: List[dict] = []
    tot = {k: 0.0 for k in ("mandays", "paid", "absent", "ot_hours",
                            "duty_hours", "gross", "ot_pay", "basic", "da",
                            "hra", "conv", "others", "pf_ee", "pf_er",
                            "esic_ee", "esic_er", "bonus", "gratuity")}
    emp_acc: Dict[str, dict] = {}
    for mk_i, mk in enumerate(months):
        rows = runs[mk]
        strength = sum(1 for u in users if _active_in_month(u, mk))
        join = sum(1 for u in users if (_dt(u.get("doj")) or date.min).strftime("%Y-%m") == mk)
        exit_ = sum(1 for u in users
                    if (_dt(u.get("exit_date") or u.get("resign_date")) or date.min).strftime("%Y-%m") == mk)
        md = paid = oth = m_gross = m_lc = 0.0
        for u in users:
            r = rows.get(u["user_id"]) or {}
            e = emp_acc.setdefault(u["user_id"], {k: 0.0 for k in tot})
            pd = _f(r.get("present_days"))
            md += pd
            paid += pd
            oth += _f(r.get("ot_hours"))
            _mg = _f(r.get("gross_paid")) or _f(r.get("monthly_gross"))
            m_gross += _mg
            m_lc += _mg + _f(r.get("pf_employer_total")) + _f(r.get("esic_employer"))
            for k, rk in (("mandays", "present_days"), ("paid", "present_days"),
                          ("ot_hours", "ot_hours"), ("duty_hours", "duty_hours"),
                          ("gross", "gross_paid"), ("ot_pay", "ot_pay"),
                          ("basic", "basic"), ("hra", "hra"),
                          ("conv", "conveyance"), ("others", "others"),
                          ("pf_ee", "pf_employee"), ("pf_er", "pf_employer_total"),
                          ("esic_ee", "esic_employee"), ("esic_er", "esic_employer")):
                v = _f(r.get(rk))
                if k == "gross" and not v:
                    v = _f(r.get("monthly_gross")) or (
                        _f(r.get("net")) + _f(r.get("total_deduction")))
                e[k] += v
                tot[k] += v
            md_days = _f(r.get("month_days")) or 0
            if md_days and pd < md_days:
                e["absent"] += md_days - pd
                tot["absent"] += md_days - pd
        wd = 26  # nominal working days for % (no second engine — display only)
        monthly.append({
            "month": _MONTHS[mk_i], "key": mk, "opening": strength,
            "joining": join, "exit": exit_,
            "closing": strength + join - exit_,
            "mandays": round(md, 1), "paid_days": round(paid, 1),
            "ot_hours": round(oth, 1),
            "gross": round(m_gross), "labour_cost": round(m_lc),
            "attendance_pct": round(md / (strength * wd) * 100, 1)
            if strength else 0.0,
        })
    for u in users:
        e = emp_acc.get(u["user_id"]) or {}
        emp_rows.append({
            "user_id": u["user_id"],
            "employee_code": u.get("employee_code") or "",
            "name": u.get("name") or "", "gender": _g(u).title(),
            "doj": str(u.get("doj") or "")[:10],
            "exit_date": str(u.get("exit_date") or u.get("resign_date") or "")[:10],
            "department": u.get("department") or "—",
            "designation": u.get("designation") or "—",
            "employment_type": _etype(u), "category": _cat(u),
            "skill": _skill(u),
            "mandays": round(e.get("mandays", 0), 1),
            "paid_days": round(e.get("paid", 0), 1),
            "absent": round(e.get("absent", 0), 1),
            "ot_hours": round(e.get("ot_hours", 0), 1),
            "basic": round(e.get("basic", 0)), "hra": round(e.get("hra", 0)),
            "other_allow": round(e.get("conv", 0) + e.get("others", 0)),
            "gross": round(e.get("gross", 0)),
            "ot_wages": round(e.get("ot_pay", 0)),
            "pf_ee": round(e.get("pf_ee", 0)), "pf_er": round(e.get("pf_er", 0)),
            "esic_ee": round(e.get("esic_ee", 0)),
            "esic_er": round(e.get("esic_er", 0)),
            "labour_cost": round(e.get("gross", 0) + e.get("pf_er", 0)
                                 + e.get("esic_er", 0)),
            "active": not (u.get("disabled") or str(
                u.get("employment_status") or "").lower() in _LEFT),
        })

    # ── employment / skill / category / department summaries
    def _sum_block(group_fn, keys):
        out = {k: {"male": 0, "female": 0, "other": 0, "total": 0} for k in keys}
        for u in users:
            k = group_fn(u)
            if k in out:
                out[k][_g(u)] += 1
                out[k]["total"] += 1
        return [{"particular": k, **v} for k, v in out.items()]

    employment = _sum_block(_etype, ["Permanent", "Contract", "Apprentice", "Trainee"]) \
        + _sum_block(_cat, ["Staff", "Worker"]) \
        + _sum_block(_desig_level, ["Supervisor", "Engineer", "Executive", "Manager"])
    g_tot = {"particular": "Total Employment",
             "male": sum(1 for u in users if _g(u) == "male"),
             "female": sum(1 for u in users if _g(u) == "female"),
             "other": sum(1 for u in users if _g(u) == "other"),
             "total": len(users)}
    employment.append(g_tot)
    skills = _sum_block(_skill, ["Highly Skilled", "Skilled", "Semi-Skilled",
                                 "Unskilled", "Unclassified"])

    dept: Dict[str, dict] = {}
    for r in emp_rows:
        d = dept.setdefault(r["department"], {
            "department": r["department"], "strength": 0, "male": 0,
            "female": 0, "mandays": 0.0, "ot_hours": 0.0, "gross": 0.0,
            "labour_cost": 0.0, "joining": 0, "exit": 0})
        d["strength"] += 1
        d["male"] += 1 if r["gender"] == "Male" else 0
        d["female"] += 1 if r["gender"] == "Female" else 0
        d["mandays"] += r["mandays"]
        d["ot_hours"] += r["ot_hours"]
        d["gross"] += r["gross"]
        d["labour_cost"] += r["labour_cost"]
        dj = r["doj"][:4]
        d["joining"] += 1 if dj == str(fy) or r["doj"][:7] in _fy_months(fy) else 0
        d["exit"] += 1 if r["exit_date"][:7] in _fy_months(fy) else 0
    departments = sorted(dept.values(), key=lambda x: x["department"])
    for d in departments:
        d["attrition_pct"] = round(d["exit"] / d["strength"] * 100, 1) if d["strength"] else 0
        for k in ("mandays", "ot_hours", "gross", "labour_cost"):
            d[k] = round(d[k], 1)

    cats = []
    for label, fn, keys in (
            ("Employment Type", _etype, ["Permanent", "Contract", "Apprentice", "Trainee"]),
            ("Employee Category", _cat, ["Staff", "Worker"]),
            ("Skill", _skill, ["Highly Skilled", "Skilled", "Semi-Skilled", "Unskilled", "Unclassified"]),
            ("Designation Level", _desig_level, ["Supervisor", "Engineer", "Executive", "Manager", "Other"])):
        for k in keys:
            grp = [r for u, r in zip(users, emp_rows) if fn(u) == k]
            if not grp:
                continue
            cats.append({"group": label, "category": k, "employees": len(grp),
                         "mandays": round(sum(r["mandays"] for r in grp), 1),
                         "ot_hours": round(sum(r["ot_hours"] for r in grp), 1),
                         "gross": round(sum(r["gross"] for r in grp)),
                         "labour_cost": round(sum(r["labour_cost"] for r in grp))})

    total_lc = tot["gross"] + tot["pf_er"] + tot["esic_er"] + tot["bonus"] + tot["gratuity"]
    avg_emp = round(sum(m["opening"] for m in monthly) / 12, 1)
    kpis = {
        "total_employment": len(users), "avg_employment": avg_emp,
        "total_mandays": round(tot["mandays"], 1),
        "avg_attendance_pct": round(sum(m["attendance_pct"] for m in monthly) / 12, 1),
        "total_gross": round(tot["gross"]),
        "total_ot_cost": round(tot["ot_pay"]),
        "total_labour_cost": round(total_lc),
        "avg_labour_cost_per_emp": round(total_lc / avg_emp) if avg_emp else 0,
        "joining": sum(m["joining"] for m in monthly),
        "exit": sum(m["exit"] for m in monthly),
        "attrition_pct": round(sum(m["exit"] for m in monthly) / avg_emp * 100, 1)
        if avg_emp else 0,
    }
    wage = {"basic": round(tot["basic"]), "da": round(tot["da"]),
            "hra": round(tot["hra"]), "conveyance": round(tot["conv"]),
            "other_allowances": round(tot["others"]),
            "gross": round(tot["gross"]), "ot_wages": round(tot["ot_pay"]),
            "bonus": round(tot["bonus"]), "employer_pf": round(tot["pf_er"]),
            "employer_esic": round(tot["esic_er"]),
            "gratuity": round(tot["gratuity"]),
            "total_labour_cost": round(total_lc)}
    statutory = [
        {"particular": "PF", "employee": round(tot["pf_ee"]), "employer": round(tot["pf_er"]),
         "annual": round(tot["pf_ee"] + tot["pf_er"])},
        {"particular": "ESIC", "employee": round(tot["esic_ee"]), "employer": round(tot["esic_er"]),
         "annual": round(tot["esic_ee"] + tot["esic_er"])},
        {"particular": "Bonus", "employee": 0, "employer": round(tot["bonus"]), "annual": round(tot["bonus"])},
        {"particular": "Gratuity", "employee": 0, "employer": round(tot["gratuity"]), "annual": round(tot["gratuity"])},
    ]

    # ── validation & data quality
    att_uids = {u for mk in months for u in att[mk]}
    sal_uids = {u for mk in months for u in runs[mk]}
    val = {"employee_master": len(users),
           "attendance": len(att_uids & uids),
           "salary": len(sal_uids & uids),
           "mismatch_attendance": sorted(
               [u.get("employee_code") or u.get("name") for u in users
                if u["user_id"] not in att_uids])[:50],
           "mismatch_salary": sorted(
               [u.get("employee_code") or u.get("name") for u in users
                if u["user_id"] not in sal_uids])[:50]}
    dq = []
    for lbl, fn in (("Missing department", lambda u: not u.get("department")),
                    ("Missing designation", lambda u: not u.get("designation")),
                    ("Missing gender", lambda u: not u.get("gender")),
                    ("Missing DOJ", lambda u: not u.get("doj")),
                    ("Missing salary (no run row all FY)",
                     lambda u: u["user_id"] not in sal_uids)):
        n = sum(1 for u in users if fn(u))
        if n:
            dq.append({"check": lbl, "count": n})
    val["data_quality"] = dq

    return {
        "company": {"name": comp.get("name"), "address": comp.get("address"),
                    "state": comp.get("state"), "code": comp.get("code") or comp.get("company_code"),
                    "company_id": cid},
        "fy": f"FY {fy}-{str(fy + 1)[2:]}", "fy_start_year": fy,
        "generated_at": now_iso(), "kpis": kpis,
        "employment_summary": employment, "skill_summary": skills,
        "attendance_summary": {
            "total_mandays": round(tot["mandays"], 1),
            "total_paid_days": round(tot["paid"], 1),
            "total_absent_days": round(tot["absent"], 1),
            "total_ot_hours": round(tot["ot_hours"], 1),
            "total_normal_hours": round(tot["duty_hours"], 1),
            "avg_attendance_pct": kpis["avg_attendance_pct"],
            "avg_mandays_per_emp": round(tot["mandays"] / len(users), 1) if users else 0},
        "monthly": monthly, "wage_summary": wage,
        "wage_monthly_avg": {k: round(v / 12) for k, v in wage.items()},
        "statutory_summary": statutory, "departments": departments,
        "employees": sorted(emp_rows, key=lambda r: r["employee_code"]),
        "categories": cats, "validation": val,
    }


async def _auth(authorization):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    return admin


def _flt(department, gender, employment_type, category, skill):
    return {"department": department or "", "gender": (gender or "").lower(),
            "employment_type": employment_type or "", "category": category or "",
            "skill": skill or ""}


# ─────────────── Official Formats (ASI-style mapping layer, Iter 687) ──────
# StatisticalReportDefinition: maps existing aggregation fields to official
# survey line items WITHOUT any new calculation engine. Built-in ASI-style
# template seeded read-only; custom formats can be added via POST /formats.

_BUILTIN_FORMATS = [{
    "definition_id": "asi_block_e",
    "name": "ASI-style — Block E: Employment & Labour Cost",
    "description": ("Annual Survey of Industries style employment/labour-cost "
                    "block mapped from existing payroll data. NOT an official "
                    "government return unless the exact notified format is "
                    "configured."),
    "builtin": True,
    "fields": [
        {"code": "E1", "label": "Average number of persons worked (male)",
         "source": "gender_male"},
        {"code": "E2", "label": "Average number of persons worked (female)",
         "source": "gender_female"},
        {"code": "E3", "label": "Total persons worked", "source": "kpis.total_employment"},
        {"code": "E4", "label": "Average annual employment", "source": "kpis.avg_employment"},
        {"code": "E5", "label": "Total man-days worked", "source": "kpis.total_mandays"},
        {"code": "E6", "label": "Wages / salaries (incl. allowances)", "source": "wage_summary.gross"},
        {"code": "E7", "label": "Overtime wages", "source": "wage_summary.ot_wages"},
        {"code": "E8", "label": "Bonus", "source": "wage_summary.bonus"},
        {"code": "E9", "label": "Employer contribution — Provident Fund",
         "source": "wage_summary.employer_pf"},
        {"code": "E10", "label": "Employer contribution — ESIC",
         "source": "wage_summary.employer_esic"},
        {"code": "E11", "label": "Gratuity provision", "source": "wage_summary.gratuity"},
        {"code": "E12", "label": "Total labour cost", "source": "wage_summary.total_labour_cost"},
        {"code": "E13", "label": "Attrition % (annual)", "source": "kpis.attrition_pct"},
    ],
}]


def _resolve_source(data: dict, src: str):
    if src == "gender_male":
        return next((r["male"] for r in data["employment_summary"]
                     if r["particular"] == "Total Employment"), 0)
    if src == "gender_female":
        return next((r["female"] for r in data["employment_summary"]
                     if r["particular"] == "Total Employment"), 0)
    cur: Any = data
    for part in src.split("."):
        cur = (cur or {}).get(part) if isinstance(cur, dict) else None
    return cur if cur is not None else 0


@router.get("/formats")
async def formats_list(authorization: Optional[str] = Header(None)):
    await _auth(authorization)
    custom = await db.statistical_report_definitions.find(
        {}, {"_id": 0}).to_list(100)
    return {"formats": _BUILTIN_FORMATS + custom}


@router.post("/formats")
async def formats_save(payload: Dict[str, Any] = Body(...),
                       authorization: Optional[str] = Header(None)):
    admin = await _auth(authorization)
    require_role(admin, ["super_admin", "sub_admin"])
    name = str(payload.get("name") or "").strip()
    fields = payload.get("fields") or []
    if not name or not isinstance(fields, list) or not fields:
        raise HTTPException(status_code=400, detail="name and fields[] required")
    doc = {"definition_id": payload.get("definition_id") or f"fmt_{uuid.uuid4().hex[:10]}",
           "name": name[:160], "description": str(payload.get("description") or "")[:500],
           "builtin": False,
           "fields": [{"code": str(f.get("code") or "")[:12],
                       "label": str(f.get("label") or "")[:200],
                       "source": str(f.get("source") or "")[:120]}
                      for f in fields][:60],
           "updated_at": now_iso()}
    await db.statistical_report_definitions.update_one(
        {"definition_id": doc["definition_id"]}, {"$set": doc}, upsert=True)
    return {"ok": True, "definition_id": doc["definition_id"]}


@router.get("/formats/{definition_id}/render")
async def formats_render(definition_id: str, company_id: str = Query(...),
                         fy_start_year: int = Query(...),
                         authorization: Optional[str] = Header(None)):
    await _auth(authorization)
    fmt = next((f for f in _BUILTIN_FORMATS
                if f["definition_id"] == definition_id), None) \
        or await db.statistical_report_definitions.find_one(
            {"definition_id": definition_id}, {"_id": 0})
    if not fmt:
        raise HTTPException(status_code=404, detail="Format not found")
    data = await _aggregate(company_id, fy_start_year, _flt("", "", "", "", ""))
    rows = [{"code": f["code"], "label": f["label"],
             "value": _resolve_source(data, f["source"])}
            for f in fmt["fields"]]
    return {"format": {"definition_id": fmt["definition_id"], "name": fmt["name"],
                       "description": fmt.get("description")},
            "company": data["company"], "fy": data["fy"], "rows": rows}


@router.get("/formats/{definition_id}/render.xlsx")
async def formats_render_xlsx(definition_id: str, company_id: str = Query(...),
                              fy_start_year: int = Query(...),
                              authorization: Optional[str] = Header(None),
                              token: Optional[str] = Query(None)):
    if token and not authorization:
        authorization = f"Bearer {token}"
    r = await formats_render(definition_id, company_id, fy_start_year, authorization)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Official Format"
    ws.append([r["format"]["name"]])
    ws.append([f"{r['company']['name']} · {r['fy']}"])
    ws.append([])
    ws.append(["Code", "Item", "Value"])
    for row in r["rows"]:
        ws.append([row["code"], row["label"], row["value"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename={definition_id}-{fy_start_year}.xlsx"})


@router.get("/annual")
async def annual(company_id: str = Query(...),
                 fy_start_year: int = Query(...),
                 department: str = Query(""), gender: str = Query(""),
                 employment_type: str = Query(""), category: str = Query(""),
                 skill: str = Query(""), compare_prev: bool = Query(False),
                 authorization: Optional[str] = Header(None)):
    admin = await _auth(authorization)
    if admin["role"] == "company_admin" and admin.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Not your firm")
    data = await _aggregate(company_id, fy_start_year,
                            _flt(department, gender, employment_type, category, skill))
    if compare_prev:
        prev = await _aggregate(company_id, fy_start_year - 1,
                                _flt(department, gender, employment_type, category, skill))
        cmp_rows = []
        for lbl, key in (("Average Employment", "avg_employment"),
                         ("Total Man-days", "total_mandays"),
                         ("Total Salary", "total_gross"),
                         ("Total OT Cost", "total_ot_cost"),
                         ("Total Labour Cost", "total_labour_cost"),
                         ("Attendance %", "avg_attendance_pct"),
                         ("Joining", "joining"), ("Exit", "exit"),
                         ("Attrition %", "attrition_pct")):
            c, p = data["kpis"][key], prev["kpis"][key]
            cmp_rows.append({"particular": lbl, "current": c, "previous": p,
                             "change_pct": round((c - p) / p * 100, 1) if p else None})
        data["prev_fy_comparison"] = cmp_rows
    return data


@router.get("/employee-detail")
async def employee_detail(company_id: str = Query(...),
                          fy_start_year: int = Query(...),
                          user_id: str = Query(...),
                          authorization: Optional[str] = Header(None)):
    """April→March monthly statistics for ONE employee (drill-down)."""
    await _auth(authorization)
    u = await db.users.find_one({"user_id": user_id, "company_id": company_id},
                                {"_id": 0})
    if not u:
        raise HTTPException(status_code=404, detail="Employee not found")
    rows = []
    tot = {k: 0.0 for k in ("days", "present", "paid", "absent", "ot_hours",
                            "gross", "ot_wages", "employer_cost")}
    for i, mk in enumerate(_fy_months(fy_start_year)):
        r = (await _run_rows(company_id, mk)).get(user_id) or {}
        er = _f(r.get("pf_employer_total")) + _f(r.get("esic_employer"))
        gross = _f(r.get("gross_paid")) or _f(r.get("monthly_gross"))
        row = {"month": _MONTHS[i], "key": mk,
               "days": _f(r.get("month_days")),
               "present": _f(r.get("present_days")),
               "paid": _f(r.get("present_days")),
               "absent": max(0.0, _f(r.get("month_days")) - _f(r.get("present_days"))),
               "ot_hours": _f(r.get("ot_hours")),
               "gross": gross,
               "ot_wages": _f(r.get("ot_pay")),
               "employer_cost": round(gross + er, 2)}
        for k in tot:
            tot[k] += row[k]
        rows.append(row)
    return {"employee": {"name": u.get("name"),
                         "employee_code": u.get("employee_code"),
                         "department": u.get("department"),
                         "designation": u.get("designation"),
                         "doj": str(u.get("doj") or "")[:10]},
            "rows": rows,
            "annual_total": {k: round(v, 1) for k, v in tot.items()}}


@router.post("/finalize")
async def finalize(payload: Dict[str, Any] = Body(...),
                   authorization: Optional[str] = Header(None)):
    admin = await _auth(authorization)
    cid = payload.get("company_id")
    fy = int(payload.get("fy_start_year") or 0)
    if not cid or not fy:
        raise HTTPException(status_code=400, detail="company_id and fy_start_year required")
    data = await _aggregate(cid, fy, _flt("", "", "", "", ""))
    ver = 1 + await db.central_stat_snapshots.count_documents(
        {"company_id": cid, "fy_start_year": fy})
    doc = {"snapshot_id": f"css_{uuid.uuid4().hex[:12]}", "company_id": cid,
           "fy_start_year": fy, "version": ver, "status": "FINALIZED",
           "generated_by": admin.get("name") or admin.get("email"),
           "finalized_at": now_iso(), "data": data}
    await db.central_stat_snapshots.insert_one(dict(doc))
    doc.pop("_id", None)
    doc.pop("data", None)
    return {"ok": True, "snapshot": doc}


@router.get("/snapshots")
async def snapshots(company_id: str = Query(...),
                    authorization: Optional[str] = Header(None)):
    await _auth(authorization)
    out = await db.central_stat_snapshots.find(
        {"company_id": company_id},
        {"_id": 0, "data": 0}).sort("finalized_at", -1).to_list(50)
    return {"snapshots": out}


@router.get("/snapshots/{snapshot_id}")
async def snapshot_detail(snapshot_id: str,
                          authorization: Optional[str] = Header(None)):
    await _auth(authorization)
    s = await db.central_stat_snapshots.find_one(
        {"snapshot_id": snapshot_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Snapshot not found")
    return s


@router.get("/annual.xlsx")
async def annual_xlsx(company_id: str = Query(...),
                      fy_start_year: int = Query(...),
                      authorization: Optional[str] = Header(None),
                      token: Optional[str] = Query(None)):
    if token and not authorization:
        authorization = f"Bearer {token}"
    await _auth(authorization)
    d = await _aggregate(company_id, fy_start_year, _flt("", "", "", "", ""))
    from openpyxl import Workbook
    wb = Workbook()

    def _sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for r in rows:
            ws.append(r)
    wb.remove(wb.active)
    k = d["kpis"]
    _sheet("Consolidated",
           ["Particular", "Value"],
           [["Company", d["company"]["name"]], ["Financial Year", d["fy"]],
            ["Generated", d["generated_at"][:19]]]
           + [[lbl, k[key]] for lbl, key in (
               ("Total Employment", "total_employment"),
               ("Average Employment", "avg_employment"),
               ("Total Man-days", "total_mandays"),
               ("Average Attendance %", "avg_attendance_pct"),
               ("Total Salary/Wages", "total_gross"),
               ("Total OT Cost", "total_ot_cost"),
               ("Total Labour Cost", "total_labour_cost"),
               ("Avg Labour Cost / Employee", "avg_labour_cost_per_emp"),
               ("Joining", "joining"), ("Exit", "exit"),
               ("Attrition %", "attrition_pct"))]
           + [["— Employment Summary —", ""]]
           + [[r["particular"], f"M {r['male']} · F {r['female']} · T {r['total']}"]
              for r in d["employment_summary"]]
           + [["— Wage Summary (Annual) —", ""]]
           + [[kk.replace("_", " ").title(), vv] for kk, vv in d["wage_summary"].items()])
    _sheet("Department-wise",
           ["Department", "Strength", "Male", "Female", "Man-days", "OT Hrs",
            "Salary Cost", "Labour Cost", "Joining", "Exit", "Attrition %"],
           [[x["department"], x["strength"], x["male"], x["female"],
             x["mandays"], x["ot_hours"], x["gross"], x["labour_cost"],
             x["joining"], x["exit"], x["attrition_pct"]] for x in d["departments"]])
    _sheet("Employee-wise",
           ["Code", "Name", "Gender", "DOJ", "Exit", "Department", "Designation",
            "Type", "Category", "Skill", "Man-days", "Paid", "Absent", "OT Hrs",
            "Basic", "HRA", "Other Allow", "Gross", "OT Wages", "PF EE", "PF ER",
            "ESIC EE", "ESIC ER", "Labour Cost"],
           [[e["employee_code"], e["name"], e["gender"], e["doj"], e["exit_date"],
             e["department"], e["designation"], e["employment_type"], e["category"],
             e["skill"], e["mandays"], e["paid_days"], e["absent"], e["ot_hours"],
             e["basic"], e["hra"], e["other_allow"], e["gross"], e["ot_wages"],
             e["pf_ee"], e["pf_er"], e["esic_ee"], e["esic_er"], e["labour_cost"]]
            for e in d["employees"]])
    _sheet("Category-wise",
           ["Group", "Category", "Employees", "Man-days", "OT Hrs", "Gross", "Labour Cost"],
           [[c["group"], c["category"], c["employees"], c["mandays"],
             c["ot_hours"], c["gross"], c["labour_cost"]] for c in d["categories"]])
    _sheet("Monthly Analysis",
           ["Month", "Opening", "Joining", "Exit", "Closing", "Man-days",
            "Paid Days", "OT Hours", "Attendance %"],
           [[m["month"], m["opening"], m["joining"], m["exit"], m["closing"],
             m["mandays"], m["paid_days"], m["ot_hours"], m["attendance_pct"]]
            for m in d["monthly"]])
    v = d["validation"]
    _sheet("Validation",
           ["Check", "Value"],
           [["Employee Master", v["employee_master"]],
            ["Attendance coverage", v["attendance"]],
            ["Salary coverage", v["salary"]],
            ["Missing attendance", ", ".join(map(str, v["mismatch_attendance"])) or "—"],
            ["Missing salary", ", ".join(map(str, v["mismatch_salary"])) or "—"]]
           + [[q["check"], q["count"]] for q in v["data_quality"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename=annual-labour-stats-{fy_start_year}.xlsx"})


@router.get("/annual.pdf")
async def annual_pdf(company_id: str = Query(...),
                     fy_start_year: int = Query(...),
                     authorization: Optional[str] = Header(None),
                     token: Optional[str] = Query(None)):
    if token and not authorization:
        authorization = f"Bearer {token}"
    await _auth(authorization)
    d = await _aggregate(company_id, fy_start_year, _flt("", "", "", "", ""))
    from utils.register_export import register_pdf
    cols = [{"key": "month", "label": "Month"}, {"key": "opening", "label": "Strength"},
            {"key": "joining", "label": "Joining"}, {"key": "exit", "label": "Exit"},
            {"key": "mandays", "label": "Man-days"}, {"key": "paid_days", "label": "Paid Days"},
            {"key": "ot_hours", "label": "OT Hrs"}, {"key": "attendance_pct", "label": "Att %"}]
    k = d["kpis"]
    sub = (f"{d['company']['name']} · {d['fy']} · Employment {k['total_employment']} · "
           f"Man-days {k['total_mandays']} · Salary ₹{k['total_gross']:,} · "
           f"Labour Cost ₹{k['total_labour_cost']:,} · Attrition {k['attrition_pct']}%")
    buf = register_pdf("Central Statistical — Annual Labour Statistics", sub,
                       cols, d["monthly"])
    return StreamingResponse(buf, media_type="application/pdf", headers={
        "Content-Disposition":
        f"attachment; filename=annual-labour-stats-{fy_start_year}.pdf"})
