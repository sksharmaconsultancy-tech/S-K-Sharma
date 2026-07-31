"""Iter 408 (user spec) — PF Contribution Type Reports (Higher PF / VPF).

ONE consolidated report engine covering the requested set via `view`:
  * all       — PF Contribution Type Report (every PF employee)
  * higher    — Higher PF Employees Report
  * vpf       — VPF Employees / VPF Summary Report
  * pending   — Higher PF Approval Pending Report
  * diff      — PF Difference Report (actual vs pure-statutory PF)

Source = the latest Compliance Salary run for the month (Monthly PF
Contribution Report), joined with the live Employee Master for the
approval/declaration workflow columns.

Endpoints (firm-scoped, super/sub/company admin):
  * GET /api/admin/reports/pf-contribution        — JSON
  * GET /api/admin/reports/pf-contribution.xlsx   — Excel
  * GET /api/admin/reports/pf-contribution.pdf    — A4 landscape PDF
"""
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException, Query
from fastapi.responses import Response

router = APIRouter(prefix="/api/admin/reports", tags=["pf-contribution"])

VIEWS = ("all", "higher", "vpf", "pending", "diff")

COLS = [
    ("employee_code", "Emp Code"), ("name", "Name"), ("uan_no", "UAN"),
    ("pf_contribution_type", "Type"), ("pf_approval_status", "Approval"),
    ("pf_declaration_available", "Declaration"),
    ("gross_paid", "Gross"), ("pf_wages", "PF Wages"),
    ("pf_ceiling_applied", "Ceiling?"), ("higher_pf_wage", "Higher Wage"),
    ("pf_employee", "PF (E)"), ("vpf_part", "VPF"),
    ("pf_employer_epf", "EPF (ER)"), ("pf_employer_eps", "EPS (ER)"),
    ("pf_employer_total", "ER Total"),
    ("statutory_pf", "Statutory PF"), ("pf_diff", "Diff vs Statutory"),
]


def _n(v: Any) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


async def _guard(authorization: Optional[str], company_id: str):
    from server import (get_user_from_token, require_role,
                        sub_admin_can_touch_company)
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    if admin.get("role") == "company_admin" and admin.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Firm not in your scope")
    if admin.get("role") == "sub_admin" and not sub_admin_can_touch_company(admin, company_id):
        raise HTTPException(status_code=403, detail="Firm not in your scope")
    return admin


async def _build(company_id: str, month: str, view: str) -> Dict[str, Any]:
    from server import db
    from utils.compliance_salary import DEFAULT_STATUTORY_CFG

    run = await db.compliance_salary_runs.find_one(
        {"company_id": company_id, "month": month},
        {"_id": 0, "rows": 1, "run_id": 1, "statutory_effective": 1,
         "statutory_cfg": 1, "finalized": 1, "generated_at": 1},
        sort=[("generated_at", -1)],
    )
    stat = (run or {}).get("statutory_effective") or (run or {}).get("statutory_cfg") or {}
    pf_cap = _n(stat.get("pf_wage_cap")) or _n(DEFAULT_STATUTORY_CFG["pf_wage_cap"])
    ee_pct = (_n(stat.get("pf_percent_employee"))
              or _n(DEFAULT_STATUTORY_CFG["pf_percent_employee"])) / 100.0

    masters: Dict[str, Dict[str, Any]] = {}
    rows_src = (run or {}).get("rows") or []
    uids = [r.get("user_id") for r in rows_src if r.get("user_id")]
    if uids:
        async for u in db.users.find(
            {"user_id": {"$in": uids}},
            {"_id": 0, "user_id": 1, "pf_contribution_type": 1,
             "pf_approval_status": 1, "pf_approval_required": 1,
             "pf_declaration_available": 1, "higher_pf_wage": 1,
             "vpf_percent": 1, "vpf_amount": 1, "vpf_enabled": 1,
             "pf_remarks": 1, "higher_pf_from": 1, "higher_pf_to": 1},
        ):
            masters[u["user_id"]] = u

    out: List[Dict[str, Any]] = []
    for r in rows_src:
        if _n(r.get("pf_employee")) <= 0 and _n(r.get("gross_paid")) <= 0:
            continue
        m = masters.get(r.get("user_id")) or {}
        ptype = str(r.get("pf_contribution_type")
                    or m.get("pf_contribution_type") or "statutory").lower()
        appr = str(m.get("pf_approval_status")
                   or r.get("pf_approval_status") or "").lower()
        pf_e = _n(r.get("pf_employee"))
        vpf = _n(r.get("vpf_amount"))
        # pure-statutory reference: 12% of capped wages (no VPF/higher)
        stat_pf = round(min(_n(r.get("pf_wages")), pf_cap) * ee_pct)
        if r.get("pf_higher_active"):
            stat_pf = round(min(_n(r.get("pf_wages")), pf_cap) * ee_pct)
        row = {
            "user_id": r.get("user_id"),
            "employee_code": r.get("employee_code"),
            "name": r.get("name"),
            "uan_no": r.get("uan_no"),
            "pf_contribution_type": ptype.upper() if ptype != "statutory" else "Statutory",
            "pf_approval_status": (appr or ("approved" if ptype != "higher" else "pending")).title(),
            "pf_declaration_available": "Yes" if (m.get("pf_declaration_available")
                                                  or r.get("pf_declaration_available")) else "No",
            "gross_paid": round(_n(r.get("gross_paid")), 2),
            "pf_wages": round(_n(r.get("pf_wages")), 2),
            "pf_ceiling_applied": "Yes" if r.get("pf_ceiling_applied") else "No",
            "higher_pf_wage": round(_n(m.get("higher_pf_wage")
                                       or r.get("higher_pf_wage")), 2),
            "pf_employee": round(pf_e, 2),
            "vpf_part": round(vpf, 2),
            "pf_employer_epf": round(_n(r.get("pf_employer_epf")), 2),
            "pf_employer_eps": round(_n(r.get("pf_employer_eps")), 2),
            "pf_employer_total": round(_n(r.get("pf_employer_total")), 2),
            "statutory_pf": stat_pf,
            "pf_diff": round(pf_e - stat_pf, 2),
            "pf_higher_active": bool(r.get("pf_higher_active")),
            "pf_higher_reason": r.get("pf_higher_reason") or "",
            "pf_reason": r.get("pf_reason") or "",
            "vpf_percent": _n(m.get("vpf_percent")),
            "pf_remarks": m.get("pf_remarks") or "",
        }
        if view == "higher" and ptype != "higher":
            continue
        if view == "vpf" and not (ptype == "vpf" or vpf > 0 or m.get("vpf_enabled")):
            continue
        if view == "pending" and not (ptype == "higher" and appr != "approved"):
            continue
        if view == "diff" and abs(row["pf_diff"]) < 0.5:
            continue
        out.append(row)
    out.sort(key=lambda x: str(x.get("employee_code") or ""))
    summary = {
        "employees": len(out),
        "statutory": sum(1 for r in out if r["pf_contribution_type"] == "Statutory"),
        "higher": sum(1 for r in out if r["pf_contribution_type"] == "HIGHER"),
        "vpf": sum(1 for r in out if r["pf_contribution_type"] == "VPF" or r["vpf_part"] > 0),
        "pending_approval": sum(1 for r in out
                                if r["pf_contribution_type"] == "HIGHER"
                                and r["pf_approval_status"] != "Approved"),
        "total_pf_employee": round(sum(r["pf_employee"] for r in out), 2),
        "total_vpf": round(sum(r["vpf_part"] for r in out), 2),
        "total_employer": round(sum(r["pf_employer_total"] for r in out), 2),
        "total_diff": round(sum(r["pf_diff"] for r in out), 2),
    }
    comp = await db.companies.find_one({"company_id": company_id},
                                       {"_id": 0, "name": 1})
    return {"month": month, "view": view, "run_id": (run or {}).get("run_id"),
            "run_locked": bool((run or {}).get("finalized")),
            "company_name": (comp or {}).get("name") or "",
            "rows": out, "summary": summary,
            "policy": {"allow_higher_pf": bool(stat.get("allow_higher_pf")),
                       "allow_vpf": stat.get("allow_vpf") is not False,
                       "vpf_max_percent": _n(stat.get("vpf_max_percent"))}}


@router.get("/pf-contribution")
async def pf_contribution_json(
    company_id: str = Query(...),
    month: str = Query(...),
    view: str = Query("all"),
    authorization: Optional[str] = Header(None),
):
    await _guard(authorization, company_id)
    if view not in VIEWS:
        raise HTTPException(status_code=400, detail=f"view must be one of {VIEWS}")
    return await _build(company_id, month, view)


@router.get("/pf-contribution.xlsx")
async def pf_contribution_xlsx(
    company_id: str = Query(...),
    month: str = Query(...),
    view: str = Query("all"),
    authorization: Optional[str] = Header(None),
):
    await _guard(authorization, company_id)
    data = await _build(company_id, month, view if view in VIEWS else "all")
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = f"PF {view}"[:28]
    ws.cell(row=1, column=1, value=f"{data['company_name']} — PF Contribution Report "
            f"({view.upper()}) — {month}").font = Font(bold=True, size=12)
    s = data["summary"]
    ws.cell(row=2, column=1, value=(
        f"Employees {s['employees']} · Statutory {s['statutory']} · Higher {s['higher']} · "
        f"VPF {s['vpf']} · Pending Approval {s['pending_approval']} · "
        f"PF(E) ₹{s['total_pf_employee']:,.0f} (incl. VPF ₹{s['total_vpf']:,.0f}) · "
        f"ER ₹{s['total_employer']:,.0f}")).font = Font(size=9)
    hr = 4
    for j, (_, lbl) in enumerate(COLS, start=1):
        c = ws.cell(row=hr, column=j, value=lbl)
        c.fill = PatternFill("solid", fgColor="1E3A8A")
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(data["rows"], start=hr + 1):
        for j, (k, _) in enumerate(COLS, start=1):
            ws.cell(row=i, column=j, value=r.get(k))
    tr = hr + len(data["rows"]) + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    for j, (k, _) in enumerate(COLS, start=1):
        if k in ("gross_paid", "pf_wages", "pf_employee", "vpf_part",
                 "pf_employer_epf", "pf_employer_eps", "pf_employer_total",
                 "statutory_pf", "pf_diff"):
            c = ws.cell(row=tr, column=j,
                        value=round(sum(_n(r.get(k)) for r in data["rows"]), 2))
            c.font = Font(bold=True)
    for j in range(1, len(COLS) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13 if j > 2 else 20
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="pf-contribution-{view}-{month}.xlsx"'})


@router.get("/pf-contribution.pdf")
async def pf_contribution_pdf(
    company_id: str = Query(...),
    month: str = Query(...),
    view: str = Query("all"),
    authorization: Optional[str] = Header(None),
):
    await _guard(authorization, company_id)
    data = await _build(company_id, month, view if view in VIEWS else "all")
    import io
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=8 * mm,
                            rightMargin=8 * mm, topMargin=8 * mm,
                            bottomMargin=8 * mm)
    h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=11)
    h2 = ParagraphStyle("h2", fontName="Helvetica", fontSize=8)
    s = data["summary"]
    flow = [
        Paragraph(f"{data['company_name']} — PF Contribution Report "
                  f"({view.upper()}) — {data['month']}", h1),
        Paragraph(
            f"Employees {s['employees']} | Statutory {s['statutory']} | "
            f"Higher {s['higher']} | VPF {s['vpf']} | Pending Approval "
            f"{s['pending_approval']} | PF(E) Rs.{s['total_pf_employee']:,.0f} "
            f"(incl. VPF Rs.{s['total_vpf']:,.0f}) | ER Rs.{s['total_employer']:,.0f}",
            h2),
        Spacer(1, 3 * mm),
    ]
    head = [lbl for _, lbl in COLS]
    body = [[r.get(k) for k, _ in COLS] for r in data["rows"]]
    _money = {"gross_paid", "pf_wages", "pf_employee", "vpf_part",
              "pf_employer_epf", "pf_employer_eps", "pf_employer_total",
              "statutory_pf", "pf_diff"}
    tot = ["TOTAL" if i == 0 else
           (round(sum(_n(r.get(k)) for r in data["rows"]), 0) if k in _money else "")
           for i, (k, _) in enumerate(COLS)]
    tbl = Table([head] + body + [tot], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#1E3A8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.2),
        ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#CBD5E1")),
        ("ALIGN", (6, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [rl.white, rl.HexColor("#F8FAFC")]),
        ("BACKGROUND", (0, -1), (-1, -1), rl.HexColor("#DDEBF7")),
    ]))
    flow.append(tbl)
    doc.build(flow)
    return Response(buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition":
                             f'inline; filename="pf-contribution-{view}-{month}.pdf"'})
