"""Iter 688 — MANUAL ATTENDANCE (Excel-style editable monthly sheet).

Additive OVERLAY on the existing attendance system — punch records are
NEVER modified or deleted. Manual marks live in db.manual_attendance;
approval requests in db.attendance_change_requests; firm-level settings
in firm_masters.manual_attendance.

  GET  /api/admin/manual-attendance/settings/{cid}       · POST same
  GET  /api/admin/manual-attendance/monthly              (grid data)
  POST /api/admin/manual-attendance/save                 (direct/approval)
  GET  /api/admin/manual-attendance/approvals            (pending list)
  POST /api/admin/manual-attendance/approvals/decide     (approve/reject)
  GET  /api/admin/manual-attendance/monthly.xlsx         (grid export)
"""
import io
import uuid
from calendar import monthrange
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query
from fastapi.responses import StreamingResponse

from server import db, get_user_from_token, require_role, now_iso  # noqa: E402

router = APIRouter(prefix="/api/admin/manual-attendance",
                   tags=["manual-attendance"])

CODES = ("P", "A", "L", "CL", "WO", "CO", "HD", "H", "EL")
_DEF = {"enabled": True, "approval_required": False, "require_reason": False,
        "maker_checker": True,
        # Iter 689 — Phase 2 approval matrix
        "approval_levels": 1,            # 1 or 2
        "level1_approver_id": "",        # empty = any admin
        "level2_approver_id": "",
        "dept_approvers": {},            # {department: user_id} overrides L1
        "rules": {"ANY": True}}          # per-change-type: "A>P": True …

RULE_KEYS = ("ANY", "A>P", "P>A", "A>L", "P>L", "P>HD", "A>HD", "WO>P")


async def _auth(authorization):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    return admin


def _mm(month: str) -> str:
    """Accept MM-YYYY or YYYY-MM → YYYY-MM."""
    s = (month or "").strip()
    if len(s) == 7 and s[2] == "-":
        return f"{s[3:]}-{s[:2]}"
    return s


async def _settings(cid: str) -> dict:
    fmdoc = await db.firm_masters.find_one(
        {"company_id": cid}, {"_id": 0, "manual_attendance": 1}) or {}
    return {**_DEF, **(fmdoc.get("manual_attendance") or {})}


@router.get("/settings/{cid}")
async def get_settings(cid: str, authorization: Optional[str] = Header(None)):
    await _auth(authorization)
    return await _settings(cid)


@router.post("/settings/{cid}")
async def save_settings(cid: str, payload: Dict[str, Any] = Body(...),
                        authorization: Optional[str] = Header(None)):
    admin = await _auth(authorization)
    require_role(admin, ["super_admin", "sub_admin"])
    doc = {k: bool(payload.get(k, _DEF[k]))
           for k in ("enabled", "approval_required", "require_reason",
                     "maker_checker")}
    doc["approval_levels"] = 2 if int(payload.get("approval_levels") or 1) == 2 else 1
    doc["level1_approver_id"] = str(payload.get("level1_approver_id") or "")[:64]
    doc["level2_approver_id"] = str(payload.get("level2_approver_id") or "")[:64]
    da = payload.get("dept_approvers")
    doc["dept_approvers"] = ({str(k)[:80]: str(v)[:64]
                              for k, v in da.items() if v}
                             if isinstance(da, dict) else {})
    rl = payload.get("rules")
    doc["rules"] = ({k: bool(rl.get(k)) for k in RULE_KEYS}
                    if isinstance(rl, dict) else {"ANY": True})
    await db.firm_masters.update_one(
        {"company_id": cid}, {"$set": {"manual_attendance": doc}}, upsert=True)
    return {"ok": True, **doc}


async def _grid(cid: str, month: str) -> dict:
    y, m = int(month[:4]), int(month[5:7])
    ndays = monthrange(y, m)[1]
    days = [f"{month}-{d:02d}" for d in range(1, ndays + 1)]
    comp = await db.companies.find_one(
        {"company_id": cid}, {"_id": 0, "attendance_policy": 1})
    firm_wo = {int(x) for x in ((comp or {}).get("attendance_policy")
                                or {}).get("weekly_off_days") or []}
    emps = await db.users.find(
        {"company_id": cid, "role": "employee", "disabled": {"$ne": True},
         # Iter 751 (user request) — sirf ACTIVE employees is report me.
         "active": {"$ne": False}},
        {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1,
         "department": 1, "employee_group": 1,
         "weekly_off_days_override": 1}).to_list(5000)
    emps.sort(key=lambda u: str(u.get("employee_code") or "").zfill(8))
    # base statuses from EXISTING punch data (read-only)
    punched: Dict[tuple, bool] = {}
    async for a in db.attendance.find(
            {"company_id": cid, "date": {"$gte": days[0], "$lte": days[-1]}},
            {"_id": 0, "user_id": 1, "date": 1}):
        punched[(a["user_id"], a["date"])] = True
    manual = {(x["user_id"], x["date"]): x async for x in
              db.manual_attendance.find(
                  {"company_id": cid,
                   "date": {"$gte": days[0], "$lte": days[-1]}}, {"_id": 0})}
    pending = {(x["user_id"], x["date"]): x async for x in
               db.attendance_change_requests.find(
                   {"company_id": cid, "status": "PENDING",
                    "date": {"$gte": days[0], "$lte": days[-1]}}, {"_id": 0})}
    # Iter 765 (user request) — ESIC Leave / Accident merge: approved ESIC
    # leave days (incl. accident-linked) auto-mark as "EL" and WIN over
    # punches and manual marks (user rule Q4a: EL jeete).
    from routes.esic_leave import esic_leave_dates_map as _el_map_fn
    try:
        el_days: Dict[str, set] = await _el_map_fn(cid, month)
    except Exception:
        el_days = {}
    today = date.today().isoformat()
    rows = []
    for u in emps:
        _ov = u.get("weekly_off_days_override")
        wo_days = ({int(x) for x in _ov} if isinstance(_ov, list) and _ov
                   else firm_wo)
        _uel = el_days.get(u["user_id"]) or ()
        cells, tot = {}, {c: 0 for c in CODES}
        for d in days:
            mk = manual.get((u["user_id"], d))
            if d in _uel:
                st, src = "EL", "esic"
            elif mk:
                st, src = mk["status"], "manual"
            elif punched.get((u["user_id"], d)):
                st, src = "P", "punch"
            elif date.fromisoformat(d).weekday() in wo_days:
                st, src = "WO", "policy"
            else:
                st, src = ("A", "auto") if d <= today else ("", "")
            pq = pending.get((u["user_id"], d))
            cells[d] = {"st": st, "src": src,
                        "pending": pq["requested_status"] if pq else None}
            if st in tot:
                tot[st] += 1
        rows.append({"user_id": u["user_id"],
                     "employee_code": u.get("employee_code") or "",
                     "name": u.get("name") or "",
                     "department": u.get("department") or "",
                     "group": u.get("employee_group") or "",
                     "cells": cells, "totals": tot})
    summary = {c: sum(r["totals"][c] for r in rows) for c in CODES}
    summary.update({
        "employees": len(rows),
        "manual": sum(1 for r in rows for c in r["cells"].values()
                      if c["src"] == "manual"),
        "pending": len(pending)})
    # Iter 762 (user request) — the firm's Holiday Master dates for this
    # month, so the grid can auto-mark the "H" columns in one click.
    holidays: Dict[str, str] = {}
    try:
        async for h in db.masters.find(
            {"type": "holiday", "date": {"$regex": f"^{month}"},
             "$or": [{"company_id": cid}, {"company_id": None},
                     {"company_id": {"$exists": False}}]},
            {"_id": 0, "date": 1, "name": 1},
        ):
            if h.get("date"):
                holidays[h["date"]] = h.get("name") or "Holiday"
    except Exception:
        holidays = {}
    return {"month": month, "days": days,
            "weekdays": [date.fromisoformat(d).strftime("%a") for d in days],
            "holidays": holidays,
            "rows": rows, "summary": summary}


@router.get("/monthly")
async def monthly(company_id: str = Query(...), month: str = Query(...),
                  authorization: Optional[str] = Header(None)):
    await _auth(authorization)
    g = await _grid(company_id, _mm(month))
    g["settings"] = await _settings(company_id)
    return g


@router.post("/save")
async def save(payload: Dict[str, Any] = Body(...),
               authorization: Optional[str] = Header(None)):
    admin = await _auth(authorization)
    cid = payload.get("company_id")
    changes = payload.get("changes") or []
    if not cid or not changes:
        raise HTTPException(status_code=400, detail="company_id and changes required")
    st = await _settings(cid)
    if not st["enabled"]:
        raise HTTPException(status_code=403,
                            detail="Manual attendance editing is disabled in Firm Master")
    if st["require_reason"] and any(not str(c.get("reason") or "").strip()
                                    for c in changes):
        raise HTTPException(status_code=400,
                            detail="Reason is required for every manual change (Firm Master rule)")
    who = admin.get("name") or admin.get("email")
    applied = queued = 0
    rules = st.get("rules") or {"ANY": True}
    dept_map = st.get("dept_approvers") or {}
    emp_depts = {u["user_id"]: (u.get("department") or "") async for u in
                 db.users.find({"company_id": cid, "role": "employee"},
                               {"_id": 0, "user_id": 1, "department": 1})}
    for c in changes:
        uid, d = c.get("user_id"), str(c.get("date") or "")[:10]
        new = str(c.get("status") or "").upper()
        if new not in CODES or not uid:
            continue
        prev = str(c.get("previous_status") or "").upper()
        # Iter 689 — per-change-type approval matrix (Firm Master rules):
        # approval needed only when ANY is on OR the exact transition is on.
        _needs = bool(st["approval_required"]) and bool(
            rules.get("ANY") or rules.get(f"{prev}>{new}"))
        # department-wise approver overrides level-1 (Firm Master)
        _appr1 = (dept_map.get(emp_depts.get(uid, ""))
                  or st.get("level1_approver_id") or "")
        audit = {"company_id": cid, "user_id": uid, "date": d,
                 "previous_status": prev or None,
                 "requested_status": new,
                 "reason": str(c.get("reason") or "")[:300],
                 "requested_by": who, "requested_by_id": admin.get("user_id"),
                 "requested_at": now_iso(), "source": "manual"}
        if _needs:
            await db.attendance_change_requests.update_one(
                {"company_id": cid, "user_id": uid, "date": d,
                 "status": "PENDING"},
                {"$set": {**audit, "status": "PENDING", "level": 1,
                          "approver_l1": _appr1,
                          "approver_l2": st.get("level2_approver_id") or "",
                          "levels": int(st.get("approval_levels") or 1),
                          "request_id": f"acr_{uuid.uuid4().hex[:10]}"}},
                upsert=True)
            queued += 1
        else:
            await db.manual_attendance.update_one(
                {"company_id": cid, "user_id": uid, "date": d},
                {"$set": {"status": new, "updated_by": who,
                          "updated_at": now_iso(), "source": "manual"}},
                upsert=True)
            await db.attendance_change_audit.insert_one(
                {**audit, "status": "APPLIED"})
            applied += 1
    if queued:
        await db.notifications.insert_one({
            "notification_id": f"ntf_{uuid.uuid4().hex[:10]}",
            "company_id": cid, "audience": "admins",
            "title": "New Attendance Change Request",
            "body": f"{queued} attendance change(s) submitted by {who} — pending approval.",
            "category": "attendance", "created_at": now_iso(), "read_by": []})
    return {"ok": True, "applied": applied, "pending": queued,
            "approval_required": st["approval_required"]}


@router.get("/approvals")
async def approvals(company_id: str = Query(...),
                    authorization: Optional[str] = Header(None)):
    await _auth(authorization)
    reqs = await db.attendance_change_requests.find(
        {"company_id": company_id, "status": "PENDING"},
        {"_id": 0}).sort("requested_at", -1).to_list(500)
    names = {u["user_id"]: u async for u in db.users.find(
        {"company_id": company_id, "role": "employee"},
        {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1,
         "department": 1})}
    for r in reqs:
        u = names.get(r["user_id"]) or {}
        r.update(name=u.get("name"), employee_code=u.get("employee_code"),
                 department=u.get("department"))
    return {"requests": reqs}


async def _apply_punch_request(r: dict, who: str, admin: dict):
    """Iter 710 — apply an APPROVED punch-repair request to db.attendance."""
    pa = r.get("punch_action")
    note = f"Approved punch repair by {who}: {r.get('reason') or ''}"[:200]
    if pa == "punch_add":
        pl = r.get("punch_payload") or {}
        await db.attendance.insert_one({
            "record_id": f"att_{uuid.uuid4().hex[:12]}",
            "user_id": r["user_id"], "company_id": r["company_id"],
            "date": r.get("date"), "kind": pl.get("kind") or "in",
            "at": pl.get("at"), "source": "manual_admin",
            "status": "approved", "approved_by": admin.get("user_id"),
            "manual_reason": note, "created_at": now_iso()})
    elif pa == "punch_edit":
        pl = r.get("punch_payload") or {}
        upd = {k: v for k, v in (("at", pl.get("at")), ("kind", pl.get("kind"))) if v}
        upd.update({"edited_by": admin.get("user_id"), "edited_at": now_iso(),
                    "edit_reason": note, "status": "approved"})
        await db.attendance.update_one(
            {"record_id": r.get("punch_record_id")}, {"$set": upd})
    elif pa == "punch_delete":
        await db.attendance.delete_one({"record_id": r.get("punch_record_id")})
    try:
        from server import invalidate_grid_cache
        invalidate_grid_cache(r["company_id"])
    except Exception:
        pass



@router.post("/approvals/decide")
async def decide(payload: Dict[str, Any] = Body(...),
                 authorization: Optional[str] = Header(None)):
    admin = await _auth(authorization)
    cid = payload.get("company_id")
    ids = payload.get("request_ids") or []
    action = str(payload.get("action") or "").upper()
    reason = str(payload.get("reason") or "")[:300]
    if action not in ("APPROVE", "REJECT") or not ids:
        raise HTTPException(status_code=400, detail="action and request_ids required")
    st = await _settings(cid)
    who = admin.get("name") or admin.get("email")
    done = 0
    for rid in ids[:500]:
        r = await db.attendance_change_requests.find_one(
            {"request_id": rid, "company_id": cid, "status": "PENDING"})
        if not r:
            continue
        # Maker-checker: maker cannot approve own request (Firm Master rule)
        if st["maker_checker"] and r.get("requested_by_id") == admin.get("user_id"):
            raise HTTPException(status_code=403,
                                detail="Maker cannot approve own request")
        # Iter 689 — designated approver enforcement (dept/level approver);
        # super admin can always decide.
        lvl = int(r.get("level") or 1)
        _designated = (r.get("approver_l1") if lvl == 1
                       else r.get("approver_l2")) or ""
        if _designated and admin.get("role") != "super_admin" \
                and admin.get("user_id") != _designated:
            raise HTTPException(
                status_code=403,
                detail=f"Only the designated Level-{lvl} approver can decide this request")
        # Iter 689 — multi-level: L1 approval escalates to Level 2.
        if action == "APPROVE" and int(r.get("levels") or 1) == 2 and lvl == 1:
            await db.attendance_change_requests.update_one(
                {"request_id": rid},
                {"$set": {"level": 2, "l1_approved_by": who,
                          "l1_approved_at": now_iso()}})
            await db.attendance_change_audit.insert_one(
                {k: v for k, v in r.items() if k != "_id"}
                | {"status": "L1_APPROVED", "decided_by": who,
                   "decided_at": now_iso()})
            done += 1
            continue
        upd = {"status": "APPROVED" if action == "APPROVE" else "REJECTED",
               "decided_by": who, "decided_by_id": admin.get("user_id"),
               "decided_at": now_iso(), "rejection_reason": reason or None}
        await db.attendance_change_requests.update_one(
            {"request_id": rid}, {"$set": upd})
        if action == "APPROVE":
            # Iter 710 — punch-repair requests apply to db.attendance.
            if r.get("punch_action"):
                await _apply_punch_request(r, who, admin)
            else:
                await db.manual_attendance.update_one(
                    {"company_id": cid, "user_id": r["user_id"], "date": r["date"]},
                    {"$set": {"status": r["requested_status"], "updated_by": who,
                              "updated_at": now_iso(), "source": "manual",
                              "approved": True}}, upsert=True)
        await db.attendance_change_audit.insert_one(
            {k: v for k, v in {**r, **upd}.items() if k != "_id"})
        done += 1
    return {"ok": True, "decided": done}


@router.get("/approver-options")
async def approver_options(company_id: str = Query(...),
                           authorization: Optional[str] = Header(None)):
    """Iter 689 — admins selectable as approvers + firm departments."""
    await _auth(authorization)
    admins = await db.users.find(
        {"$or": [{"role": {"$in": ["super_admin", "sub_admin"]}},
                 {"role": "company_admin", "company_id": company_id}]},
        {"_id": 0, "user_id": 1, "name": 1, "role": 1}).to_list(100)
    depts = await db.users.distinct(
        "department", {"company_id": company_id, "role": "employee",
                       "department": {"$nin": [None, ""]}})
    return {"approvers": admins, "departments": sorted(depts)}


@router.get("/monthly.xlsx")
async def monthly_xlsx(company_id: str = Query(...), month: str = Query(...),
                       authorization: Optional[str] = Header(None),
                       token: Optional[str] = Query(None)):
    if token and not authorization:
        authorization = f"Bearer {token}"
    await _auth(authorization)
    g = await _grid(company_id, _mm(month))
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    hdr = (["Code", "Employee Name", "Department"]
           + [d[8:10] for d in g["days"]] + list(CODES))
    ws.append(hdr)
    ws.append(["", "", ""] + g["weekdays"] + [""] * len(CODES))
    for i in (1, 2):
        for c in ws[i]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
    for r in g["rows"]:
        ws.append([r["employee_code"], r["name"], r["department"]]
                  + [r["cells"][d]["st"] for d in g["days"]]
                  + [r["totals"][c] for c in CODES])
    for row in ws.iter_rows(min_row=3):
        for c in row[3:]:
            c.alignment = Alignment(horizontal="center")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename=attendance-{_mm(month)}.xlsx"})
