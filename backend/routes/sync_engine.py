"""
Real-Time ZKTeco Multi-Device Synchronization Engine — Phase 1.

Modular, queue-based service that keeps employee records synchronised across
every ZKTeco machine of a company over the ADMS push protocol (the machines
poll the server; we queue device commands that they execute on their next
contact — so offline machines auto-catch-up the moment they reconnect).

Layers
------
* Repository:  MongoDB collections (``sync_settings``, ``sync_jobs``,
               ``sync_log``, ``sync_conflicts``) + the existing
               ``biometric_devices`` / ``biometric_device_cmds`` / ``users``.
* Service:     :func:`enqueue_employee_sync`, :func:`process_sync_queue`,
               :func:`get_sync_settings`.
* Worker:      :func:`sync_engine_loop` (runs every 30 s, drains the queue,
               retries failed jobs, reconciles device-command results).
* API:         REST endpoints (``/api/sync/*``) consumed by the dashboard.

Design notes
------------
* Auto-sync fires on employee Create / Update / Delete / Transfer / Disable
  from ``server.py`` — always wrapped so a sync hiccup never breaks the
  employee operation itself.
* Conflict rule (Phase 1): the portal is the source of truth. A conflict is
  logged only when a machine reports biometric templates for an employee the
  portal has none for (so on-device enrolments are never silently lost).
"""

import asyncio
import logging
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query

from server import db, get_user_from_token, require_role
from routes.biometric_devices import _queue_cmd, _template_to_cmd

logger = logging.getLogger("sync_engine")
router = APIRouter(prefix="/api", tags=["sync-engine"])

# Terminal states of an individual device command (biometric_device_cmds).
_CMD_TERMINAL = {"done", "failed"}

# ---------------------------------------------------------------------------
# Settings (Sync_Settings)
# ---------------------------------------------------------------------------
# Iter 584 (user spec) — FINAL SYNC RULES. Only three flows are legal:
#   1. MACHINE  → PAYROLL   (punch / ATTLOG sync)
#   2. MACHINE  → MACHINE   (device-to-device sync)
#   3. PORTAL   → DEVICE    (MANUAL employee registration / deletion ONLY)
# Automatic Employee Master → Machine sync is PERMANENTLY DISABLED/LOCKED.
SYNC_DEFAULTS: Dict[str, Any] = {
    "enable_auto_sync": False,  # LOCKED — not configurable (Iter 584)
    "machine_to_payroll_punch_sync": True,
    "machine_to_machine_sync": True,
    "manual_employee_registration": False,
    "manual_employee_delete": False,
    "sync_fingerprints": True,
    "sync_face": True,
    "sync_card": True,
    "sync_password": True,
    "sync_photos": False,
    # Iter 419 (user rule): attendance is NEVER synced between machines —
    # it only flows FROM each machine INTO the portal (ATTLOG push).
    "retry_failed": True,
    "max_retry_count": 3,
    "sync_interval": 30,  # seconds (worker cadence)
    # Iter 768 (user request) — AUTO Machine → Machine sync: when a machine
    # reports a NEW or CHANGED user, a debounced harvest + distribute run
    # starts automatically and every run is logged (machine_sync_logs).
    "machine_auto_sync": True,
}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


async def get_sync_settings(company_id: str) -> Dict[str, Any]:
    """Company sync settings, merged over defaults."""
    doc = await db.sync_settings.find_one({"company_id": company_id}, {"_id": 0}) or {}
    merged = dict(SYNC_DEFAULTS)
    merged.update({k: v for k, v in doc.items() if k in SYNC_DEFAULTS})
    # Iter 584 — automatic Employee Master → Machine sync is LOCKED OFF for
    # every firm regardless of what an old settings document says.
    merged["enable_auto_sync"] = False
    merged["auto_master_sync"] = "DISABLED"
    merged["company_id"] = company_id
    return merged


async def _sync_enabled_devices(company_id: str) -> List[dict]:
    """Registered, enabled, sync-enabled machines of the company."""
    return await db.biometric_devices.find(
        {"company_id": company_id,
         "enabled": {"$ne": False},
         "sync_enabled": {"$ne": False}},
        {"_id": 0, "serial_number": 1, "name": 1, "last_source_ip": 1},
    ).to_list(200)


# ---------------------------------------------------------------------------
# Command building (Service layer)
# ---------------------------------------------------------------------------
async def _build_commands(
    emp: dict, action: str, settings: dict,
) -> List[Dict[str, str]]:
    """Return the ordered list of ADMS commands for one employee-change.

    Each item is ``{"command": ..., "label": ...}``. USERINFO first (so the
    templates land on an existing user), then fingerprint / face templates
    when enabled by settings.
    """
    pin = str(emp.get("bio_code") or "").strip()
    if not pin:
        return []
    name = (emp.get("name") or "")[:24].replace("\t", " ")

    # Removal actions — take the user OFF the machine so they can't punch.
    if action in ("delete", "disable"):
        return [{
            "command": f"DATA DELETE USERINFO PIN={pin}",
            "label": f"Remove {name or pin} from device ({action})",
        }]

    # Create / update / transfer / enable — (re)push the user record.
    parts = [f"DATA UPDATE USERINFO PIN={pin}", f"Name={name}",
             f"Pri={int(emp.get('privilege') or 0)}"]
    if settings.get("sync_password") and (emp.get("punch_password") or emp.get("device_password")):
        parts.append(f"Passwd={emp.get('punch_password') or emp.get('device_password')}")
    if settings.get("sync_card") and (emp.get("card_no") or emp.get("card_number")):
        parts.append(f"Card={emp.get('card_no') or emp.get('card_number')}")
    cmds: List[Dict[str, str]] = [{
        "command": "\t".join(parts),
        "label": f"Sync user {name or pin}",
    }]

    # Biometric templates (from whatever we've captured for this PIN).
    want_fp = settings.get("sync_fingerprints")
    want_face = settings.get("sync_face")
    if want_fp or want_face:
        kinds = []
        if want_fp:
            kinds.append("fp")
        if want_face:
            kinds.append("face")
        templates = await db.biometric_templates.find(
            {"company_id": emp.get("company_id"), "pin": pin,
             "kind": {"$in": kinds}}, {"_id": 0},
        ).to_list(50)
        for t in templates:
            cmds.append({
                "command": _template_to_cmd(t),
                "label": f"Sync {t.get('kind')} template PIN {pin}",
            })
    return cmds


# ---------------------------------------------------------------------------
# Enqueue (called by server.py employee endpoints + manual sync APIs)
# ---------------------------------------------------------------------------
# Iter 584 — the ONLY sync types allowed to travel PORTAL → DEVICE.
_ALLOWED_PORTAL_SYNC = {"MANUAL_EMPLOYEE_REGISTRATION", "MANUAL_EMPLOYEE_DELETE"}
# Every other master-data type is rejected at the SERVICE layer.
BLOCKED_SYNC_TYPES = {
    "EMPLOYEE_MASTER_AUTO_SYNC", "EMPLOYEE_MASTER_BULK_SYNC",
    "AUTOMATIC_EMPLOYEE_CREATE", "AUTOMATIC_EMPLOYEE_UPDATE",
    "AUTOMATIC_EMPLOYEE_DELETE", "AUTOMATIC_EMPLOYEE_TRANSFER",
}


async def enqueue_employee_sync(
    company_id: str,
    user_id: str,
    action: str,
    actor: str = "system",
    force: bool = False,
    sync_type: str = "EMPLOYEE_MASTER_AUTO_SYNC",
    target_serials: Optional[List[str]] = None,
    send_fields: Optional[List[str]] = None,
) -> Optional[str]:
    """Queue a PORTAL → DEVICE job for one employee.

    Iter 584 — SERVICE-LAYER GUARD: only the two MANUAL sync types are
    permitted. Every automatic Employee-Master trigger (create / update /
    delete / transfer / …, still wired in employees_admin.py etc.) lands
    here with the default sync_type and is silently REJECTED
    (MASTER_DATA_DEVICE_SYNC_DISABLED) — the employee operation itself is
    never affected.
    """
    try:
        if sync_type not in _ALLOWED_PORTAL_SYNC:
            logger.info(
                "[sync] BLOCKED %s (%s) for %s/%s — "
                "MASTER_DATA_DEVICE_SYNC_DISABLED",
                sync_type, action, company_id, user_id)
            return None
        settings = await get_sync_settings(company_id)
        devices = await _sync_enabled_devices(company_id)
        if target_serials:
            want = {str(s) for s in target_serials}
            devices = [d for d in devices if d["serial_number"] in want]
        if not devices:
            return None
        emp = await db.users.find_one({"user_id": user_id}, {"_id": 0})
        if not emp:
            return None
        pin = str(emp.get("bio_code") or "").strip()
        if not pin:
            # Nothing to push without a machine punch number (Bio Code).
            return None
        job_id = f"sj_{uuid.uuid4().hex[:12]}"
        await db.sync_jobs.insert_one({
            "job_id": job_id,
            "company_id": company_id,
            "user_id": user_id,
            "pin": pin,
            "name": emp.get("name"),
            "action": action,
            "status": "pending",
            "attempts": 0,
            "max_attempts": int(settings.get("max_retry_count") or 3),
            "targets": [d["serial_number"] for d in devices],
            "cmd_ids": [],
            "source_type": "PORTAL",
            "destination_type": "DEVICE",
            "sync_type": sync_type,
            "send_fields": send_fields,
            "created_by": actor,
            "created_at": _now(),
            "updated_at": _now(),
            "error": None,
        })
        logger.info("[sync] enqueued %s (%s) job=%s pin=%s -> %d device(s)",
                    action, sync_type, job_id, pin, len(devices))
        return job_id
    except Exception:
        logger.warning("[sync] enqueue failed for %s/%s", company_id, user_id,
                       exc_info=True)
        return None


async def enqueue_employee_removal(
    company_id: str,
    user_id: str,
    pin: str,
    name: Optional[str] = None,
    actor: str = "system",
) -> Optional[str]:
    """Iter 584 — BLOCKED. This was the automatic machine-delete fired when
    an employee was removed from the payroll. Per the final sync rules an
    employee deletion / deactivation / transfer must NEVER touch the
    machines automatically — HR must use the manual
    'Delete Employee from Machine' action instead."""
    logger.info("[sync] BLOCKED auto machine-delete pin=%s (%s) — "
                "MASTER_DATA_DEVICE_SYNC_DISABLED", pin, company_id)
    return None
    # --- unreachable legacy body kept for reference ---
    try:
        settings = await get_sync_settings(company_id)
        if not settings.get("enable_auto_sync"):
            return None
        devices = await _sync_enabled_devices(company_id)
        pin = str(pin or "").strip()
        if not devices or not pin:
            return None
        job_id = f"sj_{uuid.uuid4().hex[:12]}"
        await db.sync_jobs.insert_one({
            "job_id": job_id,
            "company_id": company_id,
            "user_id": user_id,
            "pin": pin,
            "name": name,
            "action": "delete",
            "status": "pending",
            "attempts": 0,
            "max_attempts": int(settings.get("max_retry_count") or 3),
            "targets": [d["serial_number"] for d in devices],
            "cmd_ids": [],
            "created_by": actor,
            "created_at": _now(),
            "updated_at": _now(),
            "error": None,
        })
        logger.info("[sync] enqueued delete job=%s pin=%s -> %d device(s)",
                    job_id, pin, len(devices))
        return job_id
    except Exception:
        logger.warning("[sync] removal enqueue failed for %s", company_id,
                       exc_info=True)
        return None


# ---------------------------------------------------------------------------
# Worker (Queue drain + retry + reconcile)
# ---------------------------------------------------------------------------
async def _dispatch_job(job: dict) -> None:
    """Queue this job's device commands and open Sync_Log rows."""
    settings = await get_sync_settings(job["company_id"])
    emp = await db.users.find_one({"user_id": job["user_id"]}, {"_id": 0}) or {
        "bio_code": job.get("pin"), "name": job.get("name"),
        "company_id": job["company_id"],
    }
    # Iter 584 — manual registration sends ONLY the fields the admin picked.
    if job.get("send_fields") is not None:
        f = set(job["send_fields"] or [])
        settings = {**settings,
                    "sync_password": "password" in f,
                    "sync_card": "card" in f,
                    "sync_fingerprints": "fingerprint" in f,
                    "sync_face": "face" in f}
    cmds = await _build_commands(emp, job["action"], settings)
    if not cmds and job["action"] not in ("delete", "disable"):
        await db.sync_jobs.update_one(
            {"job_id": job["job_id"]},
            {"$set": {"status": "success", "updated_at": _now(),
                      "note": "nothing to sync (no PIN/templates)"}},
        )
        return
    cmd_ids: List[str] = []
    for serial in job["targets"]:
        for c in cmds:
            started = datetime.now(timezone.utc)
            cid = await _queue_cmd(serial, c["command"], job["created_by"], c["label"])
            cmd_ids.append(cid)
            await db.sync_log.insert_one({
                "log_id": f"sl_{uuid.uuid4().hex[:12]}",
                "job_id": job["job_id"],
                "company_id": job["company_id"],
                "user_id": job["user_id"],
                "pin": job.get("pin"),
                "device_serial": serial,
                "action": job["action"],
                "command": c["label"],
                "cmd_id": cid,
                "status": "queued",
                "response": None,
                "error": None,
                "exec_ms": int((datetime.now(timezone.utc) - started).total_seconds() * 1000),
                "ip": None,
                "created_at": _now(),
            })
    await db.sync_jobs.update_one(
        {"job_id": job["job_id"]},
        {"$set": {"status": "processing", "cmd_ids": cmd_ids,
                  "updated_at": _now()},
         "$inc": {"attempts": 1}},
    )


async def _reconcile_job(job: dict) -> None:
    """Turn a processing job into success / retry / failed by looking at the
    outcome of its queued device commands."""
    cmd_ids = job.get("cmd_ids") or []
    if not cmd_ids:
        return
    cmds = await db.biometric_device_cmds.find(
        {"cmd_id": {"$in": cmd_ids}}, {"_id": 0, "cmd_id": 1, "status": 1,
                                       "device_serial": 1, "result_return": 1},
    ).to_list(len(cmd_ids) + 5)
    by_id = {c["cmd_id"]: c for c in cmds}
    statuses = [by_id.get(cid, {}).get("status", "pending") for cid in cmd_ids]
    done = all(s in _CMD_TERMINAL for s in statuses)
    any_failed = any(s == "failed" for s in statuses)

    # Reflect per-device outcomes into the log rows.
    for c in cmds:
        if c["status"] in _CMD_TERMINAL:
            await db.sync_log.update_one(
                {"cmd_id": c["cmd_id"]},
                {"$set": {
                    "status": "success" if c["status"] == "done" else "failed",
                    "response": c.get("result_return"),
                    "error": None if c["status"] == "done" else f"device return {c.get('result_return')}",
                }},
            )

    if not done:
        return  # still waiting on the machine(s)

    settings = await get_sync_settings(job["company_id"])
    if not any_failed:
        await db.sync_jobs.update_one(
            {"job_id": job["job_id"]},
            {"$set": {"status": "success", "updated_at": _now(),
                      "synced_at": _now(), "error": None}},
        )
        # Iter 542 (user request) — stamp the employee so the Employee
        # Master list can show a "Deleted from machine" tag (a later
        # add/update re-sync clears the stamp again).
        if job.get("user_id"):
            if job.get("action") == "delete":
                await db.users.update_one(
                    {"user_id": job["user_id"]},
                    {"$set": {"machine_deleted_at": _now()}})
            elif job.get("action") in ("add", "update"):
                await db.users.update_one(
                    {"user_id": job["user_id"]},
                    {"$unset": {"machine_deleted_at": ""}})
        return
    # Some device command failed.
    if settings.get("retry_failed") and job.get("attempts", 0) < job.get("max_attempts", 3):
        await db.sync_jobs.update_one(
            {"job_id": job["job_id"]},
            {"$set": {"status": "retry", "cmd_ids": [], "updated_at": _now(),
                      "error": "one or more devices failed — retrying"}},
        )
    else:
        await db.sync_jobs.update_one(
            {"job_id": job["job_id"]},
            {"$set": {"status": "failed", "updated_at": _now(),
                      "error": "device sync failed after max retries"}},
        )
        # Phase 4 — notify admins of a permanently failed sync.
        try:
            await db.notifications.insert_one({
                "notification_id": f"n_{uuid.uuid4().hex[:10]}",
                "company_id": job.get("company_id"),
                "audience": "admins",
                "type": "sync.failed",
                "title": "❌ Employee sync failed",
                "body": (
                    f"Sync for {job.get('name') or job.get('pin')} failed on one "
                    "or more machines after retries. Open Device Sync Engine → "
                    "History to review."
                ),
                "created_at": _now(),
                "created_by": "system",
            })
        except Exception:
            pass


async def process_sync_queue() -> Dict[str, int]:
    """One worker pass: dispatch new/retry jobs, reconcile in-flight ones.
    A failure on one job never stops the others."""
    dispatched = reconciled = 0
    async for job in db.sync_jobs.find({"status": {"$in": ["pending", "retry"]}}).limit(100):
        try:
            await _dispatch_job(job)
            dispatched += 1
        except Exception:
            logger.warning("[sync] dispatch error job=%s", job.get("job_id"),
                           exc_info=True)
            await db.sync_jobs.update_one(
                {"job_id": job["job_id"]},
                {"$set": {"status": "failed", "error": "dispatch error",
                          "updated_at": _now()}},
            )
    async for job in db.sync_jobs.find({"status": "processing"}).limit(200):
        try:
            await _reconcile_job(job)
            reconciled += 1
        except Exception:
            logger.warning("[sync] reconcile error job=%s", job.get("job_id"),
                           exc_info=True)
    # Iter 419 — machine-only sync runs due for their distribute phase.
    try:
        await _process_machine_sync_runs()
    except Exception:
        logger.warning("[sync] machine-sync pass error", exc_info=True)
    return {"dispatched": dispatched, "reconciled": reconciled}


# ---------------------------------------------------------------------------
# Iter 419 — MACHINE-ONLY SYNC (user request: master data feeding pending).
# Synchronizes the machines with EACH OTHER — users + fingerprint / face
# templates — without touching or requiring the Employee Master.
# Phase 1 (harvest): every machine is asked to upload its user database and
# templates. Phase 2 (distribute, ~2 min later via the worker loop): every
# captured user + template is pushed to every machine (skipping templates on
# their origin machine). Offline machines catch up on their next poll.
# ---------------------------------------------------------------------------
async def _process_machine_sync_runs() -> int:
    processed = 0
    now = _now()
    async for run in db.machine_sync_runs.find(
            {"phase": "harvest", "distribute_at": {"$lte": now}}).limit(5):
        try:
            await _distribute_machine_sync(run)
            processed += 1
        except Exception:
            logger.warning("[sync] machine-sync distribute error run=%s",
                           run.get("run_id"), exc_info=True)
            await db.machine_sync_runs.update_one(
                {"run_id": run["run_id"]},
                {"$set": {"phase": "failed", "error": "distribute error",
                          "updated_at": _now()}},
            )
    return processed


async def _distribute_machine_sync(run: dict) -> None:
    cid = run["company_id"]
    devices = await _sync_enabled_devices(cid)
    templates = await db.biometric_templates.find(
        {"company_id": cid}, {"_id": 0}).to_list(20000)
    musers: Dict[str, dict] = {}
    async for m in db.biometric_machine_users.find({"company_id": cid}, {"_id": 0}):
        musers[str(m.get("pin") or "").strip()] = m
    pins = sorted({str(t["pin"]).strip() for t in templates} | set(musers.keys()))
    pins = [p for p in pins if p]
    queued = 0
    for d in devices:
        sn = d["serial_number"]
        for pin in pins:
            mu = musers.get(pin) or {}
            nm = (mu.get("name") or "")[:24].replace("\t", " ")
            parts = [f"DATA UPDATE USERINFO PIN={pin}", f"Name={nm}",
                     f"Pri={mu.get('pri') or 0}"]
            if mu.get("passwd"):
                parts.append(f"Passwd={mu['passwd']}")
            if mu.get("card"):
                parts.append(f"Card={mu['card']}")
            await _queue_cmd(sn, "\t".join(parts),
                             run.get("created_by") or "system:machine-sync",
                             f"Machine sync user {nm or pin}")
            queued += 1
        for t in templates:
            if t.get("device_serial") == sn:
                continue  # already enrolled on its origin machine
            await _queue_cmd(sn, _template_to_cmd(t),
                             run.get("created_by") or "system:machine-sync",
                             f"Machine sync {t.get('kind')} PIN {t.get('pin')}")
            queued += 1
    await db.machine_sync_runs.update_one(
        {"run_id": run["run_id"]},
        {"$set": {"phase": "done", "queued": queued, "users": len(pins),
                  "templates": len(templates), "devices": len(devices),
                  "distributed_at": _now(), "updated_at": _now()}},
    )
    # Iter 768 (user request) — machine-sync LOG REPORT entry completion.
    await db.machine_sync_logs.update_one(
        {"run_id": run["run_id"]},
        {"$set": {"status": "done", "queued": queued, "users": len(pins),
                  "templates": len(templates), "devices": len(devices),
                  "distributed_at": _now(), "updated_at": _now()}},
    )
    logger.info("[sync] machine-sync run=%s distributed: %d cmd(s), %d user(s), "
                "%d template(s) -> %d device(s)",
                run.get("run_id"), queued, len(pins), len(templates), len(devices))


async def maybe_auto_machine_sync(company_id: str, reason: str) -> None:
    """Iter 768 (user request) — AUTO Machine → Machine sync.

    Called when a machine uploads a NEW or CHANGED user. Debounced: if any
    machine-sync run for the firm started in the last 10 minutes, the new
    reason is appended to its log instead of starting another run."""
    if not company_id:
        return
    st = await get_sync_settings(company_id)
    if not st.get("machine_auto_sync", True) or \
            not st.get("machine_to_machine_sync", True):
        return
    cutoff = (datetime.now(timezone.utc) - timedelta(minutes=10)) \
        .isoformat().replace("+00:00", "Z")
    recent = await db.machine_sync_runs.find_one(
        {"company_id": company_id, "created_at": {"$gte": cutoff}},
        {"_id": 0, "run_id": 1})
    if recent:
        await db.machine_sync_logs.update_one(
            {"run_id": recent["run_id"]},
            {"$push": {"reasons": {"$each": [reason], "$slice": -50}},
             "$set": {"updated_at": _now()}})
        return
    devices = await _sync_enabled_devices(company_id)
    if len(devices) < 2:
        return  # nothing to distribute to
    for d in devices:
        for cmd, label in (
            ("DATA QUERY USERINFO", "Auto machine sync — query users"),
            ("DATA QUERY FINGERTMP", "Auto machine sync — query fingerprints"),
            ("DATA QUERY BIODATA", "Auto machine sync — query face/bio-data"),
        ):
            await _queue_cmd(d["serial_number"], cmd, "system:auto-m2m", label)
    run_id = f"ms_{uuid.uuid4().hex[:12]}"
    await db.machine_sync_runs.insert_one({
        "run_id": run_id,
        "company_id": company_id,
        "phase": "harvest",
        "source": "auto",
        "distribute_at": (datetime.now(timezone.utc) + timedelta(seconds=120))
            .isoformat().replace("+00:00", "Z"),
        "created_by": "system:auto-m2m",
        "created_at": _now(),
        "updated_at": _now(),
    })
    await db.machine_sync_logs.insert_one({
        "log_id": f"msl_{uuid.uuid4().hex[:12]}",
        "run_id": run_id,
        "company_id": company_id,
        "source": "auto",
        "reasons": [reason],
        "machines": [d["serial_number"] for d in devices],
        "status": "harvesting",
        "created_at": _now(),
        "updated_at": _now(),
    })
    logger.info("[sync] AUTO machine-sync started for %s — %s",
                company_id, reason)


async def sync_engine_loop():
    """Background scheduler — drains the sync queue every 30 seconds."""
    logger.info("[sync] engine loop started (30s cadence)")
    # Iter 419 — "Always sync, no approval": clear any conflicts still
    # waiting for review from before the auto-approve rule (idempotent).
    try:
        r = await db.sync_conflicts.update_many(
            {"status": "open"},
            {"$set": {"status": "approved",
                      "resolved_by": "system:auto-approve",
                      "resolved_at": _now()}})
        if r.modified_count:
            logger.info("[sync] auto-approved %d legacy open conflict(s)",
                        r.modified_count)
    except Exception:
        pass
    while True:
        try:
            await process_sync_queue()
        except Exception:
            logger.warning("[sync] engine loop error", exc_info=True)
        await asyncio.sleep(30)


# ---------------------------------------------------------------------------
# Conflict logging (called from the ZKTeco template ingest path)
# ---------------------------------------------------------------------------
async def log_template_conflict(company_id: str, pin: str, device_serial: str,
                                kind: str) -> None:
    """Iter 419 (user rule: "Always sync — no approval needed"): a machine
    holding biometric data the portal has nowhere else is captured and
    AUTO-APPROVED immediately — the template is already stored by the
    ingest path, so it syncs to other machines with zero admin action.
    A resolved audit row is kept so the History still shows what happened."""
    try:
        exists = await db.biometric_templates.count_documents(
            {"company_id": company_id, "pin": pin, "kind": kind,
             "device_serial": {"$ne": device_serial}})
        if exists:
            return
        await db.sync_conflicts.update_one(
            {"company_id": company_id, "pin": pin, "kind": kind,
             "device_serial": device_serial},
            {"$setOnInsert": {
                "conflict_id": f"cf_{uuid.uuid4().hex[:10]}",
                "reason": "machine_only_template",
                "detail": f"{kind} template present only on device {device_serial} — auto-approved",
                "created_at": _now(),
            },
             "$set": {
                "status": "approved",
                "resolved_by": "system:auto-approve",
                "resolved_at": _now(),
            }},
            upsert=True,
        )
    except Exception:
        logger.warning("[sync] conflict log failed", exc_info=True)


# ---------------------------------------------------------------------------
# REST API
# ---------------------------------------------------------------------------
def _scope_company(admin: dict, company_id: Optional[str]) -> str:
    if admin["role"] == "company_admin":
        return admin["company_id"]
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    return company_id


@router.get("/sync/settings")
async def get_settings_api(company_id: Optional[str] = Query(None),
                           authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, company_id)
    return await get_sync_settings(cid)


@router.put("/sync/settings")
async def put_settings_api(payload: dict = Body(...),
                           authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, payload.get("company_id"))
    update = {k: payload[k] for k in SYNC_DEFAULTS if k in payload}
    # Iter 584 — automatic master sync is LOCKED; the toggle is not saved.
    update.pop("enable_auto_sync", None)
    if "max_retry_count" in update:
        update["max_retry_count"] = max(0, min(10, int(update["max_retry_count"])))
    if "sync_interval" in update:
        update["sync_interval"] = max(15, min(3600, int(update["sync_interval"])))
    update["updated_at"] = _now()
    await db.sync_settings.update_one(
        {"company_id": cid}, {"$set": {"company_id": cid, **update}}, upsert=True)
    return await get_sync_settings(cid)


@router.post("/sync/employee")
async def sync_employee_api(payload: dict = Body(...),
                            authorization: Optional[str] = Header(None)):
    """Manually queue a sync for one employee. Body: {company_id?, user_id,
    action?}. ``action`` defaults to 'update'."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, payload.get("company_id"))
    raise HTTPException(
        status_code=403,
        detail="MASTER_DATA_DEVICE_SYNC_DISABLED — Employee Master → Machine "
               "sync is permanently disabled. Use Device Sync Engine → "
               "Register Employee on Machine (manual).")


def _left_query(cid: str) -> dict:
    """Employees marked LEFT (resigned/exited) who still have a Bio Code."""
    return {
        "company_id": cid, "role": "employee",
        "bio_code": {"$nin": [None, ""]},
        "$or": [
            {"exit_date": {"$nin": [None, ""]}},
            {"resign_date": {"$nin": [None, ""]}},
            {"employment_status": {"$in": [
                "resigned", "exited", "left", "terminated",
                "Resigned", "Exited", "Left", "Terminated"]}},
            {"active": False},
        ],
    }


@router.post("/sync/delete-employee")
async def sync_delete_employee_api(payload: dict = Body(...),
                                   authorization: Optional[str] = Header(None)):
    """Iter 541 (user request) — DELETE one employee from all sync-enabled
    machines. Body: {company_id?, user_id? | employee_code? | bio_code?,
    dry_run?}. dry_run resolves and returns the employee without queueing."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, payload.get("company_id"))
    q: dict = {"company_id": cid, "role": "employee"}
    if payload.get("user_id"):
        q["user_id"] = payload["user_id"]
    elif payload.get("employee_code"):
        q["employee_code"] = str(payload["employee_code"]).strip()
    elif payload.get("bio_code"):
        q["bio_code"] = str(payload["bio_code"]).strip()
    else:
        raise HTTPException(status_code=400,
                            detail="user_id, employee_code or bio_code required")
    emp = await db.users.find_one(q, {"_id": 0, "user_id": 1, "name": 1,
                                      "employee_code": 1, "bio_code": 1})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    if not str(emp.get("bio_code") or "").strip():
        raise HTTPException(status_code=400,
                            detail="Employee has no Bio Code — nothing to "
                                   "delete on the machines.")
    if payload.get("dry_run"):
        return {"ok": True, "dry_run": True, "employee": emp}
    settings = await get_sync_settings(cid)
    if not settings.get("manual_employee_delete"):
        raise HTTPException(status_code=403,
                            detail="MANUAL_EMPLOYEE_DELETE_DISABLED — enable "
                                   "'Manual Employee Delete' in Sync Settings first.")
    if not _has_manual_perm(admin, "delete"):
        raise HTTPException(status_code=403,
                            detail="You lack the BIOMETRIC_MANUAL_EMPLOYEE_DELETE permission.")
    job_id = await enqueue_employee_sync(
        cid, emp["user_id"], "delete", actor=admin["user_id"], force=True,
        sync_type="MANUAL_EMPLOYEE_DELETE")
    if not job_id:
        raise HTTPException(status_code=400,
                            detail="No sync-enabled machine registered for "
                                   "this firm.")
    return {"ok": True, "job_id": job_id, "employee": emp}


@router.post("/sync/delete-left")
async def sync_delete_left_api(payload: dict = Body(None),
                               authorization: Optional[str] = Header(None)):
    """Iter 541 (user request) — remove ALL employees marked LEFT
    (exit/resign date, resigned/left status or inactive) from the firm's
    machines. Body: {company_id?, dry_run?}. dry_run lists the candidates."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    payload = payload or {}
    cid = _scope_company(admin, payload.get("company_id"))
    emps = await db.users.find(
        _left_query(cid),
        {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1,
         "bio_code": 1, "exit_date": 1, "employment_status": 1},
    ).to_list(5000)
    if payload.get("dry_run"):
        return {"ok": True, "dry_run": True, "count": len(emps),
                "employees": emps[:200]}
    settings = await get_sync_settings(cid)
    if not settings.get("manual_employee_delete"):
        raise HTTPException(status_code=403,
                            detail="MANUAL_EMPLOYEE_DELETE_DISABLED — enable "
                                   "'Manual Employee Delete' in Sync Settings first.")
    if not _has_manual_perm(admin, "delete"):
        raise HTTPException(status_code=403,
                            detail="You lack the BIOMETRIC_MANUAL_EMPLOYEE_DELETE permission.")
    queued, skipped = [], 0
    for e in emps:
        job_id = await enqueue_employee_sync(
            cid, e["user_id"], "delete", actor=admin["user_id"], force=True,
            sync_type="MANUAL_EMPLOYEE_DELETE")
        if job_id:
            queued.append({"job_id": job_id, "name": e.get("name"),
                           "employee_code": e.get("employee_code")})
        else:
            skipped += 1
    return {"ok": True, "queued": len(queued), "skipped": skipped,
            "jobs": queued[:200]}


@router.post("/sync/all")
async def sync_all_api(payload: dict = Body(None),
                       authorization: Optional[str] = Header(None)):
    """Queue a sync for every employee (optionally filtered by department /
    group / branch). Body: {company_id?, department?, group?, branch?}."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    raise HTTPException(
        status_code=403,
        detail="MASTER_DATA_DEVICE_SYNC_DISABLED — bulk Employee Master → "
               "Machine sync (EMPLOYEE_MASTER_BULK_SYNC) is permanently "
               "disabled. Use manual Register Employee on Machine instead.")
    payload = payload or {}
    cid = _scope_company(admin, payload.get("company_id"))
    if not await _sync_enabled_devices(cid):
        raise HTTPException(
            status_code=404,
            detail="No sync-enabled machine registered for this company.")
    q: dict = {"role": "employee", "company_id": cid,
               "bio_code": {"$exists": True, "$nin": [None, ""]}}
    for field, key in (("department", "department"), ("employee_group", "group"),
                       ("branch", "branch")):
        if payload.get(key):
            q[field] = payload[key]
    emps = await db.users.find(q, {"_id": 0, "user_id": 1}).to_list(100000)
    queued = 0
    for e in emps:
        if await enqueue_employee_sync(cid, e["user_id"], "update",
                                       actor=admin["user_id"], force=True):
            queued += 1
    # Iter 419 (user report: "0 employee sync job(s) queued") — when nothing
    # queues, say exactly WHY instead of a bare zero.
    if queued == 0:
        base_q = {"role": "employee", "company_id": cid}
        total = await db.users.count_documents(base_q)
        with_bio = await db.users.count_documents(
            {**base_q, "bio_code": {"$exists": True, "$nin": [None, ""]}})
        filters = {k: payload[k] for k in ("department", "group", "branch") if payload.get(k)}
        if total == 0:
            why = ("this firm has no employees in the portal — check the "
                   "firm selected in the dropdown above.")
        elif with_bio == 0:
            why = (f"none of the {total} employee(s) of this firm have a "
                   "Bio Code (machine punch number). Set Bio Codes in the "
                   "Employee Master first, then sync again.")
        elif filters:
            why = (f"{with_bio} employee(s) have a Bio Code, but none match "
                   f"the selected filter {filters} — clear the filter and retry.")
        else:
            why = (f"{with_bio} employee(s) have a Bio Code but no job could "
                   "be queued — check that at least one machine of this firm "
                   "has Sync enabled (Devices tab).")
        return {"ok": True, "queued": 0, "employees": len(emps),
                "message": f"0 sync jobs queued — {why}"}
    return {"ok": True, "queued": queued, "employees": len(emps),
            "message": f"{queued} employee sync job(s) queued."}


@router.post("/sync/machines")
async def sync_machines_only_api(payload: dict = Body(None),
                                 authorization: Optional[str] = Header(None)):
    """Iter 419 — MACHINE-ONLY sync: synchronize all machines of the firm
    with each other (users + FP/face templates captured ON the machines).
    The Employee Master is NOT checked and NOT required. Body: {company_id?}."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    payload = payload or {}
    cid = _scope_company(admin, payload.get("company_id"))
    devices = await _sync_enabled_devices(cid)
    if not devices:
        raise HTTPException(
            status_code=404,
            detail="No sync-enabled machine registered for this company.")
    # Phase 1 — harvest: ask every machine to upload its users + templates.
    for d in devices:
        for cmd, label in (
            ("DATA QUERY USERINFO", "Machine sync — query users"),
            ("DATA QUERY FINGERTMP", "Machine sync — query fingerprints"),
            ("DATA QUERY BIODATA", "Machine sync — query face/bio-data"),
        ):
            await _queue_cmd(d["serial_number"], cmd, admin["user_id"], label)
    run_id = f"ms_{uuid.uuid4().hex[:12]}"
    await db.machine_sync_runs.insert_one({
        "run_id": run_id,
        "company_id": cid,
        "phase": "harvest",
        "source": "manual",
        # Distribute after the machines had time to upload (~2 minutes).
        "distribute_at": (datetime.now(timezone.utc) + timedelta(seconds=120))
            .isoformat().replace("+00:00", "Z"),
        "created_by": admin["user_id"],
        "created_at": _now(),
        "updated_at": _now(),
    })
    # Iter 768 (user request) — machine-sync LOG REPORT entry.
    await db.machine_sync_logs.insert_one({
        "log_id": f"msl_{uuid.uuid4().hex[:12]}",
        "run_id": run_id,
        "company_id": cid,
        "source": "manual",
        "reasons": [f"Manual sync started by {admin.get('name') or admin.get('email') or admin['user_id']}"],
        "machines": [d["serial_number"] for d in devices],
        "status": "harvesting",
        "created_at": _now(),
        "updated_at": _now(),
    })
    return {
        "ok": True, "run_id": run_id, "devices": len(devices),
        "message": (
            f"Machine-only sync started on {len(devices)} machine(s) — "
            "collecting users + fingerprints/faces from every machine now; "
            "distribution to all machines starts automatically in ~2 minutes. "
            "Employee Master is not required."
        ),
    }


@router.get("/sync/machines/logs")
async def sync_machines_logs_api(company_id: Optional[str] = Query(None),
                                 limit: int = Query(50, ge=1, le=200),
                                 authorization: Optional[str] = Header(None)):
    """Iter 768 (user request) — Machine → Machine sync LOG REPORT: every
    manual and AUTO sync run with trigger reasons, machines and counts."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, company_id)
    logs = await db.machine_sync_logs.find(
        {"company_id": cid}, {"_id": 0}
    ).sort("created_at", -1).to_list(limit)
    return {"ok": True, "logs": logs, "count": len(logs)}


@router.get("/sync/machines/status")
async def sync_machines_status_api(company_id: Optional[str] = Query(None),
                                   authorization: Optional[str] = Header(None)):
    """Latest machine-only sync run for the firm."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, company_id)
    run = await db.machine_sync_runs.find_one(
        {"company_id": cid}, {"_id": 0}, sort=[("created_at", -1)])
    return {"run": run}


@router.get("/sync/machines/overview")
async def sync_machines_overview_api(company_id: Optional[str] = Query(None),
                                     authorization: Optional[str] = Header(None)):
    """Iter 419 (user request) — every registered machine of the firm with
    LIVE on-device counts (employees / fingerprints / punch records, as
    reported by the machine itself on each heartbeat) + online status and
    pending command backlog."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, company_id)
    now = datetime.now(timezone.utc)
    devices = await db.biometric_devices.find(
        {"company_id": cid},
        {"_id": 0, "device_id": 1, "serial_number": 1, "name": 1, "brand": 1,
         "model": 1, "kind": 1, "enabled": 1, "sync_enabled": 1,
         "last_seen_at": 1, "user_count": 1, "fp_count": 1,
         "att_log_count": 1, "firmware": 1, "device_ip": 1,
         "templates_captured": 1, "total_punches_ingested": 1},
    ).to_list(200)
    machines = []
    for d in devices:
        online = False
        last = d.get("last_seen_at")
        if last:
            try:
                online = (now - datetime.fromisoformat(
                    str(last).replace("Z", "+00:00"))).total_seconds() < 180
            except Exception:
                pass
        pending_cmds = await db.biometric_device_cmds.count_documents(
            {"device_serial": d["serial_number"],
             "status": {"$in": ["pending", "sent"]}})
        machines.append({
            **d,
            "online": online,
            "pending_cmds": pending_cmds,
            # Employees ON the machine, as the machine reports on heartbeat.
            "employees_on_machine": d.get("user_count"),
        })
    machines.sort(key=lambda m: (not m["online"], m.get("name") or ""))
    return {"machines": machines}


@router.get("/sync/status")
async def sync_status_api(company_id: Optional[str] = Query(None),
                          authorization: Optional[str] = Header(None)):
    """Live dashboard aggregates."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    dev_q: dict = {}
    job_q: dict = {}
    if admin["role"] == "company_admin":
        dev_q["company_id"] = job_q["company_id"] = admin["company_id"]
    elif company_id:
        dev_q["company_id"] = job_q["company_id"] = company_id
    now = datetime.now(timezone.utc)
    devices = await db.biometric_devices.find(dev_q, {"_id": 0, "last_seen_at": 1}).to_list(500)
    online = 0
    for d in devices:
        last = d.get("last_seen_at")
        if last:
            try:
                if (now - datetime.fromisoformat(str(last).replace("Z", "+00:00"))).total_seconds() < 180:
                    online += 1
            except Exception:
                pass
    async def _count(status):
        return await db.sync_jobs.count_documents({**job_q, "status": status})
    pending = await _count("pending")
    retry = await _count("retry")
    processing = await _count("processing")
    synced = await _count("success")
    failed = await _count("failed")
    last_job = await db.sync_jobs.find(
        {**job_q, "status": "success"}, {"_id": 0, "synced_at": 1, "updated_at": 1},
    ).sort("updated_at", -1).limit(1).to_list(1)
    return {
        "total_devices": len(devices),
        "online_devices": online,
        "offline_devices": len(devices) - online,
        "pending_sync": pending + retry,
        "processing": processing,
        "employees_synced": synced,
        "failed_sync": failed,
        "current_queue": pending + retry + processing,
        "last_sync_time": (last_job[0].get("synced_at") or last_job[0].get("updated_at")) if last_job else None,
        "open_conflicts": await db.sync_conflicts.count_documents({**job_q, "status": "open"}),
    }


@router.get("/queue")
async def sync_queue_api(company_id: Optional[str] = Query(None),
                         status: Optional[str] = Query(None),
                         limit: int = Query(100, ge=1, le=500),
                         authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    if status:
        q["status"] = status
    jobs = await db.sync_jobs.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return {"jobs": jobs}


@router.get("/sync/logs")
async def sync_logs_api(company_id: Optional[str] = Query(None),
                        limit: int = Query(200, ge=1, le=1000),
                        authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    logs = await db.sync_log.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return {"logs": logs}


@router.get("/sync/conflicts")
async def sync_conflicts_api(company_id: Optional[str] = Query(None),
                             authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {"status": "open"}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    items = await db.sync_conflicts.find(q, {"_id": 0}).sort("created_at", -1).to_list(300)
    return {"conflicts": items}


@router.post("/sync/conflicts/{conflict_id}/resolve")
async def resolve_conflict_api(conflict_id: str, payload: dict = Body(None),
                               authorization: Optional[str] = Header(None)):
    """Approve (pull the machine template into the portal) or reject a
    conflict. Body: {decision: 'approve'|'reject'}."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    decision = (payload or {}).get("decision") or "approve"
    cf = await db.sync_conflicts.find_one({"conflict_id": conflict_id}, {"_id": 0})
    if not cf:
        raise HTTPException(status_code=404, detail="Conflict not found")
    if admin["role"] == "company_admin" and cf.get("company_id") != admin["company_id"]:
        raise HTTPException(status_code=403, detail="Not authorised")
    await db.sync_conflicts.update_one(
        {"conflict_id": conflict_id},
        {"$set": {"status": "approved" if decision == "approve" else "rejected",
                  "resolved_by": admin["user_id"], "resolved_at": _now()}},
    )
    return {"ok": True, "decision": decision}


@router.get("/sync/report.xlsx")
async def sync_report_xlsx(company_id: Optional[str] = Query(None),
                           authorization: Optional[str] = Header(None)):
    """Phase 4 — consolidated Sync Report: per-device counts, per-status
    totals and the most recent jobs (incl. failures)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    q: dict = {}
    if admin["role"] == "company_admin":
        q["company_id"] = admin["company_id"]
    elif company_id:
        q["company_id"] = company_id
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    jobs = await db.sync_jobs.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    logs = await db.sync_log.find(q, {"_id": 0}).to_list(20000)

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1D4ED8")

    def _hdr(ws, cols):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.font = hf
            cell.fill = hfill

    # Sheet 1 — status summary
    ws1 = wb.active
    ws1.title = "Summary"
    _hdr(ws1, ["Status", "Jobs"])
    counts: Dict[str, int] = {}
    for j in jobs:
        counts[j.get("status", "?")] = counts.get(j.get("status", "?"), 0) + 1
    r = 2
    for k in ("pending", "processing", "retry", "success", "failed", "cancelled"):
        ws1.cell(row=r, column=1, value=k)
        ws1.cell(row=r, column=2, value=counts.get(k, 0))
        r += 1

    # Sheet 2 — per-device
    ws2 = wb.create_sheet("Device-wise")
    _hdr(ws2, ["Device Serial", "Total", "Success", "Failed", "Queued"])
    dev: Dict[str, Dict[str, int]] = {}
    for lg in logs:
        d = dev.setdefault(lg.get("device_serial", "?"),
                           {"total": 0, "success": 0, "failed": 0, "queued": 0})
        d["total"] += 1
        d[lg.get("status", "queued") if lg.get("status") in ("success", "failed", "queued") else "queued"] += 1
    r = 2
    for serial, d in sorted(dev.items()):
        ws2.cell(row=r, column=1, value=serial)
        ws2.cell(row=r, column=2, value=d["total"])
        ws2.cell(row=r, column=3, value=d["success"])
        ws2.cell(row=r, column=4, value=d["failed"])
        ws2.cell(row=r, column=5, value=d["queued"])
        r += 1

    # Sheet 3 — recent jobs
    ws3 = wb.create_sheet("Jobs")
    _hdr(ws3, ["Job ID", "Employee", "PIN", "Action", "Status", "Attempts",
               "Devices", "Created", "Updated", "Error"])
    for i, j in enumerate(jobs[:2000], start=2):
        vals = [j.get("job_id"), j.get("name"), j.get("pin"), j.get("action"),
                j.get("status"), j.get("attempts"), len(j.get("targets") or []),
                j.get("created_at"), j.get("updated_at"), j.get("error")]
        for c, v in enumerate(vals, 1):
            ws3.cell(row=i, column=c, value=v)

    for ws in (ws1, ws2, ws3):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 45)

    buf = BytesIO()
    wb.save(buf)
    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sync-report.xlsx"},
    )


# ---------------------------------------------------------------------------
# Iter 584 — MANUAL EMPLOYEE DEVICE MANAGEMENT (the only Portal → Device ops)
# ---------------------------------------------------------------------------
def _has_manual_perm(admin: dict, kind: str) -> bool:
    """BIOMETRIC_MANUAL_EMPLOYEE_REGISTRATION / _DELETE permission check.
    Super admin always allowed; firm admin allowed; sub-admins need the
    dedicated permission (or full biometric_devices:write)."""
    role = admin.get("role")
    if role in ("super_admin", "company_admin"):
        return True
    if role == "sub_admin":
        perms = admin.get("permissions") or []
        return (f"biometric_manual_employee_{kind}" in perms
                or "biometric_devices:write" in perms)
    return False


def _dev_online(d: dict) -> bool:
    ts = str(d.get("last_seen") or d.get("last_contact") or "")
    if not ts:
        return False
    try:
        seen = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        return (datetime.now(timezone.utc) - seen) < timedelta(minutes=10)
    except Exception:
        return False


def _dev_row(d: dict) -> dict:
    return {
        "serial_number": d.get("serial_number"),
        "name": d.get("name") or d.get("device_name"),
        "model": d.get("model") or d.get("device_model"),
        "location": d.get("location"),
        "kind": d.get("kind"),
        "online": _dev_online(d),
        "last_seen": d.get("last_seen") or d.get("last_contact"),
    }


async def _resolve_emp(cid: str, payload_or_qs: dict) -> dict:
    q: dict = {"company_id": cid, "role": "employee"}
    if payload_or_qs.get("user_id"):
        q["user_id"] = str(payload_or_qs["user_id"]).strip()
    elif payload_or_qs.get("employee_code"):
        q["employee_code"] = str(payload_or_qs["employee_code"]).strip()
    elif payload_or_qs.get("bio_code"):
        q["bio_code"] = str(payload_or_qs["bio_code"]).strip()
    else:
        raise HTTPException(status_code=400,
                            detail="user_id, employee_code or bio_code required")
    emp = await db.users.find_one(q, {"_id": 0, "user_id": 1, "name": 1,
                                      "employee_code": 1, "bio_code": 1,
                                      "card_no": 1, "card_number": 1,
                                      "punch_password": 1, "device_password": 1,
                                      "active": 1})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found in this firm")
    return emp


@router.get("/device-sync/registration-preview")
async def manual_registration_preview(
    company_id: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),
    employee_code: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Everything the confirmation screen needs: employee + Device User ID,
    per-field availability (honest — templates are sent only when the portal
    actually HAS them), the firm's machines with online status, and on which
    machines the PIN is already registered (duplicate protection)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, company_id)
    emp = await _resolve_emp(cid, {"user_id": user_id, "employee_code": employee_code})
    pin = str(emp.get("bio_code") or "").strip()
    devices = await db.biometric_devices.find(
        {"company_id": cid, "enabled": {"$ne": False}}, {"_id": 0}).to_list(200)
    registered_on = set()
    if pin:
        registered_on = {
            m.get("device_serial")
            async for m in db.biometric_machine_users.find(
                {"company_id": cid, "pin": pin},
                {"_id": 0, "device_serial": 1})}
    fp_count = face_count = 0
    if pin:
        fp_count = await db.biometric_templates.count_documents(
            {"company_id": cid, "pin": pin, "kind": "fp"})
        face_count = await db.biometric_templates.count_documents(
            {"company_id": cid, "pin": pin, "kind": "face"})
    fields = {
        "name": {"available": bool(emp.get("name")), "value": emp.get("name")},
        "employee_code": {"available": bool(emp.get("employee_code")),
                          "value": emp.get("employee_code")},
        "device_user_id": {"available": bool(pin), "value": pin},
        "card": {"available": bool(emp.get("card_no") or emp.get("card_number"))},
        "password": {"available": bool(emp.get("punch_password") or emp.get("device_password"))},
        "fingerprint": {"available": fp_count > 0, "count": fp_count,
                        "note": None if fp_count else
                        "No fingerprint template in the portal — enroll on a "
                        "machine and run Machine Sync to capture it first."},
        "face": {"available": face_count > 0, "count": face_count,
                 "note": None if face_count else
                 "No face template in the portal — not supported until one "
                 "is captured from a machine."},
    }
    return {
        "employee": {"user_id": emp["user_id"], "name": emp.get("name"),
                     "employee_code": emp.get("employee_code"),
                     "device_user_id": pin or None,
                     "active": emp.get("active", True)},
        "fields": fields,
        "machines": [{**_dev_row(d),
                      "already_registered": d.get("serial_number") in registered_on}
                     for d in devices],
        "registered_on": sorted(registered_on),
    }


@router.post("/device-sync/manual-register-employee")
async def manual_register_employee(
    payload: dict = Body(...),
    authorization: Optional[str] = Header(None),
):
    """Iter 584 — MANUAL Employee Registration on selected machine(s).
    Body: {company_id?, user_id | employee_code, device_serials: [..],
    fields: ["card","password","fingerprint","face"], update_existing?}.
    sync_type = MANUAL_EMPLOYEE_REGISTRATION (the only permitted
    Portal → Device registration path)."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, payload.get("company_id"))
    settings = await get_sync_settings(cid)
    if not settings.get("manual_employee_registration"):
        raise HTTPException(status_code=403,
                            detail="MANUAL_EMPLOYEE_REGISTRATION_DISABLED — "
                                   "enable 'Manual Employee Registration' in "
                                   "Sync Settings first.")
    if not _has_manual_perm(admin, "registration"):
        raise HTTPException(status_code=403,
                            detail="You lack the BIOMETRIC_MANUAL_EMPLOYEE_"
                                   "REGISTRATION permission.")
    emp = await _resolve_emp(cid, payload)
    pin = str(emp.get("bio_code") or "").strip()
    if not pin:
        raise HTTPException(status_code=400,
                            detail="Employee has no Device User ID (Bio Code) "
                                   "— set it in the Employee Master first.")
    serials = [str(s) for s in (payload.get("device_serials") or []) if s]
    if not serials:
        raise HTTPException(status_code=400,
                            detail="Select at least one machine "
                                   "(device_serials).")
    firm_serials = {d["serial_number"] for d in await db.biometric_devices.find(
        {"company_id": cid}, {"_id": 0, "serial_number": 1}).to_list(200)}
    bad = [s for s in serials if s not in firm_serials]
    if bad:
        raise HTTPException(status_code=400,
                            detail=f"Machine(s) not in this firm: {bad}")
    # Duplicate registration protection.
    already = [m.get("device_serial") async for m in db.biometric_machine_users.find(
        {"company_id": cid, "pin": pin, "device_serial": {"$in": serials}},
        {"_id": 0, "device_serial": 1})]
    if already and not payload.get("update_existing"):
        raise HTTPException(
            status_code=409,
            detail=f"Employee already exists on: {', '.join(sorted(already))}. "
                   "Pass update_existing=true to update the device user "
                   "instead of creating a duplicate.")
    fields = [f for f in (payload.get("fields") or ["card", "password"])
              if f in ("card", "password", "fingerprint", "face")]
    job_id = await enqueue_employee_sync(
        cid, emp["user_id"], "add", actor=admin["user_id"], force=True,
        sync_type="MANUAL_EMPLOYEE_REGISTRATION",
        target_serials=serials, send_fields=fields)
    if not job_id:
        raise HTTPException(status_code=400,
                            detail="Could not queue — no sync-enabled machine "
                                   "matched your selection.")
    return {"ok": True, "job_id": job_id, "status": "QUEUED",
            "employee": {"name": emp.get("name"),
                         "employee_code": emp.get("employee_code"),
                         "device_user_id": pin},
            "machines": serials, "fields_sent": ["name", "employee_code",
                                                 "device_user_id"] + fields,
            "message": "Registration queued — offline machines execute the "
                       "command on their next contact."}


@router.post("/device-sync/manual-delete-employee")
async def manual_delete_employee(
    payload: dict = Body(...),
    authorization: Optional[str] = Header(None),
):
    """Iter 584 — MANUAL Delete Employee FROM MACHINE(s) only. The payroll
    employee, attendance, punches, salary and PF/ESIC data are NEVER
    touched. Body: {company_id?, user_id | employee_code,
    device_serials: [..] OR all_registered: true (+ confirm_code =
    employee code), }. sync_type = MANUAL_EMPLOYEE_DELETE."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, payload.get("company_id"))
    settings = await get_sync_settings(cid)
    if not settings.get("manual_employee_delete"):
        raise HTTPException(status_code=403,
                            detail="MANUAL_EMPLOYEE_DELETE_DISABLED — enable "
                                   "'Manual Employee Delete' in Sync Settings "
                                   "first.")
    if not _has_manual_perm(admin, "delete"):
        raise HTTPException(status_code=403,
                            detail="You lack the BIOMETRIC_MANUAL_EMPLOYEE_"
                                   "DELETE permission.")
    emp = await _resolve_emp(cid, payload)
    pin = str(emp.get("bio_code") or "").strip()
    if not pin:
        raise HTTPException(status_code=400,
                            detail="Employee has no Device User ID (Bio Code) "
                                   "— nothing to delete on the machines.")
    if payload.get("all_registered"):
        # Extra confirmation for the destructive all-machines variant.
        if str(payload.get("confirm_code") or "").strip() != \
                str(emp.get("employee_code") or "").strip():
            raise HTTPException(status_code=400,
                                detail="Type the Employee Code in confirm_code "
                                       "to delete from ALL registered machines.")
        serials = sorted({m.get("device_serial") async for m in
                          db.biometric_machine_users.find(
                              {"company_id": cid, "pin": pin},
                              {"_id": 0, "device_serial": 1})})
        if not serials:
            raise HTTPException(status_code=404,
                                detail="This Device User ID is not registered "
                                       "on any machine (per the machines' own "
                                       "user lists).")
    else:
        serials = [str(s) for s in (payload.get("device_serials") or []) if s]
        if not serials:
            raise HTTPException(status_code=400,
                                detail="Select at least one machine "
                                       "(device_serials) or pass "
                                       "all_registered=true.")
        firm_serials = {d["serial_number"] for d in await db.biometric_devices.find(
            {"company_id": cid}, {"_id": 0, "serial_number": 1}).to_list(200)}
        bad = [s for s in serials if s not in firm_serials]
        if bad:
            raise HTTPException(status_code=400,
                                detail=f"Machine(s) not in this firm: {bad}")
    job_id = await enqueue_employee_sync(
        cid, emp["user_id"], "delete", actor=admin["user_id"], force=True,
        sync_type="MANUAL_EMPLOYEE_DELETE", target_serials=serials)
    if not job_id:
        raise HTTPException(status_code=400,
                            detail="Could not queue — no sync-enabled machine "
                                   "matched your selection.")
    return {"ok": True, "job_id": job_id, "status": "QUEUED",
            "employee": {"name": emp.get("name"),
                         "employee_code": emp.get("employee_code"),
                         "device_user_id": pin},
            "machines": serials,
            "payroll_unchanged": True,
            "message": "Delete queued for the selected machine(s) ONLY — the "
                       "payroll employee, attendance and salary history are "
                       "not touched."}


@router.get("/device-sync/activity")
async def manual_device_activity(
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Recent MANUAL register/delete jobs for the dashboard section."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _scope_company(admin, company_id)
    jobs = await db.sync_jobs.find(
        {"company_id": cid,
         "sync_type": {"$in": list(_ALLOWED_PORTAL_SYNC)}},
        {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"jobs": jobs}
