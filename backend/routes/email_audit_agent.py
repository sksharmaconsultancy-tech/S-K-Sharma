"""Iter 674 — 🤖 EMAIL AUDIT AGENT (Phase 1 — strictly READ-ONLY).

Reads the mailbox already configured in Email SMTP & Notifications (same
IMAP credentials the Super Admin Mailbox uses — NO new SMTP config), then:
  Read → Identify → Company-Link → Audit → Classify → Extract → Report
  → Notify → Suggest

HARD RULES (user spec):
  * Only emails received ON/AFTER 15-Aug-2026 00:00:00 IST are processed —
    enforced at the backend (IMAP SEARCH SINCE + per-message re-check).
  * Message-ID dedupe: an email is never audited twice.
  * Phase 1 NEVER sends/replies/deletes email and NEVER touches payroll —
    IMAP is opened READ-ONLY (BODY.PEEK, readonly select).
  * Company auto-link priority: exact registered email → registered domain
    (common webmail domains excluded) → AI content identification.
    Multiple registered companies for one email → COMPANY_REVIEW_REQUIRED.
  * AI (GPT-5.4 via Emergent LLM key) classifies, extracts, summarizes and
    recommends; confidence < threshold (default 80) → REVIEW_REQUIRED.
  * Super Admin only.

Collections:
  email_agent_state       — singleton {enabled, sandbox, threshold, ...}
  company_email_registry  — registered emails per company (multi allowed)
  email_audit_records     — permanent audit records + processing timeline
"""
import asyncio
import base64
import io
import json
import logging
import os
import re
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from email.utils import getaddresses, parseaddr, parsedate_to_datetime
from typing import Optional

from fastapi import APIRouter, Header, HTTPException, Query
from pydantic import BaseModel

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
    now_iso,
)
from utils.notify import emit as notify_emit  # noqa: E402

router = APIRouter(prefix="/api/email-agent", tags=["email-audit-agent"])
logger = logging.getLogger("email_audit_agent")

IST = timezone(timedelta(hours=5, minutes=30))
CUTOFF = datetime(2026, 8, 15, 0, 0, 0, tzinfo=IST)
CUTOFF_IMAP = "15-Aug-2026"
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")

COMMON_DOMAINS = {
    "gmail.com", "googlemail.com", "outlook.com", "hotmail.com", "live.com",
    "yahoo.com", "yahoo.in", "yahoo.co.in", "rediffmail.com", "icloud.com",
    "aol.com", "protonmail.com", "proton.me", "zoho.com", "msn.com",
}
CATEGORIES = [
    "Salary Processing", "New Employee", "Employee Master Update",
    "Employee Exit", "Attendance", "Leave", "Overtime", "Advance",
    "Deduction", "PF / EPF", "ESIC", "Labour Compliance",
    "Compliance Documents", "Payroll Query", "Client Query",
    "General Information", "Complaint / Issue", "Urgent / Important",
    "Unknown / Requires Review",
]
MAX_FULL_FETCH_PER_SCAN = 20     # AI-processed emails per scan run
MAX_HEADER_SCAN = 300            # newest headers examined per scan
BODY_CHARS_FOR_AI = 5000
ATTACH_EXCERPT_CHARS = 1200

STATE_DEFAULT = {
    "_singleton": True, "enabled": False, "sandbox": False,
    "threshold": 80, "poll_minutes": 5,
    "last_scan_at": None, "last_scan_result": None,
}


async def _sa(authorization: Optional[str]) -> dict:
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin"])
    return admin


async def _state() -> dict:
    doc = await db.email_agent_state.find_one({"_singleton": True}, {"_id": 0})
    if not doc:
        doc = dict(STATE_DEFAULT)
        await db.email_agent_state.insert_one(dict(doc))
    return {**STATE_DEFAULT, **doc}


def _tl(record: dict, step: str, detail: str = "") -> None:
    record.setdefault("timeline", []).append(
        {"step": step, "detail": detail, "at": now_iso()})


# ───────────────────────── mailbox reading (sync, thread) ──────────────────

def _strip_html(html: str) -> str:
    txt = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", html or "",
                 flags=re.DOTALL | re.IGNORECASE)
    txt = re.sub(r"<[^>]+>", " ", txt)
    txt = re.sub(r"&nbsp;?", " ", txt)
    return re.sub(r"\s+", " ", txt).strip()


def _attachment_excerpt(name: str, data: bytes) -> dict:
    """Best-effort READ-ONLY peek into an attachment. Never raises."""
    out = {"readable": False, "excerpt": "", "note": ""}
    low = (name or "").lower()
    try:
        if low.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.worksheets[0]
            rows, seen_rows, dup_rows, blank_rows, codes = [], set(), 0, 0, []
            n_rows = n_cols = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 500:
                    break
                vals = ["" if c is None else str(c).strip() for c in row[:25]]
                if not any(vals):
                    blank_rows += 1
                    continue
                n_rows += 1
                n_cols = max(n_cols, len([v for v in vals if v]))
                key = "|".join(vals)
                if key in seen_rows:
                    dup_rows += 1
                seen_rows.add(key)
                for v in vals[:2]:  # candidate employee codes (first 2 cols)
                    if v and len(v) <= 20 and i > 0:
                        codes.append(v.upper())
                if i < 12:
                    rows.append(" | ".join(v[:30] for v in vals[:15]))
            wb.close()
            out.update(readable=True, excerpt="\n".join(rows)[:ATTACH_EXCERPT_CHARS],
                       note=f"{n_rows} data rows",
                       stats={"rows": n_rows, "cols": n_cols,
                              "blank_rows": blank_rows, "duplicate_rows": dup_rows,
                              "sheets": len(wb.sheetnames) if hasattr(wb, "sheetnames") else 1},
                       codes=codes[:300])
        elif low.endswith(".csv"):
            out.update(readable=True,
                       excerpt=data[:4000].decode("utf-8", "replace")[:ATTACH_EXCERPT_CHARS])
        elif low.endswith(".pdf"):
            from pypdf import PdfReader
            rd = PdfReader(io.BytesIO(data))
            txt = " ".join((p.extract_text() or "") for p in rd.pages[:2])
            out.update(readable=True, excerpt=txt[:ATTACH_EXCERPT_CHARS],
                       note=f"{len(rd.pages)} pages")
        elif low.endswith(".docx"):
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                xml = z.read("word/document.xml").decode("utf-8", "replace")
            out.update(readable=True, excerpt=_strip_html(xml)[:ATTACH_EXCERPT_CHARS])
        elif low.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                names = z.namelist()[:20]
            out.update(readable=True, excerpt="Contains: " + ", ".join(names),
                       note=f"{len(names)} file(s) listed")
        elif low.endswith((".doc",)):
            out["note"] = "Legacy .doc — metadata only"
        elif low.endswith((".jpg", ".jpeg", ".png", ".webp")):
            # Iter 685 (user request) — document photos (Aadhaar / PAN /
            # bank passbook / cheque) are OCR-read by the AI vision step in
            # _process_email. Keep the bytes (base64) for that step ONLY —
            # stripped before the record is stored (never saved to DB).
            if 0 < len(data) <= 6 * 1024 * 1024:
                out["_b64"] = base64.b64encode(data).decode()
                out["note"] = "Document image — OCR scan queued"
            else:
                out["note"] = "Image too large for OCR (>6 MB)"
        else:
            out["note"] = "Type not analyzed"
    except Exception as exc:  # corrupted / unreadable
        out["note"] = f"Unreadable: {str(exc)[:120]}"
    return out


def _parse_msg(raw: bytes) -> dict:
    """RFC822 bytes → normalized dict (read-only, attachments peeked)."""
    import email as email_lib
    msg = email_lib.message_from_bytes(raw)
    from routes.gmail_mailbox import _decode_hdr  # reuse existing decoder
    sender_name, sender_email = parseaddr(_decode_hdr(msg.get("From")))
    try:
        rec_dt = parsedate_to_datetime(msg.get("Date"))
        if rec_dt.tzinfo is None:
            rec_dt = rec_dt.replace(tzinfo=timezone.utc)
    except Exception:
        rec_dt = datetime.now(timezone.utc)
    plain = html = None
    attachments = []
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        fname = part.get_filename()
        if fname:
            fname = _decode_hdr(fname)
            try:
                data = part.get_payload(decode=True) or b""
            except Exception:
                data = b""
            peek = _attachment_excerpt(fname, data)
            attachments.append({
                "name": fname, "type": part.get_content_type(),
                "size": len(data), **peek,
            })
            continue
        try:
            payload = part.get_payload(decode=True)
            decoded = payload.decode(part.get_content_charset() or "utf-8",
                                     "replace") if payload else None
        except Exception:
            decoded = None
        if not decoded:
            continue
        if part.get_content_type() == "text/plain" and plain is None:
            plain = decoded
        elif part.get_content_type() == "text/html" and html is None:
            html = decoded
    body = (plain or _strip_html(html or "")).strip()
    return {
        "message_id": (msg.get("Message-ID") or f"<no-id-{uuid.uuid4().hex}>").strip(),
        "thread_id": (msg.get("References") or msg.get("In-Reply-To") or "").strip()[:300],
        "sender_name": sender_name or "",
        "sender_email": (sender_email or "").lower(),
        "recipient": _decode_hdr(msg.get("To"))[:400],
        "cc": ", ".join(a for _, a in getaddresses([msg.get("Cc") or ""]))[:400],
        "subject": _decode_hdr(msg.get("Subject"))[:300],
        "received_at": rec_dt.astimezone(timezone.utc).isoformat(),
        "body_text": body[:20000],
        "attachments": attachments,
    }


def _imap_fetch_new_sync(settings: dict, known_ids: set,
                         registered_emails: set) -> tuple:
    """Fetch new eligible messages since the cutoff. Returns
    (parsed_list, headers_seen, historical_skipped). READ-ONLY (PEEK).

    Iter 680 (user rules):
      * INBOX — Gmail PRIMARY category ONLY (Updates / Social / Promotions
        are excluded via X-GM-RAW "category:primary"; plain SINCE search
        is the fallback for non-Gmail servers).
      * SPAM — also scanned, but ONLY messages whose sender address is
        registered in the Company Email Registry are processed (a client
        mail wrongly landing in Spam is never missed)."""
    from routes.gmail_mailbox import _imap_connect
    box = _imap_connect(settings)
    parsed, seen, historical = [], 0, 0

    def _search_uids(primary_only: bool) -> list:
        if primary_only:
            try:
                typ, data = box.uid("SEARCH", None, "X-GM-RAW",
                                    '"category:primary"',
                                    f"(SINCE {CUTOFF_IMAP})")
                if typ == "OK":
                    return (data[0] or b"").split()
            except Exception:
                pass  # not Gmail — fall back to the whole INBOX
        typ, data = box.uid("SEARCH", None, f"(SINCE {CUTOFF_IMAP})")
        return (data[0] or b"").split() if typ == "OK" else []

    def _harvest(folder_label: str, uids: list, only_registered: bool):
        nonlocal seen, historical
        import email as email_lib
        for uid in uids[-MAX_HEADER_SCAN:][::-1]:  # newest first
            if len(parsed) >= MAX_FULL_FETCH_PER_SCAN:
                return
            typ, md = box.uid(
                "FETCH", uid,
                "(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID DATE FROM)])")
            if typ != "OK" or not md or not isinstance(md[0], tuple):
                continue
            seen += 1
            hdr = email_lib.message_from_bytes(md[0][1] or b"")
            mid = (hdr.get("Message-ID") or "").strip()
            if mid and mid in known_ids:
                continue
            if only_registered:
                _, frm = parseaddr(hdr.get("From") or "")
                if (frm or "").lower() not in registered_emails:
                    continue  # spam from unknown senders stays untouched
            try:
                rec_dt = parsedate_to_datetime(hdr.get("Date"))
                if rec_dt.tzinfo is None:
                    rec_dt = rec_dt.replace(tzinfo=timezone.utc)
            except Exception:
                rec_dt = datetime.now(timezone.utc)
            if rec_dt < CUTOFF:
                historical += 1
                continue  # IGNORED_HISTORICAL — never processed
            typ, fd = box.uid("FETCH", uid, "(BODY.PEEK[])")
            if typ != "OK" or not fd or not isinstance(fd[0], tuple):
                continue
            p = _parse_msg(fd[0][1] or b"")
            p["imap_uid"] = uid.decode()
            p["folder"] = folder_label
            parsed.append(p)

    try:
        # 1) INBOX — Primary category only.
        box.select("INBOX", readonly=True)  # READ-ONLY guarantee
        _harvest("INBOX", _search_uids(primary_only=True), only_registered=False)
        # 2) SPAM — registered company senders only.
        if registered_emails and len(parsed) < MAX_FULL_FETCH_PER_SCAN:
            for spam_box in ('"[Gmail]/Spam"', "Spam", "Junk"):
                try:
                    typ, _ = box.select(spam_box, readonly=True)
                except Exception:
                    typ = "NO"
                if typ == "OK":
                    _harvest("SPAM", _search_uids(primary_only=False),
                             only_registered=True)
                    break
    finally:
        try:
            box.logout()
        except Exception:
            pass
    return parsed, seen, historical


# ───────────────────────── company matching ────────────────────────────────

async def _match_company(sender_email: str) -> dict:
    """Priority 1: exact registered email. Priority 2: registered domain.
    Returns match dict; content match (priority 3) is applied later using
    the AI result."""
    out = {"company_id": None, "company_name": None, "match_type": "UNKNOWN",
           "confidence": 0, "review": False, "candidates": []}
    if not sender_email:
        return out
    regs = await db.company_email_registry.find(
        {"email": sender_email, "active": {"$ne": False}}, {"_id": 0}).to_list(20)
    comp_ids = sorted({r["company_id"] for r in regs})
    if len(comp_ids) == 1:
        out.update(company_id=comp_ids[0],
                   company_name=regs[0].get("company_name"),
                   match_type="EXACT_EMAIL_MATCH", confidence=99)
        return out
    if len(comp_ids) > 1:
        names = {r["company_id"]: r.get("company_name") for r in regs}
        out.update(match_type="MULTIPLE_COMPANY_MATCH", review=True,
                   candidates=[{"company_id": c, "company_name": names.get(c)}
                               for c in comp_ids])
        return out
    domain = sender_email.rsplit("@", 1)[-1]
    if domain and domain not in COMMON_DOMAINS:
        dregs = await db.company_email_registry.find(
            {"email": {"$regex": f"@{re.escape(domain)}$"},
             "active": {"$ne": False}}, {"_id": 0}).to_list(50)
        dcomp = sorted({r["company_id"] for r in dregs})
        if len(dcomp) == 1:
            out.update(company_id=dcomp[0],
                       company_name=dregs[0].get("company_name"),
                       match_type="DOMAIN_MATCH", confidence=90)
            return out
        if len(dcomp) > 1:
            names = {r["company_id"]: r.get("company_name") for r in dregs}
            out.update(match_type="MULTIPLE_COMPANY_MATCH", review=True,
                       candidates=[{"company_id": c, "company_name": names.get(c)}
                                   for c in dcomp])
            return out
    return out


# ───────────────────────── document OCR (Iter 685) ─────────────────────────

_DOC_OCR_PROMPT = """Look at this document photo from an email attachment.
Reply with STRICT JSON ONLY:
{
 "document_type": "AADHAAR CARD"|"PAN CARD"|"BANK PASSBOOK"|"CANCELLED CHEQUE"|
                  "VOTER ID"|"DRIVING LICENCE"|"UAN / PF DOCUMENT"|"ESIC CARD"|
                  "SALARY SLIP"|"PHOTO / SELFIE"|"OTHER DOCUMENT",
 "person_name": null|"name printed on the document",
 "id_number": null|"main number (Aadhaar no / PAN / account no / UAN / IP no)",
 "fields": { .. every other clearly readable field: dob, gender, father_name,
             address, ifsc, bank_name, branch, mobile, issue_date .. },
 "legible": true|false
}
Rules: extract ONLY what is clearly printed — never guess; unreadable → null."""


async def _ocr_documents(attachments: list) -> list:
    """Iter 685 (user request) — OCR-scan image attachments (Aadhaar / PAN /
    bank passbook / cheque photos) with AI vision. READ-ONLY; max 5 images;
    one failure never blocks the email. Returns document_analysis list."""
    from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
    out = []
    imgs = [a for a in attachments if a.get("_b64")][:5]
    for a in imgs:
        entry = {"file_name": a.get("name"), "document_type": "OTHER DOCUMENT",
                 "person_name": None, "id_number": None, "fields": {},
                 "legible": False, "ocr_used": True}
        try:
            chat = LlmChat(
                api_key=EMERGENT_LLM_KEY,
                session_id=f"email-ocr-{uuid.uuid4().hex[:8]}",
                system_message=("You are an OCR reader for Indian government-ID "
                                "and bank documents. Reply with valid JSON only; "
                                "never invent unreadable values."),
            ).with_model("gemini", "gemini-3-flash-preview")
            resp = await chat.send_message(UserMessage(
                text=_DOC_OCR_PROMPT,
                file_contents=[ImageContent(image_base64=a["_b64"])]))
            raw = re.sub(r"^```(?:json)?|```$", "", str(resp).strip(),
                         flags=re.MULTILINE).strip()
            m = re.search(r"\{.*\}", raw, re.DOTALL)
            if m:
                j = json.loads(m.group(0))
                entry["document_type"] = str(j.get("document_type")
                                             or "OTHER DOCUMENT")[:60]
                entry["person_name"] = (str(j.get("person_name"))[:120]
                                        if j.get("person_name") else None)
                entry["id_number"] = (str(j.get("id_number"))[:80]
                                      if j.get("id_number") else None)
                flds = j.get("fields")
                if isinstance(flds, dict):
                    entry["fields"] = {str(k)[:40]: str(v)[:160]
                                       for k, v in list(flds.items())[:15]
                                       if v not in (None, "", "null")}
                entry["legible"] = bool(j.get("legible"))
        except Exception as exc:  # noqa: BLE001 — OCR must never break the audit
            entry["fields"] = {"error": f"OCR failed: {str(exc)[:120]}"}
        out.append(entry)
    return out


def _doc_report_lines(doc_analysis: list) -> list:
    """Human report: 'Found attachment X — AADHAAR CARD of NAME · No. …'."""
    lines = []
    for d in doc_analysis:
        bits = [f"Found attachment {d.get('file_name')} — {d.get('document_type')}"]
        if d.get("person_name"):
            bits.append(f"of {d['person_name']}")
        core = " ".join(bits)
        if d.get("id_number"):
            core += f" · No. {d['id_number']}"
        extras = " · ".join(f"{k}: {v}" for k, v in (d.get("fields") or {}).items()
                            if k != "error")
        if extras:
            core += f" · {extras}"
        if not d.get("legible") and not d.get("id_number"):
            core += " · (image not clearly legible)"
        lines.append(core[:500])
    return lines


# ───────────────────────── AI analysis ─────────────────────────────────────

AI_SYSTEM = """You are the READ-ONLY Email Audit Agent of an Indian payroll/HR
consultancy. Analyze ONE client email and reply with STRICT JSON ONLY:
{
 "categories": [..],                 // subset of: %CATS%
 "priority": "normal"|"urgent",
 "action_required": true|false,
 "summary": "2-3 sentence factual summary",
 "recommendation": "what the human operator should do next (NEVER execute)",
 "possible_company": "company name mentioned/implied or null",
 "company_confidence": 0-100,        // only about possible_company
 "classification_confidence": 0-100,
 "extraction_confidence": 0-100,
 "recommendation_confidence": 0-100,
 "employee_name": null|"..", "employee_code": null|"..",
 "department": null|"..", "designation": null|"..",
 "salary_month": null|"..", "salary_year": null|"..",
 "attendance_month": null|"..", "joining_date": null|"..",
 "exit_date": null|"..",
 "extracted_data": { .. any other clearly stated facts .. },
 "missing_information": [".."],
 "findings": [{"severity":"critical"|"high"|"warning"|"normal","message":".."}],
 "email_vs_attachment": [{"field":"..","email_value":"..","attachment_value":"..","severity":"critical"|"warning"}]
}
Compare what the EMAIL BODY states against what the ATTACHMENT excerpts
show (salary, month, employee count, names) — report every mismatch in
email_vs_attachment; leave it [] when nothing conflicts.
SUMMARY STYLE (user rule): write the summary as a short human ANALYSIS
REPORT — what was received and requested, which documents were found
(name + identified type + whose document), and what the operator should
verify. NEVER output a raw metadata field list (no 'employee_name:',
'forwarded_by:', 'received_timestamp:' style dumps).
Rules: NEVER invent missing information — use null / empty. Known client
companies: %COMPANIES%. If the email clearly names one of them, set
possible_company to that exact name."""


async def _ai_analyze(parsed: dict, doc_analysis: Optional[list] = None) -> dict:
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    comps = await db.companies.find({}, {"_id": 0, "name": 1}).to_list(200)
    names = [c.get("name") for c in comps if c.get("name")]
    sysmsg = AI_SYSTEM.replace("%CATS%", ", ".join(CATEGORIES)) \
                      .replace("%COMPANIES%", "; ".join(names[:120]) or "none")
    att_lines = []
    for a in parsed.get("attachments", []):
        att_lines.append(f"- {a['name']} ({a['type']}, {a['size']} bytes)"
                         + (f"\n  excerpt: {a['excerpt'][:600]}" if a.get("excerpt") else ""))
    # Iter 685 — OCR document readings feed the analysis.
    doc_lines = _doc_report_lines(doc_analysis or [])
    user_txt = (
        f"FROM: {parsed.get('sender_name')} <{parsed.get('sender_email')}>\n"
        f"SUBJECT: {parsed.get('subject')}\n"
        f"RECEIVED: {parsed.get('received_at')}\n"
        f"ATTACHMENTS:\n" + ("\n".join(att_lines) or "none")
        + ("\nOCR DOCUMENT READINGS:\n" + "\n".join(doc_lines) if doc_lines else "")
        + "\n\nBODY:\n"
        + (parsed.get("body_text") or "")[:BODY_CHARS_FOR_AI])
    chat = LlmChat(api_key=EMERGENT_LLM_KEY,
                   session_id=f"email-audit-{uuid.uuid4().hex[:8]}",
                   system_message=sysmsg).with_model("openai", "gpt-5.4")
    resp = await chat.send_message(UserMessage(text=user_txt))
    raw = re.sub(r"^```(?:json)?|```$", "", str(resp).strip(), flags=re.MULTILINE).strip()
    m = re.search(r"\{.*\}", raw, re.DOTALL)
    if not m:
        raise ValueError(f"AI returned no JSON: {raw[:150]}")
    return json.loads(m.group(0))


# ───────────────────────── core pipeline ───────────────────────────────────

async def _process_email(parsed: dict, *, sandbox: bool, threshold: int) -> dict:
    rec = {
        "audit_id": f"ea_{uuid.uuid4().hex[:12]}",
        "message_id": parsed["message_id"],
        "thread_id": parsed.get("thread_id") or "",
        "sender_name": parsed.get("sender_name"),
        "sender_email": parsed.get("sender_email"),
        "recipient": parsed.get("recipient"),
        "cc": parsed.get("cc"),
        "subject": parsed.get("subject"),
        "received_at": parsed.get("received_at"),
        "body_text": (parsed.get("body_text") or "")[:20000],
        # Iter 685 — never store image bytes: _b64 is for the OCR step only.
        "attachments": [{k: v for k, v in a.items() if k != "_b64"}
                        for a in parsed.get("attachments", [])],
        "has_attachments": bool(parsed.get("attachments")),
        "folder": parsed.get("folder") or "INBOX",
        "sandbox": bool(sandbox),
        "company_id": None, "company_name": None,
        "company_match_type": "UNKNOWN", "company_match_confidence": 0,
        "company_candidates": [],
        "possible_company": None,
        "categories": [], "priority": "normal",
        "ai_summary": "", "ai_recommendation": "",
        "extracted": {}, "missing_information": [],
        "confidences": {},
        "status": "REVIEW_REQUIRED",
        "processing_started_at": now_iso(),
        "processing_completed_at": None,
        "error": None,
        "timeline": [],
        "created_at": now_iso(),
    }
    _tl(rec, "Email Received",
        f"From {rec['sender_email']} · {rec['subject'][:80]}"
        + (" · ⚠ found in SPAM (registered company sender)"
           if rec["folder"] == "SPAM" else ""))
    # Backend date-boundary re-check (never trust the IMAP filter alone).
    try:
        rdt = datetime.fromisoformat(str(rec["received_at"]).replace("Z", "+00:00"))
    except Exception:
        rdt = datetime.now(timezone.utc)
    _tl(rec, "Date Boundary Checked", f"Cutoff {CUTOFF.isoformat()}")
    if rdt < CUTOFF:
        rec["status"] = "IGNORED_HISTORICAL"
        rec["processing_completed_at"] = now_iso()
        _tl(rec, "Ignored Historical", "Received before 15-Aug-2026")
        await db.email_audit_records.insert_one(dict(rec))
        rec.pop("_id", None)
        return rec
    _tl(rec, "Sender Identified", rec["sender_email"] or "unknown")
    try:
        match = await _match_company(rec["sender_email"])
        _tl(rec, "Company Match Attempted", match["match_type"])
        rec["company_match_type"] = match["match_type"]
        rec["company_match_confidence"] = match["confidence"]
        rec["company_candidates"] = match["candidates"]
        if match["company_id"]:
            rec["company_id"] = match["company_id"]
            rec["company_name"] = match["company_name"]
            _tl(rec, "Company Matched",
                f"{match['company_name']} · {match['match_type']} {match['confidence']}%")
        # Iter 685 (user request) — OCR SCANNER for document-photo
        # attachments (Aadhaar / PAN / bank passbook / cheque images).
        doc_analysis: list = []
        if any(a.get("_b64") for a in parsed.get("attachments", [])):
            _tl(rec, "OCR Scanner Used",
                ", ".join(a["name"] for a in parsed["attachments"]
                          if a.get("_b64"))[:200])
            doc_analysis = await _ocr_documents(parsed["attachments"])
            for d in doc_analysis:
                _tl(rec, "Document Identified",
                    f"{d.get('file_name')} → {d.get('document_type')}"
                    + (f" ({d['person_name']})" if d.get("person_name") else ""))
        rec["document_analysis"] = doc_analysis
        rec["document_report"] = _doc_report_lines(doc_analysis)
        ai = await _ai_analyze(parsed, doc_analysis)
        _tl(rec, "Content Analyzed", "GPT-5.4 audit complete")
        if rec["has_attachments"]:
            _tl(rec, "Attachment Analyzed",
                ", ".join(a["name"] for a in rec["attachments"])[:200])
        rec["categories"] = [c for c in (ai.get("categories") or [])
                             if c in CATEGORIES] or ["Unknown / Requires Review"]
        rec["priority"] = "urgent" if ai.get("priority") == "urgent" else "normal"
        _tl(rec, "Email Classified", ", ".join(rec["categories"]))
        rec["ai_summary"] = str(ai.get("summary") or "")[:2000]
        rec["ai_recommendation"] = str(ai.get("recommendation") or "")[:2000]
        rec["possible_company"] = ai.get("possible_company")
        rec["extracted"] = {k: ai.get(k) for k in (
            "employee_name", "employee_code", "department", "designation",
            "salary_month", "salary_year", "attendance_month",
            "joining_date", "exit_date") if ai.get(k)}
        extra = ai.get("extracted_data")
        if isinstance(extra, dict):
            rec["extracted"].update({str(k)[:60]: str(v)[:200]
                                     for k, v in list(extra.items())[:20]})
        rec["missing_information"] = [str(x)[:200] for x in
                                      (ai.get("missing_information") or [])][:15]
        # Iter 683 — DATA ANALYSIS layer (Phase 1 spec).
        findings = [{"severity": str(f.get("severity") or "normal").lower(),
                     "message": str(f.get("message") or "")[:300]}
                    for f in (ai.get("findings") or []) if isinstance(f, dict)][:20]
        rec["email_vs_attachment"] = [
            {k: str(c.get(k) or "")[:120] for k in
             ("field", "email_value", "attachment_value", "severity")}
            for c in (ai.get("email_vs_attachment") or []) if isinstance(c, dict)][:15]
        for c in rec["email_vs_attachment"]:
            findings.append({"severity": c.get("severity") or "critical",
                             "message": f"MISMATCH {c['field']}: email says "
                                        f"{c['email_value']} but attachment says "
                                        f"{c['attachment_value']}"})
        stats = {"rows": 0, "blank_rows": 0, "duplicate_rows": 0}
        codes: list = []
        for a in rec["attachments"]:
            st_ = a.get("stats") or {}
            for k in stats:
                stats[k] += int(st_.get(k) or 0)
            codes.extend(a.get("codes") or [])
        matched = unmatched = 0
        if codes and rec["company_id"]:
            uniq = list({c for c in codes if c})[:300]
            found = set()
            async for u in db.users.find(
                    {"company_id": rec["company_id"],
                     "employee_code": {"$in": uniq}},
                    {"_id": 0, "employee_code": 1}):
                found.add(str(u.get("employee_code") or "").upper())
            matched = len(found)
            unmatched = max(0, len(uniq) - matched)
            # only meaningful when the sheet actually carries codes
            if matched and unmatched:
                findings.append({"severity": "high", "message":
                                 f"{unmatched} of {matched + unmatched} employee "
                                 f"codes in the attachment were NOT found in the "
                                 f"Employee Master (read-only check)"})
            _tl(rec, "Data Compared",
                f"Employee Master: {matched} matched · {unmatched} unmatched")
        if stats["duplicate_rows"]:
            findings.append({"severity": "warning", "message":
                             f"{stats['duplicate_rows']} duplicate row(s) in the sheet"})
        rec["data_analysis"] = {**stats, "employee_codes_seen": len(set(codes)),
                                "matched": matched, "unmatched": unmatched}
        rec["findings"] = findings
        if stats["rows"]:
            _tl(rec, "Data Validated",
                f"{stats['rows']} rows · {stats['blank_rows']} blank · "
                f"{stats['duplicate_rows']} duplicate")
        if findings:
            worst = ("critical" if any(f["severity"] == "critical" for f in findings)
                     else "high" if any(f["severity"] == "high" for f in findings)
                     else "warning")
            _tl(rec, "Exceptions Identified", f"{len(findings)} finding(s) · worst: {worst}")
        rec["confidences"] = {
            "company": int(ai.get("company_confidence") or 0),
            "classification": int(ai.get("classification_confidence") or 0),
            "extraction": int(ai.get("extraction_confidence") or 0),
            "recommendation": int(ai.get("recommendation_confidence") or 0),
        }
        _tl(rec, "Information Extracted",
            f"{len(rec['extracted'])} field(s), {len(rec['missing_information'])} missing")
        # Priority 3 — content-based company identification.
        if not rec["company_id"] and not match["review"] and rec["possible_company"]:
            conf = rec["confidences"]["company"]
            comp = await db.companies.find_one(
                {"name": {"$regex": f"^{re.escape(str(rec['possible_company']))}$",
                          "$options": "i"}},
                {"_id": 0, "company_id": 1, "name": 1})
            if comp and conf >= threshold:
                rec["company_id"] = comp["company_id"]
                rec["company_name"] = comp["name"]
                rec["company_match_type"] = "CONTENT_MATCH"
                rec["company_match_confidence"] = conf
                _tl(rec, "Company Matched",
                    f"{comp['name']} · CONTENT_MATCH {conf}%")
        # Primary status decision.
        if match["review"] or not rec["company_id"]:
            rec["status"] = "COMPANY_REVIEW_REQUIRED"
        elif rec["priority"] == "urgent":
            rec["status"] = "URGENT"
        elif rec["confidences"]["classification"] < threshold:
            rec["status"] = "REVIEW_REQUIRED"
        elif ai.get("action_required"):
            rec["status"] = "ACTION_REQUIRED"
        else:
            rec["status"] = "INFORMATION_ONLY"
        # Iter 683 — a CRITICAL data finding escalates the email.
        if any(f["severity"] == "critical" for f in findings) and \
                rec["status"] in ("INFORMATION_ONLY", "ACTION_REQUIRED", "REVIEW_REQUIRED"):
            rec["status"] = "URGENT"
        _tl(rec, "Audit Completed", rec["status"])
        _tl(rec, "Recommendation Generated", rec["ai_recommendation"][:120])
    except Exception as exc:
        rec["status"] = "PROCESSING_FAILED"
        rec["error"] = str(exc)[:400]
        _tl(rec, "Processing Failed", rec["error"])
    rec["processing_completed_at"] = now_iso()
    # Notification (in-app only; suppressed in sandbox).
    if not sandbox and rec["status"] in (
            "ACTION_REQUIRED", "URGENT", "COMPANY_REVIEW_REQUIRED"):
        who = rec["company_name"] or rec["sender_name"] or rec["sender_email"]
        await notify_emit(
            db, audience="super_admins", category="system",
            priority="important" if rec["status"] != "URGENT" else "critical",
            title=f"🤖 AI Email Audit — {rec['status'].replace('_', ' ').title()}",
            message=f"{who}: {rec['ai_summary'][:200] or rec['subject'][:200]}",
            action_url="/ai-command-center?tab=email-audit",
            reference_id=rec["audit_id"])
        _tl(rec, "Notification Generated", rec["status"])
    await db.email_audit_records.insert_one(dict(rec))
    rec.pop("_id", None)
    return rec


async def _run_scan(triggered_by: str) -> dict:
    st = await _state()
    settings = None
    try:
        from routes.gmail_mailbox import _smtp_settings
        settings = await _smtp_settings()
    except Exception:
        pass
    if not settings:
        res = {"ok": False, "error": "Email SMTP & Notifications is not configured/enabled",
               "at": now_iso(), "by": triggered_by}
        await db.email_agent_state.update_one(
            {"_singleton": True}, {"$set": {"last_scan_at": now_iso(),
                                            "last_scan_result": res}}, upsert=True)
        return res
    known = set()
    async for d in db.email_audit_records.find({}, {"_id": 0, "message_id": 1}):
        if d.get("message_id"):
            known.add(d["message_id"])
    # Iter 680 — registered company emails (needed for the Spam sweep).
    registered = set()
    async for r in db.company_email_registry.find(
            {"active": {"$ne": False}}, {"_id": 0, "email": 1}):
        if r.get("email"):
            registered.add(r["email"].lower())
    try:
        parsed_list, seen, historical = await asyncio.to_thread(
            _imap_fetch_new_sync, settings, known, registered)
    except Exception as exc:
        res = {"ok": False, "error": f"IMAP error: {str(exc)[:200]}",
               "at": now_iso(), "by": triggered_by}
        await db.email_agent_state.update_one(
            {"_singleton": True}, {"$set": {"last_scan_at": now_iso(),
                                            "last_scan_result": res}}, upsert=True)
        return res
    processed = failed = 0
    for p in parsed_list:
        if p["message_id"] in known:
            continue  # ALREADY_PROCESSED — duplicate protection
        known.add(p["message_id"])
        rec = await _process_email(p, sandbox=bool(st.get("sandbox")),
                                   threshold=int(st.get("threshold") or 80))
        processed += 1
        if rec["status"] == "PROCESSING_FAILED":
            failed += 1
    res = {"ok": True, "headers_seen": seen, "new_processed": processed,
           "failed": failed, "ignored_historical": historical,
           "at": now_iso(), "by": triggered_by}
    await db.email_agent_state.update_one(
        {"_singleton": True}, {"$set": {"last_scan_at": now_iso(),
                                        "last_scan_result": res}}, upsert=True)
    return res


async def email_agent_loop():
    """Background poller — every `poll_minutes` while enabled."""
    await asyncio.sleep(45)  # let the app settle after startup
    while True:
        try:
            st = await _state()
            if st.get("enabled"):
                await _run_scan("auto-poll")
            wait = max(2, int(st.get("poll_minutes") or 5)) * 60
        except Exception:
            logger.exception("[email-agent] poll failed")
            wait = 300
        await asyncio.sleep(wait)


# ───────────────────────── API models ───────────────────────────────────────

class SettingsIn(BaseModel):
    enabled: Optional[bool] = None
    sandbox: Optional[bool] = None
    threshold: Optional[int] = None
    poll_minutes: Optional[int] = None


class RegistryIn(BaseModel):
    company_id: str
    email: str
    email_type: Optional[str] = "general"
    contact_person: Optional[str] = None
    department: Optional[str] = None


class AssignCompanyIn(BaseModel):
    company_id: str


class SandboxIngestIn(BaseModel):
    sender_email: str
    sender_name: Optional[str] = ""
    subject: str
    body: str
    received_at: Optional[str] = None
    # Iter 685 — optional document photo (base64 jpg/png) to test the OCR
    # scanner end-to-end in sandbox mode.
    attachment_name: Optional[str] = None
    attachment_b64: Optional[str] = None


# ───────────────────────── endpoints (super admin only) ─────────────────────

OPERATIONAL = {"$nin": ["IGNORED_HISTORICAL"]}


@router.get("/settings")
async def get_settings(authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    st = await _state()
    st.pop("_singleton", None)
    smtp_ok = False
    try:
        from routes.gmail_mailbox import _smtp_settings
        smtp_ok = bool(await _smtp_settings())
    except Exception:
        pass
    return {**st, "smtp_configured": smtp_ok, "cutoff": "2026-08-15"}


@router.post("/settings")
async def set_settings(payload: SettingsIn,
                       authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    upd = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "threshold" in upd:
        upd["threshold"] = min(99, max(50, int(upd["threshold"])))
    if "poll_minutes" in upd:
        upd["poll_minutes"] = min(120, max(2, int(upd["poll_minutes"])))
    if upd:
        await db.email_agent_state.update_one(
            {"_singleton": True}, {"$set": upd}, upsert=True)
    return await get_settings(authorization)


@router.post("/scan")
async def scan_now(authorization: Optional[str] = Header(None)):
    admin = await _sa(authorization)
    return await _run_scan(f"manual:{admin.get('name') or admin['user_id']}")


@router.get("/registry")
async def list_registry(company_id: Optional[str] = Query(None),
                        authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    q: dict = {}
    if company_id:
        q["company_id"] = company_id
    rows = await db.company_email_registry.find(q, {"_id": 0}).sort(
        "created_at", -1).to_list(500)
    return {"entries": rows}


@router.post("/registry")
async def add_registry(payload: RegistryIn,
                       authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    email_l = payload.email.strip().lower()
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email_l):
        raise HTTPException(status_code=400, detail="Invalid email address")
    comp = await db.companies.find_one({"company_id": payload.company_id},
                                       {"_id": 0, "company_id": 1, "name": 1})
    if not comp:
        raise HTTPException(status_code=404, detail="Company not found")
    dup = await db.company_email_registry.find_one(
        {"company_id": comp["company_id"], "email": email_l})
    if dup:
        raise HTTPException(status_code=409,
                            detail="Email already registered for this company")
    row = {
        "registry_id": f"cer_{uuid.uuid4().hex[:10]}",
        "company_id": comp["company_id"], "company_name": comp["name"],
        "email": email_l, "email_type": (payload.email_type or "general")[:40],
        "contact_person": (payload.contact_person or "")[:80],
        "department": (payload.department or "")[:80],
        "active": True, "created_at": now_iso(),
    }
    await db.company_email_registry.insert_one(dict(row))
    row.pop("_id", None)
    return {"ok": True, "entry": row}


@router.patch("/registry/{registry_id}")
async def toggle_registry(registry_id: str,
                          authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    row = await db.company_email_registry.find_one({"registry_id": registry_id})
    if not row:
        raise HTTPException(status_code=404, detail="Entry not found")
    new_active = not bool(row.get("active", True))
    await db.company_email_registry.update_one(
        {"registry_id": registry_id}, {"$set": {"active": new_active}})
    return {"ok": True, "active": new_active}


@router.delete("/registry/{registry_id}")
async def delete_registry(registry_id: str,
                          authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    r = await db.company_email_registry.delete_one({"registry_id": registry_id})
    if not r.deleted_count:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"ok": True}


@router.get("/dashboard")
async def dashboard(authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    base = {"status": OPERATIONAL}
    total = await db.email_audit_records.count_documents(base)
    by_status: dict = {}
    async for d in db.email_audit_records.aggregate([
            {"$match": base}, {"$group": {"_id": "$status", "n": {"$sum": 1}}}]):
        by_status[d["_id"]] = d["n"]
    with_att = await db.email_audit_records.count_documents(
        {**base, "has_attachments": True})
    cat_counts: dict = {}
    async for d in db.email_audit_records.aggregate([
            {"$match": base}, {"$unwind": "$categories"},
            {"$group": {"_id": "$categories", "n": {"$sum": 1}}}]):
        cat_counts[d["_id"]] = d["n"]
    payroll_cats = {"Salary Processing", "PF / EPF", "ESIC", "Overtime",
                    "Advance", "Deduction", "Payroll Query"}
    employee_cats = {"New Employee", "Employee Master Update", "Employee Exit",
                     "Attendance", "Leave"}
    compliance_cats = {"Labour Compliance", "Compliance Documents"}
    group = {"payroll": 0, "employee": 0, "compliance": 0, "general": 0}
    for c, n in cat_counts.items():
        if c in payroll_cats:
            group["payroll"] += n
        elif c in employee_cats:
            group["employee"] += n
        elif c in compliance_cats:
            group["compliance"] += n
        else:
            group["general"] += n
    st = await _state()
    return {
        "total": total, "by_status": by_status, "by_category": cat_counts,
        "groups": group, "with_attachments": with_att,
        "window": {"from": "2026-08-15", "to": datetime.now(IST).strftime("%Y-%m-%d")},
        "last_scan_at": st.get("last_scan_at"),
        "last_scan_result": st.get("last_scan_result"),
        "enabled": st.get("enabled"), "sandbox": st.get("sandbox"),
    }


@router.get("/emails")
async def list_emails(status: Optional[str] = Query(None),
                      category: Optional[str] = Query(None),
                      company_id: Optional[str] = Query(None),
                      q: Optional[str] = Query(None),
                      limit: int = Query(50, le=200),
                      skip: int = Query(0, ge=0),
                      authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    qq: dict = {"status": OPERATIONAL}
    if status:
        qq["status"] = status
    if category:
        qq["categories"] = category
    if company_id:
        qq["company_id"] = company_id
    if q:
        rx = {"$regex": re.escape(q), "$options": "i"}
        qq["$or"] = [{"subject": rx}, {"sender_email": rx},
                     {"sender_name": rx}, {"ai_summary": rx}]
    rows = await db.email_audit_records.find(
        qq, {"_id": 0, "body_text": 0, "timeline": 0}).sort(
        "received_at", -1).skip(skip).to_list(limit)
    total = await db.email_audit_records.count_documents(qq)
    return {"emails": rows, "total": total}


@router.get("/emails/{audit_id}")
async def email_detail(audit_id: str,
                       authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    row = await db.email_audit_records.find_one({"audit_id": audit_id}, {"_id": 0})
    if not row:
        raise HTTPException(status_code=404, detail="Audit record not found")
    return row


@router.post("/emails/{audit_id}/assign-company")
async def assign_company(audit_id: str, payload: AssignCompanyIn,
                         authorization: Optional[str] = Header(None)):
    admin = await _sa(authorization)
    row = await db.email_audit_records.find_one({"audit_id": audit_id})
    if not row:
        raise HTTPException(status_code=404, detail="Audit record not found")
    comp = await db.companies.find_one({"company_id": payload.company_id},
                                       {"_id": 0, "company_id": 1, "name": 1})
    if not comp:
        raise HTTPException(status_code=404, detail="Company not found")
    new_status = row["status"]
    if new_status == "COMPANY_REVIEW_REQUIRED":
        new_status = "URGENT" if row.get("priority") == "urgent" else "REVIEW_REQUIRED"
    tl_entry = {"step": "Company Manually Assigned",
                "detail": f"{comp['name']} by {admin.get('name') or admin['user_id']}",
                "at": now_iso()}
    await db.email_audit_records.update_one(
        {"audit_id": audit_id},
        {"$set": {"company_id": comp["company_id"], "company_name": comp["name"],
                  "company_match_type": "MANUAL", "company_match_confidence": 100,
                  "status": new_status},
         "$push": {"timeline": tl_entry}})
    return {"ok": True, "status": new_status, "company_name": comp["name"]}


@router.get("/company-summary")
async def company_summary(authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    rows = []
    async for d in db.email_audit_records.aggregate([
            {"$match": {"status": OPERATIONAL}},
            {"$group": {"_id": {"cid": "$company_id", "name": "$company_name"},
                        "total": {"$sum": 1},
                        "action": {"$sum": {"$cond": [
                            {"$in": ["$status", ["ACTION_REQUIRED", "URGENT"]]}, 1, 0]}},
                        "review": {"$sum": {"$cond": [
                            {"$in": ["$status", ["REVIEW_REQUIRED",
                                                 "COMPANY_REVIEW_REQUIRED"]]}, 1, 0]}}}},
            {"$sort": {"total": -1}}]):
        rows.append({"company_id": d["_id"].get("cid"),
                     "company_name": d["_id"].get("name") or "Unknown",
                     "total": d["total"], "action_required": d["action"],
                     "review_required": d["review"]})
    return {"companies": rows}


@router.get("/daily-report")
async def daily_report(date: Optional[str] = Query(None),
                       authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    day = date or datetime.now(IST).strftime("%Y-%m-%d")
    start = datetime.strptime(day, "%Y-%m-%d").replace(tzinfo=IST)
    end = start + timedelta(days=1)
    qq = {"status": OPERATIONAL,
          "received_at": {"$gte": start.astimezone(timezone.utc).isoformat(),
                          "$lt": end.astimezone(timezone.utc).isoformat()}}
    rows = await db.email_audit_records.find(
        qq, {"_id": 0, "body_text": 0, "timeline": 0}).sort(
        "received_at", -1).to_list(300)
    by_status: dict = {}
    by_company: dict = {}
    for r in rows:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
        key = r.get("company_name") or "Unknown"
        by_company[key] = by_company.get(key, 0) + 1
    important = [r for r in rows if r["status"] in
                 ("URGENT", "ACTION_REQUIRED", "COMPANY_REVIEW_REQUIRED")]
    pending = [r for r in rows if r["status"] in
               ("ACTION_REQUIRED", "REVIEW_REQUIRED",
                "COMPANY_REVIEW_REQUIRED", "URGENT")]
    return {"date": day, "total": len(rows), "by_status": by_status,
            "by_company": sorted(
                [{"company": k, "count": v} for k, v in by_company.items()],
                key=lambda x: -x["count"]),
            "important": important[:20], "pending": pending[:30]}


@router.get("/exceptions")
async def exceptions(authorization: Optional[str] = Header(None)):
    await _sa(authorization)
    rows = await db.email_audit_records.find(
        {"status": "PROCESSING_FAILED"},
        {"_id": 0, "body_text": 0, "timeline": 0}).sort(
        "received_at", -1).to_list(100)
    return {"exceptions": rows}


@router.post("/sandbox-ingest")
async def sandbox_ingest(payload: SandboxIngestIn,
                         authorization: Optional[str] = Header(None)):
    """Sandbox/Test mode — feed ONE synthetic email through the full audit
    pipeline (company match + AI). Never touches the live mailbox or payroll."""
    await _sa(authorization)
    st = await _state()
    if not st.get("sandbox"):
        raise HTTPException(status_code=400,
                            detail="Enable Sandbox mode first (Settings)")
    parsed = {
        "message_id": f"<sandbox-{uuid.uuid4().hex}@test.local>",
        "thread_id": "", "sender_name": payload.sender_name or "",
        "sender_email": payload.sender_email.strip().lower(),
        "recipient": "audit@test.local", "cc": "",
        "subject": payload.subject[:300],
        "received_at": payload.received_at or datetime.now(timezone.utc).isoformat(),
        "body_text": payload.body[:20000], "attachments": [],
    }
    # Iter 685 — sandbox OCR test: attach one document photo.
    if payload.attachment_b64 and payload.attachment_name:
        try:
            _sz = len(base64.b64decode(payload.attachment_b64))
        except Exception:
            raise HTTPException(status_code=400, detail="attachment_b64 is not valid base64")
        parsed["attachments"] = [{
            "name": payload.attachment_name[:120], "type": "image/jpeg",
            "size": _sz, "readable": False, "excerpt": "",
            "note": "Document image — OCR scan queued",
            "_b64": payload.attachment_b64,
        }]
    rec = await _process_email(parsed, sandbox=True,
                               threshold=int(st.get("threshold") or 80))
    return {"ok": True, "record": rec}
