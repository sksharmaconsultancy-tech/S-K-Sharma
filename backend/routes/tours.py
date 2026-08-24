"""Iter 706 — OFFICIAL TOUR MANAGEMENT module.

Tour Request → configurable Approval Workflow → Start Tour (GPS) →
Live Tracking (offline-queue aware) → Client Visits → End Tour →
Expense Claim link (same Tour ID) → OD/TOUR Attendance posting →
Payroll (via attendance punches) → full Audit Trail.

Collections: tour_requests, tour_tracking_logs, tour_visits,
tour_attendance, tour_settings, tour_audit, tour_counters.
Reuses: users, companies, attendance, expense_claims,
approval_workflows / approval_requests (engine module key: "tour").
"""
import base64
import uuid
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query
from fastapi.responses import Response

from server import db, get_user_from_token, require_role, now_iso, logger  # noqa: E402

router = APIRouter(prefix="/api/tours", tags=["tours"])

ADMIN_ROLES = ["super_admin", "sub_admin", "company_admin"]
TOUR_TYPES = ["Official Tour", "Client Visit", "Business Development",
              "Training", "Meeting", "Other"]
STATUSES = ["draft", "submitted", "pending_approval", "approved", "active",
            "completed", "returned", "rejected", "cancelled"]
ALLOWED_MIME = {"application/pdf": ".pdf", "image/jpeg": ".jpg", "image/png": ".png"}
MAX_FILE_MB = 5

DEFAULT_SETTINGS = {
    "tracking_interval_min": 5,          # 1 / 5 / 10 / 15
    "od_counts_present": True,           # OD counts as Present
    "od_counts_paid": True,              # OD counts as Paid Day
    "od_ot_eligible": False,             # OD eligible for OT
    "holiday_during_tour": "skip",       # "skip" | "od"
    "weekly_off_during_tour": "skip",    # "skip" | "od"
    "half_day_od_hours": 4,              # duty hours posted for Half Day OD
    "expense_claim_grace_days": 7,       # claims allowed till end_date + N
}


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
async def _taudit(user: dict, tour_id: str, action: str, old: Any = None,
                  new: Any = None, remarks: str = "", gps: Optional[dict] = None,
                  device: Optional[dict] = None, stage: Optional[str] = None):
    try:
        await db.tour_audit.insert_one({
            "audit_id": f"tau_{uuid.uuid4().hex[:12]}", "tour_id": tour_id,
            "user_id": user.get("user_id"), "name": user.get("name"),
            "role": user.get("role"), "action": action, "old": old, "new": new,
            "remarks": remarks, "gps": gps, "device": device,
            "approval_stage": stage, "at": now_iso()})
    except Exception:
        logger.exception("[tour] audit failed")


async def _next_tour_no(company_id: str) -> str:
    year = datetime.now(timezone.utc).year
    doc = await db.tour_counters.find_one_and_update(
        {"company_id": company_id, "year": year},
        {"$inc": {"seq": 1}}, upsert=True, return_document=True)
    return f"TOUR-{year}-{int(doc['seq']):06d}"


async def get_tour_settings(company_id: str) -> dict:
    s = await db.tour_settings.find_one({"company_id": company_id}, {"_id": 0}) or {}
    return {**DEFAULT_SETTINGS, **{k: v for k, v in s.items() if k in DEFAULT_SETTINGS}}


def _tour_days(start_date: str, end_date: str) -> int:
    try:
        d0 = date.fromisoformat(str(start_date)[:10])
        d1 = date.fromisoformat(str(end_date)[:10])
        return max(1, (d1 - d0).days + 1)
    except (ValueError, TypeError):
        return 1


def _notify(company_id: str, target_user_id: str, title: str, body: str, tour_id: str):
    return db.notifications.insert_one({
        "notification_id": f"n_{uuid.uuid4().hex[:10]}", "company_id": company_id,
        "audience": "user", "target_user_id": target_user_id,
        "type": "tour", "title": title, "body": body,
        "meta": {"tour_id": tour_id}, "created_at": now_iso()})


async def _notify_admins(company_id: str, title: str, body: str, tour_id: str):
    async for u in db.users.find(
            {"company_id": company_id, "role": "company_admin"},
            {"_id": 0, "user_id": 1}).limit(10):
        try:
            await _notify(company_id, u["user_id"], title, body, tour_id)
        except Exception:
            pass


async def _get_tour_scoped(tour_id: str, user: dict) -> dict:
    t = await db.tour_requests.find_one({"tour_id": tour_id}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Tour not found")
    if user.get("role") in ("super_admin", "sub_admin"):
        return t
    if user.get("company_id") != t.get("company_id"):
        raise HTTPException(status_code=403, detail="Not your firm")
    if user.get("role") == "employee" and t.get("user_id") != user.get("user_id"):
        raise HTTPException(status_code=403, detail="Not your tour")
    return t


def _clean_payload(payload: dict) -> dict:
    """Whitelisted, typed tour fields from the request body."""
    dests = payload.get("destinations")
    if not isinstance(dests, list):
        dests = [d.strip() for d in str(payload.get("destinations") or "").split(",") if d.strip()]
    dests = [str(d)[:120] for d in dests][:10]
    fin_keys = ("est_travel", "est_food", "est_accommodation", "est_other")
    fin = {}
    for k in fin_keys:
        try:
            fin[k] = max(0.0, float(payload.get(k) or 0))
        except (TypeError, ValueError):
            fin[k] = 0.0
    total_est = round(sum(fin.values()), 2)
    adv_req = bool(payload.get("advance_required"))
    try:
        adv_amt = max(0.0, float(payload.get("advance_amount") or 0)) if adv_req else 0.0
    except (TypeError, ValueError):
        adv_amt = 0.0
    tour_type = payload.get("tour_type")
    if tour_type not in TOUR_TYPES:
        tour_type = "Official Tour"
    start_date = str(payload.get("start_date") or "")[:10]
    end_date = str(payload.get("end_date") or "")[:10]
    return {
        "tour_type": tour_type,
        "start_date": start_date, "start_time": str(payload.get("start_time") or "")[:5],
        "end_date": end_date, "end_time": str(payload.get("end_time") or "")[:5],
        "total_days": _tour_days(start_date, end_date),
        "from_location": str(payload.get("from_location") or "")[:120],
        "destinations": dests,
        "client_name": str(payload.get("client_name") or "")[:150],
        "contact_person": str(payload.get("contact_person") or "")[:100],
        "contact_number": str(payload.get("contact_number") or "")[:20],
        "meeting_purpose": str(payload.get("meeting_purpose") or "")[:400],
        "expected_outcome": str(payload.get("expected_outcome") or "")[:400],
        "purpose": str(payload.get("purpose") or "")[:400],
        "remarks": str(payload.get("remarks") or "")[:400],
        **fin, "total_estimated": total_est,
        "advance_required": adv_req, "advance_amount": adv_amt,
        # Attendance requirement (posted ONLY after final approval)
        "mark_od": bool(payload.get("mark_od")),
        "od_day_type": "half" if payload.get("od_day_type") == "half" else "full",
    }


def _od_preview(t: dict) -> List[dict]:
    out = []
    try:
        d0 = date.fromisoformat(t["start_date"])
        d1 = date.fromisoformat(t["end_date"])
    except (ValueError, TypeError, KeyError):
        return out
    cur = d0
    while cur <= d1 and len(out) < 62:
        out.append({"date": cur.isoformat(), "type": "OD",
                    "day": "Half Day" if t.get("od_day_type") == "half" else "Full Day"})
        cur += timedelta(days=1)
    return out


# ---------------------------------------------------------------------------
# Employee — My Tours
# ---------------------------------------------------------------------------
@router.get("/mine")
async def my_tours(authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    rows = await db.tour_requests.find(
        {"user_id": user["user_id"]},
        {"_id": 0, "attachments.data_b64": 0}).sort("created_at", -1).to_list(200)
    counts: Dict[str, int] = {s: 0 for s in STATUSES}
    for r in rows:
        counts[r.get("status", "draft")] = counts.get(r.get("status", "draft"), 0) + 1
    counts["total"] = len(rows)
    settings = await get_tour_settings(user.get("company_id") or "")
    return {"tours": rows, "counts": counts,
            "tracking_interval_min": settings["tracking_interval_min"],
            "tour_types": TOUR_TYPES}


@router.post("")
async def create_tour(payload: Dict[str, Any] = Body(...),
                      authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    cid = user.get("company_id")
    if not cid:
        raise HTTPException(status_code=400, detail="No firm linked to your account")
    fields = _clean_payload(payload)
    if not fields["start_date"] or not fields["end_date"]:
        raise HTTPException(status_code=400, detail="Start and End dates are required")
    if fields["end_date"] < fields["start_date"]:
        raise HTTPException(status_code=400, detail="End date cannot be before start date")
    if not fields["destinations"]:
        raise HTTPException(status_code=400, detail="At least one destination is required")
    t = {
        "tour_id": f"tr_{uuid.uuid4().hex[:12]}",
        "tour_no": await _next_tour_no(cid),
        "company_id": cid, "user_id": user["user_id"],
        "employee": {k: user.get(k) for k in
                     ("name", "employee_code", "department", "designation",
                      "branch", "reporting_manager")},
        **fields,
        "status": "draft", "attachments": [],
        "approval_request_id": None, "approval_history": [],
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.tour_requests.insert_one(dict(t))
    t.pop("_id", None)
    await _taudit(user, t["tour_id"], "created", None,
                  {"tour_no": t["tour_no"], "type": t["tour_type"]})
    return {"ok": True, "tour": t}


@router.put("/{tour_id}")
async def update_tour(tour_id: str, payload: Dict[str, Any] = Body(...),
                      authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    if t["status"] not in ("draft", "returned"):
        raise HTTPException(status_code=400,
                            detail=f"A {t['status']} tour cannot be edited")
    fields = _clean_payload(payload)
    fields["updated_at"] = now_iso()
    await db.tour_requests.update_one({"tour_id": tour_id}, {"$set": fields})
    await _taudit(user, tour_id, "updated",
                  {k: t.get(k) for k in fields if t.get(k) != fields[k] and k != "updated_at"},
                  {k: v for k, v in fields.items() if t.get(k) != v and k != "updated_at"})
    fresh = await db.tour_requests.find_one({"tour_id": tour_id}, {"_id": 0})
    return {"ok": True, "tour": fresh}


@router.post("/{tour_id}/submit")
async def submit_tour(tour_id: str, authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    if t["status"] not in ("draft", "returned"):
        raise HTTPException(status_code=400, detail=f"Cannot submit a {t['status']} tour")
    from routes.approvals_engine import get_active_workflow, create_approval_request
    wf = await get_active_workflow(t["company_id"], "tour")
    upd = {"status": "pending_approval", "submitted_at": now_iso(), "updated_at": now_iso()}
    if wf:
        req = await create_approval_request(
            t["company_id"], "tour", tour_id,
            f"{t['tour_no']} — {t['tour_type']} to {', '.join(t.get('destinations') or [])[:60]}",
            {"tour_type": t["tour_type"], "amount": t.get("total_estimated") or 0,
             "days": t.get("total_days") or 1,
             "department": (t.get("employee") or {}).get("department"),
             "designation": (t.get("employee") or {}).get("designation"),
             "destination": ", ".join(t.get("destinations") or []),
             "employee_code": (t.get("employee") or {}).get("employee_code")},
            user, wf)
        upd["approval_request_id"] = req["request_id"]
    await db.tour_requests.update_one({"tour_id": tour_id}, {"$set": upd})
    await _taudit(user, tour_id, "submitted", t["status"], "pending_approval")
    if not wf:
        await _notify_admins(t["company_id"], f"Tour request {t['tour_no']}",
                             f"{user.get('name')} requested a {t['tour_type']}", tour_id)
    return {"ok": True, "status": "pending_approval", "workflow": bool(wf)}


@router.post("/{tour_id}/cancel")
async def cancel_tour(tour_id: str, payload: Dict[str, Any] = Body(default={}),
                      authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    is_admin = user.get("role") in ADMIN_ROLES
    if t["status"] in ("completed", "cancelled", "rejected"):
        raise HTTPException(status_code=400, detail=f"A {t['status']} tour cannot be cancelled")
    if t["status"] == "active" and not is_admin:
        raise HTTPException(status_code=403,
                            detail="An active tour can only be cancelled by an authorized admin")
    await db.tour_requests.update_one(
        {"tour_id": tour_id},
        {"$set": {"status": "cancelled", "cancelled_at": now_iso(),
                  "cancelled_by": user["user_id"], "updated_at": now_iso(),
                  "tracking_active": False}})
    await _taudit(user, tour_id, "cancelled", t["status"], "cancelled",
                  remarks=str(payload.get("remarks") or ""))
    await _cancel_tour_attendance(t, user, reason="Tour cancelled")
    return {"ok": True, "status": "cancelled"}


# ---------------------------------------------------------------------------
# Workflow finalize hook (called from approvals_engine._finalize)
# ---------------------------------------------------------------------------
async def finalize_tour_approval(tour_id: str, final_status: str, actor: dict):
    """final_status: approved | rejected | returned."""
    t = await db.tour_requests.find_one({"tour_id": tour_id}, {"_id": 0})
    if not t or t.get("status") not in ("pending_approval", "submitted"):
        return
    new_status = {"approved": "approved", "rejected": "rejected",
                  "returned": "returned"}.get(final_status)
    if not new_status:
        return
    upd = {"status": new_status, "updated_at": now_iso()}
    if new_status == "approved":
        upd["approved_at"] = now_iso()
        upd["approved_by"] = actor.get("user_id")
        upd["approved_by_name"] = actor.get("name") or actor.get("email")
        # Iter 707 — approved advance becomes a payable entry for accounts.
        if t.get("advance_required") and float(t.get("advance_amount") or 0) > 0:
            upd["advance_payout"] = {"status": "pending",
                                     "amount": float(t["advance_amount"]),
                                     "created_at": now_iso()}
    await db.tour_requests.update_one({"tour_id": tour_id}, {"$set": upd})
    await _taudit(actor, tour_id, f"workflow_{final_status}",
                  "pending_approval", new_status, stage="final")
    try:
        await _notify(t["company_id"], t["user_id"],
                      f"Tour {t['tour_no']} {new_status.upper()}",
                      f"Your {t.get('tour_type')} request is {new_status}", tour_id)
    except Exception:
        pass
    if new_status == "approved" and t.get("mark_od"):
        await post_tour_attendance({**t, **upd}, actor)


@router.post("/{tour_id}/decide")
async def decide_tour(tour_id: str, payload: Dict[str, Any] = Body(...),
                      authorization: Optional[str] = Header(None)):
    """Direct admin decision when NO approval workflow is configured for
    the firm (fallback path). action: approve|reject|return."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    t = await _get_tour_scoped(tour_id, admin)
    if t["status"] != "pending_approval":
        raise HTTPException(status_code=400, detail=f"Tour is {t['status']}, not pending")
    if t.get("approval_request_id"):
        raise HTTPException(status_code=400,
                            detail="This tour is routed through the Approval Workflow — use the Approval Inbox")
    if t.get("user_id") == admin.get("user_id"):
        raise HTTPException(status_code=403, detail="Maker-checker: you cannot approve your own tour")
    action = (payload.get("action") or "").strip()
    remarks = str(payload.get("remarks") or "").strip()
    if action not in ("approve", "reject", "return"):
        raise HTTPException(status_code=400, detail="action must be approve|reject|return")
    if action in ("reject", "return") and not remarks:
        raise HTTPException(status_code=400, detail="Remarks are mandatory to reject or return")
    final = {"approve": "approved", "reject": "rejected", "return": "returned"}[action]
    await db.tour_requests.update_one(
        {"tour_id": tour_id},
        {"$push": {"approval_history": {
            "action": final, "by": admin["user_id"],
            "by_name": admin.get("name") or admin.get("email"),
            "remarks": remarks or None, "at": now_iso()}}})
    await finalize_tour_approval(tour_id, final, admin)
    fresh = await db.tour_requests.find_one({"tour_id": tour_id}, {"_id": 0})
    return {"ok": True, "tour": fresh}


# ---------------------------------------------------------------------------
# Start / End / Live tracking
# ---------------------------------------------------------------------------
def _gps_from(payload: dict) -> Optional[dict]:
    try:
        lat = float(payload.get("lat"))
        lng = float(payload.get("lng"))
    except (TypeError, ValueError):
        return None
    g = {"lat": lat, "lng": lng}
    try:
        g["accuracy"] = round(float(payload.get("accuracy")), 1)
    except (TypeError, ValueError):
        pass
    return g


@router.post("/{tour_id}/start")
async def start_tour(tour_id: str, payload: Dict[str, Any] = Body(default={}),
                     authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    if t.get("user_id") != user.get("user_id"):
        raise HTTPException(status_code=403, detail="Only the tour owner can start the tour")
    if t["status"] != "approved":
        raise HTTPException(status_code=400,
                            detail="Tour must be APPROVED before it can be started")
    gps = _gps_from(payload)
    device = {k: str(payload.get(k) or "")[:200] for k in ("device", "platform", "user_agent")}
    upd = {"status": "active", "started_at": now_iso(), "start_gps": gps,
           "start_device": device, "tracking_active": True, "updated_at": now_iso()}
    await db.tour_requests.update_one({"tour_id": tour_id}, {"$set": upd})
    if gps:
        await db.tour_tracking_logs.insert_one({
            "log_id": f"trk_{uuid.uuid4().hex[:12]}", "tour_id": tour_id,
            "company_id": t["company_id"], "user_id": user["user_id"],
            "lat": gps["lat"], "lng": gps["lng"], "accuracy": gps.get("accuracy"),
            "captured_at": now_iso(), "synced_at": now_iso(),
            "offline": False, "event": "tour_started"})
    await _taudit(user, tour_id, "tour_started", "approved", "active",
                  gps=gps, device=device)
    return {"ok": True, "status": "active", "tracking_active": True}


@router.post("/{tour_id}/track")
async def track_tour(tour_id: str, payload: Dict[str, Any] = Body(...),
                     authorization: Optional[str] = Header(None)):
    """Batch GPS sync. Body: {points: [{lat,lng,accuracy,captured_at,offline}]}.
    Original capture time is ALWAYS retained; synced_at = server now."""
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    if t.get("user_id") != user.get("user_id"):
        raise HTTPException(status_code=403, detail="Not your tour")
    if t["status"] != "active":
        return {"ok": True, "stored": 0, "tracking_active": False,
                "reason": f"tour is {t['status']}"}
    # Auto-stop when the approved end datetime has passed (+6h grace).
    try:
        end_dt = datetime.fromisoformat(f"{t['end_date']}T{t.get('end_time') or '23:59'}:00+00:00")
        if datetime.now(timezone.utc) > end_dt + timedelta(hours=6):
            await db.tour_requests.update_one(
                {"tour_id": tour_id}, {"$set": {"tracking_active": False}})
            return {"ok": True, "stored": 0, "tracking_active": False,
                    "reason": "approved tour period is over"}
    except (ValueError, TypeError):
        pass
    pts = payload.get("points") or []
    docs = []
    for p in pts[:500]:
        gps = _gps_from(p)
        if not gps:
            continue
        docs.append({
            "log_id": f"trk_{uuid.uuid4().hex[:12]}", "tour_id": tour_id,
            "company_id": t["company_id"], "user_id": user["user_id"],
            "lat": gps["lat"], "lng": gps["lng"], "accuracy": gps.get("accuracy"),
            "captured_at": str(p.get("captured_at") or now_iso()),
            "synced_at": now_iso(), "offline": bool(p.get("offline")),
            "event": "track"})
    if docs:
        await db.tour_tracking_logs.insert_many(docs)
        last = docs[-1]
        await db.tour_requests.update_one(
            {"tour_id": tour_id},
            {"$set": {"last_location": {
                "lat": last["lat"], "lng": last["lng"],
                "accuracy": last.get("accuracy"),
                "captured_at": last["captured_at"], "synced_at": last["synced_at"]}}})
    settings = await get_tour_settings(t["company_id"])
    return {"ok": True, "stored": len(docs), "tracking_active": True,
            "interval_min": settings["tracking_interval_min"]}


@router.post("/{tour_id}/end")
async def end_tour(tour_id: str, payload: Dict[str, Any] = Body(default={}),
                   authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    is_admin = user.get("role") in ADMIN_ROLES
    if t.get("user_id") != user.get("user_id") and not is_admin:
        raise HTTPException(status_code=403, detail="Not your tour")
    if t["status"] != "active":
        raise HTTPException(status_code=400, detail=f"Tour is {t['status']}, not active")
    gps = _gps_from(payload)
    upd = {"status": "completed", "ended_at": now_iso(), "end_gps": gps,
           "end_remarks": str(payload.get("remarks") or "")[:400],
           "tracking_active": False, "updated_at": now_iso()}
    await db.tour_requests.update_one({"tour_id": tour_id}, {"$set": upd})
    if gps:
        await db.tour_tracking_logs.insert_one({
            "log_id": f"trk_{uuid.uuid4().hex[:12]}", "tour_id": tour_id,
            "company_id": t["company_id"], "user_id": user["user_id"],
            "lat": gps["lat"], "lng": gps["lng"], "accuracy": gps.get("accuracy"),
            "captured_at": now_iso(), "synced_at": now_iso(),
            "offline": False, "event": "tour_ended"})
    await _taudit(user, tour_id, "tour_ended", "active", "completed", gps=gps)
    return {"ok": True, "status": "completed"}


@router.get("/{tour_id}/summary")
async def tour_summary(tour_id: str, authorization: Optional[str] = Header(None)):
    """Pre-End-Tour summary + post-completion recap."""
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    visits = await db.tour_visits.count_documents({"tour_id": tour_id})
    track_n = await db.tour_tracking_logs.count_documents({"tour_id": tour_id})
    expenses = await db.expense_claims.find(
        {"tour_id": tour_id}, {"_id": 0, "claim_no": 1, "amount": 1, "status": 1,
                               "category_name": 1, "expense_date": 1}).to_list(200)
    exp_total = round(sum(float(e.get("amount") or 0) for e in expenses
                          if e.get("status") not in ("rejected", "cancelled")), 2)
    # Total distance (haversine over track points, only if enough data)
    distance_km = None
    pts = await db.tour_tracking_logs.find(
        {"tour_id": tour_id}, {"_id": 0, "lat": 1, "lng": 1}
    ).sort("captured_at", 1).to_list(5000)
    if len(pts) >= 2:
        import math
        dist = 0.0
        for a, b in zip(pts, pts[1:]):
            la1, lo1, la2, lo2 = map(math.radians, (a["lat"], a["lng"], b["lat"], b["lng"]))
            h = (math.sin((la2 - la1) / 2) ** 2
                 + math.cos(la1) * math.cos(la2) * math.sin((lo2 - lo1) / 2) ** 2)
            dist += 2 * 6371 * math.asin(math.sqrt(h))
        distance_km = round(dist, 1)
    return {"tour_no": t["tour_no"], "status": t["status"],
            "started_at": t.get("started_at"), "ended_at": t.get("ended_at"),
            "total_days": t.get("total_days"), "visits": visits,
            "tracking_points": track_n, "distance_km": distance_km,
            "expenses_total": exp_total, "expenses_count": len(expenses)}


# ---------------------------------------------------------------------------
# Visits / client meetings
# ---------------------------------------------------------------------------
@router.post("/{tour_id}/visits")
async def add_visit(tour_id: str, payload: Dict[str, Any] = Body(...),
                    authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    if t.get("user_id") != user.get("user_id"):
        raise HTTPException(status_code=403, detail="Not your tour")
    if t["status"] != "active":
        raise HTTPException(status_code=400, detail="Visits can only be added on an ACTIVE tour")
    v = {
        "visit_id": f"tv_{uuid.uuid4().hex[:12]}", "tour_id": tour_id,
        "company_id": t["company_id"], "user_id": user["user_id"],
        "visit_date": str(payload.get("visit_date") or now_iso()[:10])[:10],
        "start_time": str(payload.get("start_time") or "")[:5],
        "end_time": str(payload.get("end_time") or "")[:5],
        "client_name": str(payload.get("client_name") or "")[:150],
        "contact_person": str(payload.get("contact_person") or "")[:100],
        "contact_number": str(payload.get("contact_number") or "")[:20],
        "meeting_purpose": str(payload.get("meeting_purpose") or "")[:400],
        "gps": _gps_from(payload),
        "summary": str(payload.get("summary") or "")[:800],
        "discussion": str(payload.get("discussion") or "")[:1200],
        "outcome": str(payload.get("outcome") or "")[:400],
        "next_followup": str(payload.get("next_followup") or "")[:10],
        "remarks": str(payload.get("remarks") or "")[:400],
        "attachments": [], "created_at": now_iso(), "updated_at": now_iso(),
    }
    if not v["client_name"]:
        raise HTTPException(status_code=400, detail="Client / company name is required")
    await db.tour_visits.insert_one(dict(v))
    v.pop("_id", None)
    await _taudit(user, tour_id, "visit_added", None,
                  {"visit_id": v["visit_id"], "client": v["client_name"]}, gps=v["gps"])
    return {"ok": True, "visit": v}


@router.put("/{tour_id}/visits/{visit_id}")
async def edit_visit(tour_id: str, visit_id: str,
                     payload: Dict[str, Any] = Body(...),
                     authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    if t.get("user_id") != user.get("user_id") and user.get("role") not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Not your tour")
    if t["status"] not in ("active",) and user.get("role") not in ADMIN_ROLES:
        raise HTTPException(status_code=400,
                            detail="Meetings can be edited only before tour completion")
    v = await db.tour_visits.find_one({"visit_id": visit_id, "tour_id": tour_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Visit not found")
    editable = ("visit_date", "start_time", "end_time", "client_name", "contact_person",
                "contact_number", "meeting_purpose", "summary", "discussion",
                "outcome", "next_followup", "remarks")
    upd = {k: str(payload[k] or "")[:1200] for k in editable if k in payload}
    if not upd:
        raise HTTPException(status_code=400, detail="Nothing to update")
    upd["updated_at"] = now_iso()
    await db.tour_visits.update_one({"visit_id": visit_id}, {"$set": upd})
    await _taudit(user, tour_id, "visit_edited",
                  {k: v.get(k) for k in upd if k != "updated_at"},
                  {k: u for k, u in upd.items() if k != "updated_at"})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Attachments (tour docs, visit photos / business cards) — base64 pattern
# shared with the Expense module.
# ---------------------------------------------------------------------------
async def _store_attachment(payload: dict) -> dict:
    mime = str(payload.get("mime") or "")
    if mime not in ALLOWED_MIME:
        raise HTTPException(status_code=400, detail="Only PDF / JPG / PNG files are allowed")
    b64 = str(payload.get("data_b64") or "")
    try:
        raw_len = len(base64.b64decode(b64[:20] + "==", validate=False)) and (len(b64) * 3 // 4)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid file data")
    if raw_len > MAX_FILE_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"File exceeds {MAX_FILE_MB} MB")
    return {"doc_id": f"tdoc_{uuid.uuid4().hex[:12]}",
            "name": str(payload.get("name") or "document")[:120],
            "kind": str(payload.get("kind") or "other")[:40],
            "mime": mime, "data_b64": b64, "uploaded_at": now_iso()}


@router.post("/{tour_id}/attachments")
async def add_tour_attachment(tour_id: str, payload: Dict[str, Any] = Body(...),
                              authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    if t.get("user_id") != user.get("user_id") and user.get("role") not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Not your tour")
    doc = await _store_attachment(payload)
    target = payload.get("visit_id")
    if target:
        r = await db.tour_visits.update_one(
            {"visit_id": target, "tour_id": tour_id}, {"$push": {"attachments": doc}})
        if not r.matched_count:
            raise HTTPException(status_code=404, detail="Visit not found")
    else:
        await db.tour_requests.update_one(
            {"tour_id": tour_id}, {"$push": {"attachments": doc}})
    await _taudit(user, tour_id, "attachment_added", None,
                  {"doc_id": doc["doc_id"], "name": doc["name"], "visit_id": target})
    return {"ok": True, "doc_id": doc["doc_id"], "name": doc["name"],
            "kind": doc["kind"], "mime": doc["mime"]}


@router.get("/{tour_id}/attachments/{doc_id}")
async def get_tour_attachment(tour_id: str, doc_id: str,
                              authorization: Optional[str] = Header(None),
                              token: Optional[str] = Query(None)):
    user = await get_user_from_token(authorization or (f"Bearer {token}" if token else None))
    await _get_tour_scoped(tour_id, user)
    t = await db.tour_requests.find_one(
        {"tour_id": tour_id, "attachments.doc_id": doc_id},
        {"_id": 0, "attachments.$": 1})
    doc = (t or {}).get("attachments", [None])[0]
    if not doc:
        v = await db.tour_visits.find_one(
            {"tour_id": tour_id, "attachments.doc_id": doc_id},
            {"_id": 0, "attachments.$": 1})
        doc = (v or {}).get("attachments", [None])[0]
    if not doc:
        raise HTTPException(status_code=404, detail="Attachment not found")
    return Response(content=base64.b64decode(doc["data_b64"]),
                    media_type=doc["mime"],
                    headers={"Content-Disposition": f'inline; filename="{doc["name"]}"'})


# ---------------------------------------------------------------------------
# Detail (timeline + visits + expenses + attendance + approvals)
# ---------------------------------------------------------------------------
@router.get("/{tour_id}")
async def tour_detail(tour_id: str, authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    t = await _get_tour_scoped(tour_id, user)
    for a in t.get("attachments") or []:
        a.pop("data_b64", None)
    visits = await db.tour_visits.find(
        {"tour_id": tour_id}, {"_id": 0, "attachments.data_b64": 0}
    ).sort([("visit_date", 1), ("start_time", 1)]).to_list(100)
    expenses = await db.expense_claims.find(
        {"tour_id": tour_id},
        {"_id": 0, "claim_id": 1, "claim_no": 1, "amount": 1, "status": 1,
         "category_name": 1, "expense_date": 1, "vendor": 1}).sort("created_at", -1).to_list(100)
    attendance = await db.tour_attendance.find(
        {"tour_id": tour_id}, {"_id": 0}).sort("date", 1).to_list(70)
    track_n = await db.tour_tracking_logs.count_documents({"tour_id": tour_id})
    # Timeline — merge lifecycle events, tracking milestones and visits.
    timeline: List[dict] = []
    if t.get("submitted_at"):
        timeline.append({"at": t["submitted_at"], "label": "Tour Submitted", "icon": "send"})
    if t.get("approved_at"):
        timeline.append({"at": t["approved_at"], "label": "Tour Approved",
                         "detail": t.get("approved_by_name"), "icon": "check"})
    if t.get("started_at"):
        timeline.append({"at": t["started_at"], "label": "Tour Started",
                         "detail": t.get("from_location"), "gps": t.get("start_gps"), "icon": "play"})
    for v in visits:
        vt = f"{v.get('visit_date')}T{(v.get('start_time') or '00:00')}:00"
        timeline.append({"at": vt, "label": f"Meeting — {v.get('client_name')}",
                         "detail": v.get("meeting_purpose"), "gps": v.get("gps"),
                         "icon": "people", "visit_id": v["visit_id"]})
    if t.get("ended_at"):
        timeline.append({"at": t["ended_at"], "label": "Tour Ended",
                         "detail": t.get("end_remarks"), "gps": t.get("end_gps"), "icon": "stop"})
    timeline.sort(key=lambda x: str(x.get("at") or ""))
    # Approval history — engine request if routed, else direct history.
    approval = None
    if t.get("approval_request_id"):
        approval = await db.approval_requests.find_one(
            {"request_id": t["approval_request_id"]},
            {"_id": 0, "status": 1, "current_level": 1, "levels": 1, "history": 1})
    settings = await get_tour_settings(t["company_id"])
    return {"tour": t, "visits": visits, "expenses": expenses,
            "attendance": attendance, "timeline": timeline,
            "tracking_points": track_n, "approval": approval,
            "od_preview": _od_preview(t) if t.get("mark_od") else [],
            "tracking_interval_min": settings["tracking_interval_min"]}


@router.get("/{tour_id}/tracking")
async def tour_tracking(tour_id: str, authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    await _get_tour_scoped(tour_id, user)
    pts = await db.tour_tracking_logs.find(
        {"tour_id": tour_id}, {"_id": 0}).sort("captured_at", 1).to_list(2000)
    return {"points": pts, "count": len(pts)}


@router.get("/{tour_id}/audit")
async def tour_audit(tour_id: str, authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    require_role(user, ADMIN_ROLES)
    await _get_tour_scoped(tour_id, user)
    rows = await db.tour_audit.find({"tour_id": tour_id}, {"_id": 0}).sort("at", -1).to_list(300)
    return {"audit": rows}


# ---------------------------------------------------------------------------
# Expense-claim link
# ---------------------------------------------------------------------------
@router.get("/eligible/for-expense")
async def eligible_tours_for_expense(authorization: Optional[str] = Header(None)):
    """Employee's own tours eligible for an official-tour expense claim."""
    user = await get_user_from_token(authorization)
    settings = await get_tour_settings(user.get("company_id") or "")
    grace = int(settings.get("expense_claim_grace_days") or 7)
    min_end = (datetime.now(timezone.utc) - timedelta(days=grace + 45)).strftime("%Y-%m-%d")
    rows = await db.tour_requests.find(
        {"user_id": user["user_id"], "status": {"$in": ["approved", "active", "completed"]},
         "end_date": {"$gte": min_end}},
        {"_id": 0, "tour_id": 1, "tour_no": 1, "tour_type": 1, "purpose": 1,
         "destinations": 1, "start_date": 1, "end_date": 1, "status": 1,
         "approved_by_name": 1}).sort("start_date", -1).to_list(50)
    return {"tours": rows, "grace_days": grace}


async def validate_tour_for_expense(user: dict, tour_id: str,
                                    expense_date: str) -> dict:
    """Called by the Expense module when Official Tour = YES."""
    t = await db.tour_requests.find_one({"tour_id": tour_id}, {"_id": 0})
    if not t or t.get("user_id") != user.get("user_id"):
        raise HTTPException(status_code=400,
                            detail="Select one of your own approved tours")
    if t.get("status") not in ("approved", "active", "completed"):
        raise HTTPException(status_code=400,
                            detail=f"Tour {t.get('tour_no')} is {t.get('status')} — "
                                   "only approved tours can carry expense claims")
    settings = await get_tour_settings(t["company_id"])
    grace = int(settings.get("expense_claim_grace_days") or 7)
    if expense_date:
        try:
            ed = date.fromisoformat(str(expense_date)[:10])
            d0 = date.fromisoformat(t["start_date"])
            d1 = date.fromisoformat(t["end_date"]) + timedelta(days=grace)
            if not (d0 <= ed <= d1):
                raise HTTPException(
                    status_code=400,
                    detail=f"Expense date must be within the tour period "
                           f"{t['start_date']} → {t['end_date']} (+{grace} days grace)")
        except (ValueError, TypeError):
            pass
    return {"tour_id": t["tour_id"], "tour_no": t["tour_no"],
            "tour_type": t.get("tour_type"), "purpose": t.get("purpose"),
            "destinations": t.get("destinations"),
            "start_date": t.get("start_date"), "end_date": t.get("end_date"),
            "status": t.get("status"), "approved_by_name": t.get("approved_by_name")}


# ---------------------------------------------------------------------------
# OD / TOUR attendance posting engine
# ---------------------------------------------------------------------------
async def post_tour_attendance(t: dict, actor: dict):
    """Runs ONLY after final approval. Conflict-safe: never overwrites an
    existing punch automatically — flags it for admin resolution."""
    settings = await get_tour_settings(t["company_id"])
    company = await db.companies.find_one(
        {"company_id": t["company_id"]},
        {"_id": 0, "attendance_policy": 1}) or {}
    full_hours = float((company.get("attendance_policy") or {}).get("full_day_hours") or 8.0)
    half_hours = float(settings.get("half_day_od_hours") or 4)
    duty_hours = half_hours if t.get("od_day_type") == "half" else full_hours
    posted, conflicts, skipped = 0, 0, 0
    try:
        d0 = date.fromisoformat(t["start_date"])
        d1 = date.fromisoformat(t["end_date"])
    except (ValueError, TypeError):
        return
    # Holiday / weekly-off handling
    holiday_dates: set = set()
    if settings.get("holiday_during_tour") == "skip":
        try:
            from server import holiday_dates_for_company
            holiday_dates = set(await holiday_dates_for_company(t["company_id"]))
        except Exception:
            holiday_dates = set()
    cur = d0
    while cur <= d1:
        dkey = cur.isoformat()
        cur += timedelta(days=1)
        if dkey in holiday_dates:
            skipped += 1
            continue
        if settings.get("weekly_off_during_tour") == "skip" and \
                date.fromisoformat(dkey).weekday() == 6:  # Sunday default WO
            skipped += 1
            continue
        existing = await db.attendance.count_documents(
            {"user_id": t["user_id"], "date": dkey, "status": "approved"})
        rec = {
            "record_id": f"toa_{uuid.uuid4().hex[:12]}",
            "tour_id": t["tour_id"], "tour_no": t["tour_no"],
            "company_id": t["company_id"], "user_id": t["user_id"],
            "employee_code": (t.get("employee") or {}).get("employee_code"),
            "date": dkey, "attendance_type": "OD",
            "day": "half" if t.get("od_day_type") == "half" else "full",
            "source": "OFFICIAL_TOUR", "created_by": "system",
            "approval_reference": t.get("approval_request_id") or "direct_admin",
            "approval_date": t.get("approved_at") or now_iso(),
            "expense_claim_id": None, "created_at": now_iso(),
        }
        if existing:
            rec["status"] = "conflict"
            rec["conflict_reason"] = ("Attendance already exists for this date. "
                                      "Tour attendance cannot overwrite it automatically.")
            conflicts += 1
        elif not settings.get("od_counts_present", True):
            rec["status"] = "recorded_only"
            skipped += 1
        else:
            rec["status"] = "posted"
            in_at = f"{dkey}T09:00:00Z"
            out_h = 9 + duty_hours
            out_at = f"{dkey}T{int(out_h):02d}:{int((out_h % 1) * 60):02d}:00Z"
            for kind, at in (("in", in_at), ("out", out_at)):
                await db.attendance.insert_one({
                    "record_id": f"att_{uuid.uuid4().hex[:12]}",
                    "user_id": t["user_id"], "company_id": t["company_id"],
                    "date": dkey, "kind": kind, "at": at,
                    "source": "official_tour", "status": "approved",
                    "approved_by": actor.get("user_id") or "system",
                    "manual_reason": f"OD — Official Tour {t['tour_no']}",
                    "tour_id": t["tour_id"],
                    "created_by": "system", "created_at": now_iso()})
            posted += 1
        await db.tour_attendance.insert_one(rec)
    await db.tour_requests.update_one(
        {"tour_id": t["tour_id"]},
        {"$set": {"attendance_summary": {
            "posted": posted, "conflicts": conflicts, "skipped": skipped,
            "posted_at": now_iso()}}})
    await _taudit(actor, t["tour_id"], "attendance_posted", None,
                  {"posted": posted, "conflicts": conflicts, "skipped": skipped})
    if conflicts:
        await _notify_admins(t["company_id"],
                             f"Tour {t['tour_no']} — {conflicts} attendance conflict(s)",
                             "Existing punches found on tour dates — resolve in Tour Management",
                             t["tour_id"])


async def _cancel_tour_attendance(t: dict, actor: dict, reason: str):
    """Remove system-posted OD punches when a tour is cancelled."""
    n = 0
    async for rec in db.tour_attendance.find(
            {"tour_id": t["tour_id"], "status": "posted"}, {"_id": 0}):
        await db.attendance.delete_many(
            {"tour_id": t["tour_id"], "date": rec["date"], "source": "official_tour"})
        n += 1
    if n:
        await db.tour_attendance.update_many(
            {"tour_id": t["tour_id"], "status": "posted"},
            {"$set": {"status": "cancelled", "cancelled_reason": reason,
                      "cancelled_at": now_iso()}})
        await _taudit(actor, t["tour_id"], "attendance_cancelled", None,
                      {"days_removed": n}, remarks=reason)


@router.post("/{tour_id}/attendance/resolve")
async def resolve_attendance_conflict(tour_id: str,
                                      payload: Dict[str, Any] = Body(...),
                                      authorization: Optional[str] = Header(None)):
    """Admin resolution for a conflicted OD day.
    action: keep_existing | convert_to_od | cancel_tour_attendance."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    t = await _get_tour_scoped(tour_id, admin)
    dkey = str(payload.get("date") or "")[:10]
    action = (payload.get("action") or "").strip()
    if action not in ("keep_existing", "convert_to_od", "cancel_tour_attendance"):
        raise HTTPException(status_code=400,
                            detail="action must be keep_existing|convert_to_od|cancel_tour_attendance")
    rec = await db.tour_attendance.find_one(
        {"tour_id": tour_id, "date": dkey, "status": "conflict"}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="No conflicted OD record for that date")
    if action == "keep_existing":
        await db.tour_attendance.update_one(
            {"record_id": rec["record_id"]},
            {"$set": {"status": "kept_existing", "resolved_by": admin["user_id"],
                      "resolved_at": now_iso()}})
    elif action == "cancel_tour_attendance":
        await db.tour_attendance.update_one(
            {"record_id": rec["record_id"]},
            {"$set": {"status": "cancelled", "resolved_by": admin["user_id"],
                      "resolved_at": now_iso()}})
    else:  # convert_to_od — supersede existing punches, insert OD pair
        await db.attendance.update_many(
            {"user_id": t["user_id"], "date": dkey, "status": "approved",
             "source": {"$ne": "official_tour"}},
            {"$set": {"status": "superseded_tour",
                      "superseded_by": admin["user_id"],
                      "superseded_at": now_iso()}})
        settings = await get_tour_settings(t["company_id"])
        company = await db.companies.find_one(
            {"company_id": t["company_id"]}, {"_id": 0, "attendance_policy": 1}) or {}
        full_hours = float((company.get("attendance_policy") or {}).get("full_day_hours") or 8.0)
        duty = float(settings.get("half_day_od_hours") or 4) \
            if rec.get("day") == "half" else full_hours
        out_h = 9 + duty
        for kind, at in (("in", f"{dkey}T09:00:00Z"),
                         ("out", f"{dkey}T{int(out_h):02d}:{int((out_h % 1) * 60):02d}:00Z")):
            await db.attendance.insert_one({
                "record_id": f"att_{uuid.uuid4().hex[:12]}",
                "user_id": t["user_id"], "company_id": t["company_id"],
                "date": dkey, "kind": kind, "at": at,
                "source": "official_tour", "status": "approved",
                "approved_by": admin["user_id"],
                "manual_reason": f"OD (converted) — Official Tour {t['tour_no']}",
                "tour_id": tour_id, "created_by": "system", "created_at": now_iso()})
        await db.tour_attendance.update_one(
            {"record_id": rec["record_id"]},
            {"$set": {"status": "posted", "converted": True,
                      "resolved_by": admin["user_id"], "resolved_at": now_iso()}})
    await _taudit(admin, tour_id, f"attendance_conflict_{action}",
                  {"date": dkey, "status": "conflict"}, {"date": dkey, "action": action},
                  remarks=str(payload.get("remarks") or ""))
    return {"ok": True, "action": action, "date": dkey}


# ---------------------------------------------------------------------------
# Admin — Tour Management
# ---------------------------------------------------------------------------
def _admin_cid(user: dict, company_id: Optional[str]) -> Optional[str]:
    if user.get("role") in ("super_admin", "sub_admin") and company_id:
        return company_id
    return user.get("company_id") or company_id


@router.get("/admin/list")
async def admin_tours(company_id: Optional[str] = Query(None),
                      status: Optional[str] = Query(None),
                      user_id: Optional[str] = Query(None),
                      authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    cid = _admin_cid(admin, company_id)
    q: Dict[str, Any] = {}
    if cid:
        q["company_id"] = cid
    if status and status != "all":
        q["status"] = status
    if user_id:
        q["user_id"] = user_id
    rows = await db.tour_requests.find(
        q, {"_id": 0, "attachments.data_b64": 0}).sort("created_at", -1).to_list(300)
    counts: Dict[str, int] = {}
    base = {"company_id": cid} if cid else {}
    for s in STATUSES:
        counts[s] = await db.tour_requests.count_documents({**base, "status": s})
    counts["total"] = sum(counts.values())
    return {"tours": rows, "counts": counts}


@router.get("/admin/live")
async def admin_live_tours(company_id: Optional[str] = Query(None),
                           authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    cid = _admin_cid(admin, company_id)
    q: Dict[str, Any] = {"status": "active"}
    if cid:
        q["company_id"] = cid
    rows = await db.tour_requests.find(
        q, {"_id": 0, "attachments.data_b64": 0}).sort("started_at", -1).to_list(100)
    out = []
    for t in rows:
        visits = await db.tour_visits.count_documents({"tour_id": t["tour_id"]})
        out.append({
            "tour_id": t["tour_id"], "tour_no": t["tour_no"],
            "employee": t.get("employee"), "user_id": t["user_id"],
            "tour_type": t.get("tour_type"), "destinations": t.get("destinations"),
            "start_date": t.get("start_date"), "end_date": t.get("end_date"),
            "started_at": t.get("started_at"),
            "last_location": t.get("last_location"),
            "tracking_active": bool(t.get("tracking_active")),
            "visits": visits})
    return {"active_tours": out}


@router.get("/admin/settings")
async def admin_tour_settings(company_id: Optional[str] = Query(None),
                              authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    cid = _admin_cid(admin, company_id)
    if not cid:
        raise HTTPException(status_code=400, detail="company_id is required")
    return {"company_id": cid, "settings": await get_tour_settings(cid)}


@router.post("/admin/settings")
async def save_tour_settings(payload: Dict[str, Any] = Body(...),
                             authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    cid = _admin_cid(admin, payload.get("company_id"))
    if not cid:
        raise HTTPException(status_code=400, detail="company_id is required")
    sets: Dict[str, Any] = {}
    if "tracking_interval_min" in payload:
        iv = int(payload["tracking_interval_min"] or 5)
        if iv not in (1, 5, 10, 15):
            raise HTTPException(status_code=400, detail="tracking interval must be 1|5|10|15")
        sets["tracking_interval_min"] = iv
    for k in ("od_counts_present", "od_counts_paid", "od_ot_eligible"):
        if k in payload:
            sets[k] = bool(payload[k])
    for k in ("holiday_during_tour", "weekly_off_during_tour"):
        if k in payload:
            if payload[k] not in ("skip", "od"):
                raise HTTPException(status_code=400, detail=f"{k} must be skip|od")
            sets[k] = payload[k]
    if "half_day_od_hours" in payload:
        sets["half_day_od_hours"] = min(8.0, max(1.0, float(payload["half_day_od_hours"] or 4)))
    if "expense_claim_grace_days" in payload:
        sets["expense_claim_grace_days"] = min(30, max(0, int(payload["expense_claim_grace_days"] or 7)))
    if not sets:
        raise HTTPException(status_code=400, detail="Nothing to update")
    sets["updated_at"] = now_iso()
    sets["updated_by"] = admin["user_id"]
    await db.tour_settings.update_one({"company_id": cid}, {"$set": sets}, upsert=True)
    return {"ok": True, "settings": await get_tour_settings(cid)}


# ---------------------------------------------------------------------------
# Iter 707 — Advance payout ledger (accounts team settle & track)
# ---------------------------------------------------------------------------
@router.get("/admin/advances")
async def admin_tour_advances(company_id: Optional[str] = Query(None),
                              status: Optional[str] = Query(None),
                              authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    cid = _admin_cid(admin, company_id)
    q: Dict[str, Any] = {"advance_payout": {"$exists": True}}
    if cid:
        q["company_id"] = cid
    if status and status != "all":
        q["advance_payout.status"] = status
    rows = await db.tour_requests.find(
        q, {"_id": 0, "tour_id": 1, "tour_no": 1, "employee": 1, "user_id": 1,
            "tour_type": 1, "destinations": 1, "start_date": 1, "end_date": 1,
            "status": 1, "advance_amount": 1, "advance_payout": 1}
    ).sort("created_at", -1).to_list(300)
    out = []
    for t in rows:
        exp = await db.expense_claims.find(
            {"tour_id": t["tour_id"], "status": {"$nin": ["rejected", "cancelled", "draft"]}},
            {"_id": 0, "amount": 1, "approved_amount": 1, "status": 1}).to_list(200)
        claimed = round(sum(float(e.get("amount") or 0) for e in exp), 2)
        approved = round(sum(float(e.get("approved_amount") if e.get("approved_amount") is not None
                                   else e.get("amount") or 0)
                             for e in exp
                             if e.get("status") in ("approved", "payment_pending",
                                                    "processing", "paid")), 2)
        t["expenses_claimed"] = claimed
        t["expenses_approved"] = approved
        t["balance"] = round(approved - float((t.get("advance_payout") or {}).get("amount") or 0), 2)
        out.append(t)
    counts = {}
    base = {"advance_payout": {"$exists": True}, **({"company_id": cid} if cid else {})}
    for st in ("pending", "paid", "settled"):
        counts[st] = await db.tour_requests.count_documents(
            {**base, "advance_payout.status": st})
    counts["total"] = sum(counts.values())
    return {"advances": out, "counts": counts}


@router.post("/{tour_id}/advance/pay")
async def pay_tour_advance(tour_id: str, payload: Dict[str, Any] = Body(default={}),
                           authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    t = await _get_tour_scoped(tour_id, admin)
    ap = t.get("advance_payout") or {}
    if ap.get("status") != "pending":
        raise HTTPException(status_code=400,
                            detail=f"Advance is {ap.get('status') or 'not created'} — only pending advances can be paid")
    upd = {"advance_payout.status": "paid",
           "advance_payout.paid_at": now_iso(),
           "advance_payout.paid_by": admin["user_id"],
           "advance_payout.paid_by_name": admin.get("name") or admin.get("email"),
           "advance_payout.mode": str(payload.get("mode") or "cash")[:30],
           "advance_payout.reference": str(payload.get("reference") or "")[:60],
           "advance_payout.remarks": str(payload.get("remarks") or "")[:200]}
    await db.tour_requests.update_one({"tour_id": tour_id}, {"$set": upd})
    await _taudit(admin, tour_id, "advance_paid", "pending",
                  {"amount": ap.get("amount"), "mode": upd["advance_payout.mode"],
                   "reference": upd["advance_payout.reference"]})
    try:
        await _notify(t["company_id"], t["user_id"],
                      f"Tour advance paid — {t['tour_no']}",
                      f"₹{ap.get('amount')} advance paid ({upd['advance_payout.mode']})", tour_id)
    except Exception:
        pass
    return {"ok": True, "status": "paid"}


@router.post("/{tour_id}/advance/settle")
async def settle_tour_advance(tour_id: str, payload: Dict[str, Any] = Body(default={}),
                              authorization: Optional[str] = Header(None)):
    """Final settlement: approved tour expenses vs advance paid.
    balance > 0 → payable to employee; balance < 0 → recoverable."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    t = await _get_tour_scoped(tour_id, admin)
    ap = t.get("advance_payout") or {}
    if ap.get("status") != "paid":
        raise HTTPException(status_code=400,
                            detail="Advance must be PAID before it can be settled")
    exp = await db.expense_claims.find(
        {"tour_id": tour_id,
         "status": {"$in": ["approved", "payment_pending", "processing", "paid"]}},
        {"_id": 0, "amount": 1, "approved_amount": 1}).to_list(200)
    approved_total = round(sum(float(e.get("approved_amount") if e.get("approved_amount") is not None
                                     else e.get("amount") or 0) for e in exp), 2)
    balance = round(approved_total - float(ap.get("amount") or 0), 2)
    upd = {"advance_payout.status": "settled",
           "advance_payout.settled_at": now_iso(),
           "advance_payout.settled_by": admin["user_id"],
           "advance_payout.settled_by_name": admin.get("name") or admin.get("email"),
           "advance_payout.expense_total": approved_total,
           "advance_payout.balance": balance,
           "advance_payout.settle_remarks": str(payload.get("remarks") or "")[:200]}
    await db.tour_requests.update_one({"tour_id": tour_id}, {"$set": upd})
    await _taudit(admin, tour_id, "advance_settled", "paid",
                  {"expense_total": approved_total, "balance": balance})
    return {"ok": True, "status": "settled", "expense_total": approved_total,
            "balance": balance}


# ---------------------------------------------------------------------------
# Iter 707 — Monthly tour report (per employee: days, visits, expenses, OD)
# ---------------------------------------------------------------------------
async def _tour_report_rows(cid: str, month: str) -> List[dict]:
    d0 = date.fromisoformat(f"{month}-01")
    d1 = (d0.replace(year=d0.year + 1, month=1) if d0.month == 12
          else d0.replace(month=d0.month + 1)) - timedelta(days=1)
    m_start, m_end = d0.isoformat(), d1.isoformat()
    tours = await db.tour_requests.find(
        {"company_id": cid, "status": {"$nin": ["draft", "cancelled", "rejected"]},
         "start_date": {"$lte": m_end}, "end_date": {"$gte": m_start}},
        {"_id": 0, "tour_id": 1, "tour_no": 1, "user_id": 1, "employee": 1,
         "start_date": 1, "end_date": 1, "status": 1, "advance_payout": 1}).to_list(1000)
    rows: Dict[str, dict] = {}
    for t in tours:
        uid = t["user_id"]
        r = rows.setdefault(uid, {
            "user_id": uid,
            "name": (t.get("employee") or {}).get("name") or "",
            "employee_code": (t.get("employee") or {}).get("employee_code") or "",
            "department": (t.get("employee") or {}).get("department") or "",
            "tours": 0, "tour_nos": [], "tour_days": 0, "visits": 0,
            "expenses_claimed": 0.0, "expenses_approved": 0.0,
            "od_posted": 0, "od_conflicts": 0, "advance_paid": 0.0})
        r["tours"] += 1
        r["tour_nos"].append(t["tour_no"])
        clip0 = max(date.fromisoformat(t["start_date"]), d0)
        clip1 = min(date.fromisoformat(t["end_date"]), d1)
        r["tour_days"] += max(0, (clip1 - clip0).days + 1)
        r["visits"] += await db.tour_visits.count_documents(
            {"tour_id": t["tour_id"], "visit_date": {"$gte": m_start, "$lte": m_end}})
        async for e in db.expense_claims.find(
                {"tour_id": t["tour_id"],
                 "expense_date": {"$gte": m_start, "$lte": m_end},
                 "status": {"$nin": ["rejected", "cancelled", "draft"]}},
                {"_id": 0, "amount": 1, "approved_amount": 1, "status": 1}):
            r["expenses_claimed"] += float(e.get("amount") or 0)
            if e.get("status") in ("approved", "payment_pending", "processing", "paid"):
                r["expenses_approved"] += float(
                    e.get("approved_amount") if e.get("approved_amount") is not None
                    else e.get("amount") or 0)
        r["od_posted"] += await db.tour_attendance.count_documents(
            {"tour_id": t["tour_id"], "status": "posted",
             "date": {"$gte": m_start, "$lte": m_end}})
        r["od_conflicts"] += await db.tour_attendance.count_documents(
            {"tour_id": t["tour_id"], "status": "conflict",
             "date": {"$gte": m_start, "$lte": m_end}})
        ap = t.get("advance_payout") or {}
        if ap.get("status") in ("paid", "settled"):
            r["advance_paid"] += float(ap.get("amount") or 0)
    out = sorted(rows.values(), key=lambda x: (x["name"] or "").lower())
    for r in out:
        r["expenses_claimed"] = round(r["expenses_claimed"], 2)
        r["expenses_approved"] = round(r["expenses_approved"], 2)
        r["advance_paid"] = round(r["advance_paid"], 2)
    return out


@router.get("/admin/report")
async def tour_monthly_report(company_id: Optional[str] = Query(None),
                              month: str = Query(...),
                              authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ADMIN_ROLES)
    cid = _admin_cid(admin, company_id)
    if not cid:
        raise HTTPException(status_code=400, detail="company_id is required")
    try:
        date.fromisoformat(f"{month}-01")
    except ValueError:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    rows = await _tour_report_rows(cid, month)
    totals = {k: round(sum(r[k] for r in rows), 2) for k in
              ("tours", "tour_days", "visits", "expenses_claimed",
               "expenses_approved", "od_posted", "od_conflicts", "advance_paid")}
    return {"month": month, "rows": rows, "totals": totals}


@router.get("/admin/report.xlsx")
async def tour_monthly_report_xlsx(company_id: Optional[str] = Query(None),
                                   month: str = Query(...),
                                   authorization: Optional[str] = Header(None),
                                   token: Optional[str] = Query(None)):
    admin = await get_user_from_token(authorization or (f"Bearer {token}" if token else None))
    require_role(admin, ADMIN_ROLES)
    cid = _admin_cid(admin, company_id)
    if not cid:
        raise HTTPException(status_code=400, detail="company_id is required")
    rows = await _tour_report_rows(cid, month)
    company = await db.companies.find_one({"company_id": cid}, {"_id": 0, "name": 1}) or {}
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Tour Report"
    ws.append([f"Monthly Tour Report — {company.get('name') or cid} — {month}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    hdr = ["S.No.", "Employee", "Code", "Department", "Tours", "Tour Nos",
           "Tour Days", "Visits", "Expenses Claimed", "Expenses Approved",
           "OD Days Posted", "OD Conflicts", "Advance Paid"]
    ws.append(hdr)
    for c in ws[3]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(rows, 1):
        ws.append([i, r["name"], r["employee_code"], r["department"], r["tours"],
                   ", ".join(r["tour_nos"]), r["tour_days"], r["visits"],
                   r["expenses_claimed"], r["expenses_approved"],
                   r["od_posted"], r["od_conflicts"], r["advance_paid"]])
    if rows:
        ws.append(["", "TOTAL", "", "", sum(r["tours"] for r in rows), "",
                   sum(r["tour_days"] for r in rows), sum(r["visits"] for r in rows),
                   round(sum(r["expenses_claimed"] for r in rows), 2),
                   round(sum(r["expenses_approved"] for r in rows), 2),
                   sum(r["od_posted"] for r in rows), sum(r["od_conflicts"] for r in rows),
                   round(sum(r["advance_paid"] for r in rows), 2)])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJKLM", (6, 24, 10, 16, 7, 26, 10, 8, 16, 17, 14, 12, 13)):
        ws.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="tour_report_{month}.xlsx"'})
