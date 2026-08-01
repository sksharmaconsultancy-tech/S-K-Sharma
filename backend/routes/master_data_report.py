"""Iter 91 — Route module: Master Data report (read-only).

Sidebar → Reports → Master Data. Three views over the EMPLOYEE master:

  * ``active`` — working right now (no exit/resign date on the master)
  * ``left``   — resign/exit date already set on the master
  * ``all``    — everything

Data is strictly READ-ONLY here (no edit endpoints) and can be exported
to Excel. Filters: free-text name/code search, Employee Type / Group,
firm (super admin), on-roll flag.

  GET /api/admin/reports/master-data?status=active|left|all&q=&employee_type=&company_id=&is_onroll=
  GET /api/admin/reports/master-data.xlsx?...same params...
"""
from io import BytesIO
from typing import Any, Dict, Optional

from fastapi import APIRouter, Header, HTTPException
from fastapi.responses import StreamingResponse

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
)

router = APIRouter(prefix="/api/admin/reports", tags=["master-data-report"])

_COLUMNS = [
    # Iter 331 (user request) — exact column sequence.
    ("uan_no", "UAN"),
    ("pf_no", "EPFO No."),
    ("esi_ip_no", "ESIC No"),
    ("employee_code", "Emp Code"),
    ("name", "Name"),
    ("father_name", "Father / Spouse Name"),
    ("gender", "Gender"),
    ("dob", "Date of Birth"),
    ("marital_status", "Marital Status"),
    ("phone", "Phone"),
    ("designation", "Designation"),
    ("department", "Department"),
    ("employee_type", "Type / Group"),
    ("is_onroll", "On-roll"),
    ("doj", "Date of Join"),
    ("exit_date", "Exit / Resign Date"),   # hidden on the Active tab
    ("basic", "Basic"),
    ("pf_basic", "PF Basic"),
    ("hra", "HRA"),
    ("conveyance", "Conv."),
    # ← dynamic allowance heads (from the Employee Master) inserted here
    ("monthly_gross", "Monthly Gross"),
    ("pan_no", "PAN"),
    ("pan_name", "Name As Per PAN"),
    ("aadhaar_no", "Aadhaar"),
    ("aadhaar_name", "Name As Per Aadhaar"),
    ("bank_name", "Bank"),
    ("bank_account", "Account No"),
    ("bank_ifsc", "IFSC"),
    ("upi_id", "UPI ID"),
    ("present_address", "Present Address"),
    ("district", "City"),
    ("state", "State"),
    ("pincode", "PIN"),
    ("permanent_address", "Permanent Address"),
    ("permanent_district", "Perm. City"),
    ("permanent_state", "Perm. State"),
    ("permanent_pincode", "Perm. PIN"),
    ("company_name", "Firm"),
]


def _num0(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _allow_key(head: str) -> str:
    import re as _re
    return "al_" + _re.sub(r"[^a-z0-9]+", "_", head.strip().lower()).strip("_")


async def _fetch_rows(
    admin: Dict[str, Any],
    status: str,
    q: Optional[str],
    employee_type: Optional[str],
    company_id: Optional[str],
    is_onroll: Optional[str],
) -> tuple:
    query: Dict[str, Any] = {"role": "employee"}
    if admin["role"] == "company_admin":
        query["company_id"] = admin.get("company_id")
    elif company_id:
        query["company_id"] = company_id

    if employee_type:
        query["employee_type"] = employee_type
    if is_onroll in ("true", "false"):
        query["is_onroll"] = is_onroll == "true"
    if q:
        import re as _re
        rx = {"$regex": _re.escape(q.strip()), "$options": "i"}
        query["$and"] = [{"$or": [
            {"name": rx}, {"employee_code": rx}, {"phone": rx},
        ]}]

    users = await db.users.find(query, {"_id": 0}).sort("name", 1).to_list(5000)

    # Iter 170 (user bug) — "Active" must exclude EVERY resigned/exit marker:
    # exit_date, resign_date, date_of_leaving, leaving_date or an
    # employment_status of exited/resigned/terminated/inactive/left.
    def _is_resigned(u: Dict[str, Any]) -> bool:
        if (u.get("exit_date") or u.get("resign_date")
                or u.get("date_of_leaving") or u.get("leaving_date")):
            return True
        return str(u.get("employment_status") or "").strip().lower() in (
            "exited", "resigned", "terminated", "inactive", "left")

    if status == "active":
        users = [u for u in users if not _is_resigned(u)]
    elif status == "left":
        users = [u for u in users if _is_resigned(u)]

    # Firm names for display
    cids = {u.get("company_id") for u in users if u.get("company_id")}
    names: Dict[str, str] = {}
    if cids:
        async for c in db.companies.find(
            {"company_id": {"$in": list(cids)}}, {"_id": 0, "company_id": 1, "name": 1},
        ):
            names[c["company_id"]] = c.get("name") or c["company_id"]

    # Iter 331 (user request) — salary heads from the Employee Master.
    # HRA / Conv. are fixed columns; every OTHER allowance head that is
    # set on any employee's master becomes its own dynamic column.
    def _split_allowances(u: Dict[str, Any]) -> Dict[str, Any]:
        hra = _num0(u.get("hra_amount"))
        conv = _num0(u.get("conv_amount"))
        extra: Dict[str, float] = {}
        struct_basic = 0.0
        allow_rows = list(u.get("compliance_salary_allowances") or [])
        # Employee Master structure rows (Salary Update modal) hold the
        # heads too — Basic feeds the Basic column, the rest are allowances.
        for r in (u.get("salary_structure_compliance") or []):
            if not isinstance(r, dict):
                continue
            h = str(r.get("head") or "").strip()
            if "employer" in h.lower():
                continue
            if h.lower().startswith("basic"):
                struct_basic += _num0(r.get("amount"))
            else:
                allow_rows.append(r)
        if struct_basic <= 0:
            for r in (u.get("salary_structure_actual") or []):
                if isinstance(r, dict) and str(r.get("head") or "").strip().lower().startswith("basic"):
                    struct_basic += _num0(r.get("amount"))
        for r in allow_rows:
            if not isinstance(r, dict):
                continue
            head = str(r.get("head") or "").strip()
            amt = _num0(r.get("amount"))
            if not head or amt <= 0:
                continue
            s = head.lower()
            if "hra" in s or "house" in s:
                hra += amt
            elif s.startswith("conv") or "travel" in s:
                conv += amt
            else:
                extra[head.upper()] = extra.get(head.upper(), 0.0) + amt
        return {"hra": hra, "conv": conv, "extra": extra, "struct_basic": struct_basic}

    extra_heads: Dict[str, str] = {}   # key → label (insertion sorted later)
    rows = []
    from utils.relation import father_or_spouse_display
    for u in users:
        al = _split_allowances(u)
        basic = _num0(u.get("compliance_basic")) or _num0(u.get("basic_salary")) \
            or _num0(u.get("basic_amount")) or al["struct_basic"]
        allow_sum = al["hra"] + al["conv"] + sum(al["extra"].values())
        gross = _num0(u.get("compliance_gross")) or (
            (basic + allow_sum) if (basic or allow_sum) else _num0(u.get("salary_monthly")))
        row = {
            **{k: u.get(k) for k, _ in _COLUMNS if k != "company_name"},
            # User directive — Female+Unmarried shows "D/O father", Female+
            # Married shows spouse name only.
            "father_name": father_or_spouse_display(u),
            "aadhaar_no": u.get("aadhaar_no") or u.get("aadhar_number"),
            "pan_no": u.get("pan_no") or u.get("pan_number"),
            "exit_date": (u.get("exit_date") or u.get("resign_date")
                          or u.get("date_of_leaving") or u.get("leaving_date")),
            "basic": basic or None,
            "pf_basic": _num0(u.get("pf_basic")) or None,
            "hra": al["hra"] or None,
            "conveyance": al["conv"] or None,
            "monthly_gross": round(gross, 2) or None,
            "present_address": u.get("present_address") or u.get("address"),
            "company_name": names.get(u.get("company_id") or "", u.get("company_id")),
            "user_id": u.get("user_id"),
        }
        for head, amt in al["extra"].items():
            k = _allow_key(head)
            extra_heads[k] = head
            row[k] = round(amt, 2)
        rows.append(row)

    # Assemble the final column list: fixed sequence with dynamic heads
    # inserted after Conv. — the Exit/Resign column is DROPPED on Active.
    columns = []
    for k, lbl in _COLUMNS:
        if k == "exit_date" and status == "active":
            continue
        columns.append({"key": k, "label": lbl})
        if k == "conveyance":
            for ek in sorted(extra_heads, key=lambda x: extra_heads[x]):
                columns.append({"key": ek, "label": extra_heads[ek].title()})
    return columns, rows


def _parse_common(status: str) -> str:
    s = (status or "all").lower()
    if s not in ("active", "left", "all"):
        raise HTTPException(status_code=400, detail="status must be active | left | all")
    return s


@router.get("/master-data")
async def master_data_report(
    status: str = "all",
    q: Optional[str] = None,
    employee_type: Optional[str] = None,
    company_id: Optional[str] = None,
    is_onroll: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    s = _parse_common(status)
    columns, rows = await _fetch_rows(admin, s, q, employee_type, company_id, is_onroll)
    return {
        "status": s,
        "count": len(rows),
        "columns": columns,
        "rows": rows,
    }


def _build_xlsx(rows: list, s: str, columns: list) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = {"active": "Active Employees", "left": "Left Employees", "all": "All Employees"}[s]

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1E3A8A")
    ws.cell(row=1, column=1, value="SN").font = hdr_font
    ws.cell(row=1, column=1).fill = hdr_fill
    for ci, c in enumerate(columns, start=2):
        cell = ws.cell(row=1, column=ci, value=c["label"])
        cell.font = hdr_font
        cell.fill = hdr_fill
    for ri, r in enumerate(rows, start=2):
        ws.cell(row=ri, column=1, value=ri - 1)
        for ci, c in enumerate(columns, start=2):
            v = r.get(c["key"])
            if c["key"] == "is_onroll":
                v = "On-roll" if v is not False else "Off-roll"
            ws.cell(row=ri, column=ci, value=v if v is not None else "")
    for ci in range(1, len(columns) + 2):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/master-data.xlsx")
async def master_data_report_xlsx(
    status: str = "all",
    q: Optional[str] = None,
    employee_type: Optional[str] = None,
    company_id: Optional[str] = None,
    is_onroll: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    s = _parse_common(status)
    columns, rows = await _fetch_rows(admin, s, q, employee_type, company_id, is_onroll)
    buf = _build_xlsx(rows, s, columns)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="MasterData_{s}.xlsx"'},
    )


# ---------------------------------------------------------------------------
# Iter 92 — Monthly Master-Data e-mail (Resend)
# ---------------------------------------------------------------------------

async def _email_master_data_for_firm(company_id: str, month_label: str) -> dict:
    """Build the ALL-employees master xlsx for one firm and email it to
    the firm's company-admin emails (fallback RESEND_TO_EMAIL)."""
    import base64
    import os

    from utils.iter60_features import _send_email_with_attachment

    fake_admin = {"role": "super_admin"}
    columns, rows = await _fetch_rows(fake_admin, "all", None, None, company_id, None)
    if not rows:
        return {"delivered": False, "error": "no_employees"}
    buf = _build_xlsx(rows, "all", columns)

    admins = await db.users.find(
        {"role": "company_admin", "company_id": company_id, "email": {"$nin": [None, ""]}},
        {"_id": 0, "email": 1},
    ).to_list(20)
    emails = [a["email"] for a in admins if a.get("email")]
    fallback = os.getenv("RESEND_TO_EMAIL", "").strip()
    if not emails and fallback:
        emails = [fallback]
    if not emails:
        return {"delivered": False, "error": "no_recipient"}

    company = await db.companies.find_one({"company_id": company_id}, {"_id": 0, "name": 1})
    firm = (company or {}).get("name") or company_id
    return await _send_email_with_attachment(
        to_emails=emails,
        subject=f"Monthly Master Data — {firm} — {month_label}",
        text_body=(
            f"Attached is the monthly Employee Master Data report for {firm} "
            f"({month_label}). {len(rows)} employee record(s). "
            "This is an automated read-only export."
        ),
        attachments=[{
            "filename": f"MasterData_{firm.replace(' ', '_')}_{month_label}.xlsx",
            "content": base64.b64encode(buf.getvalue()).decode(),
        }],
    )


@router.post("/master-data/email")
async def email_master_data_now(
    company_id: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    """Send the Master Data Excel to the firm's admins right now."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = admin.get("company_id") if admin["role"] == "company_admin" else company_id
    if not cid:
        raise HTTPException(status_code=400, detail="company_id is required")
    from datetime import datetime
    r = await _email_master_data_for_firm(cid, datetime.utcnow().strftime("%Y-%m"))
    if not r.get("delivered"):
        raise HTTPException(status_code=502, detail=f"Email failed: {r.get('error')}")
    return {"ok": True, **r}


async def monthly_master_data_email_loop():
    """Background task — on the 1st of every month, email each active
    firm's admins the Employee Master Data Excel. Idempotent via a
    system_flags marker per month."""
    import asyncio
    import logging
    from datetime import datetime

    log = logging.getLogger("master-data-email")
    while True:
        try:
            now = datetime.utcnow()
            month_label = now.strftime("%Y-%m")
            if now.day == 1:
                flag = await db.system_flags.find_one(
                    {"key": "master_data_email", "month": month_label},
                )
                if not flag:
                    firms = await db.companies.find({}, {"_id": 0, "company_id": 1}).to_list(200)
                    sent = 0
                    for f in firms:
                        try:
                            r = await _email_master_data_for_firm(f["company_id"], month_label)
                            if r.get("delivered"):
                                sent += 1
                        except Exception as exc:  # noqa: BLE001
                            log.warning("master-data email failed for %s: %s", f["company_id"], exc)
                    await db.system_flags.insert_one(
                        {"key": "master_data_email", "month": month_label,
                         "sent": sent, "at": now.isoformat()},
                    )
                    log.info("monthly master-data emails sent: %s firms", sent)
        except Exception as exc:  # noqa: BLE001
            log.warning("master-data email loop error: %s", exc)
        await asyncio.sleep(6 * 3600)  # check 4×/day
