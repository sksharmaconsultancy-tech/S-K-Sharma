"""Iter 177 — Labour Law Compliance Reports Module (Phase A + B).

One generic engine powers 22 statutory attendance reports (Daily
Attendance Register, Muster Roll, OT Register, Late Coming, ... ) over
the shared dataset: employees × attendance punches × attendance policy.

Common filters: company, branch/worksite, contractor, department,
designation, employee category (employee_type), gender, shift,
month/year or explicit date range.

Every export carries the statutory header block: company logo + details,
generated date/time (IST) + generated-by, page numbers and a QR
verification code (stored in ``report_verifications``).

Endpoints:
  * GET  /api/admin/labour-reports/catalogue
  * POST /api/admin/labour-reports/generate   (format: json|csv|excel|pdf)
  * GET  /api/admin/labour-reports/verify/{verify_id}
"""
import base64
import csv
import io
import os
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException

from server import db, get_user_from_token, require_role, now_iso, compute_textile_day  # noqa: E402

router = APIRouter(prefix="/api/admin/labour-reports", tags=["labour-reports"])

IST = timezone(timedelta(hours=5, minutes=30))

# ---------------------------------------------------------------------------
# Report catalogue
# ---------------------------------------------------------------------------
CATALOGUE: List[Dict[str, str]] = [
    {"key": "daily_attendance", "label": "Daily Attendance Register", "group": "Registers"},
    {"key": "muster_roll", "label": "Muster Roll (Form 25 style)", "group": "Registers"},
    {"key": "monthly_register", "label": "Monthly Attendance Register", "group": "Registers"},
    {"key": "overtime_register", "label": "Overtime Register", "group": "Registers"},
    {"key": "present_absent", "label": "Present / Absent Report", "group": "Daily Reports"},
    {"key": "late_coming", "label": "Late Coming Report", "group": "Daily Reports"},
    {"key": "early_going", "label": "Early Going Report", "group": "Daily Reports"},
    {"key": "miss_punch", "label": "Miss Punch Report", "group": "Daily Reports"},
    {"key": "in_out_punch", "label": "In-Out Punch Report", "group": "Daily Reports"},
    {"key": "half_day", "label": "Half Day Report", "group": "Daily Reports"},
    {"key": "shift_report", "label": "Shift Report", "group": "Shift Reports"},
    {"key": "shift_deployment", "label": "Shift Deployment Report", "group": "Shift Reports"},
    {"key": "dummy_shift", "label": "Dummy Shift Report", "group": "Shift Reports"},
    {"key": "night_shift", "label": "Night Shift Report", "group": "Shift Reports"},
    {"key": "double_shift", "label": "Double Shift Report", "group": "Shift Reports"},
    {"key": "weekly_off", "label": "Weekly Off Worked Report", "group": "Shift Reports"},
    {"key": "holiday_attendance", "label": "Holiday Attendance Report", "group": "Shift Reports"},
    {"key": "department_wise", "label": "Department Wise Report", "group": "Summary Reports"},
    {"key": "contractor_wise", "label": "Contractor Wise Report", "group": "Summary Reports"},
    {"key": "geofence_attendance", "label": "Geofence Attendance Report", "group": "Technology Reports"},
    {"key": "gps_attendance", "label": "GPS Attendance Report", "group": "Technology Reports"},
    {"key": "face_attendance", "label": "Face Recognition Attendance", "group": "Technology Reports"},
    {"key": "qr_attendance", "label": "QR Attendance Report", "group": "Technology Reports"},
    {"key": "biometric_attendance", "label": "Biometric (Device) Attendance", "group": "Technology Reports"},
    {"key": "device_wise", "label": "Device Wise Attendance", "group": "Technology Reports"},
    {"key": "location_wise", "label": "Location / Worksite Wise Attendance", "group": "Technology Reports"},
]
CAT_KEYS = {c["key"] for c in CATALOGUE}

# Iter 215 — fixed Dummy Shift master (report-only shifts assigned per
# employee from the Employee Master when the firm's Attendance Policy has
# "Dummy Shift Allowed" switched on).
DUMMY_SHIFTS = [
    {"name": "SHIFT A1", "start": "07:00", "end": "15:00"},
    {"name": "SHIFT B1", "start": "15:00", "end": "23:00"},
    {"name": "SHIFT C1", "start": "23:00", "end": "07:00"},
    {"name": "SHIFT A", "start": "08:00", "end": "16:00"},
    {"name": "SHIFT B", "start": "16:00", "end": "00:00"},
    {"name": "SHIFT C", "start": "00:00", "end": "08:00"},
    {"name": "GENERAL SHIFT", "start": "10:00", "end": "06:00"},
]
DUMMY_SHIFT_NAMES = {s["name"] for s in DUMMY_SHIFTS}


def _shift_display(e: dict) -> str:
    """Employee's assigned shift for reports: Shift-Master name first,
    then their own timing."""
    s = (e.get("shift_name") or "").strip()
    if s:
        return s
    ss, se = e.get("shift_start"), e.get("shift_end")
    return f"{ss} – {se}" if ss and se else (ss or se or "")


async def _auth(authorization: Optional[str], company_id: str):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    if admin.get("role") == "company_admin" and admin.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Not authorised for this firm")
    return admin


# ---------------------------------------------------------------------------
# Shared dataset loader
# ---------------------------------------------------------------------------
def _hhmm(iso: Optional[str]) -> str:
    return iso[11:16] if iso and len(iso) >= 16 else ""


def _mins(hhmm: str) -> Optional[int]:
    try:
        h, m = hhmm.split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return None


def _dates_between(d1: str, d2: str) -> List[str]:
    a = date.fromisoformat(d1)
    b = date.fromisoformat(d2)
    out = []
    while a <= b:
        out.append(a.isoformat())
        a += timedelta(days=1)
    return out


async def _load_dataset(company_id: str, filters: Dict[str, Any]):
    """Employees (after master filters) + attendance day summaries."""
    # --- date range ---
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    month = filters.get("month")  # "YYYY-MM"
    if month and not (from_date and to_date):
        y, m = int(month[:4]), int(month[5:7])
        from_date = f"{y:04d}-{m:02d}-01"
        nxt = date(y + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1)
        to_date = (nxt - timedelta(days=1)).isoformat()
    if not from_date or not to_date:
        raise HTTPException(status_code=400, detail="Provide month or from_date/to_date")
    if (date.fromisoformat(to_date) - date.fromisoformat(from_date)).days > 62:
        raise HTTPException(status_code=400, detail="Date range too large (max 62 days)")

    # --- employees ---
    q: Dict[str, Any] = {"company_id": company_id, "role": "employee"}
    for f_key, u_key in (
        ("department", "department"), ("designation", "designation"),
        ("employee_category", "employee_type"), ("gender", "gender"),
        ("contractor", "contractor_name"),
    ):
        v = (filters.get(f_key) or "").strip()
        if v:
            q[u_key] = {"$regex": f"^{v}$", "$options": "i"}
    emps = await db.users.find(q, {
        "_id": 0, "user_id": 1, "name": 1, "employee_code": 1, "department": 1,
        "designation": 1, "employee_type": 1, "gender": 1, "contractor_name": 1,
        "shift_start": 1, "shift_end": 1, "shift_name": 1, "dummy_shift": 1,
        "father_name": 1, "is_contractual": 1,
        # Iter 525 — Shift Deployment: policy-true hours/OT + day cost need
        # the employee's pay + policy-override fields.
        "employee_policy": 1, "compliance_gross": 1, "salary_monthly": 1,
        "salary_structure_actual": 1, "salary_structure_compliance": 1,
        "compliance_salary_mode": 1, "ot_applicable": 1,
        "attendance_policy_override": 1, "week_off_full_day": 1,
        "week_off_govt_holiday_enabled": 1,
    }).sort("employee_code", 1).to_list(5000)
    shift_f = (filters.get("shift") or "").strip()
    if shift_f:
        emps = [e for e in emps
                if _shift_display(e) == shift_f
                or (e.get("dummy_shift") or "") == shift_f
                or f"{e.get('shift_start') or ''}-{e.get('shift_end') or ''}" == shift_f
                or (e.get("shift_start") or "") == shift_f]
    uids = [e["user_id"] for e in emps]

    # --- attendance (approved only — statutory registers) ---
    aq: Dict[str, Any] = {
        "company_id": company_id, "user_id": {"$in": uids},
        "date": {"$gte": from_date, "$lte": to_date},
        "kind": {"$in": ["in", "out"]},
        "status": {"$nin": ["rejected", "pending"]},
    }
    branch = (filters.get("branch_id") or "").strip()
    if branch:
        aq["$or"] = [{"branch_id": branch}, {"worksite_id": branch}]
    recs_by: Dict[tuple, List[dict]] = defaultdict(list)
    async for r in db.attendance.find(aq, {
        "_id": 0, "user_id": 1, "date": 1, "kind": 1, "at": 1, "source": 1,
        "biometric_method": 1, "distance_m": 1, "outside_geofence": 1,
        "latitude": 1, "longitude": 1, "device_info": 1,
        "branch_name": 1, "worksite_name": 1, "gps_verified": 1,
    }).sort("at", 1):
        recs_by[(r["user_id"], r["date"])].append(r)

    # --- policy --- (the Attendance Policy screen edits the firm's
    # companies.attendance_policy — prefer it so Hours / OT Hours follow
    # the company attendance policy; fall back to the legacy collection.)
    comp = await db.companies.find_one(
        {"company_id": company_id}, {"_id": 0, "attendance_policy": 1})
    policy = (comp or {}).get("attendance_policy") or {}
    # Iter 532 (user request) — Late/Early measured against the worker's
    # ACTUAL Shift Master timing (employee's own timing → assigned Shift
    # Master → firm default).
    policy["_shift_map"] = {
        s["name"].strip(): (s.get("start") or "", s.get("end") or "")
        async for s in db.shift_masters.find(
            {}, {"_id": 0, "name": 1, "start": 1, "end": 1})
        if (s.get("name") or "").strip()
    }
    if not policy:
        policy = await db.attendance_policies.find_one(
            {"company_id": company_id}, {"_id": 0}) or {}

    dates = _dates_between(from_date, to_date)
    return emps, recs_by, policy, dates, from_date, to_date


def _day_summary(recs: List[dict], policy: dict, emp: dict) -> dict:
    ins = [r for r in recs if r["kind"] == "in"]
    outs = [r for r in recs if r["kind"] == "out"]
    first_in = _hhmm(ins[0]["at"]) if ins else ""
    last_out = _hhmm(outs[-1]["at"]) if outs else ""
    hours = 0.0
    # pair sequentially
    pairs = 0
    stack = None
    for r in recs:
        if r["kind"] == "in":
            stack = r
        elif r["kind"] == "out" and stack is not None:
            m1, m2 = _mins(_hhmm(stack["at"])), _mins(_hhmm(r["at"]))
            if m1 is not None and m2 is not None and m2 >= m1:
                hours += (m2 - m1) / 60.0
                pairs += 1
            stack = None
    # Iter 532 (user request) — Late/Early against the worker's ACTUAL
    # shift: employee's own timing → assigned Shift Master timing → firm
    # default. Cross-midnight (night) shifts handled with wrap-around
    # math; workers with NO known shift never get phantom 600+ min lates.
    _sm = (policy.get("_shift_map") or {}).get(
        (emp.get("shift_name") or "").strip())
    shift_start = (emp.get("shift_start") or (_sm and _sm[0])
                   or policy.get("shift_start") or "09:00")
    shift_end = (emp.get("shift_end") or (_sm and _sm[1])
                 or policy.get("shift_end") or "18:00")
    _known_shift = bool(emp.get("shift_start") or (_sm and _sm[0]))
    grace = int(policy.get("grace_minutes_late") or 0)
    late_by = 0
    if first_in:
        a, b = _mins(first_in), _mins(shift_start)
        if a is not None and b is not None:
            diff = (a - b) % 1440  # wrap-aware (night shifts)
            if grace < diff <= 720 and (_known_shift or diff <= 360):
                late_by = diff
    early_by = 0
    if last_out:
        a, b = _mins(last_out), _mins(shift_end)
        if a is not None and b is not None:
            diff = (b - a) % 1440  # wrap-aware (night shifts)
            if 0 < diff <= 720 and (_known_shift or diff <= 360):
                early_by = diff
    full_h = float(policy.get("full_day_hours") or policy.get("standard_working_hours") or 8.0)
    half_h = float(policy.get("half_day_hours") or 4.0)
    # Iter 215 — OT threshold follows the EMPLOYEE's own duty hours when
    # they have a shift timing assigned (Employee Master), else the
    # company Attendance Policy threshold.
    emp_dur = None
    _ss, _se = _mins(emp.get("shift_start") or ""), _mins(emp.get("shift_end") or "")
    if _ss is not None and _se is not None and _ss != _se:
        emp_dur = ((_se - _ss) % (24 * 60)) / 60.0
    ot_threshold = emp_dur or float(policy.get("overtime_threshold_hours") or full_h)
    ot_hours = max(0.0, hours - ot_threshold) if hours else 0.0
    status = "A"
    if recs:
        status = "P" if hours >= half_h or (ins and not outs) else "HD"
        if hours and hours < half_h:
            status = "HD"
    miss = bool((ins and not outs) or (outs and not ins))
    return {
        "first_in": first_in, "last_out": last_out, "hours": round(hours, 2),
        "pairs": pairs, "late_by": late_by, "early_by": early_by,
        "ot_hours": round(ot_hours, 2), "status": status, "miss": miss,
        "recs": recs,
    }


# ---------------------------------------------------------------------------
# Report builders — return (columns, rows)
# ---------------------------------------------------------------------------
def _emp_cols(e: dict) -> List[str]:
    return [str(e.get("employee_code") or ""), e.get("name") or "",
            e.get("department") or "", e.get("designation") or ""]


EMP_HEAD = ["Code", "Employee Name", "Department", "Designation"]


def _rate_of(e: dict) -> tuple:
    """(rate, mode) as per Employee Policy / Master — mirrors the
    compliance salary engine's rate resolution."""
    epol = e.get("employee_policy") or {}
    mode = str(e.get("compliance_salary_mode") or "").strip().lower()
    if mode not in ("daily", "hourly", "monthly"):
        mode = ""
    try:
        rate = float(epol.get("salary") or e.get("compliance_gross")
                     or e.get("salary_monthly") or 0)
    except (TypeError, ValueError):
        rate = 0.0
    if rate <= 0 or not mode:
        for src in ("salary_structure_compliance", "salary_structure_actual"):
            for r_ in (e.get(src) or []):
                if isinstance(r_, dict) and str(
                        r_.get("head", "")).strip().lower().startswith("basic"):
                    try:
                        amt = float(r_.get("amount") or 0)
                    except (TypeError, ValueError):
                        amt = 0.0
                    if rate <= 0 and amt > 0:
                        rate = amt
                    rt = str(r_.get("rate_type") or "").strip().lower()
                    if not mode and rt in ("daily", "hourly", "monthly"):
                        mode = rt
                    break
            if rate > 0 and mode:
                break
    return rate, (mode or "monthly")


def _day_cost(e: dict, policy: dict, d: str, present_days: float,
              hours: float, ot_hours: float) -> float:
    """Iter 525/526 — that day's pay for the employee as per the Firm
    Master + Employee Policy: base (Monthly/Daily/Hourly rate basis) +
    OT at the policy OT multiplier."""
    from calendar import monthrange
    rate, mode = _rate_of(e)
    full_h = float(policy.get("full_day_hours")
                   or policy.get("standard_working_hours") or 8.0)
    epol = e.get("employee_policy") or {}
    try:
        ot_mult = float(epol.get("ot_multiplier")
                        or policy.get("ot_multiplier") or 1.5)
    except (TypeError, ValueError):
        ot_mult = 1.5
    dim = monthrange(int(d[:4]), int(d[5:7]))[1]
    if mode == "daily":
        base = rate * present_days
        phr = rate / full_h if full_h else 0.0
    elif mode == "hourly":
        base = rate * hours
        phr = rate
    else:  # monthly
        base = rate * present_days / dim
        phr = rate / (dim * full_h) if full_h else 0.0
    return round(base + ot_hours * phr * ot_mult, 2)


def _present_of_status(status: str) -> float:
    return 1.0 if status == "P" else 0.5 if status == "HD" else 0.0


def _compliance_8hr_on(policy: dict) -> bool:
    """Iter 528 (user bug) — firm policy "Count Present Day @ 8 HRS —
    Compliance Salary only" is ON and Salary Allowed includes Compliance."""
    pm = policy.get("policy_master") or {}
    return bool(pm.get("compliance_present_8hr")
                and str(policy.get("salary_allowed") or "both")
                in ("compliance", "both"))


def _ot_allowed(e: dict, policy: dict) -> bool:
    """Same OT gates as the attendance engine: firm gate → per-employee
    override → legacy ot_applicable (default ON)."""
    if policy.get("firm_ot_allowed") is False:
        return False
    ov = e.get("attendance_policy_override") or {}
    if ov.get("ot_allowed") is not None:
        return bool(ov["ot_allowed"])
    v = e.get("ot_applicable")
    return True if v is None else bool(v)


def _apply_compliance_8hr(policy: dict, e: dict, eng: dict) -> tuple:
    """Iter 528 — when the firm's "Count Present Day @ 8 HRS" policy is
    YES: 8+ WORKED hrs = 1 Present Day, duty shown capped at 8 hrs, extra
    worked hrs → OT. Mirrors the Monthly Attendance Grid's
    ``_pm_8hr_reports`` rule so all reports agree.

    NOTE: ``compute_textile_day`` returns duty_hours = TOTAL worked hours
    (OT is a subset of it), so the split here is off the duty figure.
    Week-off / holiday special days ("all hours → OT") stay untouched.
    """
    hours, ot, present = (eng["duty_hours"], eng["ot_hours"],
                          eng["present_days"])
    if not _compliance_8hr_on(policy):
        return hours, ot, present
    worked = round(hours, 2)
    # Week-off / holiday SPECIAL transformations (week-off worked modes,
    # "all hours → OT", holiday-present rules) keep their own OT/present
    # math — only the DISPLAYED duty is capped at 8 hrs so Hours never
    # show above the policy full day.
    special = any(
        ("week_off" in (n_l := str(n).lower()) or "week-off" in n_l
         or "holiday" in n_l)
        for n in eng.get("notes") or [])
    if special:
        if worked > 0 and ot >= worked:  # all hours went to OT
            return 0.0, ot, present
        return round(min(worked, 8.0), 2), ot, present
    duty = round(min(worked, 8.0), 2)
    extra = round(max(0.0, worked - 8.0), 2)
    ot_new = extra if _ot_allowed(e, policy) else 0.0
    half_h = float(policy.get("half_day_hours") or 4.0)
    p = 1.0 if worked >= 8.0 else 0.5 if worked >= half_h else 0.0
    return duty, ot_new, p


def build_report(key: str, emps, recs_by, policy, dates,
                 run_by: Optional[dict] = None) -> tuple:
    weekly_offs = set(policy.get("weekly_off_days") or [])
    night_start = policy.get("night_shift_start") or "22:00"
    night_end = policy.get("night_shift_end") or "06:00"
    holidays = set(policy.get("holidays") or [])

    def day_rows(pred, extra_cols, extra_vals):
        cols = ["Date"] + EMP_HEAD + ["In", "Out", "Hours"] + extra_cols
        rows = []
        for d in dates:
            wd = date.fromisoformat(d).weekday()
            for e in emps:
                recs = recs_by.get((e["user_id"], d))
                if not recs:
                    continue
                s = _day_summary(recs, policy, e)
                # Iter 529 (user request) — Hours / OT / Status come from
                # the firm-master attendance-policy engine (incl. the
                # "Count Present Day @ 8 HRS" rule) so EVERY report in the
                # hub matches the Attendance Grid.
                eng = compute_textile_day(recs, policy, e, wd)
                if _compliance_8hr_on(policy):
                    h_, o_, p_ = _apply_compliance_8hr(policy, e, eng)
                else:
                    h_, o_, p_ = (eng["duty_hours"], eng["ot_hours"],
                                  eng["present_days"])
                    if o_ > 0 and h_ >= o_:
                        h_ = round(h_ - o_, 2)  # engine folds OT into duty
                s["hours"], s["ot_hours"] = h_, o_
                s["status"] = ("P" if p_ >= 1.0 else "HD" if p_ >= 0.5
                               else "OT" if o_ > 0
                               else "P" if h_ > 0 else s["status"])
                if not pred(s, d, e):
                    continue
                rows.append([d] + _emp_cols(e) + [s["first_in"], s["last_out"], s["hours"]]
                            + extra_vals(s, d, e))
        return cols, rows

    if key == "daily_attendance":
        # Iter 480 (user spec — Phase 2 statutory fields): + Contractor,
        # Shift, Punch Source, Late/Early minutes.
        def _shift_lbl(e):
            return (e.get("shift_name")
                    or (f"{e.get('shift_start')}-{e.get('shift_end')}"
                        if e.get("shift_start") and e.get("shift_end")
                        else ""))
        return day_rows(
            lambda s, d, e: True,
            ["Contractor", "Shift", "Source", "Late Min", "Early Out Min",
             "OT Hrs", "Status"],
            lambda s, d, e: [
                e.get("contractor_name") or "", _shift_lbl(e),
                ", ".join(sorted({(r.get("source") or "")[:14]
                                  for r in s["recs"]
                                  if r.get("source")})),
                s["late_by"] or "", s["early_by"] or "",
                s["ot_hours"], s["status"]])

    if key == "in_out_punch":
        cols = ["Date"] + EMP_HEAD + ["Punch Time", "Type", "Source", "Method"]
        rows = []
        for d in dates:
            for e in emps:
                for r in recs_by.get((e["user_id"], d), []):
                    rows.append([d] + _emp_cols(e) + [
                        _hhmm(r["at"]), r["kind"].upper(),
                        (r.get("source") or "")[:24], r.get("biometric_method") or ""])
        return cols, rows

    if key == "muster_roll":
        day_nums = [d[8:10] for d in dates]
        cols = EMP_HEAD[:2] + ["Father Name"] + day_nums + ["P", "HD", "A", "WO"]
        rows = []
        for e in emps:
            marks, p, hd, a, wo = [], 0, 0, 0, 0
            for d in dates:
                wd = date.fromisoformat(d).weekday()
                recs = recs_by.get((e["user_id"], d))
                if recs:
                    s = _day_summary(recs, policy, e)
                    marks.append(s["status"])
                    p += s["status"] == "P"
                    hd += s["status"] == "HD"
                elif wd in weekly_offs:
                    marks.append("WO")
                    wo += 1
                elif d in holidays:
                    marks.append("H")
                else:
                    marks.append("A")
                    a += 1
            rows.append(_emp_cols(e)[:2] + [e.get("father_name") or ""] + marks + [p, hd, a, wo])
        return cols, rows

    if key == "monthly_register":
        # Iter 687 (user request) — Monthly Attendance Register now shows
        # the FINALIZED Salary-Process figures only: Salary Process Days,
        # Attendance (engine present days) and OT Hours. Other columns
        # removed.
        run_map = run_by or {}
        cols = EMP_HEAD + ["Salary Process Days", "Attendance", "OT Hours"]
        rows = []
        for e in emps:
            att_p = 0.0
            for d in dates:
                recs = recs_by.get((e["user_id"], d))
                if recs:
                    eng = compute_textile_day(
                        recs, policy, e, date.fromisoformat(d).weekday())
                    _h, _o, pd_ = _apply_compliance_8hr(policy, e, eng)
                    att_p += 1.0 if pd_ >= 1.0 else (0.5 if pd_ == 0.5 else 0)
            rr = run_map.get(e["user_id"]) or {}
            rows.append(_emp_cols(e) + [
                round(float(rr.get("present_days") or 0), 1),
                round(att_p, 1),
                round(float(rr.get("ot_hours") or 0), 1)])
        return cols, rows

    if key == "present_absent":
        cols = ["Date", "Present", "Half Day", "Absent", "Present Names (codes)"]
        rows = []
        for d in dates:
            p_list, hd_c, a_c = [], 0, 0
            for e in emps:
                recs = recs_by.get((e["user_id"], d))
                if recs:
                    s = _day_summary(recs, policy, e)
                    if s["status"] == "P":
                        p_list.append(str(e.get("employee_code") or e.get("name")))
                    else:
                        hd_c += 1
                else:
                    a_c += 1
            rows.append([d, len(p_list), hd_c, a_c, ", ".join(p_list[:40])])
        return cols, rows

    if key == "late_coming":
        return day_rows(lambda s, d, e: s["late_by"] > 0, ["Late By (min)"],
                        lambda s, d, e: [s["late_by"]])
    if key == "early_going":
        return day_rows(lambda s, d, e: s["early_by"] > 0 and s["last_out"], ["Early By (min)"],
                        lambda s, d, e: [s["early_by"]])
    if key == "miss_punch":
        return day_rows(lambda s, d, e: s["miss"], ["Missing"],
                        lambda s, d, e: ["OUT missing" if s["first_in"] and not s["last_out"] else "IN missing"])
    if key == "half_day":
        return day_rows(lambda s, d, e: s["status"] == "HD", ["Shortfall"],
                        lambda s, d, e: [f"{s['hours']}h worked"])
    if key == "overtime_register":
        # Iter 526/529 — OT from the policy engine; Cost = that day's pay
        # (base + OT at the policy multiplier) per Firm Master + Employee
        # Policy. "Hours" col = regular duty; Total Worked = duty + OT.
        return day_rows(lambda s, d, e: s["ot_hours"] > 0,
                        ["OT Hours", "Total Worked Hrs", "Cost"],
                        lambda s, d, e: [
                            s["ot_hours"], round(s["hours"] + s["ot_hours"], 2),
                            _day_cost(e, policy, d, _present_of_status(s["status"]),
                                      s["hours"], s["ot_hours"])])
    if key == "double_shift":
        return day_rows(lambda s, d, e: s["pairs"] >= 2, ["Shifts (in-out cycles)"],
                        lambda s, d, e: [s["pairs"]])
    if key == "shift_report":
        # Iter 214 — Shift / live-muster report (user spec): everyone who
        # punched in, with their shift (from the Shift Master), punch-in
        # time and a blank Signature column — usable as an emergency
        # evacuation / muster roll for factories, hospitals, warehouses,
        # security teams etc. Employees currently on OT (second in-out
        # cycle after a morning first punch) get an "OT —" marker in
        # front of their name.
        multi = len(dates) > 1
        cols = ((["Date"] if multi else [])
                + ["Shift (From Master)", "Code", "Employee Name",
                   "Department", "Designation", "Punch In Time", "Signature"])
        rows = []
        for d in dates:
            for e in emps:
                recs = recs_by.get((e["user_id"], d))
                if not recs:
                    continue
                s = _day_summary(recs, policy, e)
                if not s["first_in"]:
                    continue
                shift = _shift_display(e) or "—"
                name = e.get("name") or ""
                fi = _mins(s["first_in"])
                ins_count = sum(1 for r in recs if r["kind"] == "in")
                if (s["pairs"] >= 2 or ins_count >= 2) and fi is not None and fi < 720:
                    name = f"OT — {name}"
                rows.append(([d] if multi else [])
                            + [shift, str(e.get("employee_code") or ""), name,
                               e.get("department") or "", e.get("designation") or "",
                               s["first_in"], ""])
        rows.sort(key=lambda r: (r[0], r[1], r[2]) if multi else (r[0], r[1]))
        return cols, rows

    if key == "shift_deployment":
        # Iter 525 (user request) — redesigned:
        # • Shift / Shift Timing / Machine-Device / Date columns removed
        #   (date + firm show in the report heading instead).
        # • S.No. column added.
        # • Hours & OT Hrs come from the FIRM MASTER attendance-policy
        #   engine (compute_textile_day — same engine as the Attendance
        #   Grid / payroll).
        # • NEW "Cost" column — that day's pay for the employee as per the
        #   Firm Master + Employee Policy (rate basis monthly/daily/hourly,
        #   OT paid at the policy OT multiplier).
        # • Optional grouping (Department / Designation wise) with band
        #   rows + per-group subtotals — display and ALL downloads follow.

        group_by = str(policy.get("_group_by") or "").strip().lower()
        if group_by not in ("department", "designation"):
            group_by = ""

        cols = ["S.No.", "Code", "Employee Name", "Father Name", "Department",
                "Designation", "Contractor", "In", "Out", "Hours", "OT Hrs",
                "Cost", "Status"]
        data = []  # (sort_keys..., row_values)
        for d in dates:
            wd = date.fromisoformat(d).weekday()
            for e in emps:
                recs = recs_by.get((e["user_id"], d))
                if not recs:
                    continue
                s = _day_summary(recs, policy, e)
                eng = compute_textile_day(recs, policy, e, wd)
                # Iter 528 — honour "Count Present Day @ 8 HRS" firm policy.
                hours, ot, pd_ = _apply_compliance_8hr(policy, e, eng)
                # Iter 530 (user request) — Cost shown in ROUND figures.
                cost = round(_day_cost(e, policy, d, pd_, hours, ot))
                # Status follows the SAME policy engine as Hours/OT.
                status = ("P" if pd_ >= 1.0 else "HD" if pd_ >= 0.5
                          else "OT" if ot > 0 else "P" if hours > 0
                          else s["status"])
                grp = ((e.get(group_by) or "— Not Set —") if group_by else "")
                data.append((grp, str(e.get("employee_code") or ""), d, [
                    0, str(e.get("employee_code") or ""), e.get("name") or "",
                    e.get("father_name") or "", e.get("department") or "",
                    e.get("designation") or "", e.get("contractor_name") or "",
                    s["first_in"], s["last_out"], hours, ot, cost, status,
                ]))
        data.sort(key=lambda x: (x[0], x[1], x[2]))

        # Iter 627 (user request) — "Summary Only" mode: NO individual
        # employee rows. Two sections one after another — Department-wise
        # then Designation-wise — each group row shows headcount + full
        # attendance totals (Present / Half-Day day counts, Hours, OT,
        # Cost). Display and ALL downloads (PDF/Excel/CSV) follow.
        # Iter 652 (user request) — summary shows EITHER Department-wise OR
        # Designation-wise (user picks one), not both forced together.
        if policy.get("_summary_only"):
            sg = str(policy.get("_summary_group") or "").strip().lower()
            if sg not in ("department", "designation"):
                sg = "department"

            def _agg(idx: int) -> Dict[str, dict]:
                m: Dict[str, dict] = {}
                for _g, code_, _d, r in data:
                    g = str(r[idx] or "").strip() or "— Not Set —"
                    a = m.setdefault(g, {"emps": set(), "p": 0, "hd": 0,
                                         "h": 0.0, "o": 0.0, "c": 0.0})
                    a["emps"].add(code_)
                    if r[12] == "HD":
                        a["hd"] += 1
                    else:
                        a["p"] += 1
                    a["h"] += r[9]; a["o"] += r[10]; a["c"] += r[11]
                return m

            s_cols = ["S.No.", sg.capitalize(), "Deployed",
                      "Present", "Half Day", "Hours", "OT Hrs", "Cost"]
            s_rows: list = []

            def _section(title: str, m: Dict[str, dict]) -> None:
                band = [""] * len(s_cols)
                band[0] = f"▶ {title}"
                s_rows.append(band)
                t_emps: set = set()
                t_p = t_hd = 0
                t_h = t_o = t_c = 0.0
                for n, g in enumerate(sorted(m), start=1):
                    a = m[g]
                    s_rows.append([n, g, len(a["emps"]), a["p"], a["hd"],
                                   round(a["h"], 2), round(a["o"], 2),
                                   round(a["c"])])
                    t_emps |= a["emps"]; t_p += a["p"]; t_hd += a["hd"]
                    t_h += a["h"]; t_o += a["o"]; t_c += a["c"]
                tr = [""] * len(s_cols)
                tr[0] = "GRAND TOTAL"
                tr[2] = len(t_emps); tr[3] = t_p; tr[4] = t_hd
                tr[5] = round(t_h, 2); tr[6] = round(t_o, 2); tr[7] = round(t_c)
                s_rows.append(tr)

            if data:
                if sg == "designation":
                    _section("DESIGNATION WISE SUMMARY", _agg(5))
                else:
                    _section("DEPARTMENT WISE SUMMARY", _agg(4))
            return s_cols, s_rows

        def _sum_row(label: str, n_emps: int, h: float, o: float, c: float) -> list:
            # Iter 526 (user request) — label sits in col 0 and the empty
            # cells up to "Hours" are MERGED in the PDF/Excel exports.
            r = [""] * len(cols)
            r[0] = f"{label}  ·  {n_emps} deployed"
            r[9] = round(h, 2)
            r[10] = round(o, 2)
            r[11] = round(c)  # Iter 530 — round-figure cost
            return r

        rows = []
        sno = 0
        g_h = g_o = g_c = 0.0
        g_emps: set = set()
        cur = object()
        c_h = c_o = c_c = 0.0
        c_emps: set = set()

        def _flush_group():
            nonlocal c_h, c_o, c_c, c_emps
            if group_by and c_emps:
                rows.append(_sum_row(f"SUBTOTAL — {cur}", len(c_emps), c_h, c_o, c_c))
            c_h = c_o = c_c = 0.0
            c_emps = set()

        for grp, code_, d, r in data:
            if group_by and grp != cur:
                _flush_group()
                cur = grp
                band = [""] * len(cols)
                # col 0 so the whole band row merges into ONE cell.
                band[0] = f"▶ {group_by.upper()}: {grp}"
                rows.append(band)
            sno += 1
            r[0] = sno
            rows.append(r)
            c_h += r[9]; c_o += r[10]; c_c += r[11]
            c_emps.add(code_)
            g_h += r[9]; g_o += r[10]; g_c += r[11]
            g_emps.add(code_)
        _flush_group()
        if rows:
            rows.append(_sum_row("GRAND TOTAL", len(g_emps), g_h, g_o, g_c))
        return cols, rows

    if key == "dummy_shift":
        # Iter 215 — live-muster layout grouped by the fixed DUMMY shifts
        # assigned per employee in the Employee Master (report-only,
        # optional per firm via Attendance Policy → Dummy Shift Allowed).
        multi = len(dates) > 1
        cols = ((["Date"] if multi else [])
                + ["Dummy Shift", "Code", "Employee Name", "Department",
                   "Designation", "Punch In Time", "Signature"])
        timing = {s["name"]: f"{s['start']} – {s['end']}" for s in DUMMY_SHIFTS}
        rows = []
        for d in dates:
            for e in emps:
                ds = (e.get("dummy_shift") or "").strip()
                if not ds:
                    continue
                recs = recs_by.get((e["user_id"], d))
                if not recs:
                    continue
                s = _day_summary(recs, policy, e)
                if not s["first_in"]:
                    continue
                name = e.get("name") or ""
                fi = _mins(s["first_in"])
                ins_count = sum(1 for r in recs if r["kind"] == "in")
                if (s["pairs"] >= 2 or ins_count >= 2) and fi is not None and fi < 720:
                    name = f"OT — {name}"
                dsl = f"{ds} ({timing[ds]})" if ds in timing else ds
                rows.append(([d] if multi else [])
                            + [dsl, str(e.get("employee_code") or ""), name,
                               e.get("department") or "", e.get("designation") or "",
                               s["first_in"], ""])
        rows.sort(key=lambda r: (r[0], r[1], r[2]) if multi else (r[0], r[1]))
        return cols, rows

    if key in ("department_wise", "contractor_wise"):
        # Iter 215 — grouped summaries; Hours / OT Hours / Normal Hours
        # follow the company Attendance Policy (and each employee's own
        # duty hours for the OT threshold).
        fld = "department" if key == "department_wise" else "contractor_name"
        label = "Department" if key == "department_wise" else "Contractor"
        agg: Dict[str, dict] = {}
        for e in emps:
            g = (e.get(fld) or "").strip() or "— Not set —"
            a = agg.setdefault(g, {"emps": set(), "days": 0, "hours": 0.0,
                                   "ot": 0.0, "cost": 0.0})
            for d in dates:
                recs = recs_by.get((e["user_id"], d))
                if not recs:
                    continue
                s = _day_summary(recs, policy, e)
                a["emps"].add(e["user_id"])
                a["days"] += 1 if s["status"] in ("P", "HD") else 0
                a["hours"] += s["hours"]
                a["ot"] += s["ot_hours"]
                # Iter 526 (user request) — group labour cost per policy.
                a["cost"] += _day_cost(e, policy, d,
                                       _present_of_status(s["status"]),
                                       s["hours"], s["ot_hours"])
        cols = [label, "Employees", "Days Present", "Total Hours",
                "OT Hours", "Normal Hours", "Cost"]
        rows = [[g, len(a["emps"]), a["days"], round(a["hours"], 1),
                 round(a["ot"], 1), round(a["hours"] - a["ot"], 1),
                 round(a["cost"], 2)]
                for g, a in sorted(agg.items()) if a["emps"]]
        return cols, rows
    if key == "night_shift":
        ns, ne = _mins(night_start) or 1320, _mins(night_end) or 360
        def is_night(s, d, e):
            m = _mins(s["first_in"])
            return m is not None and (m >= ns or m <= ne)
        return day_rows(is_night, ["Night Window"], lambda s, d, e: [f"{night_start}–{night_end}"])
    if key == "weekly_off":
        return day_rows(
            lambda s, d, e: date.fromisoformat(d).weekday() in weekly_offs,
            ["Day"], lambda s, d, e: [date.fromisoformat(d).strftime("%A")])
    if key == "holiday_attendance":
        return day_rows(lambda s, d, e: d in holidays, ["Holiday"],
                        lambda s, d, e: [d])
    if key == "geofence_attendance":
        cols = ["Date"] + EMP_HEAD + ["Punch", "Time", "Distance (m)", "Inside Geofence"]
        rows = []
        for d in dates:
            for e in emps:
                for r in recs_by.get((e["user_id"], d), []):
                    if r.get("distance_m") is None:
                        continue
                    rows.append([d] + _emp_cols(e) + [
                        r["kind"].upper(), _hhmm(r["at"]), r.get("distance_m"),
                        "No" if r.get("outside_geofence") else "Yes"])
        return cols, rows
    if key == "gps_attendance":
        cols = ["Date"] + EMP_HEAD + ["Punch", "Time", "Latitude", "Longitude", "GPS Verified"]
        rows = []
        for d in dates:
            for e in emps:
                for r in recs_by.get((e["user_id"], d), []):
                    if r.get("latitude") is None:
                        continue
                    rows.append([d] + _emp_cols(e) + [
                        r["kind"].upper(), _hhmm(r["at"]),
                        round(r["latitude"], 5), round(r["longitude"], 5),
                        "Yes" if r.get("gps_verified") else "No"])
        return cols, rows
    if key == "face_attendance":
        return day_rows(
            lambda s, d, e: any(r.get("biometric_method") == "face" for r in s["recs"]),
            ["Method"], lambda s, d, e: ["Face Recognition"])
    if key == "qr_attendance":
        cols = ["Date"] + EMP_HEAD + ["Punch", "Time", "Source"]
        rows = []
        for d in dates:
            for e in emps:
                for r in recs_by.get((e["user_id"], d), []):
                    if "qr" not in (r.get("source") or "").lower():
                        continue
                    rows.append([d] + _emp_cols(e) + [r["kind"].upper(), _hhmm(r["at"]), r.get("source")])
        return cols, rows
    if key == "biometric_attendance":
        cols = ["Date"] + EMP_HEAD + ["Punch", "Time", "Device Source"]
        rows = []
        for d in dates:
            for e in emps:
                for r in recs_by.get((e["user_id"], d), []):
                    src = (r.get("source") or "").lower()
                    if not ("zk" in src or "import" in src or "device" in src or "biometric" in src):
                        continue
                    rows.append([d] + _emp_cols(e) + [r["kind"].upper(), _hhmm(r["at"]), r.get("source")])
        return cols, rows
    if key == "device_wise":
        cols = ["Device / Source", "Punches", "Employees", "Dates Covered"]
        agg: Dict[str, dict] = defaultdict(lambda: {"punches": 0, "emps": set(), "dates": set()})
        for (uid, d), recs in recs_by.items():
            for r in recs:
                k2 = (r.get("source") or "app").split("_")[0][:32]
                agg[k2]["punches"] += 1
                agg[k2]["emps"].add(uid)
                agg[k2]["dates"].add(d)
        rows = [[k2, v["punches"], len(v["emps"]), len(v["dates"])]
                for k2, v in sorted(agg.items())]
        return cols, rows
    if key == "location_wise":
        cols = ["Location / Worksite", "Punches", "Employees", "Dates Covered"]
        agg2: Dict[str, dict] = defaultdict(lambda: {"punches": 0, "emps": set(), "dates": set()})
        for (uid, d), recs in recs_by.items():
            for r in recs:
                k2 = r.get("worksite_name") or r.get("branch_name") or "Main Office"
                agg2[k2]["punches"] += 1
                agg2[k2]["emps"].add(uid)
                agg2[k2]["dates"].add(d)
        rows = [[k2, v["punches"], len(v["emps"]), len(v["dates"])]
                for k2, v in sorted(agg2.items())]
        return cols, rows

    raise HTTPException(status_code=400, detail=f"Unknown report: {key}")


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------
def _csv_bytes(columns, rows) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(columns)
    for r in rows:
        w.writerow(r)
    return buf.getvalue().encode("utf-8-sig")


def _row_kind(row) -> str:
    """Iter 525 — grouped reports: '▶ …' band rows and SUBTOTAL / GRAND
    TOTAL rows get distinct styling in Excel + PDF exports."""
    for v in row:
        s = str(v or "")
        if s.startswith("▶"):
            return "band"
        if s.startswith("SUBTOTAL —") or s.startswith("GRAND TOTAL"):
            return "total"
    return ""


def _merge_end(row) -> int:
    """Iter 526 — band/total rows: merge col 0 up to (but not including)
    the first non-empty cell, so the label reads as ONE merged cell."""
    nz = [i for i, v in enumerate(row) if i > 0 and str(v or "").strip()]
    return (min(nz) - 1) if nz else len(row) - 1


def _excel_bytes(title, header_lines, columns, rows) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    r_i = 1
    for line in header_lines:
        ws.cell(row=r_i, column=1, value=line).font = Font(bold=(r_i <= 2), size=12 if r_i == 1 else 10)
        r_i += 1
    r_i += 1
    for c_i, c in enumerate(columns, start=1):
        cell = ws.cell(row=r_i, column=c_i, value=c)
        # Iter 526 (user request) — heading row plain WHITE (no blue fill).
        cell.font = Font(bold=True)
    _band_fill = PatternFill("solid", fgColor="475569")
    _tot_fill = PatternFill("solid", fgColor="E2E8F0")
    for row in rows:
        r_i += 1
        kind = _row_kind(row)
        for c_i, v in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=v)
            if kind == "band":
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = _band_fill
            elif kind == "total":
                cell.font = Font(bold=True)
                cell.fill = _tot_fill
        # Iter 526 — band/total label cells MERGED into one.
        if kind:
            me = _merge_end(row)
            if me > 0:
                ws.merge_cells(start_row=r_i, start_column=1,
                               end_row=r_i, end_column=me + 1)
        # Iter 214 — signature reports get taller rows to sign in.
        if "Signature" in columns:
            ws.row_dimensions[r_i].height = 26
    for c_i, col_name in enumerate(columns, start=1):
        letter = ws.cell(row=1, column=c_i).column_letter
        # Iter 530 (user request) — slim S.No. column.
        ws.column_dimensions[letter].width = (
            28 if col_name == "Signature" else 6 if col_name == "S.No." else 16)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _pdf_bytes(title, company, header_meta, columns, rows, verify_id,
               verify_url: Optional[str] = None) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.graphics.barcode import qr as rl_qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF

    buf = io.BytesIO()
    page = landscape(A4) if len(columns) > 8 else A4

    # Iter 526 (user request) — firm name (BIG), address, date/period and
    # the report title repeat on EVERY page header; "Verify:" text removed
    # (the QR still carries the verification code); Generated time prints
    # at the page bottom.
    logo_reader = None
    if company.get("logo_base64"):
        try:
            logo_reader = ImageReader(
                io.BytesIO(base64.b64decode(company["logo_base64"])))
        except Exception:
            logo_reader = None
    qr_widget = rl_qr.QrCodeWidget(verify_url or f"SKS-REPORT:{verify_id}")
    b = qr_widget.getBounds()
    dr = Drawing(16 * mm, 16 * mm,
                 transform=[16 * mm / (b[2] - b[0]), 0, 0, 16 * mm / (b[3] - b[1]), 0, 0])
    dr.add(qr_widget)

    def on_page(canvas, doc):
        canvas.saveState()
        top = page[1] - 6 * mm
        if logo_reader is not None:
            try:
                canvas.drawImage(logo_reader, 10 * mm, top - 16 * mm,
                                 16 * mm, 16 * mm,
                                 preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        renderPDF.draw(dr, canvas, page[0] - 26 * mm, top - 16 * mm)
        canvas.setFillColor(rl_colors.black)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(page[0] / 2, top - 5.5 * mm,
                                 company.get("name") or "")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(page[0] / 2, top - 9.5 * mm,
                                 company.get("address") or "")
        canvas.drawCentredString(page[0] / 2, top - 13 * mm, header_meta)
        canvas.setFont("Helvetica-Bold", 10.5)
        canvas.drawCentredString(page[0] / 2, top - 17.5 * mm, title)
        canvas.setStrokeColor(rl_colors.HexColor("#94A3B8"))
        canvas.setLineWidth(0.5)
        canvas.line(10 * mm, top - 19.5 * mm, page[0] - 10 * mm, top - 19.5 * mm)
        # footer
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(page[0] - 12 * mm, 8 * mm, f"Page {doc.page}")
        _gen = datetime.now(IST).strftime("%d-%m-%Y %I:%M %p")
        canvas.drawString(12 * mm, 8 * mm, f"Generated: {_gen} IST")
        # Iter 182 — brand punch line on every statutory report page
        canvas.setFont("Helvetica-Oblique", 7.5)
        canvas.setFillColorRGB(0.145, 0.388, 0.922)  # #2563EB
        canvas.drawCentredString(page[0] / 2, 8 * mm,
                                 '"Your Satisfaction is Our First Ambition"')
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=page,
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=28 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()

    story: list = []

    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=6.8, leading=8.2)
    data = [[Paragraph(f"<b>{c}</b>", cell_style) for c in columns]]
    extra_styles = []
    for r in rows[:4000]:
        kind = _row_kind(r)
        ri = len(data)
        if kind:
            data.append([Paragraph(f"<b>{v}</b>" if str(v or "") else "", cell_style)
                         for v in r])
            # Iter 526 — merge the label cells into ONE (band = whole row).
            extra_styles.append(("SPAN", (0, ri), (_merge_end(r), ri)))
            if kind == "band":
                extra_styles += [
                    ("BACKGROUND", (0, ri), (-1, ri), rl_colors.HexColor("#475569")),
                    ("TEXTCOLOR", (0, ri), (-1, ri), rl_colors.white),
                ]
            else:
                extra_styles.append(
                    ("BACKGROUND", (0, ri), (-1, ri), rl_colors.HexColor("#E2E8F0")))
        else:
            data.append([Paragraph(str(v), cell_style) for v in r])
    # Iter 214 — signature reports: give the Signature column a wide fixed
    # box and taller rows so it prints properly for physical sign-off.
    col_widths = None
    has_sig = "Signature" in columns
    has_sno = "S.No." in columns
    if has_sig or has_sno:
        # Iter 530 (user request) — slim S.No. column in the PDF too.
        avail = page[0] - 20 * mm
        sig_w = 34 * mm if has_sig else 0
        sno_w = 9 * mm if has_sno else 0
        n_other = len(columns) - int(has_sig) - int(has_sno)
        other = (avail - sig_w - sno_w) / max(1, n_other)
        col_widths = [other] * len(columns)
        if has_sig:
            col_widths[columns.index("Signature")] = sig_w
        if has_sno:
            col_widths[columns.index("S.No.")] = sno_w
    tbl = Table(data, repeatRows=1, colWidths=col_widths)
    tbl.setStyle(TableStyle([
        # Iter 526 (user request) — heading row WHITE (no blue background).
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.white),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.black),
        ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#94A3B8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F1F5F9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ] + ([
        ("TOPPADDING", (0, 1), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
    ] if has_sig else []) + extra_styles))
    story.append(tbl)
    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
@router.get("/catalogue")
async def catalogue(authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    return {"reports": CATALOGUE}


@router.get("/shift-options")
async def shift_options(company_id: str,
                        authorization: Optional[str] = Header(None)):
    """Iter 215 — shift filter choices: ONLY the shifts actually assigned
    to ACTIVE employees in the Employee Master, plus the fixed dummy
    shifts when the firm's Attendance Policy allows them."""
    await _auth(authorization, company_id)
    q = {"company_id": company_id, "role": "employee", "active": {"$ne": False}}
    shifts, dummy_used = set(), set()
    async for e in db.users.find(q, {
        "_id": 0, "shift_name": 1, "shift_start": 1, "shift_end": 1,
        "dummy_shift": 1,
    }):
        s = _shift_display(e)
        if s:
            shifts.add(s)
        ds = (e.get("dummy_shift") or "").strip()
        if ds:
            dummy_used.add(ds)
    comp = await db.companies.find_one(
        {"company_id": company_id},
        {"_id": 0, "attendance_policy.policy_master.dummy_shift_allowed": 1})
    allowed = bool((((comp or {}).get("attendance_policy") or {})
                    .get("policy_master") or {}).get("dummy_shift_allowed"))
    return {"shifts": sorted(shifts),
            "dummy_allowed": allowed,
            "dummy_shifts": [s["name"] for s in DUMMY_SHIFTS],
            "dummy_shifts_assigned": sorted(dummy_used),
            "dummy_master": DUMMY_SHIFTS}


@router.post("/generate")
async def generate(payload: Dict[str, Any] = Body(...),
                   authorization: Optional[str] = Header(None)):
    company_id = str(payload.get("company_id") or "").strip()
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id required")
    admin = await _auth(authorization, company_id)
    key = str(payload.get("report_key") or "").strip()
    if key not in CAT_KEYS:
        raise HTTPException(status_code=400, detail="Unknown report_key")
    fmt = str(payload.get("format") or "json").lower()
    filters = payload.get("filters") or {}

    # Iter 215 — the Shift / Dummy Shift muster reports are SINGLE-DAY
    # reports (live headcount / evacuation roll).
    if key in ("shift_report", "dummy_shift"):
        f_d = (filters.get("from_date") or "").strip()
        t_d = (filters.get("to_date") or "").strip() or f_d
        if not f_d or f_d != t_d:
            raise HTTPException(
                status_code=400,
                detail="This report is a single-day report — pick one date.")
        filters = {**filters, "from_date": f_d, "to_date": f_d, "month": None}
    # Iter 525 (user request) — Shift Deployment: NO month input — a single
    # day or an explicit From–To period only.
    if key == "shift_deployment":
        f_d = (filters.get("from_date") or "").strip()
        t_d = (filters.get("to_date") or "").strip() or f_d
        if not f_d:
            raise HTTPException(
                status_code=400,
                detail="Pick a date (single day) or a From–To period.")
        filters = {**filters, "from_date": f_d, "to_date": t_d, "month": None}
    if key == "dummy_shift":
        comp_pol = await db.companies.find_one(
            {"company_id": company_id},
            {"_id": 0, "attendance_policy.policy_master.dummy_shift_allowed": 1})
        allowed = bool((((comp_pol or {}).get("attendance_policy") or {})
                        .get("policy_master") or {}).get("dummy_shift_allowed"))
        if not allowed:
            raise HTTPException(
                status_code=400,
                detail=("Dummy Shift is not enabled for this firm. Switch on "
                        "'Dummy Shift Allowed' in the Attendance Policy first."))

    emps, recs_by, policy, dates, from_date, to_date = await _load_dataset(company_id, filters)
    # Iter 525 — Shift Deployment: pass the grouping choice (Department /
    # Designation wise) through to the builder.
    if key == "shift_deployment":
        policy["_group_by"] = str(filters.get("group_by") or "")
        # Iter 627 (user request) — Summary Only vs Full Data option.
        policy["_summary_only"] = bool(filters.get("summary_only"))
        # Iter 652 (user request) — Summary shows ONE grouping (dept OR desig).
        policy["_summary_group"] = str(filters.get("summary_group") or "")
        # Iter 526 (user bug — "Out Punch not showing") — In/Out punches
        # must match the ATTENDANCE REPORT (grid engine): re-fetch with a
        # ±1-day window, drop duplicate machine punches and STITCH
        # night-shift OUT punches (that land on the next calendar day)
        # back to the day the session started.
        from server import dedupe_close_punches, stitch_cross_day_ot
        _comp = await db.companies.find_one(
            {"company_id": company_id}, {"_id": 0, "attendance_config": 1})
        _uids = [e["user_id"] for e in emps]
        _qf = (date.fromisoformat(from_date) - timedelta(days=1)).isoformat()
        _qt = (date.fromisoformat(to_date) + timedelta(days=1)).isoformat()
        _aq: Dict[str, Any] = {
            "user_id": {"$in": _uids}, "date": {"$gte": _qf, "$lte": _qt},
            "kind": {"$in": ["in", "out"]}, "status": "approved",
        }
        _branch = (filters.get("branch_id") or "").strip()
        if _branch:
            _aq["$or"] = [{"branch_id": _branch}, {"worksite_id": _branch}]
        _by_user: Dict[str, Dict[str, list]] = {}
        async for r in db.attendance.find(
            _aq, {"_id": 0, "user_id": 1, "date": 1, "kind": 1, "at": 1,
                  "source": 1},
        ).sort([("user_id", 1), ("at", 1)]):
            _by_user.setdefault(r["user_id"], {}).setdefault(
                r["date"], []).append(r)
        recs_by = {}
        for _uid, _daymap in _by_user.items():
            _rep = stitch_cross_day_ot(dedupe_close_punches(
                _daymap, company_cfg=(_comp or {}).get("attendance_config")))
            for _dk, _plist in _rep.items():
                if from_date <= _dk <= to_date:
                    recs_by[(_uid, _dk)] = _plist
    # Iter 687 — Monthly Register reads FINALIZED salary-run figures.
    _run_by = None
    if key == "monthly_register":
        _mk = str(filters.get("month") or (dates[0][:7] if dates else ""))[:7]
        _run = await db.compliance_salary_runs.find_one(
            {"company_id": company_id, "month": _mk},
            {"_id": 0, "rows": 1}, sort=[("generated_at", -1)])
        _run_by = {r.get("user_id"): r for r in (_run or {}).get("rows") or []}
    columns, rows = build_report(key, emps, recs_by, policy, dates,
                                 run_by=_run_by)
    # Iter 530 (user request) — generic "Group Wise" formatting (all
    # reports; Shift Deployment has its own native grouping): ▶ band rows
    # + per-group SUBTOTALS of every numeric column + GRAND TOTAL. The
    # on-screen preview AND the PDF / Excel / CSV downloads all follow it.
    _gb = str(filters.get("group_by") or "").strip().lower()
    _gb_col = {"department": "Department", "designation": "Designation",
               "contractor": "Contractor"}.get(_gb)
    if key != "shift_deployment" and _gb_col and _gb_col in columns and rows:
        gi = columns.index(_gb_col)
        num_idx = [i for i in range(len(columns))
                   if i != gi and all(isinstance(r[i], (int, float))
                                      and not isinstance(r[i], bool)
                                      for r in rows)]
        groups: Dict[str, list] = {}
        for r in rows:
            groups.setdefault(str(r[gi] or "").strip() or "— Not Set —",
                              []).append(r)

        def _tot_row(lbl: str, rs: list) -> list:
            t: list = [""] * len(columns)
            t[0] = f"{lbl}  ·  {len(rs)} rows"
            for i in num_idx:
                t[i] = round(sum(float(r[i]) for r in rs), 2)
            return t

        new_rows: list = []
        for g in sorted(groups):
            band: list = [""] * len(columns)
            band[0] = f"▶ {_gb_col.upper()}: {g}"
            new_rows.append(band)
            new_rows.extend(groups[g])
            new_rows.append(_tot_row(f"SUBTOTAL — {g}", groups[g]))
        new_rows.append(_tot_row("GRAND TOTAL", rows))
        rows = new_rows
    label = next(c["label"] for c in CATALOGUE if c["key"] == key)
    # Iter 531 (user request) — grouped reports show the grouping name in
    # the heading of the PDF / Excel / preview too.
    _gb_label = str(filters.get("group_by") or "").strip().lower()
    if key == "shift_deployment" and filters.get("summary_only"):
        # Iter 652 — Summary heading shows the chosen grouping only.
        _sg = str(filters.get("summary_group") or "").strip().lower()
        if _sg not in ("department", "designation"):
            _sg = "department"
        label = f"{label} — {_sg.capitalize()} Wise Summary"
    elif _gb_label in ("department", "designation", "contractor"):
        label = f"{label} — {_gb_label.capitalize()} Wise"

    company = await db.companies.find_one(
        {"company_id": company_id},
        {"_id": 0, "name": 1, "address": 1, "logo_base64": 1})
    gen_at = datetime.now(IST).strftime("%d-%m-%Y %I:%M %p")
    gen_by = admin.get("name") or admin.get("email") or "Admin"
    verify_id = f"lrv_{uuid.uuid4().hex[:10]}"
    await db.report_verifications.insert_one({
        "verify_id": verify_id, "report_key": key, "report_label": label,
        "company_id": company_id, "company_name": (company or {}).get("name") or "",
        "from_date": from_date, "to_date": to_date, "rows": len(rows),
        "generated_by": admin.get("user_id"), "generated_by_name": gen_by,
        "generated_at": now_iso(),
    })
    # Iter 531 (user question) — the QR is the report AUTHENTICITY seal:
    # scanning it opens a public verification page proving this exact
    # print was generated by the portal (report, firm, period, row count,
    # generated when/by whom). Encode a real URL so any phone camera works.
    _base = (os.environ.get("APP_PUBLIC_URL") or "").rstrip("/")
    verify_url = (f"{_base}/api/admin/labour-reports/verify/{verify_id}"
                  if _base else f"SKS-REPORT:{verify_id}")
    # Iter 526 (user request) — PDF heading: single day shows "Date: …"
    # (not "Period: X to X"); Generated time moves to the page bottom and
    # "Generated by" is removed from the print.
    def _dmy(s: str) -> str:
        return f"{s[8:10]}-{s[5:7]}-{s[:4]}" if len(s or "") == 10 else s
    header_meta = (f"Date: {_dmy(from_date)}" if from_date == to_date
                   else f"Period: {_dmy(from_date)} to {_dmy(to_date)}")

    if fmt == "json":
        return {"report_key": key, "label": label, "columns": columns,
                "rows": rows[:2000], "total_rows": len(rows),
                "from_date": from_date, "to_date": to_date,
                "generated_at": gen_at, "generated_by": gen_by,
                "verify_id": verify_id}
    stem = f"{key}_{from_date}_{to_date}"
    if fmt == "csv":
        return {"filename": f"{stem}.csv",
                "file_base64": base64.b64encode(_csv_bytes(columns, rows)).decode()}
    if fmt == "excel":
        header_lines = [company.get("name") or "", label,
                        f"Period: {from_date} to {to_date}",
                        f"Generated: {gen_at} IST · By: {gen_by} · Verify: {verify_id}"]
        return {"filename": f"{stem}.xlsx",
                "file_base64": base64.b64encode(
                    _excel_bytes(label, header_lines, columns, rows)).decode()}
    if fmt == "pdf":
        return {"filename": f"{stem}.pdf",
                "file_base64": base64.b64encode(
                    _pdf_bytes(label, company or {}, header_meta, columns, rows, verify_id, verify_url)).decode()}
    raise HTTPException(status_code=400, detail="format must be json|csv|excel|pdf")


@router.get("/verify/{verify_id}")
async def verify_report(verify_id: str, format: Optional[str] = None):
    """Iter 531 — public verification page for the QR printed on every
    report. Scanning the QR opens this page: it proves the print is a
    GENUINE portal-generated report (which report, firm, period, rows,
    generated when and by whom). ?format=json returns machine-readable."""
    from fastapi.responses import HTMLResponse
    doc = await db.report_verifications.find_one({"verify_id": verify_id}, {"_id": 0})
    if format == "json":
        if not doc:
            raise HTTPException(status_code=404, detail="Unknown verification code")
        return {"ok": True, "verification": doc}
    if not doc:
        return HTMLResponse(status_code=404, content="""
<!doctype html><html><head><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Report Verification</title></head>
<body style="font-family:system-ui,Arial;background:#FEF2F2;margin:0;padding:24px;text-align:center">
<div style="max-width:420px;margin:40px auto;background:#fff;border:2px solid #DC2626;border-radius:14px;padding:28px">
<div style="font-size:44px">❌</div>
<h2 style="color:#DC2626;margin:8px 0">NOT A VALID REPORT CODE</h2>
<p style="color:#475569">This verification code was not issued by the S.K. Sharma &amp; Co. portal.
The document may be forged or tampered with.</p></div></body></html>""")

    def _dmy(s):
        s = str(s or "")
        return f"{s[8:10]}-{s[5:7]}-{s[:4]}" if len(s) >= 10 else s
    period = (_dmy(doc.get("from_date"))
              if doc.get("from_date") == doc.get("to_date")
              else f"{_dmy(doc.get('from_date'))} to {_dmy(doc.get('to_date'))}")
    rows_html = "".join(
        f"<tr><td style='padding:6px 10px;color:#64748B'>{k}</td>"
        f"<td style='padding:6px 10px;font-weight:700;color:#0F172A'>{v}</td></tr>"
        for k, v in [
            ("Report", doc.get("report_label") or doc.get("report_key")),
            ("Firm", doc.get("company_name") or doc.get("company_id")),
            ("Period", period),
            ("Data rows", doc.get("rows")),
            ("Generated by", doc.get("generated_by_name")),
            ("Generated at", _dmy(str(doc.get("generated_at") or "")[:10])
             + " " + str(doc.get("generated_at") or "")[11:16]),
            ("Verification code", verify_id),
        ])
    return HTMLResponse(content=f"""
<!doctype html><html><head><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Genuine Report — S.K. Sharma &amp; Co.</title></head>
<body style="font-family:system-ui,Arial;background:#F0FDF4;margin:0;padding:24px">
<div style="max-width:460px;margin:30px auto;background:#fff;border:2px solid #16A34A;border-radius:14px;padding:26px;text-align:center">
<div style="font-size:44px">✅</div>
<h2 style="color:#15803D;margin:8px 0">GENUINE REPORT</h2>
<p style="color:#475569;margin:0 0 14px">This document was generated by the S.K. Sharma &amp; Co. HRMS portal.</p>
<table style="margin:0 auto;text-align:left;border-collapse:collapse">{rows_html}</table>
<p style="color:#94A3B8;font-size:12px;margin-top:16px">"Your Satisfaction is Our First Ambition"</p>
</div></body></html>""")
