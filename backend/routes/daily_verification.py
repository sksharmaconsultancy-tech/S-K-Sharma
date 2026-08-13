"""Iter 420 (user request) — Daily In/Out & OT Verification Report.

For HR / Security / Supervisors to PHYSICALLY verify attendance against
biometric / mobile punch records for one day. Reuses the SAME compute
pipeline as the Attendance Grid + In/Out & OT Matrix
(``_compute_monthly_grid_data``) so every figure matches 1:1 — existing
modules are untouched.

Endpoints (firm-scoped: super_admin / sub_admin / company_admin):
  * GET  /api/admin/reports/daily-verification          — JSON (paginated)
  * POST /api/admin/reports/daily-verification/verify   — physical verify
  * GET  /api/admin/reports/daily-verification/employee — drill-down
  * GET  /api/admin/reports/daily-verification.xlsx     — Excel (colours)
  * GET  /api/admin/reports/daily-verification.csv      — CSV
  * GET  /api/admin/reports/daily-verification.pdf      — A4 (landscape /
        portrait via ?orientation=)
  * POST /api/admin/reports/daily-verification/email    — email XLSX+PDF
  * POST /api/admin/reports/daily-verification/whatsapp — WhatsApp PDF

Row colours (screen + exports):  red = missing punch · orange =
unapproved OT · yellow = late / early · blue = manual attendance ·
green = normal present · grey = absent / WO / holiday / leave.

Audit trail → ``attendance_audit_log`` (verifications + exports + sends).
Verification records → ``daily_verifications``.
"""
import csv
import io
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query
from fastapi.responses import Response

router = APIRouter(prefix="/api/admin/reports", tags=["daily-verification"])

ROW_COLORS = {          # hex fills shared by screen + xlsx + pdf
    "red": "FECACA",     # missing in/out punch
    "orange": "FED7AA",  # unapproved OT
    "yellow": "FEF08A",  # late arrival / early exit
    "blue": "DBEAFE",    # manual attendance
    "green": "DCFCE7",   # normal present
    "muted": "E2E8F0",   # absent / WO / holiday / leave
}

HEADERS = ["S.No.", "Emp Code", "Employee Name", "Father Name", "Department",
           "Designation",
           "Contractor", "Punch In", "Punch Out", "Work Hrs", "OT In",
           "OT Out", "OT Hrs", "Total Duty Hrs", "Attendance Status",
           "Signature", "Verified", "Verified By", "Remarks"]


def _now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _hm(hours: Any) -> str:
    try:
        h = float(hours or 0)
    except (TypeError, ValueError):
        return "-"
    if h <= 0:
        return "-"
    m = int(round(h * 60))
    return f"{m // 60:02d}:{m % 60:02d}"


def _to12(s: Any) -> str:
    """Iter 554 (user request) — convert an 'HH:MM' / 'HH:MM:SS' string to
    12-hour 'hh:MM AM/PM'. Non-time values ('-', '') pass through as-is."""
    txt = str(s or "").strip()
    if not txt or ":" not in txt:
        return txt
    try:
        parts = txt.split(":")
        h, m = int(parts[0]), int(parts[1])
        ampm = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        return f"{h12:02d}:{m:02d} {ampm}"
    except (ValueError, IndexError):
        return txt


def _apply_12h(data: dict) -> dict:
    """Iter 554 — rewrite all clock-time fields in the report rows to
    12-hour AM/PM format (durations like Work Hrs / OT Hrs stay HH:MM)."""
    for r in data.get("rows") or []:
        for k in ("punch_in", "punch_out", "ot_in", "ot_out"):
            r[k] = _to12(r.get(k))
    return data


async def _auth(authorization: Optional[str], company_id: Optional[str]):
    from server import get_user_from_token, require_role, sub_admin_can_touch_company
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if admin.get("role") == "company_admin":
        company_id = admin.get("company_id")
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    if admin.get("role") == "sub_admin" and not sub_admin_can_touch_company(admin, company_id):
        raise HTTPException(status_code=403, detail="Firm not in your scope")
    return admin, company_id


async def _audit(action: str, admin: dict, company_id: str, date: str,
                 detail: str, user_id: Optional[str] = None) -> None:
    from server import db
    try:
        await db.attendance_audit_log.insert_one({
            "audit_id": f"dv_{uuid.uuid4().hex[:10]}",
            "module": "daily_verification",
            "action": action,
            "company_id": company_id,
            "date": date,
            "user_id": user_id,
            "by": admin.get("user_id"),
            "by_name": admin.get("name") or admin.get("email"),
            "detail": detail,
            "at": _now(),
        })
    except Exception:
        pass


def _classify_source(p: dict) -> str:
    src = str(p.get("source") or "").lower()
    bm = str(p.get("biometric_method") or "").lower()
    if src.startswith("zkteco:"):
        return "Biometric"
    if "manual" in src or "admin" in src:
        return "Manual"
    if "import" in src:
        return "Manual"
    if bm == "face" or "face" in src:
        return "Face Recognition"
    if bm == "qr" or "qr" in src:
        return "QR"
    return "Mobile App"


async def _build(company_id: str, date: str,
                 department: Optional[str] = None,
                 designation: Optional[str] = None,
                 contractor: Optional[str] = None,
                 category: Optional[str] = None,
                 shift: Optional[str] = None,
                 group: Optional[str] = None,
                 status: str = "active",
                 source: Optional[str] = None,
                 machine: Optional[str] = None,
                 employee_code: Optional[str] = None,
                 q: Optional[str] = None,
                 exceptions_only: bool = False,
                 present_only: bool = False) -> Dict[str, Any]:
    from server import db, _compute_monthly_grid_data
    try:
        datetime.strptime(date, "%Y-%m-%d")
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")

    grid = await _compute_monthly_grid_data(
        company_id, date[:7], from_date=date, to_date=date)
    labels = grid.get("day_labels") or []
    label = labels[0] if labels else ""
    weekday = (grid.get("weekday_labels") or [""])[0]

    company = await db.companies.find_one(
        {"company_id": company_id},
        {"_id": 0, "name": 1, "logo_base64": 1, "attendance_policy": 1})
    # Iter 479 (user request) — Bio Code shown on the PDF instead of the
    # Emp Code.
    bio_map: Dict[str, str] = {}
    async for _u in db.users.find({"company_id": company_id},
                                  {"_id": 0, "user_id": 1, "bio_code": 1}):
        bio_map[_u["user_id"]] = str(_u.get("bio_code") or "")
    policy = (company or {}).get("attendance_policy") or {}
    half_day_hours = float(policy.get("half_day_hours") or 4.0)
    ot_daily_limit = policy.get("ot_daily_max_hours")  # optional cap
    try:
        ot_daily_limit = float(ot_daily_limit) if ot_daily_limit else None
    except (TypeError, ValueError):
        ot_daily_limit = None

    # ---- master extras ----------------------------------------------------
    extra: Dict[str, dict] = {}
    async for u in db.users.find(
        {"role": "employee", "company_id": company_id},
        {"_id": 0, "user_id": 1, "employee_code": 1, "department": 1,
         "designation": 1, "position": 1, "employee_type": 1,
         "contractor_name": 1, "shift_name": 1, "employee_group": 1,
         "exit_date": 1, "employment_status": 1, "ot_applicable": 1,
         "father_name": 1},
    ):
        extra[u["user_id"]] = u

    # ---- approved leaves covering the date --------------------------------
    on_leave: set = set()
    try:
        async for lr in db.leave_requests.find(
            {"company_id": company_id, "status": "approved"},
            {"_id": 0, "user_id": 1, "from_date": 1, "to_date": 1},
        ):
            if str(lr.get("from_date") or "") <= date <= str(lr.get("to_date") or ""):
                on_leave.add(lr["user_id"])
    except Exception:
        pass

    # ---- the day's raw punches (sources / machines / drill data) ----------
    # Iter 420 (user rule) — only machines LINKED to this firm.
    dev_names: Dict[str, str] = {}
    async for d in db.biometric_devices.find(
            {"company_id": company_id},
            {"_id": 0, "serial_number": 1, "name": 1}):
        dev_names[str(d.get("serial_number") or "")] = d.get("name") or ""
    punches: Dict[str, List[dict]] = {}
    async for p in db.attendance.find(
        {"company_id": company_id, "date": date},
        {"_id": 0, "user_id": 1, "at": 1, "kind": 1, "source": 1,
         "device_serial": 1, "biometric_method": 1, "status": 1},
    ):
        punches.setdefault(p.get("user_id") or "", []).append(p)

    # ---- verification records ---------------------------------------------
    verif: Dict[str, dict] = {}
    async for v in db.daily_verifications.find(
            {"company_id": company_id, "date": date}, {"_id": 0}):
        verif[v.get("user_id") or ""] = v

    term = (q or "").strip().lower()
    rows: List[Dict[str, Any]] = []
    for emp in grid.get("employees") or []:
        uid = emp.get("user_id")
        ex = extra.get(uid, {})
        dept = emp.get("department") or ex.get("department") or ""
        desig = (emp.get("designation") or ex.get("designation")
                 or ex.get("position") or "")
        etype = ex.get("employee_type") or ""
        contr = ex.get("contractor_name") or ""
        shf = ex.get("shift_name") or ""
        grp = ex.get("employee_group") or ""
        resigned = bool(ex.get("exit_date")) or str(
            ex.get("employment_status") or "").lower() in (
            "exited", "resigned", "terminated", "inactive", "left")
        if department and dept != department:
            continue
        if designation and desig != designation:
            continue
        if contractor and contr != contractor:
            continue
        if category and etype != category:
            continue
        if shift and shf != shift:
            continue
        if group and grp != group:
            continue
        if status == "active" and resigned:
            continue
        if status == "resigned" and not resigned:
            continue
        if employee_code and str(emp.get("employee_code") or "") != employee_code:
            continue
        if term and term not in f"{emp.get('name', '')} {emp.get('employee_code', '')}".lower():
            continue

        cell = (emp.get("days") or {}).get(label) or {}
        plist = sorted(punches.get(uid) or [], key=lambda x: str(x.get("at") or ""))
        srcs = sorted({_classify_source(p) for p in plist})
        serials = sorted({str(p.get("device_serial") or "")
                          for p in plist if p.get("device_serial")})
        if source and source not in srcs:
            continue
        if machine and machine not in serials:
            continue

        has_in, has_out = bool(cell.get("in")), bool(cell.get("out"))
        duty = float(cell.get("duty_hours") or 0)
        ot_h = float(cell.get("ot_hours") or 0)
        late = int(cell.get("late_min") or 0)
        early = int(cell.get("early_min") or 0)
        manual = "Manual" in srcs
        leave = uid in on_leave
        wo = bool(cell.get("weekly_off"))
        hol = bool(cell.get("holiday"))

        # ---- attendance status --------------------------------------------
        markers: List[str] = []
        if not has_in and not has_out:
            st = ("Leave" if leave else "Holiday" if hol
                  else "Weekly Off" if wo else "Absent")
        elif has_in and not has_out:
            st = "Missing Out Punch"
        elif has_out and not has_in:
            st = "Missing In Punch"
        else:
            st = "Half Day" if 0 < duty <= half_day_hours else "Present"
        if late > 0:
            markers.append("Late Coming")
        if early > 0:
            markers.append("Early Going")
        if ot_h > 0:
            markers.append("Overtime")
        if manual and (has_in or has_out):
            markers.append("Manual Attendance")

        # ---- OT approved / unapproved --------------------------------------
        ot_ok = bool(ex.get("ot_applicable", True))
        approved_ot = unapproved_ot = 0.0
        if ot_h > 0:
            if not ot_ok:
                unapproved_ot = ot_h
            elif ot_daily_limit is not None and ot_h > ot_daily_limit:
                approved_ot, unapproved_ot = ot_daily_limit, ot_h - ot_daily_limit
            else:
                approved_ot = ot_h
        if unapproved_ot > 0:
            markers.append("Unapproved OT")

        # ---- row colour (priority per spec) --------------------------------
        if "Missing" in st:
            color = "red"
        elif unapproved_ot > 0:
            color = "orange"
        elif late > 0 or early > 0:
            color = "yellow"
        elif manual and (has_in or has_out):
            color = "blue"
        elif st in ("Absent", "Weekly Off", "Holiday", "Leave"):
            color = "muted"
        else:
            color = "green"
        exception = color in ("red", "orange", "yellow", "blue") or st == "Absent"
        if exceptions_only and not exception:
            continue
        # Iter 479 (user request) — "Only Present" filter: any employee
        # with at least ONE punch (even a single/missing-out punch) counts
        # as present.
        if present_only and not (has_in or has_out):
            continue

        v = verif.get(uid) or {}
        has_ot = _hm(ot_h) != "-"
        rows.append({
            "user_id": uid,
            "employee_code": emp.get("employee_code") or "",
            "bio_code": bio_map.get(uid) or "",
            "name": emp.get("name") or "",
            "father_name": (ex.get("father_name") or emp.get("father_name") or ""),
            "department": dept, "designation": desig,
            "contractor": contr, "category": etype, "shift": shf,
            "group": grp,
            "punch_in": cell.get("in") or "-",
            "punch_out": cell.get("out") or "-",
            "work_hours": _hm(duty),
            "ot_in": (cell.get("ot_in") or "-") if has_ot else "-",
            "ot_out": (cell.get("ot_out") or "-") if has_ot else "-",
            "ot_hours": _hm(ot_h),
            "approved_ot": _hm(approved_ot),
            "unapproved_ot": _hm(unapproved_ot),
            "total_hours": _hm(cell.get("hours")),
            "status": st,
            "markers": markers,
            "late_min": late, "early_min": early,
            "sources": srcs, "machines": serials,
            "machine_names": [dev_names.get(s) or s for s in serials],
            "color": color,
            "exception": exception,
            "verified": bool(v.get("verified")),
            "verified_by": v.get("verified_by_name") or "",
            "verified_at": v.get("verified_at") or "",
            "remarks": v.get("remarks") or "",
        })

    # ---- summary dashboard --------------------------------------------------
    def _cnt(fn) -> int:
        return sum(1 for r in rows if fn(r))

    def _hm_min(s: str) -> int:
        try:
            h, m = str(s).split(":")
            return int(h) * 60 + int(m)
        except (ValueError, AttributeError):
            return 0

    tot_ot = sum(_hm_min(r["ot_hours"]) for r in rows if r["ot_hours"] != "-")
    tot_aot = sum(_hm_min(r["approved_ot"]) for r in rows if r["approved_ot"] != "-")
    tot_uot = sum(_hm_min(r["unapproved_ot"]) for r in rows if r["unapproved_ot"] != "-")
    summary = {
        "total_employees": len(rows),
        "present": _cnt(lambda r: r["status"] in ("Present", "Half Day")),
        "absent": _cnt(lambda r: r["status"] == "Absent"),
        "leave": _cnt(lambda r: r["status"] == "Leave"),
        "weekly_off": _cnt(lambda r: r["status"] == "Weekly Off"),
        "holiday": _cnt(lambda r: r["status"] == "Holiday"),
        "missing_in": _cnt(lambda r: r["status"] == "Missing In Punch"),
        "missing_out": _cnt(lambda r: r["status"] == "Missing Out Punch"),
        "late": _cnt(lambda r: r["late_min"] > 0),
        "early": _cnt(lambda r: r["early_min"] > 0),
        "with_ot": _cnt(lambda r: r["ot_hours"] != "-"),
        "total_ot_hours": _hm(tot_ot / 60.0),
        "approved_ot_hours": _hm(tot_aot / 60.0),
        "unapproved_ot_hours": _hm(tot_uot / 60.0),
        "verified": _cnt(lambda r: r["verified"]),
        "pending_verification": _cnt(lambda r: not r["verified"]),
    }
    return {
        "company": {"company_id": company_id,
                    "name": (company or {}).get("name") or "",
                    "logo_base64": (company or {}).get("logo_base64")},
        # Iter 540 — attendance calculation mode badge for operators.
        "punch_sequence": bool(
            (((company or {}).get("attendance_policy") or {})
             .get("policy_master") or {}).get("attendance_by_duty_hours")),
        "date": date, "weekday": weekday,
        "group": group or "",
        "rows": rows, "summary": summary,
        "filter_options": {
            "departments": sorted({e.get("department") or "" for e in extra.values()} - {""}),
            "designations": sorted({(e.get("designation") or e.get("position") or "") for e in extra.values()} - {""}),
            "contractors": sorted({e.get("contractor_name") or "" for e in extra.values()} - {""}),
            "categories": sorted({e.get("employee_type") or "" for e in extra.values()} - {""}),
            "shifts": sorted({e.get("shift_name") or "" for e in extra.values()} - {""}),
            "groups": sorted({e.get("employee_group") or "" for e in extra.values()} - {""}),
            "sources": ["Biometric", "Mobile App", "Face Recognition", "QR", "Manual"],
            "machines": [{"serial": s, "name": n} for s, n in sorted(dev_names.items()) if s],
        },
    }


_FILTERS = ("department", "designation", "contractor", "category", "shift",
            "group", "source", "machine", "employee_code", "q")


def _filter_kwargs(**kw) -> dict:
    return {k: v for k, v in kw.items() if v}


# ---------------------------------------------------------------------------
# JSON (paginated)
# ---------------------------------------------------------------------------
@router.get("/daily-verification")
async def daily_verification_json(
    company_id: Optional[str] = Query(None),
    date: str = Query(...),
    department: Optional[str] = Query(None),
    designation: Optional[str] = Query(None),
    contractor: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    shift: Optional[str] = Query(None),
    group: Optional[str] = Query(None),
    status: str = Query("active"),
    source: Optional[str] = Query(None),
    machine: Optional[str] = Query(None),
    employee_code: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    exceptions_only: bool = Query(False),
    present_only: bool = Query(False),
    time_format: str = Query("24h"),
    limit: int = Query(200, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    authorization: Optional[str] = Header(None),
):
    _admin, cid = await _auth(authorization, company_id)
    data = await _build(cid, date, department, designation, contractor,
                        category, shift, group, status, source, machine,
                        employee_code, q, exceptions_only, present_only)
    if time_format == "12h":
        _apply_12h(data)
    rows = data["rows"]
    data["total_rows"] = len(rows)
    data["rows"] = rows[offset:offset + limit]
    data["offset"], data["limit"] = offset, limit
    return data


# ---------------------------------------------------------------------------
# Physical verification (+ audit trail)
# ---------------------------------------------------------------------------
@router.post("/daily-verification/verify")
async def daily_verification_verify(
    payload: dict = Body(...),
    authorization: Optional[str] = Header(None),
):
    from server import db
    admin, cid = await _auth(authorization, payload.get("company_id"))
    date = str(payload.get("date") or "")
    user_id = str(payload.get("user_id") or "")
    if not date or not user_id:
        raise HTTPException(status_code=400, detail="date and user_id required")
    verified = bool(payload.get("verified"))
    remarks = str(payload.get("remarks") or "").strip()[:500]
    rec = {
        "company_id": cid, "date": date, "user_id": user_id,
        "verified": verified,
        "verified_by": admin.get("user_id"),
        "verified_by_name": admin.get("name") or admin.get("email"),
        "verified_at": _now(),
        "remarks": remarks,
    }
    await db.daily_verifications.update_one(
        {"company_id": cid, "date": date, "user_id": user_id},
        {"$set": rec}, upsert=True)
    await _audit("physical_verification", admin, cid, date,
                 f"{'VERIFIED' if verified else 'UN-VERIFIED'}"
                 f"{' — ' + remarks if remarks else ''}", user_id=user_id)
    return {"ok": True, **rec}


# ---------------------------------------------------------------------------
# Drill-down — punch timeline + 7-day history
# ---------------------------------------------------------------------------
@router.get("/daily-verification/employee")
async def daily_verification_employee(
    company_id: Optional[str] = Query(None),
    date: str = Query(...),
    user_id: str = Query(...),
    time_format: str = Query("24h"),
    authorization: Optional[str] = Header(None),
):
    from server import db, _compute_monthly_grid_data
    _admin, cid = await _auth(authorization, company_id)
    u = await db.users.find_one(
        {"user_id": user_id, "company_id": cid},
        {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1,
         "department": 1, "designation": 1, "shift_name": 1,
         "shift_start": 1, "shift_end": 1, "contractor_name": 1})
    if not u:
        raise HTTPException(status_code=404, detail="Employee not found")
    dev_names: Dict[str, str] = {}
    async for d in db.biometric_devices.find(
            {"company_id": cid},
            {"_id": 0, "serial_number": 1, "name": 1, "model": 1, "brand": 1}):
        dev_names[str(d.get("serial_number") or "")] = (
            d.get("name") or d.get("model") or str(d.get("brand") or "").upper())
    timeline: List[dict] = []
    async for p in db.attendance.find(
        {"company_id": cid, "date": date, "user_id": user_id},
        {"_id": 0, "at": 1, "kind": 1, "source": 1, "device_serial": 1,
         "biometric_method": 1, "status": 1, "latitude": 1, "longitude": 1,
         "location_name": 1, "selfie_base64": 1},
    ):
        at = str(p.get("at") or "")
        sel = p.get("selfie_base64") or ""
        timeline.append({
            "time": at[11:19] if len(at) >= 19 else at,
            "kind": p.get("kind"),
            "source": _classify_source(p),
            "machine": dev_names.get(str(p.get("device_serial") or ""), ""),
            "device_serial": p.get("device_serial") or "",
            "status": p.get("status") or "",
            "lat": p.get("latitude"), "lng": p.get("longitude"),
            "location_name": p.get("location_name") or "",
            "selfie": sel if sel else None,
        })
    timeline.sort(key=lambda x: x["time"])
    # previous 7 days mini-history via the same pipeline
    d0 = datetime.strptime(date, "%Y-%m-%d").date()
    frm = (d0 - timedelta(days=6)).isoformat()
    history: List[dict] = []
    try:
        g = await _compute_monthly_grid_data(
            cid, frm[:7], from_date=frm, to_date=date, only_user_id=user_id)
        emp = next((e for e in g.get("employees") or []
                    if e.get("user_id") == user_id), None)
        for i, dl in enumerate(g.get("day_labels") or []):
            cell = ((emp or {}).get("days") or {}).get(dl) or {}
            iso_lbls = g.get("iso_labels") or []
            history.append({
                "day": dl,
                "date": iso_lbls[i] if i < len(iso_lbls) else "",
                "in": cell.get("in") or "-",
                "out": cell.get("out") or "-",
                "hours": _hm(cell.get("duty_hours")),
                "ot": _hm(cell.get("ot_hours")),
            })
    except Exception:
        pass
    if time_format == "12h":  # Iter 554 — AM/PM option
        for t in timeline:
            t["time"] = _to12(t["time"])
        for h in history:
            h["in"], h["out"] = _to12(h["in"]), _to12(h["out"])
    return {"employee": u, "date": date, "timeline": timeline,
            "history": history}


# ---------------------------------------------------------------------------
# Exports — shared builders
# ---------------------------------------------------------------------------
def _summary_lines(data: dict) -> List[str]:
    s = data["summary"]
    return [
        f"Total {s['total_employees']} · Present {s['present']} · "
        f"Absent {s['absent']} · Leave {s['leave']} · WO {s['weekly_off']} · "
        f"Holiday {s['holiday']}",
        f"Missing IN {s['missing_in']} · Missing OUT {s['missing_out']} · "
        f"Late {s['late']} · Early {s['early']} · With OT {s['with_ot']}",
        f"OT {s['total_ot_hours']} (Approved {s['approved_ot_hours']} / "
        f"Unapproved {s['unapproved_ot_hours']}) · "
        f"Verified {s['verified']} / Pending {s['pending_verification']}",
    ]


def _row_values(i: int, r: dict) -> list:
    st = r["status"] + ("" if not r["markers"] else " · " + ", ".join(r["markers"]))
    return [i, r["employee_code"], r["name"], r.get("father_name") or "",
            r["department"], r["designation"],
            r["contractor"], r["punch_in"], r["punch_out"], r["work_hours"],
            r["ot_in"], r["ot_out"], r["ot_hours"], r["total_hours"], st,
            "", "YES" if r["verified"] else "",
            r["verified_by"], r["remarks"]]


def _build_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Verification"
    ws.append([f"Daily In/Out & OT Verification — {data['company']['name']}"])
    ws.append([f"Date: {data['date']} ({data['weekday']})"])
    for ln in _summary_lines(data):
        ws.append([ln])
    ws.append([])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(HEADERS)
    hrow = ws.max_row
    for c in ws[hrow]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E293B")
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(data["rows"], start=1):
        ws.append(_row_values(i, r))
        fill = PatternFill("solid", fgColor=ROW_COLORS[r["color"]])
        for c in ws[ws.max_row]:
            c.fill = fill
    for col, w in zip("ABCDEFGHIJKLMNOPQRS",
                      [5, 9, 24, 20, 14, 14, 16, 9, 9, 9, 9, 9, 8, 10, 30, 14, 9, 16, 20]):
        ws.column_dimensions[col].width = w
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_csv(data: dict) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"Daily In/Out & OT Verification — {data['company']['name']} — {data['date']}"])
    w.writerow(HEADERS)
    for i, r in enumerate(data["rows"], start=1):
        w.writerow(_row_values(i, r))
    return buf.getvalue()


def _build_pdf(data: dict, orientation: str = "landscape",
               group_by: str = "designation") -> bytes:
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    pagesize = landscape(A4) if orientation != "portrait" else A4
    buf = io.BytesIO()
    # Iter 479 (user request) — summary block moved to the BOTTOM of every
    # page (with page number), so the bottom margin reserves room for it.
    # Iter 523 (user request) — left/right margins reduced 8mm → 4mm.
    # Iter 541 (user request) — PORTRAIT gets proper 12mm side margins so
    # the table no longer touches the page edges; landscape keeps 4mm.
    _side = 12 * mm if orientation == "portrait" else 4 * mm
    doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=_side,
                            rightMargin=_side, topMargin=8 * mm,
                            bottomMargin=17 * mm)
    styles = getSampleStyleSheet()
    avail = pagesize[0] - 2 * _side
    # Iter 479 (user request) — header: "Daily Report" + date on the RIGHT,
    # company name on the SECOND row.
    # Iter 523 (user request) — company ADDRESS printed with the firm name.
    # Iter 554 (user request) — header CENTERED: Row 1 = Company Name
    # (+ Group when a Group filter is applied), Row 2 = "Daily Report —
    # DD-MM-YYYY (Ddd)".
    _comp_addr = ", ".join(
        s for s in (str(data["company"].get("address") or "").strip(),
                    str(data["company"].get("city") or "").strip()) if s)
    _grp = str(data.get("group") or "").strip()
    try:
        _dmy = datetime.strptime(str(data["date"]), "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        _dmy = str(data["date"])
    story: list = [
        Paragraph(
            str(data["company"]["name"]) + (f" ({_grp})" if _grp else ""),
            ParagraphStyle("t", parent=styles["Title"], alignment=1,
                           fontSize=14, leading=17, spaceAfter=0)),
    ]
    if _comp_addr:
        story.append(Paragraph(_comp_addr, ParagraphStyle(
            "a", parent=styles["Normal"], alignment=1, fontSize=8,
            leading=10)))
    story.append(Paragraph(
        f"Daily Report — {_dmy} ({data['weekday']})",
        ParagraphStyle("d", parent=styles["Normal"], alignment=1,
                       fontSize=10.5, leading=14,
                       fontName="Helvetica-Bold")))
    story.append(Spacer(1, 3 * mm))
    fs = 7.0 if orientation != "portrait" else 5.8
    # Iter 479 (user request) — PDF drops Verified / Verified By / Remarks
    # and shows the BIO CODE instead of the Emp Code.
    # Iter 523 (user request) — Attendance Status column REMOVED; the
    # Contractor column only prints when the firm actually uses
    # contractors; taller rows + wide Signature column for signing.
    _has_contr = any((r.get("contractor") or "").strip()
                     for r in data["rows"])
    # Iter 524 (user request) — Department COLUMN removed; the print is
    # GROUPED department-wise with a grey band per department, and the
    # employee's FATHER NAME is shown next to the name.
    pdf_headers = ["S.No.", "Bio Code", "Employee Name", "Father Name",
                   "Designation"] + (["Contractor"] if _has_contr else []) + \
                  ["Punch In", "Punch Out", "Work Hrs", "OT In", "OT Out",
                   "OT Hrs", "Total Duty Hrs", "Signature"]
    n_cols = len(pdf_headers)
    tdata = [pdf_headers]
    tstyle = [
        ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#1E293B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl.white),
        ("FONTSIZE", (0, 0), (-1, -1), fs),
        ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#94A3B8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Iter 435 (user request) — Employee Name column LEFT-aligned.
        ("ALIGN", (2, 1), (2, -1), "LEFT"),
    ]
    # Iter 523 (user request) — taller rows so employees can sign properly.
    _row_h = 9 * mm if orientation != "portrait" else 8 * mm
    row_heights: list = [None]
    from itertools import groupby
    # Iter 524 (user request) — grouped print with a grey band per group.
    # Iter 554 (user request) — DEFAULT grouping is now DESIGNATION-wise;
    # department-wise stays available as an option (group_by=department).
    _gb_field = "department" if group_by == "department" else "designation"
    _gb_label = _gb_field.upper()
    rows_sorted = sorted(
        data["rows"],
        key=lambda r: ((r.get(_gb_field) or "~").lower(),
                       str(r.get("employee_code") or "")))
    ridx = 0
    sno = 0
    for gval, grp in groupby(rows_sorted,
                             key=lambda r: r.get(_gb_field) or ""):
        ridx += 1
        tdata.append([f"{_gb_label} — {gval or 'No ' + _gb_field.title()}"]
                     + [""] * (n_cols - 1))
        tstyle += [
            ("SPAN", (0, ridx), (-1, ridx)),
            ("BACKGROUND", (0, ridx), (-1, ridx), rl.HexColor("#CBD5E1")),
            ("FONTNAME", (0, ridx), (-1, ridx), "Helvetica-Bold"),
        ]
        row_heights.append(5.5 * mm)
        for r in grp:
            ridx += 1
            sno += 1
            tdata.append([str(v) for v in (
                [sno, r.get("bio_code") or r["employee_code"], r["name"],
                 r.get("father_name") or "", r["designation"]]
                + ([r["contractor"]] if _has_contr else [])
                + [r["punch_in"], r["punch_out"], r["work_hours"],
                   r["ot_in"], r["ot_out"], r["ot_hours"],
                   r["total_hours"], ""])])
            row_heights.append(_row_h)
            # Iter 523 (user request) — rows with BOTH punches are NOT
            # highlighted; only problem rows keep their colour.
            _both = r["punch_in"] != "-" and r["punch_out"] != "-"
            if not _both and r["color"] != "green":
                tstyle.append(("BACKGROUND", (0, ridx), (-1, ridx),
                               rl.HexColor("#" + ROW_COLORS[r["color"]])))
    # Signature column gets a fixed generous width; the rest auto-size.
    _sig_w = 34 * mm if orientation != "portrait" else 26 * mm
    tbl = Table(tdata, repeatRows=1, rowHeights=row_heights,
                colWidths=[None] * (n_cols - 1) + [_sig_w])
    tbl.setStyle(TableStyle(tstyle))
    story.append(tbl)
    story.append(Spacer(1, 4 * mm))
    # Iter 523 (user request) — legend line removed from the print.
    sum_lines = _summary_lines(data)

    def _footer(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 6.8)
        y = 4 * mm
        for j, ln in enumerate(reversed(sum_lines)):
            canvas.drawString(4 * mm, y + j * 9, ln)
        canvas.setFont("Helvetica-Bold", 7.5)
        canvas.drawRightString(pagesize[0] - 4 * mm, y,
                               f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


async def _export_data(authorization, company_id, date, **kw):
    time_format = kw.pop("time_format", "24h")  # Iter 554 — AM/PM option
    admin, cid = await _auth(authorization, company_id)
    data = await _build(cid, date, **kw)
    if time_format == "12h":
        _apply_12h(data)
    return admin, cid, data


_EXPORT_Q = dict(
    department=Query(None), designation=Query(None), contractor=Query(None),
    category=Query(None), shift=Query(None), group=Query(None),
    status=Query("active"), source=Query(None), machine=Query(None),
    employee_code=Query(None), q=Query(None), exceptions_only=Query(False),
)


@router.get("/daily-verification.xlsx")
async def daily_verification_xlsx(
    company_id: Optional[str] = Query(None), date: str = Query(...),
    department: Optional[str] = Query(None), designation: Optional[str] = Query(None),
    contractor: Optional[str] = Query(None), category: Optional[str] = Query(None),
    shift: Optional[str] = Query(None), group: Optional[str] = Query(None),
    status: str = Query("active"), source: Optional[str] = Query(None),
    machine: Optional[str] = Query(None), employee_code: Optional[str] = Query(None),
    q: Optional[str] = Query(None), exceptions_only: bool = Query(False),
    present_only: bool = Query(False),
    time_format: str = Query("24h"),
    authorization: Optional[str] = Header(None),
):
    admin, cid, data = await _export_data(
        authorization, company_id, date, department=department,
        designation=designation, contractor=contractor, category=category,
        shift=shift, group=group, status=status, source=source,
        machine=machine, employee_code=employee_code, q=q,
        exceptions_only=exceptions_only, present_only=present_only,
        time_format=time_format)
    await _audit("report_export", admin, cid, date, "XLSX export")
    return Response(
        content=_build_xlsx(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename=Daily_Verification_{date}.xlsx"})


@router.get("/daily-verification.csv")
async def daily_verification_csv(
    company_id: Optional[str] = Query(None), date: str = Query(...),
    department: Optional[str] = Query(None), designation: Optional[str] = Query(None),
    contractor: Optional[str] = Query(None), category: Optional[str] = Query(None),
    shift: Optional[str] = Query(None), group: Optional[str] = Query(None),
    status: str = Query("active"), source: Optional[str] = Query(None),
    machine: Optional[str] = Query(None), employee_code: Optional[str] = Query(None),
    q: Optional[str] = Query(None), exceptions_only: bool = Query(False),
    present_only: bool = Query(False),
    time_format: str = Query("24h"),
    authorization: Optional[str] = Header(None),
):
    admin, cid, data = await _export_data(
        authorization, company_id, date, department=department,
        designation=designation, contractor=contractor, category=category,
        shift=shift, group=group, status=status, source=source,
        machine=machine, employee_code=employee_code, q=q,
        exceptions_only=exceptions_only, present_only=present_only,
        time_format=time_format)
    await _audit("report_export", admin, cid, date, "CSV export")
    return Response(content=_build_csv(data), media_type="text/csv",
                    headers={"Content-Disposition":
                             f"attachment; filename=Daily_Verification_{date}.csv"})


@router.get("/daily-verification.pdf")
async def daily_verification_pdf(
    company_id: Optional[str] = Query(None), date: str = Query(...),
    orientation: str = Query("landscape"),
    group_by: str = Query("designation"),
    department: Optional[str] = Query(None), designation: Optional[str] = Query(None),
    contractor: Optional[str] = Query(None), category: Optional[str] = Query(None),
    shift: Optional[str] = Query(None), group: Optional[str] = Query(None),
    status: str = Query("active"), source: Optional[str] = Query(None),
    machine: Optional[str] = Query(None), employee_code: Optional[str] = Query(None),
    q: Optional[str] = Query(None), exceptions_only: bool = Query(False),
    present_only: bool = Query(False),
    time_format: str = Query("24h"),
    authorization: Optional[str] = Header(None),
):
    admin, cid, data = await _export_data(
        authorization, company_id, date, department=department,
        designation=designation, contractor=contractor, category=category,
        shift=shift, group=group, status=status, source=source,
        machine=machine, employee_code=employee_code, q=q,
        exceptions_only=exceptions_only, present_only=present_only,
        time_format=time_format)
    await _audit("report_export", admin, cid, date, f"PDF export ({orientation})")
    return Response(content=_build_pdf(data, orientation, group_by),
                    media_type="application/pdf",
                    headers={"Content-Disposition":
                             f"inline; filename=Daily_Verification_{date}.pdf"})


# ---------------------------------------------------------------------------
# Email + WhatsApp
# ---------------------------------------------------------------------------
@router.post("/daily-verification/email")
async def daily_verification_email(
    payload: dict = Body(...),
    authorization: Optional[str] = Header(None),
):
    from routes.email_notifications import _get_settings, _smtp_send
    admin, cid = await _auth(authorization, payload.get("company_id"))
    date = str(payload.get("date") or "")
    to = str(payload.get("to") or "").strip()
    if not date or not to:
        raise HTTPException(status_code=400, detail="date and to (email) required")
    settings = await _get_settings()
    if not settings or not settings.get("username"):
        raise HTTPException(status_code=400,
                            detail="SMTP is not configured (Email Settings).")
    data = await _build(cid, date, exceptions_only=bool(payload.get("exceptions_only")))
    if str(payload.get("time_format") or "") == "12h":
        _apply_12h(data)
    s = data["summary"]
    body = (f"Daily In/Out & OT Verification Report — {data['company']['name']}\n"
            f"Date: {date}\n\n" + "\n".join(_summary_lines(data)) +
            "\n\nExcel and PDF copies are attached.")
    await _smtp_send(settings, to,
                     f"Daily Verification Report — {data['company']['name']} ({date})",
                     body,
                     attachments=[
                         {"filename": f"Daily_Verification_{date}.xlsx",
                          "content": _build_xlsx(data),
                          "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                         {"filename": f"Daily_Verification_{date}.pdf",
                          "content": _build_pdf(data),
                          "mime": "application/pdf"},
                     ])
    await _audit("report_email", admin, cid, date, f"Emailed to {to}")
    return {"ok": True, "message": f"Report emailed to {to}",
            "present": s["present"], "total": s["total_employees"]}


@router.post("/daily-verification/whatsapp")
async def daily_verification_whatsapp(
    payload: dict = Body(...),
    authorization: Optional[str] = Header(None),
):
    from utils.whatsapp_engine import (WhatsAppClient, _configured,
                                       get_settings, normalize_number)
    admin, cid = await _auth(authorization, payload.get("company_id"))
    date = str(payload.get("date") or "")
    to = normalize_number(str(payload.get("to") or ""))
    if not date or not to:
        raise HTTPException(status_code=400, detail="date and to (phone) required")
    settings = await get_settings(cid)
    if not _configured(settings):
        raise HTTPException(status_code=400,
                            detail="WhatsApp is not configured for this firm "
                                   "(WhatsApp Center → Settings).")
    data = await _build(cid, date)
    if str(payload.get("time_format") or "") == "12h":
        _apply_12h(data)
    pdf = _build_pdf(data)
    client = WhatsAppClient(settings)
    media_id = await client.upload_media(
        f"Daily_Verification_{date}.pdf", "application/pdf", pdf)
    await client.send_document(
        to, media_id, f"Daily_Verification_{date}.pdf",
        caption=f"Daily In/Out & OT Verification — "
                f"{data['company']['name']} ({date})")
    await _audit("report_whatsapp", admin, cid, date, f"WhatsApp PDF to {to}")
    return {"ok": True, "message": f"Report sent on WhatsApp to {to}"}
