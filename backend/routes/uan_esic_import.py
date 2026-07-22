"""Bulk UAN / ESIC-IP import (Iter 242 — user request).

Upload one Excel/CSV with an employee identifier column (Employee Code /
Bio / PF No / UAN-name) plus UAN and/or ESIC IP columns; the matching
employees in the firm get their ``uan_no`` / ``esi_ip_no`` filled in one
shot. Only these two identifier fields are touched — salary, attendance,
leave and the rest of the Employee Master are never modified.

Endpoints
    POST /api/admin/uan-esic-import/template.xlsx   → blank template
    POST /api/admin/uan-esic-import                 → upload + apply
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, File, Form, Header, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from server import db, get_user_from_token, now_iso, require_role  # noqa: E402

router = APIRouter(prefix="/api/admin/uan-esic-import", tags=["uan-esic-import"])

# Header aliases (lower-cased, non-alnum stripped) → canonical key.
_CODE_ALIASES = {
    "employeecode", "empcode", "code", "empno", "employeeno", "empid",
    "employeeid", "ecode", "sno", "srno",
}
_UAN_ALIASES = {"uan", "uanno", "uannumber", "uanno.", "universalaccountnumber"}
_IP_ALIASES = {
    "esicipno", "esicip", "ipno", "ipnumber", "esino", "esicno",
    "esicnumber", "insurancenumber", "insuranceno", "esiipno", "esi",
}
_BIO_ALIASES = {"biocode", "bio", "biometriccode", "machineid", "biometricid"}
_NAME_ALIASES = {"name", "employeename", "empname"}


def _norm(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


async def _admin(authorization: Optional[str], company_id: str) -> dict:
    user = await get_user_from_token(authorization)
    require_role(user, ["super_admin", "sub_admin", "company_admin"])
    if user["role"] == "company_admin" and user.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="Not your firm")
    return user


def _parse_table(raw: bytes, filename: str) -> List[Dict[str, str]]:
    """Return a list of row dicts keyed by normalised header."""
    name = (filename or "").lower()
    rows: List[List[str]] = []
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c).strip() for c in r])
    else:
        # CSV / TSV
        text = raw.decode("utf-8-sig", errors="replace")
        sniff_delim = "\t" if text.count("\t") > text.count(",") else ","
        for r in csv.reader(io.StringIO(text), delimiter=sniff_delim):
            rows.append([str(c).strip() for c in r])
    rows = [r for r in rows if any(c for c in r)]
    if len(rows) < 2:
        raise HTTPException(status_code=400,
                            detail="File has no data rows.")
    headers = [_norm(h) for h in rows[0]]
    out: List[Dict[str, str]] = []
    for r in rows[1:]:
        d = {}
        for i, h in enumerate(headers):
            if h:
                d[h] = r[i] if i < len(r) else ""
        out.append(d)
    return out


def _pick(d: Dict[str, str], aliases: set) -> str:
    for k, v in d.items():
        if k in aliases:
            return str(v or "").strip()
    return ""


@router.post("/template.xlsx")
async def template(company_id: str = Form(...),
                   authorization: Optional[str] = Header(None)):
    await _admin(authorization, company_id)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "UAN_ESIC"
    ws.append(["Employee Code", "Name", "UAN", "ESIC IP No"])
    # Pre-fill existing employees so the user just fills the blanks.
    async for u in db.users.find(
        {"company_id": company_id, "role": "employee"},
        {"_id": 0, "employee_code": 1, "name": 1, "uan_no": 1, "esi_ip_no": 1},
    ).sort("employee_code", 1):
        ws.append([u.get("employee_code") or "", u.get("name") or "",
                   u.get("uan_no") or "", u.get("esi_ip_no") or ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="UAN_ESIC_Template.xlsx"'},
    )


@router.post("")
async def do_import(
    company_id: str = Form(...),
    match_by: str = Form("employee_code"),  # employee_code | bio_code | name
    dry_run: bool = Form(False),
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None),
):
    admin = await _admin(authorization, company_id)
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file.")
    table = _parse_table(raw, file.filename or "")

    # Build lookup of existing employees for this firm.
    field = {"employee_code": "employee_code", "bio_code": "bio_code",
             "name": "name"}.get(match_by, "employee_code")
    users = await db.users.find(
        {"company_id": company_id, "role": "employee"},
        {"_id": 0, "user_id": 1, "employee_code": 1, "bio_code": 1, "name": 1},
    ).to_list(20000)

    def _key(v: str) -> str:
        return re.sub(r"\s+", " ", str(v or "").strip().lower())

    index: Dict[str, str] = {}
    for u in users:
        k = _key(u.get(field))
        if k:
            index[k] = u["user_id"]

    updated = 0
    uan_set = 0
    ip_set = 0
    not_found: List[str] = []
    invalid: List[str] = []
    seen = 0
    for row in table:
        ident = (_pick(row, _CODE_ALIASES) if field == "employee_code"
                 else _pick(row, _BIO_ALIASES) if field == "bio_code"
                 else _pick(row, _NAME_ALIASES))
        if not ident:
            continue
        seen += 1
        uid = index.get(_key(ident))
        if not uid:
            not_found.append(ident)
            continue
        uan = _pick(row, _UAN_ALIASES)
        ip = _pick(row, _IP_ALIASES)
        upd: Dict[str, Any] = {}
        if uan:
            digits = re.sub(r"\D", "", uan)
            if len(digits) == 12:
                upd["uan_no"] = digits
                upd["uan_no_source"] = "bulk_import"
                upd["uan_no_updated_at"] = now_iso()
                uan_set += 1
            else:
                invalid.append(f"{ident}: UAN '{uan}' is not 12 digits")
        if ip:
            upd["esi_ip_no"] = ip.strip()
            upd["esi_ip_no_source"] = "bulk_import"
            upd["esi_ip_no_updated_at"] = now_iso()
            ip_set += 1
        if upd and not dry_run:
            await db.users.update_one({"user_id": uid}, {"$set": upd})
        if upd:
            updated += 1

    return {
        "ok": True,
        "dry_run": dry_run,
        "match_by": field,
        "rows_read": len(table),
        "rows_with_identifier": seen,
        "employees_updated": updated,
        "uan_filled": uan_set,
        "esic_ip_filled": ip_set,
        "not_found_count": len(not_found),
        "not_found": not_found[:50],
        "invalid": invalid[:50],
        "by": admin.get("name") or admin.get("user_id"),
    }
