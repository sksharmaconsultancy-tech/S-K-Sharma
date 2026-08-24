"""Iter 292 (user request) — Monthly Employee In/Out & Overtime Matrix Report.

A BRAND-NEW report (existing reports untouched): one employee per matrix,
rows = D-In / D-Out / OT-In / OT-Out / Total Hrs / OT Hrs, columns = days
of the month. Reuses the SAME compute pipeline as the on-screen Attendance
Grid (`_compute_monthly_grid_data`) so every figure matches 1:1.

Endpoints (all firm-scoped, super/sub/company admin):
  * GET /api/admin/reports/inout-ot-matrix          — JSON (paginated)
  * GET /api/admin/reports/inout-ot-matrix.xlsx     — Excel, colours preserved
  * GET /api/admin/reports/inout-ot-matrix.pdf      — A4 LANDSCAPE, one
        employee per page, header repeated on every page
  * GET /api/admin/reports/inout-ot-matrix.csv      — plain CSV

Colour legend (matches UI): OT=light blue, Late=yellow, Missing punch=red,
Holiday=grey, Weekly off=light green, Leave=orange, Normal=white.
"""
import csv
import io
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException, Query
from fastapi.responses import Response

router = APIRouter(prefix="/api/admin/reports", tags=["inout-ot-matrix"])

# Iter 402 (user request) — row sequence: D-In, D-Out, Total Hrs, OT-In,
# OT-Out, Total OT Hrs, then Total Working Hrs (= Total Hrs + Total OT Hrs).
ROW_KEYS = [
    ("d_in", "D-In"), ("d_out", "D-Out"), ("total", "Total Hrs"),
    ("ot_in", "OT-In"), ("ot_out", "OT-Out"), ("ot", "Total OT Hrs"),
    ("grand", "Total Working Hrs"),
]

# hex colours shared by xlsx + pdf so exports match the screen exactly
FLAG_COLORS = {
    "ot": "DBEAFE",        # light blue
    "late": "FEF08A",      # yellow
    "missing": "FECACA",   # red
    "holiday": "E2E8F0",   # grey
    "weekly_off": "DCFCE7",  # light green
    "leave": "FED7AA",     # orange
}


def _row_keys(data: Dict[str, Any]) -> List[tuple]:
    """Iter 710 — in dummy mode the OT rows and the duplicate grand row are
    meaningless once hours are masked; the matrix stays a clean
    D-In / D-Out / Total Hrs set."""
    if data.get("dummy_mode"):
        return [rk for rk in ROW_KEYS if rk[0] in ("d_in", "d_out", "total")]
    return ROW_KEYS


def _f(v: Any) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _fmt_hm(hours: Any) -> str:
    """Float hours → HH:MM ('-' when zero/empty)."""
    try:
        h = float(hours or 0)
    except (TypeError, ValueError):
        return "-"
    if h <= 0:
        return "-"
    total_min = int(round(h * 60))
    return f"{total_min // 60:02d}:{total_min % 60:02d}"


def _cell_flag(cell: Dict[str, Any]) -> str:
    """Priority: missing > late > ot > leave > holiday > weekly_off > normal."""
    has_in, has_out = bool(cell.get("in")), bool(cell.get("out"))
    if (has_in and not has_out) or (has_out and not has_in):
        return "missing"
    if (cell.get("late_min") or 0) > 0:
        return "late"
    if float(cell.get("ot_hours") or 0) > 0:
        return "ot"
    if cell.get("leave"):
        return "leave"
    if not has_in and cell.get("holiday"):
        return "holiday"
    if not has_in and cell.get("weekly_off"):
        return "weekly_off"
    return "normal"


async def _auth(authorization: Optional[str], company_id: Optional[str]):
    from server import get_user_from_token, require_role, sub_admin_can_touch_company
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if admin.get("role") == "company_admin":
        company_id = admin.get("company_id")
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    if admin.get("role") == "sub_admin" and not sub_admin_can_touch_company(admin, company_id):
        raise HTTPException(status_code=403, detail="Firm not in your scope")
    return admin, company_id


# Iter 702 (user issue — "In/Out report takes too much time to load"):
# the FULL month grid (punch pairing, dedup, cross-day stitching for every
# employee) was recomputed on EVERY page change / filter / search keystroke.
# Cache the computed grid per (company, month) for a short TTL — pagination
# and filters then run on the cached data. _build only READS the grid, so
# sharing the object is safe. New punches show up within TTL seconds.
_GRID_CACHE: Dict[str, tuple] = {}
_GRID_TTL_SEC = 90.0


async def _cached_grid(company_id: str, month: str) -> Dict[str, Any]:
    import time as _t
    from server import _compute_monthly_grid_data
    key = f"{company_id}|{month}"
    hit = _GRID_CACHE.get(key)
    if hit and (_t.time() - hit[0]) < _GRID_TTL_SEC:
        return hit[1]
    data = await _compute_monthly_grid_data(company_id, month)
    _GRID_CACHE[key] = (_t.time(), data)
    if len(_GRID_CACHE) > 40:   # keep the cache tiny
        oldest = min(_GRID_CACHE, key=lambda k: _GRID_CACHE[k][0])
        _GRID_CACHE.pop(oldest, None)
    return data


async def _build(
    company_id: str,
    month: str,
    department: Optional[str] = None,
    designation: Optional[str] = None,
    employee_type: Optional[str] = None,
    contractor: Optional[str] = None,
    shift: Optional[str] = None,
    q: Optional[str] = None,
    status: str = "active",
    dummy: bool = False,
    dummy_shift: Optional[str] = None,
) -> Dict[str, Any]:
    """Compute the grid once, join extra master fields, apply filters and
    shape the per-employee 6-row matrix."""
    from server import db
    data = await _cached_grid(company_id, month)
    day_labels: List[str] = data.get("day_labels") or []
    weekday_labels: List[str] = data.get("weekday_labels") or []

    # Iter 628 (user request) — DUMMY SHIFT IN/OUT MATRIX: a strictly
    # READ-ONLY reporting layer. The matrix pipeline is identical; only the
    # D-In / D-Out DISPLAY is substituted with the employee's Dummy-Shift
    # master timings on present days. NOTHING is ever written to
    # attendance / payroll / punches.
    dummy_map: Dict[str, tuple] = {}
    dummy_dur: Dict[str, tuple] = {}  # Iter 710 — name → ("H:MM", minutes)
    if dummy:
        _dpol = await db.companies.find_one(
            {"company_id": company_id},
            {"_id": 0, "attendance_policy.policy_master.dummy_shift_allowed": 1})
        if not bool((((_dpol or {}).get("attendance_policy") or {})
                     .get("policy_master") or {}).get("dummy_shift_allowed")):
            raise HTTPException(
                status_code=400,
                detail=("Dummy Shift is not enabled for this firm. Switch on "
                        "'Dummy Shift Allowed' in the Attendance Policy first."))
        from routes.labour_reports import DUMMY_SHIFTS

        def _hm2min(s: str) -> int:
            h, mn = str(s).split(":")
            return int(h) * 60 + int(mn)
        # overnight = end at/before start → OUT lands on the NEXT calendar
        # date (marked with * in the report; no extra attendance day).
        dummy_map = {d["name"]: (d["start"], d["end"],
                                 _hm2min(d["end"]) <= _hm2min(d["start"]))
                     for d in DUMMY_SHIFTS}
        # Iter 710 (user issue — "dummy report shows ACTUAL 12-hr duty
        # hours") — pre-compute every dummy shift's FIXED duration so the
        # Total-Hrs cells are masked with it, never the real duty hours.
        for _d in DUMMY_SHIFTS:
            _mins = _hm2min(_d["end"]) - _hm2min(_d["start"])
            if _mins <= 0:
                _mins += 24 * 60
            dummy_dur[_d["name"]] = (
                f"{_mins // 60:02d}:{_mins % 60:02d}", _mins)
    # Summary counters (dummy mode only — sec 16 of the user spec).
    _dsum = {"present": 0, "week_off": 0, "holiday": 0, "absent": 0}
    _dshift_counts: Dict[str, int] = {}

    # Extra master fields not present in the grid rows.
    extra: Dict[str, Dict[str, Any]] = {}
    async for u in db.users.find(
        {"role": "employee", "company_id": company_id},
        {"_id": 0, "user_id": 1, "employee_type": 1, "contractor_name": 1,
         "shift_name": 1, "department": 1, "designation": 1, "position": 1,
         "exit_date": 1, "employment_status": 1, "dummy_shift": 1},
    ):
        extra[u["user_id"]] = u

    # Approved leaves for the month (marks the LEAVE colour). Collection may
    # be empty — that's fine.
    leave_days: Dict[str, set] = {}
    try:
        async for lr in db.leave_requests.find(
            {"company_id": company_id, "status": "approved"},
            {"_id": 0, "user_id": 1, "from_date": 1, "to_date": 1},
        ):
            f, t = str(lr.get("from_date") or ""), str(lr.get("to_date") or "")
            if f[:7] <= month <= t[:7]:
                leave_days.setdefault(lr["user_id"], set()).add((f, t))
    except Exception:
        pass

    def _on_leave(uid: str, iso: str) -> bool:
        for f, t in leave_days.get(uid, set()):
            if f <= iso <= t:
                return True
        return False

    term = (q or "").strip().lower()
    out_rows: List[Dict[str, Any]] = []
    for emp in data.get("employees") or []:
        ex = extra.get(emp.get("user_id"), {})
        dept = emp.get("department") or ex.get("department") or ""
        desig = emp.get("designation") or ex.get("designation") or ex.get("position") or ""
        etype = ex.get("employee_type") or ""
        contr = ex.get("contractor_name") or ""
        shf = ex.get("shift_name") or ""
        resigned = bool(ex.get("exit_date")) or str(
            ex.get("employment_status") or "").lower() in (
            "exited", "resigned", "terminated", "inactive", "left")
        if department and dept != department:
            continue
        if designation and desig != designation:
            continue
        if employee_type and etype != employee_type:
            continue
        if contractor and contr != contractor:
            continue
        if shift and shf != shift:
            continue
        if status == "active" and resigned:
            continue
        if status == "resigned" and not resigned:
            continue
        if term and term not in f"{emp.get('name', '')} {emp.get('employee_code', '')}".lower():
            continue
        ds = (ex.get("dummy_shift") or "").strip()
        if dummy and dummy_shift and ds != dummy_shift:
            continue
        if dummy:
            _dshift_counts[ds or "— None —"] = _dshift_counts.get(ds or "— None —", 0) + 1

        _emp_dmin = 0  # Iter 710 — dummy-mode month total (minutes)
        days: Dict[str, Dict[str, Any]] = {}
        for i, dl in enumerate(day_labels):
            cell = (emp.get("days") or {}).get(dl) or {}
            iso = f"{month}-{str(dl)[:2]}" if len(str(dl)) >= 2 else ""
            leave = _on_leave(emp.get("user_id"), iso) if iso else False
            c = dict(cell)
            c["leave"] = leave
            flag = _cell_flag(c)
            _tot_s = _fmt_hm(cell.get("duty_hours"))
            _ot_s = _fmt_hm(cell.get("ot_hours"))
            # Iter 402 (user rule) — when NO OT is counted for the day the
            # OT-In / OT-Out punches are NOT shown either (the arithmetic
            # split boundary is not a real punch and used to appear BEFORE
            # the duty OUT).
            _has_ot = _ot_s != "-"
            days[dl] = {
                "d_in": cell.get("in") or "-",
                "d_out": cell.get("out") or "-",
                "ot_in": (cell.get("ot_in") or "-") if _has_ot else "-",
                "ot_out": (cell.get("ot_out") or "-") if _has_ot else "-",
                "total": _tot_s,
                "ot": _ot_s,
                # Iter 402 — Total Working Hrs = Total Hrs + Total OT Hrs
                # (the grid's ``hours`` is already duty + OT combined).
                "grand": _fmt_hm(cell.get("hours")),
                "flag": flag,
                # hover / click details
                "detail": {
                    "date": iso,
                    "weekday": weekday_labels[i] if i < len(weekday_labels) else "",
                    "punch_count": cell.get("punches") or 0,
                    "working_hours": _fmt_hm(cell.get("hours")),
                    "break_time": _fmt_hm(cell.get("break_hours")),
                    "late_min": cell.get("late_min") or 0,
                    "early_min": cell.get("early_min") or 0,
                    "ot_hours": _fmt_hm(cell.get("ot_hours")),
                    "sources": cell.get("sources") or [],
                },
            }
            if dummy:
                # Iter 628 — day-status priority (user spec sec 10):
                # Holiday → Week Off → Present → Absent. Substitution is
                # DISPLAY ONLY; the underlying attendance is never touched.
                has_in, has_out = bool(cell.get("in")), bool(cell.get("out"))
                e_d = days[dl]

                def _mask_hours(_total: str) -> None:
                    # Iter 710 — NEVER leak the actual duty / OT hours in
                    # dummy mode; only the dummy shift's own fixed duration
                    # (or "-") is shown. DISPLAY-ONLY, database untouched.
                    e_d["total"] = e_d["grand"] = _total
                    e_d["ot"] = e_d["ot_in"] = e_d["ot_out"] = "-"
                    e_d["detail"]["working_hours"] = _total
                    e_d["detail"]["ot_hours"] = "-"
                    e_d["detail"]["break_time"] = "-"

                if cell.get("holiday"):
                    e_d["d_in"] = e_d["d_out"] = "H"
                    e_d["flag"] = "holiday"
                    _mask_hours("-")
                    _dsum["holiday"] += 1
                elif cell.get("weekly_off"):
                    e_d["d_in"] = e_d["d_out"] = "WO"
                    e_d["flag"] = "weekly_off"
                    _mask_hours("-")
                    _dsum["week_off"] += 1
                elif has_in or has_out:
                    # Iter 710 (user: "show ONLY dummy shift duty timings")
                    # — EVERY present day shows the dummy shift's own
                    # timings and fixed duration, regardless of punch count
                    # or a missing side. Actual punch times never appear.
                    if ds in dummy_map:
                        st, en, overnight = dummy_map[ds]
                        e_d["d_in"] = st
                        e_d["d_out"] = f"{en}*" if overnight else en
                    if ds in dummy_dur:
                        _dstr, _dmin = dummy_dur[ds]
                        _mask_hours(_dstr)
                        _emp_dmin += _dmin
                    if ds in dummy_map and e_d["flag"] in ("late", "missing",
                                                           "ot"):
                        e_d["flag"] = "normal"
                    _dsum["present"] += 1
                else:
                    _dsum["absent"] += 1
        totals = emp.get("totals") or {}
        if dummy:
            # Iter 710 — month totals rebuilt from the masked dummy
            # durations; actual duty / OT totals never reach the payload.
            _mt = (f"{_emp_dmin // 60:02d}:{_emp_dmin % 60:02d}"
                   if _emp_dmin else "-")
            _row_totals = {"month_total": _mt, "month_ot": "-",
                           "month_grand": _mt}
        else:
            _row_totals = {
                "month_total": _fmt_hm(totals.get("duty_hours")),
                "month_ot": _fmt_hm(totals.get("ot_hours")),
                "month_grand": _fmt_hm(_f(totals.get("duty_hours"))
                                       + _f(totals.get("ot_hours"))),
            }
        out_rows.append({
            "user_id": emp.get("user_id"),
            "employee_code": emp.get("employee_code"),
            "name": emp.get("name"),
            "department": dept, "designation": desig,
            "category": etype, "contractor_name": contr,
            # Iter 710 — the ACTUAL shift name is hidden in dummy mode so
            # the report reveals nothing about the real duty pattern.
            "shift_name": "" if dummy else shf,
            "dummy_shift": ds,
            "status": "RESIGNED" if resigned else "ACTIVE",
            "days": days,
            **_row_totals,
            "present_days": totals.get("present_days_policy",
                                       totals.get("present_days")),
        })

    company = data.get("company") or {}
    comp_doc = await db.companies.find_one(
        {"company_id": company_id},
        {"_id": 0, "name": 1, "logo_base64": 1, "attendance_policy": 1})
    y, m = month[:4], month[5:7]

    # Iter 520 (user request — "Set this report as per the FIRM ATTENDANCE
    # POLICY") — surface the firm's effective policy on the report header
    # (screen + Excel + PDF) so every figure is verifiable against it.
    _pol = (comp_doc or {}).get("attendance_policy") or {}
    from server import inject_firm_ot_flag
    _pol = await inject_firm_ot_flag(dict(_pol), company_id)
    _wd_names = ["Monday", "Tuesday", "Wednesday", "Thursday",
                 "Friday", "Saturday", "Sunday"]
    _wo_names = ", ".join(
        _wd_names[int(i)] for i in (_pol.get("weekly_off_days") or [])
        if isinstance(i, (int, float)) and 0 <= int(i) <= 6) or "None"
    _fd_h = float(_pol.get("full_day_hours")
                  or _pol.get("standard_working_hours") or 8.0)
    _grace = int(float(_pol.get("grace_minutes_late") or 0))
    _round_m = int(float(_pol.get("duty_hours_rounding_minutes") or 0))
    _slab = (_pol.get("policy_master") or {}).get("ot_slab_minutes")
    _slab = int(_slab) if _slab in (0, 30, 60) else 30
    _ot_on = _pol.get("firm_ot_allowed") is not False
    policy_summary = {
        "full_day_hours": _fd_h,
        "half_day_hours": float(_pol.get("half_day_hours") or 4.0),
        "grace_minutes_late": _grace,
        "duty_hours_rounding_minutes": _round_m,
        "ot_slab_minutes": _slab,
        "weekly_off": _wo_names,
        "ot_allowed": _ot_on,
        "line": (
            f"Firm Attendance Policy — Full Day {_fd_h:g} hrs · "
            + (f"OT beyond {_fd_h:g} worked hrs"
               + (f" (slab {_slab} min)" if _slab else " (exact)")
               if _ot_on else "OT OFF (firm master)")
            + f" · Late grace {_grace} min"
            + f" · Rounding {_round_m} min" + f" · Weekly Off: {_wo_names}"
        ),
    }

    # Iter 403 (user accepted) — day-wise OT totals across the FILTERED
    # employee set so supervisors spot heavy-OT days at a glance.
    def _hm_min(s: Any) -> int:
        try:
            h, mn = str(s).split(":")
            return int(h) * 60 + int(mn)
        except (ValueError, AttributeError):
            return 0

    day_ot_min: Dict[str, int] = {dl: 0 for dl in day_labels}
    for e in out_rows:
        for dl, v in (e.get("days") or {}).items():
            if v.get("ot") and v["ot"] != "-":
                day_ot_min[dl] = day_ot_min.get(dl, 0) + _hm_min(v["ot"])
    result = {
        "company": {"company_id": company_id,
                    "name": (comp_doc or {}).get("name") or company.get("name"),
                    "logo_base64": (comp_doc or {}).get("logo_base64")},
        "month": month, "year": y, "month_number": m,
        "payroll_period": f"01-{m}-{y} to {len(day_labels):02d}-{m}-{y}",
        "policy": policy_summary,
        "day_labels": day_labels, "weekday_labels": weekday_labels,
        "employees": out_rows,
        "day_ot_totals": {dl: _fmt_hm(mn / 60.0) for dl, mn in day_ot_min.items()},
        "month_ot_total": _fmt_hm(sum(day_ot_min.values()) / 60.0),
        "filter_options": {
            "departments": sorted({e.get("department") or "" for e in extra.values()} - {""}),
            "designations": sorted({(e.get("designation") or e.get("position") or "") for e in extra.values()} - {""}),
            "categories": sorted({e.get("employee_type") or "" for e in extra.values()} - {""}),
            "contractors": sorted({e.get("contractor_name") or "" for e in extra.values()} - {""}),
            "shifts": sorted({e.get("shift_name") or "" for e in extra.values()} - {""}),
        },
    }
    if dummy:
        result["dummy_mode"] = True
        result["report_title"] = "DUMMY SHIFT IN / OUT MATRIX REPORT"
        result["dummy_note"] = ("DUMMY SHIFT REPORT — FOR REPORTING PURPOSE "
                                "ONLY · * = OUT on the next calendar date")
        result["dummy_summary"] = {
            "total_employees": len(out_rows),
            "present_days": _dsum["present"],
            "week_off_days": _dsum["week_off"],
            "holiday_days": _dsum["holiday"],
            "absent_days": _dsum["absent"],
            "shift_counts": dict(sorted(_dshift_counts.items())),
        }
        result["filter_options"]["dummy_shifts"] = sorted(
            {(e.get("dummy_shift") or "").strip() for e in extra.values()} - {""})
    return result


_FILTER_PARAMS = dict()  # documentation only


@router.get("/inout-ot-matrix")
async def inout_ot_matrix_json(
    month: str,
    company_id: Optional[str] = None,
    department: Optional[str] = None,
    designation: Optional[str] = None,
    employee_type: Optional[str] = None,
    contractor: Optional[str] = None,
    shift: Optional[str] = None,
    q: Optional[str] = None,
    status: str = "active",
    dummy: int = 0,
    dummy_shift: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    authorization: Optional[str] = Header(None),
):
    _, cid = await _auth(authorization, company_id)
    data = await _build(cid, month, department, designation, employee_type,
                        contractor, shift, q, status,
                        dummy=bool(dummy), dummy_shift=dummy_shift)
    emps = data.pop("employees")
    total = len(emps)
    start = (page - 1) * page_size
    data["employees"] = emps[start:start + page_size]
    data["total_employees"] = total
    data["page"] = page
    data["page_size"] = page_size
    data["total_pages"] = max(1, -(-total // page_size))
    return data


# ---------------------------------------------------------------------------
# Exports — layout mirrors the screen exactly.
# ---------------------------------------------------------------------------
def _header_lines(data: Dict[str, Any], emp: Dict[str, Any]) -> List[str]:
    c = data["company"]
    return [
        f"Company: {c.get('name') or ''}",
        f"Employee: {emp.get('employee_code') or ''} — {emp.get('name') or ''}",
        f"Department: {emp.get('department') or '-'}   Designation: {emp.get('designation') or '-'}   "
        f"Category: {emp.get('category') or '-'}"
        + (f"   Contractor: {emp['contractor_name']}" if emp.get("contractor_name") else ""),
        # Iter 710 — the ACTUAL shift name is never printed in dummy mode.
        (f"Dummy Shift: {emp.get('dummy_shift') or 'None'}"
         if data.get("dummy_mode")
         else f"Shift: {emp.get('shift_name') or '-'}")
        + f"   Month: {data['month_number']}/{data['year']}   "
        f"Payroll Period: {data['payroll_period']}",
    ]


def _dummy_summary_lines(data: Dict[str, Any]) -> List[str]:
    """Iter 628 — report summary block (dummy mode, user spec sec 16)."""
    ds = data.get("dummy_summary") or {}
    sc = ds.get("shift_counts") or {}
    return [
        f"SUMMARY — Active Employees: {ds.get('total_employees', 0)} · "
        f"Present Days: {ds.get('present_days', 0)} · "
        f"Week Off: {ds.get('week_off_days', 0)} · "
        f"Holidays: {ds.get('holiday_days', 0)} · "
        f"Absent Days: {ds.get('absent_days', 0)}",
        "Shift-wise Employees: " + (" · ".join(
            f"{k}: {v}" for k, v in sc.items()) or "—"),
    ]


@router.get("/inout-ot-matrix.xlsx")
async def inout_ot_matrix_xlsx(
    month: str,
    company_id: Optional[str] = None,
    department: Optional[str] = None,
    designation: Optional[str] = None,
    employee_type: Optional[str] = None,
    contractor: Optional[str] = None,
    shift: Optional[str] = None,
    q: Optional[str] = None,
    status: str = "active",
    dummy: int = 0,
    dummy_shift: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _, cid = await _auth(authorization, company_id)
    data = await _build(cid, month, department, designation, employee_type,
                        contractor, shift, q, status,
                        dummy=bool(dummy), dummy_shift=dummy_shift)
    emps = data["employees"][:500]
    if not emps:
        raise HTTPException(status_code=404, detail="No employees match the filters")

    # Iter 293 (user request) — ONE sheet with ALL employees stacked in the
    # on-screen sort order (no per-employee sheets).
    wb = Workbook()
    ws = wb.active
    ws.title = f"InOut-OT {month}"
    ws.page_setup.orientation = "landscape"
    thin = Border(*[Side(style="thin", color="CBD5E1")] * 4)
    center = Alignment(horizontal="center", vertical="center")
    r = 1
    # Iter 628 — dummy-mode heading + read-only disclaimer.
    if data.get("dummy_mode"):
        ws.cell(row=r, column=1, value=data["report_title"]).font = Font(
            bold=True, size=13, color="7C2D12")
        r += 1
        ws.cell(row=r, column=1, value=data["dummy_note"]).font = Font(
            bold=True, size=10, color="DC2626")
        r += 2
    # Iter 520 — firm attendance policy line on top of the sheet.
    _pol_line = (data.get("policy") or {}).get("line")
    if _pol_line:
        ws.cell(row=r, column=1, value=_pol_line).font = Font(
            bold=True, size=9, color="0F3B5C")
        r += 2
    for emp in emps:
        for li, line in enumerate(_header_lines(data, emp)):
            ws.cell(row=r, column=1, value=line).font = Font(bold=(li == 0), size=10)
            r += 1
        head_row = r
        hc = ws.cell(row=head_row, column=1, value="Attendance")
        hc.fill = PatternFill("solid", fgColor="1E3A8A")
        hc.font = Font(bold=True, color="FFFFFF")
        for j, dl in enumerate(data["day_labels"], start=2):
            cell = ws.cell(row=head_row, column=j, value=str(dl)[:2])
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = center
            cell.border = thin
        for i, (key, label) in enumerate(_row_keys(data)):
            rr = head_row + 1 + i
            lc = ws.cell(row=rr, column=1, value=label)
            lc.font = Font(bold=True, size=9)
            lc.border = thin
            for j, dl in enumerate(data["day_labels"], start=2):
                d = emp["days"].get(dl) or {}
                cell = ws.cell(row=rr, column=j, value=d.get(key) or "-")
                cell.alignment = center
                cell.font = Font(size=8)
                cell.border = thin
                color = FLAG_COLORS.get(d.get("flag") or "")
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
        sr = head_row + 1 + len(_row_keys(data))
        ws.cell(row=sr, column=1,
                value=f"Month Totals — Working {emp['month_total']}"
                      + ("" if data.get("dummy_mode")
                         else f" · OT {emp['month_ot']}"
                         f" · Total Working {emp.get('month_grand') or '-'}")
                      + f" · Present Days {emp.get('present_days') or 0}").font = Font(bold=True, size=9)
        r = sr + 2  # blank separator row between employees
    # Iter 403 (user accepted) — day-wise OT totals footer.
    # Iter 710 — skipped in dummy mode (OT is fully masked there).
    if not data.get("dummy_mode"):
        ws.cell(row=r, column=1, value="DAY-WISE OT TOTALS (all filtered employees)"
                ).font = Font(bold=True, size=10)
        r += 1
        hc2 = ws.cell(row=r, column=1, value="Day")
        hc2.fill = PatternFill("solid", fgColor="1E3A8A")
        hc2.font = Font(bold=True, color="FFFFFF")
        for j, dl in enumerate(data["day_labels"], start=2):
            cell = ws.cell(row=r, column=j, value=str(dl)[:2])
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = center
            cell.border = thin
        r += 1
        lc2 = ws.cell(row=r, column=1, value="Total OT")
        lc2.font = Font(bold=True, size=9)
        lc2.border = thin
        for j, dl in enumerate(data["day_labels"], start=2):
            v = (data.get("day_ot_totals") or {}).get(dl) or "-"
            cell = ws.cell(row=r, column=j, value=v)
            cell.alignment = center
            cell.font = Font(size=8, bold=(v != "-"))
            cell.border = thin
            if v != "-":
                cell.fill = PatternFill("solid", fgColor="DBEAFE")
        r += 1
        ws.cell(row=r, column=1,
                value=f"Month OT Total: {data.get('month_ot_total') or '-'}"
                ).font = Font(bold=True, size=9)
    # Iter 628 — dummy-mode summary block (sec 16).
    if data.get("dummy_mode"):
        for line in _dummy_summary_lines(data):
            r += 1
            ws.cell(row=r, column=1, value=line).font = Font(bold=True, size=9)
    ws.column_dimensions["A"].width = 13
    for j in range(2, len(data["day_labels"]) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 6.5
    ws.freeze_panes = "B1"  # keep the Attendance-type column visible

    buf = io.BytesIO()
    wb.save(buf)
    _fn = ("dummy-shift-matrix" if data.get("dummy_mode")
           else "inout-ot-matrix")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="{_fn}-{month}.xlsx"'})


@router.get("/inout-ot-matrix.csv")
async def inout_ot_matrix_csv(
    month: str,
    company_id: Optional[str] = None,
    department: Optional[str] = None,
    designation: Optional[str] = None,
    employee_type: Optional[str] = None,
    contractor: Optional[str] = None,
    shift: Optional[str] = None,
    q: Optional[str] = None,
    status: str = "active",
    dummy: int = 0,
    dummy_shift: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    _, cid = await _auth(authorization, company_id)
    data = await _build(cid, month, department, designation, employee_type,
                        contractor, shift, q, status,
                        dummy=bool(dummy), dummy_shift=dummy_shift)
    buf = io.StringIO()
    w = csv.writer(buf)
    _dm = bool(data.get("dummy_mode"))
    if _dm:
        w.writerow([data["report_title"]])
        w.writerow([data["dummy_note"]])
        w.writerow([])
    w.writerow(["Emp Code", "Name", "Department", "Designation", "Shift"]
               + (["Dummy Shift"] if _dm else []) + ["Type"]
               + [str(d)[:2] for d in data["day_labels"]])
    for emp in data["employees"]:
        for key, label in _row_keys(data):
            w.writerow([emp.get("employee_code"), emp.get("name"),
                        emp.get("department"), emp.get("designation"),
                        emp.get("shift_name")]
                       + ([emp.get("dummy_shift") or "None"] if _dm else [])
                       + [label]
                       + [(emp["days"].get(dl) or {}).get(key) or "-"
                          for dl in data["day_labels"]])
    if _dm:
        w.writerow([])
        for line in _dummy_summary_lines(data):
            w.writerow([line])
    return Response(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition":
                 f'attachment; filename="'
                 f'{"dummy-shift-matrix" if _dm else "inout-ot-matrix"}'
                 f'-{month}.csv"'})


@router.get("/inout-ot-matrix.pdf")
async def inout_ot_matrix_pdf(
    month: str,
    company_id: Optional[str] = None,
    department: Optional[str] = None,
    designation: Optional[str] = None,
    employee_type: Optional[str] = None,
    contractor: Optional[str] = None,
    shift: Optional[str] = None,
    q: Optional[str] = None,
    status: str = "active",
    dummy: int = 0,
    dummy_shift: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    """LEGAL LANDSCAPE (Iter 293, user request) — all employees flow
    continuously (multiple per page, header above each matrix), whole month
    on one row-set, colours identical to the screen."""
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import legal, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (KeepTogether, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)
    from reportlab.lib.styles import ParagraphStyle

    _, cid = await _auth(authorization, company_id)
    data = await _build(cid, month, department, designation, employee_type,
                        contractor, shift, q, status,
                        dummy=bool(dummy), dummy_shift=dummy_shift)
    emps = data["employees"][:500]
    if not emps:
        raise HTTPException(status_code=404, detail="No employees match the filters")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(legal),
        leftMargin=8 * mm, rightMargin=8 * mm,
        topMargin=8 * mm, bottomMargin=8 * mm)
    h1 = ParagraphStyle("h1", fontSize=11, leading=14, spaceAfter=1,
                        fontName="Helvetica-Bold")
    h2 = ParagraphStyle("h2", fontSize=8, leading=10)
    flow: List[Any] = []
    ndays = len(data["day_labels"])
    page_w = landscape(legal)[0] - 16 * mm
    label_w = 20 * mm
    day_w = (page_w - label_w) / max(1, ndays)

    flag_fill = {k: rl.HexColor(f"#{v}") for k, v in FLAG_COLORS.items()}
    # Iter 628 — dummy-mode heading + read-only disclaimer on the PDF.
    if data.get("dummy_mode"):
        flow.append(Paragraph(data["report_title"], ParagraphStyle(
            "dt", fontSize=13, leading=16, spaceAfter=2,
            fontName="Helvetica-Bold", textColor=rl.HexColor("#7C2D12"))))
        flow.append(Paragraph(data["dummy_note"], ParagraphStyle(
            "dn", fontSize=8.5, leading=11, spaceAfter=3,
            fontName="Helvetica-Bold", textColor=rl.HexColor("#DC2626"))))
    flow.append(Paragraph(
        "Legend: <font backcolor='#DBEAFE'> OT </font> "
        "<font backcolor='#FEF08A'> Late </font> "
        "<font backcolor='#FECACA'> Missing punch </font> "
        "<font backcolor='#E2E8F0'> Holiday </font> "
        "<font backcolor='#DCFCE7'> Weekly off </font> "
        "<font backcolor='#FED7AA'> Leave </font>", h2))
    # Iter 520 — firm attendance policy line under the legend.
    if (data.get("policy") or {}).get("line"):
        flow.append(Paragraph(data["policy"]["line"], h2))
    flow.append(Spacer(1, 2 * mm))
    for emp in emps:
        block: List[Any] = []
        for li, line in enumerate(_header_lines(data, emp)):
            block.append(Paragraph(line, h1 if li == 0 else h2))
        block.append(Spacer(1, 1.5 * mm))
        head = ["Attendance"] + [str(d)[:2] for d in data["day_labels"]]
        body = []
        styles = [
            ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#1E3A8A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 1.6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.6),
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        ]
        for i, (key, label) in enumerate(_row_keys(data)):
            row = [label]
            for j, dl in enumerate(data["day_labels"]):
                d = emp["days"].get(dl) or {}
                row.append(d.get(key) or "-")
                fill = flag_fill.get(d.get("flag") or "")
                if fill:
                    styles.append(("BACKGROUND", (j + 1, i + 1), (j + 1, i + 1), fill))
            body.append(row)
        tbl = Table([head] + body,
                    colWidths=[label_w] + [day_w] * ndays, repeatRows=1)
        tbl.setStyle(TableStyle(styles))
        block.append(tbl)
        block.append(Spacer(1, 1.5 * mm))
        block.append(Paragraph(
            f"Month Totals — Working {emp['month_total']}"
            + ("" if data.get("dummy_mode")
               else f" · OT {emp['month_ot']} · "
               f"Total Working {emp.get('month_grand') or '-'}")
            + f" · Present Days {emp.get('present_days') or 0}", h2))
        block.append(Spacer(1, 4 * mm))
        # Keep an employee's header + matrix together on one page.
        flow.append(KeepTogether(block))
    # Iter 403 (user accepted) — day-wise OT totals footer.
    # Iter 710 — skipped in dummy mode (OT is fully masked there).
    if not data.get("dummy_mode"):
        ot_head = ["Day"] + [str(d)[:2] for d in data["day_labels"]]
        ot_vals = ["Total OT"] + [(data.get("day_ot_totals") or {}).get(dl) or "-"
                                  for dl in data["day_labels"]]
        ot_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#1E3A8A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 1.6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.6),
        ]
        for j, v in enumerate(ot_vals[1:]):
            if v != "-":
                ot_styles.append(("BACKGROUND", (j + 1, 1), (j + 1, 1),
                                  rl.HexColor("#DBEAFE")))
        ot_tbl = Table([ot_head, ot_vals],
                       colWidths=[label_w] + [day_w] * ndays)
        ot_tbl.setStyle(TableStyle(ot_styles))
        flow.append(KeepTogether([
            Paragraph(f"Day-wise OT Totals (all filtered employees) — "
                      f"Month OT Total: {data.get('month_ot_total') or '-'}", h1),
            Spacer(1, 1.5 * mm), ot_tbl]))
    # Iter 628 — dummy-mode summary block (sec 16).
    if data.get("dummy_mode"):
        flow.append(Spacer(1, 3 * mm))
        for line in _dummy_summary_lines(data):
            flow.append(Paragraph(line, h2))
    doc.build(flow)
    _fn = ("dummy-shift-matrix" if data.get("dummy_mode")
           else "inout-ot-matrix")
    return Response(
        content=buf.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition":
                 f'inline; filename="{_fn}-{month}.pdf"'})
