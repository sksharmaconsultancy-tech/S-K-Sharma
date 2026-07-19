"""Iter 204 — Employee Shift Change Management (generalized, policy-gated).

Distinct from the legacy hospital shift-SWAP flow (routes/shift_change.py):
this module implements the enterprise request→approval workflow:

  employee request → (optional two-level) approval → daily shift assignment
  → attendance engine picks the APPROVED shift for that day automatically
  (grid, IN/OUT / OT / HRS reports and payroll views all recompute from it).

Config lives on the firm's Attendance Policy under ``shift_change``:
  enabled, reason_mandatory, post_punch_allowed, auto_approve,
  time_window (any|prev_day|before_shift_start|within_2h),
  approval_levels (single|two_level).
"""
import io
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query, Request
from fastapi.responses import Response

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
    now_iso,
    sub_admin_can_touch_company,
)

router = APIRouter(prefix="/api", tags=["shift-change-v2"])

IST = timezone(timedelta(hours=5, minutes=30))

SC_DEFAULTS = {
    "enabled": False,
    "reason_mandatory": True,
    "post_punch_allowed": False,
    "auto_approve": False,
    "time_window": "any",
    "approval_levels": "single",
}


def _sc_config(company: dict) -> dict:
    raw = ((company or {}).get("attendance_policy") or {}).get("shift_change") or {}
    cfg = dict(SC_DEFAULTS)
    cfg.update({k: raw[k] for k in SC_DEFAULTS if k in raw})
    return cfg


async def _company_for(user: dict, company_id: Optional[str] = None) -> dict:
    cid = company_id or user.get("company_id")
    if not cid:
        raise HTTPException(status_code=400, detail="No firm linked")
    co = await db.companies.find_one({"company_id": cid}, {"_id": 0})
    if not co:
        raise HTTPException(status_code=404, detail="Firm not found")
    return co


async def _shifts() -> List[dict]:
    return await db.shift_masters.find({}, {"_id": 0}).to_list(200)


def _notify(user_id: str, title: str, message: str, company_id: Optional[str]):
    return db.notifications.insert_one({
        "notification_id": f"ntf_{uuid.uuid4().hex[:10]}",
        "user_id": user_id,
        "company_id": company_id,
        "type": "shift_change",
        "title": title,
        "message": message,
        "read": False,
        "created_at": now_iso(),
    })


async def _month_locked(company_id: str, month: str) -> bool:
    """Payroll lock = a finalized salary run exists for that month."""
    for coll in (db.salary_runs, db.compliance_salary_runs):
        if await coll.find_one({"company_id": company_id, "month": month,
                                "finalized": True}, {"_id": 1}):
            return True
    return False


# ---------------------------------------------------------------------------
# EMPLOYEE — config + create + my requests + cancel
# ---------------------------------------------------------------------------
@router.get("/shift-change/config")
async def shift_change_config(authorization: Optional[str] = Header(None)):
    user = await get_user_from_token(authorization)
    co = await _company_for(user)
    cfg = _sc_config(co)
    shifts = await _shifts()
    return {
        "config": cfg,
        "company_name": co.get("name"),
        "shifts": [{"shift_id": s.get("shift_id"), "name": s.get("name"),
                    "start": s.get("start"), "end": s.get("end")} for s in shifts],
        "current_shift": {
            "shift_id": user.get("shift_id"),
            "name": user.get("shift_name"),
            "start": user.get("shift_start"),
            "end": user.get("shift_end"),
        },
    }


@router.post("/shift-change/requests-v2")
async def create_request(
    request: Request,
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    user = await get_user_from_token(authorization)
    require_role(user, ["employee"])
    co = await _company_for(user)
    cfg = _sc_config(co)
    if not cfg["enabled"]:
        raise HTTPException(status_code=403, detail="Shift change requests are disabled for your firm")

    date = str(payload.get("date") or "").strip()
    shift_id = str(payload.get("requested_shift_id") or "").strip()
    reason = str(payload.get("reason") or "").strip()
    remarks = str(payload.get("remarks") or "").strip()
    if not date or len(date) != 10:
        raise HTTPException(status_code=400, detail="Pick a valid date (YYYY-MM-DD)")
    if not shift_id:
        raise HTTPException(status_code=400, detail="Pick the requested shift")
    if cfg["reason_mandatory"] and not reason:
        raise HTTPException(status_code=400, detail="Reason is mandatory")

    # Validations -----------------------------------------------------------
    if user.get("exit_date") and str(user["exit_date"]) <= date:
        raise HTTPException(status_code=400, detail="Cannot request for a resigned/exited period")
    shift = await db.shift_masters.find_one({"shift_id": shift_id}, {"_id": 0})
    if not shift:
        raise HTTPException(status_code=404, detail="Requested shift not found")
    if shift_id == user.get("shift_id"):
        raise HTTPException(status_code=400, detail="Requested shift is the same as your current shift")
    dup = await db.shift_change_requests_v2.find_one({
        "user_id": user["user_id"], "date": date,
        "status": {"$in": ["pending", "pending_final", "approved"]}}, {"_id": 1})
    if dup:
        raise HTTPException(status_code=400, detail="You already have a request for this date")
    if await _month_locked(co["company_id"], date[:7]):
        raise HTTPException(status_code=400, detail="Payroll for this month is locked — request not allowed")

    now_ist = datetime.now(IST)
    # Time-window rule
    tw = cfg["time_window"]
    shift_start_dt = None
    try:
        hh, mm = int(shift["start"][:2]), int(shift["start"][3:5])
        shift_start_dt = datetime.strptime(date, "%Y-%m-%d").replace(
            hour=hh, minute=mm, tzinfo=IST)
    except Exception:
        pass
    if tw == "prev_day" and date <= now_ist.strftime("%Y-%m-%d"):
        raise HTTPException(status_code=400, detail="Requests must be made at least one day in advance")
    if tw == "before_shift_start" and shift_start_dt and now_ist >= shift_start_dt:
        raise HTTPException(status_code=400, detail="Request must be made before the shift starts")
    if tw == "within_2h" and shift_start_dt and (
            now_ist < shift_start_dt - timedelta(hours=2) or now_ist >= shift_start_dt):
        raise HTTPException(status_code=400, detail="Request allowed only within 2 hours before shift start")

    # Post-punch rule
    punched = await db.attendance.find_one(
        {"user_id": user["user_id"], "date": date}, {"_id": 1})
    if punched and not cfg["post_punch_allowed"]:
        raise HTTPException(status_code=400,
                            detail="You have already punched on this date — post-punch shift change is not allowed")

    # Approved-leave rule
    leave = await db.leaves.find_one({
        "user_id": user["user_id"], "status": "approved",
        "from_date": {"$lte": date}, "to_date": {"$gte": date}}, {"_id": 1})
    if leave:
        raise HTTPException(status_code=400, detail="You are on approved leave on this date")

    req_no = f"SCR-{datetime.now(IST).strftime('%y%m')}-{uuid.uuid4().hex[:4].upper()}"
    doc = {
        "request_id": f"scr_{uuid.uuid4().hex[:10]}",
        "request_no": req_no,
        "flow": "request_v2",
        "user_id": user["user_id"],
        "employee_code": user.get("employee_code"),
        "employee_name": user.get("name"),
        "company_id": co["company_id"],
        "company_name": co.get("name"),
        "date": date,
        "old_shift": {"shift_id": user.get("shift_id"), "name": user.get("shift_name"),
                      "start": user.get("shift_start"), "end": user.get("shift_end")},
        "requested_shift": {"shift_id": shift_id, "name": shift.get("name"),
                            "start": shift.get("start"), "end": shift.get("end")},
        "reason": reason,
        "remarks": remarks,
        "post_punch": bool(punched),
        "status": "pending",
        "history": [{
            "action": "submitted", "by": user["user_id"], "by_name": user.get("name"),
            "at": now_iso(), "ip": (request.client.host if request.client else None),
            "device": request.headers.get("user-agent", "")[:160],
        }],
        "created_at": now_iso(),
    }
    await db.shift_change_requests_v2.insert_one(dict(doc))

    if cfg["auto_approve"]:
        await _apply_approval(doc, {"user_id": "system", "name": "Auto-approve"},
                              "Auto-approved by firm policy")
        doc["status"] = "approved"
    else:
        async for a in db.users.find(
                {"role": {"$in": ["super_admin", "sub_admin", "company_admin"]}},
                {"_id": 0, "user_id": 1, "role": 1, "company_id": 1}):
            if a["role"] == "company_admin" and a.get("company_id") != co["company_id"]:
                continue
            await _notify(a["user_id"], "New shift change request",
                          f"{user.get('name')} ({req_no}) requested "
                          f"{shift.get('name')} on {date}", co["company_id"])

    doc.pop("_id", None)
    return {"ok": True, "request": doc}


@router.get("/shift-change/requests-v2/my")
async def my_requests(
    status: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    user = await get_user_from_token(authorization)
    q: Dict[str, Any] = {"user_id": user["user_id"]}
    if status:
        q["status"] = status
    rows = await db.shift_change_requests_v2.find(q, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"rows": rows}


@router.post("/shift-change/requests-v2/{request_id}/cancel")
async def cancel_request(
    request_id: str,
    authorization: Optional[str] = Header(None),
):
    user = await get_user_from_token(authorization)
    req = await db.shift_change_requests_v2.find_one(
        {"request_id": request_id, "user_id": user["user_id"]}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req["status"] not in ("pending", "pending_final", "sent_back"):
        raise HTTPException(status_code=400, detail="Only pending requests can be cancelled")
    await db.shift_change_requests_v2.update_one(
        {"request_id": request_id},
        {"$set": {"status": "cancelled"},
         "$push": {"history": {"action": "cancelled", "by": user["user_id"],
                               "by_name": user.get("name"), "at": now_iso()}}})
    return {"ok": True}


# ---------------------------------------------------------------------------
# ADMIN — list + decide (single/bulk) + register + daily assignments
# ---------------------------------------------------------------------------
async def _admin(authorization, company_id: Optional[str] = None):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if company_id:
        if admin["role"] == "sub_admin" and not sub_admin_can_touch_company(admin, company_id):
            raise HTTPException(status_code=403, detail="Firm not in your scope")
        if admin["role"] == "company_admin" and admin.get("company_id") != company_id:
            raise HTTPException(status_code=403, detail="Not your firm")
    return admin


async def _apply_approval(req: dict, approver: dict, remarks: str):
    """Write the daily shift assignment — the attendance engine picks it up
    automatically (grid/reports/payroll views recompute per request)."""
    rs = req["requested_shift"]
    await db.daily_shift_assignments.update_one(
        {"user_id": req["user_id"], "date": req["date"]},
        {"$set": {
            "user_id": req["user_id"],
            "company_id": req["company_id"],
            "date": req["date"],
            "shift_id": rs.get("shift_id"),
            "name": rs.get("name"),
            "start": rs.get("start"),
            "end": rs.get("end"),
            "source": "shift_change_request",
            "request_id": req["request_id"],
            "assigned_at": now_iso(),
            "assigned_by": approver.get("user_id"),
        }},
        upsert=True,
    )
    await db.shift_change_requests_v2.update_one(
        {"request_id": req["request_id"]},
        {"$set": {"status": "approved", "approved_shift": rs,
                  "approved_by": approver.get("user_id"),
                  "approved_by_name": approver.get("name"),
                  "approved_at": now_iso(), "approval_remarks": remarks},
         "$push": {"history": {"action": "approved", "by": approver.get("user_id"),
                               "by_name": approver.get("name"), "at": now_iso(),
                               "remarks": remarks}}})
    await _notify(req["user_id"], "Shift change approved",
                  f"{req['request_no']}: {rs.get('name')} on {req['date']} approved. "
                  "Attendance will be calculated on the new shift.",
                  req.get("company_id"))


@router.get("/admin/shift-change/requests-v2")
async def admin_list_requests(
    company_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    month: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await _admin(authorization, company_id)
    q: Dict[str, Any] = {}
    if company_id:
        q["company_id"] = company_id
    elif admin["role"] == "company_admin":
        q["company_id"] = admin.get("company_id")
    if status:
        q["status"] = status
    if month:
        q["date"] = {"$gte": f"{month}-01", "$lte": f"{month}-31"}
    rows = await db.shift_change_requests_v2.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    counts: Dict[str, int] = {}
    for r in rows:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    return {"rows": rows, "counts": counts}


@router.post("/admin/shift-change/requests-v2/decide")
async def admin_decide(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    ids = payload.get("request_ids") or []
    action = str(payload.get("action") or "")
    remarks = str(payload.get("remarks") or "")
    if action not in ("approve", "reject", "send_back"):
        raise HTTPException(status_code=400, detail="action must be approve / reject / send_back")
    if not ids:
        raise HTTPException(status_code=400, detail="No requests selected")

    done, results = 0, []
    for rid in ids:
        req = await db.shift_change_requests_v2.find_one({"request_id": rid}, {"_id": 0})
        if not req or req["status"] not in ("pending", "pending_final", "sent_back"):
            results.append({"request_id": rid, "ok": False, "error": "Not pending"})
            continue
        if admin["role"] == "sub_admin" and not sub_admin_can_touch_company(admin, req["company_id"]):
            results.append({"request_id": rid, "ok": False, "error": "Out of scope"})
            continue
        if admin["role"] == "company_admin" and admin.get("company_id") != req["company_id"]:
            results.append({"request_id": rid, "ok": False, "error": "Not your firm"})
            continue

        if action == "approve":
            co = await db.companies.find_one(
                {"company_id": req["company_id"]}, {"_id": 0, "attendance_policy": 1})
            cfg = _sc_config(co or {})
            # Two-level flow: company_admin approval moves to pending_final;
            # super/sub admin approval is always final.
            if (cfg["approval_levels"] == "two_level"
                    and admin["role"] == "company_admin"
                    and req["status"] == "pending"):
                await db.shift_change_requests_v2.update_one(
                    {"request_id": rid},
                    {"$set": {"status": "pending_final"},
                     "$push": {"history": {"action": "approved_level1",
                                           "by": admin["user_id"], "by_name": admin.get("name"),
                                           "at": now_iso(), "remarks": remarks}}})
                results.append({"request_id": rid, "ok": True, "status": "pending_final"})
                done += 1
                continue
            await _apply_approval(req, admin, remarks)
            results.append({"request_id": rid, "ok": True, "status": "approved"})
        else:
            new_status = "rejected" if action == "reject" else "sent_back"
            await db.shift_change_requests_v2.update_one(
                {"request_id": rid},
                {"$set": {"status": new_status, "approval_remarks": remarks,
                          "decided_by": admin["user_id"], "decided_at": now_iso()},
                 "$push": {"history": {"action": new_status, "by": admin["user_id"],
                                       "by_name": admin.get("name"), "at": now_iso(),
                                       "remarks": remarks}}})
            # Rejected → original shift stays; attendance engine untouched.
            await _notify(req["user_id"],
                          f"Shift change {new_status.replace('_', ' ')}",
                          f"{req['request_no']} for {req['date']}: {new_status.replace('_', ' ')}."
                          f"{' ' + remarks if remarks else ''} Attendance stays on your original shift.",
                          req.get("company_id"))
            results.append({"request_id": rid, "ok": True, "status": new_status})
        done += 1
    return {"ok": True, "processed": done, "results": results}


def _register_xlsx(title: str, rows: List[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Change Register"
    hdrs = ["Request No", "Date", "Emp Code", "Employee", "Old Shift",
            "Requested Shift", "Reason", "Status", "Requested At",
            "Decided By", "Decided At", "Remarks"]
    ws.append([title])
    ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F3D7A")
    ws.append(hdrs)
    for c in ws[2]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3D7A")
    for r in rows:
        old = r.get("old_shift") or {}
        new = r.get("requested_shift") or {}
        ws.append([
            r.get("request_no"), r.get("date"), r.get("employee_code"),
            r.get("employee_name"),
            f"{old.get('name') or '-'} ({old.get('start') or ''}-{old.get('end') or ''})",
            f"{new.get('name') or '-'} ({new.get('start') or ''}-{new.get('end') or ''})",
            r.get("reason"), (r.get("status") or "").upper(),
            str(r.get("created_at") or "")[:16].replace("T", " "),
            r.get("approved_by_name") or r.get("decided_by") or "",
            str(r.get("approved_at") or r.get("decided_at") or "")[:16].replace("T", " "),
            r.get("approval_remarks") or "",
        ])
    for i, w in enumerate([14, 11, 9, 22, 22, 22, 24, 11, 17, 16, 17, 24], start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/admin/shift-change/register")
async def shift_change_register(
    company_id: str = Query(...),
    month: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    fmt: str = Query("json"),
    authorization: Optional[str] = Header(None),
):
    await _admin(authorization, company_id)
    q: Dict[str, Any] = {"company_id": company_id}
    if month:
        q["date"] = {"$gte": f"{month}-01", "$lte": f"{month}-31"}
    if status:
        q["status"] = status
    rows = await db.shift_change_requests_v2.find(q, {"_id": 0}).sort("date", -1).to_list(2000)
    if fmt == "xlsx":
        data = _register_xlsx(
            f"Shift Change Register — {month or 'All'}{' — ' + status.upper() if status else ''}",
            rows)
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":
                     f'attachment; filename="ShiftChangeRegister_{month or "all"}.xlsx"'})
    return {"rows": rows, "count": len(rows)}


@router.get("/admin/shift-change/daily-assignments")
async def daily_assignments_report(
    company_id: str = Query(...),
    month: str = Query(...),
    fmt: str = Query("json"),
    authorization: Optional[str] = Header(None),
):
    await _admin(authorization, company_id)
    rows = await db.daily_shift_assignments.find(
        {"company_id": company_id, "date": {"$gte": f"{month}-01", "$lte": f"{month}-31"}},
        {"_id": 0}).sort("date", 1).to_list(3000)
    uids = {r["user_id"] for r in rows}
    names = {}
    if uids:
        async for u in db.users.find({"user_id": {"$in": list(uids)}},
                                     {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1}):
            names[u["user_id"]] = u
    for r in rows:
        u = names.get(r["user_id"]) or {}
        r["employee_name"] = u.get("name")
        r["employee_code"] = u.get("employee_code")
    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Shift Assignments"
        ws.append([f"Daily Shift Assignment Report — {month}"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F3D7A")
        ws.append(["Date", "Emp Code", "Employee", "Shift", "Start", "End", "Source", "Assigned By"])
        for c in ws[2]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3D7A")
        for r in rows:
            ws.append([r.get("date"), r.get("employee_code"), r.get("employee_name"),
                       r.get("name"), r.get("start"), r.get("end"),
                       r.get("source"), r.get("assigned_by")])
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":
                     f'attachment; filename="DailyShiftAssignments_{month}.xlsx"'})
    return {"rows": rows, "count": len(rows)}
