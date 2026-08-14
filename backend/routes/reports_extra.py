"""Iter 86 - Route module: extra reports (users-log activity feed).

First endpoint extracted from the monolithic `server.py` as a proof of
concept for the modularization effort.

The endpoint pattern used here is the template for all future extracts:

  1) `router = APIRouter(prefix="/api")` - preserves the `/api` prefix
     so URLs are unchanged.
  2) Shared state (`db`, `get_user_from_token`, `require_role`) is
     imported lazily from `server` so this module doesn't need to
     duplicate the FastAPI app / motor client setup.
  3) `server.py` includes this router at the very bottom of its file,
     AFTER all shared helpers are defined.  That ordering breaks the
     apparent circular import because at the moment `server.py` runs
     ``from routes.reports_extra import router``, all names this
     sub-module needs are already bound on the `server` module object.

Endpoints:
  * GET /api/admin/users-log - Unified activity feed across:
      - company_audit_log
      - attendance_audit_log
      - salary_runs (generated_at + finalized_at)
      - compliance_salary_runs (generated_at + finalized_at)
    Filters: from_date, to_date, company_id, user_id.
"""
from typing import Optional, List
import re
from fastapi import APIRouter, Header, Query

# Shared helpers live on the `server` module.  Importing them here at
# module-load time is safe because `server.py` only pulls this
# sub-module in at the very bottom of its file - long after `db`,
# `get_user_from_token`, and `require_role` are bound.
from server import db, get_user_from_token, require_role  # noqa: E402
from shared.audit import derive_module  # noqa: E402

router = APIRouter(prefix="/api")


# Iter 572 — strip internal-ID tokens ("company_id=cmp_x, user_id=...") that
# older log rows stored in details; firm/user names are shown separately.
_ID_TOKEN = re.compile(r"\b\w*_?id=[^,\s]+[,]?\s*", re.IGNORECASE)


def _clean_details(details: str) -> str:
    return _ID_TOKEN.sub("", details or "").strip(" ,")


# Iter 572 — step-by-step human-readable narration of every audit event.
def _pretty_field(f: str) -> str:
    return (f or "").replace("_", " ").strip().title()


_SECURITY_PHRASES = [
    ("OTP_GENERATED", "generated a login OTP"),
    ("OTP_SENT_EMAIL", "was sent a login OTP by Email"),
    ("OTP_SENT_WHATSAPP", "was sent a login OTP on WhatsApp"),
    ("OTP_SENT_SMS", "was sent a login OTP by SMS"),
    ("OTP_VERIFICATION_SUCCESS", "verified the login OTP and signed in"),
    ("OTP_VERIFICATION_FAILED", "entered a wrong login OTP"),
    ("OTP_BLOCKED", "was blocked after too many wrong OTP attempts"),
    ("OTP_EXPIRED", "tried an expired login OTP"),
    ("OTP_RESEND", "requested the OTP again"),
    ("2FA_METHOD_CHANGED", "changed the 2FA verification method"),
    ("2FA_SETTINGS_UPDATED", "updated the security (2FA) settings"),
    ("TRUSTED_DEVICE_ADDED", "trusted this device for future logins"),
    ("TRUSTED_DEVICE_REVOKED", "revoked a trusted device"),
    ("TRUSTED_DEVICE_LOGIN", "signed in from a trusted device (OTP skipped)"),
    ("ALL_SESSIONS_REVOKED", "logged out from all devices"),
    ("NEW_IP_LOGIN", "signed in from a NEW IP address"),
]


def _describe(ev: dict) -> str:
    """Builds 'S.K. Sharma Admin (Super Admin) updated Employee RAJESH —
    changed Designation from Helper to Supervisor' style narration."""
    actor = ev.get("actor_name") or "System"
    if actor == "-":
        actor = "System"
    role = ev.get("actor_role") or ""
    who = f"{actor} ({role.replace('_', ' ').title()})" if role else actor
    action = ev.get("action") or ""
    t = ev.get("action_type") or "OTHER"
    module = ev.get("module") or "record"
    label = ev.get("record_label") or ""
    target = f"{module}" + (f" “{label}”" if label else "")
    failed = ev.get("success") is False

    for key, phrase in _SECURITY_PHRASES:
        if key in action:
            out = f"{who} {phrase}"
            return out + (" — FAILED" if failed and "FAILED" not in phrase.upper()
                          and "wrong" not in phrase and "blocked" not in phrase else "")

    if t == "LOGIN":
        path = ev.get("path") or ""
        how = "PIN" if "pin" in path else ("password" if "password" in path else "credentials")
        return f"{who} {'failed to sign in' if failed else 'signed in'} with {how}"

    if t == "DOWNLOAD":
        name = (ev.get("path") or action).rstrip("/").split("/")[-1].split("?")[0]
        return f"{who} downloaded “{name}”"

    changes = ev.get("changes") or []
    if t == "UPDATE" and changes:
        steps = "; ".join(
            f"{_pretty_field(c.get('field'))}: “{(c.get('old') or '—')[:60]}” → “{(c.get('new') or '—')[:60]}”"
            for c in changes[:3])
        more = f" (+{len(changes) - 3} more fields)" if len(changes) > 3 else ""
        return f"{who} updated {target} — changed {steps}{more}"
    if t == "UPDATE":
        det = _clean_details(ev.get("details") or "")
        return f"{who} updated {target}" + (f" ({det[:120]})" if det else "") + (" — FAILED" if failed else "")
    if t == "DELETE":
        return f"{who} deleted {target}" + (" — FAILED" if failed else "")
    if t == "CREATE":
        det = _clean_details(ev.get("details") or "")
        return f"{who} created {target}" + (f" ({det[:120]})" if det and not label else "") + (" — FAILED" if failed else "")
    det = _clean_details(ev.get("details") or "")
    return f"{who} — {action}" + (f" ({det[:100]})" if det else "")


def _action_type(action: str, source: str) -> str:
    """Normalize an event to CREATE/UPDATE/DELETE/LOGIN/DOWNLOAD/OTHER."""
    a = (action or "").upper()
    for t in ("CREATE", "UPDATE", "DELETE", "LOGIN", "DOWNLOAD", "AUTH"):
        if a.startswith(t):
            return "LOGIN" if t == "AUTH" else t
    if source in ("salary_runs", "compliance_salary_runs"):
        return "UPDATE" if "finalized" in a.lower() else "CREATE"
    if source == "attendance_audit_log":
        return "UPDATE"
    return "OTHER"


# Iter 85 - Users Log Report - unified activity feed.
#
# Aggregates events from four sources into a single date-filtered
# stream:
#   * company_audit_log   - admin actions (approvals, PIN changes, ...)
#   * attendance_audit_log - punch decisions (approve/reject/edit)
#   * salary_runs          - generated_at + finalized_at
#   * compliance_salary_runs - generated_at + finalized_at
#
# Filters:
#   from_date / to_date  (YYYY-MM-DD, inclusive)
#   company_id           (optional; scopes to a single firm)
#   user_id              (optional; scopes to a single actor)
@router.get("/admin/users-log")
async def users_log_report(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    company_id: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),
    module: Optional[str] = Query(None),
    action_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),  # "success" | "failed"
    search: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    scope_cid = admin.get("company_id") if admin["role"] == "company_admin" else company_id

    date_range: dict = {}
    if from_date:
        date_range["$gte"] = f"{from_date}T00:00:00"
    if to_date:
        date_range["$lte"] = f"{to_date}T23:59:59"

    def _apply(q: dict, ts_field: str) -> dict:
        if date_range:
            q[ts_field] = date_range
        if scope_cid:
            q["company_id"] = scope_cid
        return q

    events: List[dict] = []

    # 0) activity_log — the FULL automatic action trail (Iter 247): every
    #    create/update/delete + report download by any logged-in user.
    #    Iter 568 — now carries field-level old→new diffs, module, IP,
    #    device and success flag (Detailed Audit Trail).
    async for e in db.activity_log.find(_apply({}, "at"), {"_id": 0}).sort("at", -1).limit(3000):
        st = e.get("status")
        failed = (e.get("success") is False) or (isinstance(st, int) and st >= 400)
        extra = f" [FAILED {st}]" if failed else ""
        events.append({
            "at": e.get("at"),
            "actor_id": e.get("actor_id"),
            "action": e.get("action") or f"{e.get('method')} {e.get('path')}",
            "company_id": e.get("company_id"),
            "details": (_clean_details(e.get("details") or "") + extra).strip(),
            "source": "activity_log",
            "module": e.get("module") or derive_module(e.get("path") or ""),
            "success": not failed,
            "status_code": st,
            "ip": e.get("ip") or "",
            "device": e.get("device") or "",
            "method": e.get("method") or "",
            "path": e.get("path") or "",
            "record_id": e.get("record_id"),
            "record_label": e.get("record_label") or "",
            "changes": e.get("changes") or [],
            "old_values": e.get("old_values"),
            "new_values": e.get("new_values"),
        })

    # 1) company_audit_log - generic admin actions
    async for e in db.company_audit_log.find(_apply({}, "at"), {"_id": 0}).sort("at", -1).limit(1000):
        events.append({
            "at": e.get("at"),
            "actor_id": e.get("actor_id") or e.get("user_id"),
            "action": e.get("action") or e.get("kind") or "action",
            "company_id": e.get("company_id"),
            "details": e.get("details") or e.get("note") or "",
            "source": "company_audit_log",
        })

    # 2) attendance_audit_log - punch approve/reject/edit
    async for e in db.attendance_audit_log.find(_apply({}, "at"), {"_id": 0}).sort("at", -1).limit(1000):
        events.append({
            "at": e.get("at"),
            "actor_id": e.get("admin_id") or e.get("actor_id"),
            "action": f"punch.{e.get('action') or 'decision'}",
            "company_id": e.get("company_id"),
            "details": e.get("reason") or e.get("note") or "",
            "source": "attendance_audit_log",
        })

    # 3) salary_runs (Actual Salary)
    async for e in db.salary_runs.find(
        _apply({}, "generated_at"),
        {"_id": 0, "generated_by": 1, "generated_at": 1, "month": 1, "company_id": 1, "run_type": 1, "finalized_by": 1, "finalized_at": 1},
    ).sort("generated_at", -1).limit(500):
        events.append({
            "at": e.get("generated_at"),
            "actor_id": e.get("generated_by"),
            "action": "salary.generated",
            "company_id": e.get("company_id"),
            "details": f"month={e.get('month')} type={e.get('run_type', 'actual')}",
            "source": "salary_runs",
        })
        if e.get("finalized_at"):
            events.append({
                "at": e.get("finalized_at"),
                "actor_id": e.get("finalized_by"),
                "action": "salary.finalized",
                "company_id": e.get("company_id"),
                "details": f"month={e.get('month')}",
                "source": "salary_runs",
            })

    # 4) compliance_salary_runs
    async for e in db.compliance_salary_runs.find(
        _apply({}, "generated_at"),
        {"_id": 0, "generated_by": 1, "generated_at": 1, "month": 1, "company_id": 1},
    ).sort("generated_at", -1).limit(500):
        events.append({
            "at": e.get("generated_at"),
            "actor_id": e.get("generated_by"),
            "action": "compliance.generated",
            "company_id": e.get("company_id"),
            "details": f"month={e.get('month')}",
            "source": "compliance_salary_runs",
        })

    # Iter 568 — normalize every event: module, success flag, action_type.
    _src_module = {
        "company_audit_log": "Admin",
        "attendance_audit_log": "Attendance",
        "salary_runs": "Payroll",
        "compliance_salary_runs": "Compliance",
    }
    for ev in events:
        ev.setdefault("module", _src_module.get(ev.get("source") or "", "Other"))
        ev.setdefault("success", True)
        ev["action_type"] = _action_type(ev.get("action") or "", ev.get("source") or "")

    # Filter by user_id if requested (applied after aggregation)
    if user_id:
        events = [ev for ev in events if ev.get("actor_id") == user_id]
    # Iter 568 — advanced filters.
    if module:
        events = [ev for ev in events if (ev.get("module") or "").lower() == module.lower()]
    if action_type:
        events = [ev for ev in events if ev.get("action_type") == action_type.upper()]
    if status == "failed":
        events = [ev for ev in events if not ev.get("success")]
    elif status == "success":
        events = [ev for ev in events if ev.get("success")]

    # Enrich with actor + company names for a nicer display
    actor_ids = {ev.get("actor_id") for ev in events if ev.get("actor_id")}
    cids = {ev.get("company_id") for ev in events if ev.get("company_id")}
    actor_names: dict = {}
    if actor_ids:
        async for u in db.users.find(
            {"user_id": {"$in": list(actor_ids)}},
            {"_id": 0, "user_id": 1, "name": 1, "role": 1, "phone": 1},
        ):
            actor_names[u["user_id"]] = {
                "name": u.get("name") or "-",
                "role": u.get("role") or "",
                "phone": u.get("phone") or "",
            }
    company_names: dict = {}
    if cids:
        async for c in db.companies.find(
            {"company_id": {"$in": list(cids)}},
            {"_id": 0, "company_id": 1, "name": 1},
        ):
            company_names[c["company_id"]] = c.get("name") or "-"
    for ev in events:
        actor = actor_names.get(ev.get("actor_id") or "") or {}
        ev["actor_name"] = actor.get("name") or "-"
        ev["actor_role"] = actor.get("role") or ""
        ev["company_name"] = company_names.get(ev.get("company_id") or "") or "-"
        # Iter 572 — step-by-step human narration of the action.
        ev["description"] = _describe(ev)

    # Iter 568 — free-text search across the enriched events.
    if search and search.strip():
        s = search.strip().lower()
        events = [
            ev for ev in events
            if s in " ".join((
                ev.get("action") or "", ev.get("details") or "",
                ev.get("actor_name") or "", ev.get("record_label") or "",
                ev.get("path") or "", ev.get("module") or "",
                ev.get("company_name") or "", ev.get("ip") or "",
                ev.get("description") or "",
            )).lower()
        ]

    # Sort DESC by timestamp
    events.sort(key=lambda ev: (ev.get("at") or ""), reverse=True)
    events = events[:2000]

    # Iter 568 — summary cards for the Detailed Audit Trail.
    summary = {
        "total": len(events),
        "creates": sum(1 for e in events if e.get("action_type") == "CREATE"),
        "updates": sum(1 for e in events if e.get("action_type") == "UPDATE"),
        "deletes": sum(1 for e in events if e.get("action_type") == "DELETE"),
        "logins": sum(1 for e in events if e.get("action_type") == "LOGIN"),
        "downloads": sum(1 for e in events if e.get("action_type") == "DOWNLOAD"),
        "failed": sum(1 for e in events if not e.get("success")),
        "unique_users": len({e.get("actor_id") for e in events if e.get("actor_id")}),
    }
    mod_counts: dict = {}
    for e in events:
        m = e.get("module") or "Other"
        mod_counts[m] = mod_counts.get(m, 0) + 1
    summary["modules"] = dict(sorted(mod_counts.items(), key=lambda kv: -kv[1]))
    return {"events": events, "count": len(events), "summary": summary}


# Iter 247 — Excel export of the SAME filtered log (full report with
# date and time, one row per action).
@router.get("/admin/users-log.xlsx")
async def users_log_report_xlsx(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    company_id: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),
    module: Optional[str] = Query(None),
    action_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    data = await users_log_report(
        from_date, to_date, company_id, user_id,
        module, action_type, status, search, authorization)
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import Response

    wb = Workbook()
    ws = wb.active
    ws.title = "Users Log"
    ws.append(["Date", "Time", "User", "Role", "Firm", "Module", "Type",
               "Description (Step-by-step)", "Record",
               "Changes (Old → New)", "Details", "IP", "Status", "Source"])
    for c in ws[1]:
        c.font = Font(bold=True)
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    for ev in data["events"]:
        at = ev.get("at") or ""
        d = f"{at[8:10]}-{at[5:7]}-{at[0:4]}" if len(at) >= 10 else ""
        t = at[11:19] if len(at) >= 19 else ""
        chg = "; ".join(
            f"{c.get('field')}: '{c.get('old')}' → '{c.get('new')}'"
            for c in (ev.get("changes") or [])
        )[:800]
        ws.append([
            d, t,
            ev.get("actor_name") or "-",
            ev.get("actor_role") or "",
            ev.get("company_name") or "-",
            ev.get("module") or "",
            ev.get("action_type") or "",
            ev.get("description") or ev.get("action") or "",
            ev.get("record_label") or ev.get("record_id") or "",
            chg,
            (ev.get("details") or "")[:500],
            ev.get("ip") or "",
            "FAILED" if not ev.get("success") else "OK",
            ev.get("source") or "",
        ])
        if not ev.get("success"):
            for c in ws[ws.max_row]:
                c.fill = fail_fill
    for col, w in zip("ABCDEFGHIJKLMN", (12, 10, 22, 14, 24, 14, 10, 40, 18, 55, 45, 15, 9, 18)):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"users-log-{from_date or 'all'}-to-{to_date or 'all'}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
