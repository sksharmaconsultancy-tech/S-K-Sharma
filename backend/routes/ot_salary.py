"""Iter 129 — OT Salary Process (user request).

A SEPARATE salary process that pays OVERTIME hours only — available
EXCLUSIVELY for firms whose Firm/Attendance policy is Textile *Policy 2*
(where OT hours are tracked separately from present days).

Rules (user-directed):
  • OT rate basis is configurable per firm: % of BASIC or % of GROSS
    (per-day amounts from the Employee Master's Actual salary structure).
  • Recorded OT HRS are shown as "OT Duty HRS ÷ 2" (configurable divisor).
  • Payment = recorded OT HRS × (per-day base × % ÷ full-day-hours).
  • A downloadable Bank Sheet (XLSX) with account / IFSC per employee.
  • Compliance & Actual salary processes are NOT touched — OT is paid
    here ONLY (no double payment).

Config is auto-saved on the company doc under ``ot_salary_cfg``.
"""
from io import BytesIO
from typing import Any, Dict, List, Optional

import calendar as _cal

from fastapi import APIRouter, Header, HTTPException, Query
from fastapi.responses import StreamingResponse

from server import (  # noqa: E402
    db,
    get_user_from_token,
    sub_admin_can_touch_company,
    _build_ot_report_rows,
)

router = APIRouter(prefix="/api/admin/ot-salary", tags=["ot-salary"])


def _num(v: Any, default: float = 0.0) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


async def _require_admin(authorization: Optional[str], company_id: Optional[str] = None) -> dict:
    admin = await get_user_from_token(authorization)
    role = admin.get("role")
    if role not in ("super_admin", "sub_admin", "company_admin"):
        raise HTTPException(status_code=403, detail="Not authorised")
    if company_id:
        if role == "sub_admin" and not sub_admin_can_touch_company(admin, company_id):
            raise HTTPException(status_code=403, detail="Firm not in your scope")
        if role == "company_admin" and admin.get("company_id") != company_id:
            raise HTTPException(status_code=403, detail="You can only view your own firm")
    return admin


def _is_policy2(company: dict) -> bool:
    pol = company.get("attendance_policy") or {}
    return (pol.get("policy_variant") or "").strip() == "policy_2"


def _per_day_pay(emp: dict, month_days: int) -> tuple:
    """Return (per_day_basic, per_day_gross) for the employee.

    Mirrors the Actual salary engine (utils/salary_run.py): the "Basic
    Salary" structure row (with its rate_type) is the wage rate — the
    "Salary 1/2/3" rows are attendance BONUS tiers and are NOT counted.
    Gross adds the monthly Actual allowances pro-rated per day.
    """
    policy = emp.get("employee_policy") or {}
    rate = _num(policy.get("salary") or emp.get("salary_monthly"))
    mode = str(policy.get("salary_mode") or "monthly").lower()
    struct = [r for r in (emp.get("salary_structure_actual") or []) if isinstance(r, dict)]
    basic_row = None
    for r in struct:
        if str(r.get("head", "")).strip().lower().startswith("basic"):
            basic_row = r
            break
    if basic_row and _num(basic_row.get("amount")) > 0:
        rate = _num(basic_row.get("amount"))
        rt = str(basic_row.get("rate_type") or "").strip().lower()
        if rt in ("monthly", "daily", "hourly"):
            mode = rt
    if mode == "daily":
        per_day_basic = rate
    elif mode == "hourly":
        per_day_basic = rate * 8.0
    else:  # monthly
        per_day_basic = rate / max(1, month_days)
    allow = sum(
        _num(r.get("amount"))
        for r in (emp.get("actual_salary_allowances") or [])
        if isinstance(r, dict)
    )
    per_day_gross = per_day_basic + (allow / max(1, month_days))
    return per_day_basic, per_day_gross


async def _compute(company_id: str, month: str, admin: dict,
                   calc_on: str, pct: float, divide: float) -> Dict[str, Any]:
    company = await db.companies.find_one(
        {"company_id": company_id},
        {"_id": 0, "company_id": 1, "name": 1, "attendance_policy": 1, "ot_salary_cfg": 1},
    )
    if company is None:
        raise HTTPException(status_code=404, detail="Company not found")
    if not _is_policy2(company):
        raise HTTPException(
            status_code=400,
            detail="OT Salary Process is only available for firms on Textile Policy 2 "
                   "(set it in Firm Master → Attendance Policy).",
        )
    try:
        y, m = [int(x) for x in month.split("-")]
        month_days = _cal.monthrange(y, m)[1]
    except (ValueError, IndexError) as exc:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM") from exc

    pol = company.get("attendance_policy") or {}
    full_day_hours = _num(pol.get("full_day_hours") or pol.get("standard_working_hours"), 8.0) or 8.0

    _, ot_rows = await _build_ot_report_rows(company_id, month, admin)

    # Aggregate OT hours per employee
    tot_by_uid: Dict[str, float] = {}
    for r in ot_rows:
        tot_by_uid[r["user_id"]] = tot_by_uid.get(r["user_id"], 0.0) + _num(r.get("ot_hours"))

    emp_docs: Dict[str, dict] = {}
    if tot_by_uid:
        async for u in db.users.find(
            {"user_id": {"$in": list(tot_by_uid.keys())}},
            {"_id": 0, "user_id": 1, "employee_code": 1, "name": 1, "designation": 1,
             "department": 1, "salary_structure_actual": 1, "salary_monthly": 1,
             "actual_salary_allowances": 1, "employee_policy": 1,
             "bank_account_number": 1, "bank_name": 1, "ifsc_code": 1},
        ):
            emp_docs[u["user_id"]] = u

    divide = divide if divide and divide > 0 else 1.0
    pct = pct if pct and pct > 0 else 100.0
    # Iter 131 (user directive) — Firm Master OT Calculation config wins
    # when set on the Textile Policy 2 firm. Iter 131b: EITHER %Basic OR
    # %Gross (mutually exclusive, enforced at save time).
    # OT hourly rate = per-day base × % ÷ full-day hours.
    fm_pct_basic = _num(pol.get("ot_pct_basic"))
    fm_pct_gross = _num(pol.get("ot_pct_gross"))
    use_fm_cfg = fm_pct_basic > 0 or fm_pct_gross > 0
    rows: List[Dict[str, Any]] = []
    for uid, ot_total in tot_by_uid.items():
        emp = emp_docs.get(uid) or {}
        per_day_basic, per_day_gross = _per_day_pay(emp, month_days)
        if use_fm_cfg:
            base_per_day = (per_day_basic * fm_pct_basic / 100.0
                            + per_day_gross * fm_pct_gross / 100.0)
            hourly = base_per_day / full_day_hours if full_day_hours else 0.0
        else:
            base_per_day = per_day_basic if calc_on == "basic" else per_day_gross
            hourly = (base_per_day * (pct / 100.0)) / full_day_hours if full_day_hours else 0.0
        recorded = round(ot_total / divide, 2)
        amount = round(recorded * hourly)
        rows.append({
            "user_id": uid,
            "employee_code": emp.get("employee_code"),
            "name": emp.get("name"),
            "designation": emp.get("designation") or emp.get("department"),
            "ot_duty_hours": round(ot_total, 2),
            "ot_hours": recorded,
            "per_day_base": round(base_per_day, 2),
            "hourly_rate": round(hourly, 2),
            "amount": amount,
            "bank_name": emp.get("bank_name"),
            "bank_account_number": emp.get("bank_account_number"),
            "ifsc_code": emp.get("ifsc_code"),
        })
    rows.sort(key=lambda r: (str(r.get("employee_code") or "zzz"), str(r.get("name") or "")))

    # Iter 129d (user directive) — deduct employee-share ESIC on the OT
    # amount, ONLY for employees who are ESIC-eligible in the firm's
    # Compliance Salary run for the same month. No compliance run or not
    # eligible → no ESIC deduction on OT.
    import math

    from routes.compliance_settings import get_standard_compliance_cfg

    cfg_std = await get_standard_compliance_cfg()
    firm_doc = await db.companies.find_one(
        {"company_id": company_id}, {"_id": 0, "statutory_overrides": 1},
    )
    overrides = (firm_doc or {}).get("statutory_overrides") or {}
    esic_rate = _num(overrides.get("esic_percent_employee"),
                     _num(cfg_std.get("esic_percent_employee"), 0.75))
    comp_run = await db.compliance_salary_runs.find(
        {"company_id": company_id, "month": month},
        {"_id": 0, "rows.user_id": 1, "rows.esic_applicable": 1, "created_at": 1},
    ).sort("created_at", -1).to_list(1)
    eligible_uids = set()
    if comp_run:
        for cr in (comp_run[0].get("rows") or []):
            if cr.get("esic_applicable"):
                eligible_uids.add(cr.get("user_id"))
    for r in rows:
        if r["user_id"] in eligible_uids and r["amount"] > 0:
            r["esic_employee"] = math.ceil(r["amount"] * esic_rate / 100.0)
        else:
            r["esic_employee"] = 0
        r["net"] = r["amount"] - r["esic_employee"]

    totals = {
        "ot_duty_hours": round(sum(r["ot_duty_hours"] for r in rows), 2),
        "ot_hours": round(sum(r["ot_hours"] for r in rows), 2),
        "amount": round(sum(r["amount"] for r in rows)),
        "esic_employee": round(sum(r["esic_employee"] for r in rows)),
        "net": round(sum(r["net"] for r in rows)),
    }
    cfg = {"calc_on": calc_on, "pct": pct, "divide": divide}
    if use_fm_cfg:
        cfg = {
            "calc_on": "firm_master", "pct": pct, "divide": divide,
            "ot_pct_basic": fm_pct_basic, "ot_pct_gross": fm_pct_gross,
        }
    # Auto-save the config so the firm remembers its OT settings.
    await db.companies.update_one({"company_id": company_id}, {"$set": {"ot_salary_cfg": cfg}})
    return {
        "company_id": company_id,
        "company_name": company.get("name"),
        "month": month,
        "month_days": month_days,
        "full_day_hours": full_day_hours,
        "cfg": cfg,
        "rows": rows,
        "totals": totals,
        "employees_count": len(rows),
    }


@router.get("/firms")
async def ot_salary_firms(authorization: Optional[str] = Header(None)):
    """Firms eligible for the OT Salary Process (Textile Policy 2 only)."""
    admin = await _require_admin(authorization)
    q: Dict[str, Any] = {"attendance_policy.policy_variant": "policy_2"}
    if admin.get("role") == "company_admin":
        q["company_id"] = admin.get("company_id")
    firms = []
    async for c in db.companies.find(
        q,
        {"_id": 0, "company_id": 1, "name": 1, "ot_salary_cfg": 1,
         "attendance_policy.ot_pct_basic": 1, "attendance_policy.ot_pct_gross": 1},
    ):
        if admin.get("role") == "sub_admin" and not sub_admin_can_touch_company(admin, c["company_id"]):
            continue
        pol = c.pop("attendance_policy", None) or {}
        c["ot_pct_basic"] = pol.get("ot_pct_basic") or 0
        c["ot_pct_gross"] = pol.get("ot_pct_gross") or 0
        firms.append(c)
    firms.sort(key=lambda c: str(c.get("name") or ""))
    return {"ok": True, "firms": firms}


@router.get("/{company_id}/{month}")
async def ot_salary_run(
    company_id: str,
    month: str,
    calc_on: str = Query("gross"),
    pct: float = Query(100.0),
    divide: float = Query(2.0),
    authorization: Optional[str] = Header(None),
):
    admin = await _require_admin(authorization, company_id)
    if calc_on not in ("basic", "gross"):
        raise HTTPException(status_code=400, detail="calc_on must be basic|gross")
    data = await _compute(company_id, month, admin, calc_on, pct, divide)
    return {"ok": True, "run": data}


@router.get("/{company_id}/{month}/bank.xlsx")
async def ot_salary_bank_xlsx(
    company_id: str,
    month: str,
    calc_on: str = Query("gross"),
    pct: float = Query(100.0),
    divide: float = Query(2.0),
    authorization: Optional[str] = Header(None),
):
    admin = await _require_admin(authorization, company_id)
    data = await _compute(company_id, month, admin, calc_on, pct, divide)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "OT Bank Sheet"
    head_fill = PatternFill("solid", fgColor="0F2E3D")
    head_font = Font(bold=True, color="FFFFFF", size=10)

    ws.append([f"{data['company_name']} — OT Salary Bank Sheet — {month}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append([
        f"OT on: {data['cfg']['calc_on'].upper()} @ {data['cfg']['pct']:g}%  ·  "
        f"OT HRS = Duty HRS ÷ {data['cfg']['divide']:g}  ·  Full day: {data['full_day_hours']:g} hrs"
    ])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    headers = ["Sr", "Emp Code", "Name", "Bank Name", "Account No", "IFSC",
               "OT Duty HRS", "OT HRS", "Rate/Hr", "Amount", "ESIC", "Net"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(data["rows"], start=1):
        ws.append([
            i, r.get("employee_code"), r.get("name"), r.get("bank_name"),
            r.get("bank_account_number"), r.get("ifsc_code"),
            r.get("ot_duty_hours"), r.get("ot_hours"), r.get("hourly_rate"),
            r.get("amount"), r.get("esic_employee"), r.get("net"),
        ])
    t = data["totals"]
    ws.append(["", "", "TOTAL", "", "", "", t["ot_duty_hours"], t["ot_hours"], "",
               t["amount"], t["esic_employee"], t["net"]])
    last = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(row=last, column=col).font = Font(bold=True)

    widths = [5, 10, 26, 18, 20, 14, 12, 10, 10, 12, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"OT_BankSheet_{data['company_name'].replace(' ', '_')}_{month}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
