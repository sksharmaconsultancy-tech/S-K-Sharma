"""Bank Transfer File Generator — salary upload files for corporate
net-banking (no bank API needed).

Flow: Salary processed → admin picks month + bank format → downloads a
ready-to-upload NEFT/salary bulk file → uploads it in the bank's corporate
portal → bank credits salaries.

Supported bank layouts: ICICI (CIB bulk), HDFC (ENet), SBI (CMP), Axis
(iConnect), Kotak (CMS) + a Generic layout. File types: xlsx / csv / txt /
xml. Data source: the same compliance-run rows the Bank Sheet report uses.
"""
import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Optional
from xml.sax.saxutils import escape

from fastapi import APIRouter, Header, HTTPException, Query
from fastapi.responses import Response

from server import db, get_user_from_token, require_role  # noqa: E402
from routes.challans import _bank_sheet_rows  # noqa: E402

router = APIRouter(prefix="/api", tags=["bank-transfer"])

BANKS: Dict[str, Dict[str, Any]] = {
    "icici": {
        "label": "ICICI Bank (Corporate CIB Bulk Upload)",
        "headers": ["PYMT_PROD_TYPE_CODE", "PYMT_MODE", "DEBIT_ACC_NO", "BNF_NAME",
                    "BENE_ACC_NO", "BENE_IFSC", "AMT", "PYMT_DATE", "REMARK"],
    },
    "hdfc": {
        "label": "HDFC Bank (ENet Salary Upload)",
        "headers": ["Transaction Type", "Beneficiary Code", "Beneficiary Account Number",
                    "Instrument Amount", "Beneficiary Name", "Customer Reference Number",
                    "Value Date", "IFSC Code", "Narration"],
    },
    "sbi": {
        "label": "State Bank of India (CMP Bulk NEFT)",
        "headers": ["Sl No", "Debit Account No", "Beneficiary Name", "Beneficiary Account No",
                    "IFSC Code", "Amount", "Payment Date", "Narration"],
    },
    "axis": {
        "label": "Axis Bank (iConnect Bulk Payment)",
        "headers": ["SRNO", "PAYMENT MODE", "DEBIT ACCOUNT NO", "BENEFICIARY NAME",
                    "BENEFICIARY ACCOUNT NO", "IFSC CODE", "AMOUNT", "PAYMENT DATE", "NARRATION"],
    },
    "kotak": {
        "label": "Kotak Mahindra Bank (CMS Bulk Upload)",
        "headers": ["Client Code", "Payment Type", "Payment Date", "Debit Account No",
                    "Beneficiary Name", "Beneficiary Account No", "IFSC Code", "Amount", "Remarks"],
    },
    "generic": {
        "label": "Generic NEFT Format (any bank)",
        "headers": ["S No", "Beneficiary Name", "Account Number", "IFSC", "Amount",
                    "Payment Mode", "Payment Date", "Narration"],
    },
}


def _row_values(bank: str, i: int, r: Dict[str, Any], debit_acc: str,
                pay_date: str, narration: str) -> List[Any]:
    name = r.get("name_as_per_bank") or r.get("name") or ""
    acc = r.get("account_no") or ""
    ifsc = (r.get("ifsc") or "").upper()
    amt = f"{float(r.get('net_salary') or 0):.2f}"
    if bank == "icici":
        return ["PAB_VENDOR", "NEFT", debit_acc, name, acc, ifsc, amt, pay_date, narration]
    if bank == "hdfc":
        return ["N", r.get("employee_code") or f"EMP{i}", acc, amt, name,
                f"SAL{pay_date.replace('/', '')}{i:04d}", pay_date, ifsc, narration]
    if bank == "sbi":
        return [i, debit_acc, name, acc, ifsc, amt, pay_date, narration]
    if bank == "axis":
        return [i, "NEFT", debit_acc, name, acc, ifsc, amt, pay_date, narration]
    if bank == "kotak":
        return ["", "NEFT", pay_date, debit_acc, name, acc, ifsc, amt, narration]
    return [i, name, acc, ifsc, amt, "NEFT", pay_date, narration]


@router.get("/admin/bank-transfer/formats")
async def bank_transfer_formats(authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    return {"banks": [{"key": k, "label": v["label"], "headers": v["headers"]}
                      for k, v in BANKS.items()],
            "file_types": ["xlsx", "csv", "txt", "xml"]}


@router.get("/admin/bank-transfer/file")
async def bank_transfer_file(
    month: str,
    bank: str = Query("generic"),
    fmt: str = Query("xlsx"),
    company_id: Optional[str] = None,
    debit_account: Optional[str] = None,
    payment_date: Optional[str] = None,   # DD/MM/YYYY
    employee_type: Optional[str] = None,
    bank_name: Optional[str] = None,      # filter: only employees of this bank
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    bank = bank.lower()
    if bank not in BANKS:
        raise HTTPException(status_code=400, detail=f"bank must be one of {list(BANKS)}")
    if fmt not in ("xlsx", "csv", "txt", "xml"):
        raise HTTPException(status_code=400, detail="fmt must be xlsx/csv/txt/xml")

    data = await _bank_sheet_rows(admin, company_id, month, employee_type, "Bank", bank_name)
    rows = [r for r in data.get("rows", []) if float(r.get("net_salary") or 0) > 0
            and (r.get("account_no") or "").strip()]
    if not rows:
        raise HTTPException(status_code=404,
                            detail="No payable bank-mode employees found for this month. "
                                   "Run the Compliance Salary Process first and ensure "
                                   "employees have bank account details.")
    firm = await db.companies.find_one({"company_id": data.get("company_id")},
                                       {"_id": 0, "name": 1})
    firm_name = (firm or {}).get("name") or "Firm"
    debit_acc = (debit_account or "").strip()
    pay_date = (payment_date or datetime.now().strftime("%d/%m/%Y")).strip()
    mon_label = datetime.strptime(month, "%Y-%m").strftime("%b %Y").upper()
    narration = f"SALARY {mon_label}"
    headers = BANKS[bank]["headers"]
    matrix = [_row_values(bank, i, r, debit_acc, pay_date, narration)
              for i, r in enumerate(rows, 1)]
    fname = f"salary-upload-{bank}-{month}"

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Salary Upload"
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2563EB")
        for m in matrix:
            ws.append(m)
        for idx, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = \
                max(14, len(h) + 4)
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})

    if fmt in ("csv", "txt"):
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(headers)
        for m in matrix:
            w.writerow(m)
        return Response(
            content=out.getvalue(),
            media_type="text/csv" if fmt == "csv" else "text/plain",
            headers={"Content-Disposition": f'attachment; filename="{fname}.{fmt}"'})

    # XML — simple salary-payments structure (banks that accept XML vary;
    # verify the exact schema with the branch before first upload).
    lines = ['<?xml version="1.0" encoding="UTF-8"?>',
             f'<SalaryPayments firm="{escape(firm_name)}" month="{month}" '
             f'paymentDate="{escape(pay_date)}" debitAccount="{escape(debit_acc)}">']
    for i, r in enumerate(rows, 1):
        lines.append(
            "  <Payment>"
            f"<SNo>{i}</SNo>"
            f"<BeneficiaryName>{escape(r.get('name_as_per_bank') or r.get('name') or '')}</BeneficiaryName>"
            f"<AccountNumber>{escape(str(r.get('account_no') or ''))}</AccountNumber>"
            f"<IFSC>{escape((r.get('ifsc') or '').upper())}</IFSC>"
            f"<Amount>{float(r.get('net_salary') or 0):.2f}</Amount>"
            f"<Mode>NEFT</Mode>"
            f"<Narration>{escape(narration)}</Narration>"
            "</Payment>")
    lines.append("</SalaryPayments>")
    return Response(
        content="\n".join(lines),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{fname}.xml"'})
