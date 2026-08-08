"""Iter 521 (user request) — PRESENT / ABSENT REPORT as per the FIRM
ATTENDANCE POLICY.

Derives a P / HD / A / WO / H status matrix from the SAME compute pipeline
that powers the Attendance Grid (``_compute_monthly_grid_data``), so every
status matches payroll's policy-based Present Days 1:1:
  • P  — full present (policy day credit 1)
  • HD — half day     (policy day credit 0.5)
  • WO — weekly off   (per firm policy / employee override, no credit)
  • H  — holiday      (holiday master, no credit)
  • A  — absent       (no punches, or worked below the policy threshold)
  • blank — future date

Endpoints:
  GET /api/admin/reports/present-absent            (JSON)
  GET /api/admin/reports/present-absent.xlsx       (Excel matrix)
  GET /api/admin/reports/present-absent.pdf        (landscape PDF)
"""
import io
import sys
from datetime import date
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException, Query
from fastapi.responses import Response

sys.path.append("/app/backend")
from server import (db, get_user_from_token, sub_admin_can_touch_company,  # noqa: E402
                    _compute_monthly_grid_data)

router = APIRouter(prefix="/api/admin/reports", tags=["present-absent"])

STATUS_ORDER = ["P", "HD", "A", "WO", "H"]


async def _authz(authorization: Optional[str], company_id: str) -> dict:
    admin = await get_user_from_token(authorization)
    if admin.get("role") not in ("super_admin", "sub_admin", "company_admin"):
        raise HTTPException(status_code=403, detail="Not authorised")
    if admin.get("role") == "sub_admin" and not sub_admin_can_touch_company(admin, company_id):
        raise HTTPException(status_code=403, detail="Firm not in your scope")
    if admin.get("role") == "company_admin" and admin.get("company_id") != company_id:
        raise HTTPException(status_code=403, detail="You can only view your own firm")
    return admin


def _day_status(cell: Dict[str, Any], iso_date: str, today_iso: str) -> str:
    if iso_date > today_iso:
        return ""  # future date
    p = float(cell.get("present") or 0)
    if p >= 1:
        return "P"
    if p >= 0.5:
        return "HD"
    if cell.get("holiday"):
        return "H"
    if cell.get("weekly_off"):
        return "WO"
    return "A"


async def _build(company_id: str, month: str, department: str = "",
                 search: str = "") -> Dict[str, Any]:
    data = await _compute_monthly_grid_data(company_id=company_id, month=month)
    day_labels: List[str] = data.get("day_labels") or []
    weekday_labels = data.get("weekday_labels") or []
    day_full_dates: List[str] = data.get("day_full_dates") or [
        f"{month}-{str(dl)[:2]}" for dl in day_labels]
    today_iso = date.today().isoformat()

    comp = await db.companies.find_one(
        {"company_id": company_id},
        {"_id": 0, "name": 1, "attendance_policy": 1})
    _pol = (comp or {}).get("attendance_policy") or {}
    _fd = float(_pol.get("full_day_hours") or _pol.get("standard_working_hours") or 8.0)
    policy_line = (f"As per Firm Attendance Policy — Full Day {_fd:g} hrs · "
                   f"Half Day {float(_pol.get('half_day_hours') or 4.0):g} hrs · "
                   f"P=Present  HD=Half Day  A=Absent  WO=Weekly Off  H=Holiday")

    term = (search or "").strip().lower()
    dep = (department or "").strip()
    rows: List[Dict[str, Any]] = []
    day_counts: Dict[str, Dict[str, int]] = {dl: {s: 0 for s in STATUS_ORDER}
                                             for dl in day_labels}
    for e in data.get("employees") or []:
        if dep and (e.get("department") or "") != dep:
            continue
        if term and term not in f"{e.get('name', '')} {e.get('employee_code', '')}".lower():
            continue
        days: Dict[str, str] = {}
        totals = {s: 0 for s in STATUS_ORDER}
        for i, dl in enumerate(day_labels):
            cell = (e.get("days") or {}).get(dl) or {}
            iso = day_full_dates[i] if i < len(day_full_dates) else f"{month}-{str(dl)[:2]}"
            st = _day_status(cell, iso, today_iso)
            days[dl] = st
            if st:
                totals[st] += 1
                day_counts[dl][st] += 1
        rows.append({
            "employee_code": e.get("employee_code"),
            "name": e.get("name"),
            "department": e.get("department"),
            "designation": e.get("designation") or e.get("position"),
            "days": days,
            "totals": totals,
            # policy Present Days exactly as payroll counts them
            "present_days": (e.get("totals") or {}).get("present_days_policy", 0),
        })
    rows.sort(key=lambda r: ((r.get("department") or ""),
                             str(r.get("employee_code") or "")))
    return {
        "company": {"company_id": company_id, "name": (comp or {}).get("name")},
        "month": month,
        "policy_line": policy_line,
        "day_labels": day_labels,
        "weekday_labels": weekday_labels,
        "employees": rows,
        "day_counts": day_counts,
        "departments": sorted({r.get("department") or "" for r in rows} - {""}),
        "grand_totals": {s: sum(r["totals"][s] for r in rows) for s in STATUS_ORDER},
    }


@router.get("/present-absent")
async def present_absent_json(company_id: str = Query(...), month: str = Query(...),
                              department: str = "", search: str = "",
                              authorization: Optional[str] = Header(None)):
    await _authz(authorization, company_id)
    return await _build(company_id, month, department, search)


_XL_FILL = {"P": "DCFCE7", "HD": "FEF9C3", "A": "FEE2E2",
            "WO": "E0F2FE", "H": "FED7AA"}


@router.get("/present-absent.xlsx")
async def present_absent_xlsx(company_id: str = Query(...), month: str = Query(...),
                              department: str = "", search: str = "",
                              authorization: Optional[str] = Header(None)):
    await _authz(authorization, company_id)
    d = await _build(company_id, month, department, search)
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Present-Absent"
    thin = Border(*[Side(style="thin", color="CBD5E1")] * 4)
    center = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1,
            value=f"PRESENT / ABSENT REPORT — {d['company'].get('name') or ''} — {month}"
            ).font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=d["policy_line"]).font = Font(
        size=9, color="0F3B5C", bold=True)
    hdr = ["S.No", "Code", "Employee Name", "Department"] + \
        [str(dl) for dl in d["day_labels"]] + \
        ["P", "HD", "A", "WO", "H", "Present Days"]
    r = 4
    for c, v in enumerate(hdr, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=True, size=8, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F3B5C")
        cell.alignment = center
        cell.border = thin
    for i, e in enumerate(d["employees"], start=1):
        r += 1
        base = [i, e.get("employee_code"), e.get("name"), e.get("department")]
        vals = base + [e["days"].get(dl, "") for dl in d["day_labels"]] + \
            [e["totals"][s] for s in STATUS_ORDER] + [e.get("present_days")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.font = Font(size=8)
            if c > 4 and c <= 4 + len(d["day_labels"]):
                cell.alignment = center
                fill = _XL_FILL.get(str(v))
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
            elif c > 4:
                cell.alignment = center
                cell.font = Font(size=8, bold=True)
    # daily present footer
    r += 1
    ws.cell(row=r, column=3, value="Daily Present (P + HD)").font = Font(bold=True, size=8)
    for j, dl in enumerate(d["day_labels"]):
        dc = d["day_counts"][dl]
        cell = ws.cell(row=r, column=5 + j, value=dc["P"] + dc["HD"])
        cell.font = Font(bold=True, size=8)
        cell.alignment = center
        cell.border = thin
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    from openpyxl.utils import get_column_letter
    for j in range(len(d["day_labels"])):
        ws.column_dimensions[get_column_letter(5 + j)].width = 4.2
    ws.freeze_panes = "E5"
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="present-absent-{month}.xlsx"'})


@router.get("/present-absent.pdf")
async def present_absent_pdf(company_id: str = Query(...), month: str = Query(...),
                             department: str = "", search: str = "",
                             authorization: Optional[str] = Header(None)):
    await _authz(authorization, company_id)
    d = await _build(company_id, month, department, search)
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=8 * mm, rightMargin=8 * mm,
                            topMargin=8 * mm, bottomMargin=8 * mm,
                            title=f"Present-Absent {month}")
    W = landscape(A4)[0] - 16 * mm
    h1 = ParagraphStyle("h1", fontSize=12, leading=15, alignment=1,
                        fontName="Helvetica-Bold")
    h2 = ParagraphStyle("h2", fontSize=8, leading=10, alignment=1,
                        textColor=rl.HexColor("#0F3B5C"))
    flow: List[Any] = [
        Paragraph(f"PRESENT / ABSENT REPORT — {d['company'].get('name') or ''} — {month}", h1),
        Paragraph(d["policy_line"], h2), Spacer(1, 3 * mm)]
    n_days = len(d["day_labels"])
    head = ["S.No.", "Code", "Employee"] + \
        [str(int(str(dl)[:2])) for dl in d["day_labels"]] + \
        ["P", "HD", "A", "WO", "H", "Days"]
    body = [head]
    _fill = {"P": "#DCFCE7", "HD": "#FEF9C3", "A": "#FEE2E2",
             "WO": "#E0F2FE", "H": "#FED7AA"}
    styles = [
        ("GRID", (0, 0), (-1, -1), 0.3, rl.HexColor("#94A3B8")),
        ("FONTSIZE", (0, 0), (-1, -1), 5.6),
        ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#0F3B5C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.2),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.2),
    ]
    for ri, e in enumerate(d["employees"], start=1):
        row = [str(ri), str(e.get("employee_code") or ""),
               (e.get("name") or "")[:22]]
        for ci, dl in enumerate(d["day_labels"]):
            st = e["days"].get(dl, "")
            row.append(st)
            f = _fill.get(st)
            if f:
                styles.append(("BACKGROUND", (3 + ci, ri), (3 + ci, ri),
                               rl.HexColor(f)))
        row += [str(e["totals"][s]) for s in STATUS_ORDER] + \
            [f"{e.get('present_days'):g}" if e.get("present_days") is not None else ""]
        body.append(row)
    # daily present footer
    foot = ["", "", "Daily Present"]
    for dl in d["day_labels"]:
        dc = d["day_counts"][dl]
        foot.append(str(dc["P"] + dc["HD"]))
    foot += [""] * 6
    body.append(foot)
    styles.append(("FONTNAME", (0, len(body) - 1), (-1, len(body) - 1),
                   "Helvetica-Bold"))
    name_w = W * 0.105
    code_w = W * 0.042
    sno_w = W * 0.028
    tot_w = W * 0.021 * 6
    day_w = (W - name_w - code_w - sno_w - tot_w) / max(n_days, 1)
    t = Table(body, colWidths=[sno_w, code_w, name_w] + [day_w] * n_days
              + [W * 0.021] * 6, repeatRows=1)
    t.setStyle(TableStyle(styles))
    flow.append(t)
    doc.build(flow)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition":
                             f'inline; filename="present-absent-{month}.pdf"'})
