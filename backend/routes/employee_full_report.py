"""Employee Full Report — consolidated per-employee reporting.

Aggregates EVERYTHING related to a single employee into one payload:
  * Profile / master data
  * Attendance (day-wise summary + raw punches) within a period
  * Leaves within the period
  * Actual Salary rows (from salary_runs.rows)
  * Compliance Salary rows (from compliance_salary_runs.rows)
  * Documents (metadata only)
  * Tickets

Endpoints:
  * GET /api/admin/employee-report              — JSON aggregate
  * GET /api/admin/employee-report/export.xlsx  — multi-sheet workbook
  * GET /api/admin/employee-report/export.pdf   — printable PDF
"""
from typing import Optional, List, Dict, Any
from datetime import datetime
import io

from fastapi import APIRouter, Header, Query, HTTPException
from fastapi.responses import Response

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
    _load_scoped_employee_any_role,
)

router = APIRouter(prefix="/api")


def _months_between(from_date: str, to_date: str) -> List[str]:
    """['YYYY-MM', ...] inclusive between two YYYY-MM-DD dates."""
    try:
        fy, fm = int(from_date[:4]), int(from_date[5:7])
        ty, tm = int(to_date[:4]), int(to_date[5:7])
    except (ValueError, IndexError):
        return []
    out: List[str] = []
    y, m = fy, fm
    while (y, m) <= (ty, tm) and len(out) < 60:
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def _hours_between(a: Optional[str], b: Optional[str]) -> Optional[float]:
    if not a or not b:
        return None
    try:
        t1 = datetime.fromisoformat(a.replace("Z", "+00:00"))
        t2 = datetime.fromisoformat(b.replace("Z", "+00:00"))
        h = (t2 - t1).total_seconds() / 3600.0
        return round(h, 2) if h >= 0 else None
    except ValueError:
        return None


async def _gather(user_id: str, from_date: str, to_date: str, admin: dict) -> Dict[str, Any]:
    emp = await _load_scoped_employee_any_role(user_id, admin)
    emp.pop("pin_hash", None)
    emp.pop("password_hash", None)
    emp.pop("face_descriptor", None)
    emp.pop("face_base64", None)

    company = await db.companies.find_one(
        {"company_id": emp.get("company_id")}, {"_id": 0, "name": 1}
    )
    company_name = (company or {}).get("name") or "—"

    # ---- Attendance -----------------------------------------------------
    punches = await db.attendance.find(
        {"user_id": user_id, "date": {"$gte": from_date, "$lte": to_date}},
        {"_id": 0, "selfie_base64": 0},
    ).sort("at", 1).to_list(3000)

    by_date: Dict[str, Dict[str, Any]] = {}
    for p in punches:
        d = p.get("date") or (p.get("at") or "")[:10]
        row = by_date.setdefault(d, {"date": d, "first_in": None, "last_out": None,
                                     "punch_count": 0, "sources": set(), "statuses": set()})
        row["punch_count"] += 1
        if p.get("source"):
            row["sources"].add(str(p["source"]))
        if p.get("status"):
            row["statuses"].add(str(p["status"]))
        if p.get("kind") == "in" and (row["first_in"] is None or (p.get("at") or "") < row["first_in"]):
            row["first_in"] = p.get("at")
        if p.get("kind") == "out" and (row["last_out"] is None or (p.get("at") or "") > row["last_out"]):
            row["last_out"] = p.get("at")

    days: List[Dict[str, Any]] = []
    total_hours = 0.0
    for d in sorted(by_date.keys()):
        row = by_date[d]
        hrs = _hours_between(row["first_in"], row["last_out"])
        if hrs:
            total_hours += hrs
        days.append({
            "date": d,
            "first_in": row["first_in"],
            "last_out": row["last_out"],
            "hours": hrs,
            "punch_count": row["punch_count"],
            "sources": sorted(row["sources"]),
            "statuses": sorted(row["statuses"]),
        })

    attendance = {
        "days": days,
        "summary": {
            "present_days": len(days),
            "total_punches": len(punches),
            "total_hours": round(total_hours, 2),
            "avg_hours": round(total_hours / len(days), 2) if days else 0,
        },
    }

    # ---- Leaves ----------------------------------------------------------
    leaves = await db.leaves.find(
        {
            "user_id": user_id,
            "from_date": {"$lte": to_date},
            "to_date": {"$gte": from_date},
        },
        {"_id": 0},
    ).sort("from_date", -1).to_list(200)

    # ---- Salary (actual) rows -------------------------------------------
    months = _months_between(from_date, to_date)
    salary_rows: List[Dict[str, Any]] = []
    if months:
        async for run in db.salary_runs.find(
            {"month": {"$in": months}, "rows.user_id": user_id},
            {"_id": 0, "run_id": 1, "month": 1, "run_type": 1, "finalized": 1, "rows": 1},
        ).sort("month", 1):
            for r in run.get("rows") or []:
                if r.get("user_id") != user_id:
                    continue
                salary_rows.append({
                    "run_id": run.get("run_id"),
                    "month": run.get("month"),
                    "run_type": run.get("run_type") or "actual",
                    "finalized": bool(run.get("finalized")),
                    "p_days": r.get("p_days"),
                    "basic": r.get("basic"),
                    "oth_allo": r.get("oth_allo"),
                    "total_gross": r.get("total_gross"),
                    "epf": r.get("epf"),
                    "esi": r.get("esi"),
                    "tds": r.get("tds"),
                    "adv": r.get("adv"),
                    "net_pay": r.get("net_pay"),
                })

    # ---- Compliance salary rows -----------------------------------------
    compliance_rows: List[Dict[str, Any]] = []
    if months:
        async for run in db.compliance_salary_runs.find(
            {"month": {"$in": months}, "rows.user_id": user_id},
            {"_id": 0, "run_id": 1, "month": 1, "rows": 1},
        ).sort("month", 1):
            for r in run.get("rows") or []:
                if r.get("user_id") != user_id:
                    continue
                compliance_rows.append({
                    "run_id": run.get("run_id"),
                    "month": run.get("month"),
                    "present_days": r.get("present_days"),
                    "basic": r.get("basic"),
                    "hra": r.get("hra"),
                    "monthly_gross": r.get("monthly_gross"),
                    "gross_paid": r.get("gross_paid"),
                    "pf_employee": r.get("pf_employee") or r.get("epf"),
                    "esi_employee": r.get("esi_employee") or r.get("esi"),
                    "net_pay": r.get("net_pay") or r.get("net_paid"),
                })

    # ---- Documents (metadata only) --------------------------------------
    documents = await db.employee_documents.find(
        {"user_id": user_id},
        {"_id": 0, "base64": 0, "file_base64": 0},
    ).sort("uploaded_at", -1).to_list(200)

    # ---- Tickets ---------------------------------------------------------
    tickets = await db.tickets.find(
        {
            "user_id": user_id,
            "created_at": {"$gte": f"{from_date}T00:00:00", "$lte": f"{to_date}T23:59:59"},
        },
        {"_id": 0, "attachments": 0},
    ).sort("created_at", -1).to_list(200)

    return {
        "employee": {
            "user_id": emp.get("user_id"),
            "name": emp.get("name"),
            "employee_code": emp.get("employee_code"),
            "father_name": emp.get("father_name"),
            "designation": emp.get("designation"),
            "department": emp.get("department"),
            "employee_type": emp.get("employee_type"),
            "phone": emp.get("phone"),
            "email": emp.get("email"),
            "doj": emp.get("doj") or emp.get("date_of_joining"),
            "uan": emp.get("uan") or emp.get("uan_number"),
            "esic_no": emp.get("esic_no") or emp.get("esi_number"),
            "pan": emp.get("pan") or emp.get("pan_number"),
            "gender": emp.get("gender"),
            "blood_group": emp.get("blood_group"),
            "marital_status": emp.get("marital_status"),
            "is_onroll": emp.get("is_onroll", True),
            "company_id": emp.get("company_id"),
            "company_name": company_name,
        },
        "period": {"from_date": from_date, "to_date": to_date},
        "attendance": attendance,
        "leaves": leaves,
        "salary_rows": salary_rows,
        "compliance_rows": compliance_rows,
        "documents": documents,
        "tickets": tickets,
    }


def _validate_dates(from_date: Optional[str], to_date: Optional[str]) -> None:
    for v in (from_date, to_date):
        if not v or len(v) != 10:
            raise HTTPException(status_code=400, detail="from_date and to_date (YYYY-MM-DD) are required")


@router.get("/admin/employee-report")
async def employee_full_report(
    user_id: str = Query(...),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    _validate_dates(from_date, to_date)
    return await _gather(user_id, from_date, to_date, admin)


@router.get("/admin/employee-report/export.xlsx")
async def employee_full_report_xlsx(
    user_id: str = Query(...),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    _validate_dates(from_date, to_date)
    data = await _gather(user_id, from_date, to_date, admin)

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    def sheet(title: str, headers: List[str], rows: List[List[Any]], first: bool = False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
        for r in rows:
            ws.append(r)
        return ws

    e = data["employee"]
    sheet("Profile", ["Field", "Value"], [
        ["Name", e.get("name")], ["Employee Code", e.get("employee_code")],
        ["Father Name", e.get("father_name")], ["Designation", e.get("designation")],
        ["Department", e.get("department")], ["Employee Type", e.get("employee_type")],
        ["Firm", e.get("company_name")], ["Phone", e.get("phone")],
        ["Email", e.get("email")], ["DOJ", e.get("doj")],
        ["UAN", e.get("uan")], ["ESIC No", e.get("esic_no")], ["PAN", e.get("pan")],
        ["Gender", e.get("gender")], ["Blood Group", e.get("blood_group")],
        ["Marital Status", e.get("marital_status")],
        ["On-roll", "Yes" if e.get("is_onroll") else "No"],
        ["Report Period", f"{from_date} to {to_date}"],
    ], first=True)

    s = data["attendance"]["summary"]
    sheet("Attendance", ["Date", "First In", "Last Out", "Hours", "Punches", "Sources", "Statuses"], [
        [d.get("date"), d.get("first_in"), d.get("last_out"), d.get("hours"),
         d.get("punch_count"), ", ".join(d.get("sources") or []), ", ".join(d.get("statuses") or [])]
        for d in data["attendance"]["days"]
    ] + [[], ["TOTALS", "", "", s["total_hours"], s["total_punches"], f"Present days: {s['present_days']}", ""]])

    sheet("Leaves", ["Type", "From", "To", "Status", "Reason"], [
        [l.get("leave_type"), l.get("from_date"), l.get("to_date"), l.get("status"), l.get("reason")]
        for l in data["leaves"]
    ])

    sheet("Salary", ["Month", "Type", "Finalized", "P Days", "Basic", "Other Allo", "Gross", "EPF", "ESI", "TDS", "Advance", "Net Pay"], [
        [r.get("month"), r.get("run_type"), "Yes" if r.get("finalized") else "No", r.get("p_days"),
         r.get("basic"), r.get("oth_allo"), r.get("total_gross"), r.get("epf"), r.get("esi"),
         r.get("tds"), r.get("adv"), r.get("net_pay")]
        for r in data["salary_rows"]
    ])

    sheet("Compliance", ["Month", "Present Days", "Basic", "HRA", "Monthly Gross", "Gross Paid", "PF", "ESI", "Net Pay"], [
        [r.get("month"), r.get("present_days"), r.get("basic"), r.get("hra"), r.get("monthly_gross"),
         r.get("gross_paid"), r.get("pf_employee"), r.get("esi_employee"), r.get("net_pay")]
        for r in data["compliance_rows"]
    ])

    sheet("Documents", ["Category", "Name", "Uploaded At", "Status"], [
        [d.get("category"), d.get("name") or d.get("filename"), d.get("uploaded_at"), d.get("status")]
        for d in data["documents"]
    ])

    sheet("Tickets", ["Subject", "Category", "Status", "Created At"], [
        [t.get("subject") or t.get("title"), t.get("category"), t.get("status"), t.get("created_at")]
        for t in data["tickets"]
    ])

    buf = io.BytesIO()
    wb.save(buf)
    fname = f"employee-report-{e.get('employee_code') or user_id}-{from_date}-to-{to_date}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/admin/employee-report/export.pdf")
async def employee_full_report_pdf(
    user_id: str = Query(...),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    _validate_dates(from_date, to_date)
    data = await _gather(user_id, from_date, to_date, admin)

    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=14, spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, spaceBefore=10, spaceAfter=4)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8)

    def tbl(headers: List[str], rows: List[List[Any]], widths=None):
        body = [headers] + [[("" if v is None else str(v)) for v in r] for r in rows]
        t = Table(body, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#c9d4e0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#f2f6fa")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return t

    e = data["employee"]
    story = [
        Paragraph(f"Employee Full Report — {e.get('name') or ''}", h1),
        Paragraph(
            f"Firm: {e.get('company_name')} &nbsp;|&nbsp; Code: {e.get('employee_code') or '—'} "
            f"&nbsp;|&nbsp; Period: {from_date} to {to_date}", small),
        Spacer(1, 6),
        Paragraph("Profile", h2),
        tbl(["Field", "Value", "Field", "Value"], [
            ["Father Name", e.get("father_name"), "Designation", e.get("designation")],
            ["Department", e.get("department"), "Employee Type", e.get("employee_type")],
            ["Phone", e.get("phone"), "DOJ", e.get("doj")],
            ["UAN", e.get("uan"), "ESIC No", e.get("esic_no")],
            ["PAN", e.get("pan"), "On-roll", "Yes" if e.get("is_onroll") else "No"],
        ]),
    ]

    s = data["attendance"]["summary"]
    story.append(Paragraph("Attendance Summary", h2))
    story.append(tbl(
        ["Present Days", "Total Punches", "Total Hours", "Avg Hours/Day"],
        [[s["present_days"], s["total_punches"], s["total_hours"], s["avg_hours"]]],
    ))
    if data["attendance"]["days"]:
        story.append(Paragraph("Attendance (Day-wise)", h2))
        story.append(tbl(
            ["Date", "First In", "Last Out", "Hours", "Punches"],
            [[d.get("date"), (d.get("first_in") or "")[11:16], (d.get("last_out") or "")[11:16],
              d.get("hours"), d.get("punch_count")] for d in data["attendance"]["days"][:200]],
        ))

    if data["leaves"]:
        story.append(Paragraph("Leaves", h2))
        story.append(tbl(["Type", "From", "To", "Status", "Reason"], [
            [l.get("leave_type"), l.get("from_date"), l.get("to_date"), l.get("status"),
             (l.get("reason") or "")[:60]] for l in data["leaves"]]))

    if data["salary_rows"]:
        story.append(Paragraph("Actual Salary", h2))
        story.append(tbl(
            ["Month", "P Days", "Basic", "Gross", "EPF", "ESI", "TDS", "Adv", "Net Pay"],
            [[r.get("month"), r.get("p_days"), r.get("basic"), r.get("total_gross"), r.get("epf"),
              r.get("esi"), r.get("tds"), r.get("adv"), r.get("net_pay")] for r in data["salary_rows"]]))

    if data["compliance_rows"]:
        story.append(Paragraph("Compliance Salary", h2))
        story.append(tbl(
            ["Month", "P Days", "Basic", "HRA", "Monthly Gross", "Gross Paid", "Net Pay"],
            [[r.get("month"), r.get("present_days"), r.get("basic"), r.get("hra"),
              r.get("monthly_gross"), r.get("gross_paid"), r.get("net_pay")] for r in data["compliance_rows"]]))

    if data["documents"]:
        story.append(Paragraph("Documents", h2))
        story.append(tbl(["Category", "Name", "Uploaded At"], [
            [d.get("category"), d.get("name") or d.get("filename"), (d.get("uploaded_at") or "")[:10]]
            for d in data["documents"]]))

    if data["tickets"]:
        story.append(Paragraph("Tickets", h2))
        story.append(tbl(["Subject", "Status", "Created"], [
            [(t.get("subject") or t.get("title") or "")[:60], t.get("status"), (t.get("created_at") or "")[:10]]
            for t in data["tickets"]]))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    doc.build(story)
    fname = f"employee-report-{e.get('employee_code') or user_id}-{from_date}-to-{to_date}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
