"""Iter 310 — Employee Master Detail Slip (Phase 1).

Professional A4 printable slip with the employee's complete master
information + FYTD attendance & leave summary, profile-completion %,
QR code linking back to the slip, and PDF / Excel / Email exports.

Fields not yet captured on the Employee Master (Mother Name, Grade,
Cost Centre, Confirmation Date, Nominee, Education, Experience, Company
Assets) safely render "—" per the user's Phase-1 directive.

Endpoints (super_admin / sub_admin / company_admin):
  GET  /api/admin/employee-detail-slip/employees?company_id=&search=
  GET  /api/admin/employee-detail-slip/{user_id}
  GET  /api/admin/employee-detail-slip/{user_id}/slip.pdf
  GET  /api/admin/employee-detail-slip/{user_id}/slip.xlsx
  POST /api/admin/employee-detail-slip/{user_id}/email   {"to": "..."}
"""
import base64
import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query, Request
from fastapi.responses import StreamingResponse

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
)

router = APIRouter(prefix="/api/admin/employee-detail-slip",
                   tags=["employee-detail-slip"])

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _fy_start(today: date) -> date:
    """April-1 of the Financial Year containing ``today``."""
    return date(today.year if today.month >= 4 else today.year - 1, 4, 1)


def _s(v: Any) -> Optional[str]:
    s = str(v).strip() if v is not None else ""
    return s or None


def _dmy(v: Any) -> Optional[str]:
    s = _s(v)
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        return s


async def _scoped_admin(authorization: Optional[str]) -> Dict[str, Any]:
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    return admin


def _company_scope(admin: Dict[str, Any], company_id: Optional[str]) -> Optional[str]:
    if admin["role"] == "company_admin":
        return admin.get("company_id")
    return company_id or None


# Profile-completion checklist (Phase 1 — existing master fields only).
_COMPLETION_FIELDS = [
    "name", "father_name", "dob", "gender", "phone", "email", "address",
    "employee_code", "designation", "department", "employee_type", "doj",
    "uan_no", "pf_no", "esi_ip_no", "pan_no", "aadhaar_no",
    "bank_name", "bank_account", "bank_ifsc",
]


def _profile_completion(u: Dict[str, Any]) -> int:
    filled = sum(1 for k in _COMPLETION_FIELDS if _s(u.get(k)))
    # salary counts as one slot (any of the salary fields present)
    if _s(u.get("salary_monthly")) or (u.get("salary_structure_actual") or []):
        filled += 1
    return round(filled * 100.0 / (len(_COMPLETION_FIELDS) + 1))


def _salary_summary(u: Dict[str, Any]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    if u.get("salary_monthly"):
        out.append({"label": "Monthly Salary (Actual)",
                    "value": f"₹{float(u['salary_monthly']):,.2f}"})
    if u.get("compliance_gross"):
        out.append({"label": "Compliance Gross",
                    "value": f"₹{float(u['compliance_gross']):,.2f}"})
    if u.get("compliance_basic"):
        out.append({"label": "Compliance Basic",
                    "value": f"₹{float(u['compliance_basic']):,.2f}"})
    if u.get("pf_basic"):
        out.append({"label": "PF Basic Salary",
                    "value": f"₹{float(u['pf_basic']):,.2f}"})
    for r in (u.get("salary_structure_actual") or []):
        if isinstance(r, dict) and r.get("head"):
            amt = float(r.get("amount") or 0)
            rt = _s(r.get("rate_type")) or ""
            out.append({"label": str(r["head"]),
                        "value": f"₹{amt:,.2f}" + (f" / {rt}" if rt else "")})
    return out


async def _slip_data(admin: Dict[str, Any], user_id: str) -> Dict[str, Any]:
    u = await db.users.find_one(
        {"user_id": user_id, "role": "employee"},
        {"_id": 0, "pin_hash": 0, "password_hash": 0,
         "temp_pin_plaintext": 0, "temp_password_plaintext": 0},
    )
    if not u:
        raise HTTPException(status_code=404, detail="Employee not found")
    if admin["role"] == "company_admin" and \
            u.get("company_id") != admin.get("company_id"):
        raise HTTPException(status_code=403, detail="Not your firm's employee")

    firm = await db.companies.find_one(
        {"company_id": u.get("company_id")},
        {"_id": 0, "name": 1, "company_id": 1, "address": 1}) or {}
    fm = await db.firm_masters.find_one(
        {"company_id": u.get("company_id")},
        {"_id": 0, "address": 1, "pf_code": 1, "esi_code": 1}) or {}

    # ---- FYTD attendance summary (distinct punch dates) ----
    today = date.today()
    fy0 = _fy_start(today)
    fy_label = f"FY {fy0.year}-{str(fy0.year + 1)[2:]}"
    punch_dates: set = set()
    async for a in db.attendance.find(
        {"user_id": user_id,
         "date": {"$gte": fy0.isoformat(), "$lte": today.isoformat()}},
        {"_id": 0, "date": 1},
    ):
        if a.get("date"):
            punch_dates.add(str(a["date"])[:10])
    by_month: Dict[str, int] = {}
    for d in punch_dates:
        by_month[d[:7]] = by_month.get(d[:7], 0) + 1
    attendance_fytd = {
        "fy_label": fy_label,
        "from": fy0.isoformat(),
        "to": today.isoformat(),
        "days_present": len(punch_dates),
        "by_month": dict(sorted(by_month.items())),
        "first_punch_date": min(punch_dates) if punch_dates else None,
        "last_punch_date": max(punch_dates) if punch_dates else None,
    }

    # ---- FYTD leaves (approved) ----
    leaves_by_type: Dict[str, float] = {}
    async for lv in db.leaves.find(
        {"user_id": user_id, "status": "approved",
         "from_date": {"$gte": fy0.isoformat()}},
        {"_id": 0, "leave_type": 1, "from_date": 1, "to_date": 1},
    ):
        try:
            f = datetime.strptime(str(lv["from_date"])[:10], "%Y-%m-%d").date()
            t = datetime.strptime(str(lv.get("to_date") or lv["from_date"])[:10],
                                  "%Y-%m-%d").date()
            days = max(1, (t - f).days + 1)
        except (ValueError, KeyError):
            days = 1
        k = str(lv.get("leave_type") or "other").title()
        leaves_by_type[k] = leaves_by_type.get(k, 0) + days
    leaves_fytd = [{"type": k, "days": v}
                   for k, v in sorted(leaves_by_type.items())]

    # ---- sections (backend-driven layout; None → "—" in UI/exports) ----
    def F(label: str, value: Any) -> Dict[str, Any]:
        return {"label": label, "value": value}

    sections: List[Dict[str, Any]] = [
        {"title": "Personal Information", "fields": [
            F("Full Name", _s(u.get("name"))),
            F("Father / Spouse Name", _s(u.get("father_name"))),
            F("Mother Name", None),
            F("Date of Birth", _dmy(u.get("dob"))),
            F("Gender", _s(u.get("gender"))),
            F("Phone", _s(u.get("phone"))),
            F("Email", _s(u.get("email"))),
            F("Permanent Address", _s(u.get("address"))),
            F("Present Address", _s(u.get("present_address"))),
        ]},
        {"title": "Employment Details", "fields": [
            F("Employee Code", _s(u.get("employee_code"))),
            F("Biometric Code", _s(u.get("bio_code"))),
            F("Designation", _s(u.get("designation"))),
            F("Department", _s(u.get("department"))),
            F("Branch", _s(u.get("branch_name"))),
            F("Employee Group", _s(u.get("employee_type"))),
            F("Grade", None),
            F("Cost Centre", None),
            F("Date of Joining", _dmy(u.get("doj"))),
            F("Confirmation Date", None),
            F("Employment Status", _s(u.get("employment_status")) or "Active"),
            F("On-roll / Off-roll",
              "Off-roll" if u.get("is_onroll") is False else "On-roll"),
            F("Contractor", _s(u.get("contractor_name"))),
            F("Shift", _s(u.get("shift_name"))),
        ]},
        {"title": "Statutory / KYC", "fields": [
            F("UAN No.", _s(u.get("uan_no"))),
            F("PF No.", _s(u.get("pf_no"))),
            F("ESI IP No.", _s(u.get("esi_ip_no"))),
            F("PAN No.", _s(u.get("pan_no"))),
            F("Aadhaar No.", _s(u.get("aadhaar_no"))),
        ]},
        {"title": "Bank Details", "fields": [
            F("Bank Name", _s(u.get("bank_name"))),
            F("Account No.", _s(u.get("bank_account"))),
            F("IFSC", _s(u.get("bank_ifsc"))),
            F("Account Holder", _s(u.get("bank_account_name"))),
        ]},
        {"title": "Salary Information", "fields": (
            [{"label": r["label"], "value": r["value"]}
             for r in _salary_summary(u)] or [F("Salary Structure", None)]
        )},
        {"title": "Other (Phase 2)", "fields": [
            F("Nominee", None),
            F("Education", None),
            F("Experience", None),
            F("Company Assets", None),
        ]},
    ]

    return {
        "user_id": user_id,
        "employee": {
            "name": u.get("name"),
            "employee_code": u.get("employee_code"),
            "designation": u.get("designation"),
            "company_id": u.get("company_id"),
        },
        "firm": {
            "name": firm.get("name"),
            "address": fm.get("address") or firm.get("address"),
            "pf_code": fm.get("pf_code"),
            "esi_code": fm.get("esi_code"),
        },
        "sections": sections,
        "attendance_fytd": attendance_fytd,
        "leaves_fytd": leaves_fytd,
        "profile_completion": _profile_completion(u),
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


# ---------------------------------------------------------------------------
# endpoints
# ---------------------------------------------------------------------------
@router.get("/employees")
async def slip_employee_list(
    company_id: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Ordered employee list for the slip's search + prev/next navigation."""
    admin = await _scoped_admin(authorization)
    cid = _company_scope(admin, company_id)
    q: Dict[str, Any] = {"role": "employee", "disabled": {"$ne": True}}
    if cid:
        q["company_id"] = cid
    if search and search.strip():
        s = search.strip()
        q["$or"] = [
            {"name": {"$regex": s, "$options": "i"}},
            {"employee_code": {"$regex": s, "$options": "i"}},
        ]
    emps = await db.users.find(
        q, {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1,
            "designation": 1, "company_id": 1},
    ).to_list(3000)

    def _key(e):
        code = str(e.get("employee_code") or "")
        return (0, int(code)) if code.isdigit() else (1, (e.get("name") or "").lower())
    emps.sort(key=_key)
    return {"employees": emps, "count": len(emps)}


@router.get("/{user_id}")
async def slip_detail(user_id: str,
                      authorization: Optional[str] = Header(None)):
    admin = await _scoped_admin(authorization)
    return await _slip_data(admin, user_id)


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def _qr_image(data: str):
    """QR PNG bytes → ReportLab ImageReader."""
    import qrcode
    from reportlab.lib.utils import ImageReader
    img = qrcode.make(data, box_size=6, border=1)
    b = io.BytesIO()
    img.save(b, format="PNG")
    b.seek(0)
    return ImageReader(b)


def _slip_pdf(data: Dict[str, Any], qr_url: str) -> bytes:
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        BaseDocTemplate, Frame, PageTemplate, Paragraph, Spacer,
        Table, TableStyle,
    )

    W, H = A4
    BRAND = rl.HexColor("#0F3B5C")
    BAND = rl.HexColor("#EAF1F7")
    firm = data.get("firm") or {}
    emp = data.get("employee") or {}
    comp_pct = int(data.get("profile_completion") or 0)

    def _header(c, d):
        c.saveState()
        c.setFillColor(BRAND)
        c.rect(0, H - 24 * mm, W, 24 * mm, stroke=0, fill=1)
        c.setFillColor(rl.white)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(12 * mm, H - 10 * mm, (firm.get("name") or "").upper())
        c.setFont("Helvetica", 7.5)
        if firm.get("address"):
            c.drawString(12 * mm, H - 14.5 * mm, str(firm["address"])[:120])
        codes = "   ·   ".join(x for x in [
            f"PF Code: {firm.get('pf_code')}" if firm.get("pf_code") else "",
            f"ESI Code: {firm.get('esi_code')}" if firm.get("esi_code") else "",
        ] if x)
        if codes:
            c.drawString(12 * mm, H - 18.5 * mm, codes)
        c.setFont("Helvetica-Bold", 11)
        c.drawRightString(W - 34 * mm, H - 10 * mm, "EMPLOYEE MASTER DETAIL SLIP")
        c.setFont("Helvetica", 8)
        c.drawRightString(W - 34 * mm, H - 14.5 * mm,
                          f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
        c.drawRightString(W - 34 * mm, H - 18.5 * mm,
                          f"Profile Completion: {comp_pct}%")
        # QR — top right
        try:
            c.drawImage(_qr_image(qr_url), W - 30 * mm, H - 22 * mm,
                        width=20 * mm, height=20 * mm, mask="auto")
        except Exception:  # noqa: BLE001
            pass
        c.setFillColor(rl.HexColor("#666666"))
        c.setFont("Helvetica", 6.5)
        c.drawCentredString(W / 2, 6 * mm,
                            "System generated Employee Master Detail Slip — "
                            "S.K. Sharma & Co. Payroll Portal")
        c.restoreState()

    buf = io.BytesIO()
    doc = BaseDocTemplate(
        buf, pagesize=A4, leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=30 * mm, bottomMargin=12 * mm,
        title=f"Employee Detail Slip — {emp.get('name')}",
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, W - 24 * mm,
                  H - doc.topMargin - doc.bottomMargin, id="f")
    doc.addPageTemplates([PageTemplate(id="pg", frames=[frame], onPage=_header)])

    h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=12,
                        textColor=BRAND, leading=15)
    sub = ParagraphStyle("sub", fontName="Helvetica", fontSize=9,
                         textColor=rl.HexColor("#555555"), leading=12)
    sec_t = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=8.5,
                           textColor=rl.white, leading=11)
    lbl = ParagraphStyle("lbl", fontName="Helvetica", fontSize=7.8,
                         textColor=rl.HexColor("#555555"), leading=10)
    val = ParagraphStyle("val", fontName="Helvetica-Bold", fontSize=8.2,
                         leading=10)

    story: List[Any] = [
        Paragraph(f"{(emp.get('name') or '').upper()}"
                  f"  ·  Code: {emp.get('employee_code') or '—'}", h1),
        Paragraph(emp.get("designation") or "—", sub),
        Spacer(1, 3 * mm),
    ]

    def section_table(title: str, fields: List[Dict[str, Any]]) -> Table:
        # 2 field-pairs per row → 4 columns
        rows: List[List[Any]] = [[Paragraph(title, sec_t), "", "", ""]]
        pair: List[Any] = []
        for f in fields:
            v = f.get("value")
            pair += [Paragraph(str(f.get("label") or ""), lbl),
                     Paragraph(str(v) if v not in (None, "") else "—", val)]
            if len(pair) == 4:
                rows.append(pair)
                pair = []
        if pair:
            rows.append(pair + [""] * (4 - len(pair)))
        cw = [(W - 24 * mm) * p for p in (0.22, 0.28, 0.22, 0.28)]
        t = Table(rows, colWidths=cw)
        t.setStyle(TableStyle([
            ("SPAN", (0, 0), (-1, 0)),
            ("BACKGROUND", (0, 0), (-1, 0), BRAND),
            ("BOX", (0, 0), (-1, -1), 0.6, rl.HexColor("#B9C4CE")),
            ("INNERGRID", (0, 1), (-1, -1), 0.3, rl.HexColor("#DDE4EA")),
            ("BACKGROUND", (0, 1), (0, -1), BAND),
            ("BACKGROUND", (2, 1), (2, -1), BAND),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    for s in data.get("sections") or []:
        story.append(section_table(s["title"], s["fields"]))
        story.append(Spacer(1, 3 * mm))

    # Attendance + leaves FYTD
    att = data.get("attendance_fytd") or {}
    att_fields = [
        {"label": "Financial Year", "value": att.get("fy_label")},
        {"label": "Days Present (FYTD)", "value": att.get("days_present")},
        {"label": "First Punch", "value": _dmy(att.get("first_punch_date"))},
        {"label": "Last Punch", "value": _dmy(att.get("last_punch_date"))},
    ]
    for m, n in (att.get("by_month") or {}).items():
        att_fields.append({"label": f"Month {m}", "value": f"{n} day(s)"})
    story.append(section_table(
        f"Attendance Summary — {att.get('fy_label') or 'FYTD'}", att_fields))
    story.append(Spacer(1, 3 * mm))
    lv_fields = ([{"label": f"{x['type']} Leave", "value": f"{x['days']:g} day(s)"}
                  for x in (data.get("leaves_fytd") or [])]
                 or [{"label": "Approved Leaves (FYTD)", "value": None}])
    story.append(section_table(
        f"Leave Information — {att.get('fy_label') or 'FYTD'}", lv_fields))
    story.append(Spacer(1, 8 * mm))

    foot = Table([
        [Paragraph("Employee Signature", lbl),
         Paragraph("Checked By", lbl),
         Paragraph("Authorised Signatory", lbl)],
    ], colWidths=[(W - 24 * mm) / 3.0] * 3)
    foot.setStyle(TableStyle([
        ("TOPPADDING", (0, 0), (-1, -1), 26),
        ("LINEABOVE", (0, 0), (0, 0), 0.5, rl.HexColor("#999999")),
        ("LINEABOVE", (1, 0), (1, 0), 0.5, rl.HexColor("#999999")),
        ("LINEABOVE", (2, 0), (2, 0), 0.5, rl.HexColor("#999999")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(foot)
    doc.build(story)
    return buf.getvalue()


def _slip_xlsx(data: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Detail Slip"
    brand = PatternFill("solid", fgColor="0F3B5C")
    band = PatternFill("solid", fgColor="EAF1F7")
    white_b = Font(bold=True, color="FFFFFF", size=11)
    emp = data.get("employee") or {}
    firm = data.get("firm") or {}

    ws.merge_cells("A1:D1")
    ws["A1"] = f"{(firm.get('name') or '').upper()} — EMPLOYEE MASTER DETAIL SLIP"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = brand
    ws.merge_cells("A2:D2")
    ws["A2"] = (f"{emp.get('name') or ''}  ·  Code: {emp.get('employee_code') or '—'}"
                f"  ·  Profile Completion: {data.get('profile_completion')}%")
    ws["A2"].font = Font(bold=True, size=11)
    r = 4
    for s in data.get("sections") or []:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=s["title"])
        c.font = white_b
        c.fill = brand
        r += 1
        for f in s["fields"]:
            ws.cell(row=r, column=1, value=f["label"]).fill = band
            v = f.get("value")
            ws.cell(row=r, column=2,
                    value=v if v not in (None, "") else "—")
            r += 1
        r += 1
    att = data.get("attendance_fytd") or {}
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1,
                value=f"Attendance Summary — {att.get('fy_label')}")
    c.font = white_b
    c.fill = brand
    r += 1
    for k, v in [("Days Present (FYTD)", att.get("days_present")),
                 ("First Punch", att.get("first_punch_date")),
                 ("Last Punch", att.get("last_punch_date"))]:
        ws.cell(row=r, column=1, value=k).fill = band
        ws.cell(row=r, column=2, value=v if v is not None else "—")
        r += 1
    for m, n in (att.get("by_month") or {}).items():
        ws.cell(row=r, column=1, value=f"Month {m}").fill = band
        ws.cell(row=r, column=2, value=f"{n} day(s)")
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1, value=f"Leave Information — {att.get('fy_label')}")
    c.font = white_b
    c.fill = brand
    r += 1
    lvs = data.get("leaves_fytd") or []
    if not lvs:
        ws.cell(row=r, column=1, value="Approved Leaves (FYTD)").fill = band
        ws.cell(row=r, column=2, value="—")
        r += 1
    for x in lvs:
        ws.cell(row=r, column=1, value=f"{x['type']} Leave").fill = band
        ws.cell(row=r, column=2, value=f"{x['days']:g} day(s)")
        r += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 30
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    b = io.BytesIO()
    wb.save(b)
    return b.getvalue()


def _qr_link(request: Request, user_id: str) -> str:
    import os
    base = (request.headers.get("origin")
            or os.environ.get("PUBLIC_APP_URL", "")).rstrip("/")
    if base:
        return f"{base}/employee-detail-slip?user_id={user_id}"
    return f"EMPLOYEE-DETAIL-SLIP:{user_id}"


def _fname(emp: Dict[str, Any], ext: str) -> str:
    code = str(emp.get("employee_code") or "emp").replace(" ", "")
    return f"Employee_Detail_Slip_{code}.{ext}"


@router.get("/{user_id}/slip.pdf")
async def slip_pdf(user_id: str, request: Request,
                   authorization: Optional[str] = Header(None),
                   token: Optional[str] = Query(None)):
    admin = await _scoped_admin(authorization or (f"Bearer {token}" if token else None))
    data = await _slip_data(admin, user_id)
    pdf = _slip_pdf(data, _qr_link(request, user_id))
    return StreamingResponse(
        io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition":
                 f'inline; filename="{_fname(data["employee"], "pdf")}"'})


@router.get("/{user_id}/slip.xlsx")
async def slip_xlsx(user_id: str,
                    authorization: Optional[str] = Header(None)):
    admin = await _scoped_admin(authorization)
    data = await _slip_data(admin, user_id)
    return StreamingResponse(
        io.BytesIO(_slip_xlsx(data)),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition":
                 f'attachment; filename="{_fname(data["employee"], "xlsx")}"'})


@router.post("/{user_id}/email")
async def slip_email(user_id: str, request: Request,
                     payload: Dict[str, Any] = Body(...),
                     authorization: Optional[str] = Header(None)):
    admin = await _scoped_admin(authorization)
    to = str(payload.get("to") or "").strip()
    if not to or "@" not in to:
        raise HTTPException(status_code=400, detail="Valid recipient email required")
    data = await _slip_data(admin, user_id)
    pdf = _slip_pdf(data, _qr_link(request, user_id))
    from utils.iter60_features import _send_email_with_attachment
    emp = data["employee"]
    result = await _send_email_with_attachment(
        [to],
        f"Employee Master Detail Slip — {emp.get('name')} "
        f"({emp.get('employee_code') or '—'})",
        f"Please find attached the Employee Master Detail Slip for "
        f"{emp.get('name')} (Code {emp.get('employee_code') or '—'}).\n\n"
        f"— {data.get('firm', {}).get('name') or 'S.K. Sharma & Co.'}",
        attachments=[{
            "filename": _fname(emp, "pdf"),
            "content": base64.b64encode(pdf).decode(),
        }],
    )
    if not result.get("delivered"):
        raise HTTPException(status_code=502,
                            detail=f"Email failed: {result.get('error')}")
    return {"ok": True, "to": to, "email_id": result.get("email_id")}
