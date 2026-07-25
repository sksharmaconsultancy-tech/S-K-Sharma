"""Iter 292 (user request) — Employee Reports hub helpers.

  * GET /api/admin/annual-salary-statement       — JSON: employee's month-by-
        month compliance salary for a financial year (Apr–Mar).
  * GET /api/admin/annual-salary-statement.xlsx  — same as Excel.

Pay Slip / Salary Register / letters reuse existing endpoints — the hub
screen simply links to them.
"""
import io
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Header, HTTPException

router = APIRouter(prefix="/api/admin", tags=["employee-reports-hub"])

MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep",
               "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

FIELDS = [
    ("present_days", "Present Days"), ("basic", "Basic"), ("hra", "HRA"),
    ("conveyance", "Conveyance"), ("medical", "Medical"), ("special", "Special"),
    ("others", "Others"), ("ot_pay", "OT Pay"), ("gross_paid", "Gross Paid"),
    ("pf_employee", "PF (EE)"), ("esic_employee", "ESIC (EE)"),
    ("other_deduction", "Other Ded."), ("net", "Net Pay"),
]


async def _auth(authorization: Optional[str], company_id: Optional[str]):
    from server import get_user_from_token, require_role, sub_admin_can_touch_company
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    if admin.get("role") == "company_admin":
        company_id = admin.get("company_id")
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    if admin.get("role") == "sub_admin" and not sub_admin_can_touch_company(admin, company_id):
        raise HTTPException(status_code=403, detail="Firm not in your scope")
    return admin, company_id


def _fy_months(fy_start_year: int) -> List[str]:
    """FY 2025 → ['2025-04' … '2026-03']."""
    out = []
    for i in range(12):
        y = fy_start_year + (1 if i >= 9 else 0)
        m = 4 + i if i < 9 else i - 8
        out.append(f"{y:04d}-{m:02d}")
    return out


async def _collect(company_id: str, user_id: str, fy: int) -> Dict[str, Any]:
    from server import db
    emp = await db.users.find_one(
        {"user_id": user_id, "company_id": company_id, "role": "employee"},
        {"_id": 0, "name": 1, "employee_code": 1, "department": 1,
         "designation": 1, "position": 1, "uan_no": 1, "esi_ip_no": 1, "doj": 1})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found in this firm")
    company = await db.companies.find_one(
        {"company_id": company_id}, {"_id": 0, "name": 1})
    months = _fy_months(fy)
    rows: List[Dict[str, Any]] = []
    for i, m in enumerate(months):
        run = await db.compliance_salary_runs.find_one(
            {"company_id": company_id, "month": m, "finalized": True},
            {"_id": 0, "rows": 1})
        if not run:
            run = await db.compliance_salary_runs.find_one(
                {"company_id": company_id, "month": m}, {"_id": 0, "rows": 1})
        r = next((x for x in (run or {}).get("rows", [])
                  if x.get("user_id") == user_id), None)
        row = {"month": m, "label": f"{MONTH_NAMES[i]} {m[:4]}"}
        for key, _lab in FIELDS:
            row[key] = round(float((r or {}).get(key) or 0), 2)
        row["has_data"] = bool(r)
        rows.append(row)
    totals = {key: round(sum(r[key] for r in rows), 2) for key, _ in FIELDS}
    return {
        "company": {"company_id": company_id, "name": (company or {}).get("name")},
        "employee": emp, "fy": f"{fy}-{(fy + 1) % 100:02d}",
        "fields": FIELDS, "rows": rows, "totals": totals,
    }


@router.get("/annual-salary-statement")
async def annual_salary_statement(
    user_id: str, fy: int,
    company_id: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    _, cid = await _auth(authorization, company_id)
    return await _collect(cid, user_id, fy)


@router.get("/annual-salary-statement.xlsx")
async def annual_salary_statement_xlsx(
    user_id: str, fy: int,
    company_id: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    _, cid = await _auth(authorization, company_id)
    data = await _collect(cid, user_id, fy)
    wb = Workbook()
    ws = wb.active
    ws.title = "Annual Statement"
    emp = data["employee"]
    ws.append([f"{data['company']['name']} — Annual Salary Statement FY {data['fy']}"])
    ws.append([f"Employee: {emp.get('employee_code')} — {emp.get('name')}   "
               f"Dept: {emp.get('department') or '-'}   "
               f"Desig: {emp.get('designation') or emp.get('position') or '-'}   "
               f"UAN: {emp.get('uan_no') or '-'}"])
    ws.append([])
    head = ["Month"] + [lab for _k, lab in data["fields"]]
    ws.append(head)
    thin = Border(*[Side(style="thin", color="CBD5E1")] * 4)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="1E3A8A")
        c.border = thin
        c.alignment = Alignment(horizontal="center")
    for r in data["rows"]:
        ws.append([r["label"]] + [r[k] for k, _ in data["fields"]])
        for c in ws[ws.max_row]:
            c.border = thin
            c.font = Font(size=9)
    ws.append(["TOTAL"] + [data["totals"][k] for k, _ in data["fields"]])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill("solid", fgColor="DBEAFE")
        c.border = thin
    ws.column_dimensions["A"].width = 12
    buf = io.BytesIO()
    wb.save(buf)
    code = emp.get("employee_code") or user_id
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="annual-salary-{code}-FY{fy}.xlsx"'})
