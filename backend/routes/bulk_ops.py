"""Iter 202 — Bulk Operations.

  * Bulk Attendance Upload  — status grid (P/A/HD per day) OR In/Out times,
    Excel template download → preview → apply (creates approved punches).
  * Bulk Salary Revision    — select employees + % / flat amount, or Excel
    upload with new amounts. Applies to Actual and/or Compliance salary and
    logs every change in ``salary_revisions``.
  * Bulk Transfer           — move employees between firms.
  * Bulk Resignation        — set exit date for many employees at once.
  * Bulk Shift Assignment   — assign a Shift Master to many employees.

Every operation is logged in ``bulk_ops_log`` for the History tab.
"""
import base64
import calendar
import io
import uuid
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Header, HTTPException, Query

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
    now_iso,
    sub_admin_can_touch_company,
)

router = APIRouter(prefix="/api/admin/bulk-ops", tags=["bulk-ops"])

STATUS_CODES = {"P", "A", "HD", "WO", "H", "L", "OFF", ""}
PUNCH_STATUSES = {"P", "HD"}  # only these create punches


async def _auth(authorization: Optional[str], company_id: str = "", write: bool = False):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin"])
    if company_id and admin.get("role") == "sub_admin" \
            and not sub_admin_can_touch_company(admin, company_id):
        raise HTTPException(status_code=403, detail="Firm not in your scope")
    return admin


def _num(v: Any, default: float = 0.0) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _actual_basic(u: dict) -> float:
    for r in (u.get("salary_structure_actual") or []):
        if isinstance(r, dict) and str(r.get("head", "")).lower().startswith("basic"):
            return _num(r.get("amount"))
    return _num(u.get("salary_monthly"))


async def _log(admin: dict, op: str, company_id: str, count: int, detail: str):
    await db.bulk_ops_log.insert_one({
        "log_id": f"blog_{uuid.uuid4().hex[:10]}",
        "op": op,
        "company_id": company_id,
        "count": count,
        "detail": detail,
        "by": admin.get("user_id"),
        "by_name": admin.get("name"),
        "at": now_iso(),
    })


def _xlsx_response(wb) -> Dict[str, str]:
    buf = io.BytesIO()
    wb.save(buf)
    return {"file_base64": base64.b64encode(buf.getvalue()).decode()}


def _hdr_style(ws, ncols: int):
    from openpyxl.styles import Font, PatternFill
    for c in ws[1][:ncols]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1D4ED8")


async def _employees(company_id: str) -> List[dict]:
    return await db.users.find(
        {"role": "employee", "company_id": company_id},
        {"_id": 0, "user_id": 1, "employee_code": 1, "bio_code": 1, "name": 1,
         "father_name": 1, "designation": 1, "department": 1,
         "contractor_name": 1, "employee_type": 1, "doj": 1, "exit_date": 1,
         "shift_start": 1, "shift_end": 1, "shift_name": 1, "full_day_hrs": 1,
         "salary_structure_actual": 1, "salary_monthly": 1,
         "compliance_gross": 1, "is_onroll": 1},
    ).to_list(10000)


def _code_key(r):
    c = str(r.get("employee_code") or "").strip()
    try:
        return (0, float(c), "")
    except ValueError:
        return (1, 0.0, c.lower())


# ---------------------------------------------------------------------------
# Employee list for the selection UI
# ---------------------------------------------------------------------------
@router.get("/employees")
async def bulk_employees(
    company_id: str = Query(...),
    authorization: Optional[str] = Header(None),
):
    await _auth(authorization, company_id)
    emps = await _employees(company_id)
    emps.sort(key=_code_key)
    rows = []
    for u in emps:
        rows.append({
            "user_id": u["user_id"],
            "employee_code": u.get("employee_code"),
            "name": u.get("name"),
            "designation": u.get("designation"),
            "department": u.get("department"),
            "contractor_name": u.get("contractor_name"),
            "doj": u.get("doj"),
            "exit_date": u.get("exit_date"),
            "shift_name": u.get("shift_name"),
            "shift_start": u.get("shift_start"),
            "shift_end": u.get("shift_end"),
            "actual_basic": round(_actual_basic(u), 2),
            "compliance_gross": round(_num(u.get("compliance_gross")), 2),
        })
    return {"rows": rows, "count": len(rows)}


# ---------------------------------------------------------------------------
# 1) BULK ATTENDANCE UPLOAD
# ---------------------------------------------------------------------------
@router.get("/attendance-template")
async def attendance_template(
    company_id: str = Query(...),
    month: str = Query(...),           # YYYY-MM
    kind: str = Query("status"),       # status | inout
    authorization: Optional[str] = Header(None),
):
    await _auth(authorization, company_id)
    from openpyxl import Workbook
    y, m = int(month[:4]), int(month[5:7])
    days = calendar.monthrange(y, m)[1]
    emps = await _employees(company_id)
    emps = [e for e in emps if not e.get("exit_date")]
    emps.sort(key=_code_key)

    wb = Workbook()
    ws = wb.active
    if kind == "inout":
        ws.title = "InOut"
        ws.append(["Emp Code", "Name", "Date", "In Time", "Out Time"])
        _hdr_style(ws, 5)
        for e in emps:
            ws.append([e.get("employee_code") or "", e.get("name") or "",
                       f"01-{m:02d}-{y}", "09:00", "18:00"])
        for i, w in enumerate([12, 26, 14, 10, 10], start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        fname = f"bulk_attendance_inout_{month}.xlsx"
    else:
        ws.title = "Status"
        hdr = ["Emp Code", "Name"] + [str(d) for d in range(1, days + 1)]
        ws.append(hdr)
        _hdr_style(ws, len(hdr))
        for e in emps:
            ws.append([e.get("employee_code") or "", e.get("name") or ""] + [""] * days)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 26
        from openpyxl.utils import get_column_letter
        for i in range(3, days + 3):
            ws.column_dimensions[get_column_letter(i)].width = 5
        fname = f"bulk_attendance_status_{month}.xlsx"
    out = _xlsx_response(wb)
    out["filename"] = fname
    out["note"] = ("Status codes: P (Present), HD (Half Day), A (Absent), "
                   "WO (Week Off), H (Holiday), L (Leave). Only P & HD create punches.")
    return out


def _read_sheet(data: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _parse_time(v: Any) -> Optional[str]:
    import datetime as _dt
    if v is None or v == "":
        return None
    if isinstance(v, _dt.time):
        return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, _dt.datetime):
        return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, (int, float)):  # excel fraction of day
        mins = int(round(float(v) * 24 * 60)) % (24 * 60)
        return f"{mins // 60:02d}:{mins % 60:02d}"
    s = str(v).strip()
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
        try:
            t = _dt.datetime.strptime(s, fmt)
            return f"{t.hour:02d}:{t.minute:02d}"
        except ValueError:
            continue
    return None


def _parse_date(v: Any) -> Optional[str]:
    import datetime as _dt
    if v is None or v == "":
        return None
    if isinstance(v, _dt.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, _dt.date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return _dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


async def _emp_lookup(company_id: str) -> Dict[str, dict]:
    """Map employee_code / bio_code / lowercase name → employee."""
    emps = await _employees(company_id)
    lookup: Dict[str, dict] = {}
    for e in emps:
        for k in (e.get("employee_code"), e.get("bio_code")):
            k = _cell(k)
            if k:
                lookup.setdefault(f"code:{k.lower()}", e)
        n = _cell(e.get("name")).lower()
        if n:
            lookup.setdefault(f"name:{n}", e)
    return lookup


def _match(lookup: Dict[str, dict], code: str, name: str) -> Optional[dict]:
    if code and f"code:{code.lower()}" in lookup:
        return lookup[f"code:{code.lower()}"]
    if name and f"name:{name.lower()}" in lookup:
        return lookup[f"name:{name.lower()}"]
    return None


def _shift_times(emp: dict, status: str) -> tuple:
    start = _cell(emp.get("shift_start")) or "09:00"
    full_hrs = _num(emp.get("full_day_hrs"), 0) or 9.0
    sh, sm = int(start[:2]), int(start[3:5])
    hours = full_hrs / 2 if status == "HD" else full_hrs
    end_min = (sh * 60 + sm + int(hours * 60)) % (24 * 60)
    end = _cell(emp.get("shift_end"))
    if status == "HD" or not end:
        end = f"{end_min // 60:02d}:{end_min % 60:02d}"
    return start, end


@router.post("/attendance-preview")
async def attendance_preview(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    company_id = str(payload.get("company_id") or "")
    month = str(payload.get("month") or "")
    kind = str(payload.get("kind") or "status")
    await _auth(authorization, company_id)
    if not company_id or not month:
        raise HTTPException(status_code=400, detail="company_id and month are required")
    try:
        data = base64.b64decode(str(payload.get("file_base64") or ""))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid file upload")
    if not data:
        raise HTTPException(status_code=400, detail="Empty file")

    grid = _read_sheet(data)
    if not grid or len(grid) < 2:
        raise HTTPException(status_code=400, detail="Sheet has no data rows")
    lookup = await _emp_lookup(company_id)
    y, m = int(month[:4]), int(month[5:7])
    days = calendar.monthrange(y, m)[1]

    rows: List[dict] = []
    if kind == "inout":
        for raw in grid[1:]:
            raw = list(raw) + [None] * 5
            code, name = _cell(raw[0]), _cell(raw[1])
            if not code and not name:
                continue
            emp = _match(lookup, code, name)
            date = _parse_date(raw[2])
            tin, tout = _parse_time(raw[3]), _parse_time(raw[4])
            err = None
            if not emp:
                err = "Employee not matched"
            elif not date:
                err = "Invalid/missing date"
            elif not tin and not tout:
                err = "No In/Out time"
            rows.append({
                "user_id": emp["user_id"] if emp else None,
                "employee_code": code, "name": emp.get("name") if emp else name,
                "date": date, "in_time": tin, "out_time": tout,
                "status": "error" if err else "matched", "error": err,
            })
    else:
        for raw in grid[1:]:
            raw = list(raw) + [None] * (days + 2)
            code, name = _cell(raw[0]), _cell(raw[1])
            if not code and not name:
                continue
            emp = _match(lookup, code, name)
            day_cells = {}
            bad = []
            for d in range(1, days + 1):
                v = _cell(raw[1 + d]).upper()
                if v and v not in STATUS_CODES:
                    bad.append(f"day {d}: '{v}'")
                    continue
                if v in PUNCH_STATUSES:
                    day_cells[str(d)] = v
            err = None
            if not emp:
                err = "Employee not matched"
            elif bad:
                err = "Unknown codes — " + ", ".join(bad[:4])
            rows.append({
                "user_id": emp["user_id"] if emp else None,
                "employee_code": code, "name": emp.get("name") if emp else name,
                "days": day_cells, "punch_days": len(day_cells),
                "status": "error" if err else "matched", "error": err,
            })

    matched = [r for r in rows if r["status"] == "matched"]
    return {
        "rows": rows,
        "summary": {
            "total": len(rows),
            "matched": len(matched),
            "errors": len(rows) - len(matched),
            "punch_days": sum(r.get("punch_days") or
                              (1 if (r.get("in_time") or r.get("out_time")) else 0)
                              for r in matched),
        },
    }


@router.post("/attendance-apply")
async def attendance_apply(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    company_id = str(payload.get("company_id") or "")
    month = str(payload.get("month") or "")
    kind = str(payload.get("kind") or "status")
    overwrite = bool(payload.get("overwrite"))
    rows = payload.get("rows") or []
    admin = await _auth(authorization, company_id, write=True)
    if not company_id or not month or not rows:
        raise HTTPException(status_code=400, detail="Nothing to apply")

    emps = {e["user_id"]: e for e in await _employees(company_id)}
    batch = f"bulkatt_{uuid.uuid4().hex[:10]}"
    created = skipped = 0
    y, m = int(month[:4]), int(month[5:7])

    async def _insert(uid: str, date: str, kind_p: str, hhmm: str):
        nonlocal created, skipped
        at = f"{date}T{hhmm}:00Z"
        dup = await db.attendance.find_one(
            {"user_id": uid, "date": date, "kind": kind_p, "at": at}, {"_id": 1})
        if dup:
            skipped += 1
            return
        await db.attendance.insert_one({
            "record_id": f"att_{uuid.uuid4().hex[:12]}",
            "user_id": uid,
            "company_id": company_id,
            "date": date,
            "kind": kind_p,
            "at": at,
            "latitude": None, "longitude": None, "distance_m": None,
            "outside_geofence": False,
            "source": "excel_import",
            "import_batch": batch,
            "status": "approved",
            "decision_by": admin["user_id"],
            "decision_at": now_iso(),
            "decision_reason": f"Bulk attendance upload ({batch})",
            "created_at": now_iso(),
        })
        created += 1

    for r in rows:
        uid = str(r.get("user_id") or "")
        emp = emps.get(uid)
        if not emp:
            continue
        if kind == "inout":
            date = str(r.get("date") or "")
            if not date:
                continue
            if overwrite:
                await db.attendance.delete_many(
                    {"user_id": uid, "date": date, "company_id": company_id})
            else:
                existing = await db.attendance.count_documents(
                    {"user_id": uid, "date": date})
                if existing:
                    skipped += 1
                    continue
            if r.get("in_time"):
                await _insert(uid, date, "in", r["in_time"])
            if r.get("out_time"):
                await _insert(uid, date, "out", r["out_time"])
        else:
            for dstr, status in (r.get("days") or {}).items():
                try:
                    d = int(dstr)
                except ValueError:
                    continue
                date = f"{y:04d}-{m:02d}-{d:02d}"
                if overwrite:
                    await db.attendance.delete_many(
                        {"user_id": uid, "date": date, "company_id": company_id})
                else:
                    existing = await db.attendance.count_documents(
                        {"user_id": uid, "date": date})
                    if existing:
                        skipped += 1
                        continue
                tin, tout = _shift_times(emp, status)
                await _insert(uid, date, "in", tin)
                await _insert(uid, date, "out", tout)

    await _log(admin, "attendance_upload", company_id, created,
               f"{kind} upload for {month}: {created} punches created, {skipped} skipped")
    return {"ok": True, "created": created, "skipped": skipped, "batch": batch}


# ---------------------------------------------------------------------------
# 2) BULK SALARY REVISION
# ---------------------------------------------------------------------------
def _revise_actual(u: dict, mode: str, value: float) -> tuple:
    """Return (old_basic, new_basic, new_structure, new_salary_monthly)."""
    structure = [dict(r) for r in (u.get("salary_structure_actual") or []) if isinstance(r, dict)]
    old_basic = _actual_basic(u)
    if mode == "percent":
        factor = 1 + value / 100.0
        for r in structure:
            r["amount"] = round(_num(r.get("amount")) * factor, 2)
        new_monthly = round(_num(u.get("salary_monthly")) * factor, 2) \
            if u.get("salary_monthly") else u.get("salary_monthly")
    else:  # flat — add to the Basic head (or first head)
        target = None
        for r in structure:
            if str(r.get("head", "")).lower().startswith("basic"):
                target = r
                break
        if target is None and structure:
            target = structure[0]
        if target is not None:
            target["amount"] = round(_num(target.get("amount")) + value, 2)
        new_monthly = round(_num(u.get("salary_monthly")) + value, 2) \
            if u.get("salary_monthly") else u.get("salary_monthly")
    new_basic = old_basic
    for r in structure:
        if str(r.get("head", "")).lower().startswith("basic"):
            new_basic = _num(r.get("amount"))
            break
    else:
        if structure:
            new_basic = _num(structure[0].get("amount"))
    return old_basic, new_basic, structure, new_monthly


@router.post("/salary-revision")
async def salary_revision(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    """Select-and-apply mode: percent or flat, on actual / compliance / both."""
    company_id = str(payload.get("company_id") or "")
    user_ids = payload.get("user_ids") or []
    mode = str(payload.get("mode") or "percent")            # percent | flat
    target = str(payload.get("target") or "actual")         # actual | compliance | both
    value = _num(payload.get("value"))
    effective_month = str(payload.get("effective_month") or "")
    note = str(payload.get("note") or "")
    admin = await _auth(authorization, company_id, write=True)
    if not company_id or not user_ids:
        raise HTTPException(status_code=400, detail="Select at least one employee")
    if mode not in ("percent", "flat") or value == 0:
        raise HTTPException(status_code=400, detail="Enter a non-zero revision value")

    changed, results = 0, []
    for uid in user_ids:
        u = await db.users.find_one(
            {"user_id": uid, "company_id": company_id, "role": "employee"},
            {"_id": 0, "user_id": 1, "employee_code": 1, "name": 1,
             "salary_structure_actual": 1, "salary_monthly": 1, "compliance_gross": 1})
        if not u:
            continue
        updates: Dict[str, Any] = {}
        detail: Dict[str, Any] = {"user_id": uid, "employee_code": u.get("employee_code"),
                                  "name": u.get("name")}
        if target in ("actual", "both"):
            old_b, new_b, structure, new_monthly = _revise_actual(u, mode, value)
            updates["salary_structure_actual"] = structure
            if new_monthly is not None:
                updates["salary_monthly"] = new_monthly
            detail["actual_old"] = round(old_b, 2)
            detail["actual_new"] = round(new_b, 2)
        if target in ("compliance", "both"):
            old_c = _num(u.get("compliance_gross"))
            if old_c > 0:
                new_c = round(old_c * (1 + value / 100.0), 2) if mode == "percent" \
                    else round(old_c + value, 2)
                updates["compliance_gross"] = new_c
                detail["compliance_old"] = round(old_c, 2)
                detail["compliance_new"] = new_c
            else:
                detail["compliance_skipped"] = "No compliance salary set"
        if not updates:
            results.append(detail)
            continue
        updates["salary_revised_at"] = now_iso()
        updates["salary_revised_by"] = admin["user_id"]
        await db.users.update_one({"user_id": uid}, {"$set": updates})
        await db.salary_revisions.insert_one({
            "revision_id": f"srev_{uuid.uuid4().hex[:10]}",
            "user_id": uid,
            "company_id": company_id,
            "employee_code": u.get("employee_code"),
            "name": u.get("name"),
            "mode": mode, "target": target, "value": value,
            "effective_month": effective_month,
            "changes": {k: v for k, v in detail.items()
                        if k not in ("user_id", "employee_code", "name")},
            "note": note,
            "by": admin["user_id"], "by_name": admin.get("name"),
            "at": now_iso(),
        })
        changed += 1
        results.append(detail)

    await _log(admin, "salary_revision", company_id, changed,
               f"{mode} {value}{'%' if mode == 'percent' else ''} on {target} "
               f"({changed} employees){' — ' + note if note else ''}")
    return {"ok": True, "changed": changed, "results": results}


@router.get("/salary-template")
async def salary_template(
    company_id: str = Query(...),
    authorization: Optional[str] = Header(None),
):
    await _auth(authorization, company_id)
    from openpyxl import Workbook
    emps = await _employees(company_id)
    emps = [e for e in emps if not e.get("exit_date")]
    emps.sort(key=_code_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Revision"
    ws.append(["Emp Code", "Name", "Current Actual Basic", "New Actual Basic",
               "Current Compliance Gross", "New Compliance Gross"])
    _hdr_style(ws, 6)
    for e in emps:
        ws.append([e.get("employee_code") or "", e.get("name") or "",
                   round(_actual_basic(e), 2),
                   "", round(_num(e.get("compliance_gross")), 2), ""])
    for i, w in enumerate([12, 26, 20, 20, 24, 24], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    out = _xlsx_response(wb)
    out["filename"] = "bulk_salary_revision_template.xlsx"
    return out


@router.post("/salary-preview")
async def salary_preview(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    company_id = str(payload.get("company_id") or "")
    await _auth(authorization, company_id)
    try:
        data = base64.b64decode(str(payload.get("file_base64") or ""))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid file upload")
    grid = _read_sheet(data)
    if not grid or len(grid) < 2:
        raise HTTPException(status_code=400, detail="Sheet has no data rows")
    lookup = await _emp_lookup(company_id)
    rows = []
    for raw in grid[1:]:
        raw = list(raw) + [None] * 6
        code, name = _cell(raw[0]), _cell(raw[1])
        if not code and not name:
            continue
        emp = _match(lookup, code, name)
        new_actual = _num(raw[3], -1) if raw[3] not in (None, "") else -1
        new_comp = _num(raw[5], -1) if raw[5] not in (None, "") else -1
        err = None
        if not emp:
            err = "Employee not matched"
        elif new_actual < 0 and new_comp < 0:
            err = "No new amount entered"
        rows.append({
            "user_id": emp["user_id"] if emp else None,
            "employee_code": code, "name": emp.get("name") if emp else name,
            "current_actual": round(_actual_basic(emp), 2) if emp else None,
            "new_actual": new_actual if new_actual >= 0 else None,
            "current_compliance": round(_num(emp.get("compliance_gross")), 2) if emp else None,
            "new_compliance": new_comp if new_comp >= 0 else None,
            "status": "error" if err else "matched", "error": err,
        })
    matched = [r for r in rows if r["status"] == "matched"]
    return {"rows": rows, "summary": {"total": len(rows), "matched": len(matched),
                                      "errors": len(rows) - len(matched)}}


@router.post("/salary-apply-excel")
async def salary_apply_excel(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    company_id = str(payload.get("company_id") or "")
    rows = payload.get("rows") or []
    effective_month = str(payload.get("effective_month") or "")
    note = str(payload.get("note") or "")
    admin = await _auth(authorization, company_id, write=True)
    if not rows:
        raise HTTPException(status_code=400, detail="Nothing to apply")
    changed = 0
    for r in rows:
        uid = str(r.get("user_id") or "")
        if not uid:
            continue
        u = await db.users.find_one(
            {"user_id": uid, "company_id": company_id, "role": "employee"},
            {"_id": 0, "user_id": 1, "employee_code": 1, "name": 1,
             "salary_structure_actual": 1, "salary_monthly": 1, "compliance_gross": 1})
        if not u:
            continue
        updates: Dict[str, Any] = {}
        changes: Dict[str, Any] = {}
        new_actual = r.get("new_actual")
        new_comp = r.get("new_compliance")
        if new_actual is not None and _num(new_actual, -1) >= 0:
            old_b = _actual_basic(u)
            structure = [dict(x) for x in (u.get("salary_structure_actual") or [])
                         if isinstance(x, dict)]
            hit = False
            for x in structure:
                if str(x.get("head", "")).lower().startswith("basic"):
                    x["amount"] = round(_num(new_actual), 2)
                    hit = True
                    break
            if not hit:
                structure.insert(0, {"head": "Basic Salary",
                                     "amount": round(_num(new_actual), 2)})
            updates["salary_structure_actual"] = structure
            if u.get("salary_monthly"):
                updates["salary_monthly"] = round(_num(new_actual), 2)
            changes["actual_old"] = round(old_b, 2)
            changes["actual_new"] = round(_num(new_actual), 2)
        if new_comp is not None and _num(new_comp, -1) >= 0:
            changes["compliance_old"] = round(_num(u.get("compliance_gross")), 2)
            changes["compliance_new"] = round(_num(new_comp), 2)
            updates["compliance_gross"] = round(_num(new_comp), 2)
        if not updates:
            continue
        updates["salary_revised_at"] = now_iso()
        updates["salary_revised_by"] = admin["user_id"]
        await db.users.update_one({"user_id": uid}, {"$set": updates})
        await db.salary_revisions.insert_one({
            "revision_id": f"srev_{uuid.uuid4().hex[:10]}",
            "user_id": uid, "company_id": company_id,
            "employee_code": u.get("employee_code"), "name": u.get("name"),
            "mode": "excel", "target": "excel", "value": None,
            "effective_month": effective_month,
            "changes": changes, "note": note,
            "by": admin["user_id"], "by_name": admin.get("name"),
            "at": now_iso(),
        })
        changed += 1
    await _log(admin, "salary_revision", company_id, changed,
               f"Excel salary revision ({changed} employees)"
               f"{' — ' + note if note else ''}")
    return {"ok": True, "changed": changed}


# ---------------------------------------------------------------------------
# 3) BULK TRANSFER (between firms)
# ---------------------------------------------------------------------------
@router.post("/transfer")
async def bulk_transfer(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    from_company = str(payload.get("company_id") or "")
    to_company = str(payload.get("to_company_id") or "")
    user_ids = payload.get("user_ids") or []
    effective_date = str(payload.get("effective_date") or "")
    note = str(payload.get("note") or "")
    admin = await _auth(authorization, from_company, write=True)
    if admin.get("role") == "sub_admin" and not sub_admin_can_touch_company(admin, to_company):
        raise HTTPException(status_code=403, detail="Destination firm not in your scope")
    if not to_company or not user_ids:
        raise HTTPException(status_code=400, detail="Select employees and a destination firm")
    if to_company == from_company:
        raise HTTPException(status_code=400, detail="Destination is the same firm")
    dest = await db.companies.find_one({"company_id": to_company}, {"_id": 0, "name": 1})
    if not dest:
        raise HTTPException(status_code=404, detail="Destination firm not found")

    moved = 0
    for uid in user_ids:
        u = await db.users.find_one(
            {"user_id": uid, "company_id": from_company, "role": "employee"},
            {"_id": 0, "user_id": 1, "employee_code": 1, "name": 1})
        if not u:
            continue
        await db.users.update_one({"user_id": uid}, {"$set": {
            "company_id": to_company,
            "transferred_from": from_company,
            "transferred_at": now_iso(),
            "transfer_effective_date": effective_date or None,
            "transfer_note": note or None,
            "transferred_by": admin["user_id"],
        }})
        moved += 1
    await _log(admin, "transfer", from_company, moved,
               f"Transferred {moved} employees to {dest.get('name')}"
               f"{' w.e.f. ' + effective_date if effective_date else ''}")
    return {"ok": True, "moved": moved, "to_company_name": dest.get("name")}


# ---------------------------------------------------------------------------
# 4) BULK RESIGNATION (exit date)
# ---------------------------------------------------------------------------
@router.post("/resignation")
async def bulk_resignation(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    company_id = str(payload.get("company_id") or "")
    user_ids = payload.get("user_ids") or []
    exit_date = str(payload.get("exit_date") or "")
    reason = str(payload.get("reason") or "")
    admin = await _auth(authorization, company_id, write=True)
    if not user_ids or not exit_date:
        raise HTTPException(status_code=400, detail="Select employees and an exit date")
    import re
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", exit_date):
        raise HTTPException(status_code=400, detail="Exit date must be YYYY-MM-DD")
    done = 0
    for uid in user_ids:
        r = await db.users.update_one(
            {"user_id": uid, "company_id": company_id, "role": "employee"},
            {"$set": {
                "exit_date": exit_date,
                "employment_status": "resigned",
                "exit_reason": reason or None,
                "exit_set_at": now_iso(),
                "exit_set_by": admin["user_id"],
            }})
        if r.modified_count:
            done += 1
    await _log(admin, "resignation", company_id, done,
               f"Exit date {exit_date} set for {done} employees"
               f"{' — ' + reason if reason else ''}")
    return {"ok": True, "updated": done}


# ---------------------------------------------------------------------------
# 5) BULK SHIFT ASSIGNMENT
# ---------------------------------------------------------------------------
@router.post("/shift-assign")
async def bulk_shift_assign(
    payload: Dict[str, Any] = Body(...),
    authorization: Optional[str] = Header(None),
):
    company_id = str(payload.get("company_id") or "")
    user_ids = payload.get("user_ids") or []
    shift_id = str(payload.get("shift_id") or "")
    admin = await _auth(authorization, company_id, write=True)
    if not user_ids or not shift_id:
        raise HTTPException(status_code=400, detail="Select employees and a shift")
    shift = await db.shift_masters.find_one({"shift_id": shift_id}, {"_id": 0})
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    done = 0
    for uid in user_ids:
        r = await db.users.update_one(
            {"user_id": uid, "company_id": company_id, "role": "employee"},
            {"$set": {
                "shift_id": shift_id,
                "shift_name": shift.get("name"),
                "shift_start": shift.get("start"),
                "shift_end": shift.get("end"),
                "shift_assigned_at": now_iso(),
                "shift_assigned_by": admin["user_id"],
            }})
        if r.modified_count:
            done += 1
    await _log(admin, "shift_assign", company_id, done,
               f"Shift '{shift.get('name')}' ({shift.get('start')}-{shift.get('end')}) "
               f"assigned to {done} employees")
    return {"ok": True, "updated": done, "shift_name": shift.get("name")}


# ---------------------------------------------------------------------------
# History
# ---------------------------------------------------------------------------
@router.get("/history")
async def bulk_history(
    company_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    admin = await _auth(authorization, company_id or "")
    q: Dict[str, Any] = {}
    if company_id:
        q["company_id"] = company_id
    logs = await db.bulk_ops_log.find(q, {"_id": 0}).sort("at", -1).to_list(200)
    ids = {l["company_id"] for l in logs if l.get("company_id")}
    names = {}
    if ids:
        async for c in db.companies.find(
                {"company_id": {"$in": list(ids)}}, {"_id": 0, "company_id": 1, "name": 1}):
            names[c["company_id"]] = c["name"]
    for l in logs:
        l["company_name"] = names.get(l.get("company_id"))
    return {"rows": logs}
