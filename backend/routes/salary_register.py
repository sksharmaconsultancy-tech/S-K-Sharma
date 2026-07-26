"""Iter 299 — Enterprise Salary Register (Payroll → Salary Process).

A dynamic, enterprise-grade salary register over BOTH engines:

  * ``compliance`` — rows from ``compliance_salary_runs`` (PF/ESIC/PT/TDS)
  * ``actual``     — rows from ``salary_runs`` (run_type == "actual")

Columns are NOT hardcoded: a per-source ordered registry provides labels
and grouping for the known heads, and a catch-all sweep appends any
numeric head found in the run rows that isn't registered (so future
dynamic heads — e.g. "Salary 1/2/3" — appear automatically). All-zero
optional heads are pruned so the grid stays tight.

Endpoints (super_admin / sub_admin / company_admin):

  GET /api/admin/salary-register/filters?source=&company_id=
  GET /api/admin/salary-register?source=&company_id=&month=&run_id=&employee_type=&branch=&department=&contractor=&search=&sort_by=&sort_dir=&page=&page_size=
  GET /api/admin/salary-register/export.csv?...   (same filters)
  GET /api/admin/salary-register/export.xlsx?...
  GET /api/admin/salary-register/export.pdf?...
"""
import base64
import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Body, Header, HTTPException
from fastapi.responses import StreamingResponse

from server import (  # noqa: E402
    db,
    get_user_from_token,
    require_role,
)

router = APIRouter(prefix="/api/admin/salary-register", tags=["salary-register"])

# ---------------------------------------------------------------------------
# Column registries (ordered). group: info | attendance | earnings |
# deductions | employer | net. type: text | num | int
# ---------------------------------------------------------------------------

GROUPS = [
    {"key": "info", "label": "Employee"},
    {"key": "attendance", "label": "Attendance"},
    {"key": "earnings", "label": "Earnings"},
    {"key": "deductions", "label": "Deductions"},
    {"key": "employer", "label": "Employer Contributions"},
    {"key": "net", "label": "Net"},
]

_COMPLIANCE_REGISTRY: List[Tuple[str, str, str, str]] = [
    # (key, label, group, type)
    ("employee_code", "Emp Code", "info", "text"),
    ("name", "Employee Name", "info", "text"),
    ("father_name", "Father / Spouse", "info", "text"),
    ("designation", "Designation", "info", "text"),
    ("department", "Department", "info", "text"),
    ("branch_name", "Branch", "info", "text"),
    ("employee_type", "Group", "info", "text"),
    ("contractor_name", "Contractor", "info", "text"),
    ("uan_no", "UAN", "info", "text"),
    ("esi_ip_no", "ESI IP No", "info", "text"),
    ("salary_mode", "Mode", "info", "text"),
    ("rate", "Rate", "attendance", "num"),
    ("month_days", "Month Days", "attendance", "int"),
    ("present_days", "Present", "attendance", "num"),
    ("half_days", "Half Days", "attendance", "num"),
    ("ot_hours", "OT Hrs", "attendance", "num"),
    ("basic", "Basic", "earnings", "num"),
    ("hra", "HRA", "earnings", "num"),
    ("conveyance", "Conveyance", "earnings", "num"),
    ("medical", "Medical", "earnings", "num"),
    ("special", "Special", "earnings", "num"),
    ("others", "Others", "earnings", "num"),
    ("ot_pay", "OT Pay", "earnings", "num"),
    ("monthly_gross", "Monthly Gross", "earnings", "num"),
    ("gross_paid", "Gross Paid", "earnings", "num"),
    ("pf_wages", "PF Wages", "deductions", "num"),
    ("pf_employee", "PF (Emp)", "deductions", "num"),
    ("vpf_amount", "VPF", "deductions", "num"),
    ("esic_employee", "ESIC (Emp)", "deductions", "num"),
    ("pt", "Prof. Tax", "deductions", "num"),
    ("tds", "TDS", "deductions", "num"),
    ("master_deduction", "Master Ded.", "deductions", "num"),
    ("other_deduction", "Other Ded.", "deductions", "num"),
    ("total_deduction", "Total Ded.", "deductions", "num"),
    ("pf_employer_epf", "EPF (Er)", "employer", "num"),
    ("pf_employer_eps", "EPS (Er)", "employer", "num"),
    ("esic_employer", "ESIC (Er)", "employer", "num"),
    ("net", "Net Pay", "net", "num"),
]

_ACTUAL_REGISTRY: List[Tuple[str, str, str, str]] = [
    ("employee_code", "Emp Code", "info", "text"),
    ("name", "Employee Name", "info", "text"),
    ("father_name", "Father / Spouse", "info", "text"),
    ("designation", "Designation", "info", "text"),
    ("department", "Department", "info", "text"),
    ("employee_type", "Group", "info", "text"),
    ("salary_mode", "Mode", "info", "text"),
    ("doj", "DOJ", "info", "text"),
    ("basic", "Rate", "attendance", "num"),
    ("max_p_days", "Month Days", "attendance", "int"),
    ("p_days", "Present", "attendance", "num"),
    ("p_hours", "Hours", "attendance", "num"),
    ("duty_hrs", "Duty Hrs", "attendance", "num"),
    ("basic_salary", "Basic Salary", "earnings", "num"),
    ("w_basic_salary", "Weighted Basic", "earnings", "num"),
    ("oth_allo", "Other Allow.", "earnings", "num"),
    ("total_gross", "Total Gross", "earnings", "num"),
    ("epf", "EPF", "deductions", "num"),
    ("esi", "ESI", "deductions", "num"),
    ("adv", "Advance", "deductions", "num"),
    ("tds", "TDS", "deductions", "num"),
    ("net_pay", "Net Pay", "net", "num"),
]

# Row keys that must never become columns.
_EXCLUDE_KEYS = {
    "user_id", "company_id", "company_name", "is_onroll", "pf_applicable",
    "esic_applicable", "enabled_allowances", "enabled_deductions",
    "pt_state", "exit_date", "max_p_days", "month_days",
    "basic_master", "hra_master", "conveyance_master", "medical_master",
    "special_master", "others_master", "gross_master",
    "pf_basic", "stat_wage_base", "esic_wage_base", "pf_employer_total",
    "duty_hours", "compliance_basic",
}

# Columns always kept even if all-zero.
_ALWAYS_KEEP = {
    "employee_code", "name", "present_days", "p_days", "basic",
    "basic_salary", "monthly_gross", "gross_paid", "total_gross",
    "total_deduction", "net", "net_pay", "rate",
}


def _registry(source: str) -> List[Tuple[str, str, str, str]]:
    return _COMPLIANCE_REGISTRY if source == "compliance" else _ACTUAL_REGISTRY


# ---------------------------------------------------------------------------
# Iter 309 (user) — full layout editing for the dynamic Salary Register:
# choose columns · order · rename headings · widths · employees per page ·
# row height. Saved once in app_settings; applied to the GRID, Excel, CSV,
# PDF and Email exports alike.
# ---------------------------------------------------------------------------

_LAYOUT_KEY = {"key": "salary_register_module_layout"}


def _layout_catalog() -> List[Dict[str, Any]]:
    """Union of both source registries (key → default heading/width)."""
    seen: Dict[str, Dict[str, Any]] = {}
    for reg in (_COMPLIANCE_REGISTRY, _ACTUAL_REGISTRY):
        for k, lbl, _g, t in reg:
            if k not in seen:
                seen[k] = {"key": k, "heading": lbl,
                           "width": 28 if k == "name" else (22 if t == "text" else 16),
                           "numeric": t != "text"}
    return list(seen.values())


async def _saved_layout() -> Optional[Dict[str, Any]]:
    doc = await db.app_settings.find_one(_LAYOUT_KEY, {"_id": 0, "layout": 1})
    return (doc or {}).get("layout") or None


def _apply_layout(columns: List[Dict[str, Any]],
                  layout: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Select/order/rename/size the run's columns per the saved layout.
    Dynamic heads not in the catalog stay visible (appended at the end)."""
    if not layout or not layout.get("columns"):
        return columns
    by_key = {c["key"]: c for c in columns}
    catalog_keys = {c["key"] for c in _layout_catalog()}
    out: List[Dict[str, Any]] = []
    for item in layout["columns"]:
        c = by_key.get(item.get("key"))
        if not c:
            continue
        c = dict(c)
        if item.get("heading"):
            c["label"] = str(item["heading"])
        if item.get("width"):
            c["width"] = float(item["width"])
        out.append(c)
    chosen = {c["key"] for c in out}
    # dynamic extras (not configurable in the catalog) remain visible
    out.extend(c for c in columns
               if c["key"] not in chosen and c["key"] not in catalog_keys)
    return out


@router.get("/layout")
async def get_module_layout(authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "sub_admin", "company_admin"])
    doc = await db.app_settings.find_one(_LAYOUT_KEY, {"_id": 0}) or {}
    return {"layout": doc.get("layout") or None, "catalog": _layout_catalog(),
            "updated_at": doc.get("updated_at"),
            "updated_by_name": doc.get("updated_by_name")}


@router.put("/layout")
async def put_module_layout(payload: Dict[str, Any] = Body(...),
                            authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin"])
    valid = {c["key"] for c in _layout_catalog()}
    cols: List[Dict[str, Any]] = []
    for c in payload.get("columns") or []:
        if not isinstance(c, dict) or c.get("key") not in valid:
            continue
        item: Dict[str, Any] = {"key": c["key"]}
        if str(c.get("heading") or "").strip():
            item["heading"] = str(c["heading"]).strip()[:40]
        try:
            w = float(c.get("width") or 0)
            if w > 0:
                item["width"] = max(4.0, min(80.0, w))
        except (TypeError, ValueError):
            pass
        cols.append(item)
    if not cols:
        raise HTTPException(status_code=400, detail="Select at least one column")
    layout: Dict[str, Any] = {"columns": cols}
    try:
        pp = int(payload.get("per_page") or 0)
        if pp > 0:
            layout["per_page"] = max(1, min(100, pp))
    except (TypeError, ValueError):
        pass
    try:
        rh = float(payload.get("row_height") or 0)
        if rh > 0:
            layout["row_height"] = max(3.0, min(20.0, rh))
    except (TypeError, ValueError):
        pass
    from datetime import datetime, timezone
    await db.app_settings.update_one(_LAYOUT_KEY, {"$set": {
        **_LAYOUT_KEY, "layout": layout,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by_name": admin.get("name") or admin.get("email"),
    }}, upsert=True)
    return {"ok": True, "layout": layout}


@router.delete("/layout")
async def delete_module_layout(authorization: Optional[str] = Header(None)):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin"])
    await db.app_settings.delete_one(_LAYOUT_KEY)
    return {"ok": True}


def _build_columns(source: str, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Ordered registry columns + dynamic catch-all + all-zero pruning."""
    reg = _registry(source)
    reg_keys = {k for k, _, _, _ in reg}
    cols: List[Dict[str, Any]] = [
        {"key": k, "label": lbl, "group": g, "type": t} for k, lbl, g, t in reg
    ]
    # Dynamic sweep: numeric keys present in rows but not registered.
    extra: List[str] = []
    seen: set = set()
    for r in rows:
        for k, v in r.items():
            if k in reg_keys or k in _EXCLUDE_KEYS or k in seen:
                continue
            if isinstance(v, bool):
                continue
            if isinstance(v, (int, float)):
                seen.add(k)
                extra.append(k)
    net_idx = next((i for i, c in enumerate(cols) if c["group"] == "net"), len(cols))
    for k in extra:
        label = k.replace("_", " ").title()
        cols.insert(net_idx, {"key": k, "label": label, "group": "earnings", "type": "num"})
        net_idx += 1
    # Prune numeric columns that are zero/empty across ALL rows.
    if rows:
        keep: List[Dict[str, Any]] = []
        for c in cols:
            if c["type"] == "text" or c["key"] in _ALWAYS_KEEP:
                keep.append(c)
                continue
            if any(float(r.get(c["key"]) or 0) != 0 for r in rows):
                keep.append(c)
        cols = keep
    return cols


# ---------------------------------------------------------------------------
# Data fetch + filtering
# ---------------------------------------------------------------------------

def _resolve_company(admin: Dict[str, Any], company_id: Optional[str]) -> str:
    if admin["role"] == "company_admin":
        return admin.get("company_id") or ""
    if not company_id:
        raise HTTPException(status_code=400, detail="company_id is required")
    return company_id


def _run_query(source: str, company_id: str, month: str) -> Tuple[Any, Dict[str, Any]]:
    if source == "compliance":
        return db.compliance_salary_runs, {"company_id": company_id, "month": month}
    return db.salary_runs, {
        "company_id": company_id, "month": month,
        "$or": [{"run_type": "actual"}, {"run_type": {"$exists": False}}],
    }


async def _load_run(
    source: str, company_id: str, month: str, run_id: Optional[str],
) -> Optional[Dict[str, Any]]:
    coll, q = _run_query(source, company_id, month)
    if run_id:
        q = {**q, "run_id": run_id}
    return await coll.find_one(q, {"_id": 0}, sort=[("generated_at", -1)])


def _filter_rows(
    rows: List[Dict[str, Any]],
    employee_type: Optional[str],
    branch: Optional[str],
    department: Optional[str],
    contractor: Optional[str],
    search: Optional[str],
) -> List[Dict[str, Any]]:
    out = rows
    if employee_type:
        out = [r for r in out if (r.get("employee_type") or "") == employee_type]
    if branch:
        out = [r for r in out if (r.get("branch_name") or "") == branch]
    if department:
        out = [r for r in out if (r.get("department") or "") == department]
    if contractor:
        out = [r for r in out if (r.get("contractor_name") or "") == contractor]
    if search:
        s = search.strip().lower()
        out = [
            r for r in out
            if s in str(r.get("name") or "").lower()
            or s in str(r.get("employee_code") or "").lower()
            or s in str(r.get("father_name") or "").lower()
            or s in str(r.get("designation") or "").lower()
        ]
    return out


def _sort_rows(rows: List[Dict[str, Any]], sort_by: str, sort_dir: str) -> List[Dict[str, Any]]:
    rev = sort_dir == "desc"

    def keyf(r: Dict[str, Any]):
        v = r.get(sort_by)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return (0, float(v), "")
        # employee_code: numeric-aware sort
        s = str(v or "")
        if s.isdigit():
            return (0, float(s), "")
        return (1, 0.0, s.lower())

    return sorted(rows, key=keyf, reverse=rev)


def _totals(rows: List[Dict[str, Any]], columns: List[Dict[str, Any]]) -> Dict[str, float]:
    t: Dict[str, float] = {}
    for c in columns:
        if c["type"] == "text":
            continue
        t[c["key"]] = round(sum(float(r.get(c["key"]) or 0) for r in rows), 2)
    return t


async def _prepare(
    admin: Dict[str, Any],
    source: str,
    company_id: Optional[str],
    month: Optional[str],
    run_id: Optional[str],
    employee_type: Optional[str],
    branch: Optional[str],
    department: Optional[str],
    contractor: Optional[str],
    search: Optional[str],
    sort_by: Optional[str],
    sort_dir: str,
) -> Dict[str, Any]:
    """Shared loader for grid + exports. Returns run, filtered rows, columns."""
    if source not in ("compliance", "actual"):
        raise HTTPException(status_code=400, detail="source must be compliance | actual")
    if not month:
        raise HTTPException(status_code=400, detail="month (YYYY-MM) is required")
    cid = _resolve_company(admin, company_id)
    run = await _load_run(source, cid, month, run_id)
    if not run:
        return {"run": None, "rows": [], "columns": [], "company_id": cid}
    rows = _filter_rows(
        run.get("rows") or [], employee_type, branch, department, contractor, search,
    )
    columns = _build_columns(source, rows)
    # Iter 309 — saved layout (columns/order/headings/widths) applies to
    # the grid AND every export.
    layout = await _saved_layout()
    columns = _apply_layout(columns, layout)
    if sort_by:
        rows = _sort_rows(rows, sort_by, sort_dir)
    else:
        rows = _sort_rows(rows, "employee_code", "asc")
    return {"run": run, "rows": rows, "columns": columns,
            "company_id": cid, "layout": layout}


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/filters")
async def register_filters(
    source: str = "compliance",
    company_id: Optional[str] = None,
    month: Optional[str] = None,
    authorization: Optional[str] = Header(None),
):
    """Months with runs, plus distinct branch/dept/type/contractor values."""
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    cid = _resolve_company(admin, company_id)

    if source == "compliance":
        coll, base_q = db.compliance_salary_runs, {"company_id": cid}
    else:
        coll, base_q = db.salary_runs, {
            "company_id": cid,
            "$or": [{"run_type": "actual"}, {"run_type": {"$exists": False}}],
        }
    months = sorted(await coll.distinct("month", base_q), reverse=True)

    runs_for_month: List[Dict[str, Any]] = []
    branches: set = set()
    departments: set = set()
    types: set = set()
    contractors: set = set()
    if month:
        cursor = coll.find(
            {**base_q, "month": month},
            {"_id": 0, "run_id": 1, "generated_at": 1, "employees_count": 1,
             "employee_type": 1, "rows.branch_name": 1, "rows.department": 1,
             "rows.employee_type": 1, "rows.contractor_name": 1},
        ).sort("generated_at", -1)
        async for run in cursor:
            runs_for_month.append({
                "run_id": run.get("run_id"),
                "generated_at": run.get("generated_at"),
                "employees_count": run.get("employees_count"),
                "employee_type_filter": run.get("employee_type"),
            })
            for r in run.get("rows") or []:
                if r.get("branch_name"):
                    branches.add(r["branch_name"])
                if r.get("department"):
                    departments.add(r["department"])
                if r.get("employee_type"):
                    types.add(r["employee_type"])
                if r.get("contractor_name"):
                    contractors.add(r["contractor_name"])
    return {
        "months": months,
        "runs": runs_for_month,
        "branches": sorted(branches),
        "departments": sorted(departments),
        "employee_types": sorted(types),
        "contractors": sorted(contractors),
        # Iter 307 — default recipient for "Email register to firm".
        "firm_email": await _firm_email(cid),
    }


@router.get("")
async def salary_register(
    source: str = "compliance",
    company_id: Optional[str] = None,
    month: Optional[str] = None,
    run_id: Optional[str] = None,
    employee_type: Optional[str] = None,
    branch: Optional[str] = None,
    department: Optional[str] = None,
    contractor: Optional[str] = None,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: str = "asc",
    page: int = 1,
    page_size: int = 50,
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    prep = await _prepare(
        admin, source, company_id, month, run_id, employee_type,
        branch, department, contractor, search, sort_by, sort_dir,
    )
    run, rows, columns = prep["run"], prep["rows"], prep["columns"]
    if not run:
        return {
            "columns": [], "groups": GROUPS, "rows": [], "totals": {},
            "total_rows": 0, "page": 1, "page_size": page_size, "run_meta": None,
        }
    page_size = max(10, min(page_size, 500))
    page = max(1, page)
    total = len(rows)
    start = (page - 1) * page_size
    page_rows = rows[start:start + page_size]
    # strip non-column keys to shrink the payload
    col_keys = [c["key"] for c in columns]
    slim = [{k: r.get(k) for k in col_keys} for r in page_rows]
    return {
        "columns": columns,
        "groups": GROUPS,
        "rows": slim,
        "totals": _totals(rows, columns),
        "total_rows": total,
        "page": page,
        "page_size": page_size,
        "run_meta": {
            "run_id": run.get("run_id"),
            "generated_at": run.get("generated_at"),
            "employees_count": run.get("employees_count"),
            "month": run.get("month"),
            "month_days": run.get("month_days") or run.get("default_month_days"),
        },
    }


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------

def _fmt_cell(v: Any, ctype: str) -> Any:
    if v is None:
        return "" if ctype == "text" else 0
    if ctype == "num":
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return 0
    if ctype == "int":
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return 0
    return str(v)


async def _company_name(company_id: str) -> str:
    c = await db.companies.find_one({"company_id": company_id}, {"_id": 0, "name": 1})
    return (c or {}).get("name") or company_id


def _export_filename(source: str, month: str, ext: str,
                     employee_type: Optional[str] = None) -> str:
    """Iter 312 (user directive) — ``Reportname_Group_MonthYear.ext``
    e.g. ``SalaryRegister_Worker_June2026.pdf``."""
    try:
        my = datetime.strptime((month or "")[:7], "%Y-%m").strftime("%B%Y")
    except ValueError:
        my = (month or "month").replace("-", "")
    grp = "".join(w.capitalize() for w in str(employee_type).split()) \
        if employee_type else "All"
    return f"SalaryRegister_{grp}_{my}.{ext}"


@router.get("/export.csv")
async def export_csv(
    source: str = "compliance",
    company_id: Optional[str] = None,
    month: Optional[str] = None,
    run_id: Optional[str] = None,
    employee_type: Optional[str] = None,
    branch: Optional[str] = None,
    department: Optional[str] = None,
    contractor: Optional[str] = None,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: str = "asc",
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    prep = await _prepare(
        admin, source, company_id, month, run_id, employee_type,
        branch, department, contractor, search, sort_by, sort_dir,
    )
    rows, columns = prep["rows"], prep["columns"]
    if not prep["run"]:
        raise HTTPException(status_code=404, detail="No salary run found for this month")
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Sr"] + [c["label"] for c in columns])
    for i, r in enumerate(rows, 1):
        w.writerow([i] + [_fmt_cell(r.get(c["key"]), c["type"]) for c in columns])
    t = _totals(rows, columns)
    w.writerow(["", "TOTAL"] + [
        t.get(c["key"], "") if c["type"] != "text" else ""
        for c in columns[1:]
    ])
    data = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(data), media_type="text/csv",
        headers={"Content-Disposition":
                 f'attachment; filename="{_export_filename(source, month, "csv", employee_type)}"'},
    )


def _xlsx_bytes(comp: str, source: str, month: str,
                rows: List[Dict[str, Any]], columns: List[Dict[str, Any]]) -> bytes:
    """Styled Salary Register workbook (shared by download + email)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Register"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    band_fills = {
        "info": PatternFill("solid", fgColor="44546A"),
        "attendance": PatternFill("solid", fgColor="7F7F7F"),
        "earnings": PatternFill("solid", fgColor="2E7D32"),
        "deductions": PatternFill("solid", fgColor="B71C1C"),
        "employer": PatternFill("solid", fgColor="6A1B9A"),
        "net": PatternFill("solid", fgColor="1F4E79"),
    }
    white_bold = Font(bold=True, color="FFFFFF", size=9)
    center = Alignment(horizontal="center", vertical="center")

    ws.append([f"{comp} — Salary Register ({source.title()}) — {month}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns) + 1)
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

    # Band row (group headers)
    band_row = ["#"]
    for c in columns:
        band_row.append(next((g["label"] for g in GROUPS if g["key"] == c["group"]), ""))
    ws.append(band_row)
    # merge consecutive identical bands
    r_idx = 2
    col_i = 2
    while col_i <= len(columns) + 1:
        j = col_i
        val = ws.cell(row=r_idx, column=col_i).value
        while j + 1 <= len(columns) + 1 and ws.cell(row=r_idx, column=j + 1).value == val:
            j += 1
        if j > col_i:
            ws.merge_cells(start_row=r_idx, start_column=col_i, end_row=r_idx, end_column=j)
        grp = columns[col_i - 2]["group"]
        cell = ws.cell(row=r_idx, column=col_i)
        cell.fill = band_fills.get(grp, head_fill)
        cell.font = white_bold
        cell.alignment = center
        col_i = j + 1

    # Header row
    ws.append(["Sr"] + [c["label"] for c in columns])
    for ci in range(1, len(columns) + 2):
        cell = ws.cell(row=3, column=ci)
        cell.fill = head_fill
        cell.font = white_bold
        cell.alignment = center
        cell.border = border

    for i, r in enumerate(rows, 1):
        ws.append([i] + [_fmt_cell(r.get(c["key"]), c["type"]) for c in columns])
    t = _totals(rows, columns)
    total_row = ["", "TOTAL"] + [
        t.get(c["key"], "") if c["type"] != "text" else "" for c in columns[1:]
    ]
    ws.append(total_row)
    last = ws.max_row
    for ci in range(1, len(columns) + 2):
        cell = ws.cell(row=last, column=ci)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor="FFF3CD")

    # widths + number formats + freeze
    ws.column_dimensions["A"].width = 5
    for idx, c in enumerate(columns, 2):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = 16 if c["key"] in ("name", "father_name") else 11
        if c["type"] == "num":
            for rr in range(4, last + 1):
                ws.cell(row=rr, column=idx).number_format = "#,##0.00"
    name_col = next((i for i, c in enumerate(columns, 2) if c["key"] == "name"), 3)
    ws.freeze_panes = ws.cell(row=4, column=name_col + 1)

    # ---- Iter 308c (user) — second "Summary" sheet (same as the PDF's
    # last page: head-wise totals, days/net, rupees in words, signatures).
    from utils.salary_register_pdf import _num_to_words_inr
    tot = _totals(rows, columns)
    s2 = wb.create_sheet("Summary")
    sec_head_fill = PatternFill("solid", fgColor="1F4E79")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    def head_label(c: Dict[str, Any]) -> str:
        lbl_txt = str(c["label"])
        return lbl_txt if lbl_txt.lower().startswith("total") else f"Total {lbl_txt}"

    s2.append([f"SUMMARY — {comp} — {month} ({source.title()})"])
    s2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    s2.cell(row=1, column=1).font = Font(bold=True, size=12)

    def sec_rows(title_txt: str, pairs, bold_last=True):
        s2.append([])
        s2.append([title_txt])
        hc = s2.cell(row=s2.max_row, column=1)
        hc.fill = sec_head_fill
        hc.font = Font(bold=True, color="FFFFFF", size=10)
        s2.merge_cells(start_row=s2.max_row, start_column=1,
                       end_row=s2.max_row, end_column=2)
        for i, (k, v) in enumerate(pairs):
            s2.append([k, v])
            r_i = s2.max_row
            bold = bold_last and i == len(pairs) - 1
            for ci in (1, 2):
                cell = s2.cell(row=r_i, column=ci)
                cell.border = box
                if bold:
                    cell.font = Font(bold=True, size=10)
                if ci == 2 and isinstance(v, (int, float)):
                    cell.number_format = "#,##0.00"

    earn_cols = [c for c in columns if c["group"] == "earnings"]
    ded_cols = [c for c in columns if c["group"] == "deductions"]
    er_cols = [c for c in columns if c["group"] == "employer"]
    net_col = next((c for c in columns if c["group"] == "net"), None)
    gross_key = next((k for k in ("gross_paid", "total_gross", "monthly_gross")
                      if k in tot), None)
    days_key = next((k for k in ("present_days", "p_days") if k in tot), None)
    hrs_key = next((k for k in ("ot_hours", "p_hours") if k in tot), None)

    sec_rows("EARNINGS",
             [("No. Of Emp", len(rows))]
             + [(head_label(c), tot.get(c["key"], 0)) for c in earn_cols])
    if ded_cols:
        sec_rows("DEDUCTIONS", [(head_label(c), tot.get(c["key"], 0)) for c in ded_cols])
    if er_cols:
        sec_rows("EMPLOYER CONTRIBUTIONS",
                 [(head_label(c), tot.get(c["key"], 0)) for c in er_cols],
                 bold_last=False)
    tail = []
    if days_key:
        tail.append(("Total Days ->", tot.get(days_key, 0)))
    if hrs_key:
        tail.append(("Total Hours ->", tot.get(hrs_key, 0)))
    if net_col:
        tail.append(("Net Payable Amount", tot.get(net_col["key"], 0)))
    if tail:
        sec_rows("NET", tail)
    s2.append([])
    if gross_key:
        s2.append([f"RUPEES: {_num_to_words_inr(int(round(tot.get(gross_key, 0))))} (GROSS)"])
        s2.cell(row=s2.max_row, column=1).font = Font(bold=True, size=10)
    if net_col:
        s2.append([f"RUPEES: {_num_to_words_inr(int(round(tot.get(net_col['key'], 0))))} (NET PAYABLE)"])
        s2.cell(row=s2.max_row, column=1).font = Font(bold=True, size=10)
    s2.append([])
    s2.append(["Checked by ____________________", f"For {comp.upper()}"])
    s2.append(["Payment Date ____________________", "AUTHORISED SIGNATORY / MANAGER"])
    for r_i in (s2.max_row - 1, s2.max_row):
        s2.cell(row=r_i, column=2).font = Font(bold=True, size=10)
        s2.cell(row=r_i, column=2).alignment = Alignment(horizontal="right")
    s2.column_dimensions["A"].width = 42
    s2.column_dimensions["B"].width = 34

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@router.get("/export.xlsx")
async def export_xlsx(
    source: str = "compliance",
    company_id: Optional[str] = None,
    month: Optional[str] = None,
    run_id: Optional[str] = None,
    employee_type: Optional[str] = None,
    branch: Optional[str] = None,
    department: Optional[str] = None,
    contractor: Optional[str] = None,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: str = "asc",
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    prep = await _prepare(
        admin, source, company_id, month, run_id, employee_type,
        branch, department, contractor, search, sort_by, sort_dir,
    )
    if not prep["run"]:
        raise HTTPException(status_code=404, detail="No salary run found for this month")
    comp = await _company_name(prep["company_id"])
    data = _xlsx_bytes(comp, source, month, prep["rows"], prep["columns"])
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="{_export_filename(source, month, "xlsx", employee_type)}"'},
    )


def _pdf_bytes(comp: str, source: str, month: str,
               rows: List[Dict[str, Any]], columns: List[Dict[str, Any]],
               filt_bits: List[str], title_override: str = "",
               layout: Optional[Dict[str, Any]] = None) -> bytes:
    """A3-landscape Salary Register PDF (shared by download + email).
    Iter 309 — honours the saved layout's employees-per-page + row height."""
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A3),
        leftMargin=8 * mm, rightMargin=8 * mm, topMargin=8 * mm, bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    title = Paragraph(
        f"<b>{comp}</b> — {title_override or 'Salary Register'} ({source.title()}) — {month}",
        styles["Title"],
    )
    sub_bits = list(filt_bits)
    sub_bits.append(f"Employees: {len(rows)}")
    sub = Paragraph(" | ".join(sub_bits), styles["Normal"])

    band = [""]
    for c in columns:
        band.append(next((g["label"] for g in GROUPS if g["key"] == c["group"]), ""))
    header = ["Sr"] + [c["label"] for c in columns]
    t = _totals(rows, columns)
    totals_row = ["", "TOTAL"] + [
        t.get(c["key"], "") if c["type"] != "text" else "" for c in columns[1:]
    ]

    def row_cells(i: int, r: Dict[str, Any]) -> List[Any]:
        return [i] + [
            _fmt_cell(r.get(c["key"]), c["type"]) if c["type"] != "text"
            else (str(r.get(c["key"]) or ""))[:22]
            for c in columns
        ]

    n_cols = len(columns) + 1
    font_size = 6.5 if n_cols <= 28 else (5.5 if n_cols <= 36 else 4.8)
    band_colors = {
        "info": "#44546A", "attendance": "#7F7F7F", "earnings": "#2E7D32",
        "deductions": "#B71C1C", "employer": "#6A1B9A", "net": "#1F4E79",
    }
    row_h = float((layout or {}).get("row_height") or 0)
    per_page = int((layout or {}).get("per_page") or 0)

    def make_table(chunk: List[Dict[str, Any]], start: int, with_totals: bool) -> Table:
        data: List[List[Any]] = [band, header]
        for off, r in enumerate(chunk):
            data.append(row_cells(start + off + 1, r))
        if with_totals:
            data.append(totals_row)
        heights = None
        if row_h > 0:
            heights = [None, None] + [row_h * mm] * len(chunk)
            if with_totals:
                heights.append(None)
        tbl = Table(data, repeatRows=2, rowHeights=heights)
        style = [
            ("FONTSIZE", (0, 0), (-1, -1), font_size),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (-1, 1), rl_colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 1), (-1, 1), rl_colors.white),
            ("GRID", (0, 1), (-1, -1), 0.25, rl_colors.HexColor("#C8C8C8")),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 2), (-1, -2 if with_totals else -1),
             [rl_colors.white, rl_colors.HexColor("#F2F6FA")]),
            ("TOPPADDING", (0, 0), (-1, -1), 1.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]
        if with_totals:
            style.append(("BACKGROUND", (0, -1), (-1, -1), rl_colors.HexColor("#FFF3CD")))
            style.append(("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"))
        ci = 1
        while ci <= len(columns):
            grp = columns[ci - 1]["group"]
            j = ci
            while j + 1 <= len(columns) and columns[j]["group"] == grp:
                j += 1
            style.append(("SPAN", (ci, 0), (j, 0)))
            style.append(("BACKGROUND", (ci, 0), (j, 0),
                          rl_colors.HexColor(band_colors.get(grp, "#1F4E79"))))
            style.append(("TEXTCOLOR", (ci, 0), (j, 0), rl_colors.white))
            style.append(("ALIGN", (ci, 0), (j, 0), "CENTER"))
            ci = j + 1
        for idx, c in enumerate(columns, 1):
            if c["type"] == "text":
                style.append(("ALIGN", (idx, 1), (idx, -1), "LEFT"))
        tbl.setStyle(TableStyle(style))
        return tbl

    flow: List[Any] = [title, sub, Spacer(1, 4 * mm)]
    if per_page > 0 and rows:
        for start in range(0, len(rows), per_page):
            chunk = rows[start:start + per_page]
            last = start + per_page >= len(rows)
            flow.append(make_table(chunk, start, last))
            if not last:
                flow.append(PageBreak())
    else:
        flow.append(make_table(rows, 0, True))
    flow.append(PageBreak())
    flow.extend(_summary_flowables(comp, rows, columns))
    doc.build(flow)
    return buf.getvalue()


def _summary_flowables(comp: str, rows: List[Dict[str, Any]],
                       columns: List[Dict[str, Any]]) -> List[Any]:
    """Iter 308 (user) — last-page summary in the same style as the
    Compliance Salary Register (PDF Option 2): boxed sections with
    head-wise totals, days/net, amounts in words + signature strip."""
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from utils.salary_register_pdf import _num_to_words_inr

    lbl = ParagraphStyle("srs_lbl", fontName="Helvetica", fontSize=8.5, leading=11)
    lblb = ParagraphStyle("srs_lblb", fontName="Helvetica-Bold", fontSize=8.5, leading=11)

    def amt(v: Any) -> str:
        return f"{float(v or 0):,.2f}"

    tot = _totals(rows, columns)

    def sec(pairs: List[Tuple[str, str]], bold_last: bool = True) -> Table:
        d = [[Paragraph(k, lblb if (bold_last and i == len(pairs) - 1) else lbl),
              Paragraph(v, lblb if (bold_last and i == len(pairs) - 1) else lbl)]
             for i, (k, v) in enumerate(pairs)]
        t = Table(d, colWidths=[70 * mm, 36 * mm])
        t.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, rl_colors.black),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#999999")),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 1.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ]))
        return t

    earn_cols = [c for c in columns if c["group"] == "earnings"]
    ded_cols = [c for c in columns if c["group"] == "deductions"]
    er_cols = [c for c in columns if c["group"] == "employer"]
    net_col = next((c for c in columns if c["group"] == "net"), None)

    gross_key = next((k for k in ("gross_paid", "total_gross", "monthly_gross")
                      if k in tot), None)
    days_key = next((k for k in ("present_days", "p_days") if k in tot), None)
    hrs_key = next((k for k in ("ot_hours", "p_hours") if k in tot), None)

    def head_label(c: Dict[str, Any]) -> str:
        lbl_txt = str(c["label"])
        return lbl_txt if lbl_txt.lower().startswith("total") else f"Total {lbl_txt}"

    out: List[Any] = []
    out.append(Paragraph(f"SUMMARY — {comp}", ParagraphStyle(
        "srs_h", fontName="Helvetica-Bold", fontSize=11, leading=14)))
    out.append(Spacer(1, 3 * mm))
    out.append(sec(
        [("No. Of Emp", str(len(rows)))]
        + [(head_label(c), amt(tot.get(c["key"]))) for c in earn_cols],
    ))
    if ded_cols:
        out.append(Spacer(1, 4 * mm))
        out.append(sec([(head_label(c), amt(tot.get(c["key"])))
                        for c in ded_cols]))
    if er_cols:
        out.append(Spacer(1, 4 * mm))
        out.append(sec([(head_label(c), amt(tot.get(c["key"])))
                        for c in er_cols], bold_last=False))
    tail: List[Tuple[str, str]] = []
    if days_key:
        tail.append(("Total Days ->", f"{tot.get(days_key, 0):g}"))
    if hrs_key:
        tail.append(("Total Hours ->", f"{tot.get(hrs_key, 0):g}"))
    if net_col:
        tail.append(("Net Payable Amount", amt(tot.get(net_col["key"]))))
    if tail:
        out.append(Spacer(1, 4 * mm))
        out.append(sec(tail))
    out.append(Spacer(1, 5 * mm))
    if gross_key:
        out.append(Paragraph(
            f"RUPEES: {_num_to_words_inr(int(round(tot.get(gross_key, 0))))} (GROSS)", lblb))
    if net_col:
        out.append(Paragraph(
            f"RUPEES: {_num_to_words_inr(int(round(tot.get(net_col['key'], 0))))} (NET PAYABLE)", lblb))
    out.append(Spacer(1, 10 * mm))
    foot = Table([
        [Paragraph("Checked by ____________________", lbl),
         Paragraph(f"For {comp.upper()}", lblb)],
        [Paragraph("Payment Date ____________________", lbl),
         Paragraph("AUTHORISED SIGNATORY / MANAGER", lblb)],
    ], colWidths=[150 * mm, 150 * mm])
    foot.setStyle(TableStyle([
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 1), (-1, 1), 14),
    ]))
    out.append(foot)
    return out


def _filt_bits(employee_type: Optional[str], branch: Optional[str],
               department: Optional[str], contractor: Optional[str]) -> List[str]:
    bits = []
    if employee_type:
        bits.append(f"Group: {employee_type}")
    if branch:
        bits.append(f"Branch: {branch}")
    if department:
        bits.append(f"Dept: {department}")
    if contractor:
        bits.append(f"Contractor: {contractor}")
    return bits


@router.get("/export.pdf")
async def export_pdf(
    source: str = "compliance",
    company_id: Optional[str] = None,
    month: Optional[str] = None,
    run_id: Optional[str] = None,
    employee_type: Optional[str] = None,
    branch: Optional[str] = None,
    department: Optional[str] = None,
    contractor: Optional[str] = None,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: str = "asc",
    authorization: Optional[str] = Header(None),
):
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    prep = await _prepare(
        admin, source, company_id, month, run_id, employee_type,
        branch, department, contractor, search, sort_by, sort_dir,
    )
    if not prep["run"]:
        raise HTTPException(status_code=404, detail="No salary run found for this month")
    comp = await _company_name(prep["company_id"])
    title_ov = await _module_title()
    data = _pdf_bytes(comp, source, month, prep["rows"], prep["columns"],
                      _filt_bits(employee_type, branch, department, contractor),
                      title_override=title_ov, layout=prep.get("layout"))
    return StreamingResponse(
        io.BytesIO(data), media_type="application/pdf",
        headers={"Content-Disposition":
                 f'attachment; filename="{_export_filename(source, month, "pdf", employee_type)}"'},
    )


async def _module_title() -> str:
    """Saved title from Reports → PDF Report Formats → Salary Register
    (Dynamic Module)."""
    try:
        from routes.report_formats import get_report_format
        return str((await get_report_format("salary_register_module")).get("title") or "").strip()
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# Iter 307 — one-click "Email register to firm" (Resend, with attachments).
# ---------------------------------------------------------------------------

@router.post("/email")
async def email_register(
    payload: Dict[str, Any] = Body(default={}),
    authorization: Optional[str] = Header(None),
):
    """Email the filtered register (PDF + Excel) to the firm.

    Body: same filter keys as the grid (source, company_id, month, run_id,
    employee_type, branch, department, contractor, search) plus:
      * ``to``      — recipient email (defaults to the firm's email)
      * ``formats`` — subset of ["pdf", "xlsx"] (default both)
    """
    admin = await get_user_from_token(authorization)
    require_role(admin, ["super_admin", "company_admin", "sub_admin"])
    source = str(payload.get("source") or "compliance")
    month = payload.get("month")
    employee_type = payload.get("employee_type") or None
    branch = payload.get("branch") or None
    department = payload.get("department") or None
    contractor = payload.get("contractor") or None
    prep = await _prepare(
        admin, source, payload.get("company_id"), month, payload.get("run_id") or None,
        employee_type, branch, department, contractor,
        payload.get("search") or None, payload.get("sort_by") or None,
        str(payload.get("sort_dir") or "asc"),
    )
    if not prep["run"]:
        raise HTTPException(status_code=404, detail="No salary run found for this month")
    rows, columns = prep["rows"], prep["columns"]
    comp = await _company_name(prep["company_id"])

    to = str(payload.get("to") or "").strip()
    if not to:
        to = await _firm_email(prep["company_id"]) or ""
    if not to or "@" not in to:
        raise HTTPException(
            status_code=400,
            detail="No recipient email — enter one or add the firm's email in Firm Master (Header → Email 1)")

    formats = payload.get("formats") or ["pdf", "xlsx"]
    attachments: List[Dict[str, str]] = []
    if "pdf" in formats:
        pdf = _pdf_bytes(comp, source, month, rows, columns,
                         _filt_bits(employee_type, branch, department, contractor),
                         title_override=await _module_title(),
                         layout=prep.get("layout"))
        attachments.append({
            "filename": _export_filename(source, month, "pdf", employee_type),
            "content": base64.b64encode(pdf).decode(),
        })
    if "xlsx" in formats:
        xlsx = _xlsx_bytes(comp, source, month, rows, columns)
        attachments.append({
            "filename": _export_filename(source, month, "xlsx", employee_type),
            "content": base64.b64encode(xlsx).decode(),
        })
    if not attachments:
        raise HTTPException(status_code=400, detail="Pick at least one format (pdf / xlsx)")

    t = _totals(rows, columns)
    gross_key = "gross_paid" if source == "compliance" else "total_gross"
    net_key = "net" if source == "compliance" else "net_pay"
    summary = (
        f"Salary Register — {comp}\n"
        f"Month: {month}  ·  Engine: {source.title()}\n"
        f"Employees: {len(rows)}\n"
        f"Gross: ₹{t.get(gross_key, 0):,.2f}\n"
        f"Net Payable: ₹{t.get(net_key, 0):,.2f}\n\n"
        "The detailed register is attached (PDF + Excel).\n\n"
        "— Sent from the S.K. Sharma & Co. portal"
    )
    from utils.iter60_features import _send_email_with_attachment
    result = await _send_email_with_attachment(
        [to],
        subject=f"Salary Register · {comp} · {month} ({source.title()})",
        text_body=summary,
        attachments=attachments,
    )
    if not result.get("delivered"):
        raise HTTPException(
            status_code=502,
            detail=f"Email could not be sent ({result.get('error') or 'unknown error'})")
    return {"ok": True, "to": to, "email_id": result.get("email_id"),
            "attachments": [a["filename"] for a in attachments]}


async def _firm_email(company_id: str) -> Optional[str]:
    fm = await db.firm_masters.find_one(
        {"company_id": company_id}, {"_id": 0, "header.email_1": 1, "header.email_2": 1})
    hdr = (fm or {}).get("header") or {}
    email = (hdr.get("email_1") or hdr.get("email_2") or "").strip()
    if email:
        return email
    c = await db.companies.find_one({"company_id": company_id},
                                    {"_id": 0, "email": 1, "contact_email": 1})
    return ((c or {}).get("email") or (c or {}).get("contact_email") or "").strip() or None
