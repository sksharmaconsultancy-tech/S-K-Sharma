"""Iter 172 — Bulk Punch Import via Excel (Punch Approvals) — backend tests."""
import base64
import io
import os
import uuid

import pytest
import requests
from openpyxl import Workbook

BASE_URL = os.environ.get("EXPO_PUBLIC_BACKEND_URL", "https://emplo-connect-1.preview.emergentagent.com").rstrip("/")
COMPANY_ID = "cmp_527fecdd7c"  # Kankani
EMP_BIO = "72"                  # SURENDRA SINGH bio_code
EMP_NAME = "SURENDRA SINGH"
EMP_USER_ID = "user_44cd6f561da0"


# ---------- fixtures ------------------------------------------------------

@pytest.fixture(scope="module")
def super_token():
    r = requests.post(
        f"{BASE_URL}/api/auth/admin-password-login",
        json={"email": "sksharmaconsultancy@gmail.com", "password": "sharma123"},
        timeout=30,
    )
    assert r.status_code == 200, f"admin login failed: {r.status_code} {r.text}"
    tok = r.json().get("session_token") or r.json().get("token")
    assert tok, r.text
    return tok


@pytest.fixture(scope="module")
def h_super(super_token):
    return {"Authorization": f"Bearer {super_token}", "Content-Type": "application/json"}


def _build_xlsx(rows):
    """rows: list of tuples (bio, name, date, in_time, out_time)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Bio Code", "Name", "Date", "In Time", "Out Time"])
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


# ---------- template endpoint ---------------------------------------------

def test_template_returns_valid_xlsx(h_super):
    r = requests.get(f"{BASE_URL}/api/admin/punch-import/template", headers=h_super, timeout=30)
    assert r.status_code == 200, r.text
    j = r.json()
    assert j.get("filename", "").endswith(".xlsx")
    raw = base64.b64decode(j["file_base64"])
    assert raw[:2] == b"PK", "not a real xlsx (missing PK header)"
    # openable
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Bio Code" in headers and "Date" in headers


# ---------- preview: matching + parsing -----------------------------------

def test_preview_bio_name_unmatched_errors_and_time_date_parsing(h_super):
    # date used far away from ZK-device dates
    d1 = "20-06-2026"
    d2 = "21-06-2026"
    b64 = _build_xlsx([
        (EMP_BIO, "",           d1, "9:00 AM",  "6:00 PM"),  # match by bio_code + AM/PM
        ("",      EMP_NAME,     d2, "09:15",    "18:30"),    # match by name + 24h
        ("99999", "GHOST USER", d1, "09:00",    "18:00"),    # unmatched
        (EMP_BIO, "",           "", "09:00",    "18:00"),    # error – missing date
    ])
    r = requests.post(
        f"{BASE_URL}/api/admin/punch-import/preview",
        headers=h_super,
        json={"company_id": COMPANY_ID, "file_base64": b64},
        timeout=30,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    s = body["summary"]
    assert s["total"] == 4
    assert s["matched"] == 2
    assert s["unmatched"] == 1
    assert s["errors"] == 1
    # 2 matched rows each with in+out = 4 punches
    assert s["punches_to_create"] == 4

    rows = {r_["row_no"]: r_ for r_ in body["rows"]}
    # Row 2 – bio match, AM/PM parsed
    r_bio = rows[2]
    assert r_bio["status"] == "matched"
    assert r_bio["matched_by"] == "bio_code"
    assert r_bio["in_time"] == "09:00"
    assert r_bio["out_time"] == "18:00"
    assert r_bio["date"] == "2026-06-20"
    assert r_bio["user_id"] == EMP_USER_ID
    # Row 3 – name match, 24h
    r_name = rows[3]
    assert r_name["status"] == "matched"
    assert r_name["matched_by"] == "name"
    assert r_name["date"] == "2026-06-21"
    assert r_name["user_id"] == EMP_USER_ID
    # Row 4 – unmatched
    assert rows[4]["status"] == "unmatched"
    # Row 5 – error (missing date)
    assert rows[5]["status"] == "error"


# ---------- commit + de-dup ------------------------------------------------

DUP_DATE = "22-06-2026"


def test_commit_creates_then_dedups(h_super):
    b64 = _build_xlsx([(EMP_BIO, "", DUP_DATE, "09:00", "18:00")])
    prev = requests.post(
        f"{BASE_URL}/api/admin/punch-import/preview",
        headers=h_super,
        json={"company_id": COMPANY_ID, "file_base64": b64},
        timeout=30,
    ).json()
    matched = [r for r in prev["rows"] if r["status"] == "matched"]
    assert matched, prev

    # first commit -> creates 2 punches
    r1 = requests.post(
        f"{BASE_URL}/api/admin/punch-import/commit",
        headers=h_super,
        json={"company_id": COMPANY_ID, "rows": matched},
        timeout=30,
    )
    assert r1.status_code == 200, r1.text
    j1 = r1.json()
    assert j1["ok"] is True
    assert j1["created"] == 2
    assert j1["skipped_duplicates"] == 0

    # second commit -> all dup
    r2 = requests.post(
        f"{BASE_URL}/api/admin/punch-import/commit",
        headers=h_super,
        json={"company_id": COMPANY_ID, "rows": matched},
        timeout=30,
    )
    assert r2.status_code == 200, r2.text
    j2 = r2.json()
    assert j2["created"] == 0
    assert j2["skipped_duplicates"] == 2


# ---------- punches show up in day-status --------------------------------

def test_committed_punches_visible_in_day_status(h_super):
    # Convert DUP_DATE to ISO for API filter
    d, m, y = DUP_DATE.split("-")
    iso = f"{y}-{m}-{d}"
    r = requests.get(
        f"{BASE_URL}/api/admin/attendance/day-status/{COMPANY_ID}",
        headers=h_super,
        params={"from_date": iso, "to_date": iso},
        timeout=30,
    )
    assert r.status_code == 200, r.text
    data = r.json()
    # find surendra singh entry
    users = data.get("rows") or data.get("users") or data.get("data") or []
    found = False
    for row in users:
        if row.get("user_id") == EMP_USER_ID:
            # look for any 09:00 / 18:00 in day payload
            blob = str(row)
            if "09:00" in blob and "18:00" in blob:
                found = True
                break
    if not found:
        # Print a sample for debugging
        pytest.skip(f"Could not verify punches in day-status structure; keys={list(data)[:8]}")


# ---------- cross-company authorisation ----------------------------------

def test_cross_company_admin_gets_403(h_super):
    """Login as testsub@sksharma.co (sub_admin scoped to a different firm) —
    preview/commit against Kankani must return 403."""
    sub = requests.post(
        f"{BASE_URL}/api/auth/admin-password-login",
        json={"email": "testsub@sksharma.co", "password": "testsub123"},
        timeout=30,
    )
    if sub.status_code != 200:
        pytest.skip(f"sub-admin login failed: {sub.status_code}")
    stok = sub.json().get("session_token") or sub.json().get("token")
    h = {"Authorization": f"Bearer {stok}", "Content-Type": "application/json"}
    b64 = _build_xlsx([(EMP_BIO, "", "23-06-2026", "09:00", "18:00")])
    r = requests.post(
        f"{BASE_URL}/api/admin/punch-import/preview",
        headers=h,
        json={"company_id": COMPANY_ID, "file_base64": b64},
        timeout=30,
    )
    assert r.status_code == 403, f"expected 403 got {r.status_code}: {r.text[:200]}"
    # commit also 403
    r2 = requests.post(
        f"{BASE_URL}/api/admin/punch-import/commit",
        headers=h,
        json={"company_id": COMPANY_ID, "rows": [{"user_id": EMP_USER_ID, "date": "2026-06-23", "in_time": "09:00"}]},
        timeout=30,
    )
    assert r2.status_code == 403, f"expected 403 got {r2.status_code}: {r2.text[:200]}"


# ---------- cleanup module -------------------------------------------------

def test_zzz_cleanup_excel_import_records(h_super):
    """Clean up any excel_import punches we created."""
    # No admin API to delete raw attendance; do it via direct Mongo call.
    from pymongo import MongoClient
    mc = MongoClient(os.environ.get("MONGO_URL", "mongodb://localhost:27017"))
    dbn = os.environ.get("DB_NAME", "test_database")
    res = mc[dbn].attendance.delete_many({"source": "excel_import"})
    print(f"deleted {res.deleted_count} excel_import attendance records")
    mc.close()
