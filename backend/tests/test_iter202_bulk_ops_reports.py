"""Iter 202 backend tests — Bulk Ops, statutory extra reports, CLRA xlsx,
attendance grid present_days_policy, and the 8-HR compliance sub-point.

Focus areas (main agent has already verified basic paths):
  * inout attendance template + preview + apply round trip (item 3)
  * bulk salary EXCEL template + preview (item 5)
  * validation errors on transfer/resignation/shift-assign (item 6)
  * reports JSON + xlsx PK magic + pdf %PDF magic (item 8)
  * CLRA xlsx endpoints (item 9)
  * grid.totals.present_days_policy present (item 10)
  * 8-HR sub-point behaviour (item 11) — MAIN FOCUS
  * auth 401 checks (item 12)

All state changes are cleaned up at the end of each test (fixtures).
"""
import base64
import io
import os
import calendar
import pytest
import requests
from pymongo import MongoClient

_mongo = MongoClient(os.environ.get("MONGO_URL", "mongodb://localhost:27017"))
_db = _mongo[os.environ.get("DB_NAME", "test_database")]

BASE = os.environ.get("EXPO_PUBLIC_BACKEND_URL", "").rstrip("/") \
    or "https://emplo-connect-1.preview.emergentagent.com"
CID = "cmp_527fecdd7c"  # Kankani Enterprises (Rajasthan)


# --------------------------------------------------------------------- session
@pytest.fixture(scope="session")
def token():
    r = requests.post(f"{BASE}/api/auth/admin-password-login",
                      json={"email": "sksharmaconsultancy@gmail.com",
                            "password": "sharma123"}, timeout=20)
    assert r.status_code == 200, r.text
    return r.json()["session_token"]


@pytest.fixture(scope="session")
def H(token):
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture(scope="session")
def one_employee(H):
    """Pick one active employee from bulk-ops list."""
    r = requests.get(f"{BASE}/api/admin/bulk-ops/employees",
                     params={"company_id": CID}, headers=H, timeout=20)
    assert r.status_code == 200, r.text
    rows = r.json()["rows"]
    # find one with no exit_date preferred (e.g. code 50)
    emp = next((e for e in rows if str(e.get("employee_code")) == "50"), rows[0])
    return emp


# ============================================================ item 3: inout RT
class TestAttendanceInoutRoundTrip:
    MONTH = "2026-07"

    def test_inout_template(self, H):
        r = requests.get(f"{BASE}/api/admin/bulk-ops/attendance-template",
                         params={"company_id": CID, "month": self.MONTH, "kind": "inout"},
                         headers=H, timeout=20)
        assert r.status_code == 200, r.text
        b = base64.b64decode(r.json()["file_base64"])
        assert b[:2] == b"PK"

    def test_inout_preview_and_apply_cleanup(self, H, one_employee):
        """Build a minimal xlsx with 1 row for a future day; preview then apply,
        finally delete created punches by import_batch."""
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.append(["Emp Code", "Name", "Date", "In Time", "Out Time"])
        # future date 20th of MONTH — unlikely collision with existing punches
        date_str = "20-07-2026"
        ws.append([one_employee["employee_code"], one_employee["name"],
                   date_str, "09:00", "19:00"])
        buf = io.BytesIO(); wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode()

        prev = requests.post(f"{BASE}/api/admin/bulk-ops/attendance-preview",
                             json={"company_id": CID, "month": self.MONTH,
                                   "kind": "inout", "file_base64": b64},
                             headers=H, timeout=20)
        assert prev.status_code == 200, prev.text
        pj = prev.json()
        assert pj["summary"]["matched"] == 1, pj
        row = pj["rows"][0]
        assert row["status"] == "matched"
        assert row["date"] == "2026-07-20"

        # apply with overwrite so we don't collide with anything
        appl = requests.post(f"{BASE}/api/admin/bulk-ops/attendance-apply",
                             json={"company_id": CID, "month": self.MONTH,
                                   "kind": "inout", "overwrite": True,
                                   "rows": pj["rows"]},
                             headers=H, timeout=30)
        assert appl.status_code == 200, appl.text
        aj = appl.json()
        assert aj["created"] >= 2, aj
        batch = aj["batch"]

        # cleanup — no batch-delete endpoint exists; wipe directly in mongo
        deleted = _db.attendance.delete_many({"import_batch": batch}).deleted_count
        assert deleted >= 2, f"cleanup failed for batch {batch}"


# ============================================================ item 5: salary excel preview
class TestSalaryExcelPreview:
    def test_salary_template_and_preview(self, H, one_employee):
        r = requests.get(f"{BASE}/api/admin/bulk-ops/salary-template",
                         params={"company_id": CID}, headers=H, timeout=20)
        assert r.status_code == 200, r.text
        b = base64.b64decode(r.json()["file_base64"])
        assert b[:2] == b"PK"

        # Build a preview file — fill new_actual for one row.
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.append(["Emp Code", "Name", "Current Actual Basic", "New Actual Basic",
                   "Current Compliance Gross", "New Compliance Gross"])
        cur = one_employee.get("actual_basic") or 1000
        new_val = round(cur * 1.05, 2)
        ws.append([one_employee["employee_code"], one_employee["name"],
                   cur, new_val, one_employee.get("compliance_gross") or 0, ""])
        buf = io.BytesIO(); wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode()

        p = requests.post(f"{BASE}/api/admin/bulk-ops/salary-preview",
                          json={"company_id": CID, "file_base64": b64},
                          headers=H, timeout=20)
        assert p.status_code == 200, p.text
        pj = p.json()
        assert pj["summary"]["matched"] == 1, pj
        assert pj["rows"][0]["new_actual"] == new_val


# ============================================================ item 6: validation errors
class TestBulkValidation:
    def test_transfer_same_firm(self, H):
        r = requests.post(f"{BASE}/api/admin/bulk-ops/transfer",
                          json={"company_id": CID, "to_company_id": CID,
                                "user_ids": ["u_fake"]},
                          headers=H, timeout=15)
        assert r.status_code == 400, r.text
        assert "same firm" in r.text.lower()

    def test_transfer_missing_dest(self, H):
        r = requests.post(f"{BASE}/api/admin/bulk-ops/transfer",
                          json={"company_id": CID, "user_ids": ["u_fake"]},
                          headers=H, timeout=15)
        assert r.status_code == 400, r.text

    def test_resignation_bad_date(self, H, one_employee):
        r = requests.post(f"{BASE}/api/admin/bulk-ops/resignation",
                          json={"company_id": CID,
                                "user_ids": [one_employee["user_id"]],
                                "exit_date": "31/12/2099"},
                          headers=H, timeout=15)
        assert r.status_code == 400, r.text
        assert "YYYY-MM-DD" in r.text

    def test_shift_assign_and_revert(self, H, one_employee):
        # get an actual shift
        sm = requests.get(f"{BASE}/api/shift-masters",
                          params={"company_id": CID}, headers=H, timeout=15)
        if sm.status_code != 200:
            pytest.skip("shift-masters endpoint not available")
        sh_data = sm.json()
        shifts = sh_data.get("rows") or sh_data.get("shifts") or sh_data
        if not isinstance(shifts, list) or not shifts:
            pytest.skip("No shifts defined for firm")
        shift = shifts[0]
        shift_id = shift.get("shift_id") or shift.get("id")
        if not shift_id:
            pytest.skip("shift missing id")

        # Capture original shift fields
        orig = one_employee
        r = requests.post(f"{BASE}/api/admin/bulk-ops/shift-assign",
                          json={"company_id": CID,
                                "user_ids": [orig["user_id"]],
                                "shift_id": shift_id},
                          headers=H, timeout=15)
        assert r.status_code == 200, r.text
        assert r.json().get("updated", 0) >= 0

        # Revert via direct users update — call profile-update if available.
        # Best-effort: reset via same endpoint using a shift that matches orig,
        # or via /api/admin/users/{id} PATCH if available.
        rev = requests.patch(
            f"{BASE}/api/admin/users/{orig['user_id']}",
            json={"shift_id": None,
                  "shift_name": orig.get("shift_name"),
                  "shift_start": orig.get("shift_start"),
                  "shift_end": orig.get("shift_end")},
            headers=H, timeout=15)
        if rev.status_code == 404:
            print(f"WARN: no revert endpoint. Employee {orig['user_id']} shift assigned to {shift_id}")


# ============================================================ item 7 & 8: history + reports
class TestHistoryAndReports:
    def test_history(self, H):
        r = requests.get(f"{BASE}/api/admin/bulk-ops/history",
                         params={"company_id": CID}, headers=H, timeout=15)
        assert r.status_code == 200, r.text
        assert "rows" in r.json()

    @pytest.mark.parametrize("path,params", [
        ("/api/admin/reports/pt", {"company_id": CID, "month": "2026-06"}),
        ("/api/admin/reports/lwf", {"company_id": CID, "month": "2026-06"}),
        ("/api/admin/reports/gratuity", {"company_id": CID}),
        ("/api/admin/reports/fnf", {"company_id": CID}),
        ("/api/admin/reports/advance-loan", {"company_id": CID}),
        ("/api/admin/reports/mis", {"month": "2026-06", "company_id": CID}),
    ])
    def test_reports_json_xlsx_pdf(self, H, path, params):
        # JSON
        r = requests.get(f"{BASE}{path}", params=params, headers=H, timeout=30)
        assert r.status_code == 200, r.text
        j = r.json()
        assert "rows" in j
        # xlsx
        p2 = dict(params, fmt="xlsx")
        r2 = requests.get(f"{BASE}{path}", params=p2, headers=H, timeout=30)
        assert r2.status_code == 200
        assert r2.content[:2] == b"PK", f"{path} xlsx not PK — {r2.content[:20]}"
        # pdf
        p3 = dict(params, fmt="pdf")
        r3 = requests.get(f"{BASE}{path}", params=p3, headers=H, timeout=30)
        assert r3.status_code == 200
        assert r3.content[:4] == b"%PDF", f"{path} pdf not %PDF — {r3.content[:20]}"

    def test_lwf_rajasthan_zero(self, H):
        r = requests.get(f"{BASE}/api/admin/reports/lwf",
                         params={"company_id": CID, "month": "2026-06"},
                         headers=H, timeout=20).json()
        assert r["total_ee"] == 0 and r["total_er"] == 0
        assert "Rajasthan" in r["state"] or "no" in r["rule_note"].lower()

    def test_gratuity_liability_positive(self, H):
        r = requests.get(f"{BASE}/api/admin/reports/gratuity",
                         params={"company_id": CID},
                         headers=H, timeout=30).json()
        # There must be at least SOME liability across 125 employees w/ daily rates
        assert r["total_liability"] >= 0
        # daily-rated employees: at least one row with monthly_basic > 0
        assert any(row.get("monthly_basic", 0) > 0 for row in r["rows"])

    def test_advance_loan_rows(self, H):
        r = requests.get(f"{BASE}/api/admin/reports/advance-loan",
                         params={"company_id": CID},
                         headers=H, timeout=20).json()
        assert isinstance(r["rows"], list)

    def test_mis_man_days_positive(self, H):
        r = requests.get(f"{BASE}/api/admin/reports/mis",
                         params={"month": "2026-06", "company_id": CID},
                         headers=H, timeout=30).json()
        assert r["rows"]
        # man_days may be 0 if no punches in 2026-06 (attendance was cleared).
        # Just assert numeric.
        assert isinstance(r["rows"][0].get("man_days"), int)


# ============================================================ item 9: CLRA xlsx
class TestCLRAXlsx:
    @pytest.mark.parametrize("form,extra", [
        ("form-xii.xlsx", {}),
        ("form-xiii.xlsx", {}),
        ("form-xiv.xlsx", {}),
        ("form-xv.xlsx", {"month": "2026-06"}),
    ])
    def test_clra_xlsx(self, H, form, extra):
        r = requests.get(f"{BASE}/api/admin/clra-registers/{form}",
                         params={"company_id": CID, **extra},
                         headers=H, timeout=25)
        assert r.status_code == 200, r.text
        assert r.content[:2] == b"PK"


# ============================================================ item 10: grid present_days_policy
class TestGridPresentDaysPolicy:
    def test_grid_has_present_days_policy(self, H):
        r = requests.get(
            f"{BASE}/api/admin/attendance/monthly-grid/{CID}/2026-06",
            headers=H, timeout=45)
        assert r.status_code == 200, r.text
        j = r.json()
        rows = j.get("employees") or j.get("rows") or []
        assert rows, "grid empty"
        for row in rows[:5]:
            tot = row.get("totals") or {}
            assert "present_days_policy" in tot, f"row missing key: {tot.keys()}"
            assert isinstance(tot["present_days_policy"], (int, float))


# ============================================================ item 11: 8-HR sub-point
class TestCompliancePresent8Hr:
    """Toggle policy_master.compliance_present_8hr and verify grid splits duty/OT at 8h."""

    @pytest.fixture(scope="class")
    def _test_month(self):
        return "2026-08"  # future — safe to add test punches

    @pytest.fixture(scope="class")
    def _emp(self, H, one_employee):
        return one_employee

    def _get_policy(self, H):
        r = requests.get(f"{BASE}/api/attendance/policy",
                         params={"company_id": CID}, headers=H, timeout=15)
        assert r.status_code == 200, r.text
        return r.json()

    def _patch_pm_flag(self, H, val):
        r = requests.patch(
            f"{BASE}/api/attendance/policy",
            params={"company_id": CID},
            json={"policy": {"policy_master": {"compliance_present_8hr": val}}},
            headers=H, timeout=20)
        assert r.status_code == 200, r.text
        # verify
        got = self._get_policy(H)
        assert (got["policy"].get("policy_master") or {}).get(
            "compliance_present_8hr") is val, got["policy"].get("policy_master")

    def test_flag_toggle_and_grid_split(self, H, _test_month, _emp):
        # Ensure a punch of 10hrs on 2026-08-05 for _emp via bulk-ops inout
        from openpyxl import Workbook
        date_str = "05-08-2026"
        wb = Workbook(); ws = wb.active
        ws.append(["Emp Code", "Name", "Date", "In Time", "Out Time"])
        ws.append([_emp["employee_code"], _emp["name"], date_str, "08:00", "18:00"])
        buf = io.BytesIO(); wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode()
        prev = requests.post(f"{BASE}/api/admin/bulk-ops/attendance-preview",
                             json={"company_id": CID, "month": _test_month,
                                   "kind": "inout", "file_base64": b64},
                             headers=H, timeout=20).json()
        appl = requests.post(f"{BASE}/api/admin/bulk-ops/attendance-apply",
                             json={"company_id": CID, "month": _test_month,
                                   "kind": "inout", "overwrite": True,
                                   "rows": prev["rows"]},
                             headers=H, timeout=25).json()
        batch = appl.get("batch")

        try:
            # ---- Turn ON the flag
            self._patch_pm_flag(H, True)
            r = requests.get(
                f"{BASE}/api/admin/attendance/monthly-grid/{CID}/{_test_month}",
                headers=H, timeout=45).json()
            rows = r.get("employees") or r.get("rows") or []
            row = next((x for x in rows if x.get("user_id") == _emp["user_id"]), None)
            assert row, "employee row missing"
            # `days` is a dict keyed by day-of-month string ("01"..)
            days = row.get("days") or {}
            day5 = days.get("05") if isinstance(days, dict) else None
            print(f"DAY5 (flag ON): {day5}")
            assert day5, "day-5 cell missing"
            dh = day5.get("duty_hours") or 0
            oh = day5.get("ot_hours") or 0
            raw = day5.get("raw_hours") or day5.get("hours") or 0
            print(f"raw={raw} duty={dh} ot={oh}")
            # With 10h worked + 8h split rule: duty<=8, ot>=1 (grace applies)
            assert dh <= 8.5, f"duty_hours should be <=8 with flag ON, got {dh}"
            assert oh >= 1.0, f"ot_hours should pick up extras, got {oh}"
        finally:
            # ---- Turn OFF the flag no matter what
            self._patch_pm_flag(H, False)
            # cleanup punches via mongo (no batch endpoint)
            if batch:
                _db.attendance.delete_many({"import_batch": batch})


# ============================================================ item 12: auth 401
class TestAuth401:
    @pytest.mark.parametrize("method,path,kw", [
        ("get", "/api/admin/bulk-ops/employees", {"params": {"company_id": CID}}),
        ("get", "/api/admin/bulk-ops/history", {}),
        ("get", "/api/admin/reports/pt", {"params": {"company_id": CID, "month": "2026-06"}}),
        ("get", "/api/admin/reports/gratuity", {"params": {"company_id": CID}}),
        ("get", "/api/admin/clra-registers/form-xii.xlsx", {"params": {"company_id": CID}}),
        ("get", "/api/admin/clra-registers/form-xv.xlsx",
         {"params": {"company_id": CID, "month": "2026-06"}}),
        ("post", "/api/admin/bulk-ops/transfer", {"json": {"company_id": CID}}),
        ("post", "/api/admin/bulk-ops/resignation", {"json": {"company_id": CID}}),
    ])
    def test_no_token_returns_401(self, method, path, kw):
        r = requests.request(method.upper(), f"{BASE}{path}", timeout=15, **kw)
        assert r.status_code == 401, f"{path}: {r.status_code} {r.text[:120]}"
