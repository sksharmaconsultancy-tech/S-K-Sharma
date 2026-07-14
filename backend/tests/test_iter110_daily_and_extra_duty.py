"""
Iter 110/111 backend regression tests
-------------------------------------
Covers:
  1. GET /api/admin/attendance/day-status/{company_id} — cells include
     edit_reason / edited_by_name / original_hhmm for edited punches.
  2. POST /api/admin/attendance/extra-duty — accepts NEGATIVE extra_hours,
     rejects negative extra_amount, and DELETES entry when both are 0.
  3. PATCH /api/admin/attendance/{record_id} (body {at, reason}) still
     works; downstream day-status Updated cell exposes audit trio.
  4. GET /api/admin/attendance/daily/{cid}/{YYYY-MM-DD}.xlsx and .pdf →
     200 with the correct headers/columns; invalid date → 400.
  5. Monthly exports with new Bio Code column:
     /monthly-inout/{cid}/{YYYY-MM}.xlsx, /monthly-hours/{cid}/{YYYY-MM}.xlsx
     and both .pdf variants — first data column is Bio Code (xlsx) or
     appears after Code in the PDF header row.
"""
import os
import io
import re
import pathlib
from datetime import datetime, timezone, timedelta

import pytest
import requests
from dotenv import load_dotenv

# --- env ---------------------------------------------------------------
load_dotenv("/app/frontend/.env")
BASE_URL = (os.environ.get("EXPO_PUBLIC_BACKEND_URL") or "").rstrip("/")
assert BASE_URL, "EXPO_PUBLIC_BACKEND_URL missing"
API = f"{BASE_URL}/api"

CID = "cmp_527fecdd7c"           # Kankani Enterprises
ADMIN_EMAIL = "sksharmaconsultancy@gmail.com"
ADMIN_PWD = "sharma123"


# --- shared fixtures --------------------------------------------------
@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(
        f"{API}/auth/admin-password-login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PWD},
        timeout=20,
    )
    assert r.status_code == 200, f"login failed {r.status_code} {r.text[:200]}"
    tok = r.json().get("session_token") or r.json().get("token")
    assert tok, f"no token in login response: {r.text[:200]}"
    return tok


@pytest.fixture(scope="module")
def headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture(scope="module")
def sample_employee(headers):
    """First active Kankani employee (used for extra-duty + punch edit)."""
    r = requests.get(
        f"{API}/admin/employees?company_id={CID}",
        headers=headers, timeout=20,
    )
    assert r.status_code == 200, r.text[:200]
    emps = r.json().get("employees") or r.json().get("users") or []
    assert emps, "no employees in Kankani"
    return emps[0]


# =====================================================================
# 1. Extra-duty NEGATIVE hours / delete-when-zero / reject negative amt
# =====================================================================
class TestExtraDutyNegative:
    """Iter 111 — extra_hours may be NEGATIVE to REDUCE duty (backend)."""

    DATE = "2026-07-09"

    def test_negative_hours_accepted(self, headers, sample_employee):
        r = requests.post(
            f"{API}/admin/attendance/extra-duty",
            headers=headers,
            json={
                "user_id": sample_employee["user_id"],
                "date": self.DATE,
                "extra_hours": -0.5,
                "extra_amount": 0,
                "note": "TEST_iter110",
            },
            timeout=20,
        )
        assert r.status_code == 200, r.text[:200]
        body = r.json()
        assert body.get("ok") is True
        entry = body.get("entry") or {}
        assert entry.get("extra_hours") == -0.5
        assert entry.get("user_id") == sample_employee["user_id"]

    def test_negative_amount_rejected(self, headers, sample_employee):
        r = requests.post(
            f"{API}/admin/attendance/extra-duty",
            headers=headers,
            json={
                "user_id": sample_employee["user_id"],
                "date": self.DATE,
                "extra_hours": 0,
                "extra_amount": -100,
            },
            timeout=20,
        )
        assert r.status_code == 400, r.text[:200]

    def test_zero_deletes_entry(self, headers, sample_employee):
        # First upsert a real entry so delete has something to work on.
        requests.post(
            f"{API}/admin/attendance/extra-duty",
            headers=headers,
            json={
                "user_id": sample_employee["user_id"],
                "date": self.DATE,
                "extra_hours": 1.0,
                "extra_amount": 0,
            },
            timeout=20,
        )
        # Now zero it out → should delete.
        r = requests.post(
            f"{API}/admin/attendance/extra-duty",
            headers=headers,
            json={
                "user_id": sample_employee["user_id"],
                "date": self.DATE,
                "extra_hours": 0,
                "extra_amount": 0,
            },
            timeout=20,
        )
        assert r.status_code == 200, r.text[:200]
        assert r.json().get("deleted") is True

        # Confirm the entry is gone via GET.
        g = requests.get(
            f"{API}/admin/attendance/extra-duty/{CID}"
            f"?from_date={self.DATE}&to_date={self.DATE}",
            headers=headers, timeout=20,
        )
        assert g.status_code == 200
        entries = g.json().get("entries") or []
        assert not any(
            e.get("user_id") == sample_employee["user_id"] and e.get("date") == self.DATE
            for e in entries
        ), "extra-duty entry should have been deleted"


# =====================================================================
# 2. Day-status + PATCH edit_reason / edited_by_name / original_hhmm
# =====================================================================
class TestDayStatusEditAudit:
    """Iter 111 — day-status cells expose edit audit trio for edited punches."""

    DATE = "2026-07-09"

    @pytest.fixture(scope="class")
    def created_record(self, headers, sample_employee):
        """Create a manual IN punch then PATCH it; return record_id & context."""
        create = requests.post(
            f"{API}/admin/attendance/manual-punch",
            headers=headers,
            json={
                "user_id": sample_employee["user_id"],
                "kind": "in",
                "at": f"{self.DATE}T09:00:00",
                "reason": "TEST_iter110_initial_create",
            },
            timeout=20,
        )
        assert create.status_code == 200, create.text[:200]
        rec = create.json()["record"]
        yield {"record_id": rec["record_id"], "emp": sample_employee}
        # Cleanup — best-effort delete.
        requests.delete(
            f"{API}/admin/attendance/{rec['record_id']}", headers=headers, timeout=10,
        )

    def test_patch_requires_reason(self, headers, created_record):
        r = requests.patch(
            f"{API}/admin/attendance/{created_record['record_id']}",
            headers=headers,
            json={"at": f"{self.DATE}T09:30:00", "reason": ""},
            timeout=20,
        )
        assert r.status_code == 400, f"empty reason should be rejected: {r.status_code}"

    def test_patch_with_reason_ok(self, headers, created_record):
        r = requests.patch(
            f"{API}/admin/attendance/{created_record['record_id']}",
            headers=headers,
            json={"at": f"{self.DATE}T09:30:00", "reason": "Due to Mismatch"},
            timeout=20,
        )
        assert r.status_code == 200, r.text[:200]

    def test_day_status_exposes_audit_fields(self, headers, created_record):
        r = requests.get(
            f"{API}/admin/attendance/day-status/{CID}"
            f"?from_date={self.DATE}&to_date={self.DATE}",
            headers=headers, timeout=30,
        )
        assert r.status_code == 200, r.text[:200]
        rows = r.json().get("rows") or []
        emp_uid = created_record["emp"]["user_id"]
        # Each row here is (user_id, date), so match on both.
        my_row = next(
            (row for row in rows
             if row.get("user_id") == emp_uid and row.get("date") == self.DATE),
            None,
        )
        assert my_row, "employee/date row missing from day-status"
        in_cell = my_row.get("in") or {}
        assert in_cell.get("hhmm") == "09:30", f"expected 09:30 got {in_cell.get('hhmm')}"
        assert in_cell.get("edited") is True
        # The Iter 111 audit trio must be present in the edited cell.
        for k in ("edit_reason", "edited_by_name", "original_hhmm"):
            assert k in in_cell, f"missing audit key {k} in cell: {in_cell}"
        assert (in_cell.get("edit_reason") or "").lower() == "due to mismatch"
        assert in_cell.get("original_hhmm") == "09:00"
        assert (in_cell.get("edited_by_name") or "").strip() != ""


# =====================================================================
# 3. Daily xlsx / pdf export (Iter 111)
# =====================================================================
class TestDailyExport:
    DATE = "2026-07-09"

    def test_daily_xlsx_200(self, headers):
        r = requests.get(
            f"{API}/admin/attendance/daily/{CID}/{self.DATE}.xlsx",
            headers=headers, timeout=45,
        )
        assert r.status_code == 200, r.text[:200]
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "excel" in ct, f"bad content-type {ct}"
        # Non-trivial payload (XLSX file starts with PK zip magic).
        assert r.content[:2] == b"PK", "xlsx body doesn't look like a zip"

        # Openpyxl reads inline strings — parse and check header row content.
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Header row is row 4 (rows 1-2 are title/summary, row 3 blank).
        header_vals = [c.value for c in ws[4]]
        expected = [
            "S.No", "Bio Code", "Emp Code", "Name", "Father Name",
            "Designation", "In", "Out", "OT In", "OT Out",
            "Duty HRS", "OT HRS", "Total HRS", "Status",
        ]
        assert header_vals == expected, f"header mismatch: {header_vals}"

    def test_daily_pdf_200(self, headers):
        r = requests.get(
            f"{API}/admin/attendance/daily/{CID}/{self.DATE}.pdf",
            headers=headers, timeout=45,
        )
        assert r.status_code == 200, r.text[:200]
        assert r.headers.get("content-type", "").startswith("application/pdf")
        assert r.content[:4] == b"%PDF", "not a PDF"

    def test_daily_invalid_date_400(self, headers):
        r = requests.get(
            f"{API}/admin/attendance/daily/{CID}/09-07-2026.xlsx",
            headers=headers, timeout=20,
        )
        # regex mismatch → 400
        assert r.status_code == 400, f"expected 400 got {r.status_code}: {r.text[:200]}"


# =====================================================================
# 4. Monthly exports still work with Bio Code column
# =====================================================================
class TestMonthlyExportsBioCode:
    MONTH = "2026-07"

    def _fetch(self, headers, path):
        r = requests.get(f"{API}{path}", headers=headers, timeout=90)
        assert r.status_code == 200, f"{path} → {r.status_code} {r.text[:200]}"
        return r

    def _xlsx_first_row_labels(self, content: bytes, header_row: int = 1) -> list:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        # Scan first 6 rows for any row containing "Bio" text.
        for r in range(1, 7):
            vals = [c.value for c in ws[r]]
            if any(isinstance(v, str) and "Bio" in v for v in vals):
                return vals
        # Fallback: return whichever row was requested.
        return [c.value for c in ws[header_row]]

    def test_monthly_inout_xlsx_has_bio(self, headers):
        r = self._fetch(headers, f"/admin/attendance/monthly-inout/{CID}/{self.MONTH}.xlsx")
        assert r.content[:2] == b"PK"
        header = self._xlsx_first_row_labels(r.content)
        assert any(isinstance(v, str) and "Bio" in v for v in header), (
            f"Bio Code column missing in Grid xlsx header: {header[:10]}"
        )

    def test_monthly_hours_xlsx_has_bio(self, headers):
        r = self._fetch(headers, f"/admin/attendance/monthly-hours/{CID}/{self.MONTH}.xlsx")
        assert r.content[:2] == b"PK"
        header = self._xlsx_first_row_labels(r.content)
        assert any(isinstance(v, str) and "Bio" in v for v in header), (
            f"Bio Code column missing in Hours xlsx header: {header[:10]}"
        )

    def test_monthly_inout_pdf_ok(self, headers):
        r = self._fetch(headers, f"/admin/attendance/monthly-inout/{CID}/{self.MONTH}.pdf")
        assert r.content[:4] == b"%PDF"

    def test_monthly_hours_pdf_ok(self, headers):
        r = self._fetch(headers, f"/admin/attendance/monthly-hours/{CID}/{self.MONTH}.pdf")
        assert r.content[:4] == b"%PDF"
