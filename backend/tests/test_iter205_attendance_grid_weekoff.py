"""Iter 205 backend tests:
- Attendance grid: Employee Code column removed everywhere, Father Name added,
  identity columns frozen (verified in frontend). Backend must return the
  monthly-grid rows with father_name, no employee_code column reliance.
- Clock-timing totals: totals accumulated as minutes (HH:MM); when
  attendance_by_duty_hours true & halfday_threshold_rule false, present_days
  are WHOLE integers + remainder to extra hours.
- XLSX exports: monthly-inout & monthly-hours workbooks each contain 2 sheets
  (main + 'OT HRS'); trailing totals in Hours-Only are HH:MM strings.
- New week_off_worked policy: persists & sanitizes; attendance engine applies
  half_day_ot (with double_ot) and ot_only modes on the weekly-off day.
"""
import io
import os
import time
import pytest
import requests
from openpyxl import load_workbook
from pymongo import MongoClient

BASE = os.environ.get("EXPO_PUBLIC_BACKEND_URL", "https://emplo-connect-1.preview.emergentagent.com").rstrip("/")
CMP = "cmp_527fecdd7c"
USER_ID = "user_44cd6f561da0"  # code 50 SURENDRA SINGH
MONTH = "2026-06"
TEST_DAY_ISO = "2026-06-14"  # Sunday, weekly-off (weekly_off_days=[6]), no existing punches

ADMIN_EMAIL = "sksharmaconsultancy@gmail.com"
ADMIN_PASSWORD = "sharma123"

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "test_database")


@pytest.fixture(scope="session")
def token():
    r = requests.post(f"{BASE}/api/auth/admin-password-login",
                      json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=30)
    assert r.status_code == 200, r.text
    return r.json()["session_token"]


@pytest.fixture(scope="session")
def api(token):
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {token}", "Content-Type": "application/json"})
    return s


@pytest.fixture(scope="session")
def db():
    return MongoClient(MONGO_URL)[DB_NAME]


@pytest.fixture(scope="session")
def original_policy(api):
    r = api.get(f"{BASE}/api/attendance/policy", params={"company_id": CMP})
    assert r.status_code == 200
    return (r.json().get("policy") or r.json()).get("week_off_worked")


@pytest.fixture(scope="session")
def original_policy_master(api):
    r = api.get(f"{BASE}/api/attendance/policy", params={"company_id": CMP})
    return (r.json().get("policy") or r.json()).get("policy_master") or {}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _patch_policy(api, delta: dict):
    r = api.patch(f"{BASE}/api/attendance/policy",
                  params={"company_id": CMP}, json={"policy": delta})
    assert r.status_code == 200, r.text
    return r.json()


def _get_policy(api):
    r = api.get(f"{BASE}/api/attendance/policy", params={"company_id": CMP})
    assert r.status_code == 200
    return r.json().get("policy") or r.json()


def _seed_punch(db, when_iso_start: str, when_iso_end: str, batch_tag: str = "TEST_iter205"):
    """Seed a pair of approved punches for USER_ID on the given IST window.
    when_iso_* are 'YYYY-MM-DDTHH:MM' in IST -> stored as UTC ISO strings.
    Schema: record_id, user_id, company_id, date (YYYY-MM-DD), kind (in|out),
    at (UTC ISO), source, status, manual_reason (used as batch tag), created_by,
    created_at.
    """
    from datetime import datetime, timezone, timedelta
    IST = timezone(timedelta(hours=5, minutes=30))
    ids = []
    for kind, iso in (("in", when_iso_start), ("out", when_iso_end)):
        dt_ist = datetime.fromisoformat(iso).replace(tzinfo=IST)
        dt_utc = dt_ist.astimezone(timezone.utc)
        date_str = dt_ist.strftime("%Y-%m-%d")
        rid = f"att_{batch_tag}_{kind}_{int(time.time()*1000)}"
        doc = {
            "record_id": rid,
            "user_id": USER_ID,
            "company_id": CMP,
            "date": date_str,
            "kind": kind,
            "at": dt_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "source": "manual_admin",
            "status": "approved",
            "approved_by": "user_67791559822a",
            "manual_reason": batch_tag,
            "created_by": "user_67791559822a",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        db.attendance.insert_one(doc)
        ids.append(rid)
        time.sleep(0.005)
    return ids


def _cleanup_seeded(db):
    db.attendance.delete_many({
        "user_id": USER_ID,
        "manual_reason": {"$regex": "^TEST_iter205"},
    })


def _monthly_grid(api):
    r = api.get(f"{BASE}/api/admin/attendance/monthly-grid/{CMP}/{MONTH}")
    assert r.status_code == 200, r.text
    return r.json()


def _find_row(grid: dict, user_id: str):
    for row in (grid.get("employees") or grid.get("rows") or []):
        if row.get("user_id") == user_id:
            return row
    return None


def _rows(grid: dict):
    return grid.get("employees") or grid.get("rows") or []


# ---------------------------------------------------------------------------
# 1. Policy persistence & sanitization
# ---------------------------------------------------------------------------
class TestWeekOffWorkedPolicy:
    def test_patch_and_get_persists(self, api):
        delta = {"week_off_worked": {
            "mode": "half_day_ot",
            "half_day_threshold": 4,
            "full_day_threshold": 8,
            "ot_after": 0,
            "double_ot": True,
        }}
        _patch_policy(api, delta)
        pol = _get_policy(api)
        w = pol.get("week_off_worked") or {}
        assert w.get("mode") == "half_day_ot"
        assert float(w.get("half_day_threshold")) == 4.0
        assert float(w.get("full_day_threshold")) == 8.0
        assert float(w.get("ot_after")) == 0.0
        assert w.get("double_ot") is True
        # sanitizer keeps boolean toggles as booleans
        for k in ("salary_credit", "leave_adjustment", "comp_off",
                   "double_wages", "approval_required"):
            assert isinstance(w.get(k), bool)


# ---------------------------------------------------------------------------
# 2. Attendance engine on weekly-off day — half_day_ot + double_ot
# ---------------------------------------------------------------------------
class TestWeekOffAttendanceEngine:
    """Sunday 2026-06-21 08:00->14:00 IST -> 6h worked, weekoff day.
    half_day_ot: 4h half-day threshold, 8h full-day. With ot_after=0, worked=6h
    means: half day present (0.5) + OT (worked - half_day_threshold) = 2h,
    doubled by double_ot -> 4h OT. Expected: present=0.5, duty=4, ot=4.
    """
    @pytest.fixture(autouse=True)
    def _seed(self, api, db):
        # policy already set to half_day_ot + double_ot in prior test class
        _patch_policy(api, {"week_off_worked": {
            "mode": "half_day_ot",
            "half_day_threshold": 4,
            "full_day_threshold": 8,
            "ot_after": 0,
            "double_ot": True,
            "salary_credit": True,
            "leave_adjustment": False,
            "comp_off": False,
            "double_wages": False,
            "approval_required": False,
        }})
        # clean existing test seed
        _cleanup_seeded(db)
        self.ids = _seed_punch(db, f"{TEST_DAY_ISO}T08:00", f"{TEST_DAY_ISO}T14:00", "TEST_iter205_a")
        yield
        _cleanup_seeded(db)

    def test_half_day_ot_double_ot(self, api):
        grid = _monthly_grid(api)
        row = _find_row(grid, USER_ID)
        assert row, "row for test user not present in monthly grid"
        cell = (row.get("days") or {}).get("14") or (row.get("cells") or {}).get("14")
        assert cell, f"day 14 cell missing: keys={(row.get('days') or row.get('cells') or {}).keys()}"
        # cells may be dicts or strings; fetch numeric fields
        present = cell.get("present") if isinstance(cell, dict) else None
        duty = cell.get("duty_hours") if isinstance(cell, dict) else None
        ot = cell.get("ot_hours") if isinstance(cell, dict) else None
        assert present == 0.5, f"present expected 0.5 got {present} cell={cell}"
        assert abs(float(duty) - 4.0) < 0.02, f"duty expected 4.0 got {duty}"
        assert abs(float(ot) - 4.0) < 0.02, f"ot expected 4.0 (2h doubled) got {ot}"


class TestWeekOffOtOnly:
    """ot_only mode: entire worked time goes to OT, present=0, duty=0."""
    @pytest.fixture(autouse=True)
    def _seed(self, api, db):
        _patch_policy(api, {"week_off_worked": {
            "mode": "ot_only",
            "half_day_threshold": 4,
            "full_day_threshold": 8,
            "ot_after": 0,
            "double_ot": False,
            "salary_credit": False,
            "leave_adjustment": False,
            "comp_off": False,
            "double_wages": False,
            "approval_required": False,
        }})
        _cleanup_seeded(db)
        _seed_punch(db, f"{TEST_DAY_ISO}T08:00", f"{TEST_DAY_ISO}T14:00", "TEST_iter205_b")
        yield
        _cleanup_seeded(db)

    def test_ot_only(self, api):
        grid = _monthly_grid(api)
        row = _find_row(grid, USER_ID)
        assert row
        cell = (row.get("days") or {}).get("14") or (row.get("cells") or {}).get("14")
        assert cell
        assert (cell.get("present") or 0) == 0, f"present expected 0 got {cell.get('present')}"
        assert abs(float(cell.get("duty_hours") or 0)) < 0.02
        assert abs(float(cell.get("ot_hours") or 0) - 6.0) < 0.02, cell


# ---------------------------------------------------------------------------
# 3. Grid rows structure: Father Name present, employee_code no longer required
# ---------------------------------------------------------------------------
class TestGridColumns:
    def test_rows_expose_father_name(self, api):
        grid = _monthly_grid(api)
        rows = _rows(grid)
        assert rows, "no rows in monthly grid"
        # every row must have father_name attribute (may be empty str/null)
        with_father = [r for r in rows if "father_name" in r]
        assert len(with_father) == len(rows), \
            f"only {len(with_father)}/{len(rows)} rows have father_name field"

    def test_totals_are_clock_exact(self, api):
        grid = _monthly_grid(api)
        rows = _rows(grid)
        assert rows
        # find a row with meaningful data
        row = next((r for r in rows if (r.get("totals") or {}).get("hours", 0) > 0), rows[0])
        cells = row.get("days") or row.get("cells") or {}
        # sum minute-accurate: duty + ot hours per cell -> total minutes
        total_min = 0
        for _, c in cells.items():
            if isinstance(c, dict):
                d = float(c.get("duty_hours") or 0)
                o = float(c.get("ot_hours") or 0)
                total_min += round((d + o) * 60)
        totals = row.get("totals") or {}
        th = float(totals.get("hours") or (float(totals.get("duty_hours") or 0) + float(totals.get("ot_hours") or 0)))
        got_min = round(th * 60)
        # allow 1-minute rounding tolerance
        assert abs(got_min - total_min) <= 1, \
            f"totals hours off from cell sum: totals_min={got_min} cells_min={total_min} totals={totals}"


# ---------------------------------------------------------------------------
# 4. Whole-day integer present + extra remainder when attendance_by_duty_hours
# ---------------------------------------------------------------------------
class TestDutyHoursDivision:
    @pytest.fixture(autouse=True)
    def _toggle(self, api, original_policy_master):
        _patch_policy(api, {"policy_master": {
            **(original_policy_master or {}),
            "attendance_by_duty_hours": True,
            "halfday_threshold_rule": False,
        }})
        yield
        _patch_policy(api, {"policy_master": original_policy_master})

    def test_whole_days_and_extra(self, api):
        grid = _monthly_grid(api)
        rows = _rows(grid)
        assert rows
        checked = 0
        for row in rows:
            totals = row.get("totals") or {}
            pdp = totals.get("present_days_policy")
            if pdp is None:
                continue
            # whole integer
            assert float(pdp).is_integer(), \
                f"present_days_policy must be integer in duty-hours mode, got {pdp}"
            checked += 1
            if checked >= 5:
                break
        assert checked > 0, "no row exposed present_days_policy"


# ---------------------------------------------------------------------------
# 5. XLSX exports have 2 sheets + HH:MM totals
# ---------------------------------------------------------------------------
class TestXlsxExports:
    def _fetch(self, api, path):
        r = api.get(f"{BASE}{path}", timeout=60)
        assert r.status_code == 200, f"{path} -> {r.status_code} {r.text[:200]}"
        assert r.content[:2] == b"PK", "not a xlsx"
        return load_workbook(io.BytesIO(r.content), data_only=True)

    def test_monthly_hours_two_sheets(self, api):
        wb = self._fetch(api, f"/api/admin/attendance/monthly-hours/{CMP}/{MONTH}.xlsx")
        names = wb.sheetnames
        assert "Hours Only" in names, names
        assert "OT HRS" in names, names

    def test_monthly_hours_trailing_totals_hhmm(self, api):
        wb = self._fetch(api, f"/api/admin/attendance/monthly-hours/{CMP}/{MONTH}.xlsx")
        ws = wb["Hours Only"]
        # scan last 4 rows (totals) for HH:MM string
        hhmm_found = False
        for row in ws.iter_rows(min_row=max(1, ws.max_row - 4), values_only=True):
            for v in row:
                if isinstance(v, str) and len(v) >= 3 and v.count(":") == 1:
                    h, _, m = v.partition(":")
                    if h.strip().lstrip("-").isdigit() and m.strip().isdigit():
                        hhmm_found = True; break
            if hhmm_found:
                break
        assert hhmm_found, "no HH:MM string in trailing totals of Hours Only"

    def test_monthly_inout_two_sheets(self, api):
        wb = self._fetch(api, f"/api/admin/attendance/monthly-inout/{CMP}/{MONTH}.xlsx")
        names = wb.sheetnames
        assert "Grid View" in names, names
        assert "OT HRS" in names, names


# ---------------------------------------------------------------------------
# 6. Restore original policy (best-effort teardown at end of module)
# ---------------------------------------------------------------------------
def test_zzz_restore(api, original_policy, original_policy_master):
    if original_policy:
        _patch_policy(api, {"week_off_worked": original_policy})
    if original_policy_master:
        _patch_policy(api, {"policy_master": original_policy_master})
