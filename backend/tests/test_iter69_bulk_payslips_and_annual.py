"""Iter 69 backend tests — Bulk payslips (PDF/ZIP) + Annual FY report XLSX.

Under test:
  1. GET /api/admin/salary-runs/{run_id}/payslips.pdf
     - super_admin happy path → application/pdf, body starts with %PDF-
     - company_admin happy path (own firm)
     - company_admin cross-firm → 403
     - 404 for unknown run_id
     - empty run (0 rows) → 200 with placeholder PDF (not 500)
     - Content-Disposition includes .pdf filename
     - Auth guard (no token → 401/403)

  2. GET /api/admin/salary-runs/{run_id}/payslips.zip
     - super_admin happy path → application/zip, body starts with PK
     - ZIP contains N .pdf files for run with N rows; each is a valid PDF
     - Filename inside zip pattern <EmployeeCode>_<Name>_<Month>.pdf
     - company_admin cross-firm → 403
     - 404 for unknown run_id
     - Auth guard

  3. GET /api/admin/reports/annual.xlsx?fy=2025-26&company_id=...
     - super_admin with company_id → 200 XLSX, body starts with PK
     - workbook has 4 sheets: Summary, Salary — per employee, Attendance, PF & ESIC
     - super_admin without company_id → 400
     - company_admin auto-forced to own firm even if a foreign company_id is passed
     - non-existent company_id → 404
     - Auth guard

Runs against the public preview URL. Uses OTP-dev-code flow for super_admin
and admin-pin-login for company_admin (SKSCO1) — pin_hash is NOT touched.
"""
from __future__ import annotations
import io
import os
import re
import zipfile
from typing import Optional

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = (
    os.environ.get("EXPO_BACKEND_URL")
    or os.environ.get("EXPO_PUBLIC_BACKEND_URL")
    or "https://emplo-connect-1.preview.emergentagent.com"
).rstrip("/")
API = f"{BASE_URL}/api"

SUPER_ADMIN_EMAIL = "sksharmaconsultancy@gmail.com"
SKSCO_ADMIN_PHONE = "+919810000001"
SKSCO_ADMIN_PIN = "387908"


def _auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _otp_login(sess, identifier: str, channel: str = "email") -> str:
    r = sess.post(f"{API}/auth/otp/request",
                  json={"identifier": identifier, "channel": channel})
    assert r.status_code == 200, f"otp/request failed: {r.status_code} {r.text[:200]}"
    body = r.json()
    code = body.get("dev_code") or body.get("code")
    assert code, f"No dev code in response: {body}"
    r = sess.post(f"{API}/auth/otp/verify",
                  json={"identifier": identifier, "channel": channel, "code": code})
    assert r.status_code == 200, f"otp/verify failed: {r.status_code} {r.text[:200]}"
    body = r.json()
    return body.get("session_token") or body.get("token")


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------
@pytest.fixture(scope="session")
def sess():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    return s


@pytest.fixture(scope="session")
def super_token(sess) -> str:
    return _otp_login(sess, SUPER_ADMIN_EMAIL, "email")


@pytest.fixture(scope="session")
def sksco_admin_token(sess) -> tuple[str, str]:
    r = sess.post(f"{API}/auth/admin-pin-login",
                  json={"identifier": SKSCO_ADMIN_PHONE, "pin": SKSCO_ADMIN_PIN})
    assert r.status_code == 200, f"admin-pin-login failed: {r.status_code} {r.text[:200]}"
    body = r.json()
    token = body["session_token"]
    company_id = body["user"].get("company_id")
    assert company_id, f"Company admin has no company_id: {body['user']}"
    return token, company_id


@pytest.fixture(scope="session")
def sksco_company(sess, super_token) -> dict:
    r = sess.get(f"{API}/companies", headers=_auth(super_token))
    assert r.status_code == 200
    items = r.json()
    if isinstance(items, dict):
        items = items.get("companies") or items.get("items") or []
    for c in items:
        if (c.get("company_code") or "").upper() == "SKSCO1":
            return c
    pytest.skip("SKSCO1 company not found")


@pytest.fixture(scope="session")
def other_company(sess, super_token) -> Optional[dict]:
    r = sess.get(f"{API}/companies", headers=_auth(super_token))
    if r.status_code != 200:
        return None
    items = r.json()
    if isinstance(items, dict):
        items = items.get("companies") or items.get("items") or []
    for c in items:
        if (c.get("company_code") or "").upper() != "SKSCO1":
            return c
    return None


@pytest.fixture(scope="session")
def any_salary_run(sess, super_token) -> dict:
    """Return the first available salary run from the environment."""
    r = sess.get(f"{API}/admin/salary-runs", headers=_auth(super_token))
    assert r.status_code == 200, f"list salary-runs failed: {r.status_code} {r.text[:200]}"
    body = r.json()
    runs = body.get("runs") if isinstance(body, dict) else body
    if not runs:
        pytest.skip("No salary runs in environment")
    return runs[0]


@pytest.fixture(scope="session")
def sksco_salary_run(sess, super_token, sksco_company) -> Optional[dict]:
    """Return a salary run belonging to SKSCO1 if any."""
    r = sess.get(f"{API}/admin/salary-runs",
                 params={"company_id": sksco_company["company_id"]},
                 headers=_auth(super_token))
    assert r.status_code == 200
    body = r.json()
    runs = body.get("runs") if isinstance(body, dict) else body
    if not runs:
        return None
    # Filter defensively — the endpoint should already scope
    for run in runs:
        if run.get("company_id") == sksco_company["company_id"]:
            return run
    return runs[0]


@pytest.fixture(scope="session")
def foreign_salary_run(sess, super_token, sksco_company) -> Optional[dict]:
    """A salary run whose company_id is NOT SKSCO1's — for cross-firm 403 checks.

    Also accepts runs with ``company_id=None`` since those will also fail the
    ``run.company_id == admin.company_id`` guard for a company_admin.
    """
    r = sess.get(f"{API}/admin/salary-runs", headers=_auth(super_token))
    if r.status_code != 200:
        return None
    body = r.json()
    runs = body.get("runs") if isinstance(body, dict) else body
    if not runs:
        return None
    sksco_cid = sksco_company["company_id"]
    for run in runs:
        if run.get("company_id") != sksco_cid:
            return run
    return None


@pytest.fixture(scope="session")
def empty_salary_run(sess, super_token) -> Optional[dict]:
    """A salary run with 0 rows, for the placeholder-PDF check."""
    r = sess.get(f"{API}/admin/salary-runs", headers=_auth(super_token))
    if r.status_code != 200:
        return None
    body = r.json()
    runs = body.get("runs") if isinstance(body, dict) else body
    for run in runs or []:
        if len(run.get("rows") or []) == 0:
            return run
    return None


@pytest.fixture(scope="session")
def nonempty_salary_run(sess, super_token) -> Optional[dict]:
    """A salary run with >=1 row having user_id — for size / naming checks."""
    r = sess.get(f"{API}/admin/salary-runs", headers=_auth(super_token))
    if r.status_code != 200:
        return None
    body = r.json()
    runs = body.get("runs") if isinstance(body, dict) else body
    for run in runs or []:
        rows = run.get("rows") or []
        if any(r.get("user_id") for r in rows):
            return run
    return None


# ---------------------------------------------------------------------------
# 1. GET /admin/salary-runs/{run_id}/payslips.pdf
# ---------------------------------------------------------------------------
class TestBulkPayslipsPdf:
    def test_super_admin_download_pdf(self, sess, super_token, any_salary_run):
        run_id = any_salary_run["run_id"]
        r = sess.get(f"{API}/admin/salary-runs/{run_id}/payslips.pdf",
                     headers=_auth(super_token))
        assert r.status_code == 200, \
            f"payslips.pdf failed: {r.status_code} {r.text[:300]}"
        assert r.headers.get("content-type", "").startswith("application/pdf"), \
            f"unexpected content-type: {r.headers.get('content-type')}"
        assert r.content[:5] == b"%PDF-", \
            f"body is not a PDF: first bytes={r.content[:16]!r}"
        # Content-Disposition includes .pdf filename
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower(), f"content-disposition missing: {cd}"
        assert ".pdf" in cd.lower(), f"filename should end in .pdf: {cd}"
        # size threshold: >5KB if run has ≥1 row
        rows = any_salary_run.get("rows") or []
        if rows:
            assert len(r.content) > 5 * 1024, \
                f"PDF suspiciously small for {len(rows)}-row run: {len(r.content)} bytes"

    def test_company_admin_download_own_firm(
        self, sess, sksco_admin_token, sksco_salary_run
    ):
        if not sksco_salary_run:
            pytest.skip("No SKSCO1 salary run available")
        token, _ = sksco_admin_token
        run_id = sksco_salary_run["run_id"]
        r = sess.get(f"{API}/admin/salary-runs/{run_id}/payslips.pdf",
                     headers=_auth(token))
        assert r.status_code == 200, \
            f"own-firm download failed: {r.status_code} {r.text[:200]}"
        assert r.content[:5] == b"%PDF-"

    def test_company_admin_cross_firm_403(
        self, sess, sksco_admin_token, foreign_salary_run
    ):
        if not foreign_salary_run:
            pytest.skip("No non-SKSCO1 salary run available for cross-firm test")
        token, _ = sksco_admin_token
        run_id = foreign_salary_run["run_id"]
        r = sess.get(f"{API}/admin/salary-runs/{run_id}/payslips.pdf",
                     headers=_auth(token))
        assert r.status_code == 403, \
            f"expected 403 cross-firm, got {r.status_code} {r.text[:200]}"

    def test_404_for_unknown_run(self, sess, super_token):
        r = sess.get(
            f"{API}/admin/salary-runs/srun_does_not_exist_iter69/payslips.pdf",
            headers=_auth(super_token))
        assert r.status_code == 404, \
            f"expected 404 for bogus run, got {r.status_code} {r.text[:200]}"

    def test_empty_run_returns_placeholder_pdf(
        self, sess, super_token, empty_salary_run
    ):
        """Empty run (0 rows) should still return 200 with placeholder PDF."""
        if not empty_salary_run:
            pytest.skip("No empty (0-row) salary run available to test placeholder")
        run_id = empty_salary_run["run_id"]
        r2 = sess.get(f"{API}/admin/salary-runs/{run_id}/payslips.pdf",
                      headers=_auth(super_token))
        assert r2.status_code == 200, \
            f"empty run should not 500, got {r2.status_code} {r2.text[:200]}"
        assert r2.content[:5] == b"%PDF-", \
            f"empty-run body is not a PDF: {r2.content[:16]!r}"

    def test_auth_required(self, sess):
        r = sess.get(f"{API}/admin/salary-runs/x/payslips.pdf")
        assert r.status_code in (401, 403), \
            f"expected 401/403 without auth, got {r.status_code}"


# ---------------------------------------------------------------------------
# 2. GET /admin/salary-runs/{run_id}/payslips.zip
# ---------------------------------------------------------------------------
class TestBulkPayslipsZip:
    def test_super_admin_download_zip(self, sess, super_token, any_salary_run):
        run_id = any_salary_run["run_id"]
        r = sess.get(f"{API}/admin/salary-runs/{run_id}/payslips.zip",
                     headers=_auth(super_token))
        assert r.status_code == 200, \
            f"payslips.zip failed: {r.status_code} {r.text[:300]}"
        assert r.headers.get("content-type", "").startswith("application/zip"), \
            f"unexpected content-type: {r.headers.get('content-type')}"
        assert r.content[:2] == b"PK", \
            f"body is not a ZIP: first bytes={r.content[:16]!r}"
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower() and ".zip" in cd.lower(), \
            f"content-disposition wrong: {cd}"

    def test_zip_contents_match_rows_and_naming(
        self, sess, super_token, nonempty_salary_run
    ):
        """ZIP should contain N .pdf files (one per row with user_id), each
        starting with %PDF-, and named <EmployeeCode>_<Name>_<Month>.pdf.
        """
        if not nonempty_salary_run:
            pytest.skip("No salary run with user_id rows — cannot verify naming")
        rows = nonempty_salary_run.get("rows") or []
        expected_n = sum(1 for r in rows if r.get("user_id"))

        month = nonempty_salary_run.get("month") or ""
        run_id = nonempty_salary_run["run_id"]
        r = sess.get(f"{API}/admin/salary-runs/{run_id}/payslips.zip",
                     headers=_auth(super_token))
        assert r.status_code == 200
        buf = io.BytesIO(r.content)
        with zipfile.ZipFile(buf, "r") as zf:
            names = zf.namelist()
            pdf_names = [n for n in names if n.lower().endswith(".pdf")]
            assert len(pdf_names) == expected_n, (
                f"expected {expected_n} PDFs in ZIP, got {len(pdf_names)}: {names[:10]}"
            )
            # Filename pattern: <code>_<name>_<month>.pdf
            month_re = re.escape(month) if month else r"[\w\-]+"
            pattern = re.compile(rf"^.+_.+_{month_re}\.pdf$", re.IGNORECASE)
            for n in pdf_names:
                assert pattern.match(n), \
                    f"filename does not match <code>_<name>_<month>.pdf pattern: {n}"
                data = zf.read(n)
                assert data[:5] == b"%PDF-", \
                    f"file {n} inside zip is not a PDF: first bytes={data[:16]!r}"
                assert len(data) > 1024, \
                    f"file {n} suspiciously small: {len(data)} bytes"

    def test_company_admin_cross_firm_403(
        self, sess, sksco_admin_token, foreign_salary_run
    ):
        if not foreign_salary_run:
            pytest.skip("No non-SKSCO1 salary run available for cross-firm test")
        token, _ = sksco_admin_token
        run_id = foreign_salary_run["run_id"]
        r = sess.get(f"{API}/admin/salary-runs/{run_id}/payslips.zip",
                     headers=_auth(token))
        assert r.status_code == 403, \
            f"expected 403 cross-firm, got {r.status_code} {r.text[:200]}"

    def test_404_for_unknown_run(self, sess, super_token):
        r = sess.get(
            f"{API}/admin/salary-runs/srun_does_not_exist_iter69/payslips.zip",
            headers=_auth(super_token))
        assert r.status_code == 404, \
            f"expected 404 for bogus run, got {r.status_code}"

    def test_auth_required(self, sess):
        r = sess.get(f"{API}/admin/salary-runs/x/payslips.zip")
        assert r.status_code in (401, 403)


# ---------------------------------------------------------------------------
# 3. GET /admin/reports/annual.xlsx
# ---------------------------------------------------------------------------
EXPECTED_SHEETS = {"Summary", "Salary — per employee", "Attendance", "PF & ESIC"}


class TestAnnualReportXlsx:
    def test_super_admin_download_with_company_id(
        self, sess, super_token, sksco_company
    ):
        r = sess.get(f"{API}/admin/reports/annual.xlsx",
                     params={"fy": "2025-26",
                             "company_id": sksco_company["company_id"]},
                     headers=_auth(super_token))
        assert r.status_code == 200, \
            f"annual.xlsx failed: {r.status_code} {r.text[:300]}"
        ctype = r.headers.get("content-type", "")
        assert ctype.startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ), f"unexpected content-type: {ctype}"
        assert r.content[:2] == b"PK", \
            f"body is not XLSX (zip): first bytes={r.content[:16]!r}"
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower() and ".xlsx" in cd.lower(), \
            f"content-disposition wrong: {cd}"
        # Verify 4 required sheets
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        sheet_names = set(wb.sheetnames)
        missing = EXPECTED_SHEETS - sheet_names
        assert not missing, \
            f"missing sheets: {missing} — actual sheets: {sheet_names}"

    def test_super_admin_missing_company_id_returns_400(
        self, sess, super_token
    ):
        r = sess.get(f"{API}/admin/reports/annual.xlsx",
                     params={"fy": "2025-26"},
                     headers=_auth(super_token))
        assert r.status_code == 400, \
            f"expected 400 for missing company_id, got {r.status_code} {r.text[:200]}"

    def test_company_admin_scoped_to_own_firm(
        self, sess, sksco_admin_token, other_company
    ):
        """company_admin's company_id is auto-forced regardless of query param.

        Passing another firm's company_id should still yield 200 (scoped to
        SKSCO1) — NOT a 403 or foreign data.
        """
        token, own_cid = sksco_admin_token
        params = {"fy": "2025-26"}
        if other_company:
            params["company_id"] = other_company["company_id"]
        r = sess.get(f"{API}/admin/reports/annual.xlsx",
                     params=params, headers=_auth(token))
        assert r.status_code == 200, \
            f"company_admin annual xlsx failed: {r.status_code} {r.text[:300]}"
        assert r.content[:2] == b"PK", \
            f"body is not XLSX: {r.content[:16]!r}"
        # Filename should reference SKSCO1's company name, not the foreign firm
        cd = r.headers.get("content-disposition", "")
        if other_company and other_company.get("name"):
            foreign_slug = other_company["name"].replace(" ", "_")
            assert foreign_slug not in cd, (
                f"Filename leaks foreign company name — company_admin should be "
                f"forced to their own firm. cd={cd}"
            )

    def test_nonexistent_company_returns_404(self, sess, super_token):
        r = sess.get(f"{API}/admin/reports/annual.xlsx",
                     params={"fy": "2025-26",
                             "company_id": "company_does_not_exist_iter69"},
                     headers=_auth(super_token))
        assert r.status_code == 404, \
            f"expected 404 for bogus company_id, got {r.status_code} {r.text[:200]}"

    def test_auth_required(self, sess):
        r = sess.get(f"{API}/admin/reports/annual.xlsx",
                     params={"fy": "2025-26", "company_id": "any"})
        assert r.status_code in (401, 403)


# ---------------------------------------------------------------------------
# 4. Unit-level checks — exercise the utility builders directly. This locks
#    in the empty-list / multi-page / single-PDF behaviour even when the
#    live environment has no salary runs with rows.
# ---------------------------------------------------------------------------
class TestPayslipBuildersUnit:
    def test_build_payslip_pdf_single(self):
        import sys
        sys.path.insert(0, "/app/backend")
        from utils.payslip_pdf import build_payslip_pdf
        pdf = build_payslip_pdf(
            employee={"name": "Alice", "employee_code": "E0001",
                      "designation": "Dev", "department": "Eng"},
            company={"name": "Test Co", "address": "Addr",
                     "phone": "1", "email": "a@b.c"},
            row={"base_pay": 30000, "hra": 5000, "gross": 40000,
                 "total_deduction": 3000, "net": 37000,
                 "present_days": 20, "month_days": 30},
            month="2025-05",
        )
        assert pdf[:5] == b"%PDF-"
        assert len(pdf) > 5 * 1024

    def test_build_bulk_payslip_pdf_multi_page(self):
        import sys
        sys.path.insert(0, "/app/backend")
        from utils.payslip_pdf import build_bulk_payslip_pdf
        entries = [
            {"employee": {"name": "Alice", "employee_code": "E0001"},
             "row": {"gross": 40000, "net": 37000}},
            {"employee": {"name": "Bob", "employee_code": "E0002"},
             "row": {"gross": 45000, "net": 41000}},
        ]
        pdf = build_bulk_payslip_pdf(
            company={"name": "Test Co"}, month="2025-05", entries=entries,
        )
        assert pdf[:5] == b"%PDF-"
        assert len(pdf) > 5 * 1024

    def test_build_bulk_payslip_pdf_empty(self):
        import sys
        sys.path.insert(0, "/app/backend")
        from utils.payslip_pdf import build_bulk_payslip_pdf
        pdf = build_bulk_payslip_pdf(
            company={"name": "Test Co"}, month="2025-05", entries=[],
        )
        assert pdf[:5] == b"%PDF-", \
            f"empty entries should still return placeholder PDF: {pdf[:16]!r}"


class TestAnnualReportBuilderUnit:
    @pytest.mark.asyncio
    async def test_annual_xlsx_sheets_from_builder(self):
        """Call the async builder directly with a stub DB to lock in the
        exact 4 sheets required by the review."""
        import sys
        sys.path.insert(0, "/app/backend")
        from utils.annual_report import build_annual_report_xlsx

        class _Cursor:
            def __init__(self, items):
                self._items = items
            def __aiter__(self):
                self._i = 0
                return self
            async def __anext__(self):
                if self._i >= len(self._items):
                    raise StopAsyncIteration
                v = self._items[self._i]
                self._i += 1
                return v
            async def to_list(self, n):
                return list(self._items)

        class _Coll:
            def __init__(self, items):
                self._items = items
            def find(self, *a, **kw):
                return _Cursor(self._items)

        class _DB:
            salary_runs = _Coll([])
            compliance_salary_runs = _Coll([])
            payslips = _Coll([])
            users = _Coll([])

        xlsx = await build_annual_report_xlsx(
            _DB(), company_id="cmp_x", fy="2025-26", company_name="Test Co",
        )
        assert xlsx[:2] == b"PK", f"not xlsx: {xlsx[:16]!r}"
        wb = load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
        assert set(wb.sheetnames) >= EXPECTED_SHEETS, \
            f"missing sheets from builder: {EXPECTED_SHEETS - set(wb.sheetnames)}"
