import requests, json, asyncio, os, uuid

B = "http://localhost:8001/api"
r = requests.post(B+"/auth/admin-password-login", json={"email":"sksharmaconsultancy@gmail.com","password":"sharma123"})
H = {"Authorization":"Bearer "+(r.json().get("token") or r.json().get("session_token"))}
CID = "cmp_527fecdd7c"; UID = "user_44cd6f561da0"

# 1. firm weekly off = N/A + new mode full_day_min_hours (min auto 50%)
p = requests.patch(B+"/attendance/policy", params={"company_id":CID}, headers=H, json={
  "weekly_off_days": [],
  "full_day_hours": 8, "half_day_hours": 4,
  "week_off_worked": {"mode": "full_day_min_hours", "min_hours": 0, "ot_after": 0,
                      "comp_off": False, "salary_credit": True}})
print("policy save:", p.status_code)
pol = requests.get(B+"/attendance/policy", params={"company_id":CID}, headers=H).json()
polp = pol.get("policy") or pol
print("weekly_off_days:", polp.get("weekly_off_days"),
      "| wow mode:", (polp.get("week_off_worked") or {}).get("mode"),
      "min_hours:", (polp.get("week_off_worked") or {}).get("min_hours"))

# 2. employee override weekly off = Sunday
u = requests.patch(B+"/admin/user-role", headers=H,
                   json={"user_id": UID, "weekly_off_days_override": [6]})
print("emp weekly-off save:", u.status_code)

# 3. seed: Sunday 2026-06-28 worked 5h (>= min 4 -> full day), Sunday 2026-06-07 worked 3h (< min -> plain duty)
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
load_dotenv("/app/backend/.env")
async def seed():
    db = AsyncIOMotorClient(os.environ["MONGO_URL"])[os.environ.get("DB_NAME","test_database")]
    await db.attendance.delete_many({"user_id":UID,"date":{"$in":["2026-06-28","2026-06-07"]}})
    def rec(date,kind,hh,mm):
        return {"record_id":"att_"+uuid.uuid4().hex[:12],"user_id":UID,"company_id":CID,
                "date":date,"kind":kind,"at":f"{date}T{hh:02d}:{mm:02d}:00+05:30",
                "source":"biometric","status":"approved"}
    await db.attendance.insert_many([
        rec("2026-06-28","in",9,0), rec("2026-06-28","out",14,0),  # 5h Sunday
        rec("2026-06-07","in",9,0), rec("2026-06-07","out",12,0),  # 3h Sunday
    ])
asyncio.get_event_loop().run_until_complete(seed())

g = requests.get(f"{B}/admin/attendance/monthly-grid/{CID}/2026-06", headers=H).json()
row = next(x for x in g["employees"] if x["user_id"]==UID)
for d in ("28","07"):
    c = row["days"].get(d) or {}
    print(f"day{d}:", {k:c.get(k) for k in ("weekly_off","present","duty_hours","ot_hours","hours")})
# a non-override employee should NOT have weekly off on Sunday now (firm N/A)
other = next((x for x in g["employees"] if x["user_id"]!=UID and (x["days"].get("28") or {}).get("in")), None)
if other:
    print("other emp day28 weekly_off:", (other["days"].get("28") or {}).get("weekly_off"))

# 4. Hours-only XLSX duty row combines duty+OT capped 24 (check emp with OT)
import io
from openpyxl import load_workbook
resp = requests.get(f"{B}/admin/attendance/monthly-hours/{CID}/2026-06.xlsx", headers=H)
wb = load_workbook(io.BytesIO(resp.content))
print("sheets:", wb.sheetnames)

# cleanup
async def clean():
    db = AsyncIOMotorClient(os.environ["MONGO_URL"])[os.environ.get("DB_NAME","test_database")]
    await db.attendance.delete_many({"user_id":UID,"date":{"$in":["2026-06-28","2026-06-07"]},"source":"biometric"})
asyncio.get_event_loop().run_until_complete(clean())
print("cleanup done")
