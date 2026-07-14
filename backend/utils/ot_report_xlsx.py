"""
OT Report XLSX builder — Iter 77i.

Given the ``rows`` produced by ``build_ot_report_rows`` (in server.py), emit
a compact XLSX with one row per (employee × OT day). The XLSX is streamed
back to the browser by an endpoint in server.py.

The output is intentionally minimal so admins can copy-paste into their
existing payroll sheet without worrying about compliance columns.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _fmt_hours_hm(v: float | int | None) -> str:
    """Convert decimal hours to HH:MM (matches the on-screen format)."""
    if v is None or v <= 0:
        return "0:00"
    total_min = int(round(float(v) * 60))
    h, m = divmod(total_min, 60)
    return f"{h}:{m:02d}"


def build_ot_report_xlsx(
    company_name: str,
    period_label: str,
    rows: List[Dict[str, Any]],
) -> bytes:
    """Return the XLSX bytes for the OT Report.

    Each ``rows`` item shape::
        {
            "employee_code": str,
            "name": str,
            "designation": str | None,
            "bio_code": str | None,
            "date": "YYYY-MM-DD",
            "day_label": "Mon" | "Tue" | ...,
            "in": "HH:MM" | None,
            "out": "HH:MM" | None,
            "duty_hours": float,
            "ot_hours": float,
            "total_hours": float,
        }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "OT Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="B4232C")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    # Title bar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    ws.cell(row=1, column=1).value = (
        f"{company_name} — OT Report — {period_label}"
    )
    ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="B4232C")
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
    ws.cell(row=2, column=1).value = (
        f"{len(rows)} OT day(s). Duty HRS shown here are duty-only "
        "(excluding OT); Total Duty HRS = Duty + OT."
    )
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="555555")
    ws.cell(row=2, column=1).alignment = center

    headers = [
        "Emp Code", "Name", "Designation", "Bio",
        "Date", "Day", "In Punch", "Out Punch",
        "OT In", "OT Out",
        "Duty HRS", "OT HRS", "Total Duty HRS",
    ]
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=idx)
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    r = 5
    for row in rows:
        vals = [
            row.get("employee_code") or "",
            row.get("name") or "",
            row.get("designation") or "",
            (row.get("bio_code") if row.get("bio_code") not in (None, "") else ""),
            row.get("date") or "",
            row.get("day_label") or "",
            row.get("in") or "—",
            row.get("out") or "—",
            row.get("ot_in") or "—",
            row.get("ot_out") or "—",
            _fmt_hours_hm(row.get("duty_hours") or 0),
            _fmt_hours_hm(row.get("ot_hours") or 0),
            _fmt_hours_hm(row.get("total_hours") or 0),
        ]
        for idx, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=idx)
            c.value = v
            c.alignment = center if idx > 3 else left
            if idx >= 11:  # HRS columns
                c.font = Font(bold=True)
        r += 1

    # ----- Totals row --------------------------------------------------
    if rows:
        total_duty = sum(float(x.get("duty_hours") or 0) for x in rows)
        total_ot = sum(float(x.get("ot_hours") or 0) for x in rows)
        total_all = sum(float(x.get("total_hours") or 0) for x in rows)
        ws.cell(row=r, column=1).value = "TOTAL"
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.cell(row=r, column=11).value = _fmt_hours_hm(total_duty)
        ws.cell(row=r, column=12).value = _fmt_hours_hm(total_ot)
        ws.cell(row=r, column=13).value = _fmt_hours_hm(total_all)
        for c_idx in (11, 12, 13):
            c = ws.cell(row=r, column=c_idx)
            c.font = Font(bold=True, color="B4232C")
            c.alignment = center

    # ----- Column widths -----------------------------------------------
    widths = [10, 24, 20, 8, 12, 6, 10, 10, 10, 10, 10, 10, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
