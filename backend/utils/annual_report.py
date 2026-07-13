"""Annual (Financial Year) Report XLSX — Iter 69.

Compiles a multi-sheet workbook with a firm's payroll + attendance
performance for one financial year (April→March). Designed as the
year-end deliverable for CA / statutory audit and directors' review.

Sheets produced (in order):
  1. Summary     — one row per month with total gross, net, PF, ESIC,
                    TDS and headcount.
  2. Salary      — per-employee grand totals across 12 months.
  3. Attendance  — per-employee monthly present/half/leave counts.
  4. PF-ESIC     — statutory contributions grouped month-wise (both
                    employee + employer) so the CA can reconcile Form 3A
                    / Form 24Q workings.

Callers pass:
  * ``db``        — Motor db handle (async).
  * ``company_id`` — mandatory.
  * ``fy``        — e.g. "2025-26" (April 2025 → March 2026).
  * ``company_name`` — pretty label for the Summary sheet header.

Returns ``bytes`` (raw XLSX).
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BRAND = "1F4E4E"
BRAND_LIGHT = "E6EDED"
ACCENT = "C89B3C"
INK = "1E2A2A"
ZEBRA = "F7F9F9"

HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor=BRAND)
SUB_FONT = Font(name="Calibri", size=10, bold=True, color=INK)
SUB_FILL = PatternFill("solid", fgColor=BRAND_LIGHT)
CELL_FONT = Font(name="Calibri", size=10, color=INK)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=INK)
TOTAL_FILL = PatternFill("solid", fgColor=BRAND_LIGHT)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

THIN = Side(border_style="thin", color="D6DEDE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fy_months(fy: str) -> List[str]:
    """Return the 12 YYYY-MM strings for an Indian FY (Apr→Mar).

    ``fy`` accepts either ``"2025-26"`` or ``"2025-2026"``.
    """
    try:
        parts = fy.strip().split("-")
        start_year = int(parts[0])
    except Exception:
        start_year = datetime.now().year
    months = []
    for i in range(12):
        y = start_year + (0 if 4 + i <= 12 else 1)
        m = 4 + i if 4 + i <= 12 else (4 + i - 12)
        months.append(f"{y:04d}-{m:02d}")
    return months


def _apply_header(ws, row: int, headers: List[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER


def _apply_row(ws, row: int, values: List[Any], *, is_total: bool = False) -> None:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = TOTAL_FONT if is_total else CELL_FONT
        if is_total:
            c.fill = TOTAL_FILL
        c.alignment = RIGHT if isinstance(v, (int, float)) else LEFT
        c.border = BORDER


def _autosize(ws, max_col: int, min_w: int = 10, max_w: int = 30) -> None:
    for col in range(1, max_col + 1):
        L = get_column_letter(col)
        max_len = 0
        for cell in ws[L]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[L].width = max(min_w, min(max_w, max_len + 2))


def _month_label(m: str) -> str:
    try:
        return datetime.strptime(m, "%Y-%m").strftime("%b %Y")
    except ValueError:
        return m


async def build_annual_report_xlsx(
    db,
    *,
    company_id: str,
    fy: str,
    company_name: str,
) -> bytes:
    """Async — compiles all four sheets from live Mongo data."""
    months = fy_months(fy)

    # Fetch salary_runs, compliance_salary_runs, attendance for the FY.
    runs = await db.salary_runs.find(
        {"company_id": company_id, "month": {"$in": months}},
        {"_id": 0},
    ).to_list(50)
    comp_runs = await db.compliance_salary_runs.find(
        {"company_id": company_id, "month": {"$in": months}},
        {"_id": 0},
    ).to_list(50)
    payslips = await db.payslips.find(
        {"company_id": company_id, "month": {"$in": months}},
        {"_id": 0},
    ).to_list(20000)
    # Index employees by id for name lookups
    emp_ids = {p.get("employee_user_id") for p in payslips if p.get("employee_user_id")}
    for run in runs + comp_runs:
        for r in (run.get("rows") or []):
            if r.get("user_id"):
                emp_ids.add(r["user_id"])
    emps = {}
    if emp_ids:
        async for u in db.users.find(
            {"user_id": {"$in": list(emp_ids)}},
            {"_id": 0, "user_id": 1, "name": 1, "employee_code": 1,
             "designation": 1, "department": 1, "uan_no": 1, "pf_no": 1},
        ):
            emps[u["user_id"]] = u

    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:H1")
    hdr = ws.cell(row=1, column=1,
                  value=f"Annual Payroll Summary — {company_name} — FY {fy}")
    hdr.font = Font(size=14, bold=True, color=BRAND)
    hdr.alignment = CENTER
    _apply_header(ws, 3, [
        "Month", "Employees", "Gross", "PF (Emp+Empr)", "ESIC (Emp+Empr)",
        "TDS", "Total Deductions", "Net Payable",
    ])

    row_idx = 4
    totals = {
        "emp": 0, "gross": 0.0, "pf": 0.0, "esic": 0.0,
        "tds": 0.0, "ded": 0.0, "net": 0.0,
    }
    # Combine per-month values from BOTH salary_runs & comp_runs.
    for m in months:
        emp_set = set()
        gross = pf = esic = tds = ded = net = 0.0
        for run in runs:
            if run.get("month") != m:
                continue
            for r in (run.get("rows") or []):
                if r.get("user_id"):
                    emp_set.add(r["user_id"])
                gross += float(r.get("gross") or 0)
                ded += float(r.get("total_deduction") or 0)
                net += float(r.get("net") or 0)
        for run in comp_runs:
            if run.get("month") != m:
                continue
            for r in (run.get("rows") or []):
                pf += float((r.get("pf_employee") or 0) + (r.get("pf_employer") or 0))
                esic += float(
                    (r.get("esic_employee") or 0) + (r.get("esic_employer") or 0)
                )
                tds += float(r.get("tds") or 0)
        _apply_row(ws, row_idx, [
            _month_label(m), len(emp_set),
            round(gross, 2), round(pf, 2), round(esic, 2),
            round(tds, 2), round(ded, 2), round(net, 2),
        ])
        totals["emp"] = max(totals["emp"], len(emp_set))
        totals["gross"] += gross
        totals["pf"] += pf
        totals["esic"] += esic
        totals["tds"] += tds
        totals["ded"] += ded
        totals["net"] += net
        row_idx += 1
    _apply_row(ws, row_idx, [
        "TOTAL", totals["emp"],
        round(totals["gross"], 2),
        round(totals["pf"], 2),
        round(totals["esic"], 2),
        round(totals["tds"], 2),
        round(totals["ded"], 2),
        round(totals["net"], 2),
    ], is_total=True)
    _autosize(ws, 8, min_w=14, max_w=22)

    # ── Sheet 2: Salary (per-employee grand totals) ─────────────────────────
    ws2 = wb.create_sheet("Salary — per employee")
    _apply_header(ws2, 1, [
        "Emp Code", "Name", "Designation", "Department",
    ] + [_month_label(m) for m in months] + ["Total"])
    per_emp_total: Dict[str, Dict[str, Any]] = {}
    per_emp_month: Dict[Tuple[str, str], float] = {}
    for run in runs:
        m = run.get("month")
        if m not in months:
            continue
        for r in (run.get("rows") or []):
            uid = r.get("user_id")
            if not uid:
                continue
            per_emp_total.setdefault(uid, {"net_total": 0.0})
            per_emp_total[uid]["net_total"] += float(r.get("net") or 0)
            per_emp_month[(uid, m)] = float(r.get("net") or 0)

    row_idx = 2
    for uid in sorted(
        per_emp_total.keys(),
        key=lambda u: (
            (emps.get(u) or {}).get("employee_code") or "",
            (emps.get(u) or {}).get("name") or "",
        ),
    ):
        e = emps.get(uid) or {}
        vals: List[Any] = [
            e.get("employee_code") or "—",
            e.get("name") or "—",
            e.get("designation") or "—",
            e.get("department") or "—",
        ]
        for m in months:
            vals.append(round(per_emp_month.get((uid, m), 0.0), 2))
        vals.append(round(per_emp_total[uid]["net_total"], 2))
        _apply_row(ws2, row_idx, vals)
        row_idx += 1
    _autosize(ws2, 4 + 12 + 1, min_w=12, max_w=24)

    # ── Sheet 3: Attendance ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Attendance")
    _apply_header(ws3, 1, [
        "Emp Code", "Name",
    ] + [_month_label(m) + " Present" for m in months] + ["Total Present"])
    att_month: Dict[Tuple[str, str], float] = {}
    att_total: Dict[str, float] = {}
    for run in runs:
        m = run.get("month")
        if m not in months:
            continue
        for r in (run.get("rows") or []):
            uid = r.get("user_id")
            if not uid:
                continue
            pd = float(r.get("present_days") or 0)
            att_month[(uid, m)] = pd
            att_total[uid] = att_total.get(uid, 0.0) + pd

    row_idx = 2
    for uid in sorted(
        att_total.keys(),
        key=lambda u: (
            (emps.get(u) or {}).get("employee_code") or "",
            (emps.get(u) or {}).get("name") or "",
        ),
    ):
        e = emps.get(uid) or {}
        vals: List[Any] = [
            e.get("employee_code") or "—",
            e.get("name") or "—",
        ]
        for m in months:
            vals.append(round(att_month.get((uid, m), 0.0), 1))
        vals.append(round(att_total.get(uid, 0.0), 1))
        _apply_row(ws3, row_idx, vals)
        row_idx += 1
    _autosize(ws3, 2 + 12 + 1, min_w=10, max_w=22)

    # ── Sheet 4: PF-ESIC (month-wise contributions) ─────────────────────────
    ws4 = wb.create_sheet("PF & ESIC")
    _apply_header(ws4, 1, [
        "Month", "Employees",
        "PF Employee", "PF Employer",
        "ESIC Employee", "ESIC Employer",
        "PT", "TDS",
    ])
    tot4 = {"emp": 0, "pf_e": 0.0, "pf_r": 0.0, "esi_e": 0.0, "esi_r": 0.0, "pt": 0.0, "tds": 0.0}
    row_idx = 2
    for m in months:
        emp_set = set()
        pf_e = pf_r = esi_e = esi_r = pt = tds = 0.0
        for run in comp_runs:
            if run.get("month") != m:
                continue
            for r in (run.get("rows") or []):
                if r.get("user_id"):
                    emp_set.add(r["user_id"])
                pf_e += float(r.get("pf_employee") or 0)
                pf_r += float(r.get("pf_employer") or 0)
                esi_e += float(r.get("esic_employee") or 0)
                esi_r += float(r.get("esic_employer") or 0)
                pt += float(r.get("pt") or 0)
                tds += float(r.get("tds") or 0)
        _apply_row(ws4, row_idx, [
            _month_label(m), len(emp_set),
            round(pf_e, 2), round(pf_r, 2),
            round(esi_e, 2), round(esi_r, 2),
            round(pt, 2), round(tds, 2),
        ])
        tot4["emp"] = max(tot4["emp"], len(emp_set))
        tot4["pf_e"] += pf_e
        tot4["pf_r"] += pf_r
        tot4["esi_e"] += esi_e
        tot4["esi_r"] += esi_r
        tot4["pt"] += pt
        tot4["tds"] += tds
        row_idx += 1
    _apply_row(ws4, row_idx, [
        "TOTAL", tot4["emp"],
        round(tot4["pf_e"], 2), round(tot4["pf_r"], 2),
        round(tot4["esi_e"], 2), round(tot4["esi_r"], 2),
        round(tot4["pt"], 2), round(tot4["tds"], 2),
    ], is_total=True)
    _autosize(ws4, 8, min_w=14, max_w=22)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
