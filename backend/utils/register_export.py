"""Iter 357 — shared register-style PDF/Excel export builder.

Used by Labour Statistics, Annual Returns and Factory & Boilers modules.
A3-landscape professional register: title, company logo, column headers
repeated on every page, totals row, page numbers.
"""
from io import BytesIO
from typing import Any, Dict, List, Optional


def register_xlsx(title: str, subtitle: str, columns: List[Dict[str, str]],
                  rows: List[dict], totals: Optional[dict] = None,
                  form_line: Optional[str] = None) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Register"
    n = len(columns)
    thin = Side(style="thin", color="9AA0A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.cell(1, 1, title).font = Font(bold=True, size=13)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    # Iter 477 — optional statutory FORM heading line(s) between the title
    # and the subtitle ("\n"-separated → one merged row per line).
    r_sub = 2
    for ln in (form_line or "").split("\n"):
        if not ln.strip():
            continue
        ws.merge_cells(start_row=r_sub, start_column=1,
                       end_row=r_sub, end_column=n)
        c = ws.cell(r_sub, 1, ln.strip())
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="center")
        r_sub += 1
    ws.merge_cells(start_row=r_sub, start_column=1,
                   end_row=r_sub, end_column=n)
    ws.cell(r_sub, 1, subtitle).alignment = Alignment(horizontal="center")
    r_head = r_sub + 2
    for j, c in enumerate(columns, 1):
        cell = ws.cell(r_head, j, c["label"])
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = max(
            11, min(28, len(c["label"]) + 4))
    r = r_head + 1
    for row in rows:
        for j, c in enumerate(columns, 1):
            v = row.get(c["key"])
            cell = ws.cell(r, j, v)
            cell.border = border
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.##"
                cell.alignment = Alignment(horizontal="right")
        r += 1
    if totals:
        _has_lbl = any(str(v).strip().upper() == "TOTAL"
                       for v in totals.values())
        for j, c in enumerate(columns, 1):
            v = totals.get(c["key"],
                           "TOTAL" if j == 1 and not _has_lbl else None)
            cell = ws.cell(r, j, v)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
            cell.border = border
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.##"
                cell.alignment = Alignment(horizontal="right")
    ws.freeze_panes = f"A{r_head + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def register_pdf(title: str, subtitle: str, columns: List[Dict[str, str]],
                 rows: List[dict], totals: Optional[dict] = None,
                 logo_b64: Optional[str] = None,
                 empty_note: Optional[str] = None,
                 form_line: Optional[str] = None) -> BytesIO:
    import base64
    from reportlab.lib import colors as rl
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)
    buf = BytesIO()
    # Iter 432 (user request) — tighter page margins so the table gets more
    # width for the BIGGER print-friendly fonts.
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3), leftMargin=6 * mm,
                            rightMargin=6 * mm, topMargin=7 * mm,
                            bottomMargin=9 * mm, title=title)
    W = landscape(A3)[0] - 12 * mm
    story: List[Any] = []
    if logo_b64:
        try:
            raw = base64.b64decode(logo_b64.split(",")[-1])
            story.append(Image(BytesIO(raw), width=20 * mm, height=20 * mm))
        except Exception:  # noqa: BLE001
            pass
    story.append(Paragraph(title, ParagraphStyle(
        "h", fontSize=16, leading=20, fontName="Helvetica-Bold",
        alignment=1)))
    # Iter 477 — optional statutory FORM heading line(s) between the title
    # and the subtitle ("\n"-separated).
    for ln in (form_line or "").split("\n"):
        if not ln.strip():
            continue
        story.append(Paragraph(ln.strip(), ParagraphStyle(
            "f", fontSize=11, leading=14, fontName="Helvetica-Bold",
            alignment=1)))
    story.append(Paragraph(subtitle, ParagraphStyle(
        "s", fontSize=10, leading=13, alignment=1)))
    story.append(Spacer(1, 4 * mm))
    # Iter 432 (user request) — when the period has NO data (e.g. Fine
    # Register with no fines), print a clear centred line instead of an
    # empty table.
    if not rows and empty_note:
        story.append(Spacer(1, 18 * mm))
        story.append(Paragraph(empty_note, ParagraphStyle(
            "e", fontSize=14, fontName="Helvetica-Bold", alignment=1,
            textColor=rl.HexColor("#475569"))))
        doc.build(story)
        buf.seek(0)
        return buf
    data = [[c["label"] for c in columns]]
    for row in rows:
        data.append([
            (f"{v:,.2f}".rstrip("0").rstrip(".") if isinstance(v, float)
             else ("" if v is None else str(v)))
            for v in (row.get(c["key"]) for c in columns)])
    # Iter 520 (user bug — wage register heading/data OVERLAPPING) —
    # WRAP text instead of letting it draw over the next column:
    #   • every header label becomes a wrapping Paragraph;
    #   • long non-numeric data cells wrap too (names, bank A/c + IFSC);
    #   • font auto-shrinks when the register has many columns.
    n_cols = max(1, len(columns))
    base_fs = 10 if n_cols <= 12 else (8.5 if n_cols <= 18 else 7.5)
    _head_st = ParagraphStyle(
        "th", fontSize=base_fs, leading=base_fs + 1.5,
        fontName="Helvetica-Bold", alignment=1)
    _cell_c = ParagraphStyle("tdc", fontSize=base_fs,
                             leading=base_fs + 1.5, alignment=1)
    _cell_l = ParagraphStyle("tdl", fontSize=base_fs,
                             leading=base_fs + 1.5, alignment=0)

    def _is_numlike(s: str) -> bool:
        return (s.replace(",", "").replace(".", "")
                .replace("-", "").replace("%", "").isdigit())

    data[0] = [Paragraph(str(c["label"]), _head_st) for c in columns]
    for ri in range(1, len(data)):
        for j, c in enumerate(columns):
            s = data[ri][j]
            if isinstance(s, str) and len(s) > 11 and not _is_numlike(s):
                _left = "name" in str(c.get("key") or "").lower()
                data[ri][j] = Paragraph(s, _cell_l if _left else _cell_c)
    # Iter 432 (user request) — bigger print font, FIGURES CENTRED with a
    # tighter column gap; the Employee Name column is RIGHT-aligned.
    styles = [
        ("FONTSIZE", (0, 0), (-1, -1), base_fs),
        ("GRID", (0, 0), (-1, -1), 0.4, rl.HexColor("#9AA0A6")),
        ("BACKGROUND", (0, 0), (-1, 0), rl.HexColor("#DDEBF7")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    # Iter 435 (user request) — Name columns LEFT-aligned in all reports.
    for j, c in enumerate(columns):
        if "name" in str(c.get("key") or "").lower():
            styles.append(("ALIGN", (j, 1), (j, -1), "LEFT"))
    if totals:
        _has_lbl = any(str(v).strip().upper() == "TOTAL"
                       for v in totals.values())
        data.append([
            (f"{v:,.2f}".rstrip("0").rstrip(".") if isinstance(v, float)
             else ("" if v is None else str(v)))
            for v in (totals.get(c["key"],
                                 "TOTAL" if i == 0 and not _has_lbl else None)
                      for i, c in enumerate(columns))])
        styles += [
            ("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1),
             rl.HexColor("#FFF2CC")),
            ("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1),
             "Helvetica-Bold")]
    # Narrow serial / code columns, wider name + bank columns.
    weights = []
    for c in columns:
        k = str(c.get("key") or "").lower()
        lbl = str(c.get("label") or "").lower()
        if k in ("sno", "s_no", "serial"):
            weights.append(0.45)
        elif "name" in k:
            weights.append(1.7)
        elif "bank" in k or "ifsc" in k or "bank" in lbl or "ifsc" in lbl:
            weights.append(1.5)  # Iter 520 — Bank A/c / IFSC needs room
        else:
            weights.append(1.0)
    tw = sum(weights)
    tbl = Table(data, colWidths=[W * w / tw for w in weights], repeatRows=1)
    tbl.setStyle(TableStyle(styles))
    story.append(tbl)

    def _page(cv, d):
        cv.setFont("Helvetica", 8)
        from reportlab.lib.pagesizes import A3 as _A3, landscape as _ls
        cv.drawRightString(_ls(_A3)[0] - 6 * mm, 5 * mm,
                           f"Page {cv.getPageNumber()}")
    doc.build(story, onFirstPage=_page, onLaterPages=_page)
    buf.seek(0)
    return buf
