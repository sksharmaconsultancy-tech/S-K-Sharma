"""Master Sheet Automation — Iter 58.

Automates the Super Admin's monthly payroll process for S.K. Sharma & Co.:

  1. Generate a per-client Excel "master sheet" seeded with the employee
     roster + attendance snapshot for the month. The client fills in the
     blanks (Gross Salary / Advance / TDS) and returns it.
  2. Ingest that filled sheet — either in our exact template OR any random
     format the client sends — and produce a *column-matching MIS report*
     so the super admin can review & confirm mappings before import.
  3. Persist the imported data on the employee doc for the current month,
     ready to be piped into the Compliance Salary Process (Iter 56).
  4. Generate ECR (EPF) & ESIC challan files in the government-portal
     formats for the super admin to download and submit.

Random-format sheets are handled via a fuzzy column matcher powered by
`rapidfuzz`. Each column of the uploaded file is scored against a set of
"synonyms" for our canonical fields and the highest-scoring candidate
becomes the suggested mapping.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process


# ---------------------------------------------------------------------------
# Canonical fields + synonyms for the column matcher
# ---------------------------------------------------------------------------
CANONICAL_FIELDS: Dict[str, Dict[str, Any]] = {
    "employee_code": {
        "label": "Employee Code",
        "required": True,
        "synonyms": [
            "employee code", "emp code", "emp_code", "empcode", "employee id",
            "emp id", "empno", "code", "emp no", "employee number",
        ],
    },
    "name": {
        "label": "Employee Name",
        "required": True,
        "synonyms": [
            "name", "employee name", "emp name", "full name", "employee",
        ],
    },
    "doj": {
        "label": "Date of Joining",
        "required": False,
        "synonyms": [
            "doj", "date of joining", "joining date", "join date",
            "date joined",
        ],
    },
    "department": {
        "label": "Department",
        "required": False,
        "synonyms": [
            "department", "dept", "section", "division",
        ],
    },
    "designation": {
        "label": "Designation",
        "required": False,
        "synonyms": [
            "designation", "role", "position", "title", "job title", "post",
        ],
    },
    "employee_group": {
        "label": "Employee Group",
        "required": False,
        "synonyms": [
            "group", "employee group", "site", "location", "branch", "unit",
        ],
    },
    "phone": {
        "label": "Phone",
        "required": False,
        "synonyms": [
            "phone", "mobile", "contact", "phone number", "mobile number", "contact number",
        ],
    },
    "email": {
        "label": "Email",
        "required": False,
        "synonyms": [
            "email", "email id", "mail", "email address",
        ],
    },
    "days_worked": {
        "label": "Days Worked",
        "required": False,
        "synonyms": [
            "days worked", "days present", "present days", "pd", "days",
            "attendance days", "working days",
        ],
    },
    "gross_salary": {
        "label": "Gross Salary",
        "required": True,
        "synonyms": [
            "gross salary", "gross", "salary", "monthly salary", "ctc",
            "gross pay", "gross amount", "monthly gross",
        ],
    },
    "advance": {
        "label": "Advance",
        "required": False,
        "synonyms": [
            "advance", "loan", "advance amount", "loan amount",
            "salary advance", "adv",
        ],
    },
    "tds": {
        "label": "TDS",
        "required": False,
        "synonyms": [
            "tds", "income tax", "tax", "tds amount", "tds deduction",
        ],
    },
    "notes": {
        "label": "Notes",
        "required": False,
        "synonyms": [
            "notes", "remarks", "comment", "comments", "note",
        ],
    },
}


BRAND_TEAL = "1F4E4E"
BRAND_GOLD = "C89B3C"
BRAND_INK = "1E2A2A"
BG_SOFT = "F7F9F9"
LINE = "D6DEDE"


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------
def build_master_sheet_xlsx(
    *,
    company_name: str,
    month: str,   # "YYYY-MM"
    employees: List[Dict[str, Any]],
    attendance_days_by_user: Optional[Dict[str, int]] = None,
) -> bytes:
    """Return XLSX bytes for a pre-populated master sheet.

    Blank (client-fill) columns: Gross Salary, Advance, TDS, Notes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Sheet"

    # Title band
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = f"{company_name}  ·  Master Salary Sheet  ·  {month}"
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BRAND_TEAL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = (
        "Please fill Gross Salary, Advance and TDS columns for each employee, "
        "then email the completed sheet back to us for processing."
    )
    sub.font = Font(name="Calibri", size=9, italic=True, color=BRAND_INK)
    sub.fill = PatternFill("solid", fgColor=BG_SOFT)
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    headers = [
        "Employee Code", "Employee Name", "Date of Joining", "Department",
        "Days Worked", "Gross Salary", "Advance", "TDS", "Notes",
    ]
    thin = Side(border_style="thin", color=LINE)
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND_TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws.row_dimensions[3].height = 28

    # Data rows
    for r, emp in enumerate(employees, start=4):
        days_worked = (attendance_days_by_user or {}).get(emp.get("user_id"))
        row = [
            emp.get("employee_code") or "",
            emp.get("name") or "",
            (emp.get("doj") or "")[:10],
            emp.get("department") or "",
            days_worked if days_worked is not None else "",
            "",  # Gross Salary — blank
            "",  # Advance — blank
            "",  # TDS — blank
            "",  # Notes — blank
        ]
        for cidx, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=cidx, value=val)
            c.alignment = Alignment(horizontal="left" if cidx <= 4 else "right",
                                    vertical="center")
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor=BG_SOFT)

    # Column widths
    widths = [16, 30, 14, 18, 12, 14, 12, 12, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Footer band
    footer_row = 4 + len(employees) + 1
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=9)
    fc = ws.cell(row=footer_row, column=1)
    fc.value = f"Generated on {datetime.now(timezone.utc).strftime('%d %b %Y')}  ·  {len(employees)} employees"
    fc.font = Font(size=8, italic=True, color=BRAND_INK)
    fc.fill = PatternFill("solid", fgColor=BG_SOFT)
    fc.alignment = Alignment(horizontal="right")

    # Freeze the header
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Ingest / MIS column-matching
# ---------------------------------------------------------------------------
def parse_uploaded_xlsx(xlsx_bytes: bytes) -> Tuple[List[str], List[List[Any]]]:
    """Return (headers, data_rows) from a raw xlsx. Uses the first non-empty
    row as the header. Skips totally-blank leading rows."""
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    header_idx = -1
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        non_empty = sum(1 for c in cells if c)
        if non_empty >= 2:
            header_idx = i
            break
    if header_idx == -1:
        return [], []

    headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    body = []
    for row in rows[header_idx + 1:]:
        vals = list(row)
        # Trim rows that are entirely empty
        if not any((v is not None and str(v).strip() != "") for v in vals):
            continue
        body.append(vals)
    return headers, body


def match_columns(headers: List[str]) -> Dict[str, Any]:
    """For each canonical field, find the best-matching header and confidence.

    Returns:
        {
            "matches": [
                {"canonical": "gross_salary", "canonical_label": "Gross Salary",
                 "matched_header": "GROSS", "matched_index": 5,
                 "confidence": 92, "required": True}, ...
            ],
            "unmatched_required": [...],
            "unrecognised_headers": [...],
        }
    """
    normalised = [(h or "").strip() for h in headers]
    lower = [h.lower() for h in normalised]
    used_indices: set = set()

    matches = []
    unmatched_required: List[str] = []
    for canonical, cfg in CANONICAL_FIELDS.items():
        synonyms = cfg["synonyms"]
        best_score = 0
        best_idx: Optional[int] = None
        for idx, h in enumerate(lower):
            if not h or idx in used_indices:
                continue
            # Score against each synonym; take the max
            score = max(
                (fuzz.token_set_ratio(h, s) for s in synonyms),
                default=0,
            )
            if score > best_score:
                best_score = score
                best_idx = idx
        confident = best_score >= 65
        if confident and best_idx is not None:
            used_indices.add(best_idx)
            matches.append({
                "canonical": canonical,
                "canonical_label": cfg["label"],
                "required": cfg["required"],
                "matched_header": normalised[best_idx] if best_idx is not None else None,
                "matched_index": best_idx,
                "confidence": int(best_score),
            })
        else:
            if cfg["required"]:
                unmatched_required.append(cfg["label"])
            matches.append({
                "canonical": canonical,
                "canonical_label": cfg["label"],
                "required": cfg["required"],
                "matched_header": None,
                "matched_index": None,
                "confidence": int(best_score),
            })
    unrecognised = [
        {"index": i, "header": normalised[i]}
        for i in range(len(normalised))
        if normalised[i] and i not in used_indices
    ]
    return {
        "matches": matches,
        "unmatched_required": unmatched_required,
        "unrecognised_headers": unrecognised,
    }


def import_rows_via_mapping(
    headers: List[str],
    body: List[List[Any]],
    mapping: Dict[str, int],
) -> List[Dict[str, Any]]:
    """Apply a { canonical_field → column_index } mapping to the body and
    return a list of canonical-field dicts. Empty rows are dropped."""
    out: List[Dict[str, Any]] = []
    for row in body:
        rec: Dict[str, Any] = {}
        for canonical, col_idx in mapping.items():
            if col_idx is None or col_idx < 0 or col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            # Numeric-ish
            if canonical in ("gross_salary", "advance", "tds", "days_worked"):
                try:
                    rec[canonical] = float(str(val).replace(",", "").strip() or 0)
                except (TypeError, ValueError):
                    continue
            else:
                rec[canonical] = str(val).strip()
        # Rows without an employee_code AND name are useless
        if not rec.get("employee_code") and not rec.get("name"):
            continue
        out.append(rec)
    return out


# ---------------------------------------------------------------------------
# ECR (EPF) file generator — plain text, tab-separated, EPFO ECR 2.0 format
# ---------------------------------------------------------------------------
def build_ecr_text(compliance_run: Dict[str, Any]) -> bytes:
    """Return the ECR (Electronic Challan cum Return) plain-text file matching
    the EPFO portal upload format.

    Each row is `#` (hash) separated with 11 fields:
      UAN, MemberName, GrossWages, EPFWages, EPSWages, EDLIWages,
      EPFContribRemitted, EPSContribRemitted, EPFEPSDiffRemitted,
      NCPDays, RefundOfAdvances

    Any employee with pf_applicable=False is skipped.
    """
    lines: List[str] = []
    for r in (compliance_run.get("rows") or []):
        if not r.get("pf_applicable"):
            continue
        uan = str(r.get("uan") or r.get("employee_code") or "").strip() or "000000000000"
        name = (r.get("name") or "").upper()[:60]
        gross = round(r.get("gross_paid") or 0)
        epf_wages = round(r.get("pf_wages") or 0)
        eps_wages = epf_wages  # same base under new labour code
        edli_wages = epf_wages
        epf_contrib = round(r.get("pf_employee") or 0) + round(r.get("pf_employer_epf") or 0)
        eps_contrib = round(r.get("pf_employer_eps") or 0)
        epf_eps_diff = max(0, round(r.get("pf_employer_epf") or 0) - eps_contrib)
        ncp_days = max(0, int(r.get("month_days") or 30) - int(r.get("present_days") or 0))
        refund = 0
        lines.append(
            f"{uan}#{name}#{gross}#{epf_wages}#{eps_wages}#{edli_wages}#"
            f"{epf_contrib}#{eps_contrib}#{epf_eps_diff}#{ncp_days}#{refund}"
        )
    return ("\n".join(lines) + "\n").encode("utf-8")


# ---------------------------------------------------------------------------
# ESIC challan file — XLSX matching ESIC portal upload format
# ---------------------------------------------------------------------------
def build_esic_xlsx(compliance_run: Dict[str, Any]) -> bytes:
    """Return an XLSX file matching the ESIC portal's Monthly Contribution
    upload template. Columns: IP Number, Name, Days Worked, Total Monthly
    Wages, Reason Code (only if applicable), Last Working Day.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ESIC Contribution"

    headers = [
        "IP Number", "IP Name", "No. of Days for which Wages Paid/Payable",
        "Total Monthly Wages", "Reason Code for Zero Working Days",
        "Last Working Day",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND_TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    r = 2
    for row in (compliance_run.get("rows") or []):
        if not row.get("esic_applicable"):
            continue
        ip_number = (row.get("ip_number") or row.get("employee_code") or "").strip()
        name = (row.get("name") or "").upper()
        days = int(row.get("present_days") or 0)
        wages = round(row.get("esic_wage_base") or 0)
        reason_code = ""  # empty when days > 0
        last_working = ""
        ws.cell(row=r, column=1, value=ip_number)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=days)
        ws.cell(row=r, column=4, value=wages)
        ws.cell(row=r, column=5, value=reason_code)
        ws.cell(row=r, column=6, value=last_working)
        r += 1

    for i, w in enumerate([18, 30, 12, 16, 10, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
