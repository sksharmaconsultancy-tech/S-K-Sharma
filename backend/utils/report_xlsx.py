"""Shared XLSX builder for report exports (Iter 64).

We use ``openpyxl`` (already installed for the attendance sheet flow) so
Salary and Compliance reports can now be downloaded natively as Excel
files rather than CSV. The layout mirrors the CSV column order for
consistency, adds a bold header row with an accent fill, freezes the
top row, and applies column-width auto-sizing based on content.

Numeric currency-style columns are given a ``#,##0.00`` number format so
they open correctly in Excel with locale-safe formatting.
"""
from __future__ import annotations

import io
from typing import Any, Dict, Iterable, List, Optional, Sequence


# Column keys we should format as numbers with 2-decimal precision.
_NUMERIC_COLS = {
    "rate", "month_days", "present_days", "half_days", "duty_hours",
    "ot_hours", "base_pay", "bonus", "ot_pay", "gross", "advance",
    "total_deduction", "net",
    # Compliance-specific
    "pf_wage_base", "pf_employee", "pf_employer_epf", "pf_employer_eps",
    "pf_admin_charges", "esic_wage_base", "esic_employee",
    "esic_employer", "pt", "tds",
}

# Booleans/flags we render as readable strings.
_BOOL_COLS = {"is_onroll": ("On-roll", "Off-roll"),
              "pf_applicable": ("Yes", "No"),
              "esic_applicable": ("Yes", "No")}


def build_rows_xlsx(
    columns: Sequence[str],
    rows: Iterable[Dict[str, Any]],
    sheet_name: str = "Report",
    title: Optional[str] = None,
    subtitle: Optional[str] = None,
) -> bytes:
    """Return an .xlsx byte-string for the given rows.

    Parameters
    ----------
    columns:
        Ordered list of column keys to project from each row.
    rows:
        Iterable of dicts. Missing keys become empty cells.
    sheet_name:
        Excel sheet tab title (<= 31 chars).
    title:
        Optional bold header line printed on row 1 (spanned across
        columns). Data starts on the next row when supplied.
    subtitle:
        Optional second header line (italic, smaller).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Report")[:31]

    accent = PatternFill("solid", fgColor="1F3D7A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1F3D7A")
    subtitle_font = Font(italic=True, size=10, color="555555")

    cur_row = 1
    if title:
        cell = ws.cell(row=cur_row, column=1, value=title)
        cell.font = title_font
        ws.merge_cells(
            start_row=cur_row, start_column=1,
            end_row=cur_row, end_column=max(1, len(columns)),
        )
        cur_row += 1
    if subtitle:
        cell = ws.cell(row=cur_row, column=1, value=subtitle)
        cell.font = subtitle_font
        ws.merge_cells(
            start_row=cur_row, start_column=1,
            end_row=cur_row, end_column=max(1, len(columns)),
        )
        cur_row += 1

    # Header row
    header_row = cur_row
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=_pretty_col(col))
        cell.font = header_font
        cell.fill = accent
        cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Data rows
    data_start = header_row + 1
    rows_list = list(rows)
    for r_idx, row in enumerate(rows_list, start=data_start):
        for c_idx, col in enumerate(columns, start=1):
            raw = row.get(col, "")
            if col in _BOOL_COLS and isinstance(raw, bool):
                on, off = _BOOL_COLS[col]
                cell = ws.cell(row=r_idx, column=c_idx, value=on if raw else off)
            elif col in _NUMERIC_COLS:
                try:
                    val = float(raw) if raw not in (None, "") else None
                except (TypeError, ValueError):
                    val = None
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.number_format = "#,##0.00"
            else:
                cell = ws.cell(row=r_idx, column=c_idx, value=(raw if raw is not None else ""))

    # Auto column widths (rough heuristic: max content length + 2 padding)
    for i, col in enumerate(columns, start=1):
        header = _pretty_col(col)
        max_len = len(header)
        for row in rows_list:
            v = row.get(col, "")
            s = "" if v is None else str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)

    # Add totals row for numeric columns (only if we have data)
    if rows_list:
        total_row = data_start + len(rows_list)
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for i, col in enumerate(columns, start=1):
            if col not in _NUMERIC_COLS:
                continue
            total = 0.0
            for row in rows_list:
                try:
                    total += float(row.get(col, 0) or 0)
                except (TypeError, ValueError):
                    pass
            cell = ws.cell(row=total_row, column=i, value=round(total, 2))
            cell.number_format = "#,##0.00"
            cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pretty_col(key: str) -> str:
    """Turn ``employee_code`` into ``Employee Code`` for headers."""
    return " ".join(w.capitalize() for w in (key or "").split("_"))
