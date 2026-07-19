"""Monthly Attendance report builders (Iter 68).

Two XLSX outputs, both structured like the reference sheet the user provided:

  • ``build_monthly_hours_xlsx``  — one cell per day showing total working
    hours (matches the reference "Monthly Attendance Data.xls").
  • ``build_monthly_inout_xlsx``  — one cell per day showing IN time, OUT
    time and working hours (three text lines).

Both share:
  • Sticky columns: Employee Name · Emp Code · Father Name · Department · Designation
  • 1 column per day of the requested month, labelled ``<day> <weekday>``
    (e.g. "1 Mo", "2 Tu", …).
  • Trailing summary columns: Extra Hrs · Work Hrs · Week Off Hrs · GP Hrs ·
    MOT Hrs · OT Hrs · Lost Hrs · PL Days · CL Days · Remaining Hrs ·
    Tot. Days · Tot. Hrs.
  • Column-total footer row summing each day column across employees.

Working hours are computed as the sum of ``out - in`` pairs on the same day
(paired chronologically). If a punch is unmatched (only IN or only OUT) it
is ignored for hour totals but the pill still shows the raw punch time so
supervisors can see the gap.

Precision: sample shows ``12.0`` for a round shift and ``12.3`` for a
12h30m shift — so we render hours as ``H`` + ``.{floor(minutes/10)}``.
For the IN/OUT sheet, times are shown as ``HH:MM`` (24h) and hours below
each day are shown as ``HHh MMm`` for readability.
"""
from __future__ import annotations

import calendar
import io
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WEEKDAY_SHORT = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]

# Column styles ------------------------------------------------------------
_HDR_FILL = PatternFill("solid", fgColor="1F5254")   # deep teal
_HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
_ZEBRA_FILL = PatternFill("solid", fgColor="F5F7F8")
_TOTAL_FILL = PatternFill("solid", fgColor="E6EDED")
_TOTAL_FONT = Font(bold=True, color="0F3D3E", size=10)
_BORDER = Border(
    left=Side(style="thin", color="D6DCDC"),
    right=Side(style="thin", color="D6DCDC"),
    top=Side(style="thin", color="D6DCDC"),
    bottom=Side(style="thin", color="D6DCDC"),
)
_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _month_bounds(month: str) -> Tuple[int, int, int]:
    """Return (year, month, days_in_month) for a ``YYYY-MM`` string."""
    y, m = int(month[:4]), int(month[5:7])
    days = calendar.monthrange(y, m)[1]
    return y, m, days


def _fmt_hours_sample(minutes_total: int) -> str:
    """Format total minutes as ``HH:MM`` (e.g. 594 min -> "09:54").
    Grid View uses the same convention so both surfaces stay in sync.
    """
    if minutes_total <= 0:
        return "0:00"
    h = minutes_total // 60
    m = minutes_total % 60
    return f"{h:02d}:{m:02d}"


def _fmt_hours_hm(minutes_total: int) -> str:
    """Format total minutes as ``HH:MM`` for the IN/OUT sheet."""
    return _fmt_hours_sample(minutes_total)


def _parse_iso(v: Any) -> Optional[datetime]:
    """Parse an ISO timestamp field (str or datetime) to a UTC datetime."""
    if not v:
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    if isinstance(v, str):
        try:
            dt = datetime.fromisoformat(v.replace("Z", "+00:00"))
            return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        except ValueError:
            return None
    return None


def _fmt_time_ist(dt: Optional[datetime]) -> str:
    """Format the stored datetime as ``HH:MM``.

    Historical note: earlier biometric imports (source ``import:zk_*``) were
    stored with a mislabeled ``+00:00`` offset - the underlying ``H:M`` is
    already the local IST time the device recorded. Doing another ``+5:30``
    shift makes Excel disagree with the Grid View. To keep the two surfaces
    in lockstep we now format the raw H/M as-is, matching the on-screen
    ``strftime("%H:%M")`` used by the grid endpoint.
    """
    if dt is None:
        return "-"
    return dt.strftime("%H:%M")


def _pair_punches(day_punches: List[Dict[str, Any]]) -> Tuple[int, Optional[datetime], Optional[datetime]]:
    """Given a chronologically-sorted list of punches for one employee-day,
    return (total_worked_minutes, earliest_in, latest_out).

    Punches are paired IN → OUT.  Unmatched punches don't add to hours but
    do influence the earliest-in / latest-out returned so the pill on the
    IN/OUT sheet still shows the raw stamps."""
    total_min = 0
    earliest_in: Optional[datetime] = None
    latest_out: Optional[datetime] = None
    open_in: Optional[datetime] = None
    for p in day_punches:
        kind = (p.get("kind") or "").lower()
        at = _parse_iso(p.get("at"))
        if at is None:
            continue
        if kind == "in":
            if earliest_in is None or at < earliest_in:
                earliest_in = at
            if open_in is None:
                open_in = at
            else:
                # Two INs in a row — keep the first, drop the second.
                pass
        elif kind == "out":
            if latest_out is None or at > latest_out:
                latest_out = at
            if open_in is not None:
                delta = int((at - open_in).total_seconds() // 60)
                if delta > 0:
                    total_min += delta
                open_in = None
    return total_min, earliest_in, latest_out


def _summary_columns_hours() -> List[str]:
    return [
        "Extra  Hrs", "Work Hrs", "Week Off Hrs", "GP Hrs",
        "MOT Hrs", "OT Hrs", "Lost Hrs", "PL Days", "CL Days",
        "Remaining Hrs", "Present Days", "Tot. Hrs",
    ]


def _weekday_short(y: int, m: int, d: int) -> str:
    return WEEKDAY_SHORT[datetime(y, m, d).weekday()]


# ---------------------------------------------------------------------------
# Common shell builder
# ---------------------------------------------------------------------------

def _write_common_header(
    ws,
    company_name: str,
    month: str,
    y: int,
    m: int,
    days: int,
    sub_title: str,
) -> int:
    """Write the top title + column headers.  Returns the first data row."""
    total_cols = 5 + days + len(_summary_columns_hours())

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1).value = f"{company_name} — {sub_title}"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="0F3D3E")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1).value = f"Month: {month} ({days} days) · Generated: {datetime.now(timezone(timedelta(hours=5, minutes=30))).strftime('%d-%b-%Y %H:%M IST')}"
    ws.cell(row=2, column=1).font = Font(italic=True, color="475569", size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    hdr = ["Employee Name", "Emp  Code", "Father Name", "Department", "Designation"]
    for d in range(1, days + 1):
        hdr.append(f"{d} {_weekday_short(y, m, d)}")
    hdr.extend(_summary_columns_hours())

    header_row = 4
    for i, label in enumerate(hdr, start=1):
        c = ws.cell(row=header_row, column=i)
        c.value = label
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = _ALIGN_C
        c.border = _BORDER
    ws.row_dimensions[header_row].height = 32

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    for d in range(1, days + 1):
        ws.column_dimensions[get_column_letter(5 + d)].width = 8
    for k, _ in enumerate(_summary_columns_hours()):
        ws.column_dimensions[get_column_letter(5 + days + 1 + k)].width = 11
    ws.freeze_panes = f"F{header_row + 1}"
    return header_row + 1


# ---------------------------------------------------------------------------
# Report 1 — Monthly Working Hours (mirrors user sample)
# ---------------------------------------------------------------------------

def build_monthly_hours_xlsx(
    *,
    company_name: str,
    month: str,
    employees: List[Dict[str, Any]],
    punches_by_user_day: Dict[str, Dict[str, List[Dict[str, Any]]]],
    weekly_off: Iterable[int] = (6,),  # Sunday by default
) -> bytes:
    """Build the Monthly Attendance Data XLSX (hours-only cells)."""
    y, m, days = _month_bounds(month)
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Hours"

    data_row = _write_common_header(
        ws, company_name, month, y, m, days,
        sub_title="Monthly Attendance Data (Working Hours)",
    )

    # Per-day column totals across employees (in minutes)
    day_col_totals_min: List[int] = [0] * days

    for idx, emp in enumerate(employees):
        row = data_row + idx
        uid = emp.get("user_id") or emp.get("_id")
        by_day = punches_by_user_day.get(uid, {})

        ws.cell(row=row, column=1, value=emp.get("name") or "")
        ws.cell(row=row, column=2, value=emp.get("employee_code") or "")
        ws.cell(row=row, column=3, value=emp.get("father_name") or "")
        ws.cell(row=row, column=4, value=emp.get("department") or "")
        ws.cell(row=row, column=5, value=emp.get("position") or "")

        tot_min = 0
        working_min = 0
        weekoff_min = 0
        ot_min = 0
        present_days = 0
        for d in range(1, days + 1):
            date_key = f"{y:04d}-{m:02d}-{d:02d}"
            day_min, _in, _out = _pair_punches(by_day.get(date_key, []))
            cell = ws.cell(row=row, column=5 + d)
            cell.value = _fmt_hours_sample(day_min)
            cell.alignment = _ALIGN_C
            cell.border = _BORDER
            # Zebra striping
            if idx % 2 == 1:
                cell.fill = _ZEBRA_FILL
            day_col_totals_min[d - 1] += day_min
            tot_min += day_min
            if day_min > 0:
                present_days += 1
                weekday = datetime(y, m, d).weekday()
                if weekday in weekly_off:
                    weekoff_min += day_min
                else:
                    working_min += day_min
                # Overtime = anything above 8h in a shift day
                if day_min > 8 * 60:
                    ot_min += day_min - 8 * 60

        # Zebra on non-day cells too
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.alignment = _ALIGN_L
            cell.border = _BORDER
            if idx % 2 == 1:
                cell.fill = _ZEBRA_FILL

        # Summary columns
        summary_vals = [
            0.0,                              # Extra Hrs — reserved
            _fmt_hours_sample(working_min),   # Work Hrs
            _fmt_hours_sample(weekoff_min),   # Week Off Hrs
            0.0,                              # GP Hrs — reserved
            0.0,                              # MOT Hrs — reserved
            _fmt_hours_sample(ot_min),        # OT Hrs
            0.0,                              # Lost Hrs — reserved
            0.0,                              # PL Days
            0.0,                              # CL Days
            _fmt_hours_sample(tot_min % 60 if tot_min else 0),  # Remaining Hrs (minutes fraction)
            present_days,                     # Tot. Days
            _fmt_hours_sample(tot_min),       # Tot. Hrs
        ]
        for k, v in enumerate(summary_vals):
            cell = ws.cell(row=row, column=5 + days + 1 + k, value=v)
            cell.alignment = _ALIGN_C
            cell.border = _BORDER
            if idx % 2 == 1:
                cell.fill = _ZEBRA_FILL

        ws.row_dimensions[row].height = 22

    # Footer totals row (per-day column sums)
    footer_row = data_row + len(employees)
    ws.cell(row=footer_row, column=1, value=str(len(employees))).font = _TOTAL_FONT
    for c in range(1, 6):
        ws.cell(row=footer_row, column=c).fill = _TOTAL_FILL
        ws.cell(row=footer_row, column=c).border = _BORDER
    for d in range(days):
        cell = ws.cell(row=footer_row, column=5 + d + 1, value=_fmt_hours_sample(day_col_totals_min[d]))
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.alignment = _ALIGN_C
        cell.border = _BORDER
    for k in range(len(_summary_columns_hours())):
        cell = ws.cell(row=footer_row, column=5 + days + 1 + k)
        cell.fill = _TOTAL_FILL
        cell.border = _BORDER
    ws.row_dimensions[footer_row].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Report 2 — Monthly IN / OUT + Working Hours
# ---------------------------------------------------------------------------

def build_monthly_inout_xlsx(
    *,
    company_name: str,
    month: str,
    employees: List[Dict[str, Any]],
    punches_by_user_day: Dict[str, Dict[str, List[Dict[str, Any]]]],
    weekly_off: Iterable[int] = (6,),
) -> bytes:
    """Build the Monthly IN / OUT + Working Hrs XLSX.  Same layout as the
    hours report but each day cell shows 3 lines::

        09:00 IN
        21:00 OUT
        12h 00m
    """
    y, m, days = _month_bounds(month)
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly In-Out"

    data_row = _write_common_header(
        ws, company_name, month, y, m, days,
        sub_title="Monthly Attendance IN / OUT + Working Hours",
    )

    day_col_totals_min: List[int] = [0] * days

    for idx, emp in enumerate(employees):
        row = data_row + idx
        uid = emp.get("user_id") or emp.get("_id")
        by_day = punches_by_user_day.get(uid, {})

        ws.cell(row=row, column=1, value=emp.get("name") or "")
        ws.cell(row=row, column=2, value=emp.get("employee_code") or "")
        ws.cell(row=row, column=3, value=emp.get("father_name") or "")
        ws.cell(row=row, column=4, value=emp.get("department") or "")
        ws.cell(row=row, column=5, value=emp.get("position") or "")

        tot_min = 0
        working_min = 0
        weekoff_min = 0
        ot_min = 0
        present_days = 0
        for d in range(1, days + 1):
            date_key = f"{y:04d}-{m:02d}-{d:02d}"
            day_punches = by_day.get(date_key, [])
            day_min, in_dt, out_dt = _pair_punches(day_punches)
            in_txt = _fmt_time_ist(in_dt) if in_dt else "—"
            out_txt = _fmt_time_ist(out_dt) if out_dt else "—"
            hrs_txt = _fmt_hours_hm(day_min)
            # Multi-line cell content
            cell_val = f"IN {in_txt}\nOUT {out_txt}\n{hrs_txt}" if (in_dt or out_dt) else "—"
            cell = ws.cell(row=row, column=5 + d, value=cell_val)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
            )
            cell.font = Font(size=8)
            cell.border = _BORDER
            if idx % 2 == 1:
                cell.fill = _ZEBRA_FILL
            day_col_totals_min[d - 1] += day_min
            tot_min += day_min
            if day_min > 0:
                present_days += 1
                if datetime(y, m, d).weekday() in weekly_off:
                    weekoff_min += day_min
                else:
                    working_min += day_min
                if day_min > 8 * 60:
                    ot_min += day_min - 8 * 60

        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.alignment = _ALIGN_L
            cell.border = _BORDER
            if idx % 2 == 1:
                cell.fill = _ZEBRA_FILL

        summary_vals = [
            0.0,
            _fmt_hours_sample(working_min),
            _fmt_hours_sample(weekoff_min),
            0.0,
            0.0,
            _fmt_hours_sample(ot_min),
            0.0,
            0.0,
            0.0,
            _fmt_hours_sample(tot_min % 60 if tot_min else 0),
            present_days,
            _fmt_hours_sample(tot_min),
        ]
        for k, v in enumerate(summary_vals):
            cell = ws.cell(row=row, column=5 + days + 1 + k, value=v)
            cell.alignment = _ALIGN_C
            cell.border = _BORDER
            if idx % 2 == 1:
                cell.fill = _ZEBRA_FILL

        # Bigger row height for 3-line day cells
        ws.row_dimensions[row].height = 44

    # Footer totals
    footer_row = data_row + len(employees)
    ws.cell(row=footer_row, column=1, value=str(len(employees))).font = _TOTAL_FONT
    for c in range(1, 6):
        ws.cell(row=footer_row, column=c).fill = _TOTAL_FILL
        ws.cell(row=footer_row, column=c).border = _BORDER
    for d in range(days):
        cell = ws.cell(
            row=footer_row, column=5 + d + 1,
            value=_fmt_hours_sample(day_col_totals_min[d]),
        )
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.alignment = _ALIGN_C
        cell.border = _BORDER
    for k in range(len(_summary_columns_hours())):
        cell = ws.cell(row=footer_row, column=5 + days + 1 + k)
        cell.fill = _TOTAL_FILL
        cell.border = _BORDER
    ws.row_dimensions[footer_row].height = 24

    # Widen day columns to fit 3 lines of text
    for d in range(1, days + 1):
        ws.column_dimensions[get_column_letter(5 + d)].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ---------------------------------------------------------------------------
# Report 3 — Grid View XLSX (Iter 77x)
# ---------------------------------------------------------------------------
# Multi-row per-employee layout that mirrors the exact sample sheet the user
# supplied:
#
#   Emp Name | Emp Father Name | Designation | Type | D1 Mo | D2 Tu | ...
#             |                |             | D-In | 07:55 |       | ...
#             |                |             | D-Out| 19:58 |       | ...
#             |                |             | OT-In| ...   | (only if OT)
#             |                |             |OT-Out| ...   | (only if OT)
#             |                |             | T-Hrs| 12:03 | 00:00 | ...
#
# Trailing summary columns per row:
#     • Tot Wrk Hrs   (decimal hours; blank on D-In / OT-In rows)
#     • Day Total     (present-days count; blank on D-In / OT-In rows)
#
# The builder DOES NOT recompute punches — it consumes the same grid data
# returned by ``/api/admin/attendance/monthly-grid`` so what you see in the
# Grid View is exactly what you get in Excel (bounce-merge, 15-min dedup,
# OT cap, weekly-off rules, etc. all honored upstream).
# ---------------------------------------------------------------------------


def build_grid_view_xlsx(grid: Dict[str, Any]) -> bytes:
    """Build the multi-row Grid-View XLSX from a JSON grid payload.

    Parameters
    ----------
    grid:
        The exact dict returned by ``monthly_attendance_grid_json``. Must
        contain ``company``, ``day_labels``, ``day_full_dates``,
        ``weekday_labels`` and ``employees`` (each with ``days`` cells and
        ``totals``).
    """
    company_name = ((grid or {}).get("company") or {}).get("name") or ""
    month = grid.get("month") or ""
    day_labels: List[str] = list(grid.get("day_labels") or [])
    weekday_labels: List[str] = list(grid.get("weekday_labels") or [])
    employees: List[Dict[str, Any]] = list(grid.get("employees") or [])
    days_n = len(day_labels)

    wb = Workbook()
    ws = wb.active
    ws.title = "Grid View"

    # ---- Styles ------------------------------------------------------------
    hdr_fill = PatternFill("solid", fgColor="1F5254")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    zebra_a = PatternFill("solid", fgColor="FFFFFF")
    zebra_b = PatternFill("solid", fgColor="F5F7F8")
    total_font = Font(bold=True, color="0F3D3E", size=10)
    row_type_font = Font(bold=True, color="1F3D7A", size=10)
    border = Border(
        left=Side(style="thin", color="D6DCDC"),
        right=Side(style="thin", color="D6DCDC"),
        top=Side(style="thin", color="D6DCDC"),
        bottom=Side(style="thin", color="D6DCDC"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---- Title + subtitle --------------------------------------------------
    total_cols = 5 + days_n + 2  # header (incl. Bio Code) + day cols + 2 summary
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value=f"{company_name} — Grid View").font = Font(
        bold=True, size=14, color="0F3D3E"
    )
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 24

    from_date = grid.get("from_date") or ""
    to_date = grid.get("to_date") or ""
    period = (
        f"{from_date} → {to_date}"
        if (from_date and to_date and grid.get("range_mode"))
        else f"Month: {month}"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1, value=(
        f"{period} · Days: {days_n} · Generated: "
        f"{datetime.now(timezone(timedelta(hours=5, minutes=30))).strftime('%d-%b-%Y %H:%M IST')}"
    )).font = Font(italic=True, color="475569", size=10)
    ws.cell(row=2, column=1).alignment = center
    ws.row_dimensions[2].height = 18

    # ---- Header row --------------------------------------------------------
    header_row = 4
    hdrs = ["Bio Code", "Emp Name", "Emp Father Name", "Designation", "Type"]
    for i, label in enumerate(hdrs, start=1):
        c = ws.cell(row=header_row, column=i, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border
    for j, day_lbl in enumerate(day_labels):
        wk = weekday_labels[j] if j < len(weekday_labels) else ""
        c = ws.cell(
            row=header_row,
            column=6 + j,
            value=f"D{day_lbl}\n{wk}",
        )
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border
    c = ws.cell(row=header_row, column=6 + days_n, value="Duty HRS")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = center
    c.border = border
    c = ws.cell(row=header_row, column=6 + days_n + 1, value="Present Days")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = center
    c.border = border
    ws.row_dimensions[header_row].height = 32

    # ---- Column widths -----------------------------------------------------
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 8
    for j in range(days_n):
        ws.column_dimensions[get_column_letter(6 + j)].width = 8
    ws.column_dimensions[get_column_letter(6 + days_n)].width = 12
    ws.column_dimensions[get_column_letter(6 + days_n + 1)].width = 10
    ws.freeze_panes = f"F{header_row + 1}"

    # ---- Per-employee blocks ----------------------------------------------
    # Iter 77z-final — In/Out sheet now shows ONLY regular duty. OT
    # timestamps and hours belong to the dedicated OT report. So every
    # employee gets exactly 3 rows: D-In, D-Out, T-Hrs (duty-only totals).
    cur_row = header_row + 1
    for idx, emp in enumerate(employees):
        days_cell = emp.get("days") or {}
        totals = emp.get("totals") or {}

        row_types = ["D-In", "D-Out", "T-Hrs"]

        fill = zebra_a if idx % 2 == 0 else zebra_b

        for r_idx, row_type in enumerate(row_types):
            r = cur_row + r_idx

            # Left header columns only populated on the FIRST row of the block.
            if r_idx == 0:
                _bio = emp.get("bio_code")
                ws.cell(row=r, column=1, value="" if _bio in (None, "") else str(_bio))
                ws.cell(row=r, column=2, value=emp.get("name") or "")
                ws.cell(row=r, column=3, value=emp.get("father_name") or "")
                ws.cell(
                    row=r,
                    column=4,
                    value=emp.get("designation") or emp.get("department") or "",
                )
            else:
                for c_ in (1, 2, 3, 4):
                    ws.cell(row=r, column=c_, value="")

            # Type column
            tc = ws.cell(row=r, column=5, value=row_type)
            tc.font = row_type_font
            tc.alignment = center

            # Per-day cells
            for j, day_lbl in enumerate(day_labels):
                d = days_cell.get(day_lbl) or {}
                col = 6 + j
                val = ""
                if row_type == "D-In":
                    val = d.get("in") or ""
                elif row_type == "D-Out":
                    val = d.get("out") or ""
                elif row_type == "T-Hrs":
                    # Duty-only HRS per day (no OT — that lives in the OT sheet).
                    dh = float(d.get("duty_hours") or 0.0)
                    if dh > 0:
                        h = int(dh)
                        mm = int(round((dh - h) * 60))
                        if mm >= 60:
                            h += 1
                            mm -= 60
                        val = f"{h:02d}:{mm:02d}"
                    else:
                        val = "00:00"
                c = ws.cell(row=r, column=col, value=val)
                c.alignment = center
                c.border = border
                c.font = Font(size=9)

            # Trailing summary columns
            tot_wrk = ""
            day_total = ""
            # Duty-only total (excludes OT).
            combined = float(totals.get("hours") or 0.0)
            ot = float(totals.get("ot_hours") or 0.0)
            duty_only_total = round(max(0.0, combined - ot), 2)
            if row_type == "D-Out":
                tot_wrk = duty_only_total
                day_total = float(totals.get("present_days") or 0.0)
            elif row_type == "T-Hrs":
                tot_wrk = duty_only_total
                day_total = float(totals.get("present_days") or 0.0)
            else:  # D-In
                tot_wrk = 0.0
                day_total = 0.0

            sc = ws.cell(row=r, column=5 + days_n, value=tot_wrk)
            sc.alignment = center
            sc.border = border
            if row_type == "T-Hrs":
                sc.font = total_font
            if isinstance(tot_wrk, (int, float)):
                sc.number_format = "0.00"

            dc = ws.cell(row=r, column=5 + days_n + 1, value=day_total)
            dc.alignment = center
            dc.border = border
            if row_type == "T-Hrs":
                dc.font = total_font
            if isinstance(day_total, (int, float)):
                dc.number_format = "0.00"

            # Apply zebra fill + borders to all cells in the row.
            for c_ in range(1, 5 + days_n + 2):
                cell = ws.cell(row=r, column=c_)
                cell.fill = fill
                cell.border = border
                if cell.alignment is None:
                    cell.alignment = left if c_ <= 3 else center

            ws.row_dimensions[r].height = 18

        cur_row += len(row_types)

    # ---- Footer summary row (grand totals) ---------------------------------
    if employees:
        footer_row = cur_row
        ws.cell(row=footer_row, column=1, value=f"Employees: {len(employees)}").font = total_font
        for c_ in range(1, 6 + days_n + 2):
            ws.cell(row=footer_row, column=c_).fill = PatternFill("solid", fgColor="E6EDED")
            ws.cell(row=footer_row, column=c_).border = border

        grand_hrs = sum(
            max(
                0.0,
                float((e.get("totals") or {}).get("hours") or 0.0)
                - float((e.get("totals") or {}).get("ot_hours") or 0.0),
            )
            for e in employees
        )
        grand_days = sum(
            float((e.get("totals") or {}).get("present_days") or 0.0) for e in employees
        )
        c = ws.cell(row=footer_row, column=6 + days_n, value=round(grand_hrs, 2))
        c.font = total_font
        c.alignment = center
        c.number_format = "0.00"
        c = ws.cell(row=footer_row, column=6 + days_n + 1, value=round(grand_days, 2))
        c.font = total_font
        c.alignment = center
        c.number_format = "0.00"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ---------------------------------------------------------------------------
# Report 4 — Hours Only XLSX from Grid data (Iter 77z)
# ---------------------------------------------------------------------------
# Simpler variant of the Grid View sheet: ONE row per employee where each
# day cell shows the COMBINED per-day HRS = Regular Duty + OT (in HH:MM).
# Trailing columns split those totals into: Duty HRS, OT HRS, Total HRS,
# and Days so admins can reconcile at a glance.
# ---------------------------------------------------------------------------


def build_hours_only_grid_xlsx(grid: Dict[str, Any]) -> bytes:
    """One row per employee. Each day cell = duty+OT combined HH:MM.

    Consumes the same JSON payload as ``build_grid_view_xlsx`` so all the
    policy rules (bounce-merge, cross-day OT, weekly-off, dedup) are
    honored upstream.
    """
    company_name = ((grid or {}).get("company") or {}).get("name") or ""
    month = grid.get("month") or ""
    day_labels: List[str] = list(grid.get("day_labels") or [])
    weekday_labels: List[str] = list(grid.get("weekday_labels") or [])
    employees: List[Dict[str, Any]] = list(grid.get("employees") or [])
    days_n = len(day_labels)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hours Only"

    hdr_fill = PatternFill("solid", fgColor="1F5254")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    zebra_a = PatternFill("solid", fgColor="FFFFFF")
    zebra_b = PatternFill("solid", fgColor="F5F7F8")
    total_font = Font(bold=True, color="0F3D3E", size=10)
    border = Border(
        left=Side(style="thin", color="D6DCDC"),
        right=Side(style="thin", color="D6DCDC"),
        top=Side(style="thin", color="D6DCDC"),
        bottom=Side(style="thin", color="D6DCDC"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Layout (Iter 202 — user request): A=Bio Code, B=Name, C=Father,
    # D=Designation, E=Type (Duty HRS / OT HRS — one row EACH per employee,
    # day-wise, per the attendance policy), F..=days, then trailing totals.
    trail_labels = ["Duty HRS", "OT HRS", "Total Duty HRS", "Present Days", "Extra HRS"]
    total_cols = 5 + days_n + len(trail_labels)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value=f"{company_name} — Hours Only").font = Font(
        bold=True, size=14, color="0F3D3E"
    )
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 24

    from_date = grid.get("from_date") or ""
    to_date = grid.get("to_date") or ""
    period = (
        f"{from_date} → {to_date}"
        if (from_date and to_date and grid.get("range_mode"))
        else f"Month: {month}"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1, value=(
        f"{period} · Days: {days_n} · Duty HRS and OT HRS shown in SEPARATE rows "
        f"per employee, day-wise, as per the firm's attendance policy. "
        f"Generated: {datetime.now(timezone(timedelta(hours=5, minutes=30))).strftime('%d-%b-%Y %H:%M IST')}"
    )).font = Font(italic=True, color="475569", size=10)
    ws.cell(row=2, column=1).alignment = center
    ws.row_dimensions[2].height = 18

    header_row = 4
    hdrs = ["Bio Code", "Emp Name", "Emp Father Name", "Designation", "Type"]
    for i, label in enumerate(hdrs, start=1):
        c = ws.cell(row=header_row, column=i, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border
    for j, day_lbl in enumerate(day_labels):
        wk = weekday_labels[j] if j < len(weekday_labels) else ""
        c = ws.cell(row=header_row, column=6 + j, value=f"D{day_lbl}\n{wk}")
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border
    for k, label in enumerate(trail_labels):
        c = ws.cell(row=header_row, column=6 + days_n + k, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[header_row].height = 32

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    for j in range(days_n):
        ws.column_dimensions[get_column_letter(6 + j)].width = 8
    for k in range(len(trail_labels)):
        ws.column_dimensions[get_column_letter(6 + days_n + k)].width = 11
    ws.freeze_panes = f"F{header_row + 1}"

    def _fmt_hhmm(hrs: float) -> str:
        if hrs <= 0:
            return "00:00"
        h = int(hrs)
        mm = int(round((hrs - h) * 60))
        if mm >= 60:
            h += 1
            mm -= 60
        return f"{h:02d}:{mm:02d}"

    cur = header_row + 1
    grand_duty = grand_ot = grand_total = 0.0
    grand_days = 0.0
    for idx, emp in enumerate(employees):
        days_cell = emp.get("days") or {}
        totals = emp.get("totals") or {}
        fill = zebra_a if idx % 2 == 0 else zebra_b
        duty_row, ot_row = cur, cur + 1

        # Identity columns merged vertically across the Duty/OT row pair.
        _bio = emp.get("bio_code")
        id_vals = [
            "" if _bio in (None, "") else str(_bio),
            emp.get("name") or "",
            emp.get("father_name") or "",
            emp.get("designation") or emp.get("department") or "",
        ]
        for col_i, v in enumerate(id_vals, start=1):
            ws.merge_cells(start_row=duty_row, start_column=col_i,
                           end_row=ot_row, end_column=col_i)
            c = ws.cell(row=duty_row, column=col_i, value=v)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=duty_row, column=5, value="Duty HRS").font = Font(size=9, bold=True)
        ws.cell(row=ot_row, column=5, value="OT HRS").font = Font(
            size=9, bold=True, color="B45309")

        # Iter 202 (user request) — day-wise Duty HRS and OT HRS in
        # SEPARATE rows, both already policy-adjusted upstream (8-HR
        # sub-point, week-off / holiday rules, rounding, OT gates).
        for j, day_lbl in enumerate(day_labels):
            d = days_cell.get(day_lbl) or {}
            duty_h = float(d.get("duty_hours") or 0.0)
            ot_h = float(d.get("ot_hours") or 0.0)
            c = ws.cell(row=duty_row, column=6 + j, value=_fmt_hhmm(duty_h))
            c.alignment = center
            c.border = border
            c.font = Font(size=9)
            c2 = ws.cell(row=ot_row, column=6 + j, value=_fmt_hhmm(ot_h))
            c2.alignment = center
            c2.border = border
            c2.font = Font(size=9, color="B45309" if ot_h > 0 else "94A3B8")

        combined = float(totals.get("hours") or 0.0)
        ot = float(totals.get("ot_hours") or 0.0)
        duty_only = round(max(0.0, combined - ot), 2)
        present = float(totals.get("present_days") or 0.0)
        # Iter 202 — Present Days per the firm's attendance policy.
        days_int = totals.get("present_days_policy")
        if days_int is None:
            days_int = int(totals.get("total_days_int") or 0)
        extra_hrs = float(totals.get("total_extra_hrs") or 0.0)
        grand_duty += duty_only
        grand_ot += ot
        grand_total += combined
        grand_days += present
        for k, val in enumerate([
            round(duty_only, 2), round(ot, 2),
            round(combined, 2), days_int, round(extra_hrs, 2),
        ]):
            ws.merge_cells(start_row=duty_row, start_column=6 + days_n + k,
                           end_row=ot_row, end_column=6 + days_n + k)
            c = ws.cell(row=duty_row, column=6 + days_n + k, value=val)
            c.alignment = center
            c.border = border
            c.number_format = "0" if k == 3 else "0.00"
            c.font = total_font

        for r_ in (duty_row, ot_row):
            for c_ in range(1, total_cols + 1):
                cell = ws.cell(row=r_, column=c_)
                cell.fill = fill
                cell.border = border
            ws.row_dimensions[r_].height = 17
        cur += 2

    # Footer totals
    if employees:
        for c_ in range(1, total_cols + 1):
            ws.cell(row=cur, column=c_).fill = PatternFill("solid", fgColor="E6EDED")
            ws.cell(row=cur, column=c_).border = border
        ws.cell(row=cur, column=1, value=f"Employees: {len(employees)}").font = total_font
        grand_days_int = sum(
        float((e.get("totals") or {}).get("present_days_policy")
              if (e.get("totals") or {}).get("present_days_policy") is not None
              else (e.get("totals") or {}).get("total_days_int") or 0)
        for e in employees)
        grand_extra_hrs = sum(float((e.get("totals") or {}).get("total_extra_hrs") or 0.0) for e in employees)
        for k, val in enumerate([
            round(grand_duty, 2), round(grand_ot, 2),
            round(grand_total, 2), grand_days_int, round(grand_extra_hrs, 2),
        ]):
            c = ws.cell(row=cur, column=6 + days_n + k, value=val)
            c.font = total_font
            c.alignment = center
            c.number_format = "0" if k == 3 else "0.00"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
