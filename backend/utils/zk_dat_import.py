"""ZKTeco .dat biometric-punch parser (Iter 77).

The ZKTeco terminals export two flat text files with tab-separated columns:

    <bio_code>  <YYYY-MM-DD HH:MM:SS>  <status>  <verify_type>  <workcode>  <reserved>

* ``IN.dat``  contains punches for entry
* ``OUT.dat`` contains punches for exit

Both files can also be combined into a single upload where the ``kind`` is
inferred from the status column (0=IN, 1=OUT) - we handle both shapes.

Public entry-points:

* :func:`parse_zk_dat_lines` - tolerant line-level parser.
* :func:`import_zk_dat_bytes` - end-to-end: parses + inserts + dedupes into
  ``db.attendance``. Returns a stats dict for the UI.
"""
from __future__ import annotations

import re
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Line parsing
# ---------------------------------------------------------------------------

def parse_zk_dat_lines(
    text: str,
    default_kind: Optional[str] = None,
) -> List[Tuple[str, datetime, Optional[str]]]:
    """Parse one .dat file's content. Returns list of
    ``(bio_code, datetime, kind_hint)`` tuples. ``kind_hint`` is either
    ``"in"``, ``"out"`` or ``None`` (when unknown - the caller supplies
    ``default_kind`` for the whole file).

    Skips malformed lines silently; the caller can measure ``bad`` via the
    difference between input lines and output rows.
    """
    rows: List[Tuple[str, datetime, Optional[str]]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        parts = re.split(r"\s+", line, maxsplit=6)
        if len(parts) < 3:
            continue
        bio = parts[0].strip()
        ts_raw = f"{parts[1]} {parts[2]}"
        try:
            dt = datetime.strptime(ts_raw, "%Y-%m-%d %H:%M:%S").replace(
                tzinfo=timezone.utc,
            )
        except ValueError:
            continue
        # Optional inline kind from the 4th column when present.
        # ZKTeco status codes: 0=CheckIn, 1=CheckOut, 4=Overtime In, 5=Overtime Out
        kind: Optional[str] = default_kind
        if len(parts) >= 4 and parts[3].strip().isdigit():
            status = int(parts[3])
            if status in (0, 4):
                kind = "in"
            elif status in (1, 5):
                kind = "out"
        rows.append((bio, dt, kind))
    return rows


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# Iter 139 — Device "attendance record" exports (new formats).
#
# 1. Tab-separated .TXT export with header:
#        No  TMNo  EnNo  Name  Mode  INOUT  DateTime
#    e.g. "1\t1\t203\tHari om singh \t5\t0\t2024/09/03 13:45:56"
#
# 2. Binary .DAT backup ("ZoucqGENLOGData" header) — 8-byte records:
#        u16 pad | u32 seconds-since-2000-01-01 | u16 (enroll_no << 4 | flags)
#    Verified against the matching .TXT export (50,942/50,942 identical).
# ---------------------------------------------------------------------------

GENLOG_MARKER = b"GENLOGData"


def is_genlog_dat(data: Optional[bytes]) -> bool:
    """True when the bytes look like a binary GENLOG .DAT device backup."""
    return bool(data) and GENLOG_MARKER in data[:64]


def parse_genlog_records(data: bytes) -> List[Tuple[str, datetime]]:
    """Decode a binary GENLOG .DAT backup into ``(bio_code, datetime)``
    pairs. Auto-aligns the record start offset (header length varies by a
    byte or two across firmware versions)."""
    import struct
    from datetime import timedelta

    idx = data.find(GENLOG_MARKER)
    if idx < 0:
        return []
    base_start = idx + len(GENLOG_MARKER)
    epoch = datetime(2000, 1, 1, tzinfo=timezone.utc)
    # Punches can't come from the future — misaligned decodes typically
    # yield far-future years, so a tight upper bound both filters garbage
    # AND makes the alignment scoring discriminative.
    max_year = datetime.now(timezone.utc).year + 1

    def _decode(start: int) -> List[Tuple[str, datetime]]:
        body = data[start:]
        out: List[Tuple[str, datetime]] = []
        for i in range(len(body) // 8):
            rec = body[i * 8:(i + 1) * 8]
            _pad, ts, enf = struct.unpack("<HIH", rec)
            en = enf >> 4
            if en <= 0:
                continue
            dt = epoch + timedelta(seconds=ts)
            if dt.year < 2001 or dt.year > max_year:
                continue
            out.append((str(en), dt))
        return out

    # Score candidate offsets by how many of the first 200 records decode
    # to sane values; pick the best-aligned one.
    best: List[Tuple[str, datetime]] = []
    best_score = -1
    for extra in range(0, 8):
        start = base_start + extra
        sample_n = min(200, max(1, (len(data) - start) // 8))
        body = data[start:start + sample_n * 8]
        score = len(_decode_sample(body, epoch, max_year))
        if score > best_score:
            best_score = score
            best = _decode(start)
    return best


def _decode_sample(body: bytes, epoch: datetime, max_year: int) -> List[int]:
    import struct
    from datetime import timedelta
    ok = []
    for i in range(len(body) // 8):
        rec = body[i * 8:(i + 1) * 8]
        _pad, ts, enf = struct.unpack("<HIH", rec)
        if enf >> 4 <= 0:
            continue
        dt = epoch + timedelta(seconds=ts)
        if 2001 <= dt.year <= max_year:
            ok.append(i)
    return ok


def genlog_to_txt_text(data: bytes) -> str:
    """Convert a binary GENLOG .DAT backup into the tab-separated device
    .TXT shape so it can be persisted + re-read through the normal text
    pipeline (used by the 'Refresh Bio' recovery flow)."""
    lines = ["No\tTMNo\tEnNo\tName\tMode\tINOUT\tDateTime"]
    for i, (bio, dt) in enumerate(parse_genlog_records(data), start=1):
        lines.append(f"{i}\t1\t{bio}\t\t0\t0\t{dt:%Y/%m/%d %H:%M:%S}")
    return "\n".join(lines)


_DEVICE_TXT_DT_FORMATS = (
    "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M",
    "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S",
)


def parse_device_txt_lines(
    text: str,
    default_kind: Optional[str] = None,
) -> List[Tuple[str, datetime, Optional[str]]]:
    """Parse the tab-separated device .TXT export
    (``No | TMNo | EnNo | Name | Mode | INOUT | DateTime``). The header row
    is optional — without it the standard column positions are assumed.
    INOUT: even codes (0/2/4) = IN, odd codes (1/3/5) = OUT. When every
    punch carries the same INOUT value the position-based re-classify in
    :func:`import_zk_dat_bytes` fixes the pairing."""
    rows: List[Tuple[str, datetime, Optional[str]]] = []
    col: Optional[Dict[str, int]] = None
    for raw in text.splitlines():
        if not raw.strip():
            continue
        parts = raw.split("\t")
        if col is None:
            low = [p.strip().lower() for p in parts]
            if "enno" in low:
                col = {name: i for i, name in enumerate(low)}
                continue
            col = {"enno": 2, "inout": 5, "datetime": 6}
        en_i = col.get("enno", 2)
        dt_i = col.get("datetime", 6)
        io_i = col.get("inout", 5)
        if len(parts) <= max(en_i, dt_i):
            continue
        bio = parts[en_i].strip()
        if not bio or not any(ch.isdigit() for ch in bio):
            continue
        ts_raw = parts[dt_i].strip()
        dt: Optional[datetime] = None
        for fmt in _DEVICE_TXT_DT_FORMATS:
            try:
                dt = datetime.strptime(ts_raw, fmt).replace(tzinfo=timezone.utc)
                break
            except ValueError:
                continue
        if dt is None:
            continue
        kind = default_kind or "in"
        if io_i < len(parts) and parts[io_i].strip().isdigit():
            kind = "out" if int(parts[io_i]) % 2 == 1 else (default_kind or "in")
        rows.append((bio, dt, kind))
    return rows


def _looks_like_device_txt(text: str) -> bool:
    """True when the first non-empty line is the device .TXT header."""
    for line in text.splitlines():
        if not line.strip():
            continue
        cols = [c.strip().lower() for c in line.split("\t")]
        return "enno" in cols
    return False


def decode_punch_bytes(
    data: Optional[bytes],
    default_kind: Optional[str] = None,
) -> List[Tuple[str, datetime, Optional[str]]]:
    """Format-dispatching decoder: binary GENLOG .DAT, device .TXT export
    (with or without header), or the classic ZK .dat text shape."""
    if not data:
        return []
    if is_genlog_dat(data):
        return [(bio, dt, default_kind or "in")
                for bio, dt in parse_genlog_records(data)]
    text = data.decode("utf-8", errors="replace")
    if _looks_like_device_txt(text):
        return parse_device_txt_lines(text, default_kind=default_kind)
    rows = parse_zk_dat_lines(text, default_kind=default_kind)
    if not rows:
        # Headerless device .TXT fallback (tab positions are fixed).
        rows = parse_device_txt_lines(text, default_kind=default_kind)
    return rows


# ---------------------------------------------------------------------------
# Iter 106 — Excel (.xlsx / .xls) punch imports. The employer keeps IN
# punches in one sheet/file and OUT punches in another. Expected columns
# (header row optional): CODE | DATE | TIME  — or CODE | DATETIME.
# We normalise every row into the same .dat text shape so the proven
# import pipeline (mapping, dedupe, range filter) is reused 1:1.
# ---------------------------------------------------------------------------

def _cell_to_dt(date_v: Any, time_v: Any = None) -> Optional[datetime]:
    """Combine a date cell + optional time cell into a datetime."""
    from datetime import date as _date, time as _time, timedelta as _td
    d: Optional[datetime] = None
    if isinstance(date_v, datetime):
        d = date_v
    elif isinstance(date_v, _date):
        d = datetime(date_v.year, date_v.month, date_v.day)
    elif isinstance(date_v, (int, float)) and date_v > 20000:
        # Excel serial date (xlrd already converts when told; safeguard)
        d = datetime(1899, 12, 30) + _td(days=float(date_v))
    elif isinstance(date_v, str):
        s = date_v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M:%S",
                    "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d/%m/%Y %H:%M",
                    "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
            try:
                d = datetime.strptime(s, fmt)
                break
            except ValueError:
                continue
    if d is None:
        return None
    # Merge the separate time cell when the date cell had no time part.
    if time_v is not None and d.hour == 0 and d.minute == 0 and d.second == 0:
        if isinstance(time_v, datetime):
            d = d.replace(hour=time_v.hour, minute=time_v.minute, second=time_v.second)
        elif isinstance(time_v, _time):
            d = d.replace(hour=time_v.hour, minute=time_v.minute, second=time_v.second)
        elif isinstance(time_v, (int, float)) and 0 <= float(time_v) < 1:
            secs = round(float(time_v) * 86400)
            d = d.replace(hour=secs // 3600, minute=(secs % 3600) // 60, second=secs % 60)
        elif isinstance(time_v, str) and time_v.strip():
            t = time_v.strip()
            for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
                try:
                    tt = datetime.strptime(t, fmt)
                    d = d.replace(hour=tt.hour, minute=tt.minute, second=tt.second)
                    break
                except ValueError:
                    continue
    return d.replace(tzinfo=timezone.utc)


def excel_punches_to_dat_text(data: bytes, filename: str = "") -> str:
    """Convert an Excel sheet of punches into ZK .dat text lines
    (``code<TAB>YYYY-MM-DD HH:MM:SS``). Raises ValueError on unreadable
    files."""
    rows: List[Tuple[Any, Any, Any]] = []
    name = (filename or "").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        import xlrd
        book = xlrd.open_workbook(file_contents=data)
        sh = book.sheet_by_index(0)
        for r in range(sh.nrows):
            vals = sh.row_values(r)
            c0 = vals[0] if len(vals) > 0 else None
            c1 = vals[1] if len(vals) > 1 else None
            c2 = vals[2] if len(vals) > 2 else None
            # xlrd returns dates as floats — convert via xldate when possible
            if isinstance(c1, float) and c1 > 1:
                try:
                    c1 = datetime(*xlrd.xldate_as_tuple(c1, book.datemode))
                except Exception:
                    pass
            if isinstance(c2, float) and 0 <= c2 < 1:
                pass  # handled as fraction-of-day in _cell_to_dt
            rows.append((c0, c1, c2))
    else:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        for vals in ws.iter_rows(values_only=True):
            c0 = vals[0] if len(vals) > 0 else None
            c1 = vals[1] if len(vals) > 1 else None
            c2 = vals[2] if len(vals) > 2 else None
            rows.append((c0, c1, c2))
        wb.close()

    lines: List[str] = []
    for c0, c1, c2 in rows:
        if c0 is None or c1 is None:
            continue
        code = str(c0).strip()
        if code.endswith(".0"):
            code = code[:-2]
        if not code or not any(ch.isdigit() for ch in code):
            continue  # header / junk row
        dt = _cell_to_dt(c1, c2)
        if dt is None:
            continue
        lines.append(f"{code}\t{dt:%Y-%m-%d} {dt:%H:%M:%S}")
    if not lines:
        raise ValueError(
            "No punch rows found — expected columns: CODE | DATE | TIME "
            "(or CODE | DATETIME).")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Import driver
# ---------------------------------------------------------------------------

async def _build_bio_index(db, company_id: str) -> Dict[str, dict]:
    index: Dict[str, dict] = {}
    async for u in db.users.find(
        {
            "company_id": company_id,
            "role": "employee",
            "bio_code": {"$exists": True, "$ne": None},
        },
        {
            "_id": 0, "user_id": 1, "name": 1,
            "employee_code": 1, "bio_code": 1,
            # Iter 223 — shift override so imports can classify punches
            # according to the employee's shift when both files are given.
            "attendance_policy_override": 1,
        },
    ):
        key = str(u["bio_code"]).lstrip("0") or "0"
        # First-writer wins so re-imports stay deterministic.
        index.setdefault(key, u)
    return index


async def import_zk_dat_bytes(
    db,
    *,
    company_id: str,
    in_bytes: Optional[bytes] = None,
    out_bytes: Optional[bytes] = None,
    combined_bytes: Optional[bytes] = None,
    from_date: Optional[str] = None,   # YYYY-MM-DD (inclusive) - optional filter
    to_date: Optional[str] = None,     # YYYY-MM-DD (inclusive) - optional filter
    source_tag: Optional[str] = None,
    on_existing: str = "skip",  # Iter 224 — "skip" (default) | "replace"
) -> Dict[str, Any]:
    """Parse + insert ZKTeco punches into ``db.attendance``. Idempotent:
    re-running with the same file is a no-op (dedupes by user + at + kind
    + source)."""
    tag = source_tag or f"import:zk_upload_{datetime.now(timezone.utc):%Y%m%dT%H%M%S}"
    bio_index = await _build_bio_index(db, company_id)
    # Iter 175 — contractual employees: their imported punches must be
    # approved by the company first (Contractor Punch approvals).
    contractual_map: Dict[str, Optional[str]] = {}
    async for _c in db.users.find(
        {"company_id": company_id, "is_contractual": True},
        {"_id": 0, "user_id": 1, "contractor_name": 1},
    ):
        contractual_map[_c["user_id"]] = _c.get("contractor_name")

    # Iter 139 — every slot now accepts classic .dat text, the device .TXT
    # export AND the binary GENLOG .DAT backup (format auto-detected).
    # Iter 223 (user rules) —
    #   • IN.dat slot  → EVERY punch is an IN punch (status byte ignored).
    #   • OUT.dat slot → EVERY punch is an OUT punch.
    #   • Near-duplicates (same employee, same day, same kind within
    #     15 minutes) are IGNORED — only the first punch is kept.
    #   • An evening IN punch on a day that already has a morning IN
    #     lands as the 3rd punch → the pipeline reads it as OT IN.
    #   • When BOTH files are imported, punches are sanity-checked
    #     against the EMPLOYEE'S SHIFT (misordered kinds re-classified).
    rows: List[Tuple[str, datetime, Optional[str], str]] = []
    rows.extend([(b, d, "in", "in_file") for (b, d, _k) in decode_punch_bytes(in_bytes, default_kind="in")])
    rows.extend([(b, d, "out", "out_file") for (b, d, _k) in decode_punch_bytes(out_bytes, default_kind="out")])
    rows.extend([(b, d, k, "combined") for (b, d, k) in decode_punch_bytes(combined_bytes)])

    # Shift Master catalogue (for the both-files shift classification).
    shift_masters: Dict[str, dict] = {}
    async for _s in db.shift_masters.find({}, {"_id": 0, "shift_id": 1, "start": 1, "end": 1}):
        if _s.get("shift_id"):
            shift_masters[_s["shift_id"]] = _s

    def _hhmm_min(v: Optional[str]) -> Optional[int]:
        try:
            hh, mm = str(v).strip().split(":")[:2]
            return int(hh) * 60 + int(mm)
        except Exception:
            return None

    def _shift_mid_min(user: dict) -> int:
        """Midpoint of the employee's assigned shift (minutes from 00:00).
        Fallback 13:00 when no shift is assigned."""
        ov = (user or {}).get("attendance_policy_override") or {}
        sh = shift_masters.get(ov.get("shift_id") or "")
        st = _hhmm_min((sh or {}).get("start"))
        en = _hhmm_min((sh or {}).get("end"))
        if st is None or en is None:
            return 13 * 60
        dur = (en - st) if en >= st else (en + 24 * 60 - st)
        return (st + dur // 2) % (24 * 60)

    stats = {
        "total_lines": len(rows),
        "inserted": 0,
        "duplicate": 0,
        "near_duplicate": 0,
        "unmapped": 0,
        "out_of_range": 0,
        "missing_kind": 0,
        # Iter 224 (user rule) — existing-data protection counters.
        "manual_locked_days": 0,      # days kept untouched (manual master punches)
        "existing_machine_days": 0,   # days with DIFFERENT machine data → need permission
        "replaced_days": 0,           # days replaced after permission
    }
    unmapped_seen: set = set()
    # Iter 86 - Buffer punches by (user_id, date) so we can re-classify
    # kind by punch-position when the source file gave every row the
    # same status byte (very common in ZKTeco combined exports).
    pending_by_user_day: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}

    for bio, dt, kind, slot in rows:
        if kind is None:
            stats["missing_kind"] += 1
            continue
        date_str = dt.strftime("%Y-%m-%d")
        if from_date and date_str < from_date:
            stats["out_of_range"] += 1
            continue
        if to_date and date_str > to_date:
            stats["out_of_range"] += 1
            continue
        key = bio.lstrip("0") or "0"
        user = bio_index.get(key)
        if not user:
            stats["unmapped"] += 1
            unmapped_seen.add(bio)
            continue
        # Iter 86 - Re-classify kind by punch-position within the day.
        #
        # Real ZKTeco terminals often export EVERY punch with the same
        # `status` byte (e.g. always 1=CheckOut) because operators don't
        # configure the terminal function keys. Trusting that column
        # would classify every punch as "out" and the attendance grid
        # would then flag every day as `missing_punch` and show 0 hours.
        #
        # Fix: buffer by (user, date), sort by timestamp, then alternate
        # kinds - 1st=in, 2nd=out, 3rd=in (OT start), 4th=out (OT end).
        # This matches how the rest of the pipeline (`_pair_punches`,
        # `split_regular_ot_times`, `compute_textile_day`) already
        # expects the data to look.  Users who rely on genuine IN.dat +
        # OUT.dat uploads still get honored kinds because those come in
        # via `in_bytes` / `out_bytes` (default_kind="in"/"out") - the
        # position-reclassify only alters the FINAL kind when the file's
        # kinds all agree for a given (user, date) grouping.
        pending_by_user_day.setdefault(
            (user["user_id"], date_str), []
        ).append({
            "bio": bio,
            "user": user,
            "dt": dt,
            "raw_kind": kind,
            "slot": slot,
        })
    # Now flush pending groups with position-inferred kind when needed.
    for (uid, date_str), items in pending_by_user_day.items():
        items.sort(key=lambda x: x["dt"])

        # Iter 223 (user rule) — NEAR-DUPLICATE FILTER: within one
        # employee-day, a punch of the same kind landing within 15 min
        # of the previously kept punch is a double-read → ignored.
        kept: List[Dict[str, Any]] = []
        last_kept_by_kind: Dict[str, datetime] = {}
        for item in items:
            k = item["raw_kind"] or ""
            prev = last_kept_by_kind.get(k)
            if prev is not None and (item["dt"] - prev).total_seconds() <= 15 * 60:
                stats["near_duplicate"] += 1
                continue
            last_kept_by_kind[k] = item["dt"]
            kept.append(item)
        items = kept
        if not items:
            continue

        # Iter 224 (user rule) — EXISTING-DATA PROTECTION:
        #  • A day that already has MANUAL punches from the master is
        #    NEVER changed/replaced — the whole day is skipped.
        #  • A day that already has MACHINE punches (device / previous
        #    import) is NOT replaced without permission: by default the
        #    day is skipped and reported so the portal can PROMPT the
        #    admin; a re-run with on_existing="replace" (after the
        #    prompt) deletes ONLY the old machine punches and imports
        #    the new ones. Days whose new punches are all exact
        #    duplicates flow through unchanged (idempotent re-upload).
        _existing = await db.attendance.find(
            {"user_id": uid, "date": date_str},
            {"_id": 0, "at": 1, "kind": 1, "source": 1},
        ).to_list(300)
        if any(str(d.get("source") or "").startswith("manual") for d in _existing):
            stats["manual_locked_days"] += 1
            continue
        _machine = [
            d for d in _existing
            if re.match(r"^(import|zkteco|bio|excel)", str(d.get("source") or ""))
        ]
        if _machine:
            _have = {(d.get("at"), d.get("kind")) for d in _machine}
            _all_dup = all(
                (i["dt"].isoformat(), i["raw_kind"]) in _have for i in items
            )
            if not _all_dup:
                if on_existing == "replace":
                    await db.attendance.delete_many({
                        "user_id": uid,
                        "date": date_str,
                        "source": {"$regex": "^(import|zkteco|bio|excel)"},
                    })
                    stats["replaced_days"] += 1
                else:
                    stats["existing_machine_days"] += 1
                    continue

        slots_here = {i["slot"] for i in items}
        raw_kinds = {i["raw_kind"] for i in items}
        # Legacy combined-file behavior: when EVERY row carries the same
        # status byte and there are 2+ punches, alternate in/out/in/out.
        # Slot files (IN.dat / OUT.dat alone) keep their forced kinds.
        force_alternate = (
            slots_here == {"combined"} and len(items) >= 2 and len(raw_kinds) <= 1
        )

        # Iter 223 (user rule) — BOTH FILES imported: classify according
        # to the EMPLOYEE'S SHIFT. When the chronological kind sequence
        # doesn't pair cleanly (two INs or two OUTs in a row), rebuild it
        # by alternation anchored on the shift: first punch before the
        # shift midpoint starts as IN (normal day → in/out/OT-in/OT-out,
        # so an evening IN after a morning IN becomes the OT IN); a day
        # whose first punch lands after the shift midpoint starts as OUT
        # (missed morning punch / night shift spillover).
        shift_anchor: Optional[str] = None
        if "in_file" in slots_here and "out_file" in slots_here and len(items) >= 2:
            seq = [i["raw_kind"] for i in items]
            clean = all(seq[p] != seq[p + 1] for p in range(len(seq) - 1)) and seq[0] == "in"
            if not clean:
                mid = _shift_mid_min(items[0]["user"])
                first_min = items[0]["dt"].hour * 60 + items[0]["dt"].minute
                shift_anchor = "in" if first_min < mid else "out"

        for pos, item in enumerate(items):
            if force_alternate:
                kind = "in" if pos % 2 == 0 else "out"
            elif shift_anchor is not None:
                even = shift_anchor
                odd = "out" if shift_anchor == "in" else "in"
                kind = even if pos % 2 == 0 else odd
            else:
                kind = item["raw_kind"]
            dt = item["dt"]
            user = item["user"]
            # Idempotency: skip if ANY punch already exists for this
            # user at the exact same timestamp+kind — regardless of the
            # import batch tag. (Previously the query included the
            # per-upload ``source`` tag which is timestamped, so
            # re-uploading the same .dat file silently duplicated every
            # punch and broke IN/OUT pairing in the attendance grid.)
            existing = await db.attendance.find_one(
                {
                    "user_id": user["user_id"],
                    "at": dt.isoformat(),
                    "kind": kind,
                },
                {"_id": 0, "record_id": 1},
            )
            if existing:
                stats["duplicate"] += 1
                continue
            record = {
                "record_id": f"imp_{uuid.uuid4().hex[:12]}",
                "user_id": user["user_id"],
                "company_id": company_id,
                "branch_id": None,
                "branch_name": "Biometric Terminal (imported)",
                "date": date_str,
                "kind": kind,
                "at": dt.isoformat(),
                "original_at": dt.isoformat(),
                "latitude": None,
                "longitude": None,
                "distance_m": None,
                "source": tag,
                "outside_geofence": False,
                "status": "approved",
                "decision_by": "system:import",
                "decision_at": now_iso(),
                "decision_reason": f"Uploaded ZKTeco .dat ({tag})",
                "device_serial": tag,
                "device_id": None,
                "device_verify_type": None,
                "selfie_base64": None,
                "created_at": now_iso(),
            }
            # Iter 175 — contractual employee: force pending + stamp contractor.
            if user["user_id"] in contractual_map:
                record.update({
                    "is_contractual": True,
                    "contractor_name": contractual_map[user["user_id"]],
                    "status": "pending",
                    "decision_by": None,
                    "decision_at": None,
                    "decision_reason": None,
                    "pending_reason": "contractual_employee",
                })
            await db.attendance.insert_one(record)
            stats["inserted"] += 1

    stats["unmapped_bio_codes"] = sorted(
        unmapped_seen, key=lambda x: int(x) if x.isdigit() else 0,
    )[:50]
    stats["source_tag"] = tag
    return stats
