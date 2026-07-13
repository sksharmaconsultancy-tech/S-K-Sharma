"""Daily-basis attendance report builders (Iter 111).

One row per employee for a SINGLE selected date, exported as XLSX or a
portrait A4 PDF. Consumes the exact grid payload from
``_compute_monthly_grid_data`` (called with from_date == to_date) so all
policy rules (bounce-merge, dedup, OT cap, night shift, extra duty) are
honoured — the numbers match the on-screen Grid View 1:1.

Columns: S.No · Bio Code · Emp Code · Name · Father Name · Designation ·
In · Out · OT In · OT Out · Duty HRS · OT HRS · Total HRS · Status
"""
from __future__ import annotations

import io
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER

_HEADERS = [
    "S.No", "Bio Code", "Emp Code", "Name", "Father Name", "Designation",
    "In", "Out", "OT In", "OT Out", "Duty HRS", "OT HRS", "Total HRS", "Status",
]


def _hhmm(hrs: float) -> str:
    h = int(hrs)
    mm_ = int(round((hrs - h) * 60))
    if mm_ >= 60:
        h += 1
        mm_ -= 60
    return f"{h:02d}:{mm_:02d}"


def _daily_rows(grid: Dict[str, Any]) -> Tuple[List[List[Any]], Dict[str, Any]]:
    """Flatten the single-day grid into report rows + summary counters."""
    day_labels: List[str] = list(grid.get("day_labels") or [])
    label = day_labels[0] if day_labels else ""
    rows: List[List[Any]] = []
    present = absent = anomalies = 0
    tot_duty = tot_ot = 0.0
    for i, emp in enumerate(grid.get("employees") or [], start=1):
        d = (emp.get("days") or {}).get(label) or {}
        duty = float(d.get("duty_hours") or 0.0)
        ot = float(d.get("ot_hours") or 0.0)
        total = float(d.get("hours") or 0.0) or round(duty + ot, 2)
        if d.get("anomaly"):
            status = "MISS PUNCH"
            anomalies += 1
        elif duty > 0 or (d.get("in") and d.get("out")):
            status = "P"
            present += 1
        else:
            status = "A"
            absent += 1
        tot_duty += duty
        tot_ot += ot
        bio = emp.get("bio_code")
        rows.append([
            i,
            "" if bio in (None, "") else str(bio),
            emp.get("employee_code") or "",
            emp.get("name") or "",
            emp.get("father_name") or "",
            emp.get("designation") or emp.get("department") or "",
            d.get("in") or "—",
            d.get("out") or "—",
            d.get("ot_in") or "",
            d.get("ot_out") or "",
            _hhmm(duty) if duty > 0 else "00:00",
            _hhmm(ot) if ot > 0 else "",
            _hhmm(total) if total > 0 else "00:00",
            status,
        ])
    summary = {
        "present": present, "absent": absent, "anomalies": anomalies,
        "tot_duty": round(tot_duty, 2), "tot_ot": round(tot_ot, 2),
    }
    return rows, summary


def build_daily_xlsx(grid: Dict[str, Any], date_s: str) -> bytes:
    company_name = ((grid or {}).get("company") or {}).get("name") or ""
    rows, summary = _daily_rows(grid)

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"

    hdr_fill = PatternFill("solid", fgColor="1F5254")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    zebra_b = PatternFill("solid", fgColor="F5F7F8")
    total_font = Font(bold=True, color="0F3D3E", size=10)
    border = Border(
        left=Side(style="thin", color="D6DCDC"),
        right=Side(style="thin", color="D6DCDC"),
        top=Side(style="thin", color="D6DCDC"),
        bottom=Side(style="thin", color="D6DCDC"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    n_cols = len(_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value=f"{company_name} — Daily Attendance Report").font = Font(
        bold=True, size=14, color="0F3D3E",
    )
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 24

    try:
        pretty = datetime.strptime(date_s, "%Y-%m-%d").strftime("%d-%b-%Y (%A)")
    except ValueError:
        pretty = date_s
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(row=2, column=1, value=(
        f"Date: {pretty} · Present: {summary['present']} · Absent: {summary['absent']}"
        f" · Miss Punch: {summary['anomalies']} · Generated: "
        f"{datetime.now(timezone(timedelta(hours=5, minutes=30))).strftime('%d-%b-%Y %H:%M IST')}"
    )).font = Font(italic=True, color="475569", size=10)
    ws.cell(row=2, column=1).alignment = center

    header_row = 4
    for i, label in enumerate(_HEADERS, start=1):
        c = ws.cell(row=header_row, column=i, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[header_row].height = 22

    widths = [6, 10, 10, 26, 24, 18, 8, 8, 8, 8, 10, 9, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    r = header_row + 1
    for idx, row in enumerate(rows):
        for c_i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=c_i, value=val)
            c.border = border
            c.font = Font(size=9)
            c.alignment = left if c_i in (4, 5, 6) else center
            if idx % 2 == 1:
                c.fill = zebra_b
            if c_i == 14:
                if val == "P":
                    c.font = Font(size=9, bold=True, color="15803D")
                elif val == "A":
                    c.font = Font(size=9, bold=True, color="B91C1C")
                elif val == "MISS PUNCH":
                    c.font = Font(size=9, bold=True, color="B45309")
        r += 1

    # Footer totals
    if rows:
        ws.cell(row=r, column=1, value=f"Employees: {len(rows)}").font = total_font
        ws.cell(row=r, column=11, value=_hhmm(summary["tot_duty"])).font = total_font
        ws.cell(row=r, column=12, value=_hhmm(summary["tot_ot"])).font = total_font
        ws.cell(
            row=r, column=13,
            value=_hhmm(summary["tot_duty"] + summary["tot_ot"]),
        ).font = total_font
        ws.cell(
            row=r, column=14,
            value=f"P:{summary['present']} A:{summary['absent']}",
        ).font = total_font
        for c_i in range(1, n_cols + 1):
            ws.cell(row=r, column=c_i).fill = PatternFill("solid", fgColor="E6EDED")
            ws.cell(row=r, column=c_i).border = border
            ws.cell(row=r, column=c_i).alignment = center

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_daily_pdf(grid: Dict[str, Any], date_s: str) -> bytes:
    company_name = ((grid or {}).get("company") or {}).get("name") or ""
    rows, summary = _daily_rows(grid)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=8 * mm, rightMargin=8 * mm,
        topMargin=8 * mm, bottomMargin=8 * mm,
    )
    title = ParagraphStyle(
        "t", fontName="Helvetica-Bold", fontSize=13, alignment=TA_CENTER,
        textColor=rl_colors.HexColor("#0F3D3E"), spaceAfter=2,
    )
    sub = ParagraphStyle(
        "s", fontName="Helvetica-Oblique", fontSize=9, alignment=TA_CENTER,
        textColor=rl_colors.HexColor("#475569"), spaceAfter=6,
    )
    try:
        pretty = datetime.strptime(date_s, "%Y-%m-%d").strftime("%d-%b-%Y (%A)")
    except ValueError:
        pretty = date_s
    now_ist = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    story: List[Any] = [
        Paragraph(f"{company_name} — Daily Attendance Report", title),
        Paragraph(
            f"Date: {pretty} · Present: {summary['present']} · Absent: {summary['absent']}"
            f" · Miss Punch: {summary['anomalies']} · Generated: "
            f"{now_ist.strftime('%d-%b-%Y %H:%M IST')}",
            sub,
        ),
    ]

    data = [_HEADERS] + [[str(v) for v in row] for row in rows]
    # Footer
    data.append([
        f"Employees: {len(rows)}", "", "", "", "", "", "", "", "", "",
        _hhmm(summary["tot_duty"]), _hhmm(summary["tot_ot"]),
        _hhmm(summary["tot_duty"] + summary["tot_ot"]),
        f"P:{summary['present']} A:{summary['absent']}",
    ])

    col_w = [
        10 * mm, 13 * mm, 13 * mm, 42 * mm, 38 * mm, 26 * mm,
        13 * mm, 13 * mm, 13 * mm, 13 * mm, 15 * mm, 13 * mm, 15 * mm, 20 * mm,
    ]
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1F5254")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (2, -1), "CENTER"),
        ("ALIGN", (6, 1), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#D6DCDC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1), rl_colors.HexColor("#E6EDED")),
        ("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1), "Helvetica-Bold"),
    ]
    for r_i in range(1, len(data) - 1):
        if r_i % 2 == 0:
            style.append(("BACKGROUND", (0, r_i), (-1, r_i), rl_colors.HexColor("#F5F7F8")))
        status = data[r_i][-1]
        color = {"P": "#15803D", "A": "#B91C1C", "MISS PUNCH": "#B45309"}.get(status)
        if color:
            style.append(("TEXTCOLOR", (-1, r_i), (-1, r_i), rl_colors.HexColor(color)))
            style.append(("FONTNAME", (-1, r_i), (-1, r_i), "Helvetica-Bold"))

    table = Table(data, colWidths=col_w, repeatRows=1)
    table.setStyle(TableStyle(style))
    story.append(table)
    doc.build(story)
    return buf.getvalue()
