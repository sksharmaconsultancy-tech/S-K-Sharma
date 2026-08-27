"""Iter 746 — Payroll.xlsx module-status updater (user request).

Fills the 'Actions Remarks' column (D) on both sheets (HRIS + HR Modules)
with the CURRENT portal status: ✅ Available / 🟡 Partial / ❌ Not available,
colour-coded, matched by the requirement text in column C (or B for HRIS
top-section module names).
"""
import re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SRC = "/tmp/Payroll.xlsx"
OUT = "/app/Payroll_Module_Status.xlsx"

GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")


def norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


A = "✅ Available — "
P = "🟡 Partial — "
N = "❌ Not available — "

M = {
    # ── HRIS top section (function-level, column C) ──
    "current opening post": N + "recruitment module abhi nahi bana",
    "job description": N + "recruitment module abhi nahi bana",
    "hod approval, budget": N + "recruitment module abhi nahi bana",
    "interview schedule with panel team": N + "recruitment module abhi nahi bana",
    "recruitment metrics and analytics report": N + "recruitment module abhi nahi bana",
    "offer letters": A + "HR Letters module",
    "appointment letter": A + "HR Letters module",
    "promotion letter": A + "HR Letters module",
    "probation to confirmation letter": A + "HR Letters + Probation module",
    "probation to extension letter": A + "HR Letters + Probation module",
    "employees life cycle & trend": N + "performance management abhi nahi",
    "apprisal": N + "performance management abhi nahi",
    "kra": N + "performance management abhi nahi",
    "kpa": N + "performance management abhi nahi",
    "goals": N + "performance management abhi nahi",
    "feedback": N + "performance management abhi nahi",
    "job rotation / transfer": A + "Branch Transfer + history (Branch Master)",
    "user satisfaction survey": N + "survey module abhi nahi",
    "tracking performance metrics": N + "performance management abhi nahi",
    "payroll expenditure to revenue ratio": P + "Labour Cost Dashboard hai; revenue input pending",
    "absenteeism rate": A + "Present/Absent reports + analytics",
    "employees attrition rate": P + "Joined/Left data hai; attrition % card pending",
    "total overtime cost vs monthly salary overheads": A + "Labour Cost Dashboard + OT reports",
    "salary virance from previous month": P + "run compare hai; variance report pending",
    "recruitment cost & budget": N + "recruitment module abhi nahi",
    "total manpower-cetegory wise": A + "Analytics + Master Data report",
    "new joined": A + "Analytics/joining reports",
    "left": A + "Offboarding + reports",
    "total hrs": A + "Attendance Grid / hrs reports",
    "ot hrs": A + "OT reports + IN-OUT OT Matrix",
    "depttwise cost & budget": P + "dept-wise cost hai; budget input pending",
    "depttwise present & absent summary": A + "Present/Absent report (dept filter)",
    "freqent late commers": A + "Late Penalty report (late count/employee)",
    "freqent absenters": P + "absent report hai; frequency ranking pending",
    "compare chart current month & previous month or years": P + "kuch compare views hai; chart pending",
    "man days work report for unit wise as per factory license monthly/yearly": A + "Labour Statistics + Factory Compliance",
    "average number of workers employeed yearly basis": A + "Factory Annual Return",
    "number of man hours worked": A + "Labour Statistics",
    "number of man hours ot worked": A + "OT reports / Labour Statistics",
    "total wages paid during the year": A + "Wage registers + annual returns",
    "total find/deduction during the year": A + "Fine/Damage registers (statutory)",
    "statuary bonus report": A + "Bonus Registers (Form C/D) + yearly summary",
    "customized salary/wages register": A + "Report Formats + Salary Register",
    "customized attendance register": A + "Report Formats + Attendance registers",
    "esi accident recrods report": A + "Accident Register (ESIC + Factory)",
    "traking hrs worked": A + "Attendance Grid hours",
    # ── Masters & Policy ──
    "employees master data": A + "Employee Master + Bulk Import + KYC + Documents",
    "employee probation to confirmation process report": A + "Probation module (letters ke saath)",
    "default shift assignment": A + "Shift Master + default shift",
    "shift policy configuration for different shift": A + "Attendance Policy multi-shift config",
    "leave policy, master & opening": A + "Leave module + Balance Config (opening)",
    "holiday master": A + "Masters → Holiday Master",
    "week off assign": A + "Weekoff Rules",
    "alternate weekoff assign policy": A + "Alternate/Occurrence weekoff policy",
    "c-off policy configuration policy": A + "Comp-Off Ledger + policy",
    "priviledge/rights assign": A + "Roles & Permissions + Access Management",
    "reporting manager & working flow process": P + "RM field + Approval Inbox hai; manager-chain pending",
    "deptt hierarchy process & approval": P + "Dept master + approvals hai; hierarchy chart pending",
    "incremnetal/promotion type effective date": A + "Salary revision effective-date + Arrear run",
    "org. structure": N + "org chart abhi nahi",
    "employees dashboard, matrix & static analysis": A + "Analytics + dashboards",
    # ── Attendance & Policy ──
    "biometric in-out attendnace configuration": A + "Biometric devices + ZK import + Secure Punch",
    "live attendance fatching": A + "Present Today + Sync Dashboard",
    "count of present & absent as per time/shift confg.": A + "Daily Attendance + reports",
    "shift rosters daily/weekly/monthly": A + "Roster module",
    "attendnace regularization policy": A + "Punch repair + ESS regularization + approvals",
    "gate pass with in-out time deduction config.": A + "Gate Pass module",
    "miss punch/multi punch regularization": A + "Multi-punch report + manual punch + repair",
    "out station employees attendnace config./ on tour/od policy": A + "Tour/OD module (request + approval)",
    "computing of actual hrs work": A + "Attendance Grid hours",
    "computing of total hrs work": A + "Attendance Grid hours",
    "late mark policy/penatly configuration": A + "NAYA — Attendance Policy me Late Penalty (grace/slabs/auto-deduction)",
    "employees live attend. dashboard, matrix & static analysis etc.": A + "Attendance Grid + GPS dashboard",
    "various attendnace format /reports": A + "Report Formats + registers",
    # ── Overtime & Policy ──
    "over time count, hrs shift start & end calculation": A + "IN-OUT OT Matrix",
    "minimum ot count 00:30 minutes": P + "OT rounding policy hai; 30-min minimum rule confirm/add",
    "miss punch regularization with ot hrs calcuation": A + "repair ke baad OT recalc",
    "overtime hrs calculation & policy": A + "OT Salary Run + policy",
    "overtime approval process": P + "punch approvals hai; dedicated OT-hrs approval pending",
    "various ot reports & hrs report": A + "OT Report + Matrix",
    "customized ot report": A + "Report Formats",
    # ── Salary Process & Policy ──
    "salary strucutre assigned & its components.": A + "CTC/structure + compliance structure",
    "personal/loan advance": A + "Advances module (auto recovery)",
    "salary advance": A + "Advances module",
    "tour expenditure & approval process": A + "Expense Claims + Tour approval",
    "pf calculation, implementation of abry scheme etc": P + "PF calculation full; ABRY scheme pending",
    "esi calculation": A + "ESIC calc (portal-match) + accident register",
    "ot hrs. calculation & process": A + "OT Salary Run",
    "bonus, ex gratia, incentive, calculation": A + "Bonus Run + registers + yearly summary",
    "deduction ( fine, uniform, assets etc)": A + "Deduction heads + Asset recovery + Late Penalty",
    "tds calculation, form 16, tax planning, (old regime & new regime)": A + "Form 16 module (old/new regime)",
    "f& f calculator with required format.": A + "F&F Calculator + settlement",
    "required various format": A + "Report Formats center",
    # ── ESS & other ──
    "emmployees self service portal ( attendnace, leave, od, regularization)": A + "Employee PWA",
    "mobile app self services": A + "PWA + Expo app",
    "auto mail to all hod for present & absent report": A + "Attendance Email scheduler",
    "auto mail & notificaton for approval": A + "Email + push notifications",
    "employees greeting mail for birthday etc.": P + "Birthday report hai; auto greeting mail pending",
}

# long OT policy row (text truncated differently in sheets) — match by prefix
PREFIX_M = [
    ("over time policy (upto 48 hrs restrict",
     P + "OT calc full hai; 48-hr weekly cap rule pending"),
]

wb = openpyxl.load_workbook(SRC)
filled = missed = 0
for ws in wb.worksheets:
    for r in range(2, ws.max_row + 1):
        c_val = ws.cell(row=r, column=3).value
        key = norm(c_val)
        if not key:
            continue
        status = M.get(key)
        if not status:
            for pfx, st in PREFIX_M.items() if isinstance(PREFIX_M, dict) else PREFIX_M:
                if key.startswith(pfx):
                    status = st
                    break
        if not status:
            missed += 1
            print("NO MATCH:", ws.title, r, repr(str(c_val)[:70]))
            continue
        cell = ws.cell(row=r, column=4)
        cell.value = status
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=10)
        cell.fill = GREEN if status.startswith("✅") else (
            YELLOW if status.startswith("🟡") else RED)
        filled += 1
    ws.column_dimensions["D"].width = 55

wb.save(OUT)
print(f"filled {filled} rows, unmatched {missed} → {OUT}")
